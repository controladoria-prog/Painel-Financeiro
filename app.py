"""
Painel Analítico de Performance Estratégica — Controladoria B&A
=================================================================
Dashboard financeiro em Streamlit: consolida Orçado vs. Realizado,
DRE detalhada, histórico mensal e projeções de tendência.

Fontes de dados: Google Sheets (com fallback para arquivos locais em rede).
"""

import base64
import io
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import datetime
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName

# ============================================================================
# 1. CONFIGURAÇÃO DA PÁGINA
# ============================================================================
st.set_page_config(
    page_title="Controladoria B&A - Painel Financeiro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

FUSO_BR = ZoneInfo("America/Sao_Paulo")

# ============================================================================
# 2. DESIGN SYSTEM — paleta única, usada tanto no CSS quanto nos gráficos
# ============================================================================
COLORS = {
    "bg": "#0B0E14",
    "sidebar_bg": "#080A10",
    "surface": "#141824",
    "surface_alt": "#1A1F2E",
    "border": "#232838",
    "border_soft": "#1D2230",
    "text": "#F1F5F9",
    "text_muted": "#8B95A5",
    "primary": "#4C8DFF",
    "primary_soft": "rgba(76, 141, 255, 0.14)",
    "secondary": "#5B6472",
    "positive": "#3ECF8E",
    "negative": "#F76E6E",
    "warning": "#F5A623",
    "muted_line": "#94A3B8",
}

FONT_STACK = "'Inter', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif"

# Config fixa do Plotly: dashboard "travado" (sem zoom/pan/modebar) para uso executivo
CONFIG_PLOTLY_TRAVADO = {
    "staticPlot": False,
    "displayModeBar": False,
    "scrollZoom": False,
    "doubleClick": False,
    "responsive": True,
}


def cor_variacao(valor):
    """Retorna verde/vermelho conforme o sinal do valor (positivo/negativo)."""
    if pd.isna(valor):
        return COLORS["text_muted"]
    return COLORS["positive"] if valor >= 0 else COLORS["negative"]


def estilo_grafico(fig, height=400, **overrides):
    """Aplica o layout visual padrão do painel a uma figura Plotly."""
    layout = dict(
        paper_bgcolor=COLORS["surface"],
        plot_bgcolor=COLORS["surface"],
        font=dict(color=COLORS["text_muted"], family=FONT_STACK, size=12),
        margin=dict(l=20, r=20, t=30, b=60),
        height=height,
        # Separadores no padrão numérico brasileiro (vírgula decimal, ponto de
        # milhar) -- afeta eixos, indicadores e qualquer número que o próprio
        # Plotly formatar automaticamente.
        separators=",.",
        hoverlabel=dict(
            bgcolor=COLORS["surface_alt"],
            bordercolor=COLORS["border"],
            font=dict(color=COLORS["text"], family=FONT_STACK, size=12),
        ),
    )
    layout.update(overrides)
    fig.update_layout(**layout)
    return fig


def kpi_card_html(label, value, value_color, subtext="", subtext_color=None,
                   progress_pct=None, icon="📌"):
    """Gera o HTML (string) de um cartão de KPI, sem renderizar."""
    subtext_color = subtext_color or COLORS["text_muted"]
    progress_html = ""
    if progress_pct is not None:
        pct = max(0.0, min(100.0, progress_pct))
        progress_html = (
            f'<div class="progress-container">'
            f'<div class="progress-bar" style="width:{pct:.1f}%;"></div>'
            f"</div>"
        )
    return (
        f'<div class="kpi-card" style="border-top-color:{value_color};">'
        f'<div class="kpi-top">'
        f'<span class="kpi-label">{label}</span>'
        f'<span class="kpi-icon" style="background:{value_color}22;">{icon}</span>'
        f"</div>"
        f'<div class="kpi-value" style="color:{value_color};">{value}</div>'
        f'<div class="kpi-subtext" style="color:{subtext_color};">{subtext}</div>'
        f"{progress_html}"
        f"</div>"
    )


def html_compacto(html):
    """Tira quebras de linha e indentação do HTML antes de mandar pro
    Streamlit.

    Sem isso, blocos de HTML escritos de forma indentada (que é como se
    escreve pra ficar legível no código) são interpretados pelo Markdown
    como BLOCO DE CÓDIGO -- e a tela mostra as tags cruas em vez do
    resultado renderizado. Compactar resolve de forma definitiva."""
    return re.sub(r">\s+<", "><", " ".join(str(html).split()))


def render_kpi_row(cards):
    """Renderiza uma linha de cartões de KPI em flexbox — usada no cabeçalho fixo (sticky)."""
    html = "".join(kpi_card_html(**c) for c in cards)
    return f'<div class="kpi-row">{html}</div>'


# ---------------------------------------------------------------------------
# Funções de cálculo/formatação (usadas em todo o painel, inclusive no Painel
# de TV, que é montado antes da barra lateral -- por isso ficam aqui perto do
# topo, e não mais lá embaixo na seção 6 original).
# ---------------------------------------------------------------------------
def get_valor_consolidado_multi(list_dfs, termo_conta, colunas, exato_linha_sintetica=False):
    total = 0.0
    if not colunas:
        return total

    for df in list_dfs:
        if df.empty:
            continue
        col_nome = "Nome" if "Nome" in df.columns else df.columns[0]

        if exato_linha_sintetica:
            termo_limpo = str(termo_conta).strip().lower()
            mask = df[col_nome].astype(str).str.strip().str.lower() == termo_limpo
        else:
            termo_escapado = re.escape(str(termo_conta))
            mask = df[col_nome].astype(str).str.contains(termo_escapado, case=False, na=False, regex=True)

        sub_df = df[mask]
        if not sub_df.empty:
            cols_existentes = [c for c in colunas if c in sub_df.columns]
            dados_num = sub_df[cols_existentes].apply(pd.to_numeric, errors="coerce").fillna(0)
            total += dados_num.sum().sum()
    return total


def formata_brl(val):
    if pd.isna(val) or val == 0:
        return "R$ 0,00"
    return f"R$ {val:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")


def formata_m(val):
    if pd.isna(val) or val == 0:
        return "R$ 0M"
    return f"R$ {val / 1e6:,.1f}M".replace(".", ",")


def eh_grupo_sintetico(nome_linha):
    return bool(re.match(r"^\d+\s*-\s*", str(nome_linha).strip()))


def eh_linha_custos_despesas(nome_linha):
    """Retorna True se a linha pertence a algum dos grupos de Custo das
    Vendas (4), Despesas Variáveis (6) ou Despesas Operacionais (8) --
    inclusive as sublinhas deles (ex.: "6.6 - ...", "8.3.1 - ..."). Usada na
    visão "Gerencial" (foco em custos e despesas) das abas de DRE e
    Histórico Mensal."""
    m = re.match(r"^(\d+)", str(nome_linha).strip())
    return bool(m) and m.group(1) in ("4", "6", "8")


# ---------------------------------------------------------------------------
# Expandir/colapsar grupos da DRE clicando no checkbox da linha (usado nas
# abas "DRE Orçado X Realizado" e "Histórico Mensal", nas visões Sintética,
# Analítica e Gerencial).
# ---------------------------------------------------------------------------
def _numero_linha_dre(nome_linha):
    """Extrai o número hierárquico do início da linha (ex.: "8.3.1" de
    "8.3.1 - Salários"). Retorna None se a linha não começar com número."""
    m = re.match(r"^(\d+(?:\.\d+)*)", str(nome_linha).strip())
    return m.group(1) if m else None


def _linha_pertence_ao_grupo(nome_linha, numero_grupo):
    """True se `nome_linha` é uma sublinha (em qualquer profundidade) do
    grupo de número `numero_grupo` -- ex.: "8.3.1" pertence ao grupo "8"."""
    numero = _numero_linha_dre(nome_linha)
    if numero is None or numero_grupo is None:
        return False
    return numero == numero_grupo or numero.startswith(numero_grupo + ".")


def _subgrupos_nivel2(linhas_disponiveis, numero_grupo):
    """Entre as linhas da DRE disponíveis, devolve as que são subgrupos de
    NÍVEL 2 (um ponto só, ex.: "8.3 - Pessoal") do grupo `numero_grupo` --
    ex.: os subgrupos de "8 - Despesas Operacionais" (Pessoal, Ocupação,
    Comercial etc). Usado pra detalhar a composição de custos do Painel de
    TV com os grupos de verdade da DRE, em vez de nomes fixos."""
    resultado = []
    for linha in linhas_disponiveis:
        numero = _numero_linha_dre(linha)
        if not numero or "." not in numero:
            continue
        partes = numero.split(".")
        if len(partes) == 2 and partes[0] == numero_grupo:
            resultado.append(str(linha).strip())
    return resultado


def _nome_sem_numero_dre(nome_linha):
    """Remove o prefixo numérico da linha da DRE (ex.: "8.3 - Pessoal" vira
    "Pessoal") -- só para exibição mais limpa."""
    return re.sub(r"^\d+(\.\d+)*\s*-\s*", "", str(nome_linha)).strip()


def _montar_linhas_com_expansao(todas_linhas, modo, grupos_alternados):
    """Monta a lista de linhas a exibir, respeitando quais grupos estão
    expandidos/colapsados.

    modo == "sintetica": mostra só as linhas de grupo (nível 1, sem ponto no
    número). Se o número de um grupo estiver em `grupos_alternados`, esse
    grupo está "expandido" -- mostra também todas as sublinhas dele (em
    qualquer profundidade), logo depois da linha do grupo.

    modo == "expandida": mostra todas as linhas (comportamento padrão das
    visões Analítica/Gerencial). Se o grupo "pai" de nível 1 de uma sublinha
    estiver em `grupos_alternados`, esse grupo está "colapsado" -- essa
    sublinha fica escondida (só a linha do grupo em si continua visível)."""
    todas_linhas = list(todas_linhas)
    if modo == "sintetica":
        resultado = []
        for linha in todas_linhas:
            if not eh_grupo_sintetico(linha):
                continue
            resultado.append(linha)
            numero_grupo = _numero_linha_dre(linha)
            if numero_grupo in grupos_alternados:
                for sub in todas_linhas:
                    if sub != linha and _linha_pertence_ao_grupo(sub, numero_grupo):
                        resultado.append(sub)
        return resultado

    resultado = []
    for linha in todas_linhas:
        if eh_grupo_sintetico(linha):
            resultado.append(linha)
            continue
        numero = _numero_linha_dre(linha)
        grupo_pai = numero.split(".")[0] if numero else None
        if grupo_pai in grupos_alternados:
            continue
        resultado.append(linha)
    return resultado


def _processar_clique_expansao(df_tabela, evento, grupos_alternados, col_nome_linha="Conta / Linha DRE"):
    """Processa o(s) checkbox(es) clicado(s) no st.dataframe (evento de
    seleção) e alterna (liga/desliga) o grupo correspondente dentro de
    `grupos_alternados` (um set, mutado in-place). Só reage a cliques em
    linhas de GRUPO (nível 1) -- clicar numa sublinha não faz nada, já que
    ela não tem "filhos" pra expandir/colapsar. Retorna True se algum grupo
    mudou de estado (e a tabela precisa recarregar pra refletir isso)."""
    if not evento or not getattr(evento, "selection", None):
        return False
    linhas_sel = evento.selection.rows or []
    mudou = False
    for idx in linhas_sel:
        if idx >= len(df_tabela):
            continue
        nome_linha = df_tabela.iloc[idx][col_nome_linha]
        if not eh_grupo_sintetico(nome_linha):
            continue
        numero_grupo = _numero_linha_dre(nome_linha)
        if not numero_grupo:
            continue
        if numero_grupo in grupos_alternados:
            grupos_alternados.discard(numero_grupo)
        else:
            grupos_alternados.add(numero_grupo)
        mudou = True
    return mudou


def cor_valor(val):
    if pd.isna(val):
        return ""
    color = COLORS["positive"] if val >= 0 else COLORS["negative"]
    return f"color: {color}; font-weight: 500;"


# Imagem de fundo do painel visual da tela de login (colaboradores do Grupo
# Beea, selo Great Place To Work) -- redimensionada/comprimida pra ficar leve.
LOGIN_BG_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAcFBQYFBAcGBgYIBwcICxILCwoKCxYPEA0SGhYbGhkWGRgcICgiHB4mHhgZIzAkJiorLS4tGyIyNTEsNSgsLSz/2wBDAQcICAsJCxULCxUsHRkdLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCz/wAARCANrArwDASIAAhEBAxEB/8QAHAAAAQUBAQEAAAAAAAAAAAAAAAECAwQFBgcI/8QAWxAAAQMCBAMFBAYGBgUJBAkFAQACAwQRBRIhMQZBUQcTImFxFDKBkRVCUnKhsSMzNDVUgggkJVNi0RZ0krLBFzZDc3WUs9LwJjdE4Rg4VWSTosLD8SdGVoOV/8QAGwEBAQEBAQEBAQAAAAAAAAAAAAECAwQFBgf/xAA2EQEBAAIBBAADBAgHAQADAAAAAQIRMQMEEiEFE0EUMlFxBhUWIjNSYaE0QnKBkbHRwTXw8f/aAAwDAQACEQMRAD8A8nwemo3UEXewRl5aDcjdX30VA1v7PET6KOmo2PooCw5fANFKaQxML3u8LdV1mnntu2fLQ0MYzOp2ZjyVSSlp5BYwst0sp5XmSQu+SRrHONgrqG6qfR9J/cMS/R9J/cNWg2mJGqd7PZPS+2b9H0f8OxJ9H0n9wxaRp0wwEJ6PagKCk/h2J4w+k/h2KyYyOSALJqJupqKmow4NdTxkHqFojD6A7QRfJZFyNk4SEG9ypqHtr/RlF/DR/JJ9G0P8NH8lThrnNNnm7VoMmEjbgobqP6Nob/s0fyR9HUP8NH8lI5+qaXobpv0bQ/w0fyS/RtD/AAsfySiTRIZCmobppw+gH/wsfyUb6Cit+zR/JSF6YXp6N1AcPo837NH8lIzDqTu2/wBXZsjMrMf6pvotSM21B9HUn8OxH0fSfw7FZQrqM7qt9H0n8OxH0fSfw7FZQmjdVvo+k/h2I+jqT+HYrKE0u6rfR1J/DsR9HUn8OxWbqCWrihNibu6BNG6b9HUn8OxH0fSfw7FJDURztu06jcFSpo3Vb6PpP4diPo+k/h2KyhNQ3Vb6PpP4diPo+k/h2KyhNQ3Vb6PpP4diPo+k/h2KzZCahuq30fSfw7EfR1J/DsVlCahuq30dSfw7EfR1J/Ds+SsoTUN1W+jqT+HYop8Po+5cfZ2K8opxeFwTUTdZBoKRxsKdiljwqkd71OxXGsaB0SBxGyTGLuq5w2hjsfZ2JfZqMHSnYpnROdqSnxQty3Kul3UbaKkIv7OxL9H0h/8Ah2K0MvJO8IF1dL5VVGG0n9wz5Jfo2k/h2KV84b0VR9Wbkhw0809LupTh9F/cMUclFRMF+4YbbgbqtLOXmxJsFCZHdTpzUN1fFNhxAtFGQfJApsPv+ojWdmJ52RnO11fR5VqNoaF20DFI3DqM7QRrIEjm7ONlLFLI25anpfKtM4dR2/Z2JBQUg/8Ah2KJlW6Rt75VMJrszO6K+mt04YfRkfs7Ev0dSfw7E6OS7b8lK1wKuou6g+j6T+HYkOHUn8OxWkiaN1V+jqT+HYj6OpP4disoRZap/R1Jnd/V2ckv0dSfw7FY+u74JVyy1tztu1b6OpP4diDh1Hb9nYrKRZ1E3Vb6OpP4diPo6k/h2KZ0mV5ba5tf4KGSraI8xva2g6q6N004fSW/UMSewUn9wxVXVUptrlH+HRNfO4jLmJ6qaXdPmjomaMp2O81FlpS8f1eMeSQ2AGt0HxEXKuobq2KOje1uWBlinsoaO36hmipd4WmwOymFQ5ovupqG6tDD6M//AA7FHUYZSPp3tFOy9rhLBVB7tTYcrqz3zBYEixNlfTO6rSUdG+mjd7Oyzxrv0UcWF0xddtM2xFtrK3SG0BYde7cQpxqrti2qTMFovrQtJ6jRTtwyiBJFNGCfJWAUt0TdVZcPpMn7OxZVRh9IKkn2dm/Rb0nuFZNXpOfmtSellqKOhpBf+rs1CxMWhbBWhkTQxuQGwW419nA30WHjbz9ID7gW/VjWO9uroDeggP8AgCgxGpJtC0+qdRyBmGQuPJgWe5xe8k7krhOGrydHGXmyuxwho2TaeOzL81YAUtakIGgILQn2QoqPKmlimSWUFZ0Y6KJzPJXCLqJzVU0pvbZRlWnhVnBaiUNKtU0xjeAToVTG6kB0VrLXKao4JM8Q6hPKAukuhIUCEpEFCBqtxfqm+iq81aj/AFTfRWM09CRKtIRCE172sYXE6AXQMmmbGbFzRzJKquxNo+rf0VKondPISTpyCh2UVdnr3SANAsLa2VMHxXuUA8kFFSRyugdmYdfNW/pJxaPALqgBcIBQasNe15yvGU8lbBuLrBbYjXdWWVZYRva2ovoiaayVZcWIPDvGAW+S0IpmStu0/AqiRCEKIEIQqEQlQgRMm1jIT0yTRhQRBmmqTLlcpA640TXOFr7lVRe40CiJeDbZKZLc03OLoozOYLk3TX1Bta+p2Cill1trY81BK8gCyzsJMTmBzXKjJ30QTfzSFVS5jbXmm3QjRAJbpqEDuafcAaEqLmlGqCRrja1yntmePDclqhS5tkNtKCozeA6Ot03Vpjjy1WSJdABy2PNW4qxrdHA5ep5LcalaDXp4N1CHNc27TdIJCDYqtJykSB10qKZ9d3wSo+s74JL6LjeXO8i/isgkAEnkmgku0VWrqsjgxp+9fQKRFaerkM4LdMvIqFzrt8W/IJr35nuOlyboNrXuqE94k9E3c3uml2hCQHVA9xJKMxCQ3FkE6bbooBO6cHHbVGcNbtqhpuVA9pa3UmxUwlOUBrrka2VcAEm9koaQ3RFXqaoe9zi1gJfqbcrK3HKCbHQ9OayGOexwLTsrJqHSMDQLHm5GbGoCDzShZ4L6VwLnh8Z/BXYZGytzMN2qudh7hdpWXWj9KD1C1MwG5WdXMu5pG2y3CKQOixsZ1rgf8AWyARosTGHWrgP8AV26Y8t8SWwqnaN3MChjGaQBI116SnHSMKWlbeT0XL6LeV9gs1PCQDROAWWwlslCLIGoKcUllAwpjgpCE0hBVcOqruFirsjbhVX67qxKgO6cCkcNULTC3Rvs4tOxVxZcbskjXdCtTQi6BvJIU5IgRIUpQgRWYv1bfRVlaj/VN9FYlOQhRzTMhYS4620C0ycSG7qhXzZmZRtsqlRUSVDru0HIKG5RQT5pMpO2qkZETurUUI6LO2pFMMdbZPEbnbC602QtA2UoaOQACbXTIMDwNim924i4BW1lCR0LXDUJs8WHYgouVpyUTDsLEqlPA+E33CqWIweSniqHxAZTpdVhqVI0hpsURtxSCWJrxzCeoKRwNO0Ag7qdVkIQlsiEQhCAUU5tC74KVRVAvA7/ANc0VA2Q202UbpL3Gymiiuy6r1TMrhl2VUZr2Tsotuq+u6DKRoshJNXEHZQPN/NK5/i0Cad/JFhEiexhdqnd05niIuEa0j80hHNSsZnfoDZHdEyEN5Js0i+CDop2RhzyxwUklO2NhJ5Js0ppQDySnUJtlUB3S8kFLyQA2sFLGC5waCPUKEa807M6+iDQpZXsc9ps5jTY2Vx7CPEsmKTu2kHc8wtiCVs8II9CtytyiN3VS3uoHsLHXCe1+yrR31nfBRPcRcDUeSkvq70Cr1Ly2MNaLl52G643lzqOSqEQIbq7a3RZsmYm7ze53SucXB1xY800XLbHZEDSOqHuBNmovySbaoGEItZOvzTw26CPdOAOynEAIUjadTbWlTIeYTSLHRaIiFrWUctKHajRTa+Ks0G1yLpS4ctCnOhezUXUViN91UOa7XUpzCQ9Rjwm/JOJFrg2KIs993jA1wFr69VY70MAbCPUBUDltf8AG6mbK5l8pFvNalSxoNfnaHDZNnZni9NVRbJLqWvtm3slMU795St+cY8SOyDUkk/Jc9jUkYxD3fqD6y6I0t2gE7cwNVz+NUbfpAau9wLnlm6Yz204jemiP+EK1TOyAusSqVKb0kR/whasAa2FotuLqfRdezm1LeYsp2ysdsVXc1jugKa2Ox0Kyq6gkDcqNpICSS5bZUOdMxvNRmqZyCi7rXVODGt3CgcKkHkUrZWONtilDmgaAIOV41Cihw00VSVuqsi7fCTcckyRlxdCxRdukSvFnJo0XRzOutOF2aFp8lmbBXqJ14SDyKgsJEqREIkKVCBFZi/VN9FXViP9U30ViU5Uq9mcAAj081dVWsja8X2I1utIypbAaaHmEkTbm5RK9zj4iTba6kiHgWa1EjBcq5E3RVY9SrkajolATrJG7JeaBwASkBABS5VRGQoahgdEQVZIUFS090SOSJWMW2ukuQnObqSm2OyrC3QECcAnXcLWGqw6Z2WdvmbLcaLCyrNKhCEAiyEKBFHUENhJPl+alVes/ZH/AA/NUJHJ4d9CoZ25lWhmtYHkrDpLjMqqu7whQuNynyPCiWA5rc2nMobA50oZzulboQVZhmia4uedToFVg9mt4QCCdNFabFZlrXGyfE9kgu03UoGiy7KPs7o5btAyncKtNDLG8ll8pPJa9gkLAUSxm0dO/MXvBF+qfV+6GAEuOui0MtlXqYS+zmmzgi/RjuB2IsgDkVZfETuNfIJCwgWV2x4qx3RfRLl3PRM5qslRyRdCoW+isU1TJA67bEcwVVBupoml8gaOWui1CN5jmzRBzdQQoXtLTojDxamt/iKsFoK06K3fhjXOIPIKGaYWbID4mnaySud3d2DZ1rqpKdGtvsNlyvLFRSHNIXdUnI6JHGzuqS9jc6hRDtAU1Le6TpdFK0XPkrEbL8kxmpsArcTNFm1qQrWbaKVsaVjVM1qw2jEad3SlDU6wRVV8Nhe1ws+WICR46bLYcN1n1URz5hoVYl9s46DVAbpeynfEOW5Vd12nTZbc6dcWsdk9oLRbSyibc7KWxLRqLhA+O+ew0CuxPD26bhZwDtdbFW6Z7hJkd00KItWXP42P7QH3AujyGw6FYGNxn6QH3AsZcNY8mUf7DB9wLQu5sTd9lQoG5qKnHVoWwWA2Ftlr6J9Va72x59LHqnQzFz8pU+RtrWKdHGL3sEJErPdSOCeNAkIuo0hecjbk2UDi8sLg0ADqrTmgixF0wtaQbhWJVLvnX/yU8Ty8J5jj6D5JzYwNhZSkPb0KHDROa2yRwsFmtM6pFnaKL6oU9Vuq/IBdJw50+12qzQus9zeuqr/VUtIbTjzRGikSpECWSJShEIrEX6pvoq6sxfqmeisKcmTC8L7C+hT0jr5Tbey0yxa4tdOMhBAaNk2L9X8UyZ2a2lnC4Kc24jAHNZraxEQDqQrsb2EaOF1mx00kp1NgpfZ3wa3RdtNuqc1uqq08126nVXG6i/JGkUtTHEbG5PQKL29hdbKdVM5rGklwuVA6ohYfEGhVDu/LtQnhwkBaeaVskTxoQnZGnUJRjVMTopCD10UF1r10QfTlw3GqyDujNTUljUxg7XW4BosCAXnYNrndb42VjNCEtkWQIhKkQCiqBeBwUqjqDaBxQYpGV5CkMnhATJNXXTeSgCczklilboUOPRQKLkgdVZNGbb6qCAZpmgdbq86dgvqbDcgI3jFZjJYZNBZaULi5guVWD2v1BDgrMXuo6SJUx0jGe86yUmwKpSxF7rlyCy6qhH100Txv2cCqRpXE6FBpXAeaJ7W3gOFwoXRgqBj5IT1HMKyHB7bjZRqXahUR5DcDQqALTkYHtsVnyx908jkrGMpow7oQhaYIAArEJyuvbfc8lBa6twvaGluW+YWWsVjXp2hsIy89VIo4P1DQdwLKVabZuI2EzdLmw0VB4LACeavYictXG7Q2ABHzVSpa4jONGtXK8sVXJ121S5SPJNduErjcBRALBPa242TBqrMTdlGokijt5K2wWCjY3RStGqxa6Q9uilCY0KQBQKE7kiwQgSygnjzN9FYUcj2tHi2RWU65mMfUi6gkblmc3Q2U8j+9qCYxpzKZI2MNzXJcN7rcc6rk2cbJ3vO3skcAXWCVzMtgfwVZLa9yXWN7J8HeOJLXC97KHW9lPFIYHB2UHzQXaWclxZIPEFlY2CcQ/kCuBzw8yujILzcLMxdzpK7M2QNGQaFZy4ax5WMIbenpzyDAVsgLIwX9mh/6sLXB0QLZA12TS7W19Eoe07FRTym3SF4Sd4w6BwugcCLpS3yUL3WeLKZrrhAZUEJbpCgRMfsnJjzdqiqVUFW5qxU7qvzW5w55JBsnwm0zfVMCfB+vb6qo0wkSjZBQIUiDuhAisxfqmeirKzF+qb6KxKchLZFlpll19IAHTN9bWVdgHdNJWzKwSRlp0B3WPlAaGg3AJspWocKh4FmcksksphGe9z1Fk+OJwbdRS3uLlRs2nJ7weq2oz4AsqmaHShawabXsosMlaXNNgqZpg7R97ei0AbmyjffMVV1tAKaIG+Z2brdTMbbQEkJzW3UobZEQyNzRubbcLKgpu/kLTdoHNbEmgJVWgGUPdb6yIaMNbFKxzSTY7HmroFgnvvz+aarGMoRCEKshFkIUCKOf9S5SqCrdlpnnpb81VZc4s6wUF1YcM2qrnQkKUKghNCkY3vHtYNybKC3QxeF0hG+gUkjD3L4xYBysMYGMDBsNEFt1Nu0kkUIad8cmYuvy0WjEMrdU0M1TibCybXRJH2ChcbanRSO1THxZzc7hFN76MDV415pWvHI3CZPEZWNGYNym40UDIJIXaOzDor6SWrbmNcOSYyPJcclIL21RsEaRuVWoaHclaeoHi6RKokWTVYezRQEWK05WALTw6Jry8kabFZrBc67LdooTFDc6F2tui1CcrAaGiwFglQhabZ2INzzhth7o3WdKXtHdknKDstepgEslwbOAAusmqiMU5aTcgrleWLygJuEM3Q7U6BIN1ESg2U8TlWUkR8QUrUaLCLKTM1rS4lQx6JXsMjw29mjUrDY9qfu1uiX29zd2qwwMjbYAJCWu+qD8E2ukAxC/1VOypzW81GY2E3yJQwDayUiyDcXVWoGc2JsFZjIIsoZo892nmpCqTXNe9kcQAvsSoZGOdI5rgARppzVoUwhN2Eg7XUJZ+muTpsts6VGnc+SVt3CztOifLG+NxJ0B/FMDtNbqsHCzH2KDmcRzSFu1t0ljfzQSske7w5ic2iycXZeuvt4QtaJoD2hxNlk4/NH9JjIA4d23VZy4ax5XcMdkpaY/4QtgO0WHQm2HwH/AFsxnPGD1Ct4JyVpu9PLBuNFGfAwlRslc8m1zZRpMWEjdIIQNt0md40sUZ3jkdENJAzL5pQbFRd67TM3ROa9rtQVBLmSXTRqUqA3TXbJ2ya82Cgo1B8dlANU+V15E0DVdJw5U5TUovMPJQlWqNviLlRdSJUFQIUiVIgTmrMX6tvoq4FzZWYhaNvorEpyEIWmQdRbqslrLG3Qn81rLOk8NQ8dHKVrHlYjYCNlSqoiyU6bq+17WRZidFRqKkOJso6UMJgcywzX3sVph7pIxkdl8yLrHY5xOgVxvfGIAGx9UIuNMgFnlrvMCycWXGqpNjmA1fc+qlD5QLaFBMHd2fF7vVTBwIFlnyPlsQWnXndSUrpGXa/3TqPJBLUn9E6x1UdAW904Gw1TZfETc6IoCA2Q8geiItAkgi2l9LoRmDifLqiysYypEJUiqBCEKAVav/YpPh+asqtX/ALFJ8PzQZbZNFG4XNwEinjALbpyK6uYay9Q5x+qFVlbZ2iu4Z/0nwUXHlo2SEWTuSjlBLdDqFHc4N8N1G7QqN8zw3QahRtmla8ZmAtPMFBYbqnWTIngvJbsp7XQRZb8klgOSlLUwoGOTCU526jcUU15UJFzopHG6GAXuUEZic8EAKrJA9mpGnULVjiz6u0HIdUwxAtf0Vl9s2Sq1JQPe9riQGfmtkaaKtQlxhF2kAK0urMmoQoQUKqjcPG74LHrz/WSNSWnVbDtZD8FjVriKolcryxeVY2uUBtgkO6XMFECkh1kACiLtFNTaPupVjQaNAn3ypsfJT2BC5uqpNNlCgM0r3ZfEBbTKr5ga87JW04aVZZEsRZGFjMgIdbxOvzUzAQ2xN/MpwYnZbBLdrIivlcpCbi4UbrBK02CilLc2ygdCc1x4XHy0VkaHyTZb2Nt7KppmzguY7QAje2yr2aIyOatyW7g9SqRWoxkc05SC4X8kZsp0Fv8AgkvsUuXMS6+p1VQ8OscwvYdVj4swurQczR4RutcGzrdVh4y8Mrg0AO8A1WcuGseWlRfu+H7gWjRS2uw/BZ1F+74fuBTscWEOHJa+jG/bWLQVEwd3I6w35J0EokjDgfVK5l9Qs8Nw7vLn3bJHSXvpZMs5KGE7o16MdmeAOScyAN11UrRbklUQjW2QgmybmUClV55MrSnvkABN1RmkznyVkS1GDd6UbpGauSjUro5HFXqMWjv1VA+8tOmbaBoUVKhKUiBEhSoQDfeCsR/q2+igbvdTxfq2+i1GachCFQLNrrsqc3JwC0lUxCIvjDgL5d1KRDkfNStya2Oqglg7n3hcqail7p2U7FLVHM+yjavTuD3APOUX5LUjZTtYCXAnncrOijBcrIyxjxBGouGWBlsrc3oFBLGamwyd2BzG6fCQ7YKyQAEVBHA2MAAX8ynPFvFyASueAoZZ/CbIyjOjXEndMpRJ3TiwaOOoUUsmZuUK9TMyQNCrNSMDgPEbm/LYJyEIyEiVIgRCEKgVbEP2KT4fmFZVev8A2N/w/NBiJ7SWpGDxKyYg5ikgruIKtYcbSPHkqpjLXWKu0bCw5uRCi48rt0hcOqY9pc0gGygjbJrqNFHee1mzSlIb0UHjH1fkm969n1Si+K00NaLAWTwVVbVsJs7wqcG4uiJCdFG4pSdFE46oQ1xUTjdOcUw6o0RPc0d2GjcprRdwHVTth8RebkoJIRliyn4KNmZws0auKdmJbYak6eiswxiNluZVk2xwfayEqF1ZNKEqRFRu/WO9AsvEmWexwGrrhah/WH0Cp4g0GAP2LTouV5YvLJNsxCTLyQ4akp3JEJkUkfgeE3kLHRK0E+oUrUaEbtArDXXVGN/hUzH6rDoutNk4WPNQNfdSNco0kskcdEt9E1x5IivfM4pzTY6pjmOa8kC4KCx793EDoFBYdIMiYTokaywsTqnEi2qopVQaGE21us9w0J1VyvNo7DmVUaS5tjstzhzyI0jIAnAg2ytsmD3iLpzSW6B1gVpk57wHXA2XP4029eLn6gW84C+97Ln8acPpD+QLGXDeHLZof3fD9wKcbKCh/d8P3ApxsFqOd5OglMUnlzWq05gCsY7lXKSpsAxxUsWVfyoITBILboLx1WWzjokzJjpBZQulvsgle9ROkACYSbJpB3QMke56hdoFOWqB61GciMGhKdzSN931ShaYG5WtELRNHksoe8FrM9xvoopxSJUigCkSosgGmxU8f6pvoq6sR/qm+i1EpyEIWkCHNDmlp2KEIMmojNPLbkdQmCXOLH4LWmhbPHld8+ixqiB9PJZ23I9VNLKsRHROe4E3JVRkllK2Q7lRrbQp3gN0Uxku0lZrJS3bmh1QWgjdU2mkmOaw1JUDpblNZmkdYC7nfgr0NIyOznDM7zUOUMEDnuDntsAr7dkAWSjmk9mXBUIQtMBCEKBLIQkVAq9f+xSfD81YUNY3NSPHp+aDIhZnN1baBcC6hgjI1CfZwk1VgkdG126swMaGWVQNeSLDRWWNc0KE5Ob0SOjN9EMBBN1IFnTvKhtIDokLC466KwhRdoG07b3Iv6qUCwSpjnIFLlG4oJTHO0QhrimJHO1SA3KrRxuxzHDe91YEzicojcXHXRQt8Un5LSiaGsFtytSRztMhhyDM7Vx19FMhC0mwhCECJE5NKoid759AqlfrSnyN1bd759AoZxnjLeq5Xlm8sTzSj8UuQgIA8agbrZSxOuywNio9nWTs+UjTY3uqJmHKbX3UzSqkkl3EjS5vop4pMwWbGpVlr1M16rN1UgBWW9rIfYJM19VG0EhNkEmmUj0Ki7TXTS8N5qqZJdiA1Aa5277K6ak2ndO1u2p6lMEpkeByKQRNbe5BT2Msb29E9LZIp1591vxVVtizfVS1rs1SQPq6KC1jZbjz2+ykWdayUkAaDXmlBaQb7pB81UKLEXvzXPY2D9IbfUC6BpGaxGiwcbAOIDX6gWcuG8OWzQ/u+D7gU45KvQ/u+D7gVjn6KzhzvJrveTU47gpFUW6dxLQCVYy3CrQC1laCw6RGWlGRS2SWUVGGJcuifZIRoggk2VWTdW5NlXcLlajNN2ARyKCjqtMF5q6K1rWAZSbBUU7ldFXm1rDu0hSCojd9ZZgKcHKaRp96z7QRnb9oLOznkU4S9QppWgHC41HzViP9U30WR3gstaA3p4/uhaiU9CELSBLySIQKFWrmNfT6jYqyo6gXp3+iDDdGWnwpt3De6sOF0sceZZ21pAHGynhp3yHaw6lWoqa1tFbZFbkm2tI6eBsTbD58yrIaEBtk+yBpCQc04qLNldrsiWbh5QjdC0wEIQgEiEiAUVScsDipVBW/sj/h+aIq0pBJHmpJGAvFlXiBbqFMxxL9UFlkYtdStakjIyqQWVa0hlGVt0wOVotBFjqFTmifCC5niYPmFmumNSZkmZVRUt9EjqoKNLRKhe7XdV3VROyiMjnFRVl0gAUTpL7KOxO5TmRkobAN05ouU7JYJbWCqpYBeUBaQ2WfSNzS3V8bLpOGLyVASX1ShVCpEqSyiBNTkiqoXe+fQKOQXPwUjv1h9AmuFwuWXLN5ZMkfdktd9U7KJwAFxqVbqYyx7pLZgfwVUi7Mw56KCMfml+rbmgtttskG6oHMskjeWusnG7rC6a9tjcKC5HLeystcCFnRuuLKdjnDms2NyrzXJXHVV2yhSZweazpo+wduEgiHRAcE8OCLsCIJSLBOuFDM+zDl3shWTK7NO53Uplja9tUrgbn8UC9l0caBo65CdeztBZIWkWBQ43A6oCxDgTssDG/3hy9wLoswcwBzfiucxu4xC3+ALOXDpjy26H93wfcCnG6goT/Z8H3AphotRyvJD7yGi7rI6qWBl3XSk5TsbaxU7U1rdFIAsOhQiyAlUU3RIU6ya/ZBA/dQkblTE3B6KCV41F9VqM1GkQdwjmtuYCc1N2ShFBshKdkiIW6S6EbIFBW5T/s0f3QsILdp/wBmj+6FYlSIQhUCVCOaATZf1L/QpSQ1pcSABzKpTYjG68cYLri2bkhFQalSRaOTGe8pi21iFh0Xoz4QpWlQRHwhTNQPCVIEqoQqJ41UpTXC6hEDi8N8DgD5i4TYqtriWy2jeDbyKfK5sTC5xtZZL355HO6laxjOTbDg4XBBHklWLDUSQE5HWB5clcjxFp/WNt5jVa0xteSKNk8cg8DwSpFFIoav9lffy/NTqvWfsj/h+aCjG4mzQrcVPc3Khpo9b2WgwABANZYJwSnZMJIRo9BALT5hNDrpSfCUGa6Np5Jndt6BSi6QhYdkJYOQRkU2VPEaKhbEpQywUgagiwRELlE9wa0k7KR7gBcqlUOLiBf+XorjN1MrqNDDrueXOO+wWgSsqhcWDM/QdSrvtMb9GvBPkV2s0543aYHVPChY+6lGylbOQkuglRAkKLhCohf+sPoE26V/6w+gTSuWXLN5U8QJyNPK6oX0K0a5wEFidSdAs3cDXVQLe42+SZsU8C1rFIbZfRAC3rZKSC226YDZK1uYHqFQjTkde2itMcCNCq1jb/gnNaLXY6xHIqa2b0tht07uyq7JnN0cCrMc7Hcws2VuWUAPHNOBdzKlBaQls0jRZaIASN00tvcFOYcrrHZOdYm4RWbNA5r3GygtbZbEkecAqrJTDotSs3FUuC3oVHY3UrgWXBGiYbtAvqCtRz4OIs1c/jTP6+L/AGAuhFnC4Go6rAxsE4hr9gLOXDpjy1KH9gg+4FYJ2Cr0OmHwfcCn5LU4cryD5K5TFjmgbEKoAnA2NwbJZsl00w1LbRVIqwt0kF/NWmTRvGjgsWOkoSpTtohZUhIVaefXKBdSTyZGG26rMjvq7mtRKYSXCznG3QKMsN9BorJMTBqb+igklzaNFgtRimOsgDQFJyTmatWmTQlGhslIsk3QOQkB5IQCEIQIt6n/AGaL7oWDZbsLg2kjLiAA0bqxKlRewuqM2JNbpE3Mep2VGWeWY3e8ny5LQ1pKyCPeQE9G6qpJiht+ijt5uP8AwWfdJdQSSzyzu/SPJHTkkj0cb7lNbqCpI7FwzGyaIniF3WVoNzKONgA8PPmrMbbDVYdSx7aqXMoXODQka++5QW2pb2VZ1XHG3VwCqyYiPqgu/BVGk5wHNVZ62OK4BzO6LNkq5ZNM1h0ChsSrpm5JZqh877uOnIdFGiyOa252lRdCQoFuRrdTMrJoxYPJHQ6qC6EF5mJPHvxg+midPVxTUrmtJDtND6rPSjUqWLK0qVzcoVxpusqC4dotKM+EKRqJUhF0XRdFQuBaUGTw7qQ2KrTWaL7IhlkWSg31QsO5WtUgCa0p11EFrKGaQRtuTZJNUBgs3VyoSyG5c52pW5jti5yCWUuN3Gw5NUPelt8ot5phJck1WuOGN7OL3O94kp0TzG8OGllGE4aJBv0wuwHrqrIFln4bVtkYInWDwNPNXy5a5dJdwqaU0uSXKqnXReybcprnIpHn9IfQKGedsEZc4+g6okmbEHPcdBb4qlGx1bN3smkY2HVcry53lXkdJI8Sv0ze76JvO4VzErNZHYc9FTbqog0vYI0vuhx10Ueu6ilKVhIPwQRYi5uEXtsqF1JRY/8AzSh2iS5SBwJAtyTw6Mv8TTby3Ud0i0ylEjmnQkjzU8dQB712+ap3Sh5Clkq7sa7HRyN3DvRLl6HRZTX2NxcHqFZbPMI81g4bXIWLg3M19ug2SSBoaSTos81k/Kw9AoXzSSCz3kjokwW5iZ4kk02SMIuWuFweibdGw0+a6a05275WX0hY3M0kWF7Fctjd/pDc+4F0xmfLHfMSWjZYGORD29pJYSY2nVZy4dJZb6X6H93wfcCsBV6DWgg+4FZAspOHO8hKiyLKoRCVFkCh7xs4p3tEoFs5TLJFNGznPe86uJTbna5QiyoQpE6yQoECczmm805m5QOIuEzmn3TSLFAEICEqAKLIRyQInPlfIGhziQ0WATUh1NgrEKDcoJQNAmk3WwJUIHVQA3Tgb7preqcFYh7JHs91xCl9rm+3dQJENpjUyHd34Jhke7clMTk0bpNUZUoSohLW2SpEqoEISKIVIhCBUnNKm80U7klj1eExx1spINaho/8AWyEXIY7K4wWUUY0Uhe2NuZ7gAo2lBRcWus6XFWNuIm5j1OgVCasmnPieQOg2Qa09dBF9bMegWbU1zpmloaGg/NVXEEDy800DVKNtjSaZjv8ACEwysbu8D4rMdNI8BpebDSyjICxpvya4qoR/0jU2ScOb4DoeaoQxXOZ2ysFwa27tgumOH1rnl1LxA9zQzM469FTe4uddOllMjvJMVyu+GcZr3SXRdCXZZbASpAUt0Q5jyxwINiNbhbdJUioiufeGhCwlLDM6F4c0qxZdOhACXKFSZVgsBU0dQ1+xWq6py1QvboU/PomPdoosUJqYz1IB0jaNVaDQ0BrRYDkkBu53wTlyvLneVHEdo9eaqn0V6tjdJCCwXLTfzVC+uu6iGlIXXblTjumlnMmysUAAkAfFFgksBzS6K6QZrFKCLpNEtk0ES62QlV0ECEqECtJabhTRTWN/dP4KBF1U0vd3HI3M3wO6cioTAHbHK7yGnxTIpSzfZaNLJH3YbzvqTzUppmPjfGbOFim+oWtLlnkDGhpsNQqMlNZ9oyLHkeS0bQMIa/MNCFlYxGH1wc3UFg2NlsPhewBxbdp58lzmNF7cQs1xAyBYy1prH3Wrh4/qEH3ArSioGj6Pg+4FPlCkS8moT8o6oyhEMQn5QjKEDEJ+UIyi6BlkilyD/wDk2SZY+cjR8UEaN08taBcEH0Sta12qCEjVOb73wUvdt80CMX5opiR26l7tvmgxi/NREQS8ypO7FuaO7CoiQVL3Y80d0PNBCgC1z1U3cjzULhqQuvTwufCUl7oQELr8jNNg+SU6NSIOqvyMzZRslCS6Lp8nNNlQkui6fIzNnITbounyMw9FkzMUZinyM0SIUeYozFPkZiRCjzlGcp8jMSIUecozlPk5B5QN0zOUZinyMgo94lS0+lQ02UN9E5srmuBG4T5GatSWoZTxXdqeQ6rJnnfO8lx+HRTxM9rqbSuO3JWThkH2n/NcM5cbqtz2yUq1PoyH7T/mj6NgBHif81jZpnFobGDe7ncuiaLAHqtR1BC51yXH4ppw6E83/NXZpmKSKPO652Cv/R0P2n/NSNo42tsHOsrLPqll1qKhcI266Doq0kjpDc7dFpPoI3m5c/5pPo6H7T/mlz3wmOOmWUvJaf0bD9p/zR9Gw/af81nbWmYkWp9HQ9X/ADS/RkP2n/NXYyr2SrU+i4PtP+aPoyD7T/mpsZYSrT+jIftP+aX6Mh+0/wCavlDSjE8g25K3E/LZSDDYR9Z/zTxRRj6z0tJuHNnHVI6YdUCkYPru/BKaRlwMzvwTzb8qbE7NmI1TnOawXe6ycymbGCGudr6KGSlbn1e4+qztlFLWG1oxbzOqzy8l5FrrSNIzqU0UEQJN3a+aCjsml19lfNFGdLu+aT2KNvN3zWtxFENPNOsrnsrBzck9mZ1cm4KZSA2V32VnUo9kjP2vmm4KlrhCuikj5F3zSijj6uU2qlZACvexM6u+aX2KPq5XcFAhFlf9jZ1d80nsbOrk8oKScHOZ7pI9Fb9iYebk72Jv+JNwVopyw6kjzCtQtu3PmBPMc9UnsLOrk9kJiFmuKlqaKA1zwwnR2lvI81yOOwvbiVspPgC64tde5Pppsudxtp+kB4j7gWbZprGXfpaoP3fB9wKwq1B+74PuBWLqzhLyVCBqbbodlZ75seg3RAhR96C6wb80d6RcAmyKeXAC5URlcdtAnHxt10KYIZHbDT1QNvfndCmFNl9+Ro9DdH6BnMvKBsYPzVqR7XAFrcttwoGva51xa/RP30QLfVKEiUIhUJUtrhA1Fk4BLlRTbIT7JCiEG6qO94+quKm73j6r19tzWaRCEL2shCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEILWH/ALUPulaZWZh/7UPulaZXzu4++6YkKQpSkK87REiUpFAJQkSqgSpFG+oji952vQIJkKk7EAPdjJ9Sk+kHX9xtvVDa8nBURiDebSPQqeOriksM2U9CoJkqEIBAQgbKhUIQgEfWCEfWCBUx8Ejzma5oHQp6a2VhlLO7LiNSQUDfZZvtMQaeYC/hKaa+njkIc2RpG6BX0hGrnoiOYSQx53Aa7BV21OZt7WPMWUtXVwzRNbE43Bv4h5KKCFwYXBw1PRVDmylxs0An4hOGci5aB5XQ1rmu1It5JyJslndB80oa/wCyPmnBSWszNsFTaOzhuy/xTwerD80gljJtmCk0te4spo2A4fYKXM37LkNGYaa+iUsLRqhswlvR3yTc7Q4NIdc+SfZIdkXZS+Nou64HUhOEsX2gqEriS3RwF+au09MGlr3EuJGgKzfSz2n7u4B0TXMA3U3dgjUAlQyNs4gQXA5rO6vpG7Kudxtv9ofyBbxc5tv0BAXPY7J/aWkRIyDXqs5X03jPaWh/d8H3ArAFzuAq9D+74PuBSyDRdpw53k99QGgti0vu7mVXJJN+akZA92p8I6lOLWAZWa9XFBEPCPNSQx94435BJZu1lcpISATbQhTZpULADobJQHAaG4Vx9MQ65TXQluoCbXxVC3mmlgO3yVhzGl1xoVG5tz5qohy21UsbtNUFt/8AimEZSgmBTwoWO1seamaLoHgFODNNUrQpAFA3KkLVLlSWCCKyLKQhNIQNVF3vH1WiGElZz/fd6r19rzWMiIQhe5kIQhAIQhAIQhAIQhAIQhAIQhAIQhAIQhAIQhAIQhBaoP2r+VaazKD9q/lWmvndx990x4IUhSlIvO0RIlKRQCbJKImZinjzWbVSmWSzfdbsqCSskcTlNgoQ1zjsSVYhpSbF2nkrccIaNAptdKAp5CnCleTqVohiXIm10y30rm6qIgg2utgt5KCWma++lirtLFenrHxHK/Vv5LSY9r23abhZEkLojrq1WqGTx5TuiL6VIEqAQhCAR9ZKEn1ggVLGwXc7S6RZ9T3pqnhhtZoOyHBslNNNWSNiN7G+6bFGQ4sfD3jgeTtlC6aZjyDbN1tulbUTBpc22/RVDnx+LRhb5XurdPpEqIne6QF2hvyV+DWL4oyU7hI6wbcpXua25d8PNVXyOcfERbogkM5F7N9FE6RxF3nTkCm5rEHkEjjmdfki6KCSd7JJC6wFzZDHXtyKUvDTZwuE2HMmljYA15Gt9FM2ukJAf4h6bKDO0kDayCG3ve102umgyRrxdpSnYrObLkdoSD1Vlk5cwh3wKbZ0VzgHxlwzDeyvZO9isCRfos6b32jyWnES6JtrBYvLURCB+W/eO+aUwy5Ce9db1UrhamLh7wCSdxbSk7G1kVXySWB71y5zHS5uIgGYg5ByXSzNDYQ4E5guXxmISV4doSWC+q423Xt2wkqegA+j4LfYCsi4KrUH7vg+4FYG69DheSOJJ5+SU+EW/DqUoF5PTVTUsPezZzsFasmz6eizAOk58lea0NbYCwCUDRLZc7WyZbpMo6J9kWUVVkgab6WVOSIh1lqEKKSIOYRzWpUsZZHzG6RzeY2P4KV7TvzG/mmX0sdituaG1tFZgcH6HcKBzeXRI1xY8OG4QaTWp9k2J4ewOHNOuoENydLAdSkyai5ueqcNNToEF4+qLoFykppyjnr5JbOfudOgSBuXRAlyTpos1/vu9Vp21WY/9Y71K9fa81jI1CEL3MhCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEILVB+1fyrTWZQftX8q0rr53cffdMQkQk5LzNApEIVEVTJ3cBsdXaKvTxADMQn1AL5rD6oTgRG3UqVYlFgnB3koGzMvqrMbo3+64FTTRA5OunZAUhjPJAmYJCbpe6smltkET2Agg7Kk4GnlBbpY6K+VXqW6BIlX2OzMDr3uE5Q0zs1O09NFKtMlQhCBQk+slCT6yBeSqzNka8yAZmN94K1yVeeZrGyMzC5A0QZkjHRzDQt5hXJQH2DW2JA23JSVZbK+JwNy0WNirEUzf0IJFxe5RGVYhxvvdXIpmtjaMxBvdQVIAncRzN00Czb35KokfLd3vEi90zMC652Kie4GxslaC82ARYe4k6DUFGV432S91INQLp7JHtOV7SR1U21IhA1Ti02tyVowxv5JO5I0vp6KbXxUtQnBxtrqrD4OigLCx2oNk2nibe4/JSRv0CZIwxm/JID0VRaa/OQL67LSjnZG0NuDZYwcQ7RWmjOAprZtpGphczKXCyHzxSNylzSOiqOpXizch1SspH57W1tfVPE2mkkZI3KTceq5rG8oxDQD3BzW1MwskLToRyXN43+8N/qBc8sfTpjfa9Qfu+D7gVkcz0Veg/d8H3ArA2PourneStGjj5WWhRR5KcE7nVUmjwELRptYGeimS4prIshLZZaIhKgqhiQp9kxzg3coKVRHlfmGxVVzdT57eq0JXMe0i6puABIOy1GbEI1FuYUbm6g8lIQWP1KVzdfIqslgl7t2XXKVa7wnYWVEa3B3VmF5c2x3CgmsTqTdStAtsowpGoHNJSOGqAczrMBcfJPdTy2Be4NB5DUq6TaFxAGpssx/6x3qVvR08LNmZj9p2pWDKLTPHRxXq7W+6zTUIQvcyEIQgEIQgEIQgEIQgEIQgEIQgEIQgEIQgEIQgEIQgtUH7T/KtJZtB+1fyrS5r53cffdMSFIUIXmaIhCEFYG8jz5p7I22u8XKbGNPVEshboASjUiQwRP8AqAFMNIWHMwqJ0sjH2DmWtfZTskeGNc62o5IJYHnVrtwpr2F1VY896RyUsjwGqKZJU2HhF1CaqQ/VHwCmDmtGwJStlBGhbfoCqiuKi7rOCdI0PaFKcsmhaFGG5bjoop9HpG4dCrN1Xptn+qnVYKhIEqBQk+t8Et0n1j6KgVCqphNUOIdY2GivquGl1c+5s2wQU6ijLGsLbWOhPQqN9K9jOXh1uCr1Y1zYgAdHOFwdVZdAxzCDu4W0VRhNBLkPfc22AUkje6me3XQ2ULh4iiHNBcQAr8UIa1Vadt3hXwWt3Wa6SHCK6cIQAo/aGjS4ThKHbFRo4x67hKIxzJSZtUx0gbugmyMITDEOirmssbAFPZUlyaQyaIFpFlSczK4haRcHhUqhha66sSxCXFoV7DWiZzw4kFouLKiRdWKSV0JJbzVrDcdfvIxvufwTWguqXkiwAAVH26XTTbzS+2yXuBuoug8B0sxJF9gsLGYb1w8VrRtW2KpwJs0arAxmscK/b6gWb6ax5T0H7vg+4FZGx9FWoB/Z8H3ArTPeXRm8nN3I8loUhvTs9Fnt0d6hW6WZrYspGrSQpVi6EtlB7SPsqRkuc7KKfZNc4NGqcdlC8Au1UUwyvfowJREN3G5SlzWCyifOdbDbTU2VEpDQNgq88QcLgJGukcQbC3qrGXTVQumaWZm25jbzUYvYg7hW5mZXX+RUErb+Ib8wt7c7EDr3zKRrrEPG6Rp1tuChzCw2+SDTip5ZWhwaGg83Kw2jY2znkyeR2+Sr0FYSzunkEjY3VrO4ixNgnDBwc1pIAyg8gmPBfcW0TycrMwaXEEaKXQgjnZOV0rNcLAFYMusr/vFbTxlkPkbrFk/WO9SvT2s1lWTUIQvegQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCC1QftX8q0Ss2g/af5VpFfO7j77piQpEqRedoJrnW5XTlHIb6KLAzbRJlBSs2AUoao0j7tv2QmkdAFPluo5BqAAqpsQsbp0ovaycxtgmv0O6gY38fNQmmNwbAi9zbdW2tBF0pj00V2mlaPO13iGnJSHXVK4GyTkoHwC0ZPUqVMjsGAA3snrTBUoSIUCpB7x9EXRzKBbhZ1VUvhqjkAOgWgsytic6cuA0sNVQ19dJILOa217qQYnIPqhVMhz5eaQtLTY7qofNL3ry61rph/NFk52sbfiiJqYWbdWQzNuoqcWhB6lOfLlNm6krLrOCmlaTfZDY8hURllzWuFOHGxDuXNQiRp0THNzFTRRlzL3sojdriCjRAxjeiddtlC5jntN735apjYHgk3sfVEWdEyZmeMoY14Hi+YU3KxQrJIIWhSwNNG6R178rKnMzJI4dCtOExspGsza6E3TK+nOcmSU+SLPY7pGxXbfXTVWqiRkkJY0g9EjJGsiDfnZJ/VVEAnkufxv8AeGx9wLqIxlYNdbrAx1hdiV9NWBTLWmseU1B+74PuBWRuq9B+74PuBWF0YvKToQp6N7QXNNuqr38HoQmkkSnLoQVKRpvlYwan4JGTB+rVE2K40vY6m/MqVrMoAvsstpmuuFE866JQbBMduougBe3klMec3O6Gp4KoQMDVcFA4tB7wa+Sq30Wuz3B6BeHu+tn09eL9d+jXwzt+/wAupO4m9a+rOkwlzxbvW/JQjBJP79n+yVtMY+WQMY1z3uNg1ouSell1NJ2acWVsHfMwsxtOoE0jWE/Am68mPc9fLh+p7j4F8G7fV62sfzy084dw9IXXFQwfylSfQcmWxnYf5SuqxbAMWwKQMxKhlpi42aXDwu9CNCs5S931sbq116X6N/Cutj5dPHc/pWH9ASh1xUs028JVxuHyhlnStLutloIWftvW/F0/ZX4b/J/eqbKKRhP6RpFtrJDRSd6HiRotpsVdQn2zrfifsr8N/kv/ADVCTD3yEkyNF+gVB3DsrnE+0M1N/dK3kLpj8Q6+HFP2U+Gfyf3rA/0dl/iWf7JR/o7L/EM/2St4kXSrf6z7j8f7J+yvwz+S/wDNYH+jsv8AEM/2Sj/R2T+JZ/slb6E/Wncfj/Y/ZT4Z/Jf+awP9HZP4ln+yVBWYM+jpXTGZrw22gB5rplQxr91S+rfzXft/iPXz6uOOV9W/g8PxH9Gvh/Q7TqdXDC7ktnu/g5ZCEL9S/kgQhCAQhCAQhCAQhCAQhCAQhCAQhCCzQftP8q0is2h/aR90rRK+d3H33TEiS6VIvM0VQyO8RA2U11G4XBI2RYI+SmFrKFpsE/Noo0luAFC59muc1oc4crpHP81HcF11V2np5Q6O7m5T06JomjkkMfiv6aJrSA3YozONwHW9EE0Y8IT7aqFjrDRSXUUj2qF3RSvOihJ8QvsiJWNAF+oTwo2nwhPvoqxTktwmpVUKk2JQkG5QLdQT2YwvOxFlMonyMIDDYnexU0K8cbQBK67umijqInvLXNBcedhsrQcLjYDkp2EAjK62mwUx2uXqMd0b2Nu5rm+oTmsa5gu7KLG2nNWMQnzgRaktOpTacAxbDRbZk3U8TLRNaeif3Q3SNKlCw6xGRYck22Y+QUxAsor2PkirUWjFFIwON+aI32akzhxtsoGWtonBt+SGu1sVKCECBoCQ2slPkmuOiIpzxd5PpppcpJGZed1ORmf8EyZv6MHoVUsN7pwhLt7J0EEkm2o9VO8ZaZ2n1U6l0hHms2opFkjSRZwsbLn8azmvGp9wLrmtD3OcTz5LncWj/rv8oWMs/TWE9p6D93wfcCsKvQfu+D7gVkLvOHO8hxs34oDrynzKa8aJL+IlUjYjHgCUprDdo9E+yw6Gc0xzSQpBYO1SucLbIqu1zmna/kpmkPbcKMBx30HJSRjK2yAI0Wwz3W+gWQdlrs0Y0+QXzO/4xfvv0M+/1vyn/wBeu9mWC4fg3DFRxdiTQ5zQ8xki/dsboSPMnRYOJ9r3ENVXOfRGGjpwfBH3YebeZPNdXgUD+IewuTD6KzqhkT4sg5ua/Nb4j815fw3WYfhXEsM+M0hqKWIubLCWBxva2x6Fc88rhjhjjdS/V7uz6PS7vrdz3Hc4eeeNsmP9JxqOzm7VqfFuFKihxvCWVdW8ZRkOWN2nvdWkeS5LhbhKt4tq56ehlhidAwPPek2IJtyXrHC8/BvFs08dBwwxjYAC+SWmYG3PK4J1VLgNtE3tQ4nZQRsjpmBrGNjFmixANviCt3pXPLG55bjzdL4jh2fS6+Ha9K9PKSX3dye5Na+nLz3FuA8VwPh36XrjFHGZRE2K5LzckB3kNFLwv2d4zxRTe1Rd3S0hNmzTX8f3Rz9Uzi7irFOIMbqqSqqnNoW1BYyAe60B1gfMr1/ijCqR3DFFhgx4YDStAaHNsO8Abo29x6rlh0unllbOI9vd/FO97To9PDq2TPqXe5N6nr6fWvLuIeyzG8CoH1sckVfBGLv7oEOaOtjuPRcvg2DV2P4lHQ4dCZZn69GtHUnkF7Pwm3BuGG1EUnGcOI00wsIZnts08yNTy5Kr2eU1HQw8VVmGtZMY6l7Ycut2NaXNA8rlbvbY2zXqOHT+P9z0uh1fP96zXjlqyXfr3P6OaqOxfG4qF0sNbSTTBt+6FxfyBK5/iLgPFOGcGp8RrpIMk7xGI2k5mkgnXTyVE8XY+cVOJfSlSKnNm/WHL6ZdreS9O7WZ5Kns+wqeVmSSSeNzm9CY3EhY8OlnjlcZrT1XuviXa9x0Ol1+pjlM79J/Zb4R7O24ZwzWMroaKavqmO7mfLmLGuZYC5Gmp5Lyfifhas4Sr4qKtlhlkkjEgMRNrXtz9F6f2d4pX1fZ1jM9RWTzTQ94I3veSWAR6WPKy8eq8SrMSkZNXVc1VIGgB0ry4gdLlOv4fLx1GfguPd3v+v8AMzlkvv8Ar+GvwQIQELwv2sCz8a/dUvq381oLPxr91S+rfzXp7T+Ph+b5nxf/AAPW/wBN/wCnLoQhfuX8BCEIQCLp8MMlRPHDCwySyuDGMbu5xNgB8V9D4dwxwj2O8KQ4xxJDHiGMS+EAsEjjJa+SJp0AHNx9eYC5dTq+HrmtY47fPBhlEYkMTww/WLTb5pg121XvDf6SFM6oMc3CzvYzobVIL7fdLbfC6scVcD8Ndo3BcnFfB0LKeuY1z3RRsDBKW+9G9g2f0I303vdc/nZSzzx014S8V8/oQpqSknr62GkpYnTVE7xHHG3dzibAL0cOaFIXNBsXNB8yvY5uGOBezCjh/wBLI38QY/LGH+wRH9FED12Ft9Xb8gqo7ZcEpmiGj7OsGjpxs15aTb/8P/Ncfm3L7mO43465rydC9epeI+yvjBxpMY4bPDVTKfBVUz7MDvMtAt8W29FU7FOGsM4kxDiGlr6KmrCykb3Dp2B4jeXOAcPw1CXq6ltlmjw3628sSBzSbBzSfIr1yefs87NHChGHDjDHYrCeaVwFPE8btG438iep5LMx/tYpuIeHa7CpOEMLpHVEeWKopwA+E3Bv7vlysk6uWX3cfR4yc1wFD+0j7pWiV3fGWB4Xh/ZpwTX0eH09PV1sJNRNGwB8p7sHxHnquEXi62XnltqTREFaPDsMVTxRhcE8bZYpauJj2OFw4F4BB8l2XHODYbQdtVNhtLQwQULpaUGCNgawhxGbTzXHbWnnmvRNcDlK9s40x3grg/iaXB38BUNW6NjH960MaDmF7WyrCHaHwI7STs4pA075Sy/+6ptdSfV5cPcQ1rnaNXp/EnDHDGP8BTcYcIwS0Ao35KuiefC3a9tTYjMDobEdE3hvgfAsE4Yh4p43mlbTT2NLQR3D5uYvbU33sLWGpKbb08xkdFG6znDN5myRr2nZwXqg7U+G8PJjwrs/w5kXWctLj6+E/mpIuMuz/iSX2fH+D4sL73wmro3C8Z6+ENP4H0TZp5ULOF9D8UEXGo+S6/j7gE8I10E1NUCtwqubnpanQkjfKSNL2III3C9b4S7PuDJ+z/DMVxLBaZ8jqNs88zs1z4bucbFS00+dLFovy6p4N9V7y2XsStoMOt9yZatHwB2bcX4TJLg1NAWA5O/o5Xtcx3oTv6hNmnzi7ZMAu5a+NYDUYVxXV4FHmqZ4Kk08eVushv4bDqbhey8MdjGB4JhX0hxbKypmDc8sbpe7p4fInTMR1vZNmngbDcKQaL304v2MxSeyimwkgaZ20Ti3/ay/io+IOyPhviXBjifB80VNOW5ohDJnp5T9k/ZPp8Qr5M+LwhKrVHguJ12OR4PT0znVz5e5ER0IcDYg9LWN/Re6YP2TcJ8K4T7fxPPFWSsbeWSofkgYejW6X+O/RN6Zk28BSDcr30Y32Nyy+y9xhTQTYPNE5rf9rL+KocY9kGFV2EPxjhB4DgzvRTsk7yKZtvqHWx+JB2TyXxeIlZlXfv8AnsF3HZ9gdBxHxrS4biheKR7JHyZX92RlaTqeWoXrNRhvYxgsnd1DcHfK0WIdK6od8dSrvSTHb5nu7zSh7h9Y/NfSEM3Yli0wpGQYOyRxytD4nQX9HED81y/af2MUWBYLNj3Dz5W00HjqKWR+fKwn3mO3sL6g305pMi4V4uSXHU3Vyl/Uqq5gY3fVXKYXhB87K1J6re4b4SxriuSojwekFS6mDXS3kawNBJA1cR0Ky3sdDK+N1i5ji05TcXBtoRuu+4Ur8a4e4CxykoOHsXdXYqWtZVspXmOOK1iQbXvYut6rz9zHRuLHtLHM0LSLEeVlh1BJIsFE8EAt2vzUg3Q4t5lURMjIZbNf1SdyGPzNuTz1UmYbJQ5o3QKGWF76pWkpxc0jQqMmx0UDr6ocdE0+SCEDQHFwsnvJAsACSlYBY3NkmQvnyg2GVVm1G6pkHhcwFI2tc0WDAAOSQtcZni98uirtGaQ3Ntd1dGlttcG/9GFhYxWf17wx6ZQtVkYLiPeAWdisTWVbRp7gWMpNLjymoP2CD7gVjkq1B+74PuBWbrpHO8hwu0KPqpDqwnoox/wVI1KV/eQNPkp1l4fNllMTjvqFp3WK3sEJuYc0SODW6qu6Qko3jNpy4EouqwN0t3bAo1cdLC2me430CxBoFtt/Vj7oXzO/4xfu/wBDPv8AV/Kf9um4N4zrOEK9z4m9/STW76Am1+hHQr0OTins1x53teJUsUVS7V3fU7g6/q3defcDcJx8X4tUUctU+mEUXeZmNDr62tqu3d2K0Ny1mPSd5yBjb/muXR+b4epLP6vo/F/1XO6vzc8sOp9bjv8AuhxvtPwnDMJfhvCNIIswI77u8jGX3LRuT5lc92a8VYdwxi1dU4pJIG1EYaC1heSc1zdZHFvB+IcI1rY6rLNBICYp2Dwvty8j5LS4q4Dbw3w5QYoK51QastGQsy5btzb3Wbl1fLyv+V36fa/DJ287fDO2db/NzbZ75+jla+ZlTiVVMy/dyyve24toXEr0jBOP8Cxjh2LBOMKZ0jYQAyfKXB1tibag+a8vCVefDq5YXcfd7z4Z0e86ePTz3PHiz1Y9PxPF+zfC8GqabCsNbiM87bDM12h5EvdqLeS5vgTjR/COJymSMy0NTYSxt3BGzh6LlNULV6+VymU9aefp/Behj0c+j1Lc5lzu7evVWPdmUU5xiOkbPW37wRNicCXeYPh3WV2h8dYRxTwxSUtEZRUMnbLIx0ZAb4SCAeepXm2qNVvLucrLNT28/R/R7odLq4da55ZXG+t3f+zvezvjig4cpqzDsWikdR1Ls2Zjc1rixBHQhYfGk/Ds+LRO4ah7qkEVntylvjuevkueQud61uHhXuw+F9Lp93e7wtlvM36v+wQhC4vqhZ+NfuqX1b+a0Fn41+6pfVv5r09p/Hw/N8z4v/get/pv/Tl0IQv3L+AhCEIOq7MIYp+1Lh5k1svtYdY8yASPxA+S7T+kZUzP42wymcSIYqEPYOWZz3Zj/wDlHyXluEYlPg2M0eJ0366jmbMy+xLTe3xtb4r6H4x4boO2jhCgxzh6riZX07SGtkNhrYuhkt7pBtY/8CvJ1b4dWZ3h1x942R82L3f+jZUTEcQUtyYAYZAOQccwPzAHyXBs7FuPH1fcHBMmtjI6ojyDzvm/4L1rDqTD+wzs1qpqyqiqMYrCXBrf+mltZrGg65W7k+p5qdfqY54+OPu0wxsu6+feJIYqbirFoYbd1HWTNZboHldZ2I00FT2s4Z39j3TJZWA83Bht+ZPwXByyPmmfLI7NI9xc49STcn5rR4ax2o4Z4locZpWh0tHKH5Ts8bOb8QSPivRljbh4sS+9r3H9VUVnaJj0tS5xl9tkZ4uQacrR6AAKzwRwM/jWSsazGKHDPZQ03qjbPmvtrysvQuJ+BcO7Ui/ivgivp3VdQ0OrMPmcGOEltT/hd1vodwVwU3ZNxzHJkfw1VON92ljh8wVyx6kuExl1Y1cbv8XWQdgFbVS93T8WYPNJa+WPM429AVN2J076DEuNaYvu+moXR5m6atc8XHyV7sd7OuJuGuOW4xi+FCho2U0sZe+Rl7utbQEnkoOyQh3EXaA4EEGmlII1/wCkkXHPK2ZS3fDckmrp4q3VoJNyRdKkZ7jfQJV73F7Jx9/7ouz7/qD/AOEF5qvV6TDz2mdjeC02Dyxvxfh8mKWlc8NLhly6X6jKRy3C5WPst40mm7sYBO03tme9jW/O6+PndWyu93fbG4X/AOd+D/67D/vhd32i/wDv+pf+uo/zatDh7s3wrhHFcPreL8YgFa6eP2Wgp3EudJmGUk7nW2wA81ndon/v+pP+uo/zasc1dajo+0fsr4g4r40nxXD30Yp5Io2ASylrrtFjpYrl29g/FheM0+GsHNxmcQP/AMqg7aaqoi7TapsdRMxvcQ+FshA93oCuAdWVThZ1TO4dDI4/8VJvRbNvWcV+iez/ALNqjhIYrBiOK4xUNM4gN2xNcWhxPQZW2F9SSqfb3I9nFWFUTCRSwUIMTRsCXEEj4NaF5U/3umi9jAw3ti4Woaf22Gh4rwyPusspsKhv/EG19LkG+lijUu3j4TifDZdbV9lPGlHKYzgcstjbNC9j2nzGq0MF7HOKa+YHEKePCaRusk9S9vhHOzQdT62CDWqHvq/6NVG6pOd1LX93C47hokcAB8CQvWODYYKjsqwqGpIbTyYc1khzZbNLLHXlpfVeN9o/E2FNwfDuDeHZO9w3DCHSzjUSyC+x52JcSdiT5L2Lg+k9v7J8Lo3OLBUYa2Iute2Zlr/is1Y48dn3ZUAP7Yh//wCoP811/C+FYDgeC1dNwdJSVcjj3jgazPmfawzOGYgfBcKP6PFLlA/0im/7oP8AzLqeBezKi4BrqrERiktVJND3RMjGxMa29yfXQIrzvhDCsRb/AEgAOI4WRV5dNWFrTdjnFhylp5jU29F6rx1wthPFdDTUWMYvPh9Mx5kEcU7IhKbaZswN7cvVeUdoXHdOztaoMWwZ8dT9DxiJ0jTdkpLnF7Qelja4XpFQzhPte4bZF7QHSM8bQCG1FM+2uh/HkVf6s/0c4Ox7gUf/ANw1H/e4f8l1fBuA8OcEw1UOH4730NS4PdHPVRua1w5i1rE8/QLhKj+j0/vT7Pj8fd8u8pNR8nLQw/sDwajImxfF5qmNmrmsjbAwjzJJNlbq/VJNfRd4eosOl7f8eqqd0UtqJk0bmEOAe/KHkEc9PxW9xvwhgfFdRStxnGZ6RsDSY4GTsjaSfrWcNTyuvCOGOJP9CePXV9MwvpI5ZIZImH34S4iwPMiwI62Xt2M4Dwp2tYPBUwVrXSwg93UQkd5Hfdrmnl5H4JZol3HPf8j/AAL/AP5DUf8Ae4f8l2fCWH4Bwdg7sNosbZPTmQyt7+pjcWE2uBa2ml/ivOpP6PM3enu+IIjH1dSG/wDvK8zsPwHB6GesxTE5asxROcGkNgjzAG19z+Kl1fqT19HDz8EVXFnapjmG4G6FlJFUPkfOXXjjY435b3JIAH/BdvF2McFcOsbU8Q468uLbHvZ2U0Z62G/4rE7AeIcOw6qxDCqyaOnqK3u5IXPOUPLQQWX663A56rrOOexaPjLiaXGhjs9LJK1rTE+EStaALeHUWHOyt50TjbHkpOwyiDmSyUc5YNTnmluPUaFd3xo6lk7IcWdR2NI7DHGHf3Mnh312suLpexThDhpjq/iTGH1UDG3c2oe2nhIG97G59LrteOJKZ3ZHjD6IN9lOGPMORtm5MnhsOlrKX2s9PkKX3neS9w7BeFKKrpaziKtgFUaWTuqdhbmyvDcznW5nUAeq8Okdndf/AIL3v+j5jkEuA4pw46fuasymphIPiLXNDSW+bSAfit5cOeNlqau4y7WpMZdUUXDc0NEH3ZTOpM1230DnXve25Fla7WeHoMc4Ag4rdh5oMVgZG+ojc2z8rrBzHdS0nQ/5rl65vbLh+LPoPaMXq2tdljqIAx0cjb6OzW0+NrK1x3hvFmB8Dsl4h4zkqZK7LG7DcgOY7kZgdQ224Fr+qw67eRPLtQ1RgOJ1KnG9ynWHRaRCGm3vILSR7xKm26JQfRVdqxjfyKkaCN1KSkKiEPJB2QSo5H5WE9AiBk0Fjm0cCntmiEpeJBqLWsqzGh0VzqbX1UQjHdA2F0ntirgLMznd6PEeigMcbZLd8L77FRlg8OgThG0yWst60fTaeJgOjZ26eSwMcbM3ELd6D4BzWz3TQNtLrn8aOWvAB+oFzvtvHJrUQtQwj/AFMo6TxYbTuHNgUi6OV5ObrcdVFa1x0UjN02XR5PUXQQvBaWvBseRWpR1YqGWOkg3HVZ8gvA224UTQ5jw9h1GtxyUXbde0ObYqAxuB2uo6fEGSANlIa/ryKvC26jpjnpWDHdE9sVvVTpEauW0R0C22e430CxXXJsAtpmjG+gXzO/4xfvf0L+/1fyn/AHXpXYn/AM6a7/Vf/wBQWTxPw9xBUceYlLQ4bWkvq3OilZG4DyIdt8Vrdih/9qK7/Vf/ANQWlxD2uYrhPEFfh0FBSOZTSuja95dc26hcsfH5Mud17e3uMu5nxfq/ZsJlfGc/7Lvaw7uezzD6eue11cXxi/MuDPGVB2pf+7fA+t4v/CXmPEHEWI8SVrqzEZu8eAWsaBZrB0A5L07tT17OMDHnF/4S18ydSZ2fhHGdhn2PU7Pp9S7yuWVv4TevTK4O4Awo8PDiHiebu6RwzsjLyxuXkXEam/IBa9Dwl2fcVV8Ywaocx8JzSwMe4d430dr8QpuJKGp4n7IcKdg8Zm7psT3Qx6khrbEW5kHl5LlezLhvGDxlTVz6SempqXMZJJWFgOhGUX3vdWSY3HCY7n4uN63U7jo9bu+p3Fwzxtkxl1PX00zONuGoMM45GC4PC/LIyIRsc4uJc7zK7WPgfgzhDD4H8T1TZ6qYfXe4NvzytbrYdSqXE9VBRdvFDPUENiZ3GZx2F2kA/Mpva1w3jFZxDFiNLSzVdK6FsY7lpd3ZBNwQOt1nxxx8spN6r0faOt3H2btup1bhjlhu3erb+G1LjXhHhmDAGY3gGJRMjebNgMuYSa65L63HQpvBvAWHVGAu4i4lqDBh4u5jA7LmaPrOO9idgFzldwTxDh+BtxWqw+SKm3dc+KMX3c3ldeu4RWUtV2TUMzMLGMRwwtD6RtiXFpsRY8xvZOnjM87csdai993XW7Ts8en0evepvLVylm5Pw3+P9awIOGuz7i9ktHgMzqWuY0uaQXgm3PK73gvLcUw2owjFaigqm5Zqd5Y62x8x5Ear1fBuKKM175sI7PahlRA0kvja1jmjnqQPkvOeMcabxBxRU4g2jfRl4ax0Uh8QLRY381z7iY+Ev1/o9nwPqd1O5y6WW709f5spbL/t+LDQhC8T9kFn41+6pfVv5rQWfjX7ql9W/mvT2n8fD83zPi/+B63+m/8ATl0IQv3L+AhCEIBaWCcRYxw3WGqwfEaihlIs4xO0d94HQ/ELNXpnAXY9LxXw87HsTxZuE4cS7unZA5zw02LiSQGtvcfBc+pljjP3msZbfSo7tw48dAYvpWFpIt3jaVmb52t+C4vFcZxLHa51ZildPW1DtDJM/MQOg6DyC2uPeHcI4Zx6KiwXGG4vTOgEjpmuY4B5cQW+HTYA/Fcwp08cNeWMW28UIQvUsL4H4Kqux+bHqnGzHjDYXvLO/aAyQE5Yu73N9B11utZZzDW0k3w8zpaupoaltRSVEtNM3aSJ5Y4fEarfj7ReMomZG8T4oG/6wT+ap8KcM1vGHEdPg1A+KKona5wfMTkaGtubkAnl0VXHMInwHHq3Cal8ck1HKYXujJLSRzF7Gyl8Mrq8p71tNiPFGPYxGY8RxqvrGHdktQ5zT6i9iqtBiuIYX33sFdUUnftySdzIWZ29DbcKoha8ZwbCEIWkaOB19XhuJtqaGqmpZ2tIEkLyx1ulwurl7QOLZoe6fxFiBba2ktifiNVxlD+0j0K0SvndxP33XG+khq6l1Z7W6oldU5g/vnPJfmGoObe6fU4nXVeICuqKyearBaRPI8l4I2N99FWSLzLtYrcQrMSqjU11VLVTuABkleXOIG2pVdCEQ13vj0SFzmuD2OLXtNw4GxBSON5U540UdI26bj3iukjEcXEWJMaNAO/JH43UGIcU43jERjxHF62rjO7JpnOafhsspouE8QtOuVFAd0VtmM4hDE2NmIVkbGizWsneAB0AB0VLug0c/mtvgvB48b42wjDpI+8imqW940i92DxOv5WBSkiicbxV22J11uvtL/8ANQ1FdWVTMlRWVM7eksznj8Sur7TanCJuM56TA6Clo6Kh/QXp2BoleD4nG2+ug9FyDhYKHAgFg5TxyPhlbLE90cjdWvY4tcPQjVQQ6lw9FKtOd5bDOLuJI2ZGY/iTW9BUv/zVGtxXEcRFq3EKqqA1tNM54+RKqoVTZOSkp6iekmE1PNJBKNnxvLXD4jVMQg1zxhxI2PIOIMSA/wBZf/msbEMUrsSlBra2oqi3YzSuf8rlJM/JH5nZUc1yhtLmVp3FGPUrRDBjmIxRgABjKp4A/FUQVG8RGS8mY25BAysqqqum76sqZqqT7c0hefmSVI7FcRfAYH4jWOhLcpjNQ8tLelr2t5KJ7ydGiw8wka2PKcxObkqhitUdXPQzQ1NJNLT1EJzMlicWuaeoIVYOLDpz35qSJge8h0vdgDnzQd3D20cdQ0ohGMh9tM76aMvt621XPYpjuJY/We2YrXTVs5Fg+V17DoBsB5BYRJB98noVpRMa+lH2gN1LJ9G5lQDdPFlA1+tlM111ls4Ac0oaEmYWRmRSkJhICRzxzKhdJc6IJHPUMx/QuJTmtLt0lSP6u5EvCRsRMIABNxoU0RPbCB3ZuPJU2yPaLNcR8U4VEo/6R3zWtOW1nK7vGl0Z+ScAWuN2kX8lWFZMNpHfNOFdMPrEqm/WkxF2uHMX2XL44x/0gMp0yBdJ7fN1HyC5vHqgyYnmcBcsCxlNNY3baws5sLiYfqsCkVfDnhlHTk7FgBVp9r3Gy39GbyQaJZG52X6Jl1I03bZAy94QqzxYq2BYOCrvF23RUSngrJoNGuu37JUFrGyOaqOghmE0LXgbjbopMpPOyy6GfuAWu1adRbkrvt0YGzj8FnxrXkshoC1W+6PQLnH18hHgaGeupXQxn9E088oK+b8QxsmO37/9CrvPrflP/re4X4preE6+aroY4ZHyx92RKCRa9+SzsUxCXFcVqcQna1stTIZHBu1z0Xe8NYf2dT8P0TsYqwzEXt/St7148V+g06LqcX4G7P8AAqeKbEw+ljmOVhdO/wARtfl5Ly49DPLHXlNPu9T4x2nb91cr0c/mX1vXOvw9+3h3UdV0mPcb4jxFglLhdVDAyGlLSwsBzGzcoumcX02BRY+yLhmQz0hib7pc8l+txrr0WNPQ1lIwOqaSeFp2dJG5oPzXnvlhvGV9vG9v3fy+v1MdZT3N+rP9m7wtx1i/CYdFSOZNSvOYwS3LQeo6Lcre2LH6mohfDBTUzInZiwAu7zyJ6ei8/QmPX6mM1Kz1fg/ZdbqXq9Tpy5VqcR4/VcTYw7EqtkcczmNYRGCBYbLpcD7WMewehZSSthro4xZjpbh4HS43XDIUnVzxvlK69X4Z2vW6WPR6mEuOPH9HTcUcfYzxVGIKl0cFKNe5huA49XHmq/DPGWL8KyO9hla6GQ3fDKLsJ6+R9FgoT5ufl5b9rPh3azo/Z5054fg9Eru2XHamldFT0tLSSOFu9bdxHmL6Lz6aaWomfNM90kkji5znG5JPMpiEz6uWf3qvafDu27Lf2fCY7CEIXN7gs/Gv3VL6t/NaCz8a/dUvq3816e0/j4fm+Z8X/wAD1v8ATf8Apy6EIX7l/AQhCEHT8AYJgnEHExouIMS+jaLuHyd93jY/GC2wu7Tmfkvo2u4Y4Zd2SwYBLjXs+A93G1tYJmDvG58w8VspzH5r5LXv/FYH/wBFjC/+ppP98LxdxjfLG7dsL6rkK/stwzFOPKHAuD8YFfSSUxqKurdI2UU4DrfVtra1hzJW7iPCvY9wlWnCMZxLEKvEGWEzmOee7NueQWHpqVN/RrEHf8Q7e0ZYLfcu/wD4qhjeOdk8PEGIR4lwli765lRIKhxldcyZjmP6zmdfis3LLz8N31+CyTW2dx72T0OE8Mx8VcK4i/EcGc1r3teQ5zGHQPa4AXF9CCLj5pcL7OcCrOwmq4ukZUnFYYZ3tIlsy7Hlo8NugHNdH/ymcJw9ndfgWC8NY3Hh0tPPEwmLPGxzgbkuLjYAm/krHD//ANU6v/1eq/8AFcp55+Ml/Ekm/TU7K8E7PafGI6zhzEn1eMMo/wBNG6VzgwOy5jYtFtdN+aw+O8E7K34nj1TU43K3HiZXuhEzgBPbRtstt7c1hf0d3tbx/XscbF1A6w62e264ztMa5vaZxGCDf22Q/PUfmrOnfm2bpcv3eHX9nvZFS41w5/pNxTiDsOwktL2Ma4MLmDd7nnRrb7cyuhouB+yDimoOF4HjU8deQe7yzvu8gchI2zvQLR7UM5/o94QcPv7Jlo+8ybd3k0vblmy/FeEYCKw8SYZ7Dm9r9qi7nLvnziyuPl1Jc/LRdY2TTR434NruB+In4XWPbMxze9gnYLNlYTvbkb6EdfgsGn7v2qLvgTDnbnDTYltxe3wXuX9JTuP/AGeAy9/ee/XJ4P8AivCmfrG+o/NejpZ3Pp7rnlNZPoLGuxDhrCsQp8S+lJcNwOCJzqt802Z7nXGQNJGlxfqdrBSUHAHZpxbDLS8PYtOK2Jt8wlcXfeLHgZh6WVjt/wDaxwpgxjzey+1/prbX7s5L/G6837LfaD2nYL7Nmzd67Pb7GU5r+Vv+C+Xbcvdrt63rSjPwVi8XHJ4UbG19f3vdtI0YW2v3l/s5df8A5r0mo7O+zrg6CCLirGppa2VubKHub8QxgJA8yunPsY/pDi/7R9C6X+1n5eeVePdrPtI7UMY9pzXzM7u+3d5Blt5b/is7tNSe3W492TYNifDcmO8D4k6tjjBcYHP7wPtuGu0IcOhWZ2WcBYNxrg+LOrjUNq6V7WxOjlytAc02uLa6grof6PDqh0WO2cXUgfCB07yzr/G2X8Fa7CJo5cR4xMAHs4rWiLLtlzSW/BN69LJPVYHEXZxwpwvwPWyVWLsqOI4oA4R+0tbaTTRsY1I157ryZ+3xUuIzPqsYramZxfLLUSPc52pJLjzUW7dEV7HgHZNgfEHZxhmMx1c1DVTMbJUzSSZo2taSJCG8tAbXK0cG4G7LuIi/DMJxWoqK1jC7OJnB5tu4BzQ1w9Ap6Z7o/wCi6S0kE0jmm3QzWP4Feb9mb3xdpuBljiLzlptzBY66iqeO8G1+E8bP4aib7TUulayAgW70O1afLTfpYr0b/k04H4NoYJeL8bkfWSjRscjmNJ55WtBcQOpW1ibIG/0kcJMoF3YeS2/N+WQD8AVwHbUKgdpM5mzd37PF3N72y21t/NmQ4dDX9lPDnEWAy4nwPijp5I7/AKB8mdryPq6gOa71XIdnPAA41xqqhrJpaWkoQDMGDxucSQGi+g2Nzysup/o/ip+n8Yy5vZvZ2Z97Z83h+Nsyo0PE2J8J9pfEmI4RhM2I4ZLVysqGRRuLfC4m4cAQCCXb8iiLtDw52S41jf0Hh9biMVeS6Nkpe4NkcL3ALhY7HcC6884wwMcI8T1WETVLZmxWdHKB77HC7T5Hr6L1GmquyvtIr2wS0D8LxaqdZpDTC57z0c27HH1Gq8z7QOCJOCeJxhrZzVQTsEtPJbxOaSRYgfWBFtN9FZUym2GCHNBaQR1CFnVtFX4TUCOqpamjc4ZmtljdGSOoBGqkp68GzZrA/aW3JdskLdLJWuDhcEEeSjnfkZ5nQIK1RLd2mw0CgzAcwq80mZ5sdFGFUXmub9oJuveaC4PNVeSt0/6n4lRYWx6IspEIpmQEbBI6FoANlJuLJSLtCCDuh5pchA0e4fFS5fNGTzQFrsQHEbJzNiEgFistwZ3W2SZ3KTKgNRUdi7dObGpA3VOA1UQ0Nso6kXgcFPZRStzMLeqq3hnJQggg2OhCAtuIRZKhAi5/G/3h/IF0CxsYgHtwzuynINFnJvDldoGvkoIByyBXWtLed1FQAfR8H3Ap7qzhm8kU0NtL+ihT4zYEKiRwyvcFVGxHmp3u2J5aKuNj6qKjKanP3SDdVlYiOllKoo+qkBHMr0Rincl1cX6pn3R+S5IvAG4XWxfqWfdH5L5HxPjF/Q/0H/idb8p/9T0/7TH99v5r1/to/wCb2D/9cf8AcXkEH7TH99v5r17toI/0ewjY2mP+4vn9H+Fm/R/Ff/yXafnf+oZwjSYZwV2ejiqsp/aKyoaHM01FzZrW9L7kqbhztOh4pxZuC4xhcLI6u7IyCXtJ+y4Hr1UfCc+Hcc9nI4YqanuKymaGjbNobtcBzHIp3DHZd/o1jDMYxnEqd8VGTJGGAtF7e84np0Xrnn+78v7v1fmuv9l33H27fzt3x5/208+474ei4a4snoqe4p3NEsQJuWtdy+BBXOrpOPuIIeJOLp6ymN6ZjRFE61szW31+JJXNr5vV153x4fv/AIZ829p0/n/e1NhCELm+gEIQgEIQgEIQgFn41+6pfVv5rQWfjX7ql9W/mvT2n8fD83zPi/8Aget/pv8A05dCEL9y/gIQhCAXquO9ouB4j2IUXCkBqvpKCOBrs0No7scC6zr+S8qS5T0PyXPPDHLVv0als4b/AAXxhX8EcRR4rQtbIC3u5oXmzZYzu0nkeYPIheoYnxt2RcWTtxPHcFrocQc0d7ljcC8jkXMdZ3qddF4hY9D8kZT0PyWc+ljld71VmVnp6pxv2sYdW8KnhTg/DH4bhBGSR7wGFzL3LWtB0udyTcowrtFwOj7CarhKX2r6TlhnY20N47vkLh4r9CvKyCNwQkU+RjrUPO726Dgriyo4L4rp8Zp4hOGBzJYSbd5G7cX5HYjzC9XxXtI7KuI+8rcU4endiEjLZ5KQOde1hdzXa2XhCFc+ljnfL6mOVk09V7Pe12nwPAf9G+JqB2IYRYsjc1oe5jDuxzTo5vTmPNdDQ8e9kXC9Q7E8DwGd1e0HJlpnAtJ5Bz3Wb6heEoWb2+NuyZ10PG3GVfxxxG/FK1oiaG93BA03bEwagX5nW5PNc+z9Y31H5pErTZwPQrtMZjNRm3d2+r+0vi3D+GaLCqfGMLbieF4k58NREQC4WaCHAHQ/h5FcnR9o3ZpwhBNV4BhEwrJW2t3RY4+Re8mw9Fxna12lYRxzhuFQYVDWRPo5HveZ2BoILQNLEry0kk3OpXyb07OXe5/g6ys7QsXm45PFMMjWV4l7xungDQLBlvs5dF6XUdqHZvxpS08nF+AzMrYW5dIjIB1yvYQcvkV4QlU8WZlY9m4j7Y8Ew3hd/D/AeFvoIpGFhqHRiMRhwsS1upLiPrO/FU+xzj7BeCKHFI8W9pvVujdH3MWfRocDfXTcLyQrRp9Yh1UuOmpd3aeRwlnme3Z8jnD0JJSR66Jse5Tm+GUeaivUo+P8FHYl/oqfafpLuTH+q/R37zN71+i5Hg7FabA+MsMxOrz+z0s2eTI3M62UjQfFYmXmlU1pXb8fca0+NcdUmPYFLPEaWGMMfIzK4Pa5x26a/HVdY7tM4K4sw+CPjDBHiqhHvsjMjfPK5pDgD0P4rxskbJE0bev4j2rcPcP4FJhvA+FOpnyA/pnx921hP1rElzndL/iuW7Ou0X/Q2rq4a+nkrMPriHShpBe1/Nwvobi9x6LiCmPCaNvW6PH+yHCsWZjFFhuJe1xP72KLI/Ix3IgF2UeXILmsU7R24p2n4fxTX0RFFQPa1lOCHOEYvr0LruLvgAuHAAVjBcDrOLeJaXAsPdGyeoLvHISGts0kk25WCSFydt2x9pGBcb0GGUeDxzyOppHSvnmi7uwLbZQDqep5aBeUrq+N+z6v4CkpI8Sr6KearDnNjp3OLmtH1jcDS+nzXKXXScenPK3fs+OZ8JuxxapJKx8o8QANrXCrXRoqyQiyL2TibhFgeagbmCtQvcItBcXVfL5XVul/UD1KWkHeu+wl711vdTzIxu7vkmOqRbwtv6qe2jnSEW8P4pTKAzxCyrune7mB6BRkk7klXSbWHVDQNASUxskkrrBRK/BD3cIPN26l9E9mxNLTYm5TyNUh0Kk3F1jboA3zTrAJAUt7qASpEIFKjdq5PKaBdyKr1bGizgNToqwb0Kmq5M8uUbNUI0K6zhypOaVLlvrzSEEHVVD4SGvuRmI2CxMbucQuTrkC2Y9X6FY2N2+kBv7gWMo1jy06E/2fB9wKYqCg/d8H3Ap1r6JeSXUkYBuVGVJESdL2QhJjrlHxUbhYBTuaAEyQWF1FVXbobujzToxcLcjNSMFgnpGhOXWOdJfQ/wCS6+LSFn3R+S5Arr4v1LPuj8l8j4nxi/of6Eff635T/wCpmxSubnbG8t+0Gm3zSy1VRUNDZp5ZQ3YPeXW+a9b7HqqHEeHsUwacNcWHM0EXOR4sfxC82oMEln4whwUtJf7X3Dh5B2v4Ar5t6VmONxvL9j0viWOfX63S6uOr0ve/xn4s0Ceme2VolhcPdeLtPwKnnxLE8QaIp6yqqWjZjpHPHyXt3argsdRwG+WCMB1A9sosPq+6fwP4LjexjDG1XElZWyMDmUsOUXGmZx/yC6XoXHqTpy+q8PS+N9Lr9jn32WE3h61/xr3/AF285e1zHZXtLT0IskXScYVM3EfH1f7FC+fNKYYmRtzEhumgHndMqOAOKaakNRJg1R3QFzlyucB5gG64XpXdmPuR9np/EOlOnhl18pjllJdW/i55F1JDTy1FQynijc6Z7gxrANSei1f9D+IRXsojhFUKh7e8EZZrlva99gL9ViYZXiPV1O56PS+/nJ+dZDI5JATHG94G+VpKavXex6DLhGPRzMGeOQNIPIhpBXluH4TX4xVvp8PpJaqW2YsjFyB1XXLpaxxy/F87ofFMep3HW6WUknT173ztUQrFdh9XhlW6lrYHU87PejfuFXXCzXL6+GeOcmWN3KEIQjQWfjX7ql9W/mtBZ+NfuqX1b+a9Pafx8PzfM+L/AOB63+m/9OXQhC/cv4CEIQgD7p9F9l8LNo6Xs8wiqniiEcWHRSPcWA2AjBJXxofdPovsTDdex6m/7GH/AIC8Pd/5XbpfVkDtf7OiLjFYdf8A7q//AMqP+V/s6/8AtSH/ALq//wAq+VWQTd239DLsPqFL3E39zJ/sFWdrh+J8yu97ZeJMF4o4wpazAqhs9NHRNic5sZZZ4e8kWIHIhefJSCDYixHIpF6sMfCeMcrd3YQvTuxrgjh/jafFqfGY53y0rY3xd1MY/Ccwde2+oC6eLsAhqONsR76omouHKcsMFnAyy3YC4Bx2AN7k+nK655dfDHK41qYWzbwpC+nabse7Na6N1PSN7+VouXRYg57x5mzv+C8u7TuyGfgmD6Vw2okrcILg1/eD9JATtmtoWnbNprus49zhldcLenY8zQtnhXhfEeMOIIMIw1gMsl3Pkf7kTBu53kPxJAXv2G9hHBmEYeJMYlnrntF5JppzBGD5BpFh6krfU62PTuqY4Wvmg7JAvoLiHsL4dxbCpqrg7EMlVEDliNQJ4ZD9kuuS0nrf4KtxD2OcP4H2VVOLyQVTcYpqBsshNQS0S2Gbw7WuTovD1erM8txqYWPBkq9J7G+DMB41xTFKLGY53vp4WSw91MY7DMQ69t+S7OP+j/RT8cVofNU03D0LIzCxr7yyuI8Tcx2AI3tfW3JcrlITG14EVdpn/owvTO2Ts8wXhBmBx4BSVDZq6SSNwdK6UyEBuUAHnd3JdXwZ2AUUOHR1PFFRNNVSAONLA/IyLyLhq49bWHqpcppqY3bxCLdPk0e09F9DVXZFwFiQfS4bUmlqwNO4rO9c0jqxxN/wWdwv2HUHc1sfErJZpo6ksgkgmLGyRZQQ6w8yd+ixtvxeJNNwEPNgreLU0dDjmIUkIIip6mWJgJucrXkC59AqEhWkKzXUpea9W7KOAOHeL+HKqoxOOodVU9SY7xzFgy5QRoPitjCuxTCqWsraviGtcyibO9tNC2YMHd3OUvf1I5Cym108RumOX0dN2N8FYnQ3oGzQlwsyeCqMgv8AEkFeCcb4BVcF8QzYTVkSOAD4pGiwkjOzvLYi3UFJdpZph1NQImWHvHYLuuw2rwjDeNavFMYxGlomU1I4RuqJAzM5zgDa++gPzVDs87LsS7QZ5KuSY0OFxOyvqS3M57ubGDn5k6DzXr47GOznBo2R4i5z5HDR9XXmMu8wAWj5BW2T0zJb7eDcfcUy8Y8Z12LOLhC53d07D9SJujR8dz5krnF9F49/R94fxCgM/DldNRTlt4xJJ38L/idR6gn0XgONYPX8P4xU4XicBp6umdlkadvIg8wRqCtY2fRnKX6qKOS9x4B7DMPruHYMc4qqZo21EXfspo3902OMi4L373trYWstyi4B7HManGHYdXU81W7RrYcScZHHyudU8oTCvnJKvRe1LsrfwE6CuoqmSrwmpf3YdKB3kT7XDXEaEEA2Omy85Vl2lmi3UzHkRBoNtVDt6lPb7qqF3KEJEAhBQgfAzvJmt81qOAy2VCg1n+C0SLgrGTpigcNUrNrJcviCUCzysqLJLJ5aksoECEtkIGkXUc0ohZ/iOyfJI2JhcVnvkMj8x+S1IluiblLeybeyTUldHM7N0RqdyUWCVArTlNwsLGn/ANoDQe4FubLn8bP9ofyBZyax5bFD+74PuBTlV6H93wfcCsKzhLyY42N0+J3iumvb4UREtNwgtOFwoJz4bKUbKtKbuUVH5KVgsFGBdymaNF1xYpw0CLo5IXWOYXXxfqWfdH5LkCF18P6ln3R+S+P8U4xf0T9B/wCJ1vyn/wBdt2VYn9Hcd08TnZWVjHQH13H4hd/QcMiHtrr8QyWhbTipaeWd/hP5OXidHVPosQp6qMkPgkbI23UG6+g8d4xwmn4TrMQp62mfUuprsY2QF5cRoLb6Ery9tZcdZfT29/6Q9LrdLupn0Z/Fx8b/AMxk8K4+3i6v4rwqZ2aOR7jCDt3ZGT8wD8VU4Lpzwl2ZYtiNQ3u5y+V1yNTl8DfxH4rgezfGmYNxvSy1EgZDUB0MribAX2J+IC7ztY4ioH8KRYfh9XBMamcF4heDZo8RvbqbLeHUmWHzLzNvH3fYdTo93Ow6c/c6lwv/AByr9nNPTYDwFiPFUsQmqyJHXO4a3l5XO6xsH7W8dfj8PtwglpJpQx8TI8paCbeE76X5p/ZtxdhtNhlTw3jjmx0lSXd3I/Rvi0c1x5dQVvUXZ/wdgmItxafGhLTQnvI45JWZQRtcjV1lMfLLHG9O6/F17idDodx3GPfdO5ZZfdut+vpJ+Clx3gVNh3aVgNfTMEft07TK0CwL2uHi+IP4LoO0XjqbhGamgoaWKWrqWFxklvZrQbct9VwvEHGMHEvaPhU8LsmH0U7GxvfpfxAueb7DQfJSdsVdS1+P4e+lqYqhjaZwJjeHAHNsbJl1JjjncL9Todhl1uv2nR7zG2eN3/fUv9m/2OvfV4bj8jgM80wcRsLlpP8AxUkrIOyngXPGwT4vXnL3gF257X3+y3kOazOyPEqKhwbGW1VXDA57hlEjw2/hPVJwfxRh/EnD83C3E0zA1rP0FRI62g28R2c3keYWsMp8vGb9+9OHedDqfbetn429LHLHyk+s1/fTzCpqJaqpkqaiR0k0ri973HUk80wG40XTUuEwYTxNNRzTxVQjcwxyxEOD4y4ZnNtfxZeW+6p8R9yZmvYLOc8luYWeWWbbMN977/ivn5dO68rfb910e9wy6mPS6eP7tnqsVCELk+mFn41+6pfVv5rQWfjX7ql9W/mvT2n8fD83zPi/+B63+m/9OXQhC/cv4CEIQgD7p9F9ncM1MdF2bYTVSgmODDIpHZRc2bECbfJfGJ90+i+xMMBPY/SgAknBmgAf9QvD3nvTt0vq5Qf0guCi0H2fFNdf2Rv/AJkf/SC4K/h8U/7o3/zL5rZQ1ndt/qdTsP8AoXf5J3sNX/B1P/4Lv8lr7N0/xPPJPjlZFiPEOI1sAcIqmpklYHCxyucSLj4qinyRSQuyyxPjJF7PaWn8Uxeuak9OX1ey/wBG/wD51Y1/qTP/ABFof0iOKKyGroOG6eZ0VPJD7VUBjiO8u4ta0+QsTbrbos7+jf8A86sa/wBTZ/4i0P6Q/C9ZNWUPElNC6amjh9lqCwEmOzi5rj5HMRf0XhuvtH7ztN+Hp4lQVlThddFWUEz6WpicHMliOVzT8F9e4fPHx72VxTVcTLYth/6Rg1AcWkG3o4XXyFQ0dTiddFR0ED6qpmcGxxRDM5xPkF9eYdTx8B9lkUNZKy2FYeTK4aAuDSTb1cbJ3WvWuTp797cD/RxwmOHAsYxR7QZ5KkUubmGsaCR83Lgu2/iirxrj6rwszO9gwsiGOEE5S+wLnEczc29Au8/o5YsyfAcZwp7gJ46kVVuZD2hpPwLfxXB9t/DFXgvH9XijoXewYmRNHMB4Q+wDmk8jcX9Cph/HvkuX3PTkuD+Ja3hHiakxOhkcwMkaJYwbNljJ8TXDmLfIr6k7T3B/ZPj7mm7XUZIPlovlrhHhut4s4kpMMoInSF8jTK8C7YmA+JzjyFvmvqbtQaGdk+PsaLNbRuAHloufc6+Z6Onvxrx7+jif/bjFf+z/AP8Acaul7euOcYwGsw7BsIrZKHv4nVE0kJs9wzZWtB5DQ3tuua/o4f8APjFf+z//ANxqZ/SN/wCf2Hf9nN/8R64WfvLPuqPZBJXcVdqlBJi9dVYg3D45KtgqJnSBrwAAQCdNSD8Au9/pCcT1mFYHh+C0UzofpEvfO5jrOMbLeD0Jdr6WXmvYhisWGdqVC2Z4Yytjkpbn7ThdvzLbfFelf0hOGKzFMEw7GaKF030c57J2saXOEb7HNpyBbr6pfvEv7r5/oq59JOyRjnRyMN2yMOVzT1BGt19d9m/EUvFHAWHYjUODqggxSuH1nMcWk/G1/ivjuGN9ROyGFplkecrWMGZzj0AG6+vuy7h2o4X7O8Mw2sZ3dXZ00zPsue4uy+oBA+CZwwtfNvEf/OrF/wDXZ/8AxHLMc26l4kri3i/GWubtXz7f9Y5Z/wBIMts5NG49/wD6Pn7nxv8A1mP/AHFxvbXjM9dx9PQSSE0uHxsZHHfw3c0Oc63XW3wXV/0dJxPg+OkAi1TH/uLy/tlkee1nG2Fxyh8en/8ArapJurlf3XT9hXEctPx8cGjlcaWugkJj+qHsGYOA62uFpf0lqRjarAKwD9I9k8RPkC1w/Mrj+wz/AN7uG9O5n/8ADK7j+kuP0HDgG+aot8mK61l6SXeL0ilZDwF2TNdTxsAwvDjLY7OkDMxJ9XH8V8jYpiVZjWITV+JTvqqqZxc+SQ5jc9Og8l9cUz4OPeyVogkZbFMOMVzs2QsykH0cPwXyPieG1mDYjLQYlTvpKqFxa+OQWNx06jzTBM9/R6d2CcVVtBxrHgDqhzsPxBj8sTiS1kjW5g5o5XAIPVbf9JPBoWvwbGIwGzTCSklI+sAMzfld3zWH2CcK11dxrHj7qdzcPw9j7TOFmvkc3KGt62BJPT4ro/6QuIQV2KcO8OxzRsmdIZpHOOkYeQxpPl7x+CX7y/5fbveGMSwftG7L2UEdTYTUQpKqKN4EkDsuU6ctrg7FeV4l/R/4mwmsZWYFidJXmCRskWe8EgLTcHW7b6DmsvH+ybins8webiSnxmK9KWhzqF0jJGtJtmvYaDmmcPdu3F2EzRNr54sXpQQHMnYGyEeT28+lwU1foWzjJn8ecWcf1FO7AeLzJEwvEndSUrI8xadC1zRqNeRtquDGpX1h2xYdR4t2UYlU1ENpKSJtVA5ws6N4I08rgkFfP/AvZpjXHr6h+Hugp6WmcGSVE5IbmIvlAAuTbXy0VxvpnLG7chzunt91dDxrwJi/AmJxUmJiJ7Z2l0M8LiWSAb7gEEX1B6rnW7LX0Y4Ki6OSEAjdBSKiWmf3dQ08lqXWNexWnBKJYgeY0Kxk3jUp3RbW6EvJZaKmuKcmkIETJZWwsuTryCJp2Qt6u5BZsj3SPzONyrIzaJJXSvu74DogbJuyXkujBdylG6QaBKLAIFshF0l7lAqwMaH9f/kC39lgY0T7f/IFnJrHls0ELzh8G3uDmrPcv8vmkw793QfcCs81qT0xb7VnQPc21h80jaeRvT5qyhXRtHkcG2NvmoHwvJF7a+atpLZpWN6lTUNohQVDTcs+RQYZG7scPgtgbJVZlpdbYmUpQ09R81sFodoQCmGniO8bfkr51PFlZHeXzXWxaQx/dH5LENHCdgR6FZLsbxBjy1tS4BpsNBt8ly6nZdTvvWFk0/Tfo/8AGej8Iyzy60t8tcf0doksPJcX9O4l/En5D/JH07iX8UfkP8lw/UHcfzR+r/bjsf5Mv7f+u00QABsuL+ncS/ij8h/kj6dxL+KPyH+SfqDuP5oftx2P8mX9v/XafFFh5fJcX9O4l/FH5D/JH07iX8SfkP8AJP1B3H80T9t+wv8Aky/t/wCu10SAAbLi/p3Ev4o/If5I+ncS/iT8h/kr+oO4/mi/tx2P8mX9v/XaEA//AMIsFxf07iX8SfkP8kfTuJfxR+Q/yU/UHcfzQ/bjsf5Mv7f+u0sBtolNybk3K4r6dxL+KPyH+SPp3Ev4o/If5K/qDuP5oftv2P8AJl/b/wBdqhcV9O4l/FH5D/JH07iX8SfkP8lP1B3H80X9uOx/ky/t/wCu1VPFIXVGHyRx2LiRubc1y307iX8UfkP8lbwzFa2qxBkU05fG4G4sOi3h8F6/QynUtnr28ne/pj2fc9vn0McMt5Sz6fX/AHH0RVdGf7SPoiq6M/2lvIXu+0ZP5xqMH6IqujP9pH0RVdGf7S3kJ9oyNMA4TVkEZWf7S+keHe1nhjDOF8LoKiSrE1NSxRSZadxGZrQDY89QvC9iluuXUzvU5axvjw+hB2z8Jf31b/3ZyX/ln4S/va3/ALs5fPaS65eEb+ZXQdr2MUvGnFtLiOEF76eKjbA4zN7shwe87HlZwXBfRFV0Z/tLdui69GHVyxnjHO+7t6D/AEd6Kal4oxl0gbZ1GwCxv/0i7rivtHm4Q4/+jq2j9rweekjechHeRPJcCbH3gQBp5adFy3YT/wA48V/1Vn++qPbaf/byL/U4/wDecuGWXnnbXSXWLvqftO4Co43VFM5sMrtS2OiLXn5D/ivLe07tIxHjWn+icMpjRYQHBz+8kHeVBG2a2gaOlzc79Fx90XWsP3b5M3Pc0j4VxLG+D+IIcWw10Qlju17Hu8MrDu13kfwIC99oO2XhXEcNBxmKWgfYd5DLF3zL+RaDf4gLwZV64/1N/wAPzV6uXn7vK9PnT2PiHtt4fwnC5aXg/DxJVSAhsvcCCGM/aI3cR0tbzUNZ2vcPY12ZS4Pic9YcUqcOMEzhTEtMxZYm+2rtV4Yhed7/AJWLt+x7izDOCeJa6vxh0rYZ6TuW9zGZDmzg7DyCb2v8VYbxrxVR4hhDpXQQ0ghcZYzGc2dx2PkQuKQtb97T5GOtIImzwTMlieY5I3BzHtdYtI1BB6r3/g/t6oZcPjpeKoJYKpgDTVQMzxy+ZaNWnra4/JeDoUt2ToYx9HV3bBwBhTX1eG0/tlY4XAp6MRucfN7gLfisbhHt2pT9JycUGWF8lQH0sdPCZGxx5QMtx0Ivc7kleFIUX5UXuLKimxTjHFq/Drmkqqp80WcZTlcb6jlqSsjuJOg+asoWvKxm9vi9P7G+PsE4Hw3FYMZfUNfVTskj7mEyCwZY3ttquM7RMWpeJuPsTxfDi91LUuYWGRuR2jGg3B21BWGhTf1X5GOtOk7Msco+E+PqPF8TMjaWGOVrjGzO67mEDQea6Xtm46wbjmLB24M+d5o3SmXvoTHo4Nta++xXmyE3d7PkY607Ps47TcT4DkfSyQCuwmZ2d9OX5XRu5uYeR6g6HyXsUfbB2fYtE2TEC5j2jRlXRF5afIgOC+akJfazoyPoPH+3vAsPonQ8PUU1fUWswyM7mBnn1PoAPULwDGsSxLiDGKjFMSn7+rqHZnvJt6ADkANAOSjQkukvRxr2/gzt1pBg0eG8XUszpY2d0aqJglbM21vG3cG29rg+S1Y+Mex2hqRiFNQ0IqWHM0x4acwPUDLYFfPaFF+Vi9L7UO1uTjHDzguEUslLhjnB0skxAknsbgWF8rb68ydNla7GO0fCuD8MrcIxvPBFNP7RFOxheAS0NLXAaj3RY+q8qQrv1pPk4729D7ZuPKDjfEMOgwhr30lAJHGaRuTvHPyjQHWwDea8y7twGtvmp3E5tFKxoc9oNtV0x4eTPGeVVe6eA020KO5eTsPmtOVjXHKbaaqJ0bWjMOSu2JjFI08nQfNHs8l9h81ZMoFt/kjvLnQE+gVZVvZpL7D5qSGOaJ9wBbmLqbOb+64jrZF3X0Y75JoTCQWvqpLGwNtCqoz7FhUsFQ3uAHHUaELOo1Npb/pAz6xF1HUd8wWY0et1KyaNzyRrbS6eXNcQd01D2zXUFSfE4NJP+JR+xzfZHzWq6S4Isoj0WksZ3sst9h80vsspOw+av2TmtuEZZ3s0vQfNBppOg+a0cgRkHRBndxIBaw+aQQvHT5rQ7sJDGOgRVAxP6D5rn8bjd9IcvcC67IOi57G2/wBobfUCzk1jy08P/d8H3ArJKr0H7vg+4FY5LccryRCLIO17Ig5ojLRUsLnAAam6a42aXW2WfI8vddFkasmKRM0aC6yhdi0h91jQs7dLbRZbWn4jUOHvZfRMFZODfvXXUFinAILP0lUXHiGnksxxu9x6klWrKp9Y+q+r8O5yefrcRZoMOrcUq20tBSTVdQ/3YoWF7j8Ar+LcJcQ4DA2fFsFrqGFxsJJoSG/PYL1iDE3dlPYhg+JYPFCMc4jOd1W9gcY2FpcAL9G2AG1ySuWw3tv4pgo62jxg0+PU1XEY+6rIxZpPPwgXHkfwXqnW6udt6eM1Lrlm4Y4+sr7ecJF6XgfZzgUPCFJxNxpj78GpcRcfY4IIs8jx12Jtztba1ypqnsjpxxbw5BQ4ya7h7iB7m09fG0Z2kNLi0ja+m/r0W/tfTls2z8rJ5clAubDU9AvTeKezfhjhTF6DDq3i5plmrDHVBrWl1JBlJD3AfW2+a6ntk4Y4Uw3AsKqKLE4qHEKakj9jpYIQz2ppe28pIG9iSs3u8d4zH6r8m6tv0eMYpgWKYJVx0uJ4fUUdRKwSMjlYQ5zTsQFoDgXin2iOn/0exETSRGdrDCQ4xjd1umoXtzuHYv8Al3wOLiTiebEamnoo6mibPCxnevzP/R2aLaZc197q3OyqxrtfxOHB+O6ts1PQTmaOOBjhS2kYO5AcPO997hee97l6k/Df1dPkx8zkWNjpZIvSeEuzXD8R4Wk4r4tx04LhD5TFC4NDpJ3XIJ58wdgSbFNxbsyw6p4iwSg4R4kpsZixgnIHECSnaNS94H1QAeQNxay9f2rp7s//AI5fKy1t5whezM7JOBqnH2cO03HT342yTJLEYRlcR7zW8swF9LlZ1J2Owzca8QUVTjHseA4Bl9or5WgON2B1gNtjqfTqpO86X/7F+Tk83w/B8RxYzDD6GoqzAzvJe5jL8jeptsFT6L6P7N+GuF8LpuJsQ4W4kOMU0mHugljkZkkicA4g7DQi/JeecPdmWBw8E0PEfGvEL8Fp8QH9UhiYHPe22jjoTtrYDa191zx7zHd39Nfn/wALejdR5ktDBP3tF6O/Jd7xt2X4XwzwFTcS4bj5xaGrqWxxOZGGsdG7NY33uLWK4DB3BmKRuJsLO/JdM+rj1ejllj/VjxuGU26tF1VfWxtNi8X6AKpLibrERjKOpX57T07aZkDTqQPVN9ojBALgL9SsR1ZIRuOt7aqAylx8Rurpd1vS1MI0MgunOqoWtzGQfDVc/wB4QDa58l3mI4bwxwS+DDsZw+uxrFzDHNUNjqvZoIM7Q4MFgXONjqUuoSW+2JHOyQaED1KkBJ2Gi024Vw9xJg2JV3DcNbhuI4XAaqWhqZROyaEEBzmPsCHC+xC56kxFr25ZSA7ryKc8LdxfsUmqGua4Xa4EeRSqHKamrKqic59LUzU7nCzjFIWEjobJtTVVNZKJKmeWeQC2aR5cbdLlR3F7XF+iCRzICgRGqXQIuOoCpomqgrR/U3/D81ZVavIFE+5A23PmpW+n96MlCEoaXEBoJJNgBqSVxfVIhdpPgPD3CMccXE3teJYw5oe/DaOQRMpwRcCWUgnNbk0aKOGfgLF5RTS4fifDzn2DaplV7XEw9XscAbdbFE8nHoWpxJw/WcMYxLh9YWPcGiSKWM5mTRnVr2nmD/mut4npOCeFsfkwiXh/E6t8McTzM3EsgcXxtf7uQ296yHl+Dz5C7bDsN4N4qq2YbhzMSwPE6g5KY1M7ainkfyY4gBzSToD1WXQzcO4ZHJRY5w/X1WIQyvZI+KvELRY2y5cp1HXmh5OdQvQcYg4HwnC8FrTw5icoxWlNSGjE7d3Z5blvk12us1mF4Fj/AA9j9bg9DVUVVhjIamOGWp74uhvll1sNiQUTycghKGue4NaMzjoAOZ5LtO0Lg6h4Xiwt+HSuma9jqesJdmyVUeXOB0HiGnki2zhxSFv8E4FBxDxXT0lY5zKGNr6ire02LYmNLna8tgPisOZ0TqiR0DSyFziWNcbkNvoL+iLv6GISIBB2IPoUUqEJLi9ri/S+qIVC1eFqGnxPjDB6CqZ3lPVVsMMrQSMzXPAIuNtCq+NU0VHj+IU0LcsUNTLGwXvZoeQB8ghv2pISXF7XF+l0qAEZfcpbd3z1Clgbdp9UkzDpodrLtjw+fn/EpZHm4c3W6a973MIA+Sd4u7FhqPxS3kcLZdfRBGZ2mHW9xZK6dmUG53UTqeUNymM38kGCUtDe7IcE3U1EpnF26EoM7e8IAOyjMErrfoyMu91Hez72U3U/dTioAJuNPVQNcGX10JTCd72umF1juqz5J21Bbe3PqlNS++6r5gUCxFkTayKl9wST5qeOoY42Jt6rPBtzulDgLlE21nOjaLnbqntsRpa3kslsz7WDjZOE00bLMksiNWyMqxjWVH965TMnlLRmkcT6qq0iE0hUO8efru+aMx6n5oLpsuexsj6Q3HuBaJOu657G/wB4fyBZyax5dBQfu+D7gVhV6D93wfcCsLf0cryEnJBQiIKqTKzKDuqVrqWou6c/gnRQOeegUt03IiyFODD0KuspWg8yp2wNAWfJvxZ4iJGyUQOHJaIiF04xttsp5L4swQm+yzyPGfUroe6F1z8n6x/3ivq/Dbu5PN3E1I9c4V4h4Y4z7OafgjizEfompw9+agrnWy21sCToLAkWNri2qWbgLs34Twusqsc4vZjsr4i2npsOe0PzciMpdr66dV4+l0GwA9F7vs1lvjlZL705fN9e4+ieE+PIMe7PMJwzD+JMKwDGcMYIJY8Uha+OVgFmlpJHK23O+ip13GjY+0nhOiruMMKxWio6h1RUS09KIIoH929vvgkEHN8Oa8BOosQCjlZcvsOG7ZeWvn3Tou0GrhxDtCx+qpZ2VEE1ZI6ORjszXjkQeYXofaVPw9xfwZgnEFFxHSNq8Mooqd+HOP6V7i5odzuLa8joF40lXoy6G/HV+6xOprf9XvfEvE+B1H9IThPFIMXo5KCmpAyaobKDGw/pdCeW4+apcC8R4NRdt/GWI1OK0kFFVRTiCeSUBkl5GkZTz0F14ghcvsck1v6a/vtfnXe3sfD9Xw5x/wBllDwhimOw4DieETufTyzkCOZpzdSBs4g630B1uoaA8IdlPH2A11BxCcclZnjxF0DA6ONrhlzNIO4uPDc6BeRfC6Xkr9l5nldX6fmfN+uvb3ei4S4JwvtCp+Lzx7hz6B1Z7XDTh47zO43DXG+jQTe5A03Vh/FnDGN8T8dcK4jjENLQY7JHJS4ixwMeYRsBBO27RvpoQvn+w6BLyss3s/L72VrXztcR9BcE4Xwj2eUnEEU3GuG4hiNdROYBG8NjawB1gDcguJO11jM/0d7U+znh6hqeJKTAcawGH2d0dW4BkjcrW5hci9w0HTW9wV4tpbYJNDuAU+yXfn5Xf4p871rXp7Rx5U8MYd2H4dw3gfEFNi01JXhzy14zPN3lzg37NzovHqRxbUtINiAVCN7qWnNpwfVb+V8ro5Te+Wbl55Srz39TqeaYdrpDqlyE6L4T0movqpRA5PFN5qbVXIuCDzXdzcR8McXw054pjrsNxaCFkBxKia2VkzWizTJEbG4HNpXI0sbYK6CaWITxRyNc+N2oe0EEt+IuF2PGHANUcWlxbhmgkr+H6499SyUbDI2IEXMbgLlpabixUumptRreA5jhVRi3DuMUnEFDTNzTmmzRzws6vidrb0upMOhoOFuDaLiCroKfEsTxaWVtDFUguggjjOV0rm/XcXGwB0FrrU7PsGxThPFpuKMapZ8OwqkpZmv9paY/anOYWtia06vuSDtYWVKGhn407OsKpMKjNRivDzpo5aNusksEjs4exv1srrgga6rO119VnBONqfGsQiw3iTC8OfSVDxH39JSsp5oLmwc0ste19je4XQcO4JRYdjPFeGY1C2qhwykdI4t0JMcjSMp3bmFgbciVxPDvA2NVeLRS4hQ1OFYZTSNlq62riMUcTGkF2rrXdpYAa6rssMxxnEeKce4tEHNhqsNmfGHbhudgbf4AJf6E/qOH+JI8cx2lwbFcIwv6OrpBThkFI2J8Bdo1zHjxXBtuTdQ4nxC3hrF58IwnDcNfSUMhhe6qpGyyVLm6Oc9x1FzfRtrCyxOFf+eeC/6/B/4jU3ibTi3F/wDXJv8AfKmvaeXp0GJcK0+I8eYXRYY32Okxmniq2tvcQNc0l4HW2V1vgqtRxbDQ1D6fAsIw2CgjdlYailbPLKB9Z7n3Nz0FrLaqMXhwLiXgqvqQTTx4TE2Ww1yO7xjj8nE/Bc1inBmNUFaY4KGor6V5vT1VLGZY5mfVILb8uW4Sf1X8lrFKeixvhV/EFFRRUFXSTtp62CG4ifnHgkYPq3III25rVxn6RwPhrD6rhrB6Sqw19JHJNWGhZVufMf1gkLgS2x0tpZUKynfwtwNUYXiAMWJ4vURSuprjPBDHcgvHIuJ0HkosVwbiLhmrNTwvPic9FO1slNWUYcRK0gaOy3FwbggrN4bxvuOS4jxLC8WfR1lDQMoKt0RFbDCzJCZQ4+KMX0uLXHIq92aUsNb2mYDDO0Oj9pD7HYlrS4fiAtHtAbNJgmA1eNU8dNxLO2U1bRGI5HxAju3ytFrPOvIEhcng2Kz4HjlFilNYzUczZmg7GxvY+RGnxWH0J7xMxSsnxHGKysqXF01RO+R5PMlxK2MK4Ex/GcNir6GCmfTy3yl9ZDGdDY+FzgRqOi1OIuEZcYqJuIOE4H4nhVY4zOhpxnmo3u1dHIwaixvYgWIWVhnAPEuKzBseDVEEY1fUVUZhhjA3LnuAA/NDfpPxbg/FOHYPhbuIGxGkp2OpKN7JopCGjxFt2Ekgeey6XtG4L4mxfjiprcPwKtq6WSCnySxR3a60LAbG/UELm+M8ToBR4Zw5g84qcPweN4dUtFhUTvN5Hjy2A9Fe7VKqoj7RaxjKiZjRBTeFsjgB+gZyBRmb+iThzgXF8GxyixniOD6EwugnZUSzVb2sc7KcwYxt7ucbW0C5PG8RGMcRYhiQZ3YrKmScNP1Q5xNvxVKSSSV2aV75HdXuLj8ymjcI3Jd7rruMv+avBP8A2W//AMZyr9nmJQ4dxrSR1Z/qWIB1BUgmwMcoy6+hIPwVnjL/AJq8E/8AZTv/ABnLkLkatJBGoI5FExm8XX8K8P8Asnac2gxFv6LBppaiqvtkgBcfnlb/ALSnpK6XingfimKpcX1dPVNxxl97OcWTAfBzT8F0fEs0A4LreM4pG+08T0dPQFo3bKCfaT8RE35lcZ2eVsVJxvRw1RtSYi19BUfclaWfgS0/BGeZtbwT+xOzPHcXPhqMVkbhNOb65PfmI+AaFxy7TtDi+gxg/CLZGvODUt6hzdnVEpzvPyyhcWjWPv29AoRgtB2VYdjtbhsFbXwVtRSU8MjbRyk5XB8trFwYA6zeeboqNBxpBiddFRcR4NhdRhs7xG91NRsp5oATbPG9gBuN7G91qYfgMnEPYtRU9G9r8SixWolgpiQHVADG52svu4AggbmxsufwXgXG67E2Csw+ow6iheHVVXVxGKKBgPiLi62tthuUZ9e9tOg4Kp8M4k4jGLRyV9Fw4LuhhuHVTnOtE3TUA7utyBUY4uxd0uSThLCZKEmxpBhOVuXoHgZwfO63sO4wON8XcXU+G4lJhU2OZPo6p7zurPiNmMLvq52gj42WT3nax9Iexd9xR7RfLlzygeub3bed7In5pGYFS4N2r8JzYeyWPD8TnpK2nimvniDpAHRuJ3LXAj0sqVHg1FiXGvElbiokdhmEvqKuojjdldL+lLWxg8sziNel1bikqx2tcM01dj8+O1NNV0zZpXyGRsUhkBdGx1zmA015m6bhNTTTcV8XYBV1DKVmNmanimkNmMmbMXx5jyBItfzRfbN/0+qWyZIsB4fjor2FIcPY5tuhefGT53UPFmE0EVFhWPYTCabD8XjefZi4u9nmY60jATu25BHOxUEnA/FMeI+xO4fxE1Ga1mwOcD5hw8JHney0eM3RYVgeB8KsmjnqcMEs9Y+Jwcxs8rgTGCN8oAB80a9bmnKQzCMEWJvropDVAcnKsd0hF+S6S+nz+r9+rXtQ+wUe1/4T81UDfVLZXbms+1f4T80hrLD3dPVVjoNVA999tlRYlq3S6Xs3yULni1gbqHU+ScANjdAp8024TiwA32SHKECAjklG6TN5Jc2iBzQNynC51NrKMXPNSsY47IGOAGrQQEheTYFOlzjrZQEm6CUi4GystiIG6qw+KRrepWs2IW1VRV7tGQK0Yx0Sd2ByRVXIFz2NsH0h/IF1WUdFz+NsH0ht9QLOTWPLTof3fB9wKwN1BRfu+D7gU/JbjleQUnK6XdRzuyxHXdBXaO9mLuV1cY0AqvC2zArLTqudrtilCfZRtKeCstnjZFkgS3RSgLmJP1r/ALx/NdOHLmJP1r/vH819b4Z97J5O54hiEIX2niCEIQCEIQCEIQCEIQCEIQCEIQClpwXTgBRK1h4vXMHkfyXLr/w8vyax5i7FCTuFabCBupA0DkluvzT2m5AiwTtEWHVRSW8lZosTr8McXUFfVUbjuYJnR3+RVfYE30Gq6h/BTKJkLMZ4hwzCauaNsopZRJI9jXC7c+RpDSQb2KWz6rNubxCurcUkEldWVFZINnTyukI+ZKzyHwyiSJ7o5G6hzCWkehGy3cawGtwLFBQ1TWSPka2SGSB2dkzHe65hG4K1ZeBm0QbFjXEOE4RWOaHeyTOfJKy+wfkaQw+RKnpdVyFbi2JYkGtr8Rq6sM0aJ53SAelyVYwiqmhdLDHI9jZmFrg1xGZvQ9QrHEfDFdw/JD7R3MsFSzvKepgkEkM7erXD8QbELTw/gyamoqavxjGsOwFtUzvKeOrL3zSMOz+7YCQ08ibXV3GLKgje+J7ZGPcx7SHNc02II2IPVD3uke6SRznucbuc43JPUlaUnBWMthqKunrqOtw+OllqW1tO8vjf3dszNgWv8Q0cAseqwSvp+GqDGHztdDXTSwsjF8zTHa5PKxurqVm7ieSeSUMEsr5O7bkYHOJyt6DoFNS4tW0MbmUuIVNMx3vNincwH1AKoYFgFbj1ZPTwSsidDSy1RMl7FsbbkacyruDcEYrjuEHEKSSBsLKgwTGV+QQgMz9492wbbTrdW4ycpMreFeSoaSXyS5nO1Jc65KcMcraGjkZh+J1NLn3bBM6O/wAitMdnrqunmdhHEOGYvPTsMj6anzskLRuWh4Ga3ksfCOGK7HK32XD2iSUNMjnPcGMjYPee9x0DR1Usmq3hbMptkyyyTyullkdJI83c97i5xPmTqrOIYXWYU+nbWQmE1MDKmK7gc0bxdrtDzHxXQN4Hgq3+zYZxVgmIYgdGUsb3sMh+yx72hrj5XF1P2jQTR4ngFO+J7Z24JRRujLSHB2UjLbrfRed9fy96cnS1lTQ1AnpKmammG0kMhY4fEKzXY7i+KRiPEMVrqxg+rPUPkHyJW2/gdtBlixviPCMHrC0ONJM58ksd9g8MBDD5ErLx3hyuwB0DqgwT0tU0vp6umkEkM4G+V3UcwbEIblrJ3Uk9RNVTGWomkmkIAL5HFzjYWGp8lGumpeCpBhtPXYxjGH4FDVN7ynZVl7pZWfbEbASG+ZtdFtkcyhdTNwDiDKOpr6atoq7DYKaSpbW0zy+N+QtBj2u1/iBs4bKHDOCq/FcFp8WiqKWKhfJKyaed5YymEeW5e63POLAXJRPKMCSommZGyWWSRkTcsbXOJDBe9gOQv0Ua6ifgnvMNqqzBsdw3HBRM72ohps7JY2Dd+V4GZo5kLJwPAK/iGsfT0LIwImGWaaV4jihYN3vcdAP/AEEJYouqJ307Kd00joY3FzIy45Wk7kDYEpjXFrg5pIcDcEaELqouCIa54p8L4pwTEK46MpWPfE6Q/ZY57Q1xPqFzFRTzUlVLTVET4Z4XFkkbxZzXA2II6oss+hJ55amd808r5pXm7nvcXOcfMndMVnDsOrMXxGGgoKd9TVTuyxxs3cf/AFzXRO4Ko4pfZp+McAirQcph7yRzQ77JlDcgP4IlshKyoj/5J8GiZM0Tx4rUvLWv8bQWNsbbj1WDV4xieIRNirMRrKqNnusmne9o9ATZPxPAsRwfGXYZWUrmVl2hrG+PvA73S0j3geRG63JOBmUDhBjPEuD4TW2uaSV75ZI/J+RpDT5X0RPUcmr7scxZ1F7G7Fa40trdyah+S3TLeymx7h2u4eli9q7qWnqGGSnqaeQSQzt6scN/MGxC2p+z6ow+UvxbGMOwyic1hhqpy8iozMa/9GxoLnWDgCbWBRbZy5OKR8ErJInujkjIcxzDYtI2II2KR7nSPc+Rxe55u4uNyT1K3Mc4UqMGw+nxKGtpMUwypeY46ukcSzONSxwIBa62tiFNQ8ICTB6bE8UxqgwamrMxphUB8kkwabFwYwEht9LlDc5ZLcaxVlGaRmJ1raY/9CKh4Z8r2VJbOO8NVGBxUtUKqlxCgrA7uKylcXRvLfeabgFrhzBCuUvBcow6nrsYxfD8Cgq295Ttq3OdLKz7YjYCcvmbXQ3HOMaXXsn92ei3sS4TqsJoI8Siq6PE8Nmf3baujkzsD7e44EAtd5EKXCuFKnEcOOJVNbR4XhwcWNqqx5a2Rw3axoBc4jnYWHVdZrT5vV3c7pzndHojuyuqrODpG4bPX4VitBjVNTDNP7KXCSFv2nRuAdl8xeyzpcDmj4YgxwyxmCaqfSiMXzBzWhxJ5W1V3HPVc7VEg5Qq4aSbAEny1WzguBzcS479H000cUhilmDpL5bRsLyNOoC7Dsx4ew+TExXu4lwwTSYbVF1GWyd7DeJwJd4beEG5sfRS3TUm3mzWlzg0Akk2AHNXsVwbEMAxF1BilHLSVTGhxjkGoBGhV7GOHqDCqCOek4owzFnlwb3NI2QPAt73iaBbT8VqYtwbjM3EeN0lZijK/EcLo21jy973OnjyNcQwnm1rhoeQNklNOQ5c00qdtO9wvo1u9yV0r+A6+HHaPBmBtXiNTTsqHQRm3chwzWeTYCzbEnYXVtNOTtfZPEZ6Eld3/oNh75fZmcW4Aay+URCSQMJ6d7ly/wDBYmI4PVYNiMtBX0zqephNnsdy6HoQeR5qbXxYsVK95F/CFoR07GMAASnSyeHC1lF0hlga9pBAWXUQGJ9jtyK2jYqCWNrxYi6bLGTGS1wcNwbrZikEkYd1CyJm5HmwtZOppXNftmHRaY018w6phc3qPmqMlY5psYWC/VRGukOzGD4INHO3qPmufxtw+kNx7gWgKyQg6M08lzuNVkpxDdvuDks5cN4z26WhH9nwfcCnUFB+74PuBWNL+a3HG8ktbRVqk3e1vxVxrdNVSm1qwlWcpWiwAUrbdVCyN0m2gUxp7DQrDrEzbHmpA26qBhad7KzG6wtdRuH5bIslvdKopmVYdXQTQuc+2dhJN28lvE6aKFz3Ak8l6e37jLoXccup05nNVzfJC16ikjmJcG5HdQN1nTU0kB8TbjqNl93o930+r64rw59K4oUIQvU5BKhd7wDgOEYlwtjuI4jTYdNNRzU0cJxGufSwgPz5ruadzlFlz6vUnSx8q1jj5XTgkL0riHs7wyPF8aqocQZhGEYdTUtRmc11S2Tvh/0TgbvbmFgSs8dmM0nDlNiseLRtbJLTxSe000kEbO+Ng4Pd7wadzYDoSuc7rp2b21enk4RKvRJ+zV2CcTRUFTJ7YyWjq5h39PLA28UTnBzXAkPGgIIJ8wue4zwmjwirwhlFEY21OE0tVJdxdmkey7jr1PJXHuMM8pjilwsm65tCELuwEJVMylkdbMCweY1Wc88cJvKrJbwht8Vo4dSSCZszrNaL2HMpjGRQm1xm81pQObls0gkbr5Xcd3cpccZ6d8OnJfaUpqdbqjM0c18x3IRfdIQPNPzt5G6MwKimsjc9wZHmLnGwA1JPku2rcb4e4mnD+JabEcKxdjGwy1lHaRjywZQXwusQbDXKeS42KV8EzJonFkkbg9rhuCDcFdXW45wpxBWSYli+G4pS4jMc9R7BNGYpn83APF2E7kaqWNRs8M4FJQ9pvC4qsRGLYZMx0uHzgHK5rWvLWhrvdIf9XqsGpbwXU1U09Ri3Ej55ZHPkc6ihuXE3JP6RV8X4pdU1eF/RMD8NpcGblomd5ne05sxe53NxOp5cldrMZ4Qxud9fiuEYlRV8pzTfR0sfcyvO7g14uy+5sSFNVdxZkxLhh/CtHgNJJitVG3FYqkS1dMxjYmu8MjQWuO41t5LB7THVB7T8c9ozZm1JYwHlGAAwDyy2snY5xDDXYfT4ThdD9HYTTPMrYXSd5JLIRYySO0u62gAFgFPPxbhuJ09PHxNgP0pUU8YijrIKk08zmDZrzYh9uRIurJr2l9zToOz+SpPCPGUYLvZXYYHuHLPewPra6pYmA7sr4fts2vqwfIkNK18Dx8T8D8Sew4bFhWDQ0Xcsia8yGWeR7QC951e7KDpsByWBguP0UOCz4HjVHNV4dLMKiN0EgZNBIBYubfQgjQgp7t25+p6WuzuMnG8SIGjcJqyf9hLhtRJB2P4uxjiBPiVPG+x3GQut82hWML4swPh580GEYZWOp6yGSGqnqpGOme1zHBrWhvha0OIcdybLDgxeOLgqqwUxPMs1ZFUiTTKA1jmkeuqt3ak1JpZ4EmfB2g4E+M2JrI2H0ccp/AlX6iMUfZ9xl7KLPOIw00hA1EOd+noXAD4LBwDEWYRxHh2JSMdIykqGTOY213Brr2F1ap+J/ozEsTqX0bKzD8RL21VHK4gSxufmtcbOG4I2KzlK30rJY4wuIFw4gjUEHUea9mqC6t7buFZa9ofVOwqGYtfrmnEL3C/8w+a4VmJcEUMzaykwXFquoYczKauqI/Z2u5Zi0ZngHlpdQ8RcZ1eO45hWNtdJBilHTxskmFgHSse4h7QNhYjTyK5PqX3XO1E01RVTTVD3Pmle58jnG5Libkn4rrMOJn7HscZUEmOlxKmfS3+rI8OEgHq0AlJVY3wfjdU+vxTBsToq6U55hhs8fcyvO7g14uy51sCQs/H+JYsRw6lwjC6H6MwikcZGQGTvHyyEWMkj9MzrabWA2Rfd16Y9CyOTEaWOa3dPmY19/slwB/C66HtLlmk7SscE1x3VQYY27BsbQAwDytZcuuvl4mwLiGCB3FGH1xxGCNsPt+HysD52NFm94x+hcBpmBCLfV2tcB1FQOE+OKcOd7M7C+8c2+mcPAB9bEqGSokj7DYIGuIZLj0heL75YGkX+Jv8ABSU/G+F4ZhWI4JhWEzQYXXUssUrpZRJUTyuAyPe6wAa23ut6ncrCkxyJ/AcGAdy/vosQfWmW4ylro2sy23vcXRnVt21uylxPaThsRPgnbNC8cnNdE64P/rkiD+r9idS6nNnVOMsiqiNCWNhzMaeouSfULK4Ox2Hhni6hxeeGSeKmLiY4yA512Ful9OaXhziMYKyroqyibiOFV7Q2qpHvLM2U3a9rh7r2nYoWXbDzFoLgSCNQRyIXYdp3j4ppah/7VU4ZST1WliZXR+InzIAKjgxPgfD521dNguL187DmZT11TGIA7lmLBmeL8tLrn8XxWrxzF6nE6+TvampeXvdsPIAcgBYAdAjXN26TgZzoMG4uq6c2rYMJPcuA8TWuka2Qjp4fwK4/bQaAaALSwDHavh3F2V9II3kNdHJFKLxzRuFnMcOYIWya3gKSQ1DsGxyF5OY0kVZGYfQPLc4HwJ80T3K2eAqnFK/izhn6Rpy6GjpakYZJJDbvCxjnNbmPvBrtumy89klkmkdLK9z5JCXvc7dzjqSfO63MS4vr63HqLE6VseHDDQxlDBB7lMxpuAL766kne60KrGODMYqX12IYLilBVynPNHh1RH3EjjuQHi7LnW2qJNy70IC6fsWxMT6spcWh9lJHul8bu8A9QAT5qbtOqJZcaweJ7yWQYNRtYL6C7Ln8VkcRcSR4tQUuFYbQjDMIo8zoqcSGRz5He9JI76zjttYDZN4qx2LiDEqWphhkhbBRQUpDyCSY2ZSdORQku9tbAiZeyfiyNxu2GpopmX+q7OWk/EJIMcwerwTD8G4rwuvjNBGWUtbRuDJWxOOYB0b9Hi5uDosrDMehoeEMdwd8Mj5MTdTlkgIys7t5cb89brSPEeA45h1FFxLQV4rKGBtNHWYfIwGSJvuh7H6XA0uChqpK7h+CGgwusoccfinDMte2F+dhidTym2YPjOgJb9Yb2Wzx8zhKbj3FvpPEsfiqo5u6McVHCY42tADWsJkBygWtoFzGM8Q4fJw/FgGBUU9LhzZ/apZKmUSTTy5coJsAGgDYBXZOJsB4hp4DxRhtccRgibD7fh8rA+drRZveMfoXAaZgRdE1eVqjxfhTCuGeIKDD6rHKx+I0oa2OopI2xska4OZIS15IsdL+ai7QpTDNw/SwuDaKLBqd8AA0OcEvd6l2/os/EOIsOgwSowfhrDp6Omqy32upqpRJUVAabhmgDWsvrYb8yn4fxFh9TgdNg/EmGz1tNRE+y1NLKI6iBpNyzUFrmX1AO3IrceLqX96xS4bxHFaHHYJMFa6aucHRtibD3veAtIc0s+sLX08lrVrz/wAjFA4G2XGp2mw2Pct0/BNbxFguA01QOF8OrY66ojdCcQr5WukiY4WcI2MGVpI0zEk9FT4fx+ko8GrMCxmglrcJq5GzWgk7uWCVosHsJ020IO6vLnwf2VNfJx3cC4joaxzvIdw4fmQm9lzieKZ231OE1oA8+5Kv4Jxnw3whiZOCYRXyxVLXQ1k9bKx0zoi0ju42t8LRctJNyTlA0XLcJ4vU8P8AENNilM1kj6cEOjf7sjSC1zT5EEhFjOhgc5rSTyC9N4uxl3DvbdLirG5xT+z94z7cZp4w9vxaSFz+Lz8J1FC92EYbilDWOcCGS1DJIGC+oHhDvRM4uxyHiPimpxWGF8MczY2hjyCRljaw7ebVOWuG1T8GU1L2nTQVDs+A0DTipk5PpAA9g/mu1vrdLhmKVWK8O9oGOyG2I1TYMxbuyGSYh7Qelg1voq1RxqZ+z2Lh8UxFYMsElXfV9M1xeyLro5x+AWLgGPz8O4k6oZDHUQTxugqaaW/dzxO3Y63zvyIRWdGdhy6Ls+LnmfhbhCqnJdVyUEjHuPvOjZKRGSfS4HkqLK7gNs3fjBscc6+b2N1ZH3Xp3mXPb4X81FX8RDGOJafEsUpI30cLo2CiiJaxsDLWib0Fr69Sojn5rnZRFzgtfiauw3EOIaqqwegOHUEhBipyblmgB+ZubeayWkX1V2miNmkB8QuFM1wcNN0oAItbRNdHl1aqM+s0mITKaN3eX2AU1c2z2u66Ip75SeqpJ7V6k/pj5KIAnQLQFJmkL3bHklFG0PzAm3RaZvKjkLWklc5jQ/tD+QLtXQhzC3Zchj0Lm4mW5HGzRrZZyaw5dNQH+z4PuBWLKCh8OHQfcCn1K1HC8i6rSC9SPMKwVEW/1jMOiVrHlMDkYLJpmDfeN/RJJciyMmZgHRYdzw4PbexHwSBxaR5ojb3TSGki6QgD0ClFhrrlSHZV43Ka9wospjpLJgkzIkbfVQSXyDLcXViVbDhbUJr2Ne3QXUMZkLC65IB0vzU8ZuL8inByoVNBG5udngd05LOkhki95pA68luVGjLdSp6eMGmDXAEO1IK+h0O9z6frL3Hn6nRl4cyt3h/i/EeHKGso6SGhqKetcx8sVXTNmaSy+UgO2tmKZU4Mx93U7sh+ydlkTQS078srC0+fNfWw63S681/Z5bjlhW9ifHGN4xT4jBWTQuixGOCKRjIWsaxkJvG1gGjQDyHVW5+0riCopnRPNDmk7jvJRSN7yUw27suJ3sBba1uS5JC6fJ6f4M+eTqp+0PG55YXMFFTRwNnayGnpxHHeZhZI7L1IPw5BYmK4xV4zLTSVjmOdS00dJHlbltGwWaD1NuaoKxTUU9U60bPDzcdAnh0+n+9rRcrl6V1bpcOnq9Wtys+07Za1LhMMFnSfpXjmdh8FoX28l4ur3v06bUx/FUpMMp6Wzrd5J9p3L0CrV2lU5al1nVzbyl7fQr5medyu8q64T36VooWyv8V7Wt8Fahp46fNkB16lR0jSXFWnDRZ88ta36ddTlC9zibXULs/JTusNSo3SgbAn0CwqEh43U0bzl1Tc9wTY2HUJw1GiCYOuEBwBKawGya7dRo50tkwVAUb97G6AxpVZWGvDlq0PEUWE0Qgfw/guIuzlwmraYySC9tL5hoLbeaxQwsNwU6UZzEOrkNunxTibEcaooKWc09PSQeKOlpIWwwtdzOUbnzN1lpuwRda04W7LdLmTLourpNpAVDVn+qP+H5p91FVG9M/4fmsZT07dK/vxnIQhed9sIQhAIQhAIQhAIQhECEIRQhCEAhCEAhCEAhCEAhCEE9OzMHG/NTd35qGnkaxrszgNeam7+P7bfmuuPD5nV+/R3fmkMXml76L+8b80d/F9tvzWnJQq6fJ4x8VLRx5Ib83FS1EjXRHLZ2mqIRaFvos1qCQ2CrmR5OgPr0V3JmKQwA81lpnP77Nq5xUxDu5bYkuO4KsGn13JTmxNG60GRRltr7qzbwqO2ql1yjRZFWcEMJAJ9FSc57bO0WnYE2Kjkpr6gqirDUPLdW3A3I5KyyUPGhukbCWAjKBfeyVsIYb2QRVjA6AE/VN/gmCtiawAQ6DbVWKgXhcLclmW1srDS0cQBsO6/FBrD/dH5qFseW3hubqV+W2g+CVND2x52iHzWDjNTI6vByAeAc1ut0jFwubxoA19x9gLGV9N4T26SiH9nwD/AABTtFx5KCgP9Qg5+AKw33bc12jzXkhCjaPGbqTdM+sSpk3hyU2OwTh0TmMuE/Jbkudd0R0CjcblPefmhjDe5RD4wbKUFPa0AbJHM3IRrRjmghMDNVIDdBB3CmzRhbpZKwWRYlAaVRHM3MAOpVtgytAUD7ggjcaqdpLmgqxzyOSSRMljLJGhzTyIS2S3W5bLuMWbY9Vg1rupj/I4/kVnso6l8xiELg8b3FrLqDulGy93T77qYzV9uV6ONu2ZSYNHHZ057w/ZG3/zWlGA2JoAAA5BKN1XDnW3B9QvPl1c+pd5UyxmM9Jy8Dn8tUneDoVEHO5W+STxuaTcAeSy5py5o3BHqqcjLtkeeZ0UoaB6pjxneBfZTJvBHSgtBVki4TBoVIFltBLFm3OiidECwNvYDZXbAprmA8k2qo1ojjLRrfe6dHHYKfugOSdZAxrUx7dVOAmObqoqk+Jxvbe6dFGAx2cWJ2AVkxghHdFEQQl2rXKY2D2GwJF08RgJMhLx0CsTX0IKmQ/VCO/l+yPkpLgdEmbWzWE/BN1vxwkM7+X7AQ6eVrgAB8k43v4vkEXCSpcJuejRVPHvMCSWpbJC5tiCU8gEKGVgDCeazbdOuHTx3KgQhC4vphCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIBCEIGSMzC5Ufdt6KZ2kfldMJ8l1nDx5z96mZW9EojJeMqf8FYgjzAlVyy9QkLPC4OClj0YB0CdkGo6pjdNDyWa5xK0qQG6iCUOUVISAFE6Qk2CHOumhh3VUrn5Bqhk+1yqUrXTSak2HJSsjcCBsEFsODnaKUG4WfHC+KoJzE3V5iB+UBMcFIUwlEQTD9E4eSqU9OS8lw2V126QuPkrEt0ikZZtwq7g46ABWze6Tnf8A4JYkqv3b7Bqw8ZhArgP8A5LpLkc1z2Nk/SHvfUCzlPTWN9tWg1w+D7gU4B+SgoP3fB9wK0NjddI4Xk1ya0XdbzTtykZq4kJW8OVhg0SkaJGnRJI8Bq5V6Ihe05wRyTC13eB7SQQpM1wnM1N+SqJGudYX0TT3gffOMvSyd5oFiVFJcFwTgjTklailsm81JyTCiEDczwFYsoIv1gVhajnkQpCnWRa60ybuEvNIN7JSoAjmN1Ua4Ea6WVxZpdllJ5FbxrGc3Fpr2sdYuA0StcGk6ixKr96xvP5KMnNGR0N1vblpazDNoNCmuFnZr+qibI8jTL8lIHZhY2v5FZyax9FunXTL6aJLrLaYFO0UIclLkVI51gmMcX3Kie4uNk8StjZqgm5JCFEJw4aWKUvHMgIFDtU+6hcbEEJwdcIHkp0Ju1452UV0CVsb2g80TX4JYQI4S8gdUsDy4PcTcqCplAjDW9UUslw5pV/ouvWwXkk3SE+agLy1xB0IQHX5rDvJNLGYc9PNMl/VFDXadU2UgtJt6pb6XCWZRChCFyfQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQOZlfmYTbYp3s7Ptn5KEOyzB3lZTCe31fxXWcPndTKzKj2dv2z8lNF4GZSou/02SGoI5D5quVtqyXAi2qjOkh1vzUPtDugTmvLndFKRMi/JNB5J2iy0LdU/MAFC42UTphewBJ8lRMRd2yeW6CyqufM7ZpASt74DW/zRYtkEapQ5VGvlYb2JCmbM1x10PQoLF7pExuicSoiCodZmm91Vznqfmp6o3yhVlpKUuJ5n5ppJ6pbapCgS/mufxs/wBofyBb5XP41+3j7gWcuGsOXR0H7BT/AHArR0Cq0GtBTj/AFZvqukcLyBukjGU2Cd1TWutNbyTLhrC+0rTrZRz3vfcJ40cia1lyd/orCUE6cuqe2U3sDuoiwg6bJwYOa0YrDXOA5JweDv8AMKAMCeIsx8IIHqo2l2FwU9uyY2INSsBDtVGUl9Ex26cU0oFjAMoupwwciR6FRQsBOborAC1HO8m5XjZ9/UJC543aD6FSJrh4SqiMuIdfK4fBL3jS4eIfFScklh0BQKLEi2qzLxF2sjdCQbjzWj3TL+7b00WK/wDWPHmVZdJZtcbTtNnNII8joldAALhzdfNLRuvDl6FSsYCCLLo4fXSpl11ubdFIBn9D0U7WhriLJ1h0Cp5aQFpA3SKaT3FCsX03jdi9k0lKUx5IGijSRoaAmSNDlEHEuUgsNyiGgZU1xva6m8J5oLWnS4RSttltfRA0UbwWjySREkoLF1RleXyE3VqR2WM+aqkaqV0wn1N16ou4bGydZFlnbpo27juUocQUtkllUPZLYqVz8zSq6ka67bKN48wIQhc3uCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQMd74ShH1kq6Th83q/foJTTqg7oVcyjdOBs66a3e6UoLAN04OsoNY7dCnh1wstHnUpQABomBKNSgC4Dqk71o3upMl90CFpOoT2Ea5pCdlb0R3OXZKinN0Q78E1RyyWbYboiKY5pPTRMsgHRKqhpCQ7JyRBEVz2NOtX/yBdGRqudxsf2h/IFMuG8eXSUAth8A5ZBorIsVXohfD4PuBWG76LpHmvIGgS0jBJPLfawCQ7KWgHgkd1clXHlE/wABLTuEONwirka6o8LgQBY2UTX3K56d5UjRqpA1vNqjG104Eo3EwDfshLdRi6cFFOQNEWTHPsdUNnkqNxTRISShx8JRFqFtox81Kms9weietuQSEXFkqEDW6tSpNneqU7IAbrEeP0riB9YrbWLKSZHa6XKLEkL3RE20vyIursL82qoQuyzNdstD6wXWOGXJ5bqk+adIQ0XuAoe9ba4IKsYsLKLNAVe6kc8u1KjKxfbpjNQJLX3SX1TgVGid03mLpO5aNgpEII+7b6I7tpCR8ZOt09jLDdFN7pqcAG8k5TQxZhncNOSM8KchDyPELDzTMnmPmtMxt6D5JO5Z9kfJXw21OrJ6ZmXX/wCaLFaRgYfqN+ST2aP7A+Sng186M4g9EZStA0sf2Qk9kj6J4U+biz8pSgEK8aRnQ/NQSsjYDluT6qXHTp085llNIUIQuL6IQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCBOZQjmkJXScPm9X79CQpUFVzK0IuMwudzYJHPDG3J5JlM10s3eO2GyC45oLbHZV3BzDpsrR2UT26LLRjJtbHdStmaq748yiLXt21QaAlFkve+az2yuA1B0S98Sb2V0NISXGqY6UAKkJ3nQNN09kT5PfNh0QSSz2YSFB37Tve6kq25YAANAdbKndWe2bVgSC9wVJmFr3Cp6ozFXxTa3cX01SFVg6xvdSslvofmpYuzyNFzmNj+0f5AulOq53Gx/aA+4FjLh0x5dHQj+z4L/YCn56KGgH9nwebAlmqo4vCPE7oF1jz3lI8tYwuJAVM1xbS9zGCLkkuUE0z5j4jp0UdtUqyLNKRkPW6kN2kEbFV4TlKtt8Q9ViusSMdcKVmt7KrkdGbt1HRPZNbQghRqXSxm1sgOUGcE3F7oBkJtl0Q2lkltzUfikd5JRA5xu86dFM1gaLAWTa8kawAJrzopFXqH5GeZ2Unul4S01c1o7uQ7Gwcr4IcLg3HULnXO0UlNVPhluw+E7tOxXWxy230KGGqin0a4ZvsndTKBHNuEDolSAWKii2oWLJ77vUrbWQ8Ma92a51KslvB5THlEwEkZRchW3zlrOQPkbqDvsos1oaFE5xcbldpjp58svJI6ZzvecSpWOvGFU57p0cmV1idCpkYrd0hTQbpVzrrCFIHWTkhCgcHhOzBQOaRsm94RuFRZLgkzKv3oSh7ne6CURM54aNUwSTfVkdbyKilDmFpJv1Q3UHY+q3MdpctJu8qQwv7w2HmkFXUH65PwCia5wGh+CcC1x8QF+qXG/QnUn1iUVc43cPi1Ht8w5NPwTSwOba9ra66qJ0bmamxHUKWWNY3DJZGISfYYU4Yg46d0D6FVmjvXaN26bKdsQZss7rfhikfO97QAA3rrdQvHg1UgbrYDVOljLaZzj/61Uu7PbeFxxykiohCFxfSCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQNPvIQfeQuk4fO6v36RI9+UXSOfbZQuN9ytSOVp7IzPIASco3V5rQwBoFgFWonjM5vMq0VmrCppRdB2WVRkJC3ROKaqIyzySthJ3TinMdyKbErIwBsnW8SS+micNkCOaH3aRoRZZhbleWnkbLVbusyUgzvI+0VrFnI1IlSLbIslSJLoJ4n65TssPGx/aH8gWu02N1lY0M1eD/gC5ZumHK7TVMslBAC6wyDQbKUANFzZV8PH9nwOP2ApnOv6LpOGLyNSbpbJbWHqhVDtMuisU7r6FV2m+ifC60ixW4vWTmgE6pW6hNHgf5LDcTNYOicGgIaRlui+qjUFkJUx72sbmcbBRTXuDGkuOyzZZDI8uKfPOZnbWaNgoHusLLtjNOOV2Y7xO8gnRi5vyTCeQUoOUALTIeDcOabOHNX6DES4iKc68nLPcbqK9pLpR1KQlUqCp72LITchWysqeFiPfeRwvzK2W7rCkYe9eQeZW8PTnn7KdAbpo2Q1+tjqlK6uQ5JpShHJFiWOTSxUoKp6hPbIR5rlcfwdJkt3RdQtlBUl1hspSWCEIFa0dAnE2TQQBuo5JeQ1RDZXZngdE1hsfJN6p3ILtjNOWV2V2hR5pXalItsnZyDqU5snxCiKUE2tdRFyNzbWHh8k46usNfNVGvsd1app2NJz6dCsXCcuk6lnpYiiDW5nWAVaqqu8vGz3PzSVFQZXZW6NHJV7WWMuHTpz9+bCEIXnfYCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQCEIQRyPLdAoi4lOm98eiiuu2PD5nV+9S3SHVJdJchVyOY4scCNwr7ZBIwOB3WcpIpTG7/AAndSzbUq9dIEjSHC4OiVYaLa6QiyW6Q6qBtkmS6clDrckEjBYJ3JNbcp4FygHOyMJ6BZV7m60Kx2WG3Nyz10xZpeSRCFpkFIlKRQGxWVirx7YLnUMC1VgY04/SH8gWcuG8OWvR2+joByDApgLkD5qGi/dsH3ArMQ0JPNbjN5BGqSydzQ3UlVDeasMhvGJGm6hc2ydDO6I23bzCzY1K0ITdqV7bi6qNqWsfcA2KlNWCNLLnqum4lBIS96BuVTdUEnQ3UZcXblWY7Z8lt9WG6N1KqySOkPiPw5JqQkjZbmMjNypCbJh1Kcm28JJVZJGLvTibyHySxCzbprNXOKKa4m6aNSh2pSjRBPDIY3XDrHqtSnrBIA1+juvVY17JzXaaqDoQdVjSHxu63KBVTABgebfim31Jvqt4MZmkIBOxTuVk3Zbcykao3RdCoQpOWicRom7H1WQWShxGxRrfRFhzTS7L3r+qM7uqSyLJo8qXOTuUJLAeaX4Jo2UJ1zYpqXlotMnbi6OiaNWhLyQDt0bJCEIHD8ktzf1TW6koJ1B6IiVlr3t7wQ4JIzoPIpziuWXD0dL70NQhC8z7AQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCAQhCCCceIeigsVNOfGPRM5Ltjw+Z1fv0xG6ckI0WnLRC3RCUG26HN5hQPjmdGeo6K0yVrxofgqO6LEFSxqVoJQLqkyd7fMeasR1LL6m3qs6WVJYg6pwUlmvbcEFAaFlStClDbC6YC1ou4gAdVUqqzMCyM+HmeqsmzaKrl72U22GgUCELoxfZQhARdVAUiUpFALn8b/eA+4F0B0XP43+8B9wKZcN4ctaiGaigA+wFetYKphbf7PhcfsBWytThm8hIwalKNkjTqqhSmWupEmxQNCVOR8kDUtj0SoKBLdUh2snX0SW5oC1gmPOluqc86KPd4QK92Vnmmjwx+qR5u5OcPDqoGBCQbIQKNSngcgkaLC6GnxIHlwYNN1Ix1wOqrO94qb3cp3uFrFnI836JPNOBu26bs6xO625hKCSk2KCgckKEfNAg8kHdGyPRAWRf4pRYougWwRZJzSoBBGiVIVUK0eFLfRIPdslCAvokQLX1S8kAzR46JHggoG4SvOvqiHMOtvNPf7yYG6ApXA5mlcs+Ho6X3oVCS6VeZ9gIQhE2EJLoQKhCEAhIlQCEiVAIQhAIQhAIQhAIQhDYQhCCvUNu8a8lEBZTTe+PRR2XbHh8zq/fpCkv1SpN1XMpF0gPJANkEIoIugEjQhANwlOqqAhNSjolsDuoEDiNiR6J3fSfbd803Yo0KALnO3JPqgJLJQgEoSIQKhCEAlCRKga7dYGNft4+4F0DuSwMb/eH8gUy4bx5dVSNYMOgOgOQck4ub0uoKVxNFB0DAnrM4W32lzR/ZQcpGijQrpkpSWCVIqBSMmsfE0H4KO6FFXGljhcAH4JS1v2R8lTa4tNwbKxHMHaHdRT8rfsj5JMrfsj5J6RQNLW/ZHySZW/ZHyTkiBMrfsj5JMreg+SchA3K37I+SMo6D5JySyoTKOg+SMo+yPklsiygTKPsj5JWtGQaD5IQ0+ALWLORcrfsj5Jr3Mibq1pd6bIfJ3Ytpm/JZs8/eEhp8PM9VpkTTZ3Wbt16qG5QgAHTmgLnqnNa9/u3+aeyEuF+SmOVjRl2CWrpEKd1ruNkohsdXXT3PuLhMHO5U2aK6Ozbs3TZYn2DrHzUneBgsAnh5fqFNmlM3A1uEX81blY12jhbzUD4SDdpuFraaR3vzKc1pG6UMI+r8U4BUKBYbpzS5nulJdCA1JuhCQ7IBITbdNLr7JQ0nzQIXE+Smp25p2BRhhPJWaVtqhh9fyUvBF0sAdsNkFoLdh8lJoSkI1XLbp7QyNGUCwTHsLi2wtorDgL+iT/JZ2vtXbCL6qQtF9gnItqs+VqjQN2CTM2xNh8kk5sBZQA/itSJalLw5w8I+SW4dqQLBQGVrCSSoJKhztG6BakS1cM0bel011a0eENB87KhmO5QDzV0m19srTvZSixFwAssPKsQVGXwnZZuOmpkuFoO4CYYQdtE9rg5txqnLG6quYnDkmEW5K2kLQeS1MkqqkU7oRyKjMbhyWtp7MsiyWySyqbqaAAh1wN+ilyjoPkooNneqmW2CZW9B8kZG/ZHySoRSFjT9UfJJlb9kfJOukQJlb9kfJGVvQfJKhEJlb9kfJJlHQfJKopZ2RC5Nz0CKkyt+yPkjK3oPkqft5vpH+KPb7bs/FEXMrfsj5JMrfsj5Kq2vHNpA9VYjmZKPCdeiKflb0HyRlb0HyRsi6IMo6D5JC0dB8kqCikyjoPkjKOg+SVCBMo6D5LnsbaPpDYe4OS6Jc9jf7w/kCzWseWjRuzUUJ/wBTgKvh/7BB9wK2heTcqLJyLK7DbJE+yaQgS2l0o1SAkFLt6FAmoPmEoPyQ77Q5JNjpsUFiOS3hJ9CplTB1VuAZm2J2UUWSKR7cpCYUCJEqECIQhAiEqEAAbqESEXaLX+1yCnA1WbVzNJ7tgsB7x6lWM5G1E+clrScvM9VAkSrTIFra3ViFml7hvoEyGEvJJFwNVYeBGwBotdStQhdYaJlswTTrzKe3KLXKioXXva+2iMvMusnO0kuEjmXGmiIUBp3dyTxJbnpbZQZdU5psLX3VEua7CDq4p+QBtwTpqozZttL9SnMJcwDaxv6qBWuBGoOiWQA6gWNrph0ceXNSRSA6OsQrtKhulTnREnwC6tQUDnayXA6LUZVGse/3Wkp7ImA5pXXt9ULTkpz3GSLwnoOaospr1PdSHKTzVNoMrMxyj4KVjOjLlacVBCz3gXeqsNDGjwMAsobZDaWY+7GfVLDC5lQM2hC2g0bqpVxtYRIORUvFWcmAi6C6yrmbWwTmuJOq89jtDyblIUIKypErbX1KRCBkzhY32VCSY5rN0spaua3gBVIrtjHOnF9/VJdIEq0gugm+iEIFItZHulBuUO5ILEMxZtsrsb2vbcLLBsVPBLkda+hXPLFqVoJpeGlOBDm3Chl3WJGtpQ8EaFLuqgcRzUjZSN1q4ptM6Np5KJ0J5KRsgPNPBupvQihaW5gd1Ik+sUq7T3HP6hBQhUIhCjllbE25QSKKWdkQ1Nz0CpSVMj9jYdFDckXde6CaWqfISB4W+W6hcSTqdUl9LIPRUGoQADukKDbpqgcfdGgASX6EppJKRBcgq3CzXkEdVdBBFwbhY4ViGodFZt7t5qDRQkBDgCNbpUAhCEAufxsf2h/IF0C5/G/3h/IFK1jy0MOH9nwfcCtKDD/AN3QfcCsLP0W8lsiyUbJQqG2TSFIUhCKi5pSLehSkIBGxRk3ZIRy5ckX5FLvcfJUIDceinilym6r7G6cDY2UF0y94LbJNb6qs19irDXZmopUiVIgRCVFkCBKiyEAsaX9a/7xWysaX9a/7xVjORqEKWKB0hG1lplZpCO6JDdlC5+vUKYAxsLbgC3Lmq/ruopHFKwOKdHEXuV2OANClWRUEZO6mbCN91ZDAOSUADkstaVjTh24THUoaLtFyrhFklroKORxFiU0xOGnJXi1IWBNmlQvblcDfQKIPu7pZWZYQ4GwsfJUxo7Q2WozWjTDKLucAFpx+OywmSNBGYFwG9lcbUNZHdjySevJbjNapaB6qhXsLS2VotYqSnqu8OUnVTysEkZaeYsrENhk7yJr/tKQHxdD+ao0Lyx74X8tQrM8oiylwu0/gqlWA6w8vyVet1pX/D81Ix9xcEEFRVX7M/4fmpZ6Jyzm6G6miJOpUN9VLFsuFdolRbzQhcmiWSPOVhPROUFSbQnzVgzpTmcSUrIXOF06OPM7XZWwLDRdLdJJtWFMeqkbSjmVME4LPlW/GK/sg6pfZPNWgEtk8qviqGj6FI6jPIq6E6yeVTxZb4Ht5KPULXyg7qtU07Q3M0WVmW2bibSS3dkPNSzbqlEcsrT5q7Irr2yhslQiyqBOEhCahNCxG7Ncp6ig2d6qVanDN5CEIVDXGwJ6LNq5C6Y66DktMhZdSwsIuOaCNvlukLvJIDZK1pJVCAoGyeGAHVOZE53uhTa6qIbotrqpu763B6JrondE2aqI2vohPLDayaWEC9im0IngEG3VM2TgbqovUb80ZF/d5KyqFC+0rmnmNFfUUIQhALn8b/eH8gXQLn8b/eH8gUrWPLToP3dB9wKwq9B+74PuBWFmNXkoTgmhOVCpCbIKY65UDXOudEgBvdLZA3VQjm87pAUrymjRWIDvZLuLoOoSXVQo1HmFPE7TyO6gTmHK6x2WVW7Ismxm4ynlspLIpoCVCECISosgQDVYsv65/wB4rcG6w5f1z/vFajORi0YgKaG1hmdqo6OnBu+QWHIlJO65KVIa5xdqSka3M4WTLm9grdOyx2UE0MYa3QKdDQLJbLLZLBFkI5oGFFtE7dIQim2QQnBNJG6gZZU6qPK/MNnKeSodezQkyOkiLSd+S3Eqow21unDdRuBa8tcpSQ4AW2HJac08Uha4G+oK2mOEkYLeY2WA3QqWKV8UgcNLJssXKphhqWyhWZbSU4O4ISSZammzN9VXjmtTPiO429FpkjHujNgdtLJ807XwObaxNlE1oASP90peD6oraKaLZRKWPZcK7RIhCWy5NEVarPhDeqsqrVe+FYiOIWUqbG2zdU+4CtbgATwE0PbdStIKjUJZKnWuEAIoA0SpRbql06oEsmyi7CFJbTRNcPCUSsx7LPurT9Wj0UMo1Ux9wei6OKNCVC0hEJUGyIkg2d6qVRQbO9VKrGQhCFQiz68+No8loWWfWi849EVXbGXGwWjBTBjRfdQU7bvFwr7dlm1uT6m9yw65QnCMDYBSBLyUXaB8begUbmCysSDRQuuixC5gIUL2BWConIqu+IEXChItore7VGYw4nqtuViOAls7D5rVWZ3ZY5p6FafJECEW80IBc9jf7w/kC6Fc9jf7w/kClax5amH6YdBf7AU5eFVofFhtPb7AVgMAWYt5KHp4ddR2AQDZUTBNfolYbokBWVRlISAEOBsmlvzWmaQuugJwYUWsqgCQix9UqDqEUnJKk5pQqi5A5gZc+8E8uD9RsqTXWKstIBzcnb+qzVPIQlslAQJZFkqdpyCrOzQNQsZ0TnyyHQAONyfVbXNUW08hbICAA43B+KsQ+Rx7hgJ1trYqjI65VqeQ6AtA0+JVUgElKsMa7xK7C7UBUGnUg81PC8tdYm6miNMGw3Sg3ChZIOamGyy2OaUNugkNbcqJ0xt4QUglNgm31Va8jjqbBSM03KVTymOTjqmOBUEL3tZqVGamwu1hI6qOUOc/a4HJSRNkDC21mu6rf0Ztpkw75gkaNeYUWYk6q9FG1osB8VHLThpc5ux/NJUsRhtrG+nVODhYlNY27U4MNtAtIdDO5rsuYgXvZTiS82cCzdiFWbF4gXELR7lhoWlg1G6sSmhh7wN6nRSVDY44iwN1P1ilpsr2AkXczQJZ484B6peEkUrKWNL3QA1T2M8lxrrAGk7KQQ9ShgsVKsWKaGtHJZ1Yc1TYLSdoFmzEOq9EinW0sgNA1KUqJzXPdvokaSlrDsUA5FXZE4Sa6tSuuDodFbBba7MnnQKCEnRTSHw6LLZpc3qnNaCqRBcTc2Swd6XkAkAK6Z2v5SDoltooopXHRw1U6iqEzbSKR+w9E2o1nA81JMLW9FuOdQoslshaYJZFk4C6WyBYha/wUiYzc/BPWoxeQhCFQKpUsvM0+StqOVt3NRqcmRRZTcqwxRt5oMwabNBcfJYnt04TgJ1lW9tt7zHD4J7alkg0PzVZPcLqNzbJxfooJZ2sGqKRwULxZMdVOJ8ATSZXC5cAmjZwUzYmuF26FQNcbgOt6q1GQLAG91qM1E9t1ZGyY9vIKRGSISpEAuexv94fyBdCuexv94fyBZrWPLQws3wyDyYFbVDC3BtFCDsWBWnO5ApIXk4uASB1yoyCT1UjGW3VErN1KRcKICylB0WK3DC2ybZTJhCSmjEhTyE0hVkwpvOycQktcWVCH8UA7IvokutM07zViDxXbyKrA/intOU6boRfs0Nte5QoonZ/joVNYgaogQgBOARCAaqqS4t0B05hWXSNabXuoXFph8IcCeZVGfNrmPIKK+hsVPOwtOtlX22WVMd4XA2vZWGhrgHN3HJQEGyWN5abt+KovRglWgqtPKHnaytrNbhrhmNuSZI4Rt0UjtAqE0hllyNIHmUCmfxdU5lQ1zstiCq7LtlaQ3Ub+asNY57sxAHwVsSVaYAd057OiA2zE4G4WWlbu25vNPyDnqnOZdMN9r2QKBZI4XbqLhNa87FDigkdFEykD2bnTVVXXOila8GBzTyNwo3e6VvfpiGEjTnZaNDOHxmF1geXmssmxsU6ORzJGuadQrCtKMGGpcFM8OzE/Vty5FQyzRvax4PiI2SMqS8tZYW5q3hiHEKSNul0lk5ugXB2KnBIEqlDJLhhWcP1xK0pBdh9Fnf9KVK1DiEgFkqVRuGk8gmZbnVPOgugCwuUNHsCmtcJkdrqb0RVd8WU3AT2ZQE69zYpO710KGi2udE8HRKAAEjtAgg7rPOSeSSXU+ikB8ZTHg3VjNnpGAlAQE4BbciDQpcuqcAlsgaBYlKg+8ULc4YvIQhCqBMcQbap/JNI8A6jmpW8YABa5Ubp7OyNHwG5UzNQm900PzgWPVZjam6ou/LbdSxtBI03TjTN73M1oBUzI8upVSAtsxZsxvJZab/1ZWba8l+hUinxQZn5QLnpyUD5HGQtFtPJXmgZbjRRPjBJN1pNVVDjzViGQk6lM7u5GmyfG2zxojKy9waA47DVRCrBdtYJ09u7d0VIi7Aed7IRpg3FwhIwWjaOgSlVkLnsb/eH8gXQrn8aH9ofyBZyax5XoWZKWAjYsBVtoBbe26ijaXYZTnLazQpITdmvJPov1OyhKhCAUjVG5waLkpWvB1Bus1qVKmu2UUjXudo7TopGg5QCooBRZLlS2TYjc3S6jG+qsEKF4sVZUpjhZ2mxTSny7A+aZe7VuOdCcNRbmm3S+arKaKQtOmiug+EOcQFnje6niJecu5RU5mA0a0lRue5x1OnRSNiPM/JPDGt2CIgbG5x0bZTtgJhFrXy3HqnXVKKpfG9zcxy31VSw2ogyg5wNdiqDrNda61atmaNsl819LXWU4kXBHNSqbufJNYcr77hOvZ1gUoaL2dr5IqaCRpeLaLSB0CzWRAWLfkr0ZuweSxa3D3tLgoxTsbra6mB0QEEOQcgAnNbYqYAEJh0VAemgSjyUBk/SAHZOkqA33RdRUrrZdN1A82Cb3zubbKN7822yBji697pWvvoUGwGpsmMsXILVLEJy9pNrC6Y6JxNgdkxrjG64Q6Vx5rc4YvqnOhZe7nXRnY0eAKIk3ShpKvDJXSEqSka41TdCd/yTWsAO1/VWqIEVbD6/kpaulix5ghOG6tEA7hN7tvRZ8WtoUXU3dt6JQxo5KeJ5ILEg6brPlj7uU6LY0VKuj0Dx8VLi1jkpgpbpt0t1zdA4Xao3Xc0B3LmlL9bJLcyrIbSRggeScI3B+YOJ8kjZG6a2UzTcXBuikIcDcqRuoSXBFilaop10jkqa4oGtbd906Vng806LUp7gtRjKqgF06ycW2KUDlzWnMgCcxhcbAXU0dOTq7QKw1jWCwCsibZ8rCySx6BMU9Z+0fyhQLbAQhCASOHgt0slQdWEKVrGiPZS5VCw6KYHRZboypH6J2llFK4NN3GwVA/WMrOk8Ei0HTM7v1WdNI0vKKmjdcKQNuNVWgeM1uSsF1lAjgGhRxHx3SSPuhnuqxKdO68TvNQwt7yRrfO5Ukh90W81NTxd2y53KrPCZCEirJOSwMa/eH8gW+sDGv3h/IFLw1jy2aGPvMOgaHaFgTYhke5h5Iw11sPpz0aFJVNyTZ27FScE90IKBqARsUhUUjgHCxQxgB6JUhfZFSA2KcCFBnuEl7KadJFlJmF7KEXcOiUREm+YqJcdJ1FLuntuNCo5feCsZpknuhR7GyfKfFboogbldI5087oCNwkvzVZPBsVI1xa4EFQ3T2nkg0BM3ICTvyT87bXuqLTyU7LFh+0FF0sDWyx5A50pDXWOZakb3OO1hvdZT79463UoixmPcZLE6bqpISLW3VlriWjlfmkMVwb6+igouGvRO6c0+aPI+x9VHtpyVVab+rPVWY9G6bFVKeQOOVx1/NXI47CwJ+Kw3tI0p19FGNCnXSqdeyY92iQuUch0UDWeKS6nyAqCL3lI+YRjqVVK9mm4soHOa0+8D5JrpHSOtqfIJWUznav8ACFTWuUJLpCSBsPkiH3fNWpMjGZWjQ8gqzPC+yH9Ur+RKZcdUsniaE0CyS6Y0ddOb1TCnR7pasiZguLqxSftTfj+ShGynpP2lvx/JIVopUIW3MIQhRQop2Z4XBSpDqEpGId/RKdQlnHdzOHmoy/RcXfZupJsEmQn3iU5rk8IQ0QjTxFSsicPccU0NBKmaLKNGvL26kXUkTszbhF7i1lFEcjyFBZJUbjrZKXaJrRc3QqxCLBK/dLGC46KZsQGpXSRxtQCF0hB2CnZC1g2uU9KtyMbCEIWkUaz9o/lCgU9Z+v8AgFAoBCEIBCEIQ1mikB0Uezk4FYdZUoKZI0OGqTNZBkaNSbKiCaFpaLaeiqSU2V291clljNtSq0tQ0nYoGMYGEKR7tN1XMoulzZmgou4cTdPYobqVmyrFTNhzODydFONEjBZg9E5VkIQhAi5/Gj/aH8gXQLnsb/eH8gWcmseWrhuuHwDqwK+5gkp/Fo4BQYc1rcOgtp4ArVvNZ36NarPjfZ2QnRSlQ1DcspsnRyZ22PvBaNnF1gVEXE7KQi6QssNEahgvzKcEmU3TmtR1lSMueSlB0TW6BOXOpbsbqF2snkFPyVUn3itYsZGE3JUY3T0zmukcqkbsi24SN3Sk8+iqAJybtqlvbRVEjHdVYY62vVVL31UrHX0U01Ksa335rPcfGfVaDdbeaznfrHepUD2OsLbqeJ7XPDTuQqt7OBUkRIu4+iImrWeAaC/VUHN8JN9b7LRaDJZrwQ3qqczWtks0XCCEscBfZW6aq0yv1PI9VBlswgkgnko2+B1781Fl00i9wAc5osehUlw5tws90g97OTbkimqe7dkefB+Slje13YprxcJ+h1GyLaLKqznOa05N0kZGYd5Hc83XUxGV48050YOo0Ku1K18TQLb+SZJMbWtZNLHDYhMLHHcovowvJ53SsbY35p7IlJ3YAVLUT9LJqfLyTBsjJCU+PkotypWDZQTtPVWKT9pb8fyVMHVXKT9pZ8fySJWkhCF0YCEIQCRKmudYKJIzK5v6YqmVbrH5pNFUK5V2hQntNyotbp7Roml2nbun5gFAHJS8KaXaXNdNJF1HnKdvqmk2fmuFNTAOeAVAFYpriRJyXhfDQNglTQ4EJHSBq7OPJyFXdI5x6JtyNQVLkaW0KGKa/hcplZRRrP2j+UKBT1n7R/KFAiBCEIBCEIGuGl+iA7ROtc2UJu12qzY3jT3nNoDZRupWabn4pzdSpLEiyjas+OLLsb+qgkbGB4W/irUlPmPkoXQABVVPKCVJcAWCc5mUpMuiOfBBqVK3ko2jVPafEFUXh7gSpkbw4EcwnlVAhCQohFz+N/vD+QLoVz2N/vD+QKVvHluUH7BAD9gKyN9VTw03w2C+4aAVdAv8Vy03VeuZazhtsqNy11xutKqN6ci2yzX7rpGKnjlDx5qRUrkG4U0c3J/zV0kqaycEgIKVZbSNsnKIXCeHaLNjRXGzCqm6lkkzDKPim20VnpmmOFlDzU7tlAd10jnTxyKcmN1b6J7dQCqg8kDoUHRJuL80DgfmntOoKjvcXCVpugusku3TcKiGOkkNhrdTxPs4JpcdRdRboop42+++/kE/vGtGVjbeqiSqocXuduVDM0kCwUiNwgiIOUeLXkmPaCbAahTloI9ENYOlrqCq8FoN9woyLHqvQuGeyPiLiWhGIyOgwrDXAEVNY7Lmb1a3e3mbBar+y7gunJhqO0vDhMDYhrGWv/tqbjXjXl0NSYrNOrenRXWyNe3M03C77E+wzFPYZK/h7GsNx6lYzOe6eGPta55lv4hYPAXA0vGNJjVRFiApPoqATFhhMne3DzYWIt7nnupdVqbjnzYp7Tpqu7wTsjxGpwuPFuIcRpeG6B7cwNWf0pFvskgD4m/km4/wVwnhvDtZX4TxvT4pVUzA4Uoa0GTxAG2t+d+aztrTiMoKC0BdNxPwa/hnhnAsYfXtqG4xH3giEWXuvCHWvc397yXLd4CN0Ck2UbnK9g1AcZx+gwtswhdW1DIBIW5gzM617c91r41wRJhHaRFwia9sr5ZYYvaREWgd4Brlvyv1Qco43KaTqvYKnsJoqKoMFVx1hsEzbXZLEGOHTQyKI9iNBI0iDjzCJZPqtygXPwkP5Js1XkrWqQBdTxf2dY5wSIpcQZFPRzHKyqp3ZoyehuAQfVR8JcBY9xnM8YXTNFPGbSVMxyRMPS+5PkAU2OcDfErVH+0s+P5L0h3ZVw3h5EWK9oWGwVNvFHG1pDT8XX/AKRnY77Wz2nhrifDMbDASY2uDHfAgkfOySljhkKeuoarDK2SjrYH09REbPjkFiF3eGdjWN4rhNJiEOI4eyKqhbM1ry+4DgCAbN31XS2Oclrz1C9M/5C+IP/tPDPnJ/wCVVMR7FeKKSkfLTTUFa5ov3cUjmvPkMwAv8VPKL415294YLkqnLUXNgkkMntD45WuY+MlrmuFiCNCCF1nC3ZbxHxXC2rp4I6Sid7tRVEta/wC6ACSPPZYtakcTMLP9VEdl7bJ/R3qnw5v9I4RLb3fZHZb+ub/guG4t7LOI+EKZ1XUwx1dC33qmlJc1n3gQCPXZZa04lANk8t0Xe8MdjfE/EdIysdHFhtLILsfVEhzx1DBrb1srs04IOSggr2eT+jnVtgJi4jgfLbRrqVzWn45j+S4Di3s64g4MIkxGnbJSPOVtVA7NHfodi0+oTZpzAGqeBqtLhzh2v4nxqLC8NbG6plDnDvH5GgAXJJXo9J/R/wAelANTiuHQA7hgfIR+ACmzTykbBW6dlm5uq9Tn/o+4rGL0+N0UzvsyRPZ+V1xfE3COM8IytjxSk7tj9I5mHNG/0d18jYqxMtsCQkNJCRr8zdUSX7tV45LEgpbpJNpy4BKDmGige4AXTopQBqmzSR2hViGXOLHcKp3gIKewllnXutSpYSs/aP5QoFLUvD5rjoFGATstuZEJ/d9TZPygNV0zcpEPwKXKbXTs4zFj9CENeS/INt7q+KeZ8bcrbndV3tu5w81aOyrE/pHBSxcL7Q3LDrspWyBDmAjVQPjc3YrDukkntsoHyaKJ2YHVRknmmjZxddyLpoKLqsFJsnR+8ozqnMdl1QTwv/TO10urJe0HU2VCC5Jd53ViU5o/MbLenLeqsAgi4SqnHLz6pxmd1TxPPSyVz+Nj+0P5AtyOQuAuAsPG7HEP5AsZT06YZzbXw8gUcWhsWAq2HW0Asq1JE9uHU7rbNCsX0WY2SRpfE/0VF7D3YctDW9uoVYR3gtbUaFajNUiUApzm2KZZVEjJHMPUdFO2W/JVAbKRpSxZVnvOiaXE80gTllo0BSBNTlA1/NQO2Vhw0Vc7LUZpGmxCkGh8iouSeDcea0yk3Fk0filvz6pHDW4QB0NxsltzCa09U7b0QOa69iENN033Tolhjke42YcvXkgelY0vcGjcp7YNbOeAeg1SgRsAkzu3tcBTYjeMjy29yEhKKhuWW42cLpoSB4K67sv4fh4m7QsPoqpofSszVEzDs9rBfKfIm3wXHtPJdn2U49Bw72jYfVVT2x002amke42DM4sCT0zWS8Ljyn7U+MK7iXiuuoe/czDKGZ0EFOw2Z4TlLiOZJB9BZcI6mlkbaGJ78vJjCbfJdv2o8JVvDPGNdPJA76PrpnT084HgIccxZfk4EnTpqqPB/H2M8Ee2fRIpz7Zk7zvoy73b2tqPtFScelvPti4XU41gs5qsNfW0cuUguia4ZgdwRaxHkV6j/R+nNBR8W1bWBzoKWKQNOly3vTY/JdJ2V9qGP8ZcXyYZigpBA2lfMO5jLXXDmga3OmpWF2QtPcdoBduYXfnMsW/i3J9Y8w4g4jxXirE3Yji1S6eZ+rWn3Ix9lreQWU69tVM1tmN9AggEWIHqg9T7VNey/gL/AFb/APaYvJMvkV672msdU9kfA1VCM8EcIjc8bAmNoAP+yfkvJDcDW6Ra2+Bh/wD1C4e/7Qg/3wu842/+stSf63R/k1c72X8KYzjnGOF4hR0bzRUVVHNNUvGWMBrgSAeZ02C6Pjb/AOsrSH/71R/k1T6rOGJ22xtd2sYiS1p/RQ7j/AFwQjbfRjb+QXvnaH2q1fC/G1XhUWCYbVsiZG4SztJeczb62XMO7dsTAvFw5gsbhs7u3G34pCxo4XSV7v6OVdS4oJGGprGMw9s2h8UjA2wPLNmt8VH2u4tJwphuF8DYLIaSiipRJUmIlrpSSQAT0JBcepK4biLtD4g4oxGlqsSqWllJK2aGnibkja4G97cz5ldr2zYW/H4MK42wphqMOqaVrJns8XdEEkF3Qalp6EIfR5Iyw2FlewytqMMxKKspJpKeeI5myRus4H1VNouLjX0U1LTzVVVHTU8T5ppTlZGwZnOPQAKsvXuPJmcVdlmC8YFjGYgx/stSWttn1cP94XH3ivUsAZVTdleHR0b8lU/CYxE4G1nmIZTflqvLOOIm8J9kWB8J1Bb9JTye1TsDgcmpcb/EgD7pXqOCS1MHZNQS0Qc6qZhDHQhrcxLxEMunPXks1ucvNBwr2yZR/bUt7fx4/wAl6DwYMe4c4VqJeOMWgc+OUvbLJKD3cdho59gCb3XljOL+2R7R/VK5unPDB/5V3vZpivHWL1NdFxfQZaNsbTFJNTCFxffUW5i3koryvh2gouPu2yoLGZsMnqpaxwOmeNpuB/MbfMr3LjWp4nosHhpeEMNjmqZPCZS5jW07ANLNcQCTy5Cy4PEZMI4I7fqSaFkVLSV9MGz5bNZG+QkZvLVrSfVd/wAbS8V02GxVPCjKaeWMnvoJWZnPbyLdRqOnO6JHlEmA9sT3GQVOJCTfStYB8r2Xp3AknFVZgNRRca4exs7D3bZCWOFRGRrmDSRfket15u/tJ7S4Ze7mwYsf0OHPWlh/E/a5i8rW0mD08TTvJU0vdMHxcfyVIyOEuzyij7cMTw2ojEtBhH9aiieLhwdYxg9bZvjlC9H7QanjbuYqLhCgv3jS6asL2BzOQa0OO/Mm3Sy837NuOJqntdrZ8efDHVYpCKTOwZGd4wjKPjlI9bL0nj+s43w5kNZwrBT1kAaRPA6LPKDfRzdRccrbqK8yZgHbPFMJWVOJFwN/FWxuHyJsvWuHoMXx/guSg40w2OKqlDoJ2gtLZWcn6EgH8iLryh3af2nNl7s4MQ/7P0bJdatFxX2uVtLNVHCKWlp4I3SPkqqbuhYAk2Bdc7cgg83w6txHgHj2pFBEyeuopZaRjZGl4JN23sNzay7yGHtl4kf3xmqcOhJ0DnMpgPRts3zWr2MU0ON4vjvE9eyKTFZZxYhthEHC5LRyufyVzj7ijj2ix2aiwTC5oqFoHd1MNMZnSXGpvqBrpaysiI8K4R7UqWdkr+KYRlN8s8zpmnyIyrq+0qkjrOzLFm1TGPfFB3wIGjXtsQR8brzbC6Hta4hmDXV+IYfC7QzVDxCGjqGgXPwC9N47jfF2X4vHJIZHtoi1zz9YgC5+KEfKlVMY7CxsVXY/fRXahoO4umMha6M30SxiVUc+/PQJS/w6HVSezNzak2TpIA6VoGjbarPi1uHQDwf8VJf7PzQ1uY22aErW94bDRvVbZtMysu7nslaSNLW9U7wiQ5RoAEEBw813x4ebO+zXs8NxukjdcjW6e08juFCf0c5adna/Faczqhl2527t1UFNLeZ1zrmPyI0Vrlbks4ju8RDBoHaJVjTOoVOQ5agf4hZWY3Z2g9Qq9Yw5MwGrdVmxrG6qQbJrmpInh8YcOaduub0oXsB5Ku9gVxw0Vd4N0VVLbJLKYsuUZAAqmkJCY42BUzmpscXeTBvxKT2zfSVjMsYuh7rtt1KllAbooHbt9V1s082902I3B9VMyMucAoqZuYfFXcoYy6sTK+wBc6bBc9jhtiNrj3AujYLNuVzmNtviN7fUCzlwYctmhnL8Np/EdWDRWYznZqqWHMHsUAcf+jBFlcaQxzri/MLlHssSjUC2tlBnDO9ad9wl9oLtALBVp/fvfUhVETjqU3dKTomk2VZBCBcIBS2VRKxylBVdhIUwJKzWpT0ZgkA6lLZRSHzURABKlITHDmkERCGkhPc3RRea2wmB/FDTyKja7qpOV+fNAjhroUNdyKUi/wDwTLaoJVHLJICGuccttAla5SVLLxRvHTVSh2c5I5Ru02PolcLtkYOuYKGndma6M8xopmO0Y48vCVFI7x0rTzabJgUkYs58XXZRX0B6bqxKU6G/JP3TNwlabiyo9D4b7X8YwbDGYVidLTY7hrQGiKrF3NaOQdY3HqCtB3HXZrO4yTdnYbI7UhkrbX+BH5Ly1Cz4xqZV6vD2wYZgnh4U4Kw/DJCCDNI7M63TwgH4XWHwfx4OF4cebLQGrfjEWQlkgYIz47m1jf3/AMFxcTcrcxGpT7qWRqWlAsAOgslypme26cJAort+D+0efh3CpMExLDYcawWQkmmm3ZffKSCLX1sRvtZbDePezuheKij7PWOqG6tErmFoPxv+S8vL0u+6ml276t7Z8frMWoHiGKhwujnZKaGj8Ala1wOUuOtjbbQeSyMb44ZjPaZDxYKB0LY5YZPZjLcnuwBbNbnbouWLAQozHrumjbf434nbxhxZUY02kNIJ2Mb3RfnIyttvYLn0ZSEqBF1fB/aLjnBueCjdHVUEpJko6gZoyTuRzafTQ8wuVynonBoGqG3qD+0HgHEXmfEezyIVDveMEjQ0n4ZVJH2sYVgzDHwnwdRYXM4Ee0SkPcNPIX/FeW8rJYHDvmi6aNtPEcVrsbxaauxKpfU1M2rnv/IDYAcgF6Fhnbdi2EYPR4bFg1FLHSQsga90rwXBrQLn5LzTLaYO5FPcNQArpnb1f/l8xe37lob/APXPUE3b3jsrSyDCcPicfrOc99vhcLy8IjGpKml8qtYzitbjeIzV+IzuqKmY3e93pYADYC3JdXwv2ucR8O0raN7osSpIwGsjqb5mDoHjW3rdcTIUyMaFDb2qP+kHHkAk4ckL+eWrFvxas+v7e8VlzChwekpr7Ome6Uj4aBeSNH6RSO3U0vlT5nd/O6Z9s7nF5I01Jv8AmvQOH+2TiLBqVlLVNhxWFlg105LZAOmcb/EXXnd0t1dJuvbB2/xZLnh2TN/rYt/urDxrttxbE6KopKfDKSlinY6Muc50jwCLG2wvYrzAXKfoAmoeVbPDPFeJ8I4j7ZhsrRmblkikF2SNGwI/4jVeiRdv88dN+n4ejdKecdSWt+RaSvH82d9uQTXayDoFbJUlsek4z25cSVsDoaCCkw3MP1jAZHj0LtPwUOLds+KY3w/UYNPhdK1tTD3L5hI4u2962115y83ckYAJhdZXdPk1cbpWaMISON3FK0aKojc3wXUjW2YTzIQ8eEJJH5ECbjLy5qQFoFr28lANtU9rSdSECPIMhA00FikB0uNxuE5zQXOaegUYcWnMdxofRdseHlz+8c47OHJR1QuxrxyTzYPy8nbJj9adzT9VbZp8TszAVVq47VUEnnZSU7rNCfVtzQ5gNWkFSkpaXWnb8fzUjxdpUFI4d0G81aGyKzQDTTkfUcbjyVnfUJZ42uaQRodPRQQlwJY692rnlHowy3EpKjcFIQmluiw6Icuqa5TFqYYyqISFap4u6iMhGrvySRQZ3jN7o3UtUfB4dlvGfVx6mWppTkdmfdAiJfGToHEj8FJDFmdmOqsTx3gOX3m+IeoW64YqVCAMzD7wKsO8coaOSiDP60Zm+65mb4qenbcFx5lWGXJ7xYBc/jTP6+PuBdGRdwXP44SMRH3As5Lhyt0crnUNKSABl5K291rOKr0U0TMOg7tlzkGpSFxedSVykeu32f3gF7aqJzszjdBNgo9VWQUh0S3umndVAlB1SAXKNlUSNUw2Vdp1U7CstQ9u9k+1kwC/qnqKEhanJEVC7w6H4KFwsrLxdqhIOUgqxmo+Se1yjulutMpdttig6+qa12ieBpcahAw/irTB3sBjJtpdV7AnoVNGcpHyUojbHDE/MZ7kcmhS97CC4hjiCq1TGWynKCbp8cEzh+rd8VFWBJE57XWLSOaikaBK4cjqEezzAe7f0SzA92xxBDhoVeENASe65JqgklUSXTo25n+Q3UYKssbkj8zupVkSapL6pt00usstpQAeSXIOiiY+5U10DbWSWKVCgQpp3Q4poKKejKEJUCZdEqE0lBHO7ZgsM2t1JSxATtdbr+SilBIBbuPxUtLKHTNGx6fBVF1wLQPIpWHNI49NEr/cKZAfC53moiRxyhNj8LPVRySsLrZwlM0YsA8Khz9k5rbMUd72FxqpTtYLIY39YlkSMF3EpXIIw7WykAUVvEpmIp4Fkx7k5zlC83RCxmwJQ03eSmg+FRyOyxuPkimzVUUN7m56BVBVGaQuHhy7KGCDvZC958Kkkpx3l2ENbzWtQKzEn38TQfRXqarZPfQtt1WY5kY8MYuebipoJAzwuFz1ClI0nysuBe/oonODpMxOii8TthlCeyO2+qztuYnte0HmVYabjayql7Gm258lI2RxHhFgmzxLIbSn0CY+1w7kdClHic6+4smkkEg813w9x4up6yprrmEj6zCguzMe4bFt01j7SWPPwlNYf0Mg6aLowIf1atN8TLKtD+qU8JvooRBCWtmcwnxA8+auAXCz62L9IHoa15YC2R4+Kml3IuSFjWnO4AeahEkL7OzXI6KGSlzU5JN3jW6ZFF+jzNOvNXS7/BY71ubQ6JwII0VYxlkmmxU1jlBGhWbj+C4dSy6p9k5reZ2UUT8zsttVYPhasYx2zy1DdC26jeQ6IAEHRK1xEF3CxN1lP300XSuEm2pCLN1T5Jo4x4nD0WbG2Qt9429VPFSi93apo3IM39VAFwHONh5K2xuWNo8lXkF6ljBsFa5qs8j6y5rHLuxIkfYC6W9gudxf9tHmwFZy4ax5S0Jvh0FvsBWBoFy0eIVUTGxsmc1gGg0TvpStt+0O/BcdvZcXSHUppXNfSdbY/wBYd+CQ4nW/xDvwV2ng6YBJZc2cTrLftDvwSfSdZb9od+CeR4unaLBMO65xuJ1lv17vwTPpOsv+vd+CeR4OnB1UsfvLlPpOsv8Ar3fgpYsUrcx/rDvwS5ExdcywOqVcyMTrP7934JfpOs/v3fgseTfi6ZIua+k6z+/d+CPpOs/v3fIJ5Hi6J3L1UEr/ABFo2XNS4vXiQgVLregUf0nWW/Xu+QW5Wbi6RKFzP0nWf37vkEHE6z+/d+Cu2fB0ye1/muW+k6y36934JRiVZ/fu/BPI8HV3v5j8VINh6LkRidYNp3fgrLsTrLN/Tu90cgnkeLqHyPEF2GzgqnfyOd4nkj1WJHidYbgzutboFROJVmY/p3b+Sz5aPHbrM50IJ+ama8yNcxxubXF1yLcTrLfr3fgpIsUre9b/AFh34K+SeLpgRbzS3F1y8mJ1gkd/WHb+SQYpW2/aHfgnkeLrKduaS52GqtEXXJUuJ1nduPfu38lN9KVv8Q78FLk3MHTEWUblzhxSt/iHfgmnFK3+/d8gp5L4ujGhUocuU+k6z+/d+CcMUrf4h34KeR4urukcVzP0pW2/aHfgmuxOtv8AtDvwV8l8XSE3QN1y/wBJ1n9+78E4YnWf37vwU2eLqQnXXKjFK2/7Q75BOOKVv8Q78E8jxdPfRNJuuZ+k6z+/d+CacUrf4h34J5L4unT6cfp2n/1suU+lK3+Id+CkgxSt71v9Yd+HRPIuLtXEBhJOllmGWWcmOLRt91gV2K1wgNql/wCCdR4lVtp22ncL+QUuTn4tx1FMATmuVVuc2p1CpOxSt7z9od8gsebE6z2l/wCndv0CmOezxdUHOy3DiCFNHVzM0vmv1XKfSdZcjv3fgntxOsyNPfu38lfI07CKvZs4ZSrAe1+xBXAS4nWd879O7foEsOK1wcf6y9Xa+LvSNU8GwXCxY1iJeQat9vgrzcVrSwXqHfgp5L4uqc+4UZN1zH0pW3/aHfgj6Trf4h34KXJfF05IAVCqqC9xY06BYc2KVojNqh3yCzHYpW3P9Yd+CsyS4upZIWNICYC5+hOi5c4pW3/aHfgl+lK3+Id+C1ani61sJI97RX4IY2R3sPVcVBitbmt7Q78FPV4tXNiAFS8fJcssvenO4+3VOqYxOGt25pXvdI7K3QcyuPpcQqicxmcTffRXm4nWX/Xu+QVt07dPGukZGGa/ikkq44dL3K5mqxWuZAS2ocD8FjnFK1xuahxPwVlbymne084mzvGmtlK4CRum642hxOs9mf8Ap3e95dFPFilbf9od+C9ON1Hh6mP71b73Fr9dwng/oZD1IXKz4nWd679O78E5mJ1ns/692/kr5M+DrIP1YUsZs5crFidZkH6d34JzcUrc/wC0O/BPJPF09U3PEfLVQ05uz0WG/E6zIf07tugVODE6wOd+ndv5J5Fxdg3VhCrxeGRzFgxYnWa/1h34KF+J1gqB+nd+CvkeLqi27QOibf6q536TrLftDvwTPpOsuP07vwTyZuLo3PZTNLnavdyVV1ZK91gQ0eQXMT4pWuldeocfkohiVYD+vd+Cz5O86bqO9kf7zyR0TcuYrn/pGrDR+nd+CVmJVn9+78Fds+NdRALaK0xcozEqz+/d8gphidZkP9Yd+Cu3Px9t2I56xx6K1fxFcfTYnWd879O78FY+lK3Mf6w78FPJq46dM53hPmsLGrNrwP8AAFTdilb/ABDvwUFZUSzzB8jy52UC5UypjPb/2Q=="

# ============================================================================
# 3. ESTILIZAÇÃO CSS GLOBAL
# ============================================================================
st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

        html, body, [class*="css"] {{
            font-family: {FONT_STACK};
        }}

        .stApp {{
            background: linear-gradient(180deg, {COLORS["bg"]} 0%, #0D111A 100%);
            color: {COLORS["text"]};
        }}

        [data-testid="stSidebar"] {{
            background-color: {COLORS["sidebar_bg"]};
            border-right: 1px solid {COLORS["border"]};
        }}
        [data-testid="stSidebarUserContent"] {{
            padding-bottom: 80px !important;
            padding-top: 0.8rem !important;
        }}
        /* Aproveita o espaço vazio que sobrava no topo (acima do primeiro
           elemento) tanto na barra lateral quanto no conteúdo principal. */
        div[data-testid="stAppViewContainer"] > .main div.block-container,
        .block-container {{
            padding-top: 1.2rem !important;
        }}
        div[data-baseweb="popover"] {{
            z-index: 999999 !important;
        }}
        /* Garante que listas suspensas (ex: seletor de mês) sempre caibam na tela,
           com rolagem interna, em vez de serem cortadas pela barra de tarefas.
           Vários seletores cobrem diferentes versões do Streamlit/BaseWeb. */
        ul[data-baseweb="menu"],
        div[data-baseweb="menu"],
        div[data-testid="stSelectboxVirtualDropdown"],
        div[data-testid="stSelectboxVirtualDropdown"] ul,
        div[data-baseweb="popover"] div[role="listbox"],
        div[data-baseweb="popover"] ul[role="listbox"] {{
            max-height: 260px !important;
            overflow-y: auto !important;
        }}

        /* Header nativo do Streamlit — transparente */
        header[data-testid="stHeader"] {{
            background: transparent !important;
            pointer-events: none;
        }}
        header[data-testid="stHeader"] * {{
            pointer-events: auto;
        }}
        [data-testid="stAppDeployButton"] {{
            display: none !important;
        }}
        [data-testid="stHeaderActionElements"] {{
            position: fixed !important;
            top: 10px;
            right: 10px;
            z-index: 999999;
        }}

        /* Sidebar: marca + rótulos de filtro */
        .sidebar-brand {{
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 4px 0 16px 0;
            margin-bottom: 12px;
            border-bottom: 1px solid {COLORS["border"]};
        }}
        .sidebar-brand .brand-logo {{
            width: 30px; height: 30px; border-radius: 50%;
            object-fit: cover;
            box-shadow: 0 0 10px {COLORS["primary_soft"]};
            flex-shrink: 0;
        }}
        .sidebar-brand span.title {{
            font-size: 14px; font-weight: 700; color: {COLORS["text"]};
            letter-spacing: 0.2px;
        }}
        .sidebar-brand span.subtitle {{
            display: block; font-size: 11px; color: {COLORS["text_muted"]};
        }}
        section[data-testid="stSidebar"] label p {{
            font-size: 12.5px !important;
            font-weight: 600 !important;
            color: {COLORS["text_muted"]} !important;
            text-transform: uppercase;
            letter-spacing: 0.4px;
        }}

        /* Faixa de contexto — fina, uma linha só, para não repetir um bloco
           grande e genérico em cima de todas as abas (libera espaço vertical
           pra cada aba usar com informação própria/diferente) */
        .top-status-strip {{
            display: flex;
            align-items: center;
            flex-wrap: wrap;
            gap: 8px;
            padding: 6px 4px 10px 4px;
            border-bottom: 1px solid {COLORS["border"]};
            margin-bottom: 12px;
            font-size: 12px;
            color: {COLORS["text_muted"]};
        }}
        .top-status-strip b {{ color: {COLORS["text"]}; font-weight: 600; }}
        .top-status-strip .chip {{
            display: inline-block;
            background: {COLORS["primary_soft"]};
            color: {COLORS["primary"]};
            border-radius: 20px;
            padding: 1px 10px;
            font-weight: 700;
            font-size: 11px;
            letter-spacing: 0.2px;
        }}
        .top-status-strip .sep {{ color: {COLORS["border"]}; }}

        /* Linha de cartões de KPI em flexbox (usada no cabeçalho fixo) */
        .kpi-row {{
            display: flex;
            gap: 14px;
            align-items: stretch;
        }}
        .kpi-row .kpi-card {{
            flex: 1 1 0;
            min-width: 0;
            margin-bottom: 0;
        }}

        /* Cartões de KPI */
        .kpi-card {{
            background-color: {COLORS["surface"]};
            padding: 16px 18px;
            border-radius: 10px;
            border: 1px solid {COLORS["border"]};
            margin-bottom: 10px;
            transition: transform 0.15s ease, border-color 0.15s ease;
        }}
        .kpi-card:hover {{
            transform: translateY(-2px);
            border-color: {COLORS["primary"]};
        }}
        .kpi-top {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 8px;
        }}
        .kpi-label {{
            font-size: 11px !important;
            font-weight: 700 !important;
            color: {COLORS["text_muted"]} !important;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .kpi-icon {{
            font-size: 13px;
            opacity: 0.95;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 26px;
            height: 26px;
            border-radius: 8px;
        }}
        .kpi-value {{
            font-size: 22px !important;
            font-weight: 700 !important;
            letter-spacing: -0.3px;
            line-height: 1.2;
        }}
        .kpi-subtext {{
            font-size: 12px;
            font-weight: 500;
            margin-top: 4px;
            color: {COLORS["muted_line"]};
        }}

        .progress-container {{
            width: 100%;
            background-color: {COLORS["border"]};
            border-radius: 3px;
            height: 6px;
            margin-top: 10px;
            overflow: hidden;
        }}
        .progress-bar {{
            height: 100%;
            background: linear-gradient(90deg, {COLORS["primary"]}, #7AB0FF);
            border-radius: 3px;
        }}

        /* Chart / section captions */
        .section-title {{
            font-size: 14px;
            font-weight: 700;
            color: {COLORS["text"]};
            margin-bottom: 2px;
        }}

        /* Abas */
        button[data-baseweb="tab"] {{
            background-color: {COLORS["surface"]};
            border-radius: 6px;
            color: {COLORS["text_muted"]} !important;
            padding: 7px 16px;
            margin-right: 4px;
            border: 1px solid {COLORS["border"]};
            font-size: 13px;
            font-weight: 500;
        }}
        button[data-baseweb="tab"][aria-selected="true"] {{
            background-color: {COLORS["primary_soft"]} !important;
            color: {COLORS["text"]} !important;
            font-weight: 700;
            border-color: {COLORS["primary"]} !important;
        }}
        hr {{
            border-color: {COLORS["border"]} !important;
        }}

        /* Travar primeira coluna em tabelas */
        div[data-testid="stDataFrame"] div[role="grid"] div[role="row"] div[role="gridcell"]:first-child,
        div[data-testid="stDataFrame"] div[role="grid"] div[role="row"] div[role="columnheader"]:first-child {{
            position: sticky;
            left: 0;
            background-color: {COLORS["surface"]} !important;
            z-index: 3;
            border-right: 1px solid {COLORS["border"]};
        }}

        /* Scrollbar discreta */
        ::-webkit-scrollbar {{ height: 8px; width: 8px; }}
        ::-webkit-scrollbar-track {{ background: {COLORS["bg"]}; }}
        ::-webkit-scrollbar-thumb {{ background: {COLORS["border"]}; border-radius: 4px; }}
        ::-webkit-scrollbar-thumb:hover {{ background: {COLORS["secondary"]}; }}

        .footer-note {{
            text-align: center;
            font-size: 11px;
            color: {COLORS["text_muted"]};
            margin-top: 28px;
            padding-top: 14px;
            border-top: 1px solid {COLORS["border"]};
        }}

        /* ==================== RESPONSIVIDADE MOBILE ==================== */
        @media only screen and (max-width: 768px) {{
            /* A coluna "fixa" (position: sticky) da primeira coluna das
               tabelas dá bug em navegadores mobile ao arrastar a tabela pro
               lado (texto cortado/invisível) -- desliga só no celular; a
               tabela ainda rola normalmente, só sem a coluna presa. */
            div[data-testid="stDataFrame"] div[role="grid"] div[role="row"] div[role="gridcell"]:first-child,
            div[data-testid="stDataFrame"] div[role="grid"] div[role="row"] div[role="columnheader"]:first-child {{
                position: static !important;
                z-index: auto !important;
            }}

            /* Colunas nativas do Streamlit (st.columns) empilham em vez de
               espremer lado a lado -- afeta filtros, formulários, botões etc. */
            div[data-testid="column"] {{
                width: 100% !important;
                flex: 1 1 100% !important;
                min-width: 100% !important;
            }}

            /* Reduz o espaço em branco que sobra no topo no celular. */
            div[data-testid="stAppViewContainer"] > .main div.block-container,
            .block-container {{
                padding: 0.6rem 0.8rem !important;
            }}

            /* Linhas de cartões de KPI (nosso HTML customizado em flexbox,
               usado nas abas de DRE/Previsões/Relatório) NÃO são pegas pela
               regra de coluna acima -- sem isso, ficavam 4-5 cartões
               espremidos numa tela de celular, ilegíveis. Agora empilham um
               por linha, ocupando a largura toda. */
            .kpi-row {{
                flex-direction: column !important;
                gap: 10px !important;
            }}
            .kpi-row .kpi-card {{
                flex: 1 1 100% !important;
                width: 100% !important;
            }}
            .top-status-strip {{ font-size: 11px !important; padding: 5px 2px 8px 2px !important; }}
            .kpi-card {{ padding: 12px !important; }}
            .kpi-value {{ font-size: 20px !important; }}
            .kpi-label {{ font-size: 10.5px !important; }}
            .section-title {{ font-size: 13px !important; }}

            /* Abas: viram uma faixa horizontal que rola, em vez de quebrar
               linha ou cortar o texto. */
            div[data-baseweb="tab-list"] {{
                display: flex !important;
                overflow-x: auto !important;
                white-space: nowrap !important;
                padding-bottom: 5px !important;
            }}
            button[data-baseweb="tab"] {{
                font-size: 11px !important;
                padding: 5px 10px !important;
            }}
            div[data-testid="stDataFrame"] {{ overflow-x: auto !important; }}

            /* Radios horizontais (Sintética/Analítica/Gerencial etc.) --
               deixa as opções quebrarem linha em vez de cortar. */
            div[role="radiogroup"] {{
                flex-wrap: wrap !important;
                row-gap: 6px !important;
            }}

            /* Tela de escolha de painel (Controladoria x Financeiro),
               logo após o login. */
            .hub-wrap {{ margin-top: 32px !important; }}
            .hub-title {{ font-size: 21px !important; }}
            .hub-card {{ padding: 22px 16px 14px 16px !important; margin-bottom: 10px; }}

            /* Tela de login. */
            .login-visual-panel {{ display: none !important; }}
            .login-split-left {{ min-height: auto !important; padding: 3vh 4% 2vh 4% !important; }}
            .login-hero-title {{ font-size: 28px !important; }}

            /* Painel de TV: mesmo pensado pra tela grande, se alguém abrir
               no celular, os grupos de cartões/barras/detalhamento também
               precisam empilhar em vez de espremer. */
            .tv-kpi-grid, .tv-atg-grid {{ flex-direction: column !important; gap: 10px !important; }}
            .tv-header {{ flex-direction: column !important; align-items: flex-start !important; gap: 8px; }}
            .tv-cost-row, .tv-rank-row {{ flex-wrap: wrap !important; }}
        }}

        /* Ajustes extras para telas bem pequenas (celulares na vertical). */
        @media only screen and (max-width: 480px) {{
            .kpi-value {{ font-size: 18px !important; }}
            .hub-card .icone {{ font-size: 32px !important; }}
            .hub-card h3 {{ font-size: 16px !important; }}
        }}

        /* Tela de login (acesso restrito) */
        /* ==================== TELA DE LOGIN (split-screen) ==================== */
        .login-split-left {{
            padding: 2.5vh 8% 0 4%;
        }}
        .login-split-left .login-badge {{
            display: flex; align-items: center; gap: 10px; margin-bottom: 26px;
        }}
        .login-split-left .login-badge img {{
            width: 40px; height: 40px; border-radius: 50%;
            background: #FFFFFF; padding: 5px; object-fit: contain;
            box-shadow: 0 4px 14px rgba(76,141,255,0.3);
        }}
        .login-split-left .login-badge span {{
            font-size: 13px; font-weight: 700; color: {COLORS["text_muted"]};
            letter-spacing: 0.3px;
        }}
        .login-hero-title {{
            font-size: 34px; font-weight: 800; color: {COLORS["text"]};
            margin: 0 0 6px 0; line-height: 1.12; letter-spacing: -0.5px;
        }}
        .login-hero-title .dot {{ color: {COLORS["primary"]}; }}
        .login-hero-sub {{
            font-size: 13px; color: {COLORS["text_muted"]}; margin-bottom: 22px; max-width: 380px;
        }}
        .login-field-label {{
            font-size: 12px; font-weight: 700; color: {COLORS["text_muted"]};
            text-transform: uppercase; letter-spacing: 0.4px; margin: 0 0 6px 2px;
        }}

        /* Painel visual à direita — como não temos uma foto de verdade pra
           usar, montamos um "céu noturno" abstrato só com CSS (gradientes +
           pontinhos de estrela + silhueta de montanha via clip-path),
           seguindo a composição da referência. */
        .login-visual-panel {{
            position: relative;
            height: 640px;
            max-height: 80vh;
            border-radius: 26px;
            overflow: hidden;
            margin: 2.5vh 4% 0 0;
            background-image:
                linear-gradient(180deg, rgba(11,14,20,0.1) 0%, rgba(11,14,20,0.35) 60%, rgba(11,14,20,0.85) 100%),
                url(data:image/jpeg;base64,{LOGIN_BG_B64});
            background-size: cover;
            background-position: center 20%;
            box-shadow: 0 24px 70px rgba(0,0,0,0.5);
        }}
        .login-visual-panel .shooting-star,
        .login-visual-panel .mountains {{
            display: none;
        }}
        .login-visual-panel .panel-footer {{
            position: absolute; left: 28px; bottom: 24px; right: 28px;
            display: flex; justify-content: space-between; align-items: flex-end;
            color: rgba(241,245,249,0.9);
        }}
        .login-visual-panel .panel-footer .marca {{
            font-size: 15px; font-weight: 800; letter-spacing: 0.3px;
        }}
        .login-visual-panel .panel-footer .marca small {{
            display: block; font-size: 11px; font-weight: 500; color: rgba(241,245,249,0.6); margin-top: 2px;
        }}

        /* Campos com "borda em gradiente" (dupla camada de background) --
           border simples não suporta gradiente, então usamos esse truque:
           uma camada de fundo sólida por cima (clip: padding-box) e o
           gradiente só aparece na faixa da borda (clip: border-box).
           Tudo escopado só pra coluna do formulário de login (via :has),
           pra não vazar esse estilo pro resto do painel. */
        div[data-testid="column"]:has(.login-hero-title) div[data-testid="stTextInput"] input {{
            background-image:
                linear-gradient({COLORS["bg"]}, {COLORS["bg"]}),
                linear-gradient(135deg, {COLORS["primary"]} 0%, {COLORS["positive"]} 100%) !important;
            background-origin: border-box !important;
            background-clip: padding-box, border-box !important;
            border: 2px solid transparent !important;
            border-radius: 12px !important;
            color: {COLORS["text"]} !important;
            padding: 12px 16px !important;
            font-size: 14.5px !important;
        }}
        div[data-testid="column"]:has(.login-hero-title) div[data-testid="stTextInput"] input:focus {{
            box-shadow: 0 0 0 3px {COLORS["primary_soft"]} !important;
        }}
        div[data-testid="column"]:has(.login-hero-title) div[data-testid="stTextInput"] label p {{
            display: none !important; /* usamos .login-field-label no lugar */
        }}
        div[data-testid="column"]:has(.login-hero-title) div[data-testid="stCheckbox"] label p {{
            color: {COLORS["text_muted"]} !important;
            font-size: 12.5px !important;
        }}
        .login-forgot-hint {{
            font-size: 12px; color: {COLORS["text_muted"]}; margin: 2px 0 18px 2px;
        }}
        div[data-testid="column"]:has(.login-hero-title) .stButton > button {{
            background: linear-gradient(135deg, {COLORS["primary"]} 0%, {COLORS["positive"]} 100%) !important;
            border: none !important;
            border-radius: 30px !important;
            color: #FFFFFF !important;
            font-weight: 700 !important;
            font-size: 15px !important;
            padding: 13px 0 !important;
            margin-top: 6px;
            box-shadow: 0 10px 26px rgba(76,141,255,0.35);
            transition: transform 0.12s, box-shadow 0.12s;
        }}
        div[data-testid="column"]:has(.login-hero-title) .stButton > button:hover {{
            transform: translateY(-1px);
            box-shadow: 0 12px 30px rgba(76,141,255,0.45);
        }}
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================================
# 3.1 CONTROLE DE ACESSO (login com e-mail / senha + perfis)
# ============================================================================
LOGO_BEEA_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wgARCADIAMgDASIAAhEBAxEB/8QAHAABAAMBAAMBAAAAAAAAAAAAAAYHCAUCAwQB/8QAGAEBAQEBAQAAAAAAAAAAAAAAAAECAwT/2gAMAwEAAhADEAAAAdUgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFS2W0oP23N7qQ+MvtQHtL6ULfUoqpbVVxU1zqBQfMNHodVMuh1D/ACWaDQmbZ0CgeGfNC5p1jS8Hm0Il8qx5Pb3i/vI59c+aDz5oDWPOn7fzoTvi9qvrm8KnkMmmqE1XRNvWds+PHSg9EZ40PrAZ2Ai0pgVlf+6YfusRiRTrozRC4tE+p24JCUB2LbrfWZL+cOyM69jnQdbDpXszW5qn9vcc/oGdgAMH7wwf05XvGY/6NZ607hHBPotWnOsnB07nSaS1fcmfLH1mXxLx+CXQME7VQZ1ZVFWXW+8bKk1fWDx9ASgAMH7w+PWMs+vVH7ZjPQFsfpiWd6W4lmatH9v786xv1tZcW5zt1rt7q5fjuxeKZfrje6yEWD4efPoCgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/EACoQAAEEAgECBAYDAAAAAAAAAAUCAwQGAQcAEBYSFyA2ERQVMTdwIzA1/9oACAEBAAEFAv3JabosdMRUbIXx5cT2eClWgMX2EYfEm0SbHds51g9y/TZY6b02TMfhV6TKexQK8BO2AZ2PYeTH7RSsTLTDhAI67Tc8eWszPHxdqq6araGLPC9CleBOsGcT5vHbsEYdYuoWS9sRpL9lSnCE82b/AK3TaftmX+N9fGYEOs9xieX+2D5YYzAdw4hCW0dBqMB9o+j78rE3FJsuFYVh2kBH3bMFhBbbf/dXTaX8UtC0uIzn4c2VZ4E2DL/G9MpAw4C8sQnBdPEB3Nlw3oz4gxGNwuS5bMCPUPHZLn6T1bhWNhNFOiefQro7yBr2SohZas+bMdDYWMfgIqNnD47HPl+EtbRMhngrjlVqgVyvhujzKJDUvW6o0n6HdE8b13NJOwIEcZF9Mq9BIUnzEr/MbDr+cwSMYmzwpcRAdeNoBcqFWAebTyXMYgMyNlg2FQ9ig5i23EuomkIw1p/YgFjKdmg1ZGWEcZ/ptXuRnVI9xm0a7iBA+qXXUn9k2t0amp0J+xNK1OL8Fgr8ylEq/Z2ydbJkyF0ND9SxsNGNUIRH17ZXhRW712RZYMXUbOMSdZhWUlYWa8XrRJRgF67V7kRrs8tEjXR5DWvrTHGy9kJUm2iwdlmQO2LhyRTLRMxABEQ1GEx5kqf2xcOdsXDkeiWFuXerw6BdiDLNbktamIq5ZAK64RoHtH12n3I3tUo22/tQu81Tq/LLmL5T1WJgZYDFNextuZ4ZmzjM3gBLkuuWOmT63Mg7TKRm3dtTs4pdmL2Cx7OASUkwuxSIWDI2MdJ5KxZsWXQPaPrVCjrV8hG5iFHxz7dH4zMpHbQj4xx8WH1kBB8tTdeFs5SjCE8cCDnVMRWYuHIrLqkIS2n9Z//EACMRAAIBAwMFAQEAAAAAAAAAAAABERIhMQJBURAgIlBxA4H/2gAIAQMBAT8B9q3BL4KrE8EjsSVEkvjseUOdhu1+iyzVg3seTFZ9NPVxuePJCJRCZGlZFBJZlK7P0P4T4i+CsI3shxI/gsdmpSUvkWm0FLWGJPcpjBD3ZTeUUTv7j//EACMRAAIBAgYCAwAAAAAAAAAAAAABEQIhEBIgMUFRUHEiYYH/2gAIAQIBAT8B8rEkLsgjsgVyCCCNC2ZbkWD2RTucFkPbCrFfRclkMmCWXILozaKD9I+Q/Y7jOLsUwL2PRS4JXQ6rmZcobM3ZK6M1oZmjzH//xABPEAACAQIDAwQJDQ4GAwAAAAABAgMABAUREhMhMRQiQVEQMkJxgZGhsdEGFSAjMENhcpOywdLhJDM0RFJTYnBzdIWUo+IlNTZjgpKiwvD/2gAIAQEABj8C/XIuGYZDyvEn3dYT7a2mI421tn71EScvAMhWqD1QTrJ3iP8A2q1tb5xfWMraTN22Xh4+OsIZLiaKAc6RInI1AMOjpraWz+tGGHtX7pvpPkFahjk+2/KKf3VgUUN1NFnufZuV17149lZLeaSCTbqNUTFTwNcoErifkCvtdXOz0DfnSXkePXESsSNLSv0eGv8AUc3ysnppLm5ulxKy1ZNqOrz7xSYq5OxkUFE7pieitvHOMJsG7XSciR5zWpvVBPr69B+tW3tr/wBdLdN7Rvmxy7x+g0ZEGynj3SxdX2exJ6qxTFJufcM2QJ6M95+jsPG9+iuh0sNLbj4qjijv0aSRgqrpbeT4KwKNxqR2CsD0jWKCqNKjcAOxgXxj517K/t18xr+HL8wVDHPfW0Egd+ZJKqnjX+aWfy6+mnw+0nW7nnZfvW8DI58a9SuCXGajSpkXqLvvpUUBVUZADo7NxbQ82G4U80fCuvz+yvsLvjsreY+1ytw/RPhoEHMHpFPI9grO51MdTbz46wFLKAQK8qFgCTnzxXqf+Ovzx2cGuCPa0dsz4VpXUhlYZgjp7C4dbS7eYSB2ZN6rln01/Dl+YKiurkSmVmYc18umu0n+UoS29mu1HCRyWI8dYdjMC6uSuA3wb81pLm2cMrcR0qeo9h57iRYokGZZqvsb0FbaPNUJ72Q8nshHdJzl7SVNzLWnCsbyi6EkJUeLeKybGIFHWD/bVve4ni0t3LC4dV48DnxNYZeRzRxpasCytnmedn2XtLkc07ww4qesVssOxdWth2qucsvAQayxXGfaelIyW8m4VyTD9MdxrDNcT72Ir1r1rteSiDX0Z6cqjs5XWR1ZjqThvPZaKVBJGwyZW4GjcYLiMmHse4zOXjFaRi8BHXn/AGUr41i8lyo97jJPlPDxUlvbRCKFOCj2UtvNe6JY2Ksuyc5HxV+H/wBF/RX4f/Rf0VtrWdJ4+tD2DHcXa7Ue9x84+SsvugfCY/trOzukmI4rwYeA9gzXEyQRDunOQrJZZZ/hjj9NBTcNbk/nky8tB0YOh3hlOYNbS6njt04ZyNlWXLDIf0I29Fb5Jl+ExV9x3cczfkcG8R9xxP8AeH89I5u7nnKD3Poqa8ivZNUeXNly52+po1J2TQEuOjiMqTDrRzHNKuqSReKr1CuVTS8mtM9xyzZ+9XNubsN1ll9FQuk5IPOhuE3V65TcwxKdvl0FeNKozZpG0wwA7kH/AN00OXXkry9IgyUDximfDbmRpFGeyny53hFR4fMx5JO2jS3cP0VbW9uyIVl1s0h6MjX3TiDseqJMq34nLC3+46eitFteJPs8mjuIGqzu5Pvkic7vjcfcMT/eH89Kw0ZEZ/fqLmJZdO/JZczUdi9nHHyhgnKEz1Z9GedXOrgVQr3tNRS2JnNqe00XIA8Wqvxr+bH1qAnglmA4bS4Vsv8AyrHIbqEws3PUagd27PhUcVgWF02enS+g+Ovxr+bH1q/Gv5sfWqOY2RzVw2rap19+lsrILyorqaRt+j7a220mkgPdyyaU8A9Fe2XdsnxdR+iuSPKJm0B9SjKsP7zfPPuGJ/vD+elQW1pkoy7VvrUyLHbREjtlQ5jy1ayJEwtopA7ykc3IGkuLXLlsIyyPvi9VPAFMQzza3uE3d+t9hCT8Y1ogEVtn+aTNvLVot6GeSSHKUS8Tnxzo3FqsktqG1Rzx8U7/AFd+gk8UN1l3ZGlvJXtdlbof0iTScpkY2qo3NjTJAa9co42ktpEAdl7gjrqO0WKCaKPcusHMVsbZUhZui3jzby51/iCyLcyDWdqed4aw/vN88+4EtBGSekoK/B4v+grdBEP+A7OmaJJV6nXOs/Wuz+QWvaLaKH9mgXs6prG2lbreIGs0w20U9YhWslAUdQ7Gp7C2dutoVrKGFIR1IuVaniRz1sudaVUKvUP1af/EACcQAQABAwIGAgMBAQAAAAAAAAERACExQVEQYXGBkaGx8CAwwdFw/9oACAEBAAE/If8AsluEERIcEGYX2NaFK7+wCvhmkMOXJv8AqVg5FppvIPSrNWuEwEIWTmmkqRIh7iejuU0ecJn2oIY2PoDg31zvxwc0KMMlE7Ggs5uTOtSZV9pKKOnJQhQEawndHcuUkwNsFbrzPRqZoszlSeaQ2rLPnNXr/wCgwmO9UhVCWZuEdVf8SXwFqJEXNSvl+HB45U2hhKBWY14QHdRX8YJASj0hgIA2/KeNw0jfiEZv0XhOD1Booy0SwEG9SMmYaAHmBFGjINADAcTZmriNl0P4oBG41FwdsDCyOwYXROtE2NIkjTd6pvDK0kKEwEWq/g7K9WiMXPgfFGjIlIHDQBVgMrQEMWVBG5vpWGkUuUYkEC1BuTlP/KI0zk25JjtUtx0CLPKZO5RP2y88GjwlgseAoWqgmVB9Yq9Tf8np6w7N25NqmdgtPo0p7TNvqoFnRMGAStpMAVhJCSRZHTi5NbTfjmVqzaMHSDs0pt1h5+jegAMnAoZJC2SwBaixO5ptCW8Wp6+mTUa8U8SKkHRK11FewGmOTNRudK/zdSflmIdvgUF9kfJd3n+WLEzAYSSNcxTAFs70z8rRew7Oz14MJWkE3Z0d4qfoel9SpYIJu9RE8MpoT+ynMQ1MfCsYUpD4SHdoPc2QNxo+8sCFdictMAFo73ZURF3EeqnwQShQO0+v0/X76PTUQ3FEzMTjMEERe/qubKYRI9vlqTODIRYTRYe3Wrvyg7DnkObT2Q+zJPT/AHUlobzkyJo3LYvUzSDsBTDqQxzoVvuUS12jekMktpRfVQfWJHQEQ9SlA4bWsjaWyc+VTy9gAhaDOaPupgk7q06IvmN4hTRqvCWFhNpo7jQSBZTyP6Pr99F3HDkotWGE7A1Lz4eR2um0ukBOKeZXc5B8jRpNbNQoliLzwQvStHZE0gnlesXbnQqJYKOlmQSaTrxQQkUqpmSU1iDkfAOANdV/7V4+sT/6aKOHly/irQPwC829fpe/f76TRAJ0aRAdTvMJZ6p2EPASIHVYioBjS0GS/RGY6tNUsgTdhZOo3o3VUM8UFNrRT3H4oAG5C21OtXuBpb8lt0dnxUZjRKTrKPVPzfP9spcZFtMtMZ7tNSIhMKOgSL0RRpTFOJEnxWNCBb9JelNIA3pjh1aa/peoE6VBa+vfyl5U3P8ACgBAQcOWM8fDS0kn10r7FpscUi1lv5SuVc2fiixTgICkmlaZlh+K5f1Z9VoZ2AqFiWDgP+af/9oADAMBAAIAAwAAABDzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzjnWf3yFbpX/zz72Jxqn843z7zywLT+YELx31zzyqSzsPOfeVzzzyvK3EUtXLbfzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz/8QAIBEBAAMAAgICAwAAAAAAAAAAAQARIRAxQVEgUGFxof/aAAgBAwEBPxD7UBbPVDSyS70hZT1FS41QC5Y1Ikw2VdwIlnOUeuAijw5TB3Np+EQKaJVDjq/vnCK8P6hVREsWOmMCxsiDtjP2TrnoRMnGWUDyjU0ZN6lzQKEb6WwtXiKw/CnXCEShigZtXLq3UL7gvRUV2gKA+3//xAAeEQEBAQEBAQADAQEAAAAAAAABEQAhMRAgQVBRof/aAAgBAgEBPxD+qJQ+Atg6T1mAcKmOaul8cN680PjkRj97AxgAifQKJuVf9wtTOh+ej7Vy/uf8zbuG8MLjC+ZgjhPhjdfr8Pbh7vXRqXJBXu44M3FFMT6hmLOtwv4V35kaNTuASBNCQXMoYDgwfBlrf6//xAAnEAEBAAIBAwMFAAMBAAAAAAABEQAhMUFRYRBxgSAwkaGxcMHR8f/aAAgBAQABPxD/ACjfqv3TUqoztpD0igRTszfUmNvoGTy92cl+AT7sn4c4sVaqgpH0ClVgrjcsHbU0BImjZxlNeCtc0Se7kKjGqIpBeXP+zBeuylUAdfU5d/VE+W1tGsYw1ehmj2M9Ub9qXa1twLpBfKmkbmCFV0CGF2famNArOwDSUaDM7GWwCcSgXANwYiaBdsSl+Uvfhmny88h7t4M7VdI27dD/AEM3qQRkeG0IYNEeN/QQSrQ7BcJfYxK3O1R4EcPoz4eBSQkaRNYfdCBMEoCoNvXClxlUUeEUwC0QwxABoAJD65zg8UYJAKCnQTZrERcMh5nfHmbIRF246tk91D7iSDtuc4DNCQqAOAAA9SKaRklQaJN2GH0mUAiPUz44eyBQb4ABgIKhAwB2InJhzp41KAnarrD7+XBI7hrt9BicyNQRH7qD3YDJDQqgnIiNwGxqiAd3ES5mAUYYqdUEajr0PLXJzdwh6GAFANqT+MKy69xQgfkjjOEGCjgDhFnuHXC/t1tHuwP50lEfRziBoj+rwBtYAuRGaNPZVggWPafVGXWIp51orqCuZQRdAlam9IGfMXxkJ40pl4Tf3gPydIKmBhepRMB8GqBMknDNzb6mwsuADUiAqbIijpwyvEbu0A/jk6YUaoRR7Mm+ZHZxXXngLwiqVEWpcBu4LbnjyVxZhqWhCEAQOh9TkHDLRRpExJ9rfO7AHxffh090ax7v6ZFML5pJQd9reTJhsPryjtHalV5fqUht2zUBRHYp6dzTRQQnyuuNzdTHmhzwaA79GYy9n+Qj8PBh75I+/Z+mSF8qN3AJdWTz6cUKPT0Kir0Db0wm8o734d/uY4GJfXz8iAw9ywCHANI9zNcVrhrscAsK6c5QJC/hBfDhtg5dH7P6z3pkQYhHzp9p07lgIEFD549aUObByjmNNrRyKUg68nfIkPbvZsOJQpGbSimwE6gGm2rlGFAETdsQGM/KnIGSVPEe+NxAKuoXGGylA5qBc0oeQg9h8JiOGbpNw6CCrmKwAA6ARdaAVO8rsY4ym34WYt0oLKnOXtqp24N4R1DR52LNccfREKiE+TEE/fgsKvvD2wXhXG3upZ4uCpMp12rZhGn4cGAHUgEOgqnS/Zdb8UEmxTXs5Y+MvArQV1wbemKjH1uRNDhvuK24UVc8GtPH5twd5zDhBJICIbzwZ7ydRUnlDSw4yNpalW0Y32zRnCxYxMSxRpy43ngz/BnuayZAJOy0vfGKBm1w3pI8gE08IsdlUDHYCCI1BExgS8oT87ci7FNEcQq6/r7QH06IKqQAXu1iU+G4ycsHlHjBEx9FNutICu66FyMGDrDRaUSYaGURMUW3i0NAGkALYYYI3dw91f3iDazLvQS32Dh8FFoUB22IR75OR7YUHKM7mmioTFF8nKn4OKyBoufB/XO4N6YXG1dW3xitAI9NRwiFq0ZqoJ2sCmhgVlry4e5TcnokXyB7YaUh23KFC1qEnH2QLBeOAeVUq+gI2UdAcGAAQAgen/grtxMdotq5e03P+VeqU8qVe9LggC4DPkvC3nDwdgNGAEQR0jiRqqDfK3i1E5VvgGa0eaWHBUXLTpk3NYGjb/jT/9k="

CHAVE_LOCALSTORAGE_LOGIN = "beea_login_v1"


def obter_usuarios_cadastrados():
    """Le a lista de usuarios dos Secrets, aceitando tres formas de configuracao
    (todas podem coexistir e sao somadas):

    1) Usuario "de topo" (campos soltos no inicio do arquivo de Secrets):
        email = "controladoria@grupobeea.com.br"
        senha = "Richards23*"
        papel = "admin"

    2) Lista de tabelas:
        [[users]]
        email = "..."
        senha = "..."
        papel = "admin"

    3) Sub-tabelas [usuarios.<qualquer_nome>], aceitando tanto a chave
       'papel' quanto 'perfil':
        [usuarios.coordenador_financeiro]
        email = "coordenador.financeiro@grupobeea.com.br"
        senha = "Montoia6037"
        perfil = "visualizacao"

    A chave do dicionario retornado e sempre o e-mail em minusculas. Se
    nenhum usuario for encontrado em nenhum desses formatos, retorna
    dicionario vazio (painel fica aberto, sem bloqueio de login -- util
    apenas para testes locais sem Secrets configurados)."""
    usuarios = {}

    def _add(email, senha, perfil):
        email = str(email or "").strip().lower()
        senha = str(senha or "")
        if not email or not senha:
            return
        usuarios[email] = {
            "email": email,
            "senha": senha,
            "perfil": str(perfil or "visualizacao").strip().lower(),
        }

    email_topo = st.secrets.get("email", None)
    senha_topo = st.secrets.get("senha", None)
    if email_topo and senha_topo:
        _add(email_topo, senha_topo, st.secrets.get("papel", st.secrets.get("perfil", "admin")))

    for u in st.secrets.get("users", []):
        _add(u.get("email", ""), u.get("senha", ""), u.get("papel", u.get("perfil", "visualizacao")))

    for _chave, dados in dict(st.secrets.get("usuarios", {})).items():
        _add(
            dados.get("email", ""),
            dados.get("senha", ""),
            dados.get("papel", dados.get("perfil", "visualizacao")),
        )

    return usuarios


# Trecho de JS reaproveitado em todos os componentes abaixo: tenta achar a
# janela "de verdade" (a aba do navegador) mesmo quando o painel roda dentro
# de um iframe (que é como o Streamlit renderiza os componentes -- e é
# também o motivo mais comum de "lembrar de mim" não funcionar de verdade:
# usar só `window.top` quebra sempre que o navegador bloqueia o acesso
# entre frames, o que muitos navegadores modernos fazem por padrão).
_JS_ACHAR_JANELA = """
function _janelaAlvo() {
    const candidatas = [window.top, window.parent, window];
    for (const w of candidatas) {
        try {
            if (w && w.localStorage) { w.localStorage.length; return w; }
        } catch (e) { /* sem acesso a essa janela, tenta a próxima */ }
    }
    return null;
}
function _lerCookie(nome) {
    const partes = ("; " + document.cookie).split("; " + nome + "=");
    if (partes.length === 2) return partes.pop().split(";").shift();
    return null;
}
"""


def _salvar_credenciais_no_navegador(email, senha):
    """Grava e-mail/senha (ofuscados em base64 -- isso NAO e criptografia) no
    navegador, para preencher o login sozinho na proxima vez que a pessoa
    abrir o painel. Usa DOIS mecanismos em paralelo (localStorage da janela
    "de verdade" + cookie de longa duração no documento do próprio
    componente) porque, dependendo do navegador/implantação, um dos dois
    pode estar bloqueado -- assim o recurso continua funcionando mesmo se
    um deles falhar."""
    email_b64 = base64.b64encode(email.encode("utf-8")).decode("ascii")
    senha_b64 = base64.b64encode(senha.encode("utf-8")).decode("ascii")
    components.html(
        f"""
        <script>
        {_JS_ACHAR_JANELA}
        const dados = JSON.stringify({{ e: "{email_b64}", s: "{senha_b64}" }});
        try {{
            const w = _janelaAlvo();
            if (w) w.localStorage.setItem('{CHAVE_LOCALSTORAGE_LOGIN}', dados);
        }} catch (e) {{}}
        try {{
            const validade = new Date();
            validade.setDate(validade.getDate() + 90);
            document.cookie = '{CHAVE_LOCALSTORAGE_LOGIN}=' + encodeURIComponent(dados) +
                '; expires=' + validade.toUTCString() + '; path=/; SameSite=Lax';
        }} catch (e) {{}}
        </script>
        """,
        height=0,
        width=0,
    )


def _esquecer_credenciais_no_navegador():
    """Apaga o e-mail/senha salvos no navegador (usado ao clicar em Sair,
    ou quando as credenciais salvas nao sao mais validas)."""
    components.html(
        f"""
        <script>
        {_JS_ACHAR_JANELA}
        try {{ const w = _janelaAlvo(); if (w) w.localStorage.removeItem('{CHAVE_LOCALSTORAGE_LOGIN}'); }} catch (e) {{}}
        try {{ document.cookie = '{CHAVE_LOCALSTORAGE_LOGIN}=; expires=Thu, 01 Jan 1970 00:00:00 UTC; path=/;'; }} catch (e) {{}}
        </script>
        """,
        height=0,
        width=0,
    )


def _tentar_autologin_via_url():
    """Se a URL trouxer credenciais (?le=&ls=) vindas do navegador, tenta
    logar automaticamente com elas. Retorna True se conseguiu logar.

    IMPORTANTE: remove só as chaves 'le'/'ls' da URL -- nunca usar
    st.query_params.clear() aqui, porque isso apagaria TAMBÉM outros
    parâmetros importantes da URL, como o "?modo=tv" do Painel de TV
    (era exatamente por isso que o link do Painel de TV caía sempre
    no painel normal: o parâmetro "modo" era apagado antes mesmo de
    ser lido)."""
    le = st.query_params.get("le")
    ls = st.query_params.get("ls")
    if "le" in st.query_params:
        del st.query_params["le"]
    if "ls" in st.query_params:
        del st.query_params["ls"]
    if not le or not ls:
        return False

    try:
        email_salvo = base64.b64decode(str(le)).decode("utf-8").strip().lower()
        senha_salva = base64.b64decode(str(ls)).decode("utf-8")
    except Exception:
        return False

    usuarios = obter_usuarios_cadastrados()
    usuario = usuarios.get(email_salvo)
    if usuario and senha_salva == usuario["senha"]:
        st.session_state["usuario_logado"] = {"email": usuario["email"], "perfil": usuario["perfil"]}
        return True

    # Credenciais salvas nao sao mais validas (ex.: senha trocada) -- apaga.
    _esquecer_credenciais_no_navegador()
    return False


def _pedir_autofill_via_localstorage():
    """Roda no maximo uma vez por sessao: verifica (via JS) se ha
    credenciais salvas no navegador (localStorage OU cookie, o que estiver
    disponível) e, se houver, recarrega a pagina com elas na URL para
    tentarmos o autologin."""
    components.html(
        f"""
        <script>
        {_JS_ACHAR_JANELA}
        try {{
            let salvo = null;
            const w = _janelaAlvo();
            if (w) {{
                try {{ salvo = w.localStorage.getItem('{CHAVE_LOCALSTORAGE_LOGIN}'); }} catch (e) {{}}
            }}
            if (!salvo) {{
                const viaCookie = _lerCookie('{CHAVE_LOCALSTORAGE_LOGIN}');
                if (viaCookie) salvo = decodeURIComponent(viaCookie);
            }}
            if (salvo) {{
                const dados = JSON.parse(salvo);
                if (dados.e && dados.s) {{
                    const janelaUrl = (w || window).location;
                    const url = new URL(janelaUrl.href);
                    if (!url.searchParams.get('le')) {{
                        url.searchParams.set('le', dados.e);
                        url.searchParams.set('ls', dados.s);
                        (w || window).location.replace(url.toString());
                    }}
                }}
            }}
        }} catch (e) {{}}
        </script>
        """,
        height=0,
        width=0,
    )


def checar_login():
    """Exibe uma tela de login (e-mail + senha) e retorna True somente apos
    autenticacao valida. Guarda o usuario logado em
    st.session_state['usuario_logado']. Se a pessoa marcar "Lembrar de
    mim", salva e-mail/senha no navegador para nao precisar digitar de
    novo da proxima vez. Se nao houver usuarios configurados nos Secrets,
    o painel fica aberto (sem bloqueio)."""

    usuarios = obter_usuarios_cadastrados()
    if not usuarios:
        return True

    if st.session_state.get("usuario_logado"):
        return True

    if _tentar_autologin_via_url():
        st.rerun()

    if not st.session_state.get("_autofill_login_tentado"):
        st.session_state["_autofill_login_tentado"] = True
        _pedir_autofill_via_localstorage()

    def validar_login():
        email_digitado = st.session_state.get("campo_email", "").strip().lower()
        senha_digitada = st.session_state.get("campo_senha", "")
        lembrar = st.session_state.get("campo_lembrar", True)
        usuario = usuarios.get(email_digitado)
        if usuario and str(senha_digitada) == usuario["senha"]:
            st.session_state["usuario_logado"] = {
                "email": usuario["email"],
                "perfil": usuario["perfil"],
            }
            st.session_state["login_invalido"] = False
            if lembrar:
                st.session_state["_credenciais_para_salvar"] = (usuario["email"], senha_digitada)
            else:
                st.session_state["_esquecer_credenciais"] = True
        else:
            st.session_state["usuario_logado"] = None
            st.session_state["login_invalido"] = True

    col_form, col_visual = st.columns([1, 1.05])
    with col_form:
        st.markdown(
            html_compacto(f"""
            <div class="login-split-left">
                <div class="login-badge">
                    <img src="data:image/jpeg;base64,{LOGO_BEEA_B64}" alt="Grupo Beea" />
                    <span>GRUPO B&amp;A</span>
                </div>
                <div class="login-hero-title">Faça seu Login<span class="dot">.</span></div>
                <div class="login-hero-sub">Acesso restrito à Controladoria — Painel Financeiro</div>
            </div>
            """),
            unsafe_allow_html=True,
        )
        st.markdown('<div class="login-field-label">E-mail</div>', unsafe_allow_html=True)
        st.text_input("E-mail", key="campo_email", placeholder="seu.email@grupobeea.com.br", label_visibility="collapsed")
        st.markdown('<div class="login-field-label">Senha</div>', unsafe_allow_html=True)
        st.text_input(
            "Senha",
            type="password",
            key="campo_senha",
            on_change=validar_login,
            placeholder="Digite sua senha",
            label_visibility="collapsed",
        )
        st.checkbox("Lembrar de mim neste navegador", value=True, key="campo_lembrar")
        st.markdown(
            '<div class="login-forgot-hint">Esqueceu a senha ou ainda não tem acesso? '
            "Fale com o administrador da Controladoria.</div>",
            unsafe_allow_html=True,
        )
        if st.button("Entrar", use_container_width=True):
            validar_login()
            st.rerun()
        if st.session_state.get("login_invalido", False):
            st.error("E-mail ou senha incorretos. Tente novamente.")

    with col_visual:
        st.markdown(
            """
            <div class="login-visual-panel">
                <div class="shooting-star"></div>
                <div class="mountains"></div>
                <div class="panel-footer">
                    <div class="marca">Controladoria B&amp;A
                        <small>Painel Financeiro</small>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


    return False


# ============================================================================
# 4. CARREGAMENTO DE DADOS (Google Sheets com fallback local em rede)
# ============================================================================
@st.cache_resource
def obter_caminhos_excel():
    url_orc = "https://docs.google.com/spreadsheets/d/1x68Eg_6LlSKeFJEGmfhyBfcGgheSrVsl/export?format=xlsx"
    url_real = "https://docs.google.com/spreadsheets/d/12I0vGpYU_KNhGxAHOMHWAQu3Xkz_EsUZ/export?format=xlsx"

    caminho_base = r"G:\Meu Drive\Grupo B&A\Escritorio\Financeiro\COORDENAÇÃO FINANCEIRA\ORÇAMENTO\ORÇAMENTO 2026\CONTROLADORIA"
    path_orc_local = os.path.join(caminho_base, "ORCAMENTO 2026 - REV.1.xlsx")
    path_real_local = os.path.join(caminho_base, "REALIZADO 2026.xlsx")

    try:
        xls_orc = pd.ExcelFile(url_orc)
        path_orc = url_orc
        path_real = url_real
    except Exception:
        path_orc = path_orc_local if os.path.exists(path_orc_local) else "ORCAMENTO 2026 - REV.1.xlsx"
        path_real = path_real_local if os.path.exists(path_real_local) else "REALIZADO 2026.xlsx"
        xls_orc = pd.ExcelFile(path_orc)

    xls_real = pd.ExcelFile(path_real)

    # Abas auxiliares/de referência: usadas para buscar dados (DE_PARA de
    # loja x centro de custo, lançamentos do DIÁRIO, mapeamento de plano de
    # contas x linha da DRE etc.), mas que NÃO são lojas/unidades -- por isso
    # nunca devem aparecer no filtro de lojas do painel nem do relatório.
    abas_ignorar = [
        "Sint Ebt loja", "CONS 25X26 V.1", "CONS 25X26 V.2",
        "DE_PARA", "DIÁRIO", "DIARIO", "DRE_MENSAL", "Tabela_Contas", "Tabela_Lojas",
    ]

    # IMPORTANTE: a lista de abas/lojas disponíveis precisa ser a UNIÃO das abas
    # dos dois arquivos (Orçado e Realizado), e não só do Orçado. Uma loja nova
    # (ex.: "LJ ASSAI 23157") pode já existir como aba no Realizado 2026 mas
    # ainda não ter sido criada no Orçado -- se olharmos só para o Orçado, essa
    # loja nunca aparece em lugar nenhum do painel nem do relatório, mesmo tendo
    # dados reais lançados. Mantemos a ORDEM das abas do Orçado primeiro (para
    # não bagunçar o layout já existente) e acrescentamos ao final quaisquer
    # abas que só existam no Realizado.
    abas_orc = list(xls_orc.sheet_names)
    abas_real = list(xls_real.sheet_names)
    abas_uniao = abas_orc + [a for a in abas_real if a not in abas_orc]

    abas_validas = [sheet for sheet in abas_uniao if sheet not in abas_ignorar]

    return abas_validas, path_orc, path_real


@st.cache_resource
def obter_dados_fluxo_caixa():
    """Carrega a aba "Fluxo de Caixa 2026" do Painel Financeiro.

    Lê o CSV PUBLICADO da aba (Arquivo > Compartilhar > Publicar na web >
    aba "Fluxo de Caixa 2026" > CSV). Esse caminho é o que funciona pra
    essa planilha: ela é grande demais (perto do limite de 10 milhões de
    células do Google) pros endpoints normais de export/download, que
    devolviam erro 432 ou estouravam o tempo. O CSV publicado é servido
    pronto, sem o Google precisar converter nada na hora -- baixa leve e
    rápido. Como está marcado "republicar automaticamente quando houver
    alterações", ele reflete as atualizações da planilha sozinho.

    Retorna (df, erro); erro é None se carregou certo."""
    URL_CSV_FLUXO_PUBLICADO = (
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vQB1ygqIm_-o7-314Gynm9H-2Ey2dh_McIIhapZ-9fjAaJc3D14TqRRJSWozuq1RQ"
        "/pub?gid=474479631&single=true&output=csv"
    )
    try:
        df = pd.read_csv(URL_CSV_FLUXO_PUBLICADO)
        df = df.dropna(how="all").dropna(axis=1, how="all")
        df.columns = [str(c).strip() for c in df.columns]
        return df, None
    except Exception as e:
        return None, str(e)


def _detectar_coluna(df, candidatos_nome, tipo_esperado=None):
    """Tenta achar, entre as colunas de `df`, uma que bata (por nome
    aproximado) com algum dos `candidatos_nome`. Se `tipo_esperado` for
    "data", também aceita qualquer coluna que já venha como datetime ou que
    a maioria dos valores consiga converter para data. Devolve o nome da
    coluna encontrada, ou None."""
    colunas_norm = {str(c).strip().lower(): c for c in df.columns}
    for candidato in candidatos_nome:
        candidato_norm = candidato.strip().lower()
        for nome_norm, nome_original in colunas_norm.items():
            if candidato_norm in nome_norm:
                return nome_original
    if tipo_esperado == "data":
        for c in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[c]):
                return c
            try:
                convertido = pd.to_datetime(df[c], errors="coerce")
                if convertido.notna().mean() > 0.7:
                    return c
            except Exception:
                continue
    if tipo_esperado == "numerico":
        for c in df.columns:
            if pd.api.types.is_numeric_dtype(df[c]):
                return c
    return None


try:
    with st.spinner("Conectando às planilhas financeiras..."):
        abas_disponiveis, path_orc, path_real = obter_caminhos_excel()
except Exception as e:
    st.error(f"Erro ao carregar as planilhas: {e}")
    st.stop()


def _ler_aba_ou_vazio(path, aba, colunas_modelo=None):
    """Lê uma aba de uma planilha. Se a aba não existir naquele arquivo
    específico (ex.: loja nova que só tem aba no Realizado, ainda sem aba
    correspondente no Orçado), retorna um DataFrame vazio com as mesmas
    colunas do modelo de referência, em vez de derrubar a loja inteira do
    relatório/painel."""
    try:
        return pd.read_excel(path, sheet_name=aba)
    except Exception:
        if colunas_modelo is not None:
            return pd.DataFrame(columns=colunas_modelo)
        return pd.DataFrame()


@st.cache_data(ttl=60)
def carregar_dados_abas(path_o, path_r, lista_abas):
    """Carrega Orçado e Realizado de cada aba/loja. Cada lado é lido de forma
    INDEPENDENTE: se a loja não existir no Orçado (ex.: loja nova, ainda sem
    orçamento cadastrado) mas existir no Realizado (ou vice-versa), ainda
    assim o lado que existe é carregado normalmente -- o outro lado entra
    como zero. Antes, se faltasse em UM dos dois arquivos, a loja inteira
    era descartada dos dois lados, e por isso lojas novas como a
    "LJ ASSAI 23157" desapareciam de todo o relatório e do Resumo."""
    dfs_o = []
    dfs_r = []
    colunas_modelo_o, colunas_modelo_r = None, None
    for aba in lista_abas:
        df_o = _ler_aba_ou_vazio(path_o, aba, colunas_modelo_o)
        df_r = _ler_aba_ou_vazio(path_r, aba, colunas_modelo_r)
        if not df_o.empty and colunas_modelo_o is None:
            colunas_modelo_o = df_o.columns
        if not df_r.empty and colunas_modelo_r is None:
            colunas_modelo_r = df_r.columns
        dfs_o.append(df_o)
        dfs_r.append(df_r)
    return dfs_o, dfs_r


@st.cache_data(ttl=60)
def carregar_dados_por_loja(path_o, path_r, lista_lojas):
    """Carrega os dados de Orçado/Realizado de cada loja SEPARADAMENTE (uma aba
    por loja), para permitir a divisão por loja no relatório Excel — independente
    do modo de visão escolhido na barra lateral (Consolidado ou Unidades).
    Cada lado (Orçado/Realizado) é lido de forma independente -- uma loja sem
    aba no Orçado ainda entra no relatório com os valores do Realizado (e
    Orçado zerado), em vez de ser descartada por completo."""
    dados_por_loja = {}
    for loja in lista_lojas:
        df_o = _ler_aba_ou_vazio(path_o, loja)
        df_r = _ler_aba_ou_vazio(path_r, loja)
        if df_o.empty and df_r.empty:
            continue
        dados_por_loja[loja] = (df_o, df_r)
    return dados_por_loja


# ---------------------------------------------------------------------------
# 4.1 PLANO DE CONTAS — fonte auxiliar para a aba "Plano de Contas" do relatório
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300)
def carregar_tabela_contas(path_r):
    """Lê a aba Tabela_Contas da planilha Realizado 2026: relação entre
    Natureza Financeira (plano de contas) e Linha DRE."""
    try:
        df = pd.read_excel(path_r, sheet_name="Tabela_Contas")
    except Exception:
        return pd.DataFrame(columns=["Natureza Financeira", "Linha DRE"])
    df.columns = [str(c).strip() for c in df.columns]
    return df


def montar_mapa_planos_por_dre(df_tabela_contas):
    """A partir da Tabela_Contas, retorna {linha_dre: [planos_de_conta...]}."""
    mapa = {}
    if df_tabela_contas.empty or "Linha DRE" not in df_tabela_contas.columns or "Natureza Financeira" not in df_tabela_contas.columns:
        return mapa
    for linha_dre, grupo in df_tabela_contas.groupby("Linha DRE"):
        planos = grupo["Natureza Financeira"].dropna().astype(str).str.strip().unique().tolist()
        mapa[str(linha_dre).strip()] = planos
    return mapa


# ---------------------------------------------------------------------------
# 4.1.1 TABELA_LOJAS — DE_PARA entre o nome da Loja (aba) e o Centro de Custos
# usado na DIÁRIO. Essencial para o detalhamento de lançamentos: a DIÁRIO
# identifica a loja pelo "Centro de Custos", que normalmente NÃO é igual ao
# nome da aba da loja (ex.: aba "LJ ASSAI 23157" pode ter Centro de Custos
# "23157" ou outro código) -- sem esse DE_PARA, o filtro por loja no DIÁRIO
# nunca encontra nada, e por isso os lançamentos apareciam sempre vazios.
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300)
def carregar_tabela_lojas(path_r):
    """Lê a aba Tabela_Lojas da planilha Realizado 2026: relação (DE_PARA)
    entre o nome da Loja e o Centro de Custos correspondente."""
    df = None
    for nome_aba in ("Tabela_Lojas", "Tabela Lojas", "TabelaLojas", "tabela_lojas", "DE_PARA"):
        try:
            df = pd.read_excel(path_r, sheet_name=nome_aba)
            break
        except Exception:
            continue
    if df is None:
        return pd.DataFrame(columns=["Loja", "Centro de Custos"])
    df.columns = [str(c).strip() for c in df.columns]
    return df


def montar_mapa_loja_centro_custo(df_tabela_lojas):
    """A partir da Tabela_Lojas (ou DE_PARA), retorna {nome_da_loja:
    centro_de_custo}. Tenta reconhecer as colunas pelo nome (tolerante a
    várias variações, inclusive um DE_PARA genérico "De"/"Para"); se não
    encontrar, cai para as duas primeiras colunas da aba. Monta o mapa nos
    DOIS sentidos (loja->CC e CC->loja) e devolve a união, para não depender
    de qual das duas colunas veio primeiro na planilha."""
    mapa = {}
    if df_tabela_lojas is None or df_tabela_lojas.empty:
        return mapa

    cols_lower = {str(c).strip().lower(): c for c in df_tabela_lojas.columns}
    nomes_loja = ["loja", "nome da loja", "unidade", "nome loja", "nome unidade", "de"]
    nomes_cc = [
        "centro de custos", "centro de custo", "cod centro de custo", "código centro de custo",
        "centro custo", "cc", "cod cc", "código cc", "para",
    ]
    col_loja = next((cols_lower[c] for c in nomes_loja if c in cols_lower), None)
    col_cc = next((cols_lower[c] for c in nomes_cc if c in cols_lower), None)

    if col_loja is None or col_cc is None:
        if len(df_tabela_lojas.columns) >= 2:
            col_loja, col_cc = df_tabela_lojas.columns[0], df_tabela_lojas.columns[1]
        else:
            return mapa

    for _, linha in df_tabela_lojas.iterrows():
        loja = str(linha[col_loja]).strip()
        centro_custo = str(linha[col_cc]).strip()
        if loja and centro_custo and loja.lower() != "nan" and centro_custo.lower() != "nan":
            mapa[loja] = centro_custo
            mapa.setdefault(centro_custo, loja)
    return mapa


# ---------------------------------------------------------------------------
# 4.2 DIÁRIO — lançamentos detalhados, fonte principal da aba "Plano de Contas"
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300)
def carregar_diario(path_r):
    """Lê a aba DIÁRIO da planilha Realizado 2026: lançamentos detalhados,
    com Valor Bruto, Competência (data do lançamento), Plano de Contas,
    Centro de Custos (loja/unidade), Linha DRE (a linha da DRE a que aquele
    plano de contas pertence) e, quando disponíveis, Número, Cliente /
    Fornecedor e Histórico. É a fonte usada para montar as abas "Plano de
    Contas" e "Lançamentos" do relatório em Excel."""
    colunas_saida = [
        "Valor Bruto", "Competência", "Plano de Contas", "Centro de Custos", "Linha DRE",
        "Número", "Cliente / Fornecedor", "Histórico", "Mês",
    ]

    apelidos_coluna = {
        "valor bruto": "Valor Bruto",
        "competência": "Competência",
        "competencia": "Competência",
        "plano de contas": "Plano de Contas",
        "centro de custos": "Centro de Custos",
        "centro de custo": "Centro de Custos",
        "linha dre": "Linha DRE",
        "número": "Número",
        "numero": "Número",
        "nº": "Número",
        "n°": "Número",
        "num": "Número",
        "cliente / fornecedor": "Cliente / Fornecedor",
        "cliente/fornecedor": "Cliente / Fornecedor",
        "fornecedor / cliente": "Cliente / Fornecedor",
        "fornecedor/cliente": "Cliente / Fornecedor",
        "cliente": "Cliente / Fornecedor",
        "fornecedor": "Cliente / Fornecedor",
        "histórico": "Histórico",
        "historico": "Histórico",
        "descrição": "Histórico",
        "descricao": "Histórico",
    }

    df = None
    for nome_aba in ("DIÁRIO", "DIARIO", "Diário", "Diario"):
        try:
            df = pd.read_excel(path_r, sheet_name=nome_aba)
            break
        except Exception:
            continue
    if df is None:
        return pd.DataFrame(columns=colunas_saida)

    df.columns = [str(c).strip() for c in df.columns]
    renomeio = {c: apelidos_coluna[c.strip().lower()] for c in df.columns if c.strip().lower() in apelidos_coluna}
    df = df.rename(columns=renomeio)

    colunas_necessarias = ["Valor Bruto", "Competência", "Plano de Contas", "Centro de Custos", "Linha DRE"]
    if any(c not in df.columns for c in colunas_necessarias):
        return pd.DataFrame(columns=colunas_saida)

    # Colunas extras (opcionais) para dar mais contexto na aba "Lançamentos".
    # Se a DIÁRIO não tiver alguma delas, entra em branco -- não impede o resto.
    colunas_extras = ["Número", "Cliente / Fornecedor", "Histórico"]
    for col_extra in colunas_extras:
        if col_extra not in df.columns:
            df[col_extra] = ""

    df = df[colunas_necessarias + colunas_extras].copy()
    df["Valor Bruto"] = pd.to_numeric(df["Valor Bruto"], errors="coerce").fillna(0)
    df["Plano de Contas"] = df["Plano de Contas"].astype(str).str.strip()
    df["Centro de Custos"] = df["Centro de Custos"].astype(str).str.strip()
    df["Linha DRE"] = df["Linha DRE"].astype(str).str.strip()
    for col_extra in colunas_extras:
        df[col_extra] = df[col_extra].fillna("").astype(str).str.strip()

    # Competência normalmente vem como data (dd/mm/aaaa) -> convertemos para "mm/aaaa"
    # para casar com as colunas de mês (ex.: "01/2026") usadas no resto do painel.
    competencia_dt = pd.to_datetime(df["Competência"], errors="coerce", dayfirst=True)
    mes_formatado = competencia_dt.dt.strftime("%m/%Y")
    mask_sem_data = mes_formatado.isna()
    if mask_sem_data.any():
        # Se não deu para converter em data, usa o próprio texto da célula
        # (cobre o caso de a Competência já vir como "mm/aaaa" digitada).
        mes_formatado = mes_formatado.copy()
        mes_formatado[mask_sem_data] = df.loc[mask_sem_data, "Competência"].astype(str).str.strip()
    df["Mês"] = mes_formatado

    # IMPORTANTE: filtra só por Plano de Contas preenchido -- NÃO exige Linha
    # DRE preenchida, porque planos de contas "fora da DRE" (ex.: Mercadorias
    # no modelo de Compras) legitimamente não têm Linha DRE, e precisam
    # continuar aparecendo aqui para serem encontrados pelo relatório.
    df = df[df["Plano de Contas"] != ""]
    return df.reset_index(drop=True)


def _normalizar_texto(txt):
    """Normaliza texto para comparação tolerante (maiúsculas, sem espaços duplicados)."""
    return re.sub(r"\s+", " ", str(txt or "").strip().upper())


def _normalizar_nome_aba(nome):
    """Normaliza nome de ABA especificamente para reconhecer visões
    consolidadas (ex.: "CONSOLIDADO - G&A") mesmo com pequenas diferenças de
    digitação na planilha -- maiúsculas, espaços colapsados, variações de
    hífen/traço (-, –, —) tratadas como o mesmo caractere, e espaçamento
    irregular ao redor do hífen (ex.: "X- Y", "X -Y", "X-Y") normalizado
    para o mesmo padrão "X - Y"."""
    texto = str(nome or "").strip().upper()
    texto = re.sub(r"[\u2010-\u2015\u2212]", "-", texto)
    texto = re.sub(r"\s*-\s*", " - ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def _nome_departamento_curto(nome_modelo):
    """Extrai só o nome do departamento a partir da chave do modelo de
    relatório -- ex.: "📣 Relatório de Custos - MKT" vira "MKT". Usado no
    seletor de Visão por Departamento pra não repetir "Relatório de Custos"
    toda hora."""
    if nome_modelo == "Controladoria":
        return nome_modelo
    texto = re.sub(r"^[^\w]+", "", str(nome_modelo), flags=re.UNICODE).strip()
    if " - " in texto:
        texto = texto.split(" - ")[-1].strip()
    return texto or nome_modelo


def _resolver_termo_departamento(termo, linhas_disponiveis):
    """Resolve um termo do modelo de relatório contra as linhas REAIS da
    DRE atual. Dois formatos de termo:
    - Texto comum (ex.: "6.11 - Catálogos e Revistas"): casa por substring
      de texto, como sempre funcionou.
    - "PREFIXO:6.24.2": casa pela HIERARQUIA NUMÉRICA -- a própria linha
      6.24.2 e QUALQUER linha abaixo dela (6.24.2.1, 6.24.2.3.5 etc.), sem
      precisar listar cada sublinha manualmente. Útil quando um
      departamento é dono de todo um ramo da árvore da DRE (ex.: MKT é
      dono de 6.24.2 pra baixo, mas não de 6.24.1, que é de outra área)."""
    termo = str(termo).strip()
    if termo.startswith("PREFIXO:"):
        prefixo = termo.split("PREFIXO:", 1)[1].strip()
        return [l for l in linhas_disponiveis if _linha_pertence_ao_grupo(l, prefixo)]
    termo_norm = termo.lower()
    return [l for l in linhas_disponiveis if termo_norm in str(l).strip().lower()]


def _linhas_raiz_do_conjunto(linhas):
    """Dentro de um conjunto de linhas já resolvidas, devolve só as
    "raízes" -- linhas que não têm nenhuma outra linha do MESMO conjunto
    como ancestral. Necessário porque um modelo de departamento pode listar
    uma linha de grupo e as sublinhas dela ao mesmo tempo (ex.: "8.3 -
    Pessoal" e "8.3.1 - Salários"); como o valor da linha de grupo já
    inclui o das sublinhas, somar todo mundo junto contaria o mesmo custo
    mais de uma vez. Usado em qualquer total/soma/gráfico de composição do
    Modo Departamento -- a tabela de detalhe (que lista linha por linha,
    sem somar) continua usando o conjunto completo normalmente."""
    numeros = {l: _numero_linha_dre(l) for l in linhas}
    raizes = []
    for l, num in numeros.items():
        if num is None:
            raizes.append(l)
            continue
        tem_ancestral_no_conjunto = any(
            outro_num is not None and outro_l != l and _linha_pertence_ao_grupo(l, outro_num) and outro_num != num
            for outro_l, outro_num in numeros.items()
        )
        if not tem_ancestral_no_conjunto:
            raizes.append(l)
    return raizes


def _filhos_diretos_do_conjunto(linhas, numero_pai):
    """Dentro de um conjunto de linhas, devolve as que são filhas DIRETAS
    (um nível abaixo, não neta) de `numero_pai` -- ex.: filhas de "6.24.2"
    são "6.24.2.1", "6.24.2.2" etc., mas NÃO "6.24.2.1.1" (que é neta)."""
    profundidade_pai = len(str(numero_pai).split("."))
    resultado = []
    for linha in linhas:
        numero = _numero_linha_dre(linha)
        if not numero or numero == numero_pai or not _linha_pertence_ao_grupo(linha, numero_pai):
            continue
        if len(numero.split(".")) == profundidade_pai + 1:
            resultado.append(linha)
    return resultado


def _linhas_composicao_do_conjunto(linhas):
    """Para gráficos de COMPOSIÇÃO/ranking (não soma/total -- ver
    _linhas_raiz_do_conjunto para isso): parte das raízes do conjunto, mas
    quando uma raiz tem filhas diretas dentro do MESMO conjunto, usa as
    filhas dela em vez da própria raiz -- dá uma composição mais detalhada
    e útil (ex.: em vez de uma fatia única "6.24.2 - Marketing Regional",
    mostra "Eventos", "Produção e Propaganda", "Mkt Digital" etc.). Se a
    raiz não tiver nenhuma filha no conjunto (é uma "folha"), mantém a
    própria raiz como está."""
    raizes = _linhas_raiz_do_conjunto(linhas)
    resultado = []
    for r in raizes:
        numero_r = _numero_linha_dre(r)
        filhas = _filhos_diretos_do_conjunto(linhas, numero_r) if numero_r else []
        resultado.extend(filhas if filhas else [r])
    return resultado


def _linhas_folha_do_conjunto(linhas):
    """Devolve apenas as linhas FOLHA -- as que não têm nenhuma sublinha
    dentro do mesmo conjunto. É o nível mais analítico disponível.

    Usado na Curva ABC e na detecção de anomalias: com as linhas de topo
    (4 - Custo das Vendas, 6 - Despesas Variáveis, 8 - Despesas
    Operacionais) a análise fica inútil, porque três contas gigantes
    explicam tudo e não apontam onde agir. Nas folhas aparecem as contas
    de verdade (Salários, Energia, Frete, Aluguel...)."""
    numeros = {l: _numero_linha_dre(l) for l in linhas}
    folhas = []
    for linha, numero in numeros.items():
        if numero is None:
            folhas.append(linha)
            continue
        tem_filha = any(
            outro_num is not None and outro_num != numero
            and _linha_pertence_ao_grupo(outro_l, numero)
            for outro_l, outro_num in numeros.items()
        )
        if not tem_filha:
            folhas.append(linha)
    return folhas


def _fator_proporcional_mes_corrente(colunas_periodo, meses_cols_ref, data_hoje):
    """Quando o período analisado inclui o MÊS CORRENTE, ele ainda não
    terminou -- comparar o realizado parcial contra o orçamento inteiro do
    mês faz o desvio parecer sempre péssimo.

    Devolve (coluna_do_mes_corrente, fracao_decorrida, dias_decorridos,
    dias_no_mes). Se o período não inclui o mês corrente, devolve
    (None, 1.0, 0, 0) -- ou seja, sem ajuste."""
    coluna_mes_atual = f"{data_hoje.month:02d}/{data_hoje.year}"
    if coluna_mes_atual not in colunas_periodo or coluna_mes_atual not in meses_cols_ref:
        return None, 1.0, 0, 0
    dias_no_mes = pd.Period(f"{data_hoje.year}-{data_hoje.month:02d}", freq="M").days_in_month
    dias_decorridos = data_hoje.day
    return coluna_mes_atual, dias_decorridos / dias_no_mes, dias_decorridos, dias_no_mes


def _orcado_proporcional(list_orc, termo, colunas_periodo, meses_cols_ref, data_hoje, exato=False):
    """Orçado do período com o mês corrente ajustado pelos dias já
    decorridos. Ex.: no dia 12 de um mês de 31 dias, entra 12/31 do
    orçamento daquele mês, e os meses anteriores entram inteiros.

    É esse número que deve ser comparado com o realizado até hoje --
    comparar contra o orçamento cheio de um mês que ainda está correndo
    superestima o desvio e faz qualquer alerta disparar sempre."""
    coluna_atual, fracao, _, _ = _fator_proporcional_mes_corrente(
        colunas_periodo, meses_cols_ref, data_hoje
    )
    if coluna_atual is None:
        return get_valor_consolidado_multi(list_orc, termo, colunas_periodo, exato_linha_sintetica=exato)

    colunas_fechadas = [c for c in colunas_periodo if c != coluna_atual]
    valor_fechado = (
        get_valor_consolidado_multi(list_orc, termo, colunas_fechadas, exato_linha_sintetica=exato)
        if colunas_fechadas else 0.0
    )
    valor_mes_atual = get_valor_consolidado_multi(
        list_orc, termo, [coluna_atual], exato_linha_sintetica=exato
    )
    return valor_fechado + valor_mes_atual * fracao


def _mask_loja_por_centro_custo(df_diario, candidatos_loja):
    """Retorna a máscara (True/False por linha) de correspondência da coluna
    "Centro de Custos" do DIÁRIO contra um ou mais candidatos de loja.

    Aceita tanto uma string única quanto uma lista de candidatos -- isso é
    usado para tentar, na ordem, o Centro de Custos resolvido pela
    Tabela_Lojas (DE_PARA) E o nome da própria loja/aba, sem depender de um
    único mapeamento estar 100% correto (se o DE_PARA não achar nada, ainda
    tentamos casar direto pelo nome da loja, e vice-versa)."""
    if isinstance(candidatos_loja, str):
        candidatos_loja = [candidatos_loja]
    candidatos_norm = [c for c in (_normalizar_texto(c) for c in candidatos_loja) if c]

    centros_norm = df_diario["Centro de Custos"].map(_normalizar_texto)

    # 1ª tentativa: correspondência exata com qualquer um dos candidatos.
    for cand in candidatos_norm:
        mask = centros_norm == cand
        if mask.any():
            return mask

    # 2ª tentativa: correspondência "contém" (num sentido ou no outro) com
    # qualquer um dos candidatos.
    for cand in candidatos_norm:
        mask = centros_norm.apply(lambda x: (cand in x) or (x in cand) if x else False)
        if mask.any():
            return mask

    return pd.Series(False, index=df_diario.index)


def _filtrar_diario_por_loja_e_conta(df_diario, loja, conta):
    """Filtra o DIÁRIO pelo Centro de Custos (loja) e pela Linha DRE (conta),
    com correspondência tolerante (exata primeiro, depois por contém).
    "loja" pode ser uma string ou uma lista de candidatos (ver
    _mask_loja_por_centro_custo)."""
    if df_diario is None or df_diario.empty:
        return df_diario

    conta_norm = _normalizar_texto(conta)

    mask_loja = _mask_loja_por_centro_custo(df_diario, loja)

    linhas_norm = df_diario["Linha DRE"].map(_normalizar_texto)
    mask_conta = linhas_norm == conta_norm
    if not mask_conta.any():
        mask_conta = linhas_norm.apply(lambda x: (conta_norm in x) or (x in conta_norm) if x else False)

    return df_diario[mask_loja & mask_conta]


def _filtrar_diario_por_loja_e_plano(df_diario, loja, plano):
    """Filtra o DIÁRIO pelo Centro de Custos (loja) e pelo nome exato do
    Plano de Contas, IGNORANDO a Linha DRE -- usado para planos de contas
    que fazem parte de um modelo de relatório mas não têm linha da DRE
    correspondente (ex.: "Mercadorias" no modelo de Compras). "loja" pode
    ser uma string ou uma lista de candidatos."""
    if df_diario is None or df_diario.empty:
        return df_diario

    plano_norm = _normalizar_texto(plano)

    mask_loja = _mask_loja_por_centro_custo(df_diario, loja)

    planos_norm = df_diario["Plano de Contas"].map(_normalizar_texto)
    mask_plano = planos_norm == plano_norm
    if not mask_plano.any():
        mask_plano = planos_norm.apply(lambda x: (plano_norm in x) or (x in plano_norm) if x else False)

    return df_diario[mask_loja & mask_plano]


def montar_composicao_plano_direto(df_diario, loja, plano, mapa_meses):
    """Retorna os valores mês a mês (na ordem de mapa_meses) de um Plano de
    Contas específico, direto do DIÁRIO, para uma loja -- sem passar pela
    Linha DRE."""
    df_filtrado = _filtrar_diario_por_loja_e_plano(df_diario, loja, plano)
    if df_filtrado is None or df_filtrado.empty:
        return [0.0] * len(mapa_meses)
    return [df_filtrado.loc[df_filtrado["Mês"] == m_col, "Valor Bruto"].sum() for m_col in mapa_meses.values()]


def montar_composicao_diario(df_diario, loja, conta, mapa_meses):
    """A partir do DIÁRIO, retorna {plano_de_contas: [valores_por_mes...]}
    para uma loja (Centro de Custos) e uma linha da DRE (conta) específicas,
    já na ordem de mapa_meses."""
    df_filtrado = _filtrar_diario_por_loja_e_conta(df_diario, loja, conta)
    if df_filtrado is None or df_filtrado.empty:
        return {}

    composicao = {}
    for plano, grupo in df_filtrado.groupby("Plano de Contas"):
        composicao[plano] = [grupo.loc[grupo["Mês"] == m_col, "Valor Bruto"].sum() for m_col in mapa_meses.values()]
    return composicao


# ============================================================================
# 4.3 PAINEL PARA TV — visão executiva "tech", cheia de KPIs, gauges e
# ranking de lojas, para deixar exibida numa tela/TV da empresa. Acessada por
# uma URL própria (?modo=tv), separada da sessão normal de quem está
# navegando no painel -- assim dá pra deixar aberta numa TV/monitor sem
# interferir (nem ser afetada) pelos filtros que um usuário está usando na
# sua própria sessão. Se auto-atualiza a cada 60s e tem botão de tela cheia.
# ============================================================================
ABAS_CONSOLIDADAS_TV = [
    "DRE CONSOLIDADO", "ABPR CONSOLIDADO", "VD CONSOLIDADO",
    "LJ CONSOLIDADO", "ABPR + VD", "LJ - G&A", "CONSOLIDADO - G&A",
]


def _obter_aba_consolidada_padrao(lista_abas):
    """Mesma regra de padrão usada na barra lateral (Visão Consolidada):
    a primeira visão consolidada, na ordem em que aparece nas abas
    disponíveis -- garante que o Painel de TV mostre exatamente o mesmo
    escopo que a pessoa vê por padrão no painel principal, evitando
    números "diferentes" entre os dois."""
    consolidadas_normalizadas = {_normalizar_nome_aba(n) for n in ABAS_CONSOLIDADAS_TV}
    consolidadas = [a for a in lista_abas if _normalizar_nome_aba(a) in consolidadas_normalizadas]
    if consolidadas:
        return consolidadas[0]
    return lista_abas[0] if lista_abas else None


def renderizar_painel_tv(path_orc, path_real, abas_disponiveis):
    meses_cols_tv = [
        "01/2026", "02/2026", "03/2026", "04/2026", "05/2026", "06/2026",
        "07/2026", "08/2026", "09/2026", "10/2026", "11/2026", "12/2026",
    ]
    nomes_meses_tv = [
        "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
        "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
    ]

    aba_padrao_tv = _obter_aba_consolidada_padrao(abas_disponiveis)
    if not aba_padrao_tv:
        st.error("Não foi possível carregar dados para o Painel de TV.")
        return

    agora = datetime.now(FUSO_BR)
    idx_mes_atual = min(max(agora.month - 1, 0), len(nomes_meses_tv) - 1)

    # ---- Seletores próprios do Painel de TV (visão e mês) -- minimalistas,
    # ficam no espaço vazio do cabeçalho, ao lado do relógio. O Painel de TV
    # NÃO acompanha mais os filtros do painel principal: a escolha é só
    # daqui mesmo, e persiste entre os ciclos de atualização automática.
    #
    # O CSS abaixo é global de propósito: esta função renderiza a página
    # inteira do modo TV e retorna antes do painel normal, então não há
    # risco de afetar o painel principal -- e um seletor global funciona de
    # forma bem mais confiável do que tentar mirar a coluna específica. ----
    st.markdown(
        f"""
        <style>
            div[data-baseweb="select"] > div {{
                min-height: 30px !important;
                font-size: 11px !important; letter-spacing: 0.2px;
                padding: 2px 8px !important;
                background: transparent !important;
                border: 1px solid {COLORS["border"]} !important;
                border-radius: 6px !important;
                color: {COLORS["text_muted"]} !important;
                display: flex !important; align-items: center !important;
            }}
            /* O texto do valor selecionado precisa de linha cheia, senão as
               letras com descida (g, ç, p) ficam cortadas na borda de baixo. */
            div[data-baseweb="select"] div[data-baseweb="select"] > div > div,
            div[data-baseweb="select"] > div > div {{
                line-height: 1.5 !important; overflow: visible !important;
            }}
            div[data-baseweb="select"] svg {{ width: 13px !important; height: 13px !important; opacity: 0.5; }}
            div[data-testid="stSelectbox"] {{ margin-bottom: 0 !important; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    col_head_a, col_head_filtro_visao, col_head_filtro_mes, col_head_b = st.columns(
        [3.5, 1.15, 0.8, 1.1], vertical_alignment="center"
    )
    with col_head_filtro_visao:
        aba_escolhida = st.selectbox(
            "Visão", abas_disponiveis,
            index=abas_disponiveis.index(aba_padrao_tv) if aba_padrao_tv in abas_disponiveis else 0,
            key="tv_sel_visao", label_visibility="collapsed",
        )
    with col_head_filtro_mes:
        mes_escolhido_tv = st.selectbox(
            "Mês", nomes_meses_tv, index=idx_mes_atual,
            key="tv_sel_mes", label_visibility="collapsed",
        )

    abas_para_tv = [aba_escolhida]
    list_df_orc_tv, list_df_real_tv = carregar_dados_abas(path_orc, path_real, abas_para_tv)

    df_ref_tv = list_df_real_tv[0] if list_df_real_tv else pd.DataFrame()
    colunas_validas_tv = [m for m in meses_cols_tv if m in df_ref_tv.columns]
    m_map_tv = {n: c for n, c in zip(nomes_meses_tv, meses_cols_tv) if c in colunas_validas_tv}

    # Acumula (YTD) até o mês escolhido no seletor -- se aquele mês ainda
    # não tiver dado na planilha, cai pro último mês que realmente existe.
    idx_mes_escolhido = min(nomes_meses_tv.index(mes_escolhido_tv), len(m_map_tv) - 1) if m_map_tv else 0
    cols_ytd = list(m_map_tv.values())[: idx_mes_escolhido + 1]
    legenda_periodo_tv = (
        f"Acumulado até {nomes_meses_tv[idx_mes_escolhido].capitalize()}/{list(m_map_tv.values())[idx_mes_escolhido].split('/')[-1]}"
        if m_map_tv else "Sem dados"
    )

    # ---- CSS "quiosque tech": some com sidebar/header, grade de fundo, glow ----
    st.markdown(
        f"""
        <style>
            [data-testid="stSidebar"], header[data-testid="stHeader"], footer {{ display: none !important; }}
            div[data-testid="stAppViewContainer"] {{ padding-top: 0 !important; }}
            div[data-testid="stAppViewContainer"] > .main {{ padding-top: 0 !important; }}
            section.main > div.block-container, .block-container {{
                padding: 0.5rem 1.8rem 0.4rem 1.8rem !important; max-width: 100% !important;
            }}
            div[data-testid="stVerticalBlock"] {{ gap: 0.5rem !important; }}
            .stApp {{
                background:
                    radial-gradient(circle at 15% 0%, rgba(76,141,255,0.09) 0%, transparent 42%),
                    radial-gradient(circle at 90% 100%, rgba(62,207,142,0.06) 0%, transparent 45%),
                    linear-gradient(180deg, #0A0D16 0%, #05070c 100%) !important;
            }}
            @keyframes tv-pulse {{ 0%,100% {{ opacity: 1; }} 50% {{ opacity: 0.35; }} }}
            @keyframes tv-marquee {{ 0% {{ transform: translateX(100%); }} 100% {{ transform: translateX(-100%); }} }}
            .tv-header {{
                display: flex; justify-content: space-between; align-items: center;
                padding: 2px 4px 10px 4px; border-bottom: 1px solid {COLORS["border"]}; margin-bottom: 10px;
            }}
            .tv-header .brand {{ display:flex; align-items:center; gap:12px; }}
            .tv-header img.logo {{ width: 38px; height: 38px; border-radius: 50%; box-shadow: 0 0 14px rgba(76,141,255,0.35); }}
            .tv-header h1 {{ font-size: 23px; font-weight: 800; color: {COLORS["text"]}; margin: 0; letter-spacing: 0.3px; }}
            .tv-header .sub {{ color: {COLORS["text_muted"]}; font-size: 12.5px; margin-top: 2px; }}
            .tv-live-pill {{
                display: inline-flex; align-items: center; gap: 6px; background: rgba(62,207,142,0.12);
                border: 1px solid {COLORS["positive"]}; color: {COLORS["positive"]}; border-radius: 20px;
                padding: 3px 12px; font-size: 11px; font-weight: 700; letter-spacing: 0.6px; margin-left: 12px;
            }}
            .tv-live-pill .dot {{ width: 7px; height: 7px; border-radius: 50%; background: {COLORS["positive"]}; animation: tv-pulse 1.4s infinite; }}
            .tv-kpi-grid {{ display: flex; gap: 14px; margin-bottom: 12px; }}
            .tv-kpi {{
                flex: 1; background: {COLORS["surface"]};
                border: 1px solid {COLORS["border"]}; border-radius: 12px; padding: 14px 16px;
                border-top: 3px solid var(--tv-accent, {COLORS["primary"]});
            }}
            .tv-kpi .lbl {{ font-size: 10.5px; font-weight: 700; letter-spacing: 0.6px; text-transform: uppercase; color: {COLORS["text_muted"]}; }}
            .tv-kpi .val {{ font-size: 25px; font-weight: 800; margin-top: 5px; letter-spacing: -0.5px; }}
            .tv-kpi .sub {{ font-size: 11.5px; margin-top: 4px; color: {COLORS["muted_line"]}; display:flex; align-items:center; gap:4px; }}
            .tv-section-title {{
                font-size: 12.5px; font-weight: 700; color: {COLORS["text_muted"]}; text-transform: uppercase;
                letter-spacing: 0.6px; margin: 2px 0 8px 2px; border-left: 3px solid {COLORS["primary"]}; padding-left: 8px;
            }}
            .tv-panel {{
                background: {COLORS["surface"]};
                border: 1px solid {COLORS["border"]}; border-radius: 12px; padding: 12px 16px;
                margin-bottom: 12px; height: 100%;
            }}
            /* ---- Barras de atingimento minimalistas (no lugar dos velocímetros) ---- */
            .tv-atg-grid {{ display: flex; gap: 28px; }}
            .tv-atg-item {{ flex: 1; }}
            .tv-atg-head {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 8px; }}
            .tv-atg-title {{ font-size: 12.5px; color: {COLORS["text_muted"]}; font-weight: 700; text-transform: uppercase; letter-spacing: 0.4px; }}
            .tv-atg-pct {{ font-size: 20px; font-weight: 800; font-family: 'Consolas','Courier New',monospace; }}
            .tv-atg-track {{ position: relative; height: 9px; background: {COLORS["border"]}; border-radius: 6px; }}
            .tv-atg-fill {{ height: 100%; border-radius: 6px; }}
            .tv-atg-marker {{ position: absolute; top: -4px; bottom: -4px; width: 2px; background: {COLORS["warning"]}; }}
            .tv-atg-foot {{ display: flex; justify-content: space-between; margin-top: 7px; font-size: 11.5px; color: {COLORS["muted_line"]}; }}
            /* ---- Lista de composição de custos (ao lado do donut) ---- */
            .tv-cost-row {{
                display: flex; align-items: center; gap: 10px; padding: 8px 2px;
                border-bottom: 1px dashed {COLORS["border_soft"]};
            }}
            .tv-cost-dot {{ width: 11px; height: 11px; border-radius: 3px; flex-shrink: 0; }}
            .tv-cost-nome {{ flex: 1.3; font-size: 15px; color: {COLORS["text"]}; font-weight: 600; }}
            .tv-cost-pct {{ flex: 0.6; font-size: 13.5px; color: {COLORS["text_muted"]}; text-align: right; }}
            .tv-cost-val {{ flex: 1; font-size: 15px; color: {COLORS["text"]}; text-align: right; font-family: 'Consolas','Courier New',monospace; }}
            .tv-cost-desvio {{ flex: 0.9; font-size: 13.5px; text-align: right; font-family: 'Consolas','Courier New',monospace; }}
            .tv-rank-row {{ display:flex; align-items:center; gap:10px; padding: 8px 2px; border-bottom: 1px dashed {COLORS["border_soft"]}; }}
            .tv-rank-name {{ flex:1; font-size:14.5px; color:{COLORS["text"]}; font-weight:600; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
            .tv-rank-bar-bg {{ flex:1.1; background:{COLORS["border"]}; border-radius:4px; height:7px; overflow:hidden; }}
            .tv-rank-bar-fill {{ height:100%; border-radius:4px; background: linear-gradient(90deg, {COLORS["secondary"]}, {COLORS["warning"]}); }}
            .tv-rank-pct-rec {{ flex:0.6; font-size:12.5px; color:{COLORS["text_muted"]}; text-align:right; white-space:nowrap; }}
            .tv-rank-val {{ font-size:13.5px; color:{COLORS["muted_line"]}; width: 130px; text-align:right; font-family:'Consolas','Courier New',monospace; }}
            .tv-ticker-wrap {{
                overflow: hidden; white-space: nowrap; border-top: 1px solid {COLORS["border"]};
                border-bottom: 1px solid {COLORS["border"]}; padding: 8px 0; margin-top: 4px; background: rgba(255,255,255,0.015);
            }}
            .tv-ticker {{ display:inline-block; padding-left: 100%; animation: tv-marquee 90s linear infinite; font-size: 15px; color: {COLORS["text_muted"]}; }}
            .tv-ticker b {{ color: {COLORS["text"]}; }}
            .tv-ticker .tv-tick-sep {{ color: {COLORS["primary"]}; margin: 0 30px; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ---------------- Cálculo dos indicadores ----------------
    rec_bruta_real = get_valor_consolidado_multi(list_df_real_tv, "1 - Receita Operacional Bruta", cols_ytd)
    rec_liq_real = get_valor_consolidado_multi(list_df_real_tv, "3 - Receita Operacional Liquida", cols_ytd)
    rec_liq_orc = get_valor_consolidado_multi(list_df_orc_tv, "3 - Receita Operacional Liquida", cols_ytd)
    ebitda_real = get_valor_consolidado_multi(list_df_real_tv, "11 - EBITDA", cols_ytd)
    ebitda_orc = get_valor_consolidado_multi(list_df_orc_tv, "11 - EBITDA", cols_ytd)
    margem_ebitda = (ebitda_real / rec_liq_real * 100) if rec_liq_real else 0
    margem_ebitda_orc = (ebitda_orc / rec_liq_orc * 100) if rec_liq_orc else 0
    desvio_ebitda = ebitda_real - ebitda_orc
    pct_desvio_ebitda = (desvio_ebitda / abs(ebitda_orc) * 100) if ebitda_orc else 0
    pct_atingimento_rec = (rec_liq_real / rec_liq_orc * 100) if rec_liq_orc else 0
    pct_atingimento_eb = (ebitda_real / ebitda_orc * 100) if ebitda_orc else 0

    cmv_tv = abs(get_valor_consolidado_multi(list_df_real_tv, "4 - ", cols_ytd, exato_linha_sintetica=True)) or \
        abs(get_valor_consolidado_multi(list_df_real_tv, "4 - Custo das Vendas", cols_ytd))
    cmv_tv_o = abs(get_valor_consolidado_multi(list_df_orc_tv, "4 - ", cols_ytd, exato_linha_sintetica=True)) or \
        abs(get_valor_consolidado_multi(list_df_orc_tv, "4 - Custo das Vendas", cols_ytd))
    desp_var_tv = abs(get_valor_consolidado_multi(list_df_real_tv, "6 - Despesas Variáveis", cols_ytd))
    desp_var_tv_o = abs(get_valor_consolidado_multi(list_df_orc_tv, "6 - Despesas Variáveis", cols_ytd))
    desp_op_tv_kpi = abs(get_valor_consolidado_multi(list_df_real_tv, "8 - Despesas Operacionais", cols_ytd))
    desp_op_tv_o = abs(get_valor_consolidado_multi(list_df_orc_tv, "8 - Despesas Operacionais", cols_ytd))
    deprec_tv = abs(get_valor_consolidado_multi(list_df_real_tv, "13 - Depreciação e Amortização", cols_ytd))
    total_saidas_tv = cmv_tv + desp_var_tv + desp_op_tv_kpi + deprec_tv
    total_custos_desp_tv = cmv_tv + desp_var_tv + desp_op_tv_kpi

    with col_head_a:
        st.markdown(
            html_compacto(f"""
            <div class="tv-header">
                <div class="brand">
                    <img class="logo" src="data:image/jpeg;base64,{LOGO_BEEA_B64}" alt="Grupo Beea" />
                    <div>
                        <h1>Grupo B&amp;A · Painel Executivo <span class="tv-live-pill"><span class="dot"></span>AO VIVO</span></h1>
                        <div class="sub">{legenda_periodo_tv} · atualiza sozinho a cada 90s</div>
                    </div>
                </div>
            </div>
            """),
            unsafe_allow_html=True,
        )
    with col_head_b:
        # Relógio de verdade "ao vivo": roda dentro de um componente (só ele
        # consegue executar JavaScript de fato) e atualiza a cada segundo.
        components.html(
            f"""
            <div style="text-align:right; font-family:'Consolas','Courier New',monospace;
                        color:{COLORS['text_muted']}; padding-top:2px;">
                <div id="tvClockLive" style="color:{COLORS['primary']}; font-size:24px; font-weight:800; letter-spacing:2px;">--:--:--</div>
                <div id="tvDateLive" style="font-size:11px;"></div>
            </div>
            <script>
            function _tvAtualizarRelogio() {{
                var agora = new Date();
                function pad(n) {{ return String(n).padStart(2, '0'); }}
                var elC = document.getElementById('tvClockLive');
                var elD = document.getElementById('tvDateLive');
                if (elC) {{ elC.textContent = pad(agora.getHours()) + ':' + pad(agora.getMinutes()) + ':' + pad(agora.getSeconds()); }}
                if (elD) {{
                    var dias = ['Domingo','Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira','Sexta-feira','Sábado'];
                    elD.textContent = dias[agora.getDay()] + ', ' + pad(agora.getDate()) + '/' + pad(agora.getMonth() + 1) + '/' + agora.getFullYear();
                }}
            }}
            _tvAtualizarRelogio();
            setInterval(_tvAtualizarRelogio, 1000);
            </script>
            """,
            height=54,
        )

    def _tv_kpi(cor_var, label, valor, sub, accent, icone=""):
        return (
            f'<div class="tv-kpi" style="--tv-accent:{accent};">'
            f'<div class="lbl">{label}</div>'
            f'<div class="val" style="color:{cor_var};">{valor}</div>'
            f'<div class="sub">{icone} {sub}</div></div>'
        )

    seta_rec = "▲" if rec_liq_real >= rec_liq_orc else "▼"
    seta_eb = "▲" if ebitda_real >= ebitda_orc else "▼"
    pct_custos_sobre_receita = (total_custos_desp_tv / rec_liq_real * 100) if rec_liq_real else 0

    st.markdown(
        '<div class="tv-kpi-grid">'
        + _tv_kpi(COLORS["text"], "Receita Bruta (YTD)", formata_m(rec_bruta_real),
                  "Antes de deduções", COLORS["muted_line"], "💰")
        + _tv_kpi(cor_variacao(rec_liq_real - rec_liq_orc), "Receita Líquida (YTD)", formata_m(rec_liq_real),
                  f"{seta_rec} {pct_atingimento_rec:.0f}% do orçado ({formata_m(rec_liq_orc)})", COLORS["primary"], "")
        + _tv_kpi(cor_variacao(ebitda_real - ebitda_orc), "EBITDA (YTD)", formata_m(ebitda_real),
                  f"{seta_eb} {pct_atingimento_eb:.0f}% do orçado ({formata_m(ebitda_orc)})", COLORS["positive"], "")
        + _tv_kpi(cor_variacao(margem_ebitda - margem_ebitda_orc), "Margem EBITDA", f"{margem_ebitda:.1f}%",
                  f"Orçado: {margem_ebitda_orc:.1f}% · Desvio: {formata_brl(desvio_ebitda)}", COLORS["secondary"], "")
        + _tv_kpi(cor_variacao(-(pct_custos_sobre_receita)), "Custos + Despesas / Receita", f"{pct_custos_sobre_receita:.1f}%",
                  f"{formata_m(total_custos_desp_tv)} sobre {formata_m(rec_liq_real)}", COLORS["warning"], "")
        + "</div>",
        unsafe_allow_html=True,
    )

    # ---------------- Atingimento vs. Orçado (barras minimalistas) ----------------
    def _tv_barra_atingimento(titulo, valor_real, valor_orc, pct_atg, cor):
        escala_max = max(150.0, pct_atg + 20.0)
        largura = max(2.0, min(100.0, pct_atg / escala_max * 100))
        marca = min(98.0, 100.0 / escala_max * 100)
        return (
            '<div class="tv-atg-item">'
            f'<div class="tv-atg-head"><span class="tv-atg-title">{titulo}</span>'
            f'<span class="tv-atg-pct" style="color:{cor};">{pct_atg:.0f}%</span></div>'
            f'<div class="tv-atg-track"><div class="tv-atg-fill" style="width:{largura:.1f}%; background:{cor};"></div>'
            f'<div class="tv-atg-marker" style="left:{marca:.1f}%;"></div></div>'
            f'<div class="tv-atg-foot"><span>{formata_m(valor_real)} realizado</span>'
            f'<span>Meta: {formata_m(valor_orc)}</span></div></div>'
        )

    st.markdown(
        '<div class="tv-panel"><div class="tv-atg-grid">'
        + _tv_barra_atingimento("🎯 Atingimento de Receita Líquida", rec_liq_real, rec_liq_orc, pct_atingimento_rec, COLORS["primary"])
        + _tv_barra_atingimento("🎯 Atingimento de EBITDA", ebitda_real, ebitda_orc, pct_atingimento_eb, COLORS["positive"])
        + "</div></div>",
        unsafe_allow_html=True,
    )

    # ---------------- Evolução Mensal (grande) + Composição de Custos (detalhada) ----------------
    cgtv1, cgtv2 = st.columns([1.6, 1])

    with cgtv1:
        st.markdown('<div class="tv-section-title">📈 Evolução Mensal — Receita vs. EBITDA</div>', unsafe_allow_html=True)
        rec_m_tv, eb_m_tv, rot_m_tv = [], [], []
        for m_nome, c in m_map_tv.items():
            rec_m_tv.append(get_valor_consolidado_multi(list_df_real_tv, "3 - Receita Operacional Liquida", [c]))
            eb_m_tv.append(get_valor_consolidado_multi(list_df_real_tv, "11 - EBITDA", [c]))
            rot_m_tv.append(m_nome.capitalize()[:3])
        fig_tv_line = go.Figure()
        fig_tv_line.add_trace(go.Scatter(
            x=rot_m_tv, y=rec_m_tv, name="Receita Líquida", mode="lines+markers+text",
            text=[formata_m(v) for v in rec_m_tv], textposition="top center",
            textfont=dict(size=10, color=COLORS["text_muted"]),
            line=dict(color=COLORS["primary"], width=2.5), marker=dict(size=5, color=COLORS["primary"]),
        ))
        fig_tv_line.add_trace(go.Scatter(
            x=rot_m_tv, y=eb_m_tv, name="EBITDA", mode="lines+markers+text",
            text=[formata_m(v) for v in eb_m_tv], textposition="bottom center",
            textfont=dict(size=10, color=COLORS["text_muted"]),
            line=dict(color=COLORS["positive"], width=2.5, dash="dot"), marker=dict(size=5, color=COLORS["positive"]),
        ))
        estilo_grafico(
            fig_tv_line, height=430,
            margin=dict(l=20, r=20, t=20, b=50),
            xaxis=dict(showgrid=False, fixedrange=True, tickfont=dict(size=13, color=COLORS["text_muted"])),
            yaxis=dict(showgrid=False, showticklabels=False, fixedrange=True, zeroline=False),
            legend=dict(orientation="h", yanchor="bottom", y=-0.16, xanchor="center", x=0.5, font=dict(size=13)),
        )
        st.plotly_chart(fig_tv_line, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        # ---- Desvio Mensal de EBITDA (barras) — usa o espaço abaixo do
        # gráfico de linha e complementa a evolução com "quanto acima/abaixo
        # do orçado cada mês ficou", mês a mês. ----
        st.markdown('<div class="tv-section-title">📐 Desvio Mensal — EBITDA Real vs. Orçado</div>', unsafe_allow_html=True)
        desvio_m_tv, cores_desvio_tv = [], []
        for m_nome, c in m_map_tv.items():
            eb_real_m = get_valor_consolidado_multi(list_df_real_tv, "11 - EBITDA", [c])
            eb_orc_m = get_valor_consolidado_multi(list_df_orc_tv, "11 - EBITDA", [c])
            d = eb_real_m - eb_orc_m
            desvio_m_tv.append(d)
            cores_desvio_tv.append(COLORS["positive"] if d >= 0 else COLORS["negative"])
        fig_tv_desvio = go.Figure()
        fig_tv_desvio.add_trace(go.Bar(
            x=rot_m_tv, y=desvio_m_tv, marker=dict(color=cores_desvio_tv),
            text=[formata_m(v) for v in desvio_m_tv], textposition="outside",
            textfont=dict(size=10, color=COLORS["text_muted"]),
        ))
        estilo_grafico(
            fig_tv_desvio, height=160,
            margin=dict(l=20, r=20, t=10, b=30),
            xaxis=dict(showgrid=False, fixedrange=True, tickfont=dict(size=12, color=COLORS["text_muted"])),
            yaxis=dict(showgrid=False, showticklabels=False, fixedrange=True, zeroline=True,
                       zerolinecolor=COLORS["border"], zerolinewidth=1),
            showlegend=False,
            bargap=0.35,
        )
        st.plotly_chart(fig_tv_desvio, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

    with cgtv2:
        st.markdown('<div class="tv-section-title">🥧 Composição de Custos & Saídas</div>', unsafe_allow_html=True)
        fig_tv_donut = go.Figure(data=[go.Pie(
            labels=["CMV", "Desp. Variáveis", "Desp. Operacionais", "Deprec./Amort."],
            values=[cmv_tv, desp_var_tv, desp_op_tv_kpi, deprec_tv], hole=0.66,
            marker=dict(colors=[COLORS["primary"], COLORS["muted_line"], COLORS["secondary"], COLORS["border_soft"]],
                        line=dict(color=COLORS["surface"], width=2)),
            textinfo="percent", textfont=dict(size=11), showlegend=False,
        )])
        fig_tv_donut.add_annotation(
            text=f"<b>{formata_m(total_saidas_tv)}</b><br><span style='font-size:10px;color:{COLORS['text_muted']}'>Total Saídas</span>",
            showarrow=False, font=dict(color=COLORS["text"], size=13, family=FONT_STACK),
        )
        estilo_grafico(fig_tv_donut, height=200, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig_tv_donut, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        # Detalhamento: cada categoria com % da receita líquida e desvio vs. orçado
        # (é o "mais detalhe de custos" pedido -- não só o donut).
        categorias_custo = [
            ("CMV", cmv_tv, cmv_tv_o, COLORS["primary"]),
            ("Despesas Variáveis", desp_var_tv, desp_var_tv_o, COLORS["muted_line"]),
            ("Despesas Operacionais", desp_op_tv_kpi, desp_op_tv_o, COLORS["secondary"]),
        ]
        linhas_custo = ['<div class="tv-panel" style="padding-top:8px;">']
        for nome_cat, v_real, v_orc, cor_cat in categorias_custo:
            pct_da_receita = (v_real / rec_liq_real * 100) if rec_liq_real else 0
            desvio_cat = v_orc - v_real  # custo menor que orçado = positivo (bom)
            cor_desvio = cor_variacao(desvio_cat)
            linhas_custo.append(
                '<div class="tv-cost-row">'
                f'<div class="tv-cost-dot" style="background:{cor_cat};"></div>'
                f'<div class="tv-cost-nome">{nome_cat}</div>'
                f'<div class="tv-cost-pct">{pct_da_receita:.1f}% rec.</div>'
                f'<div class="tv-cost-val">{formata_m(v_real)}</div>'
                f'<div class="tv-cost-desvio" style="color:{cor_desvio};">{formata_m(desvio_cat)}</div>'
                "</div>"
            )
        linhas_custo.append(
            f'<div style="text-align:right; font-size:10px; color:{COLORS["text_muted"]}; margin-top:4px;">'
            f"Desvio = orçado − realizado (positivo é favorável)</div></div>"
        )
        st.markdown("".join(linhas_custo), unsafe_allow_html=True)

        # ---- Detalhamento de Despesas Operacionais pelos grupos reais da
        # DRE (Pessoal, Ocupação, Comercial etc.) -- é o "de onde vem o
        # custo" que o donut sozinho não mostra. ----
        col_nome_tv = "Nome" if "Nome" in df_ref_tv.columns else df_ref_tv.columns[0]
        linhas_dre_tv = df_ref_tv[col_nome_tv].dropna().astype(str).unique()
        subgrupos_despop = _subgrupos_nivel2(linhas_dre_tv, "8")

        detalhe_despop = []
        for sub in subgrupos_despop:
            v_sub = abs(get_valor_consolidado_multi(list_df_real_tv, sub, cols_ytd, exato_linha_sintetica=True))
            if v_sub:
                detalhe_despop.append((_nome_sem_numero_dre(sub), v_sub))
        detalhe_despop.sort(key=lambda x: x[1], reverse=True)
        top_despop = detalhe_despop[:5]

        if top_despop:
            st.markdown(
                '<div class="tv-section-title" style="margin-top:10px;">🏢 Despesas Operacionais — Principais Grupos</div>',
                unsafe_allow_html=True,
            )
            max_despop = max(v for _, v in top_despop) or 1.0
            linhas_despop = ['<div class="tv-panel" style="padding-top:8px;">']
            for nome_grp, v_grp in top_despop:
                pct_do_despop = (v_grp / desp_op_tv_kpi * 100) if desp_op_tv_kpi else 0
                # % que o grupo representa da Receita Líquida -- mesmo cálculo
                # já usado no card "Resumo Gerencial" (tv-cost-pct) logo acima.
                pct_da_receita = (v_grp / rec_liq_real * 100) if rec_liq_real else 0
                pct_barra = max(3, min(100, v_grp / max_despop * 100))
                linhas_despop.append(
                    '<div class="tv-rank-row">'
                    f'<div class="tv-rank-name" title="{nome_grp}">{nome_grp}</div>'
                    f'<div class="tv-rank-bar-bg"><div class="tv-rank-bar-fill" style="width:{pct_barra:.0f}%;"></div></div>'
                    f'<div class="tv-rank-pct-rec">{pct_da_receita:.1f}% rec.</div>'
                    f'<div class="tv-rank-val">{formata_m(v_grp)} · {pct_do_despop:.0f}%</div>'
                    "</div>"
                )
            linhas_despop.append("</div>")
            st.markdown("".join(linhas_despop), unsafe_allow_html=True)

    # ---------------- Ticker de destaques + controles de tela cheia (rodapé) ----------------
    meses_com_receita = {
        m_nome: get_valor_consolidado_multi(list_df_real_tv, "3 - Receita Operacional Liquida", [c])
        for m_nome, c in m_map_tv.items()
    }
    meses_validos = {m: v for m, v in meses_com_receita.items() if v != 0}

    # Projeções simples (extrapolação linear do acumulado do ano até agora)
    # e outros números que NÃO aparecem em nenhum card/lista da tela --
    # o letreiro é pra trazer coisa nova, não repetir o que já tá visível.
    n_meses_decorridos = idx_mes_atual + 1
    n_meses_restantes = max(0, 12 - n_meses_decorridos)
    media_mensal_receita = (rec_liq_real / n_meses_decorridos) if n_meses_decorridos else 0
    media_mensal_ebitda = (ebitda_real / n_meses_decorridos) if n_meses_decorridos else 0
    projecao_receita_ano = media_mensal_receita * 12
    projecao_ebitda_ano = media_mensal_ebitda * 12

    destaques = []
    if meses_validos:
        melhor_mes = max(meses_validos, key=meses_validos.get)
        destaques.append(f"🏆 Melhor mês em receita: <b>{melhor_mes.capitalize()}</b> ({formata_m(meses_validos[melhor_mes])})")
        if len(meses_validos) > 1:
            pior_mes = min(meses_validos, key=meses_validos.get)
            destaques.append(f"📉 Pior mês em receita: <b>{pior_mes.capitalize()}</b> ({formata_m(meses_validos[pior_mes])})")
    destaques.append(f"🔮 Projeção de Receita Líquida no ano: <b>{formata_m(projecao_receita_ano)}</b> (ritmo atual)")
    destaques.append(f"🔮 Projeção de EBITDA no ano: <b>{formata_m(projecao_ebitda_ano)}</b> (ritmo atual)")
    destaques.append(f"📅 Ticket médio mensal de Receita Líquida: <b>{formata_m(media_mensal_receita)}</b>")
    destaques.append(f"⏳ Faltam <b>{n_meses_restantes}</b> mês(es) para fechar o ano")
    destaques.append(f"🗓️ Acumulado de <b>{n_meses_decorridos}</b> mês(es) no período (Jan a {nomes_meses_tv[idx_mes_atual].capitalize()})")

    ticker_html = f'<span class="tv-tick-sep">·</span>'.join(destaques)
    st.markdown(
        f"""<div class="tv-ticker-wrap"><div class="tv-ticker">{ticker_html}</div></div>""",
        unsafe_allow_html=True,
    )

    # Botão de tela cheia + aviso -- tudo dentro do MESMO componente (só ele
    # roda JavaScript de verdade). O aviso/botão somem sozinhos assim que a
    # página entra em tela cheia (por esse botão OU pelo F11), e reaparecem
    # ao sair (Esc). Se o navegador bloquear o pedido automático de tela
    # cheia, mostra um aviso na hora (em vez de falhar calado) recomendando
    # o F11, que é sempre a forma mais confiável.
    components.html(
        f"""
        <div id="tvFsWrap" style="display:flex; align-items:center; justify-content:center; gap:12px;
                    font-family:{FONT_STACK}; padding-top:6px;">
            <button id="tvFsBtn" style="background:{COLORS['primary']}; color:#fff; border:none;
                    border-radius:20px; padding:6px 18px; font-size:12.5px; font-weight:700; cursor:pointer;">
                ⛶ Tela Cheia
            </button>
            <span id="tvFsMsg" style="color:{COLORS['text_muted']}; font-size:11.5px;">
                (F11 também funciona · Esc para sair)
            </span>
        </div>
        <script>
        function _tvAlvo() {{
            try {{ if (window.top && window.top.document) {{ window.top.document.title; return window.top; }} }} catch (e) {{}}
            try {{ if (window.parent && window.parent.document) {{ window.parent.document.title; return window.parent; }} }} catch (e) {{}}
            return window;
        }}
        function _tvEmTela() {{
            try {{
                var d = _tvAlvo().document;
                return !!(d.fullscreenElement || d.webkitFullscreenElement || d.msFullscreenElement);
            }} catch (e) {{ return false; }}
        }}
        function _tvAtualizarUI() {{
            var wrap = document.getElementById('tvFsWrap');
            if (wrap) wrap.style.display = _tvEmTela() ? 'none' : 'flex';
        }}
        var btn = document.getElementById('tvFsBtn');
        if (btn) {{
            btn.addEventListener('click', function () {{
                var msg = document.getElementById('tvFsMsg');
                try {{
                    var w = _tvAlvo();
                    var el = w.document.documentElement;
                    var pedido = el.requestFullscreen ? el.requestFullscreen()
                        : (el.webkitRequestFullscreen ? el.webkitRequestFullscreen() : null);
                    if (pedido && pedido.catch) {{
                        pedido.catch(function () {{
                            if (msg) msg.textContent = 'Bloqueado pelo navegador -- pressione F11';
                        }});
                    }} else if (!pedido) {{
                        if (msg) msg.textContent = 'Não suportado aqui -- pressione F11';
                    }}
                }} catch (e) {{
                    if (msg) msg.textContent = 'Não foi possível -- pressione F11';
                }}
            }});
        }}
        try {{ _tvAlvo().document.addEventListener('fullscreenchange', _tvAtualizarUI); }} catch (e) {{}}
        try {{ _tvAlvo().document.addEventListener('webkitfullscreenchange', _tvAtualizarUI); }} catch (e) {{}}
        document.addEventListener('fullscreenchange', _tvAtualizarUI);
        _tvAtualizarUI();
        setInterval(_tvAtualizarUI, 1500);
        </script>
        """,
        height=44,
    )

    st.markdown(
        html_compacto(f"""
        <div style="text-align:center;margin-top:2px;color:{COLORS['text_muted']};font-size:11px;">
            Painel para exibição (somente leitura) · Atualiza automaticamente a cada 90 segundos (acompanha os filtros do painel principal) ·
            <a href="?" style="color:{COLORS['text_muted']};">Sair do modo TV</a>
        </div>
        """),
        unsafe_allow_html=True,
    )

    # Atualização automática SEM recarregar a página (evita perder a tela
    # cheia e evita cair de novo na tela de login a cada ciclo, como
    # acontecia com o <meta refresh> -- esse rerun acontece dentro da mesma
    # sessão/aba, sem navegação de página).
    time.sleep(90)
    st.rerun()


if st.query_params.get("modo") == "tv":
    renderizar_painel_tv(path_orc, path_real, abas_disponiveis)
    st.stop()

if not checar_login():
    st.stop()

# Apos um login bem-sucedido nesta mesma execucao: salva ou apaga as
# credenciais no navegador, conforme a caixa "Lembrar de mim".
_creds_pendentes = st.session_state.pop("_credenciais_para_salvar", None)
if _creds_pendentes:
    _salvar_credenciais_no_navegador(*_creds_pendentes)
if st.session_state.pop("_esquecer_credenciais", False):
    _esquecer_credenciais_no_navegador()

usuario_atual = st.session_state.get("usuario_logado") or {"email": "", "perfil": "admin"}
eh_admin = usuario_atual["perfil"] == "admin"


MODELOS_RELATORIO = {
    "🛒 Relatório de Custos - Compras": {
        "linhas_dre": [
            "6.6 - Material de Embalagem",
            "6.11 - Catálogos e Revistas",
            "6.13 - Amostras",
            "6.14 - Flaconetes",
        ],
        # O plano de contas "Mercadorias" faz parte do modelo de Compras, mas
        # não tem uma "Linha DRE" correspondente na Tabela_Contas/DIÁRIO --
        # por isso ele é puxado à parte, direto pelo nome do Plano de Contas,
        # mesmo sem nenhuma linha da DRE selecionada apontar para ele.
        "forcar_planos_contas": ["Mercadorias"],
        "permitir_lancamento_manual": False,
    },
    "🚚 Relatório de Custos - Suprimentos": {
        "linhas_dre": [
            "6.8 - Serviço de Entrega",
            "8.1.3 - Limpeza e Conservação",
            "8.1.4 - Manutenção e Reparos",
            "8.5.3 - Combustível",
            "8.5.4 - Manutenção Veículos",
            "8.6.1 - Material de Escritório",
            "8.6.4 - Copa e Cozinha",
            "8.7.1 - Despesas com Viagens",
            "8.8.10 - Serviços de Transporte",
            "8.8.11 - Outros Serviços Terceirizados",
        ],
        "forcar_planos_contas": [
            "Adiantamento de Benfeitorias em Imóvel Próprio",
            "Adiantamento de de Benfeitorias em Imóveis de Terceiros",
            "Benfeitorias em Imóveis de Terceiros",
            "Benfeitorias em imóvel próprio",
            "Ativo - Padronização de Franquias",
        ],
        "permitir_lancamento_manual": False,
    },
    "👥 Relatório de Custos - RH": {
        "linhas_dre": [
            "6.1 - Comissões sobre Vendas",
            "8.3 - Pessoal",
            "8.3.1 - Salários",
            "8.3.1.1 - Salário",
            "8.3.1.2 - Adiantamento Salarial",
            "8.3.1.3 - 13º Salário",
            "8.3.1.4 - Hora Extra",
            "8.3.1.5 - DSR",
            "8.3.1.6 - Férias",
            "8.3.1.7 - Descontos Gerais Sobre a Folha",
            "8.3.2 - Encargos Sociais",
            "8.3.2.1 - INSS",
            "8.3.2.2 - FGTS",
            "8.3.2.3 - IRRF Salários",
            "8.3.2.4 - Contribuição Sindical / Assistencial",
            "8.3.2.5 - Desconto INSS",
            "8.3.2.6 - Desconto IRRF",
            "8.3.3 - Benefícios",
            "8.3.3.1 - Vale Transporte",
            "8.3.3.2 - Vale Refeição",
            "8.3.3.3 - Plano de Saúde",
            "8.3.3.4 - Ajuda de Custo para Funcionários",
            "8.3.3.5 - Outros Benefícios",
            "8.3.3.6 - Desconto sobre Beneficios",
            "8.3.3.7 - Prêmios / Bônus",
            "8.3.4 - Movimentação de Pessoal",
            "8.3.4.1 - Indenizações / Rescisões",
            "8.3.4.2 - Multa de FGTS",
            "8.3.4.3 - Exames Admissionais / Demissionais",
            "8.3.4.4 - Recrutamento e Seleção",
            "8.3.4.5 - Descontos Sobre Rescisões",
            "8.3.4.6 - Temporários / Estagiários",
            "8.3.5 - Uniformes",
            "8.3.6 - Treinamento",
            "8.3.7 - Pro Labore",
            "8.3.8 - Saúde e Segurança do Trabalho",
            "8.3.9 - Outras Despesas com Pessoal",
            "8.3.10 - Contratação Pessoa Jurídica",
            "8.3.11 - Encontro e Confraternização de Time",
            "8.6.6 - Outras Despesas Administrativas",
            "8.8.2 - Auditoria / Consultoria",
        ],
        "forcar_planos_contas": [],
        # Várias linhas da DRE do RH não têm um Plano de Contas correspondente
        # no DIÁRIO/Tabela_Contas (o valor é lançado direto na linha). Nesses
        # casos, em vez de deixar a composição vazia, mostramos uma linha
        # "Lançado Manualmente" com o valor da própria linha da DRE.
        "permitir_lancamento_manual": True,
    },
    "📣 Relatório de Custos - MKT": {
        "linhas_dre": [
            # PREFIXO:6.24.2 pega a linha "6.24.2 - Marketing Regional -
            # Gestão CP" e TUDO abaixo dela na hierarquia (6.24.2.1,
            # 6.24.2.1.1, etc.), automaticamente -- é o ramo que o
            # coordenador de MKT de fato gerencia.
            #
            # Fica de fora, de propósito:
            # - "6.24 - Esforços de Marketing" (só consolida os dois ramos,
            #   incluiria custo que não é do coordenador de MKT)
            "PREFIXO:6.24.2",
        ],
        # "6.24.1 - Marketing Regional - Gestão GB" e tudo abaixo dela
        # (gestão da indústria, não do coordenador de MKT) -- não entram
        # nos KPIs/gráficos de gestão do departamento, mas aparecem no
        # relatório e num bloco à parte no painel, só pra conhecimento.
        "linhas_informativas": ["PREFIXO:6.24.1"],
        "forcar_planos_contas": [],
        # Linhas de grupo (ex.: "6.24 - Esforços de Marketing") que não
        # tiverem um Plano de Contas próprio já caem como "Lançado
        # Manualmente" automaticamente (mesma regra geral de todos os
        # modelos) -- não precisa de nada especial aqui.
        "permitir_lancamento_manual": False,
    },
}



# ============================================================================
# 4.5 SELEÇÃO DE PAINEL — Controladoria x Financeiro (tela entre o login e
# o painel de fato). O Financeiro ainda não existe -- mostra um aviso de
# "em construção" com a opção de ir para a Controladoria, que é o único
# painel ativo por enquanto.
# ============================================================================
if "painel_escolhido" not in st.session_state:
    st.session_state["painel_escolhido"] = None

if st.session_state["painel_escolhido"] is None:
    st.markdown(
        html_compacto(f"""
        <style>
            [data-testid="stSidebar"], header[data-testid="stHeader"] {{ display: none !important; }}
            .hub-wrap {{ max-width: 760px; margin: 64px auto 0 auto; text-align: center; }}
            .hub-title {{ font-size: 26px; font-weight: 800; color: {COLORS['text']}; margin-bottom: 6px; }}
            .hub-sub {{ font-size: 14px; color: {COLORS['text_muted']}; margin-bottom: 34px; }}
            .hub-card {{
                background: linear-gradient(160deg, {COLORS['surface']} 0%, {COLORS['surface_alt']} 100%);
                border: 1px solid {COLORS['border']}; border-radius: 16px; padding: 30px 22px 18px 22px;
                text-align: center; height: 100%;
            }}
            .hub-card .icone {{ font-size: 40px; margin-bottom: 8px; }}
            .hub-card h3 {{ color: {COLORS['text']}; font-size: 18px; margin: 4px 0 8px 0; }}
            .hub-card p {{ color: {COLORS['text_muted']}; font-size: 12.5px; margin-bottom: 4px; min-height: 36px; }}
        </style>
        <div class="hub-wrap">
            <div class="hub-title">👋 Bem-vindo(a) de volta</div>
            <div class="hub-sub">Escolha qual painel você quer acessar</div>
        </div>
        """),
        unsafe_allow_html=True,
    )

    col_hub_esp1, col_hub_a, col_hub_b, col_hub_esp2 = st.columns([0.6, 1, 1, 0.6])
    with col_hub_a:
        st.markdown(
            '<div class="hub-card"><div class="icone">📊</div><h3>Controladoria</h3>'
            "<p>DRE, orçado x realizado, histórico mensal, previsões e relatórios.</p></div>",
            unsafe_allow_html=True,
        )
        if st.button("Acessar Controladoria", use_container_width=True, type="primary", key="btn_hub_controladoria"):
            st.session_state["painel_escolhido"] = "controladoria"
            st.rerun()
    with col_hub_b:
        st.markdown(
            '<div class="hub-card"><div class="icone">💰</div><h3>Financeiro</h3>'
            "<p>Fluxo de caixa e demais indicadores financeiros.</p></div>",
            unsafe_allow_html=True,
        )
        if st.button("Acessar Financeiro", use_container_width=True, key="btn_hub_financeiro"):
            st.session_state["painel_escolhido"] = "financeiro"
            st.rerun()

    st.stop()

EMAILS_FINANCEIRO_PERMITIDOS = {
    "controladoria@grupobeea.com.br",
    "coordenador.financeiro@grupobeea.com.br",
    "diretoria.financeira@grupobeea.com.br",
}

# Colunas reais da aba "Fluxo de Caixa 2026" (confirmadas pelo usuário):
COL_FIN_VALOR = "Valor.1"
COL_FIN_MODALIDADE = "Modalidade"
COL_FIN_CANAL = "Canal.1"
COL_FIN_MOVIMENTO = "Movimento"
COL_FIN_DATA_LIQUIDACAO = "Data Liquidação"
COL_FIN_NUMERO = "Número"
COL_FIN_PLANO_CONTAS = "Plano de Contas"
COL_FIN_HISTORICO = "Histórico"
COL_FIN_GRUPO_DESPESA = "GRUPO DESPESA"
COL_FIN_VENCIMENTO = "Vencimento.1"

MESES_PT_FIN = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]
MESES_PT_ABREV_FIN = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def _rotulo_mes_pt(periodo):
    """Converte um Period/Timestamp mensal em "AGO/2026" (em português --
    o strftime do servidor devolveria "AUG", que confunde quem lê)."""
    return f"{MESES_PT_ABREV_FIN[periodo.month - 1]}/{periodo.year}"


def _rotulo_mes_pt_extenso(periodo):
    return f"{MESES_PT_FIN[periodo.month - 1]}/{periodo.year}"


def _classificar_movimento_fin(nome_movimento):
    """Classifica cada Movimento em uma de quatro categorias:
    - "aplicacao": aplicações financeiras -- ficam FORA de todos os totais,
      somas e tabelas principais do painel (a pedido da área), aparecendo
      só num bloco separado, para conhecimento.
    - "saldo": posição de dinheiro disponível (caixa, banco).
    - "entrada"/"saida": o fluxo de fato (a receber / a pagar).
    Saldo e fluxo nunca são somados juntos: um é posição, o outro é
    movimentação, e misturar os dois gera um total sem significado."""
    texto = str(nome_movimento).strip().lower()
    if "aplica" in texto:
        return "aplicacao"
    if any(p in texto for p in ["caixa", "banco"]):
        return "saldo"
    if "receber" in texto:
        return "entrada"
    if "pagar" in texto:
        return "saida"
    return "outro"


def _pivot_fluxo_fin(df, coluna_periodo, coluna_valor, coluna_movimento, ordem_periodos):
    """Monta a tabela Movimento x Período com a agregação CORRETA para cada
    tipo de linha:

    - Linhas de FLUXO (a receber / a pagar): SOMA do período. Faz sentido --
      são movimentações que se acumulam ao longo dos dias.
    - Linhas de SALDO (caixa, banco, aplicação): pega a POSIÇÃO DO ÚLTIMO
      DIA com movimento dentro daquele período, NÃO a soma. Saldo é uma
      foto do momento: somar o saldo de todos os dias do mês daria um
      número gigante e sem sentido (era o que acontecia antes -- o saldo
      aparecia inflado porque cada dia era somado ao anterior).
    """
    resultado = {}
    for movimento, grupo in df.groupby(coluna_movimento):
        tipo = _classificar_movimento_fin(movimento)
        linha = {}
        for periodo in ordem_periodos:
            do_periodo = grupo[grupo[coluna_periodo] == periodo]
            if do_periodo.empty:
                linha[periodo] = 0.0
            elif tipo in ("saldo", "aplicacao"):
                # posição do último dia com movimento dentro do período
                ultima_data = do_periodo["Data Efetiva"].max()
                linha[periodo] = do_periodo.loc[
                    do_periodo["Data Efetiva"] == ultima_data, coluna_valor
                ].sum()
            else:
                linha[periodo] = do_periodo[coluna_valor].sum()
        resultado[movimento] = linha

    pivot = pd.DataFrame.from_dict(resultado, orient="index")
    pivot = pivot.reindex(columns=ordem_periodos, fill_value=0.0)
    return pivot


def _saldo_posicao_atual_fin(df, coluna_valor):
    """Saldo disponível = posição do ÚLTIMO DIA com movimento no recorte
    filtrado (não a soma de todos os dias). Se houver mais de um canal, soma
    os canais NAQUELE dia, que é o saldo total da empresa naquela data."""
    linhas_saldo = df[df["Tipo Movimento"] == "saldo"]
    if linhas_saldo.empty:
        return 0.0, None
    ultima_data = linhas_saldo["Data Efetiva"].max()
    valor = linhas_saldo.loc[linhas_saldo["Data Efetiva"] == ultima_data, coluna_valor].sum()
    return valor, ultima_data


@st.cache_resource(show_spinner="Preparando os dados do fluxo de caixa...")
def preparar_fluxo_caixa(base_data):
    """Faz TODO o trabalho pesado uma vez só e guarda em cache: leitura do
    CSV, conversão de ~650 mil valores e datas do formato brasileiro, e a
    classificação dos movimentos.

    Antes isso rodava a cada clique (trocar um filtro, mudar de aba), o que
    deixava o painel lento sem necessidade -- os dados são os mesmos. O
    cache é por `base_data` porque só essa escolha muda o resultado; os
    demais filtros são aplicados depois, sobre o DataFrame já pronto.

    Retorna (df, erro, total_lido, linhas_sem_data)."""
    df_fluxo, erro = obter_dados_fluxo_caixa()
    if erro or df_fluxo is None or df_fluxo.empty:
        return None, erro or "Sem dados", 0, 0

    colunas_esperadas = [
        COL_FIN_VALOR, COL_FIN_MODALIDADE, COL_FIN_CANAL, COL_FIN_MOVIMENTO,
        COL_FIN_DATA_LIQUIDACAO, COL_FIN_VENCIMENTO,
    ]
    faltando = [c for c in colunas_esperadas if c not in df_fluxo.columns]
    if faltando:
        return None, "COLUNAS_FALTANDO:" + ", ".join(faltando), 0, 0

    df = df_fluxo.copy()

    # O CSV publicado entrega tudo como TEXTO no formato brasileiro (valor
    # com vírgula decimal e ponto de milhar, data dd/mm/aaaa).
    valores_texto = (
        df[COL_FIN_VALOR].astype(str).str.strip()
        .str.replace(r"[R$\s]", "", regex=True)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df[COL_FIN_VALOR] = pd.to_numeric(valores_texto, errors="coerce").fillna(0)

    df[COL_FIN_DATA_LIQUIDACAO] = pd.to_datetime(
        df[COL_FIN_DATA_LIQUIDACAO], errors="coerce", dayfirst=True
    )
    df[COL_FIN_VENCIMENTO] = pd.to_datetime(
        df[COL_FIN_VENCIMENTO], errors="coerce", dayfirst=True
    )

    if str(base_data).startswith("Vencimento"):
        df["Data Efetiva"] = df[COL_FIN_VENCIMENTO].fillna(df[COL_FIN_DATA_LIQUIDACAO])
    else:
        df["Data Efetiva"] = df[COL_FIN_DATA_LIQUIDACAO].fillna(df[COL_FIN_VENCIMENTO])

    total_lido = len(df)
    df = df.dropna(subset=["Data Efetiva"])
    sem_data = total_lido - len(df)

    df["Tipo Movimento"] = df[COL_FIN_MOVIMENTO].map(_classificar_movimento_fin)
    df["Liquidado"] = df[COL_FIN_DATA_LIQUIDACAO].notna()
    # Colunas de texto viram "category": ocupam bem menos memória e deixam
    # os agrupamentos por canal/modalidade bem mais rápidos.
    for coluna_texto in (COL_FIN_CANAL, COL_FIN_MODALIDADE, COL_FIN_MOVIMENTO):
        df[coluna_texto] = df[coluna_texto].astype("category")

    return df, None, total_lido, sem_data


if st.session_state["painel_escolhido"] == "financeiro":
    st.markdown(
        """<style>[data-testid="stSidebar"], header[data-testid="stHeader"] { display: none !important; }</style>""",
        unsafe_allow_html=True,
    )

    email_atual_financeiro = str(usuario_atual.get("email", "")).strip().lower()
    if email_atual_financeiro not in EMAILS_FINANCEIRO_PERMITIDOS:
        st.markdown(
            html_compacto(f"""
            <style>
                .fin-bloqueio {{
                    max-width: 520px; margin: 90px auto 0 auto; text-align: center;
                    background: linear-gradient(160deg, {COLORS['surface']} 0%, {COLORS['surface_alt']} 100%);
                    border: 1px solid {COLORS['border']}; border-radius: 18px;
                    padding: 40px 34px 34px 34px;
                    box-shadow: 0 18px 50px rgba(0,0,0,0.45);
                }}
                .fin-bloqueio .selo {{
                    width: 66px; height: 66px; margin: 0 auto 18px auto; border-radius: 50%;
                    background: rgba(247,110,110,0.10); border: 1px solid rgba(247,110,110,0.35);
                    display: flex; align-items: center; justify-content: center; font-size: 28px;
                }}
                .fin-bloqueio h2 {{
                    color: {COLORS['text']}; font-size: 20px; font-weight: 800;
                    margin: 0 0 10px 0; letter-spacing: 0.2px;
                }}
                .fin-bloqueio p {{
                    color: {COLORS['text_muted']}; font-size: 13.5px; line-height: 1.65; margin: 0 0 6px 0;
                }}
                .fin-bloqueio .rodape {{
                    margin-top: 22px; padding-top: 16px; border-top: 1px solid {COLORS['border']};
                    color: {COLORS['text_muted']}; font-size: 11.5px;
                }}
                .fin-bloqueio .rodape b {{ color: {COLORS['text']}; }}
            </style>
            <div class="fin-bloqueio">
                <div class="selo">🔒</div>
                <h2>Painel Financeiro restrito</h2>
                <p>
                    Este painel contém informações de fluxo de caixa e está liberado apenas
                    para a Controladoria e a área Financeira.
                </p>
                <p>
                    Se você precisa desse acesso para o seu trabalho, fale com a Controladoria
                    que a liberação pode ser avaliada.
                </p>
                <div class="rodape">
                    Você está conectado como <b>{usuario_atual.get('email', '—')}</b>
                </div>
            </div>
            """),
            unsafe_allow_html=True,
        )
        col_voltar_esp1, col_voltar, col_voltar_esp2 = st.columns([1, 1.2, 1])
        with col_voltar:
            st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
            if st.button("← Ir para o Painel de Controladoria", use_container_width=True, type="primary"):
                st.session_state["painel_escolhido"] = "controladoria"
                st.rerun()
        st.stop()

    # ---- Cabeçalho compacto: marca e ações na MESMA linha ----
    # O Streamlit deixa uma folga generosa entre blocos; num cabeçalho isso
    # vira espaço morto. O CSS abaixo aperta esses espaçamentos só aqui.
    st.markdown(
        f"""
        <style>
            /* Só o estilo do cabeçalho -- nada que mexa no espaçamento geral
               dos blocos. A tentativa anterior de apertar o "gap" global fez
               títulos e tabelas se sobreporem pelo painel inteiro. */
            .fin-topo {{ display: flex; align-items: center; gap: 11px; }}
            .fin-topo img.logo {{ width: 36px; height: 36px; border-radius: 50%; }}
            .fin-topo h1 {{
                font-size: 18px; font-weight: 800; color: {COLORS['text']};
                margin: 0; letter-spacing: 0.2px; line-height: 1.2;
            }}
            .fin-topo .sub {{ color: {COLORS['text_muted']}; font-size: 11px; margin-top: 1px; }}
            .fin-barra {{
                display: flex; align-items: center; gap: 7px; flex-wrap: wrap;
                color: {COLORS['text_muted']}; font-size: 10.5px; margin: 3px 0 0 0;
            }}
            .fin-barra .chip {{
                background: {COLORS['primary_soft']}; color: {COLORS['primary']};
                border-radius: 5px; padding: 3px 9px; font-size: 10.5px;
                font-weight: 700; letter-spacing: 0.5px;
            }}
            .fin-barra b {{ color: {COLORS['text']}; font-weight: 600; }}
            .fin-barra .sep {{ opacity: 0.4; }}
            /* Radio da base de data: fonte menor, sem mexer em margens */
            div[data-testid="stRadio"] label p {{ font-size: 11.5px !important; }}
            div[data-testid="stRadio"] > label p {{
                font-size: 10px !important; text-transform: uppercase;
                letter-spacing: 0.4px; opacity: 0.65;
            }}
            div[data-testid="stRadio"] > div {{ gap: 0.6rem !important; }}
            /* Faixa divisória fina abaixo do cabeçalho */
            .fin-divisor {{
                border-bottom: 1px solid {COLORS['border']};
                margin: 6px 0 12px 0;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Tudo numa linha só: marca à esquerda, base de data no meio e as ações
    # à direita. Antes isso ocupava duas linhas com bastante espaço morto.
    col_marca_fin, col_base_fin, col_atualizar_fin, col_trocar_fin = st.columns(
        [3.2, 2.4, 1.1, 1.1], vertical_alignment="center"
    )
    with col_marca_fin:
        st.markdown(
            html_compacto(f"""
            <div class="fin-topo">
                <img class="logo" src="data:image/png;base64,{LOGO_BEEA_B64}" alt="Grupo Beea"/>
                <div>
                    <h1>Painel Financeiro</h1>
                    <div class="fin-barra">
                        <span class="chip">FLUXO DE CAIXA 2026</span>
                        <span class="sep">·</span>
                        <span>{datetime.now(FUSO_BR).strftime('%d/%m/%Y às %H:%M')}</span>
                    </div>
                </div>
            </div>
            """),
            unsafe_allow_html=True,
        )
    with col_base_fin:
        # A base de data define em que mês o lançamento cai. O padrão é
        # VENCIMENTO porque é o eixo que a tabela dinâmica da planilha usa --
        # é o que faz o painel bater número a número com ela.
        base_data_fin = st.radio(
            "Base de data:",
            ["Vencimento (igual à planilha)", "Liquidação (caixa efetivo)"],
            horizontal=True, key="fin_base_data",
            help=(
                "Vencimento: cada lançamento cai no mês em que vence -- mesma regra da tabela dinâmica "
                "da planilha. Liquidação: cai no mês em que o dinheiro de fato entrou ou saiu (para o que "
                "ainda não liquidou, usa o vencimento como previsão)."
            ),
        )
    with col_atualizar_fin:
        if st.button("🔄 Atualizar", use_container_width=True, key="fin_btn_atualizar"):
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()
    with col_trocar_fin:
        if st.button("🔀 Trocar Painel", use_container_width=True, key="fin_btn_trocar"):
            st.session_state["painel_escolhido"] = None
            st.rerun()

    st.markdown('<div class="fin-divisor"></div>', unsafe_allow_html=True)


    # Todo o trabalho pesado (leitura + conversão de ~650 mil linhas) fica
    # em cache: só refaz se a base de data mudar ou se você clicar em
    # "Atualizar Dados".
    df_fin, erro_fluxo, total_linhas_lidas, linhas_sem_data = preparar_fluxo_caixa(base_data_fin)

    if erro_fluxo and str(erro_fluxo).startswith("COLUNAS_FALTANDO:"):
        st.error(
            "Algumas colunas esperadas não foram encontradas na planilha: "
            + str(erro_fluxo).split("COLUNAS_FALTANDO:", 1)[1]
            + ". Confira se o nome está escrito exatamente igual na aba."
        )
        st.stop()

    if erro_fluxo:
        st.error(
            "Não consegui carregar os dados do Fluxo de Caixa. O painel lê o CSV publicado da aba "
            "\"Fluxo de Caixa 2026\" -- confirme na planilha, em Arquivo > Compartilhar > Publicar na web, "
            "se a publicação continua ativa e se a opção \"Restringir acesso\" está DESMARCADA."
        )
        with st.expander("Detalhe técnico do erro"):
            st.code(str(erro_fluxo))
        st.stop()

    if df_fin is None or df_fin.empty:
        st.warning("O CSV publicado foi lido, mas não trouxe nenhuma linha com data válida.")
        st.stop()

    if eh_admin:
        with st.expander("🔧 Diagnóstico da planilha (visível só para administrador)"):
            st.write(
                f"**Linhas lidas do CSV:** {total_linhas_lidas} · "
                f"**Com data válida:** {len(df_fin)} · **Descartadas por falta de data:** {linhas_sem_data}"
            )
            data_min_diag = df_fin["Data Efetiva"].min()
            data_max_diag = df_fin["Data Efetiva"].max()
            st.write(f"**Período encontrado nos dados:** {data_min_diag:%d/%m/%Y} até {data_max_diag:%d/%m/%Y}")
            st.write("**Movimento → como o painel classificou:**")
            st.dataframe(
                df_fin.groupby([COL_FIN_MOVIMENTO, "Tipo Movimento"], observed=True)[COL_FIN_VALOR]
                .agg(["count", "sum"]).reset_index()
                .rename(columns={"count": "Qtd. lançamentos", "sum": "Soma (R$)"}),
                use_container_width=True, hide_index=True,
            )
            st.write("**Canais:** " + ", ".join(sorted(df_fin[COL_FIN_CANAL].dropna().astype(str).unique())))
            st.write("**Modalidades:** " + ", ".join(sorted(df_fin[COL_FIN_MODALIDADE].dropna().astype(str).unique())))
            st.dataframe(df_fin.head(15), use_container_width=True, hide_index=True)

    tab_fin_mensal, tab_fin_diario, tab_fin_tesouraria, tab_fin_analises = st.tabs(
        ["📅 Fluxo Mensal", "🗓️ Fluxo Diário", "🏦 Tesouraria", "📊 Análises"]
    )

    # ---------------- MENSAL ----------------
    with tab_fin_mensal:
        # ================= FILTROS =================
        data_min_fin = df_fin["Data Efetiva"].min().date()
        data_max_fin = df_fin["Data Efetiva"].max().date()

        # Padrão pedido: começa no 1º dia do MÊS ANTERIOR ao atual e vai até o
        # último dia do ANO atual (limitado ao que existe de dado na planilha).
        hoje_fin = datetime.now(FUSO_BR).date()
        primeiro_dia_mes_anterior = (hoje_fin.replace(day=1) - pd.Timedelta(days=1)).replace(day=1)
        ultimo_dia_ano_atual = hoje_fin.replace(month=12, day=31)
        padrao_ini_fin = max(min(primeiro_dia_mes_anterior, data_max_fin), data_min_fin)
        padrao_fim_fin = min(max(ultimo_dia_ano_atual, data_min_fin), data_max_fin)

        # O calendário do st.date_input segue o idioma do servidor (por isso vinha
        # "July", "Su/Mo/Tu") e não há como forçar português nele por
        # configuração. Em vez de brigar com o componente, o período é escolhido
        # por seletores de MÊS -- que ficam 100% em português e, para fluxo de
        # caixa, são até mais práticos. O ajuste fino por dia fica disponível
        # logo abaixo, para quem precisar de um recorte exato.
        meses_disponiveis_fin = sorted(df_fin["Data Efetiva"].dt.to_period("M").unique())
        rotulos_meses_fin = [_rotulo_mes_pt_extenso(p) for p in meses_disponiveis_fin]

        periodo_ini_padrao = pd.Period(padrao_ini_fin.strftime("%Y-%m"), freq="M")
        periodo_fim_padrao = pd.Period(padrao_fim_fin.strftime("%Y-%m"), freq="M")
        idx_ini_padrao = meses_disponiveis_fin.index(periodo_ini_padrao) if periodo_ini_padrao in meses_disponiveis_fin else 0
        idx_fim_padrao = meses_disponiveis_fin.index(periodo_fim_padrao) if periodo_fim_padrao in meses_disponiveis_fin else len(meses_disponiveis_fin) - 1

        col_f1, col_f2, col_f3, col_f4 = st.columns([1.2, 1.2, 1, 1])
        with col_f1:
            mes_ini_sel_fin = st.selectbox(
                "Do mês:", rotulos_meses_fin, index=idx_ini_padrao, key="fin_mes_ini",
                help="Por padrão começa no mês anterior ao atual.",
            )
        with col_f2:
            mes_fim_sel_fin = st.selectbox(
                "Até o mês:", rotulos_meses_fin, index=idx_fim_padrao, key="fin_mes_fim",
                help="Por padrão vai até o fim do ano atual.",
            )
        with col_f3:
            opcoes_canal_fin = ["Todos"] + sorted(df_fin[COL_FIN_CANAL].dropna().astype(str).unique().tolist())
            canal_sel_fin = st.selectbox("Canal:", opcoes_canal_fin, key="fin_canal_sel")
        with col_f4:
            opcoes_modal_fin = ["Todas"] + sorted(df_fin[COL_FIN_MODALIDADE].dropna().astype(str).unique().tolist())
            modal_sel_fin = st.selectbox("Modalidade:", opcoes_modal_fin, key="fin_modal_sel")

        periodo_ini_fin = meses_disponiveis_fin[rotulos_meses_fin.index(mes_ini_sel_fin)]
        periodo_fim_fin = meses_disponiveis_fin[rotulos_meses_fin.index(mes_fim_sel_fin)]
        if periodo_ini_fin > periodo_fim_fin:
            st.warning("O mês inicial está depois do mês final — inverti os dois para continuar.")
            periodo_ini_fin, periodo_fim_fin = periodo_fim_fin, periodo_ini_fin

        data_ini_fin = periodo_ini_fin.start_time.date()
        data_fim_fin = periodo_fim_fin.end_time.date()

        with st.expander("📆 Ajuste fino por dia (opcional)"):
            st.caption(
                "Por padrão o período pega do primeiro ao último dia dos meses escolhidos acima. "
                "Use aqui se precisar de um recorte exato dentro desses meses."
            )
            col_dia_a, col_dia_b = st.columns(2)
            with col_dia_a:
                dia_ini_fin = st.number_input(
                    f"Dia inicial (em {mes_ini_sel_fin})", min_value=1, max_value=31, value=1, step=1, key="fin_dia_ini",
                )
            with col_dia_b:
                ultimo_dia_mes_fim = periodo_fim_fin.end_time.day
                dia_fim_fin = st.number_input(
                    f"Dia final (em {mes_fim_sel_fin})", min_value=1, max_value=31,
                    value=ultimo_dia_mes_fim, step=1, key="fin_dia_fim",
                )
            # Se a pessoa digitar um dia que não existe naquele mês (ex.: 31 em
            # fevereiro), ajusta pro último dia válido em vez de dar erro.
            dia_ini_valido = min(int(dia_ini_fin), periodo_ini_fin.end_time.day)
            dia_fim_valido = min(int(dia_fim_fin), periodo_fim_fin.end_time.day)
            data_ini_fin = periodo_ini_fin.start_time.date().replace(day=dia_ini_valido)
            data_fim_fin = periodo_fim_fin.start_time.date().replace(day=dia_fim_valido)

        df_fin_periodo = df_fin[
            (df_fin["Data Efetiva"].dt.date >= data_ini_fin)
            & (df_fin["Data Efetiva"].dt.date <= data_fim_fin)
        ].copy()
        if canal_sel_fin != "Todos":
            df_fin_periodo = df_fin_periodo[df_fin_periodo[COL_FIN_CANAL].astype(str) == canal_sel_fin]
        if modal_sel_fin != "Todas":
            df_fin_periodo = df_fin_periodo[df_fin_periodo[COL_FIN_MODALIDADE].astype(str) == modal_sel_fin]

        # As APLICAÇÕES ficam fora de tudo que é total/soma/tabela principal --
        # são exibidas só num bloco à parte, no fim da aba mensal.
        df_aplicacoes_fin = df_fin_periodo[df_fin_periodo["Tipo Movimento"] == "aplicacao"].copy()
        df_fin_view = df_fin_periodo[df_fin_periodo["Tipo Movimento"] != "aplicacao"].copy()

        if df_fin_view.empty:
            st.warning("Nenhum lançamento encontrado para os filtros selecionados.")
        else:

            # ================= KPIs =================
            saldo_posicao_fin, data_saldo_fin = _saldo_posicao_atual_fin(df_fin_view, COL_FIN_VALOR)
            entradas_fin = df_fin_view.loc[df_fin_view["Tipo Movimento"] == "entrada", COL_FIN_VALOR].sum()
            saidas_fin = df_fin_view.loc[df_fin_view["Tipo Movimento"] == "saida", COL_FIN_VALOR].sum()
            fluxo_liquido_fin = entradas_fin + saidas_fin  # saídas já vêm negativas

            # "Ainda a receber" = o que está classificado como PROJETADO na coluna
            # Movimento. Antes isso era deduzido da coluna "Data Liquidação" estar
            # vazia, mas boa parte dos lançamentos marcados como "Realizado" vem sem
            # essa data preenchida na planilha -- e assim eles entravam aqui por
            # engano, inflando o número. O Movimento é a classificação confiável.
            mascara_projetado_fin = df_fin_view[COL_FIN_MOVIMENTO].astype(str).str.contains(
                "projetad", case=False, na=False
            )
            a_receber_projetado_fin = df_fin_view.loc[
                (df_fin_view["Tipo Movimento"] == "entrada") & mascara_projetado_fin, COL_FIN_VALOR
            ].sum()
            a_receber_realizado_fin = entradas_fin - a_receber_projetado_fin

            st.markdown(
                f'<div class="section-title">📊 Resumo do Período '
                f'<span style="font-weight:400;font-size:12px;color:{COLORS["text_muted"]};">'
                f'({data_ini_fin:%d/%m/%Y} a {data_fim_fin:%d/%m/%Y})</span></div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                render_kpi_row([
                    dict(label="SALDO EM CAIXA / BANCO", value=formata_brl(saldo_posicao_fin),
                         value_color=cor_variacao(saldo_posicao_fin),
                         subtext=(f"Posição em {data_saldo_fin:%d/%m/%Y}" if data_saldo_fin is not None else "Sem posição no período"),
                         icon="🏦"),
                    dict(label="ENTRADAS (A RECEBER)", value=formata_brl(entradas_fin),
                         value_color=COLORS["positive"], subtext="Realizado + projetado no período", icon="📥"),
                    dict(label="SAÍDAS (A PAGAR)", value=formata_brl(saidas_fin),
                         value_color=COLORS["negative"], subtext="Contas a pagar no período", icon="📤"),
                    dict(label="FLUXO LÍQUIDO DO PERÍODO", value=formata_brl(fluxo_liquido_fin),
                         value_color=cor_variacao(fluxo_liquido_fin), subtext="Entradas − saídas", icon="⚖️"),
                    dict(label="A RECEBER PROJETADO", value=formata_brl(a_receber_projetado_fin),
                         value_color=COLORS["warning"],
                         subtext=f"Realizado no período: {formata_brl(a_receber_realizado_fin)}", icon="⏳"),
                ]),
                unsafe_allow_html=True,
            )
            st.caption(
                "ℹ️ **Aplicações não entram em nenhum número desta tela** — ficam num bloco separado no fim da aba "
                "Fluxo Mensal, só para conhecimento. E saldo (caixa/banco) é **posição**, não movimento: por isso "
                "aparece separado das entradas e saídas e não entra no fluxo líquido."
            )
            st.markdown("<br>", unsafe_allow_html=True)

            df_m = df_fin_view.copy()
            df_m["PeriodoMes"] = df_m["Data Efetiva"].dt.to_period("M")
            meses_ordenados_m = sorted(df_m["PeriodoMes"].unique())
            rotulos_meses_m = {p: _rotulo_mes_pt(p) for p in meses_ordenados_m}

            pivot_m = _pivot_fluxo_fin(df_m, "PeriodoMes", COL_FIN_VALOR, COL_FIN_MOVIMENTO, meses_ordenados_m)

            # A coluna final tem significado diferente conforme o tipo de linha:
            # para fluxo é a SOMA do período; para saldo é a ÚLTIMA posição
            # (somar saldos de meses diferentes não faria sentido).
            totais_finais_m = {}
            for movimento in pivot_m.index:
                if _classificar_movimento_fin(movimento) in ("saldo", "aplicacao"):
                    serie = pivot_m.loc[movimento]
                    nao_zerados = serie[serie != 0]
                    totais_finais_m[movimento] = nao_zerados.iloc[-1] if not nao_zerados.empty else 0.0
                else:
                    totais_finais_m[movimento] = pivot_m.loc[movimento].sum()

            pivot_m.columns = [rotulos_meses_m[p] for p in pivot_m.columns]
            pivot_m["TOTAL / ÚLT. POSIÇÃO"] = pd.Series(totais_finais_m)
            pivot_m.index.name = "Movimento"

            # TOTAL GERAL: soma de todas as linhas em cada mês -- mesmo cálculo
            # do "Total Geral" da tabela dinâmica da planilha.
            linha_total_geral_m = pivot_m.sum(axis=0)
            pivot_m_exibicao = pd.concat([pivot_m, pd.DataFrame([linha_total_geral_m], index=["TOTAL GERAL"])])

            st.markdown('<div class="section-title">📋 Movimentos por Mês</div>', unsafe_allow_html=True)
            st.dataframe(
                pivot_m_exibicao.style.format(formata_brl).map(cor_valor),
                use_container_width=True,
            )
            st.caption(
                "As linhas de **caixa e banco** mostram o **saldo do último dia** de cada mês (é uma posição, "
                "uma foto do momento). As linhas de **contas a receber e a pagar** mostram a **soma** do mês "
                "(são movimentações que se acumulam). Por isso a última coluna traz o total acumulado para o "
                "fluxo e a posição mais recente para o saldo."
            )

            # ---- Reserva de caixa: o que sobra DEPOIS de pagar tudo ----
            # A regra da área: pagando todas as contas do mês, ainda tem que
            # sobrar ~30% do dinheiro disponível. Por isso o disponível inclui
            # os recebíveis (caixa + banco + a receber projetado + realizado),
            # e o indicador é a SOBRA sobre esse disponível.
            colunas_meses_m = [rotulos_meses_m[p] for p in meses_ordenados_m]
            serie_total_geral = linha_total_geral_m[colunas_meses_m]

            movimentos_a_pagar = [m for m in pivot_m.index if _classificar_movimento_fin(m) == "saida"]
            serie_a_pagar = (
                pivot_m.loc[movimentos_a_pagar, colunas_meses_m].sum(axis=0)
                if movimentos_a_pagar else pd.Series(0.0, index=colunas_meses_m)
            )
            # Mantém o "a pagar" negativo: é saída de dinheiro, e assim a coloração
            # padrão das tabelas já o marca em vermelho automaticamente.
            serie_obrigacoes = -serie_a_pagar.abs()
            # Disponível = tudo que há antes de descontar o que se deve
            serie_disponivel_total = serie_total_geral - serie_a_pagar
            serie_sobra = serie_total_geral  # já é disponível − a pagar
            serie_pct_sobra = pd.Series(
                [
                    (sobra / disp * 100) if disp else 0.0
                    for sobra, disp in zip(serie_sobra.values, serie_disponivel_total.values)
                ],
                index=colunas_meses_m,
            )

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(
                '<div class="section-title">🛡️ Reserva de Caixa — sobra depois de pagar tudo (meta: 30%)</div>',
                unsafe_allow_html=True,
            )

            LINHA_PCT_SOBRA = "% de sobra"
            df_reserva_m = pd.DataFrame(
                [serie_disponivel_total, serie_obrigacoes, serie_sobra, serie_pct_sobra],
                index=["Disponível", "A pagar no mês", "Sobra depois de pagar tudo", LINHA_PCT_SOBRA],
            )

            def _cor_pct_meta_fin(linha):
                """Na linha de % de sobra a cor não segue o sinal (todos são
                positivos), e sim a META: abaixo de 30% fica vermelho."""
                if linha.name == LINHA_PCT_SOBRA:
                    return [
                        f"color: {COLORS['positive'] if v >= 30 else COLORS['negative']}; font-weight: 600;"
                        for v in linha
                    ]
                return [cor_valor(v) for v in linha]

            st.dataframe(
                df_reserva_m.style.format(
                    {c: formata_brl for c in colunas_meses_m}
                ).format(
                    {c: (lambda v: f"{v:.1f}%".replace(".", ",")) for c in colunas_meses_m},
                    subset=pd.IndexSlice[[LINHA_PCT_SOBRA], :],
                ).apply(_cor_pct_meta_fin, axis=1),
                use_container_width=True,
            )

            st.caption(
                "**% de sobra** = (disponível − a pagar) ÷ disponível. Ou seja: pagando todas as contas do mês, "
                "quanto sobra em caixa em proporção ao que havia disponível. É esse número que deve ficar em "
                "30% ou mais."
            )

            # ---- Gráfico executivo: entradas x saídas, resultado e reserva ----
            df_fluxo_only = df_m[df_m["Tipo Movimento"].isin(["entrada", "saida"])]
            if not df_fluxo_only.empty:
                entradas_mes = df_fluxo_only[df_fluxo_only["Tipo Movimento"] == "entrada"].groupby("PeriodoMes")[COL_FIN_VALOR].sum()
                saidas_mes = df_fluxo_only[df_fluxo_only["Tipo Movimento"] == "saida"].groupby("PeriodoMes")[COL_FIN_VALOR].sum()
                entradas_mes = entradas_mes.reindex(meses_ordenados_m, fill_value=0)
                saidas_mes = saidas_mes.reindex(meses_ordenados_m, fill_value=0)
                rotulos_x_m = [rotulos_meses_m[p] for p in meses_ordenados_m]

                valores_entrada = [float(v) for v in entradas_mes.values]
                valores_saida = [abs(float(v)) for v in saidas_mes.values]
                resultado_mes = [e - s for e, s in zip(valores_entrada, valores_saida)]
                pct_sobra_grafico = [float(serie_pct_sobra[c]) for c in rotulos_x_m]

                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(
                    '<div class="section-title">📊 Entradas x Saídas, Resultado e Reserva por Mês</div>',
                    unsafe_allow_html=True,
                )

                # Barras com preenchimento bem translúcido e contorno forte: a
                # cor ainda identifica entrada/saída, mas dá pra enxergar as
                # linhas que passam por cima delas (com barras sólidas, as
                # linhas de resultado e de % sumiam atrás).
                fig_es = go.Figure()
                fig_es.add_trace(go.Bar(
                    name="Entradas", x=rotulos_x_m, y=valores_entrada,
                    marker=dict(color="rgba(62,207,142,0.28)",
                                line=dict(color=COLORS["positive"], width=1.8)),
                    text=[formata_m(v) for v in valores_entrada],
                    textposition="outside", textfont=dict(size=9, color=COLORS["text_muted"]),
                    hovertemplate="Entradas: R$ %{y:,.2f}<extra></extra>",
                ))
                fig_es.add_trace(go.Bar(
                    name="Saídas", x=rotulos_x_m, y=valores_saida,
                    marker=dict(color="rgba(247,110,110,0.28)",
                                line=dict(color=COLORS["negative"], width=1.8)),
                    text=[formata_m(v) for v in valores_saida],
                    textposition="outside", textfont=dict(size=9, color=COLORS["text_muted"]),
                    hovertemplate="Saídas: R$ %{y:,.2f}<extra></extra>",
                ))
                # Resultado do mês (entradas − saídas), na mesma escala das barras
                fig_es.add_trace(go.Scatter(
                    name="Resultado do mês", x=rotulos_x_m, y=resultado_mes,
                    mode="lines+markers", line=dict(color=COLORS["primary"], width=3),
                    marker=dict(size=9, line=dict(color=COLORS["bg"], width=2)),
                    hovertemplate="Resultado: R$ %{y:,.2f}<extra></extra>",
                ))
                # % de sobra x meta de 30%, no eixo da direita
                cores_pontos_sobra = [
                    COLORS["positive"] if v >= 30 else COLORS["negative"] for v in pct_sobra_grafico
                ]
                fig_es.add_trace(go.Scatter(
                    name="% de sobra", x=rotulos_x_m, y=pct_sobra_grafico, yaxis="y2",
                    mode="lines+markers+text", line=dict(color=COLORS["warning"], width=2.5),
                    marker=dict(size=10, color=cores_pontos_sobra,
                                line=dict(color=COLORS["bg"], width=2)),
                    text=[f"{v:.0f}%" for v in pct_sobra_grafico],
                    textposition="top center", textfont=dict(size=10, color=COLORS["text"]),
                    hovertemplate="Sobra: %{y:.1f}%<extra></extra>",
                ))
                fig_es.add_trace(go.Scatter(
                    name="Meta 30%", x=rotulos_x_m, y=[30] * len(rotulos_x_m), yaxis="y2",
                    mode="lines", line=dict(color=COLORS["muted_line"], width=1.5, dash="dash"),
                    hoverinfo="skip",
                ))

                teto_barras = max(valores_entrada + valores_saida) if (valores_entrada + valores_saida) else 1
                teto_pct = max(pct_sobra_grafico) if pct_sobra_grafico else 60
                estilo_grafico(
                    fig_es, height=470, barmode="group", bargap=0.3, bargroupgap=0.08,
                    xaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickangle=-25,
                               automargin=True, tickfont=dict(size=10)),
                    yaxis=dict(
                        title=dict(text="R$", font=dict(size=10, color=COLORS["text_muted"])),
                        gridcolor=COLORS["border"], fixedrange=True, tickformat=",.0f",
                        zerolinecolor=COLORS["border"],
                        range=[min(0, min(resultado_mes) * 1.15), teto_barras * 1.20],
                    ),
                    yaxis2=dict(
                        title=dict(text="% de sobra", font=dict(size=10, color=COLORS["warning"])),
                        overlaying="y", side="right", showgrid=False, fixedrange=True,
                        ticksuffix="%", tickfont=dict(size=9, color=COLORS["warning"]),
                        range=[0, max(60, teto_pct * 1.4)],
                    ),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5,
                                font=dict(size=10)),
                    margin=dict(l=70, r=70, t=30, b=90),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_es, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                st.caption(
                    "Barras = entradas e saídas do mês (eixo da esquerda, em R$). Linha azul = resultado do mês "
                    "(entradas − saídas). Linha laranja = % de sobra, no eixo da direita, com os pontos em verde "
                    "quando atingem a meta de 30% (tracejado cinza) e em vermelho quando ficam abaixo."
                )

            # ---- Aplicações: fora de todos os totais acima, só pra conhecimento ----
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="section-title">💼 Aplicações (fora dos números acima)</div>', unsafe_allow_html=True)
            if df_aplicacoes_fin.empty:
                st.caption("Nenhum lançamento de aplicação no período/filtros selecionados.")
            else:
                st.caption(
                    "Estes valores **não entram** em nenhum KPI, tabela, total ou gráfico desta tela — "
                    "estão aqui apenas para acompanhamento da posição aplicada."
                )
                df_ap = df_aplicacoes_fin.copy()
                df_ap["PeriodoMes"] = df_ap["Data Efetiva"].dt.to_period("M")
                meses_ap = sorted(df_ap["PeriodoMes"].unique())
                pivot_ap = _pivot_fluxo_fin(df_ap, "PeriodoMes", COL_FIN_VALOR, COL_FIN_MOVIMENTO, meses_ap)
                ultimas_posicoes_ap = {}
                for movimento in pivot_ap.index:
                    serie_ap = pivot_ap.loc[movimento]
                    nao_zerados_ap = serie_ap[serie_ap != 0]
                    ultimas_posicoes_ap[movimento] = nao_zerados_ap.iloc[-1] if not nao_zerados_ap.empty else 0.0
                pivot_ap.columns = [_rotulo_mes_pt(p) for p in pivot_ap.columns]
                pivot_ap["ÚLT. POSIÇÃO"] = pd.Series(ultimas_posicoes_ap)
                pivot_ap.index.name = "Movimento"
                st.dataframe(pivot_ap.style.format(formata_brl).map(cor_valor), use_container_width=True)
                st.caption("Aplicação é posição, não movimento — cada mês mostra o saldo aplicado no último dia daquele mês.")

    # ---------------- DIÁRIO ----------------
    with tab_fin_diario:
        # Esta aba tem os PRÓPRIOS filtros e cards -- ela olha o dia a dia de
        # um mês, então faz pouco sentido herdar o recorte de vários meses da
        # aba mensal.

        meses_disp_d = sorted(df_fin["Data Efetiva"].dt.to_period("M").unique())
        if not meses_disp_d:
            st.info("Sem dados para montar o fluxo diário.")
        else:
            rotulos_d = [_rotulo_mes_pt_extenso(p) for p in meses_disp_d]
            # Abre no mês do DIA ANTERIOR, não no de hoje: o movimento do dia
            # corrente costuma estar incompleto, e no dia 1º do mês isso
            # levaria a um mês praticamente vazio. Se esse mês não existir
            # nos dados, cai no mais recente disponível.
            data_ontem_d = datetime.now(FUSO_BR).date() - pd.Timedelta(days=1)
            periodo_mes_ontem = pd.Period(data_ontem_d.strftime("%Y-%m"), freq="M")
            idx_mes_padrao_d = (
                meses_disp_d.index(periodo_mes_ontem) if periodo_mes_ontem in meses_disp_d
                else len(rotulos_d) - 1
            )

            # Mesmo motivo do bloco de dias abaixo: com `key`, o `index` só
            # vale se a sessão ainda não tiver um valor guardado. Grava o
            # padrão na primeira abertura e deixa a escolha da pessoa mandar
            # dali em diante.
            if "fin_mes_sel" not in st.session_state:
                st.session_state["fin_mes_sel"] = rotulos_d[idx_mes_padrao_d]

            col_d1, col_d2, col_d3 = st.columns([1.3, 1, 1])
            with col_d1:
                mes_sel_d = st.selectbox(
                    "Mês:", rotulos_d, key="fin_mes_sel",
                    help="Abre no mês do dia anterior; a tabela começa a partir de ontem.",
                )
            with col_d2:
                opcoes_canal_d = ["Todos"] + sorted(df_fin[COL_FIN_CANAL].dropna().astype(str).unique().tolist())
                canal_sel_d = st.selectbox("Canal:", opcoes_canal_d, key="fin_canal_sel_diario")
            with col_d3:
                opcoes_modal_d = ["Todas"] + sorted(df_fin[COL_FIN_MODALIDADE].dropna().astype(str).unique().tolist())
                modal_sel_d = st.selectbox("Modalidade:", opcoes_modal_d, key="fin_modal_sel_diario")

            periodo_d = meses_disp_d[rotulos_d.index(mes_sel_d)]
            df_d = df_fin[df_fin["Data Efetiva"].dt.to_period("M") == periodo_d].copy()
            if canal_sel_d != "Todos":
                df_d = df_d[df_d[COL_FIN_CANAL].astype(str) == canal_sel_d]
            if modal_sel_d != "Todas":
                df_d = df_d[df_d[COL_FIN_MODALIDADE].astype(str) == modal_sel_d]
            # Aplicações ficam fora, igual ao resto do painel
            df_d = df_d[df_d["Tipo Movimento"] != "aplicacao"]

            if df_d.empty:
                st.info("Nenhum lançamento nesse mês para os filtros selecionados.")
            else:
                df_d["DiaOrd"] = df_d["Data Efetiva"].dt.normalize()

                # Recorte de dias no mesmo estilo do "ajuste fino por dia":
                # dia inicial e final, em vez de uma lista de etiquetas.
                ultimo_dia_mes_d = periodo_d.end_time.day
                # No mês do dia anterior, a visão abre a partir de ONTEM e vai
                # até o fim do mês -- é a leitura de quem quer ver o que vem
                # pela frente (o que já passou está fechado). Nos demais
                # meses, mostra o mês inteiro.
                if periodo_d == periodo_mes_ontem:
                    dia_ini_padrao_d = min(data_ontem_d.day, ultimo_dia_mes_d)
                else:
                    dia_ini_padrao_d = 1
                dia_fim_padrao_d = ultimo_dia_mes_d

                # IMPORTANTE: quando um campo tem `key`, o Streamlit ignora o
                # `value` e usa o que já está guardado na sessão. Por isso o
                # padrão precisa ser gravado direto no session_state -- senão
                # um valor antigo fica preso e o padrão nunca aparece. Refaz
                # também quando a pessoa troca de mês, porque o dia inicial e
                # o final válidos mudam de um mês pro outro.
                if st.session_state.get("_fin_mes_dias_ref") != mes_sel_d:
                    st.session_state["fin_dia_ini_diario"] = dia_ini_padrao_d
                    st.session_state["fin_dia_fim_diario"] = dia_fim_padrao_d
                    st.session_state["_fin_mes_dias_ref"] = mes_sel_d

                with st.expander("📆 Recortar dias do mês (opcional)"):
                    col_dd1, col_dd2 = st.columns(2)
                    with col_dd1:
                        dia_ini_d = st.number_input(
                            "Do dia", min_value=1, max_value=ultimo_dia_mes_d, step=1,
                            key="fin_dia_ini_diario",
                            help="No mês corrente, começa no dia anterior — o que já passou está fechado.",
                        )
                    with col_dd2:
                        dia_fim_d = st.number_input(
                            "Até o dia", min_value=1, max_value=ultimo_dia_mes_d, step=1,
                            key="fin_dia_fim_diario",
                            help="Por padrão vai até o último dia do mês.",
                        )
                dia_ini_valido_d, dia_fim_valido_d = sorted([int(dia_ini_d), int(dia_fim_d)])
                # Guarda o MÊS INTEIRO antes de cortar os dias: é dele que sai
                # o saldo acumulado dos dias anteriores ao recorte (senão o
                # saldo do primeiro dia exibido começaria do zero) e é ele que
                # alimenta o gráfico, que mostra sempre o mês completo.
                df_d_mes = df_d.copy()
                df_d = df_d[
                    (df_d["DiaOrd"].dt.day >= dia_ini_valido_d)
                    & (df_d["DiaOrd"].dt.day <= dia_fim_valido_d)
                ]

                if df_d.empty:
                    st.info("Nenhum lançamento nos dias selecionados.")
                else:
                    dias_ordenados_d = sorted(df_d["DiaOrd"].unique())
                    rotulos_dias_d = [pd.Timestamp(d).strftime("%d/%m") for d in dias_ordenados_d]

                    # ---- KPIs próprios da visão diária ----
                    entradas_d = df_d.loc[df_d["Tipo Movimento"] == "entrada", COL_FIN_VALOR].sum()
                    saidas_d = df_d.loc[df_d["Tipo Movimento"] == "saida", COL_FIN_VALOR].sum()
                    fluxo_liquido_d = entradas_d + saidas_d
                    saldo_final_d, data_saldo_final_d = _saldo_posicao_atual_fin(df_d, COL_FIN_VALOR)

                    df_d_fluxo = df_d[df_d["Tipo Movimento"].isin(["entrada", "saida"])]
                    por_dia_liquido_d = (
                        df_d_fluxo.groupby("DiaOrd")[COL_FIN_VALOR].sum().reindex(dias_ordenados_d, fill_value=0)
                        if not df_d_fluxo.empty else pd.Series(dtype=float)
                    )
                    if not por_dia_liquido_d.empty:
                        dia_melhor_d = por_dia_liquido_d.idxmax()
                        dia_pior_d = por_dia_liquido_d.idxmin()
                        rotulo_melhor_d = pd.Timestamp(dia_melhor_d).strftime("%d/%m")
                        rotulo_pior_d = pd.Timestamp(dia_pior_d).strftime("%d/%m")
                        valor_melhor_d = por_dia_liquido_d.max()
                        valor_pior_d = por_dia_liquido_d.min()
                    else:
                        rotulo_melhor_d = rotulo_pior_d = "—"
                        valor_melhor_d = valor_pior_d = 0.0

                    dias_negativos_d = int((por_dia_liquido_d < 0).sum()) if not por_dia_liquido_d.empty else 0

                    st.markdown(
                        f'<div class="section-title">📊 Resumo de {mes_sel_d} '
                        f'<span style="font-weight:400;font-size:12px;color:{COLORS["text_muted"]};">'
                        f'(dias {dia_ini_valido_d} a {dia_fim_valido_d})</span></div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        render_kpi_row([
                            dict(label="SALDO EM CAIXA / BANCO", value=formata_brl(saldo_final_d),
                                 value_color=cor_variacao(saldo_final_d),
                                 subtext=(f"Posição em {data_saldo_final_d:%d/%m/%Y}"
                                          if data_saldo_final_d is not None else "Sem posição no recorte"),
                                 icon="🏦"),
                            dict(label="ENTRADAS NO PERÍODO", value=formata_brl(entradas_d),
                                 value_color=COLORS["positive"], subtext=f"{len(dias_ordenados_d)} dias com movimento", icon="📥"),
                            dict(label="SAÍDAS NO PERÍODO", value=formata_brl(saidas_d),
                                 value_color=COLORS["negative"], subtext="Contas a pagar dos dias", icon="📤"),
                            dict(label="FLUXO LÍQUIDO", value=formata_brl(fluxo_liquido_d),
                                 value_color=cor_variacao(fluxo_liquido_d), subtext="Entradas − saídas", icon="⚖️"),
                            dict(label="MELHOR / PIOR DIA",
                                 value=f"{rotulo_melhor_d} / {rotulo_pior_d}",
                                 value_color=COLORS["text"],
                                 subtext=f"{formata_m(valor_melhor_d)} · {formata_m(valor_pior_d)}", icon="📈"),
                            dict(label="DIAS COM FLUXO NEGATIVO", value=str(dias_negativos_d),
                                 value_color=COLORS["negative"] if dias_negativos_d else COLORS["positive"],
                                 subtext=f"de {len(dias_ordenados_d)} dias com movimento", icon="⚠️"),
                        ]),
                        unsafe_allow_html=True,
                    )
                    st.markdown("<br>", unsafe_allow_html=True)

                    # Estrutura igual à da planilha: os canais como grupos, e
                    # dentro de cada um os movimentos, com subtotal por canal
                    # e Total Geral no fim.
                    def _agrega_por_dia(df_origem):
                        serie = df_origem.groupby("DiaOrd")[COL_FIN_VALOR].sum()
                        return serie.reindex(dias_ordenados_d, fill_value=0.0)

                    linhas_tabela_d = []
                    indices_tabela_d = []
                    estilo_linhas_d = []

                    # O mesmo movimento aparece em vários canais, o que geraria
                    # rótulos repetidos -- e o Styler do pandas não aceita
                    # índice duplicado. Daí o sufixo invisível.
                    def _rotulo_unico_d(texto, posicao):
                        return texto + ("\u200b" * posicao)

                    canais_ordenados_d = sorted(df_d[COL_FIN_CANAL].dropna().astype(str).unique())
                    for canal in canais_ordenados_d:
                        df_canal = df_d[df_d[COL_FIN_CANAL].astype(str) == canal]
                        linhas_tabela_d.append(_agrega_por_dia(df_canal))
                        indices_tabela_d.append(_rotulo_unico_d(canal, len(indices_tabela_d)))
                        estilo_linhas_d.append(("canal", canal))

                        for movimento in sorted(df_canal[COL_FIN_MOVIMENTO].dropna().astype(str).unique()):
                            df_mov = df_canal[df_canal[COL_FIN_MOVIMENTO].astype(str) == movimento]
                            linhas_tabela_d.append(_agrega_por_dia(df_mov))
                            indices_tabela_d.append(_rotulo_unico_d(f"    {movimento}", len(indices_tabela_d)))
                            estilo_linhas_d.append(("movimento", movimento))

                    linhas_tabela_d.append(_agrega_por_dia(df_d))
                    indices_tabela_d.append(_rotulo_unico_d("TOTAL GERAL", len(indices_tabela_d)))
                    estilo_linhas_d.append(("total", "TOTAL GERAL"))

                    pivot_d = pd.DataFrame(linhas_tabela_d, index=indices_tabela_d)
                    pivot_d.columns = rotulos_dias_d

                    # Coluna final: soma do período pro fluxo; última posição
                    # pras linhas de saldo (somar saldo de dias diferentes não
                    # faz sentido).
                    totais_finais_d = []
                    for posicao, (tipo_linha, nome_limpo) in enumerate(estilo_linhas_d):
                        if tipo_linha == "movimento":
                            if _classificar_movimento_fin(nome_limpo) in ("saldo", "aplicacao"):
                                serie_linha = pivot_d.iloc[posicao]
                                nao_zerados = serie_linha[serie_linha != 0]
                                totais_finais_d.append(nao_zerados.iloc[-1] if not nao_zerados.empty else 0.0)
                            else:
                                totais_finais_d.append(pivot_d.iloc[posicao].sum())
                        else:
                            if tipo_linha == "canal":
                                df_escopo = df_d[df_d[COL_FIN_CANAL].astype(str) == nome_limpo]
                            else:
                                df_escopo = df_d
                            fluxo_escopo = df_escopo[df_escopo["Tipo Movimento"].isin(["entrada", "saida"])][COL_FIN_VALOR].sum()
                            saldo_escopo = 0.0
                            df_saldo_escopo = df_escopo[df_escopo["Tipo Movimento"] == "saldo"]
                            if not df_saldo_escopo.empty:
                                ultimo_dia_saldo = df_saldo_escopo["DiaOrd"].max()
                                saldo_escopo = df_saldo_escopo.loc[
                                    df_saldo_escopo["DiaOrd"] == ultimo_dia_saldo, COL_FIN_VALOR
                                ].sum()
                            totais_finais_d.append(fluxo_escopo + saldo_escopo)

                    pivot_d["TOTAL / ÚLT. POSIÇÃO"] = totais_finais_d
                    pivot_d.index.name = "Canal / Movimento"

                    # ---- Linha SALDO: igual à da planilha ----
                    # Cada dia soma o Total Geral daquele dia com o saldo do
                    # dia anterior. IMPORTANTE: o acúmulo parte do INÍCIO DO
                    # MÊS, não do primeiro dia exibido -- se você recorta a
                    # partir do dia 11, o saldo do dia 11 já tem que trazer
                    # tudo que veio dos dias 1 a 10. Por isso o cálculo usa o
                    # mês inteiro e só depois recorta as colunas visíveis.
                    total_geral_mes_dia = (
                        df_d_mes.groupby("DiaOrd")[COL_FIN_VALOR].sum().sort_index()
                    )
                    saldo_acumulado_mes = total_geral_mes_dia.cumsum()
                    saldo_dias_exibidos = [
                        float(saldo_acumulado_mes.get(dia, 0.0)) for dia in dias_ordenados_d
                    ]
                    linha_saldo_d = saldo_dias_exibidos + [
                        saldo_dias_exibidos[-1] if saldo_dias_exibidos else 0.0
                    ]
                    pivot_d.loc[_rotulo_unico_d("SALDO (acumulado)", len(pivot_d))] = linha_saldo_d
                    estilo_linhas_d.append(("saldo_acumulado", "SALDO (acumulado)"))

                    st.markdown(
                        f'<div class="section-title">📋 Fluxo Diário por Canal — {mes_sel_d}</div>',
                        unsafe_allow_html=True,
                    )

                    def _estilo_tabela_diaria(df_tabela):
                        """Destaca canais, total geral e saldo acumulado. Usa
                        posição, não o nome da linha, então funciona mesmo com
                        rótulos parecidos entre canais."""
                        estilos = pd.DataFrame("", index=df_tabela.index, columns=df_tabela.columns)
                        for posicao, (tipo_linha, _) in enumerate(estilo_linhas_d):
                            for coluna in df_tabela.columns:
                                valor = df_tabela.iloc[posicao][coluna]
                                base = cor_valor(valor)
                                if tipo_linha == "canal":
                                    base += (
                                        f" background-color: {COLORS['surface_alt']};"
                                        " font-weight: 700; border-top: 1px solid rgba(139,149,165,0.35);"
                                    )
                                elif tipo_linha == "total":
                                    # Total Geral: faixa mais clara e borda
                                    # grossa em cima, pra fechar visualmente o
                                    # bloco dos canais.
                                    base += (
                                        " background-color: rgba(139,149,165,0.20);"
                                        " font-weight: 800; border-top: 2px solid rgba(139,149,165,0.75);"
                                    )
                                elif tipo_linha == "saldo_acumulado":
                                    # Saldo: destaque azul mais forte, é a
                                    # linha mais consultada da tabela.
                                    base += (
                                        " background-color: rgba(76,141,255,0.26);"
                                        " font-weight: 800; border-top: 2px solid rgba(76,141,255,0.85);"
                                    )
                                estilos.iloc[posicao, df_tabela.columns.get_loc(coluna)] = base
                        return estilos

                    # Largura confortável em todas as colunas de valor -- sem
                    # isso o Streamlit espreme as colunas quando o mês tem
                    # muitos dias, e os números ficam ilegíveis.
                    config_colunas_d = {
                        coluna: st.column_config.NumberColumn(coluna, width="medium")
                        for coluna in pivot_d.columns
                    }
                    st.dataframe(
                        pivot_d.style.format(formata_brl).apply(_estilo_tabela_diaria, axis=None),
                        use_container_width=True,
                        column_config=config_colunas_d,
                        height=668,  # 18 linhas visíveis, sem precisar rolar
                    )
                    st.caption(
                        "Cada coluna é um dia. As linhas em destaque são os canais (com o subtotal do canal) "
                        "e, recuadas abaixo, os movimentos de cada um. A linha **SALDO (acumulado)** soma o "
                        "Total Geral do dia com o saldo do dia anterior — é a posição de caixa ao longo do "
                        "mês. Na última coluna, contas a receber/pagar trazem a **soma do período** e "
                        "caixa/banco trazem o **saldo do último dia** com movimento."
                    )

                    # ---- Gráfico: movimento do dia + fluxo acumulado ----
                    # O gráfico mostra sempre o MÊS INTEIRO, mesmo quando a
                    # tabela acima está recortada em alguns dias -- a curva do
                    # mês só faz sentido completa. Ao trocar de mês, ele passa
                    # a mostrar todos os dias do novo mês.
                    df_mes_fluxo_graf = df_d_mes[df_d_mes["Tipo Movimento"].isin(["entrada", "saida"])]
                    if not df_mes_fluxo_graf.empty:
                        dias_grafico_d = sorted(df_d_mes["DiaOrd"].unique())
                        rotulos_grafico_d = [pd.Timestamp(d).strftime("%d/%m") for d in dias_grafico_d]

                        por_dia_liquido_graf = (
                            df_mes_fluxo_graf.groupby("DiaOrd")[COL_FIN_VALOR].sum()
                            .reindex(dias_grafico_d, fill_value=0)
                        )
                        acumulado_d = por_dia_liquido_graf.cumsum()

                        # Entradas e saídas separadas por dia: mostra o
                        # volume bruto que passou, não só o líquido -- dois
                        # dias podem ter o mesmo resultado com movimentações
                        # de tamanhos bem diferentes.
                        entradas_por_dia_d = (
                            df_d_mes[df_d_mes["Tipo Movimento"] == "entrada"]
                            .groupby("DiaOrd")[COL_FIN_VALOR].sum()
                            .reindex(dias_grafico_d, fill_value=0)
                        )
                        saidas_por_dia_graf = (
                            df_d_mes[df_d_mes["Tipo Movimento"] == "saida"]
                            .groupby("DiaOrd")[COL_FIN_VALOR].sum().abs()
                            .reindex(dias_grafico_d, fill_value=0)
                        )
                        media_liquida_d = por_dia_liquido_graf.mean()

                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown(
                            f'<div class="section-title">📈 Movimento Diário — {mes_sel_d} (mês completo)</div>',
                            unsafe_allow_html=True,
                        )
                        fig_ac = go.Figure()
                        # Entradas para cima, saídas para baixo: o eixo zero
                        # separa visualmente o que entra do que sai.
                        fig_ac.add_trace(go.Bar(
                            name="Entradas", x=rotulos_grafico_d, y=list(entradas_por_dia_d.values),
                            marker=dict(color="rgba(62,207,142,0.30)",
                                        line=dict(color=COLORS["positive"], width=1.3)),
                            hovertemplate="Entradas: R$ %{y:,.2f}<extra></extra>",
                        ))
                        fig_ac.add_trace(go.Bar(
                            name="Saídas", x=rotulos_grafico_d, y=[-v for v in saidas_por_dia_graf.values],
                            marker=dict(color="rgba(247,110,110,0.30)",
                                        line=dict(color=COLORS["negative"], width=1.3)),
                            hovertemplate="Saídas: R$ %{y:,.2f}<extra></extra>",
                        ))
                        # Resultado líquido de cada dia
                        fig_ac.add_trace(go.Scatter(
                            name="Resultado do dia", x=rotulos_grafico_d, y=list(por_dia_liquido_graf.values),
                            mode="lines+markers", line=dict(color=COLORS["warning"], width=2, dash="dot"),
                            marker=dict(size=6, line=dict(color=COLORS["bg"], width=1.5)),
                            hovertemplate="Resultado: R$ %{y:,.2f}<extra></extra>",
                        ))
                        # Acumulado no eixo da direita: a curva do caixa ao
                        # longo do mês, que é a leitura mais importante aqui.
                        fig_ac.add_trace(go.Scatter(
                            name="Acumulado", x=rotulos_grafico_d, y=list(acumulado_d.values), yaxis="y2",
                            mode="lines", line=dict(color=COLORS["primary"], width=3.5),
                            fill="tozeroy", fillcolor="rgba(76,141,255,0.10)",
                            hovertemplate="Acumulado: R$ %{y:,.2f}<extra></extra>",
                        ))
                        # Referência: média diária do resultado
                        fig_ac.add_hline(
                            y=media_liquida_d, line=dict(color=COLORS["muted_line"], width=1.2, dash="dash"),
                            annotation_text=f"média/dia {formata_m(media_liquida_d)}",
                            annotation_position="top left",
                            annotation_font=dict(size=9, color=COLORS["text_muted"]),
                        )
                        # Marca o pior dia, que costuma ser o ponto de atenção
                        if not por_dia_liquido_graf.empty:
                            idx_pior = list(por_dia_liquido_graf.values).index(por_dia_liquido_graf.min())
                            fig_ac.add_annotation(
                                x=rotulos_grafico_d[idx_pior], y=por_dia_liquido_graf.min(),
                                text=f"pior dia<br>{formata_m(por_dia_liquido_graf.min())}",
                                showarrow=True, arrowhead=2, arrowsize=0.8,
                                arrowcolor=COLORS["negative"], ax=0, ay=38,
                                font=dict(size=9, color=COLORS["negative"]),
                                bgcolor="rgba(11,14,20,0.85)", bordercolor=COLORS["negative"], borderwidth=1,
                            )

                        estilo_grafico(
                            fig_ac, height=460, barmode="relative", bargap=0.25,
                            xaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickfont=dict(size=9),
                                       tickangle=-45, automargin=True),
                            yaxis=dict(
                                title=dict(text="Movimento do dia (R$)", font=dict(size=10, color=COLORS["text_muted"])),
                                gridcolor=COLORS["border"], fixedrange=True, tickformat=",.0f",
                                zerolinecolor=COLORS["text_muted"], zerolinewidth=1.5,
                            ),
                            yaxis2=dict(
                                title=dict(text="Acumulado (R$)", font=dict(size=10, color=COLORS["primary"])),
                                overlaying="y", side="right", showgrid=False, fixedrange=True,
                                tickformat=",.0f", tickfont=dict(size=9, color=COLORS["primary"]),
                            ),
                            legend=dict(orientation="h", yanchor="bottom", y=-0.32, xanchor="center", x=0.5,
                                        font=dict(size=10)),
                            margin=dict(l=80, r=80, t=30, b=95),
                            hovermode="x unified",
                        )
                        st.plotly_chart(fig_ac, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                        st.caption(
                            "Barras verdes (para cima) = entradas do dia; barras vermelhas (para baixo) = saídas. "
                            "Linha pontilhada laranja = resultado líquido do dia. Linha azul cheia = fluxo "
                            "acumulado no eixo da direita. A linha tracejada marca a média diária. "
                            "O gráfico mostra sempre o **mês inteiro**, mesmo quando a tabela acima está "
                            "recortada em alguns dias."
                        )

    # ---------------- TESOURARIA ----------------
    with tab_fin_tesouraria:
        st.caption(
            "Visão de tesouraria: para onde o caixa caminha nas próximas semanas, quanto de capital "
            "a operação exige e o risco de depender de poucos pagadores."
        )

        col_t1, col_t2 = st.columns(2)
        with col_t1:
            opcoes_canal_t = ["Todos"] + sorted(df_fin[COL_FIN_CANAL].dropna().astype(str).unique().tolist())
            canal_sel_t = st.selectbox("Canal:", opcoes_canal_t, key="fin_canal_sel_tesouraria")
        with col_t2:
            semanas_horizonte_t = st.slider(
                "Semanas à frente:", min_value=4, max_value=26, value=13, step=1,
                key="fin_semanas_tesouraria",
                help="13 semanas é o padrão de mercado em tesouraria (um trimestre).",
            )

        df_t = df_fin[df_fin["Tipo Movimento"] != "aplicacao"].copy()
        if canal_sel_t != "Todos":
            df_t = df_t[df_t[COL_FIN_CANAL].astype(str) == canal_sel_t]

        if df_t.empty:
            st.info("Nenhum lançamento para o canal selecionado.")
        else:
            hoje_t = pd.Timestamp(datetime.now(FUSO_BR).date())

            # =============================================================
            # 1. PREVISÃO DE CAIXA — PRÓXIMAS SEMANAS
            # =============================================================
            st.markdown(
                f'<div class="section-title">📆 Previsão de Caixa — próximas {semanas_horizonte_t} semanas</div>',
                unsafe_allow_html=True,
            )

            # Saldo de partida: posição de caixa/banco mais recente
            saldo_inicial_t, data_saldo_inicial_t = _saldo_posicao_atual_fin(df_t, COL_FIN_VALOR)

            # A projeção olha só o que está À FRENTE e ainda não liquidou --
            # é o compromisso futuro, que é o que a tesouraria precisa cobrir.
            fim_horizonte_t = hoje_t + pd.Timedelta(weeks=semanas_horizonte_t)
            df_futuro_t = df_t[
                (df_t["Tipo Movimento"].isin(["entrada", "saida"]))
                & (df_t["Data Efetiva"] >= hoje_t)
                & (df_t["Data Efetiva"] < fim_horizonte_t)
            ].copy()

            if df_futuro_t.empty:
                st.info(
                    "Não há lançamentos futuros dentro do horizonte escolhido. "
                    "Aumente o número de semanas ou verifique se há contas a receber/pagar lançadas à frente."
                )
            else:
                # Agrupa por semana (segunda-feira como início)
                df_futuro_t["InicioSemana"] = (
                    df_futuro_t["Data Efetiva"] - pd.to_timedelta(df_futuro_t["Data Efetiva"].dt.weekday, unit="D")
                ).dt.normalize()

                semanas_t = []
                inicio_semana_atual = (hoje_t - pd.Timedelta(days=hoje_t.weekday())).normalize()
                for i in range(semanas_horizonte_t):
                    semanas_t.append(inicio_semana_atual + pd.Timedelta(weeks=i))

                entradas_sem = (
                    df_futuro_t[df_futuro_t["Tipo Movimento"] == "entrada"]
                    .groupby("InicioSemana")[COL_FIN_VALOR].sum().reindex(semanas_t, fill_value=0)
                )
                saidas_sem = (
                    df_futuro_t[df_futuro_t["Tipo Movimento"] == "saida"]
                    .groupby("InicioSemana")[COL_FIN_VALOR].sum().reindex(semanas_t, fill_value=0)
                )
                liquido_sem = entradas_sem + saidas_sem
                saldo_projetado_sem = saldo_inicial_t + liquido_sem.cumsum()

                rotulos_sem = [
                    f"S{i + 1}\n{pd.Timestamp(s).strftime('%d/%m')}" for i, s in enumerate(semanas_t)
                ]

                # KPIs da projeção
                menor_saldo_t = saldo_projetado_sem.min()
                idx_menor_t = int(saldo_projetado_sem.values.argmin())
                semana_critica_t = pd.Timestamp(semanas_t[idx_menor_t]).strftime("%d/%m/%Y")
                semanas_negativas_t = int((saldo_projetado_sem < 0).sum())
                saldo_final_t = saldo_projetado_sem.iloc[-1]

                st.markdown(
                    render_kpi_row([
                        dict(label="SALDO DE PARTIDA", value=formata_brl(saldo_inicial_t),
                             value_color=cor_variacao(saldo_inicial_t),
                             subtext=(f"Posição em {data_saldo_inicial_t:%d/%m/%Y}"
                                      if data_saldo_inicial_t is not None else "Sem posição registrada"),
                             icon="🏦"),
                        dict(label="ENTRADAS PREVISTAS", value=formata_brl(entradas_sem.sum()),
                             value_color=COLORS["positive"],
                             subtext=f"Nas próximas {semanas_horizonte_t} semanas", icon="📥"),
                        dict(label="SAÍDAS PREVISTAS", value=formata_brl(saidas_sem.sum()),
                             value_color=COLORS["negative"],
                             subtext=f"Nas próximas {semanas_horizonte_t} semanas", icon="📤"),
                        dict(label="SALDO AO FIM DO HORIZONTE", value=formata_brl(saldo_final_t),
                             value_color=cor_variacao(saldo_final_t),
                             subtext="Partida + entradas − saídas", icon="🎯"),
                        dict(label="MENOR SALDO PROJETADO", value=formata_brl(menor_saldo_t),
                             value_color=cor_variacao(menor_saldo_t),
                             subtext=f"Na semana de {semana_critica_t}", icon="⚠️"),
                    ]),
                    unsafe_allow_html=True,
                )

                if semanas_negativas_t:
                    st.error(
                        f"🚨 **{semanas_negativas_t} semana(s) com saldo projetado negativo.** "
                        f"O ponto mais crítico é a semana de {semana_critica_t}, com "
                        f"{formata_brl(menor_saldo_t)}. Vale antecipar recebimentos ou renegociar "
                        "vencimentos antes disso.".replace("$", "\\$")
                    )
                else:
                    st.success(
                        f"✅ Nenhuma semana com saldo negativo no horizonte. O menor ponto é "
                        f"{formata_brl(menor_saldo_t)}, na semana de {semana_critica_t}."
                        .replace("$", "\\$")
                    )

                # Gráfico: barras de entradas/saídas + linha do saldo projetado
                fig_t = go.Figure()
                fig_t.add_trace(go.Bar(
                    name="Entradas", x=rotulos_sem, y=list(entradas_sem.values),
                    marker=dict(color="rgba(62,207,142,0.30)",
                                line=dict(color=COLORS["positive"], width=1.3)),
                    hovertemplate="Entradas: R$ %{y:,.2f}<extra></extra>",
                ))
                fig_t.add_trace(go.Bar(
                    name="Saídas", x=rotulos_sem, y=list(saidas_sem.values),
                    marker=dict(color="rgba(247,110,110,0.30)",
                                line=dict(color=COLORS["negative"], width=1.3)),
                    hovertemplate="Saídas: R$ %{y:,.2f}<extra></extra>",
                ))
                cores_saldo_t = [
                    COLORS["negative"] if v < 0 else COLORS["primary"] for v in saldo_projetado_sem.values
                ]
                fig_t.add_trace(go.Scatter(
                    name="Saldo projetado", x=rotulos_sem, y=list(saldo_projetado_sem.values), yaxis="y2",
                    mode="lines+markers", line=dict(color=COLORS["primary"], width=3),
                    marker=dict(size=9, color=cores_saldo_t, line=dict(color=COLORS["bg"], width=2)),
                    hovertemplate="Saldo: R$ %{y:,.2f}<extra></extra>",
                ))
                fig_t.add_hline(
                    y=0, yref="y2", line=dict(color=COLORS["negative"], width=1.5, dash="dash"),
                    annotation_text="saldo zero", annotation_position="right",
                    annotation_font=dict(size=9, color=COLORS["negative"]),
                )
                estilo_grafico(
                    fig_t, height=460, barmode="relative", bargap=0.28,
                    xaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickfont=dict(size=9)),
                    yaxis=dict(title=dict(text="Movimento da semana (R$)",
                                          font=dict(size=10, color=COLORS["text_muted"])),
                               gridcolor=COLORS["border"], fixedrange=True, tickformat=",.0f",
                               zerolinecolor=COLORS["text_muted"]),
                    yaxis2=dict(title=dict(text="Saldo projetado (R$)",
                                           font=dict(size=10, color=COLORS["primary"])),
                                overlaying="y", side="right", showgrid=False, fixedrange=True,
                                tickformat=",.0f", tickfont=dict(size=9, color=COLORS["primary"])),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5),
                    margin=dict(l=80, r=80, t=25, b=90),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_t, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

                with st.expander("📋 Ver a projeção semana a semana"):
                    df_proj_t = pd.DataFrame({
                        "Semana": [f"S{i + 1} — {pd.Timestamp(s).strftime('%d/%m/%Y')}"
                                   for i, s in enumerate(semanas_t)],
                        "Entradas": entradas_sem.values,
                        "Saídas": saidas_sem.values,
                        "Líquido": liquido_sem.values,
                        "Saldo projetado": saldo_projetado_sem.values,
                    })
                    st.dataframe(
                        df_proj_t.style.format({
                            "Entradas": formata_brl, "Saídas": formata_brl,
                            "Líquido": formata_brl, "Saldo projetado": formata_brl,
                        }).map(cor_valor, subset=["Entradas", "Saídas", "Líquido", "Saldo projetado"]),
                        use_container_width=True, hide_index=True, height=min(520, 60 + 35 * len(df_proj_t)),
                    )
                st.caption(
                    "A projeção parte do saldo de caixa/banco mais recente e acumula os lançamentos "
                    "com vencimento em cada semana. Considera apenas o que ainda está à frente de hoje."
                )

            st.markdown("<br>", unsafe_allow_html=True)

            # =============================================================
            # 2. NECESSIDADE DE CAPITAL DE GIRO
            # =============================================================
            st.markdown('<div class="section-title">⚙️ Necessidade de Capital de Giro (NCG)</div>', unsafe_allow_html=True)
            st.caption(
                "NCG = o que a operação tem a receber menos o que tem a pagar. Quando é positiva, a empresa "
                "precisa bancar essa diferença com capital próprio ou crédito enquanto o dinheiro não entra."
            )

            # Direitos e obrigações em aberto (não liquidados)
            df_aberto_t = df_t[df_t[COL_FIN_DATA_LIQUIDACAO].isna()]
            a_receber_ncg = df_aberto_t.loc[df_aberto_t["Tipo Movimento"] == "entrada", COL_FIN_VALOR].sum()
            a_pagar_ncg = abs(df_aberto_t.loc[df_aberto_t["Tipo Movimento"] == "saida", COL_FIN_VALOR].sum())
            ncg_t = a_receber_ncg - a_pagar_ncg
            capital_disponivel_t = saldo_inicial_t
            folga_t = capital_disponivel_t - max(ncg_t, 0)
            cobertura_ncg = (capital_disponivel_t / ncg_t * 100) if ncg_t > 0 else None

            st.markdown(
                render_kpi_row([
                    dict(label="A RECEBER EM ABERTO", value=formata_brl(a_receber_ncg),
                         value_color=COLORS["positive"], subtext="Direitos ainda não liquidados", icon="📥"),
                    dict(label="A PAGAR EM ABERTO", value=formata_brl(-a_pagar_ncg),
                         value_color=COLORS["negative"], subtext="Obrigações ainda não quitadas", icon="📤"),
                    dict(label="NCG", value=formata_brl(ncg_t),
                         value_color=COLORS["warning"] if ncg_t > 0 else COLORS["positive"],
                         subtext="Positiva = a operação consome caixa", icon="⚙️"),
                    dict(label="CAPITAL DISPONÍVEL", value=formata_brl(capital_disponivel_t),
                         value_color=cor_variacao(capital_disponivel_t),
                         subtext="Saldo em caixa e banco", icon="🏦"),
                    dict(label="COBERTURA DA NCG",
                         value=(f"{cobertura_ncg:.0f}%" if cobertura_ncg is not None else "n/a"),
                         value_color=(COLORS["positive"] if (cobertura_ncg or 0) >= 100 else COLORS["negative"]),
                         subtext=(f"Folga de {formata_brl(folga_t)}" if ncg_t > 0
                                  else "NCG negativa: a operação se financia sozinha"),
                         icon="🛡️"),
                ]),
                unsafe_allow_html=True,
            )

            if ncg_t > 0 and capital_disponivel_t < ncg_t:
                st.warning(
                    f"⚠️ O capital disponível cobre apenas {cobertura_ncg:.0f}% da necessidade de capital "
                    f"de giro. Faltam {formata_brl(ncg_t - capital_disponivel_t)} para bancar o ciclo "
                    "operacional sem recorrer a crédito.".replace("$", "\\$")
                )
            elif ncg_t <= 0:
                st.success(
                    "✅ A NCG está negativa: as obrigações em aberto superam os direitos, ou seja, a "
                    "operação está sendo financiada pelos próprios fornecedores — situação confortável "
                    "para o caixa, desde que os prazos sejam honrados."
                )
            else:
                st.success(
                    f"✅ O capital disponível cobre a NCG com folga de {formata_brl(folga_t)}."
                    .replace("$", "\\$")
                )

            st.markdown("<br>", unsafe_allow_html=True)

            # =============================================================
            # 3. CONCENTRAÇÃO DE RECEBIMENTO
            # =============================================================
            st.markdown('<div class="section-title">🎯 Concentração de Recebimento</div>', unsafe_allow_html=True)
            st.caption(
                "Quanto do que a empresa recebe depende de poucas fontes. Concentração alta significa que "
                "a falha de um único pagador ou canal derruba o caixa do período."
            )

            df_entradas_t = df_t[df_t["Tipo Movimento"] == "entrada"]
            if df_entradas_t.empty:
                st.info("Sem recebimentos no recorte selecionado.")
            else:
                col_conc_a, col_conc_b = st.columns(2)

                # --- Por canal ---
                with col_conc_a:
                    st.markdown("**Por canal**")
                    receb_canal = (
                        df_entradas_t.groupby(COL_FIN_CANAL, observed=True)[COL_FIN_VALOR]
                        .sum().sort_values(ascending=False)
                    )
                    receb_canal = receb_canal[receb_canal > 0]
                    if not receb_canal.empty:
                        total_canal = receb_canal.sum()
                        maior_canal_pct = receb_canal.iloc[0] / total_canal * 100
                        fig_conc_canal = go.Figure(data=[go.Pie(
                            labels=[str(i) for i in receb_canal.index], values=list(receb_canal.values),
                            hole=0.55,
                            marker=dict(colors=[COLORS["primary"], COLORS["positive"], COLORS["warning"],
                                                 COLORS["secondary"], COLORS["muted_line"]]),
                            textinfo="percent", texttemplate="%{percent:.1%}",
                            textfont=dict(color=COLORS["text"], size=11),
                            hovertemplate="%{label}: R$ %{value:,.2f}<extra></extra>",
                        )])
                        fig_conc_canal.add_annotation(
                            text=f"<b>{maior_canal_pct:.0f}%</b><br>"
                                 f"<span style='font-size:9px;color:{COLORS['text_muted']}'>no maior canal</span>",
                            showarrow=False, font=dict(color=COLORS["text"], size=14, family=FONT_STACK),
                        )
                        estilo_grafico(
                            fig_conc_canal, height=300,
                            legend=dict(orientation="h", yanchor="top", y=-0.05, xanchor="center",
                                        x=0.5, font=dict(size=9)),
                            margin=dict(l=10, r=10, t=15, b=10),
                        )
                        st.plotly_chart(fig_conc_canal, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

                # --- Por modalidade ---
                with col_conc_b:
                    st.markdown("**Por modalidade de recebimento**")
                    receb_modal = (
                        df_entradas_t.groupby(COL_FIN_MODALIDADE, observed=True)[COL_FIN_VALOR]
                        .sum().sort_values(ascending=False)
                    )
                    receb_modal = receb_modal[receb_modal > 0]
                    if not receb_modal.empty:
                        TOP_MODAL_T = 6
                        if len(receb_modal) > TOP_MODAL_T:
                            outros_t = receb_modal.iloc[TOP_MODAL_T:].sum()
                            receb_modal = pd.concat([receb_modal.iloc[:TOP_MODAL_T],
                                                     pd.Series({"Outros": outros_t})])
                        total_modal = receb_modal.sum()
                        maior_modal_pct = receb_modal.iloc[0] / total_modal * 100
                        fig_conc_modal = go.Figure(data=[go.Pie(
                            labels=[str(i) for i in receb_modal.index], values=list(receb_modal.values),
                            hole=0.55,
                            marker=dict(colors=[COLORS["positive"], COLORS["primary"], COLORS["warning"],
                                                 COLORS["negative"], COLORS["secondary"],
                                                 COLORS["muted_line"], "#6B7280"]),
                            textinfo="percent", texttemplate="%{percent:.1%}",
                            textfont=dict(color=COLORS["text"], size=11),
                            hovertemplate="%{label}: R$ %{value:,.2f}<extra></extra>",
                        )])
                        fig_conc_modal.add_annotation(
                            text=f"<b>{maior_modal_pct:.0f}%</b><br>"
                                 f"<span style='font-size:9px;color:{COLORS['text_muted']}'>na maior modalidade</span>",
                            showarrow=False, font=dict(color=COLORS["text"], size=14, family=FONT_STACK),
                        )
                        estilo_grafico(
                            fig_conc_modal, height=300,
                            legend=dict(orientation="h", yanchor="top", y=-0.05, xanchor="center",
                                        x=0.5, font=dict(size=9)),
                            margin=dict(l=10, r=10, t=15, b=10),
                        )
                        st.plotly_chart(fig_conc_modal, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

                # Índice de concentração (HHI simplificado) por canal
                if not receb_canal.empty:
                    participacoes = (receb_canal / receb_canal.sum() * 100)
                    hhi = float((participacoes ** 2).sum())
                    if hhi >= 5000:
                        nivel_conc, cor_conc = "ALTA", COLORS["negative"]
                    elif hhi >= 2500:
                        nivel_conc, cor_conc = "MODERADA", COLORS["warning"]
                    else:
                        nivel_conc, cor_conc = "BAIXA", COLORS["positive"]

                    st.markdown(
                        render_kpi_row([
                            dict(label="MAIOR CANAL", value=f"{participacoes.iloc[0]:.0f}%",
                                 value_color=cor_conc,
                                 subtext=f"{participacoes.index[0]} · {formata_brl(receb_canal.iloc[0])}",
                                 icon="🥇"),
                            dict(label="DOIS MAIORES JUNTOS",
                                 value=f"{participacoes.iloc[:2].sum():.0f}%",
                                 value_color=cor_conc, subtext="Do total recebido", icon="🥈"),
                            dict(label="NÍVEL DE CONCENTRAÇÃO", value=nivel_conc,
                                 value_color=cor_conc,
                                 subtext=f"Índice HHI: {hhi:,.0f}".replace(",", "."), icon="⚖️"),
                        ]),
                        unsafe_allow_html=True,
                    )
                    st.caption(
                        "O índice HHI mede concentração: acima de 5.000 indica dependência alta de poucas "
                        "fontes; abaixo de 2.500, receita bem distribuída. Quanto mais concentrado, maior o "
                        "impacto no caixa se um canal falhar."
                    )


    # ---------------- ANÁLISES ----------------
    with tab_fin_analises:

        meses_disp_a = sorted(df_fin["Data Efetiva"].dt.to_period("M").unique())
        rotulos_a = ["Todo o período"] + [_rotulo_mes_pt_extenso(p) for p in meses_disp_a]
        col_a1, col_a2, col_a3 = st.columns(3)
        with col_a1:
            mes_sel_a = st.selectbox("Mês:", rotulos_a, key="fin_mes_sel_analise")
        with col_a2:
            opcoes_canal_a = ["Todos"] + sorted(df_fin[COL_FIN_CANAL].dropna().astype(str).unique().tolist())
            canal_sel_a = st.selectbox("Canal:", opcoes_canal_a, key="fin_canal_sel_analise")
        with col_a3:
            opcoes_modal_a = ["Todas"] + sorted(df_fin[COL_FIN_MODALIDADE].dropna().astype(str).unique().tolist())
            modal_sel_a = st.selectbox("Modalidade:", opcoes_modal_a, key="fin_modal_sel_analise")

        df_a = df_fin[df_fin["Tipo Movimento"] != "aplicacao"].copy()
        if mes_sel_a != "Todo o período":
            periodo_a = meses_disp_a[rotulos_a.index(mes_sel_a) - 1]
            df_a = df_a[df_a["Data Efetiva"].dt.to_period("M") == periodo_a]
        if canal_sel_a != "Todos":
            df_a = df_a[df_a[COL_FIN_CANAL].astype(str) == canal_sel_a]
        if modal_sel_a != "Todas":
            df_a = df_a[df_a[COL_FIN_MODALIDADE].astype(str) == modal_sel_a]

        if df_a.empty:
            st.info("Nenhum lançamento para os filtros selecionados.")
        else:
            df_a["DiaOrd"] = df_a["Data Efetiva"].dt.normalize()
            entradas_a = df_a.loc[df_a["Tipo Movimento"] == "entrada", COL_FIN_VALOR].sum()
            saidas_a = df_a.loc[df_a["Tipo Movimento"] == "saida", COL_FIN_VALOR].sum()
            saldo_a, data_saldo_a = _saldo_posicao_atual_fin(df_a, COL_FIN_VALOR)
            dias_com_mov_a = max(df_a["DiaOrd"].nunique(), 1)

            # ---- Prazos e cobertura ----
            df_prazo_a = df_a[
                df_a[COL_FIN_DATA_LIQUIDACAO].notna() & df_a[COL_FIN_VENCIMENTO].notna()
            ].copy()
            if not df_prazo_a.empty:
                df_prazo_a["DiasAteLiquidar"] = (
                    df_prazo_a[COL_FIN_DATA_LIQUIDACAO] - df_prazo_a[COL_FIN_VENCIMENTO]
                ).dt.days
            saidas_prazo_a = df_prazo_a[df_prazo_a["Tipo Movimento"] == "saida"] if not df_prazo_a.empty else pd.DataFrame()
            entradas_prazo_a = df_prazo_a[df_prazo_a["Tipo Movimento"] == "entrada"] if not df_prazo_a.empty else pd.DataFrame()

            prazo_medio_pagto_a = saidas_prazo_a["DiasAteLiquidar"].mean() if not saidas_prazo_a.empty else None
            prazo_medio_receb_a = entradas_prazo_a["DiasAteLiquidar"].mean() if not entradas_prazo_a.empty else None
            pct_pago_em_dia_a = (
                (saidas_prazo_a["DiasAteLiquidar"] <= 0).mean() * 100 if not saidas_prazo_a.empty else None
            )
            saida_media_diaria_a = abs(saidas_a) / dias_com_mov_a
            entrada_media_diaria_a = entradas_a / dias_com_mov_a
            dias_de_caixa_a = (saldo_a / saida_media_diaria_a) if saida_media_diaria_a else None

            saidas_por_dia_a = df_a[df_a["Tipo Movimento"] == "saida"].groupby("DiaOrd")[COL_FIN_VALOR].sum().abs()
            if not saidas_por_dia_a.empty and saidas_por_dia_a.sum():
                pct_top3_a = saidas_por_dia_a.nlargest(3).sum() / saidas_por_dia_a.sum() * 100
                rotulo_pico_a = pd.Timestamp(saidas_por_dia_a.idxmax()).strftime("%d/%m/%Y")
                valor_pico_a = saidas_por_dia_a.max()
            else:
                pct_top3_a, rotulo_pico_a, valor_pico_a = 0.0, "—", 0.0

            # ---- Contas a pagar EM ABERTO ----
            # Data de Liquidação em branco significa que ainda NÃO foi pago
            # (não é dado faltando). A previsão de pagamento é o próprio
            # vencimento, e a política interna é quitar em até 10 dias.
            PRAZO_INTERNO_DIAS = 10
            hoje_analise = pd.Timestamp(datetime.now(FUSO_BR).date())
            df_pagar_a = df_a[df_a["Tipo Movimento"] == "saida"].copy()
            df_pagar_aberto = df_pagar_a[df_pagar_a[COL_FIN_DATA_LIQUIDACAO].isna()].copy()
            df_pagar_quitado = df_pagar_a[df_pagar_a[COL_FIN_DATA_LIQUIDACAO].notna()].copy()

            valor_aberto_a = abs(df_pagar_aberto[COL_FIN_VALOR].sum())
            valor_quitado_a = abs(df_pagar_quitado[COL_FIN_VALOR].sum())
            total_pagar_a = valor_aberto_a + valor_quitado_a
            pct_quitado_a = (valor_quitado_a / total_pagar_a * 100) if total_pagar_a else 0

            if not df_pagar_aberto.empty:
                venc_aberto = df_pagar_aberto[COL_FIN_VENCIMENTO]
                mask_vencido = venc_aberto < hoje_analise
                valor_vencido_a = abs(df_pagar_aberto.loc[mask_vencido, COL_FIN_VALOR].sum())
                qtd_vencido_a = int(mask_vencido.sum())
                # A vencer dentro do prazo interno de 10 dias
                mask_prazo_interno = (
                    (venc_aberto >= hoje_analise)
                    & (venc_aberto <= hoje_analise + pd.Timedelta(days=PRAZO_INTERNO_DIAS))
                )
                valor_prazo_interno_a = abs(df_pagar_aberto.loc[mask_prazo_interno, COL_FIN_VALOR].sum())
                qtd_prazo_interno_a = int(mask_prazo_interno.sum())
            else:
                valor_vencido_a = qtd_vencido_a = 0
                valor_prazo_interno_a = qtd_prazo_interno_a = 0

            sub_prazos, sub_aberto, sub_estrutura = st.tabs(
                ["⏱️ Prazos e Cobertura", "📌 Contas a Pagar", "🧾 Estrutura e Composição"]
            )

            with sub_prazos:
                st.markdown('<div class="section-title">⏱️ Prazos e Cobertura de Caixa</div>', unsafe_allow_html=True)
                st.markdown(
                    render_kpi_row([
                        dict(label="PRAZO MÉDIO DE PAGAMENTO",
                             value=(f"{prazo_medio_pagto_a:+.1f} dias" if prazo_medio_pagto_a is not None else "—"),
                             value_color=(COLORS["positive"] if (prazo_medio_pagto_a or 0) <= 0 else COLORS["negative"]),
                             subtext=(f"Entre os {len(saidas_prazo_a)} títulos já pagos"
                                      if prazo_medio_pagto_a is not None else "Nenhum título pago no recorte"),
                             icon="📤"),
                        dict(label="PRAZO MÉDIO DE RECEBIMENTO",
                             value=(f"{prazo_medio_receb_a:+.1f} dias" if prazo_medio_receb_a is not None else "—"),
                             value_color=(COLORS["positive"] if (prazo_medio_receb_a or 0) <= 0 else COLORS["warning"]),
                             subtext=(f"Entre os {len(entradas_prazo_a)} já recebidos"
                                      if prazo_medio_receb_a is not None else "Nenhum recebimento liquidado"),
                             icon="📥"),
                        dict(label="PAGAMENTOS EM DIA",
                             value=(f"{pct_pago_em_dia_a:.0f}%" if pct_pago_em_dia_a is not None else "—"),
                             value_color=(COLORS["positive"] if (pct_pago_em_dia_a or 0) >= 90 else COLORS["warning"]),
                             subtext="Pagos até o vencimento (dos já pagos)", icon="✅"),
                        dict(label="DIAS DE CAIXA",
                             value=(f"{dias_de_caixa_a:.1f} dias" if dias_de_caixa_a is not None else "—"),
                             value_color=(COLORS["positive"] if (dias_de_caixa_a or 0) >= 30 else COLORS["negative"]),
                             subtext=f"Ao ritmo de {formata_m(saida_media_diaria_a)}/dia de saídas", icon="🛡️"),
                        dict(label="MÉDIA DIÁRIA (ENT. / SAÍ.)",
                             value=f"{formata_m(entrada_media_diaria_a)} / {formata_m(saida_media_diaria_a)}",
                             value_color=COLORS["text"], subtext=f"Em {dias_com_mov_a} dias com movimento", icon="📊"),
                        dict(label="CONCENTRAÇÃO DAS SAÍDAS", value=f"{pct_top3_a:.0f}%",
                             value_color=(COLORS["negative"] if pct_top3_a >= 50 else COLORS["text"]),
                             subtext=f"Nos 3 maiores dias · pico {rotulo_pico_a} ({formata_m(valor_pico_a)})", icon="🎯"),
                    ]),
                    unsafe_allow_html=True,
                )
                st.caption(
                    "**Prazo médio** considera só o que já foi liquidado, comparando a data de pagamento com a "
                    "de vencimento: negativo significa antecipar, positivo significa atrasar. **Dias de caixa** "
                    "estima por quantos dias o saldo atual cobriria as saídas no ritmo médio do período. "
                    "**Concentração** mostra o quanto do desembolso está espremido em poucos dias."
                )
                st.markdown("<br>", unsafe_allow_html=True)


            with sub_aberto:
                # ---- Contas a pagar em aberto ----
                st.markdown(
                    '<div class="section-title">📌 Contas a Pagar em Aberto '
                    f'<span style="font-weight:400;font-size:12px;color:{COLORS["text_muted"]};">'
                    f'(posição em {hoje_analise:%d/%m/%Y} · prazo interno de {PRAZO_INTERNO_DIAS} dias)</span></div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    render_kpi_row([
                        dict(label="TOTAL A PAGAR NO RECORTE", value=formata_brl(total_pagar_a),
                             value_color=COLORS["text"],
                             subtext=f"{len(df_pagar_a)} títulos no período filtrado", icon="🧾"),
                        dict(label="JÁ QUITADO", value=formata_brl(valor_quitado_a),
                             value_color=COLORS["positive"],
                             subtext=f"{pct_quitado_a:.0f}% do total · {len(df_pagar_quitado)} títulos", icon="✅"),
                        dict(label="AINDA EM ABERTO", value=formata_brl(valor_aberto_a),
                             value_color=COLORS["warning"],
                             subtext=f"{len(df_pagar_aberto)} títulos sem liquidação", icon="⏳"),
                        dict(label="VENCIDO E NÃO PAGO", value=formata_brl(valor_vencido_a),
                             value_color=(COLORS["negative"] if valor_vencido_a else COLORS["positive"]),
                             subtext=f"{qtd_vencido_a} títulos com vencimento passado", icon="🚨"),
                        dict(label=f"A VENCER EM ATÉ {PRAZO_INTERNO_DIAS} DIAS",
                             value=formata_brl(valor_prazo_interno_a),
                             value_color=COLORS["text"],
                             subtext=f"{qtd_prazo_interno_a} títulos dentro do prazo interno", icon="📆"),
                    ]),
                    unsafe_allow_html=True,
                )
                st.caption(
                    "Título **sem data de liquidação** é título ainda não pago — a previsão de desembolso é o "
                    "próprio vencimento. **Vencido e não pago** é o que passou da data e continua em aberto; "
                    f"**a vencer em até {PRAZO_INTERNO_DIAS} dias** é o desembolso que entra no prazo interno "
                    "combinado com fornecedores e já deve estar reservado no caixa."
                )
                st.markdown("<br>", unsafe_allow_html=True)


            with sub_estrutura:
                # ---- Estrutura e composição ----
                qtd_lanc_a = len(df_a)
                qtd_entradas_a = int((df_a["Tipo Movimento"] == "entrada").sum())
                qtd_saidas_a = int((df_a["Tipo Movimento"] == "saida").sum())
                ticket_entrada_a = entradas_a / qtd_entradas_a if qtd_entradas_a else 0
                ticket_saida_a = abs(saidas_a) / qtd_saidas_a if qtd_saidas_a else 0
                cobertura_a = (entradas_a / abs(saidas_a) * 100) if saidas_a else 0
                liquidez_imediata_a = (saldo_a / abs(saidas_a) * 100) if saidas_a else 0

                projetado_a = df_a.loc[
                    (df_a["Tipo Movimento"] == "entrada")
                    & df_a[COL_FIN_MOVIMENTO].astype(str).str.contains("projetad", case=False, na=False),
                    COL_FIN_VALOR,
                ].sum()
                realizado_a = entradas_a - projetado_a
                pct_realizado_a = (realizado_a / entradas_a * 100) if entradas_a else 0

                st.markdown('<div class="section-title">🧾 Estrutura das Movimentações</div>', unsafe_allow_html=True)
                st.markdown(
                    render_kpi_row([
                        dict(label="LANÇAMENTOS NO PERÍODO", value=f"{qtd_lanc_a:,}".replace(",", "."),
                             value_color=COLORS["text"],
                             subtext=f"{qtd_entradas_a} entradas · {qtd_saidas_a} saídas", icon="🧮"),
                        dict(label="TICKET MÉDIO DE ENTRADA", value=formata_brl(ticket_entrada_a),
                             value_color=COLORS["positive"], subtext="Valor médio por recebimento", icon="📥"),
                        dict(label="TICKET MÉDIO DE SAÍDA", value=formata_brl(ticket_saida_a),
                             value_color=COLORS["negative"], subtext="Valor médio por pagamento", icon="📤"),
                        dict(label="COBERTURA ENTRADAS/SAÍDAS", value=f"{cobertura_a:.0f}%",
                             value_color=(COLORS["positive"] if cobertura_a >= 100 else COLORS["negative"]),
                             subtext="Acima de 100% = entra mais do que sai", icon="⚖️"),
                        dict(label="LIQUIDEZ IMEDIATA", value=f"{liquidez_imediata_a:.0f}%",
                             value_color=(COLORS["positive"] if liquidez_imediata_a >= 100 else COLORS["warning"]),
                             subtext="Saldo em caixa ÷ contas a pagar", icon="💧"),
                        dict(label="RECEBÍVEIS JÁ REALIZADOS", value=f"{pct_realizado_a:.0f}%",
                             value_color=(COLORS["positive"] if pct_realizado_a >= 50 else COLORS["warning"]),
                             subtext=f"Projetado: {formata_m(projetado_a)}", icon="⏳"),
                    ]),
                    unsafe_allow_html=True,
                )
                st.markdown("<br>", unsafe_allow_html=True)

                # ---- Gráficos de composição ----
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    st.markdown('<div class="section-title">🏢 Saídas por Canal</div>', unsafe_allow_html=True)
                    saidas_canal_a = (
                        df_a[df_a["Tipo Movimento"] == "saida"].groupby(COL_FIN_CANAL, observed=True)[COL_FIN_VALOR].sum().abs().sort_values()
                    )
                    if not saidas_canal_a.empty and saidas_canal_a.sum():
                        fig_canal = go.Figure(data=[go.Bar(
                            x=list(saidas_canal_a.values), y=list(saidas_canal_a.index), orientation="h",
                            marker=dict(color="rgba(247,110,110,0.35)", line=dict(color=COLORS["negative"], width=1.5)),
                            text=[formata_m(v) for v in saidas_canal_a.values],
                            textposition="outside", textfont=dict(size=10, color=COLORS["text_muted"]),
                            hovertemplate="%{y}: R$ %{x:,.2f}<extra></extra>",
                        )])
                        estilo_grafico(
                            fig_canal, height=300,
                            xaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True,
                                       range=[0, saidas_canal_a.max() * 1.28]),
                            yaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickfont=dict(size=10)),
                            margin=dict(l=10, r=20, t=20, b=30),
                        )
                        st.plotly_chart(fig_canal, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                    else:
                        st.caption("Sem saídas no recorte selecionado.")

                with col_g2:
                    st.markdown('<div class="section-title">💳 Movimentação por Modalidade</div>', unsafe_allow_html=True)
                    modal_a = df_a.groupby(COL_FIN_MODALIDADE, observed=True)[COL_FIN_VALOR].apply(lambda s: s.abs().sum()).sort_values(ascending=False)
                    modal_a = modal_a[modal_a > 0]
                    if not modal_a.empty:
                        TOP_MODAL = 6
                        if len(modal_a) > TOP_MODAL:
                            outros_modal = modal_a.iloc[TOP_MODAL:].sum()
                            modal_a = pd.concat([modal_a.iloc[:TOP_MODAL], pd.Series({"Outros": outros_modal})])
                        fig_modal = go.Figure(data=[go.Pie(
                            labels=list(modal_a.index), values=list(modal_a.values), hole=0.55,
                            marker=dict(colors=[COLORS["primary"], COLORS["positive"], COLORS["warning"],
                                                 COLORS["negative"], COLORS["secondary"], COLORS["muted_line"], "#6B7280"]),
                            textinfo="percent", texttemplate="%{percent:.1%}",
                            textfont=dict(color=COLORS["text"], size=11),
                            hovertemplate="%{label}: R$ %{value:,.2f}<extra></extra>",
                        )])
                        fig_modal.add_annotation(
                            text=f"<b>{formata_m(modal_a.sum())}</b><br><span style='font-size:10px;color:{COLORS['text_muted']}'>Volume total</span>",
                            showarrow=False, font=dict(color=COLORS["text"], size=13, family=FONT_STACK),
                        )
                        estilo_grafico(
                            fig_modal, height=300,
                            legend=dict(orientation="h", yanchor="top", y=-0.05, xanchor="center", x=0.5, font=dict(size=9)),
                            margin=dict(l=10, r=10, t=20, b=10),
                        )
                        st.plotly_chart(fig_modal, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                    else:
                        st.caption("Sem movimentação no recorte selecionado.")

                # ---- Maiores despesas por grupo ----
                if COL_FIN_GRUPO_DESPESA in df_a.columns:
                    # Mantém o sinal negativo: são saídas de caixa. Com abs() o
                    # valor virava positivo e a coloração pintava despesa de
                    # verde, como se fosse entrada.
                    grupo_a = (
                        df_a[df_a["Tipo Movimento"] == "saida"]
                        .groupby(COL_FIN_GRUPO_DESPESA, observed=True)[COL_FIN_VALOR].sum().sort_values()
                    )
                    grupo_a = grupo_a[grupo_a < 0].head(10)
                    if not grupo_a.empty:
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown('<div class="section-title">📉 Maiores Grupos de Despesa</div>', unsafe_allow_html=True)
                        total_saidas_grupo = abs(grupo_a.sum())
                        df_grupo_a = pd.DataFrame({
                            "Grupo de Despesa": grupo_a.index.astype(str),
                            "Valor (R$)": grupo_a.values,
                            "% do total": [abs(v) / total_saidas_grupo * 100 for v in grupo_a.values],
                        })
                        st.dataframe(
                            df_grupo_a.style.format({
                                "Valor (R$)": formata_brl,
                                "% do total": lambda v: f"{v:.1f}%".replace(".", ","),
                            }).map(cor_valor, subset=["Valor (R$)"]),
                            use_container_width=True, hide_index=True,
                        )
                        st.caption("Os 10 grupos que mais consumiram caixa no recorte selecionado.")

    st.stop()

# Se chegou até aqui, painel_escolhido == "controladoria" -> segue para o
# painel normal, abaixo.



# ============================================================================
# 5. BARRA LATERAL — FILTROS
# ============================================================================
st.sidebar.markdown(
    f"""
    <div class="sidebar-brand">
        <img class="brand-logo" src="data:image/jpeg;base64,{LOGO_BEEA_B64}" alt="Grupo Beea" />
        <div>
            <span class="title">Controladoria B&A</span>
            <span class="subtitle">Painel Financeiro</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.sidebar.markdown("**🔎 Escopo da Análise**")
if eh_admin:
    with st.sidebar.expander("🔧 Abas detectadas nas planilhas"):
        st.caption(
            "Se uma aba nova (ex.: um consolidado que você acabou de criar) não "
            "aparecer nos filtros, confira aqui o nome exato como o painel está "
            "lendo -- pode ser uma pequena diferença de digitação/espaço em "
            "relação ao nome esperado."
        )
        st.write(list(abas_disponiveis))

# ---- Visão por Departamento -- troca o painel inteiro pra focar só nas
# linhas da DRE relevantes a um departamento específico (reaproveita as
# linhas já definidas nos modelos de relatório -- MKT, Compras,
# Suprimentos, RH -- como fonte da verdade de "quais linhas pertencem a
# esse setor"). Disponível para todos os perfis de acesso.
#
# Cada e-mail abre o painel já no departamento correspondente ao seu acesso
# (mas a pessoa ainda pode trocar manualmente pelo seletor, se quiser ver
# outra visão). Quem não está no mapa abre na Controladoria, como sempre.
MAPA_EMAIL_DEPARTAMENTO = {
    "coordenador.compras@grupobeea.com.br": "🛒 Relatório de Custos - Compras",
    "coordenador.marketing@grupobeea.com.br": "📣 Relatório de Custos - MKT",
    "gerente.logistica@grupobeea.com.br": "🚚 Relatório de Custos - Suprimentos",
    "pessoas.cultura@grupobeea.com.br": "👥 Relatório de Custos - RH",
}

departamento_ativo = None
linhas_departamento_ativo = []
st.sidebar.markdown("---")
st.sidebar.markdown("**🏢 Visão por Departamento**")
st.sidebar.caption("Filtra o painel inteiro (visão, tabelas e relatório) para o departamento escolhido.")
opcoes_departamento = ["Controladoria"] + list(MODELOS_RELATORIO.keys())
departamento_padrao_usuario = MAPA_EMAIL_DEPARTAMENTO.get(usuario_atual["email"].strip().lower())
indice_departamento_padrao = (
    opcoes_departamento.index(departamento_padrao_usuario)
    if departamento_padrao_usuario in opcoes_departamento
    else 0
)
departamento_sel = st.sidebar.selectbox(
    "Ver painel como:", opcoes_departamento, key="departamento_sel",
    index=indice_departamento_padrao,
    format_func=_nome_departamento_curto,
)
if departamento_sel != "Controladoria":
    departamento_ativo = departamento_sel
    linhas_departamento_ativo = MODELOS_RELATORIO[departamento_sel]["linhas_dre"]

modo_visao = st.sidebar.radio(
    "Modo de Visão:",
    ["Visão Consolidada", "Selecionar Unidade"],
)

abas_consolidadas_permitidas = [
    "DRE CONSOLIDADO",
    "ABPR CONSOLIDADO",
    "VD CONSOLIDADO",
    "LJ CONSOLIDADO",
    "ABPR + VD",
    "LJ - G&A",
    "CONSOLIDADO - G&A",
]

_consolidadas_normalizadas = {_normalizar_nome_aba(n) for n in abas_consolidadas_permitidas}
opcoes_consolidadas = [a for a in abas_disponiveis if _normalizar_nome_aba(a) in _consolidadas_normalizadas]
if not opcoes_consolidadas:
    opcoes_consolidadas = abas_disponiveis

opcoes_unidades = [a for a in abas_disponiveis if _normalizar_nome_aba(a) not in _consolidadas_normalizadas]
if not opcoes_unidades:
    opcoes_unidades = abas_disponiveis

if modo_visao == "Visão Consolidada":
    visao_sel = st.sidebar.selectbox("Selecione a Visão:", opcoes_consolidadas)
    abas_para_carregar = [visao_sel]
    label_visao = visao_sel
else:
    lojas_sel = st.sidebar.multiselect(
        "Selecione as Lojas/Unidades:",
        options=opcoes_unidades,
        default=opcoes_unidades[:3] if len(opcoes_unidades) >= 3 else opcoes_unidades,
    )
    if not lojas_sel:
        st.warning("Selecione ao menos uma unidade.")
        st.stop()
    abas_para_carregar = lojas_sel
    label_visao = f"Soma de {len(lojas_sel)} Unidades"

with st.spinner("Carregando dados das abas selecionadas..."):
    list_df_orc, list_df_real = carregar_dados_abas(path_orc, path_real, abas_para_carregar)

if not list_df_real or not list_df_orc:
    st.error("Erro ao ler dados das abas.")
    st.stop()

# 5.1 FILTRO DE PERÍODO / MESES
st.sidebar.markdown("---")
st.sidebar.markdown("**📅 Período**")

nomes_meses = [
    "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
    "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
]
meses_cols = [
    "01/2026", "02/2026", "03/2026", "04/2026", "05/2026", "06/2026",
    "07/2026", "08/2026", "09/2026", "10/2026", "11/2026", "12/2026",
]

df_ref = list_df_real[0]
colunas_validas = [m for m in meses_cols if m in df_ref.columns]
m_map = {
    m_nome: m_col
    for m_nome, m_col in zip(nomes_meses, meses_cols)
    if m_col in colunas_validas
}

# ---- Resolve as linhas do Departamento ativo (se algum estiver selecionado)
# contra as linhas REAIS da DRE atual -- mesma lógica de correspondência
# usada na aba de Emitir Relatório (ver _resolver_termo_departamento).
col_nome_dre_dept = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
linhas_dre_todas_painel = list(df_ref[col_nome_dre_dept].dropna().astype(str).unique()) if not df_ref.empty else []
linhas_departamento_resolvidas = []
linhas_departamento_raiz = []
linhas_departamento_informativas = []
if departamento_ativo:
    for termo in linhas_departamento_ativo:
        linhas_departamento_resolvidas.extend(_resolver_termo_departamento(termo, linhas_dre_todas_painel))
    linhas_departamento_resolvidas = list(dict.fromkeys(linhas_departamento_resolvidas))
    # "Raiz" = só as linhas sem outro membro do mesmo conjunto como
    # ancestral -- é essa lista que deve ser usada em qualquer SOMA/TOTAL
    # (KPIs, gráficos), pra não contar o mesmo custo duas vezes quando o
    # modelo lista uma linha de grupo e as sublinhas dela juntas. A tabela
    # de detalhe (que lista linha por linha, sem somar) continua usando
    # linhas_departamento_resolvidas normalmente.
    linhas_departamento_raiz = _linhas_raiz_do_conjunto(linhas_departamento_resolvidas)
    # Linhas "informativas" (ex.: no MKT, a gestão GB da indústria) -- não
    # são geridas pelo departamento, mas precisam aparecer no relatório e
    # em algum lugar do painel só pra conhecimento dos valores.
    for termo in MODELOS_RELATORIO.get(departamento_ativo, {}).get("linhas_informativas", []):
        linhas_departamento_informativas.extend(_resolver_termo_departamento(termo, linhas_dre_todas_painel))
    linhas_departamento_informativas = list(dict.fromkeys(linhas_departamento_informativas))

tipo_periodo = st.sidebar.radio(
    "Modo do Período:",
    ["Mês Selecionado", "Múltiplos Meses", "ANO COMPLETO (2026)"],
)

if tipo_periodo == "ANO COMPLETO (2026)":
    cols_kpi = list(m_map.values())
    cols_graficos = list(m_map.values())
    label_periodo_kpi = "ANO COMPLETO (2026)"
    label_periodo_graf = "ANO COMPLETO (2026)"
elif tipo_periodo == "Mês Selecionado":
    if not m_map:
        # Nenhum mês com dados válidos para o escopo selecionado (ex.: uma
        # loja/aba sem nenhuma coluna de mês reconhecida) -- evita quebrar o
        # app tentando montar um selectbox vazio.
        st.sidebar.warning("Nenhum mês com dados disponível para este escopo.")
        cols_kpi = []
        cols_graficos = []
        label_periodo_kpi = "Sem dados no escopo selecionado"
        label_periodo_graf = "Sem dados no escopo selecionado"
    else:
        # Abre por padrão no mês atual (real), não num índice fixo.
        idx_mes_real = datetime.now(FUSO_BR).month - 1
        idx_default = min(max(idx_mes_real, 0), len(m_map) - 1)
        mes_ref = st.sidebar.selectbox("Mês Desejado:", list(m_map.keys()), index=idx_default)
        if mes_ref not in m_map:
            # Segurança extra: se por qualquer motivo (ex.: opções do
            # selectbox mudaram entre execuções) o valor devolvido não
            # estiver mais no mapa atual, cai para o padrão em vez de
            # quebrar com ValueError.
            mes_ref = list(m_map.keys())[idx_default]
        idx = list(m_map.keys()).index(mes_ref)

        cols_kpi = list(m_map.values())[: idx + 1]
        cols_graficos = [m_map[mes_ref]]

        label_periodo_kpi = f"Acumulado YTD até {mes_ref}"
        label_periodo_graf = f"Mês de {mes_ref}"
else:
    meses_mult = st.sidebar.multiselect(
        "Selecione os Meses:",
        list(m_map.keys()),
        default=list(m_map.keys())[: min(7, len(m_map))],
    )
    cols_kpi = [m_map[m] for m in meses_mult if m in m_map]
    cols_graficos = cols_kpi
    label_periodo_kpi = "Meses Selecionados"
    label_periodo_graf = "Meses Selecionados"

st.sidebar.markdown("---")
if st.sidebar.button("🔄 Atualizar Dados", use_container_width=True):
    st.cache_data.clear()
    st.cache_resource.clear()
    st.rerun()

st.sidebar.caption(f"Última atualização: {datetime.now(FUSO_BR).strftime('%d/%m/%Y às %H:%M')}")

st.sidebar.markdown("---")
st.sidebar.markdown("**🖥️ Painel para TV**")
st.sidebar.caption("Abre uma visão executiva (somente leitura, atualização automática) para deixar exibida numa tela/TV da empresa.")
st.sidebar.markdown(
    f"""
    <a href="?modo=tv" target="_blank" style="text-decoration:none;">
        <div style="
            background:{COLORS['primary_soft']}; color:{COLORS['primary']}; border:1px solid {COLORS['primary']};
            border-radius:8px; padding:8px 12px; text-align:center; font-weight:600; font-size:13.5px;
            margin-bottom: 6px;">
            📺 Abrir Painel de TV
        </div>
    </a>
    """,
    unsafe_allow_html=True,
)

st.sidebar.markdown("---")
if st.sidebar.button("🔀 Trocar Painel", use_container_width=True):
    st.session_state["painel_escolhido"] = None
    st.rerun()

st.sidebar.markdown("---")
perfil_label = "Administrador" if eh_admin else "Visualização"
st.sidebar.caption(f"👤 {usuario_atual['email']}  ·  Perfil: **{perfil_label}**")
if st.sidebar.button("🚪 Sair", use_container_width=True):
    st.session_state["usuario_logado"] = None
    st.session_state["painel_escolhido"] = None
    _esquecer_credenciais_no_navegador()
    st.rerun()


# ============================================================================
# 6. FUNÇÕES DE SUPORTE (cálculo e formatação) -- ver seção 2 para as funções
# get_valor_consolidado_multi / formata_brl / formata_m / eh_grupo_sintetico /
# cor_valor, que foram movidas para perto do topo do arquivo (logo após
# render_kpi_row) para ficarem disponíveis também para o Painel de TV,
# que é montado antes desta seção.
# ============================================================================

# ============================================================================
# 6.1 GERAÇÃO DE RELATÓRIO EXCEL (formatado, para a aba de Emissão)
# ============================================================================
_THIN = Side(style="thin", color="FFD9DDE3")

EXCEL_STYLE = {
    "fill_title": PatternFill(fill_type="solid", start_color="FF0B0E14", end_color="FF0B0E14"),
    "fill_header": PatternFill(fill_type="solid", start_color="FF1A1F2E", end_color="FF1A1F2E"),
    "fill_zebra": PatternFill(fill_type="solid", start_color="FFF3F5F9", end_color="FFF3F5F9"),
    "fill_group": PatternFill(fill_type="solid", start_color="FFEFF3FA", end_color="FFEFF3FA"),
    "fill_total": PatternFill(fill_type="solid", start_color="FFDCE8FF", end_color="FFDCE8FF"),
    "font_title": Font(color="FFFFFFFF", bold=True, size=13, name="Calibri"),
    "font_header": Font(color="FFFFFFFF", bold=True, size=11, name="Calibri"),
    "font_normal": Font(color="FF1F2937", size=10.5, name="Calibri"),
    "font_bold": Font(color="FF1F2937", bold=True, size=10.5, name="Calibri"),
    "font_pos": Font(color="FF1B8A5A", size=10.5, name="Calibri"),
    "font_neg": Font(color="FFC0392B", size=10.5, name="Calibri"),
    "font_caption": Font(color="FF6B7280", italic=True, size=9.5, name="Calibri"),
    "border": Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN),
}


def _fonte_por_valor(valor):
    return EXCEL_STYLE["font_pos"] if valor >= 0 else EXCEL_STYLE["font_neg"]


def _escrever_titulo(ws, texto, linha, n_colunas):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_colunas)
    cell = ws.cell(row=linha, column=1, value=texto)
    cell.font = EXCEL_STYLE["font_title"]
    cell.fill = EXCEL_STYLE["fill_title"]
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha].height = 26


def _escrever_legenda(ws, texto, linha, n_colunas):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_colunas)
    cell = ws.cell(row=linha, column=1, value=texto)
    cell.font = EXCEL_STYLE["font_caption"]
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha].height = 16


def _escrever_cabecalho_matriz(ws, linha, primeiro_rotulo, mapa_meses, fill_header=None):
    """Escreve um cabeçalho no formato: [primeiro_rotulo, Jan, Fev, ..., Total Ano]."""
    headers = [primeiro_rotulo] + [m.capitalize() for m in mapa_meses.keys()] + ["Total Ano"]
    fill = fill_header or EXCEL_STYLE["fill_zebra"]
    for col, texto in enumerate(headers, start=1):
        cell = ws.cell(row=linha, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_bold"]
        cell.fill = fill
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
    return len(headers)


def _escrever_linha_matriz(ws, linha, rotulo, valores_mes, total, negrito=False, fill=None, colorir_por_sinal=False):
    """Escreve uma linha: [rotulo, valor_mes_1, valor_mes_2, ..., total]."""
    valores = [rotulo] + list(valores_mes) + [total]
    for col, val in enumerate(valores, start=1):
        cell = ws.cell(row=linha, column=col, value=val)
        cell.border = EXCEL_STYLE["border"]
        if fill:
            cell.fill = fill
        if col == 1:
            cell.font = EXCEL_STYLE["font_bold"] if negrito else EXCEL_STYLE["font_normal"]
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            valor_numerico = val if isinstance(val, (int, float)) else 0
            if colorir_por_sinal:
                cell.font = _fonte_por_valor(valor_numerico)
            else:
                cell.font = EXCEL_STYLE["font_bold"] if negrito else EXCEL_STYLE["font_normal"]
            cell.number_format = '"R$" #,##0.00'
            cell.alignment = Alignment(horizontal="right")


def _escrever_linha_flat(ws, linha, campos, valores_mes, total, negrito=False, fill=None, colorir_por_sinal=False):
    """Escreve uma linha em formato tabular "achatado": [campo_1, campo_2,
    ..., valor_mes_1, ..., total]. `campos` é uma lista de rótulos (ex.:
    [loja, conta, tipo]) -- diferente de _escrever_linha_matriz, que só
    aceita um único rótulo. Usado nas abas que precisam de AutoFilter do
    Excel (uma "loja" por coluna própria, não misturada dentro do texto)."""
    n_campos = len(campos)
    valores = list(campos) + list(valores_mes) + [total]
    for col, val in enumerate(valores, start=1):
        cell = ws.cell(row=linha, column=col, value=val)
        cell.border = EXCEL_STYLE["border"]
        if fill:
            cell.fill = fill
        if col <= n_campos:
            cell.font = EXCEL_STYLE["font_bold"] if negrito else EXCEL_STYLE["font_normal"]
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            valor_numerico = val if isinstance(val, (int, float)) else 0
            if colorir_por_sinal:
                cell.font = _fonte_por_valor(valor_numerico)
            else:
                cell.font = EXCEL_STYLE["font_bold"] if negrito else EXCEL_STYLE["font_normal"]
            cell.number_format = '"R$" #,##0.00'
            cell.alignment = Alignment(horizontal="right")


_CARACTERES_INVALIDOS_ABA = set('[]:*?/\\')


def _nome_aba_seguro(texto, usados, max_len=31):
    """Gera um nome de aba do Excel válido (até 31 caracteres, sem os
    caracteres proibidos pelo Excel) e garante que não colida com nenhum
    outro nome de aba já usado no mesmo workbook (guardados em `usados`,
    em minúsculas)."""
    limpo = "".join(c for c in str(texto) if c not in _CARACTERES_INVALIDOS_ABA).strip()
    limpo = limpo[:max_len] if limpo else "Aba"
    base = limpo
    sufixo = 1
    while limpo.lower() in usados:
        sufixo += 1
        marcador = f"_{sufixo}"
        corte = max_len - len(marcador)
        limpo = f"{base[:corte]}{marcador}"
    usados.add(limpo.lower())
    return limpo


def _ref_aba(nome):
    """Formata o nome de uma aba para uso em fórmulas do Excel (entre aspas
    simples, com aspas simples internas escapadas)."""
    return "'" + str(nome).replace("'", "''") + "'"


def _filtrar_diario_completo(df_diario, loja, conta, plano):
    """Filtra o DIÁRIO por loja (Centro de Custos) + Linha DRE (conta) +
    Plano de Contas exato -- usado para montar a aba de detalhamento dos
    lançamentos ao "clicar" numa linha da aba Plano de Contas."""
    df_filtrado = _filtrar_diario_por_loja_e_conta(df_diario, loja, conta)
    if df_filtrado is None or df_filtrado.empty:
        return df_filtrado
    plano_norm = _normalizar_texto(plano)
    planos_norm = df_filtrado["Plano de Contas"].map(_normalizar_texto)
    mask = planos_norm == plano_norm
    if not mask.any():
        mask = planos_norm.apply(lambda x: (plano_norm in x) or (x in plano_norm) if x else False)
    return df_filtrado[mask]


def _criar_aba_lancamentos(wb, nomes_usados, titulo_bloco, df_lancamentos):
    """Cria uma aba nova com o detalhamento dos lançamentos do DIÁRIO que
    compõem um valor (usada para o "clique" que abre o detalhamento a partir
    da aba Detalhe Mensal ou Plano de Contas). É basicamente uma cópia da
    própria aba DIÁRIO, só que filtrada para aqueles lançamentos específicos
    -- por isso mantém as mesmas colunas dela (Competência, Centro de Custos,
    Linha DRE, Plano de Contas, Valor Bruto). Retorna o nome da aba criada."""
    nome_aba = _nome_aba_seguro(f"Lc_{titulo_bloco}", nomes_usados)
    ws = wb.create_sheet(nome_aba)
    ws.sheet_properties.tabColor = "FF4C8DFF"

    n_col = 5
    _escrever_titulo(ws, f"Lançamentos — {titulo_bloco}"[:250], 1, n_col)
    _escrever_legenda(ws, f"{len(df_lancamentos)} lançamento(s) encontrados na aba DIÁRIO (Realizado).", 2, n_col)

    linha = 4
    headers = ["Competência", "Centro de Custos", "Linha DRE", "Plano de Contas", "Valor Bruto (R$)"]
    for col, texto in enumerate(headers, start=1):
        cell = ws.cell(row=linha, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_header"]
        cell.fill = EXCEL_STYLE["fill_header"]
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col in (1, 2, 3, 4) else "center")
    linha += 1

    df_ordenado = df_lancamentos.sort_values("Competência", na_position="last")
    for i, (_, row) in enumerate(df_ordenado.iterrows()):
        data_val = row["Competência"]
        data_txt = data_val.strftime("%d/%m/%Y") if pd.notna(data_val) and hasattr(data_val, "strftime") else str(data_val)
        valores = [data_txt, row["Centro de Custos"], row["Linha DRE"], row["Plano de Contas"], row["Valor Bruto"]]
        fill = EXCEL_STYLE["fill_zebra"] if i % 2 == 1 else None
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=linha, column=col, value=val)
            cell.border = EXCEL_STYLE["border"]
            if fill:
                cell.fill = fill
            if col == 5:
                cell.font = EXCEL_STYLE["font_normal"]
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = EXCEL_STYLE["font_normal"]
                cell.alignment = Alignment(horizontal="left")
        linha += 1

    total_valor = df_ordenado["Valor Bruto"].sum()
    cell_total_lbl = ws.cell(row=linha, column=1, value="TOTAL")
    cell_total_lbl.font = EXCEL_STYLE["font_bold"]
    cell_total = ws.cell(row=linha, column=5, value=total_valor)
    cell_total.font = EXCEL_STYLE["font_bold"]
    cell_total.number_format = '"R$" #,##0.00'
    cell_total.alignment = Alignment(horizontal="right")
    for col in range(1, n_col + 1):
        ws.cell(row=linha, column=col).fill = EXCEL_STYLE["fill_total"]
        ws.cell(row=linha, column=col).border = EXCEL_STYLE["border"]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A5"
    return nome_aba


def montar_relatorio_excel(
    contas_sel,
    dfs_real,
    dfs_orc,
    mapa_meses,
    colunas_ano,
    escopo_label,
    dados_por_loja=None,
    mapa_planos_dre=None,
    df_diario=None,
    forcar_planos_contas=None,
    permitir_lancamento_manual=False,
    mapa_loja_centro_custo=None,
):
    """Gera um relatório Excel formatado com três planilhas:
    - Resumo: total do ano por conta, CONSOLIDADO (não muda com a divisão por loja).
    - Detalhe Mensal: contas da DRE nas linhas, meses nas colunas, dividido por loja.
    - Plano de Contas: composição (planos de contas) de cada linha da DRE, por loja,
      mês a mês. Fonte principal: aba DIÁRIO da planilha Realizado 2026 (df_diario).
      Se o DIÁRIO não estiver disponível para uma loja/conta, cai para o método
      antigo (Tabela_Contas + soma nas abas por loja).

    forcar_planos_contas: lista de nomes de Plano de Contas que devem ser
    incluídos na aba "Plano de Contas" à parte, puxados direto pelo nome
    (ignorando a Linha DRE) -- usado por modelos como o de Compras, onde
    "Mercadorias" faz parte do relatório mesmo sem linha da DRE correspondente.

    permitir_lancamento_manual: quando True, uma Linha DRE sem nenhum Plano
    de Contas correspondente no DIÁRIO/Tabela_Contas aparece como
    "Lançado Manualmente", com o valor da própria linha da DRE -- usado pelo
    modelo de RH, que tem várias linhas lançadas direto, sem plano de contas.

    mapa_loja_centro_custo: {nome_da_loja: centro_de_custo}, vindo da
    Tabela_Lojas (DE_PARA). A DIÁRIO identifica a loja pelo Centro de Custos,
    que normalmente é diferente do nome da aba da loja -- por isso, sempre
    que formos filtrar a DIÁRIO por loja, resolvemos primeiro o nome da loja
    para o Centro de Custos correspondente por esse mapa (usando o próprio
    nome da loja como alternativa, caso não haja correspondência no DE_PARA).
    """
    dados_por_loja = dados_por_loja or {}
    mapa_planos_dre = mapa_planos_dre or {}
    forcar_planos_contas = forcar_planos_contas or []
    mapa_loja_centro_custo = mapa_loja_centro_custo or {}
    diario_disponivel = df_diario is not None and not df_diario.empty

    def _cc(loja):
        """Resolve os candidatos de Centro de Custos para uma loja: o valor
        mapeado pela Tabela_Lojas (DE_PARA) E o próprio nome da loja/aba,
        nessa ordem. A busca no DIÁRIO tenta os dois, então funciona mesmo
        que o DE_PARA esteja incompleto/não encontrado para alguma loja, ou
        caso o Centro de Custos já seja igual ao nome da loja em alguns
        casos."""
        centro_mapeado = mapa_loja_centro_custo.get(loja)
        candidatos = [c for c in (centro_mapeado, loja) if c]
        return candidatos or [loja]
    wb = Workbook()
    gerado_em = f"{escopo_label} · Gerado em {datetime.now(FUSO_BR).strftime('%d/%m/%Y às %H:%M')}"

    # ---------------- ABA "RESUMO" ----------------
    ws1 = wb.active
    ws1.title = "Resumo"

    _escrever_titulo(ws1, "Relatório de DRE — Orçado vs. Realizado (Resumo Anual)", 1, 5)
    _escrever_legenda(ws1, gerado_em, 2, 5)

    linha_header = 4
    headers = ["Conta / Linha DRE", "Realizado (R$)", "Orçado (R$)", "Desvio (R$)", "Desvio (%)"]
    for col, texto in enumerate(headers, start=1):
        cell = ws1.cell(row=linha_header, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_header"]
        cell.fill = EXCEL_STYLE["fill_header"]
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
    ws1.row_dimensions[linha_header].height = 20
    ws1.freeze_panes = f"A{linha_header + 1}"

    linha = linha_header + 1
    for i, conta in enumerate(contas_sel):
        v_real = get_valor_consolidado_multi(dfs_real, conta, colunas_ano)
        v_orc = get_valor_consolidado_multi(dfs_orc, conta, colunas_ano)
        desvio = v_real - v_orc
        desvio_pct = (desvio / abs(v_orc) * 100) if v_orc != 0 else 0.0

        fill = EXCEL_STYLE["fill_zebra"] if i % 2 == 1 else None
        valores = [conta, v_real, v_orc, desvio, desvio_pct]
        for col, val in enumerate(valores, start=1):
            cell = ws1.cell(row=linha, column=col, value=val)
            cell.border = EXCEL_STYLE["border"]
            if fill:
                cell.fill = fill
            if col == 1:
                cell.font = EXCEL_STYLE["font_normal"]
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col in (2, 3):
                cell.font = EXCEL_STYLE["font_normal"]
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 4:
                cell.font = _fonte_por_valor(desvio)
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = _fonte_por_valor(desvio)
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="right")
        linha += 1

    for col, largura in zip(range(1, 6), [46, 18, 18, 18, 14]):
        ws1.column_dimensions[get_column_letter(col)].width = largura

    # ========================================================================
    # Dados auxiliares: lojas individuais x visões consolidadas, e o filtro
    # único da DIÁRIO que alimenta tanto a aba "Lançamentos" quanto a aba
    # "Plano de Contas".
    # ========================================================================
    lojas_ordenadas = sorted(dados_por_loja.keys()) if dados_por_loja else []
    meses_chaves = list(mapa_meses.keys())
    n_meses = len(meses_chaves)
    n_col_matriz = 1 + n_meses + 1  # rótulo + meses + Total Ano

    LOJAS_CONSOLIDADAS = {
        "DRE CONSOLIDADO", "ABPR CONSOLIDADO", "VD CONSOLIDADO",
        "LJ CONSOLIDADO", "ABPR + VD", "LJ - G&A", "CONSOLIDADO - G&A",
    }
    _LOJAS_CONSOLIDADAS_NORM = {_normalizar_nome_aba(n) for n in LOJAS_CONSOLIDADAS}
    lojas_individuais = [l for l in lojas_ordenadas if _normalizar_nome_aba(l) not in _LOJAS_CONSOLIDADAS_NORM]

    def _lojas_do_grupo_consolidado(nome_grupo):
        """Para uma "loja" que na verdade é uma visão consolidada (ex.: "ABPR
        + VD"), devolve as lojas INDIVIDUAIS que fazem parte dela -- usado
        para somar corretamente a composição de Plano de Contas (que vem da
        DIÁRIO, que só reconhece lojas individuais, nunca uma visão
        consolidada). Grupos definidos explicitamente:
        - ABPR CONSOLIDADO = ABPR 23427 + ABPR ZNS 24527
        - VD CONSOLIDADO = VD - GUAJARA 23859 + VD - LESTE 21506 + VD - MACHAD 21691 +
          VD - MATRIZ 13967 + VD - VST ALEG 21497
        - ABPR + VD = ABPR CONSOLIDADO + VD CONSOLIDADO
        - LJ - G&A = as 13 lojas "LJ ..." (sem ESCRIT MATRIZ, sem lojas VD)
        - LJ CONSOLIDADO = ESCRIT MATRIZ 6037 + LJ - G&A (sem lojas VD)
        - CONSOLIDADO - G&A = ABPR CONSOLIDADO + VD CONSOLIDADO + LJ - G&A
        - DRE CONSOLIDADO = ABPR CONSOLIDADO + VD CONSOLIDADO + LJ CONSOLIDADO
          (na prática, todas as 21 lojas individuais)
        """
        grupo_abpr = ["ABPR 23427", "ABPR ZNS 24527"]
        grupo_vd = [
            "VD - GUAJARA 23859", "VD - LESTE 21506", "VD - MACHAD 21691",
            "VD - MATRIZ 13967", "VD - VST ALEG 21497",
        ]
        grupo_lj_ga = [
            "LJ ARAUJO 12606", "LJ ASSAI 23157", "LJ GUAJARA 23809", "LJ IG SHOP 20330",
            "LJ JATUARANA 6040", "LJ JK 12478", "LJ MACHAD 21462", "LJ MARECHAL 6039",
            "LJ NV ERA 18539", "LJ PVH1 11927", "LJ PVH2 14625", "LJ QDB 910332", "LJ SETE 6052",
        ]
        grupo_lj_consolidado = ["ESCRIT MATRIZ 6037"] + grupo_lj_ga
        grupo_dre_consolidado = grupo_abpr + grupo_vd + grupo_lj_consolidado
        grupo_consolidado_ga = grupo_abpr + grupo_vd + grupo_lj_ga

        mapa_grupos = {
            "ABPR CONSOLIDADO": grupo_abpr,
            "VD CONSOLIDADO": grupo_vd,
            "ABPR + VD": grupo_abpr + grupo_vd,
            "LJ - G&A": grupo_lj_ga,
            "LJ CONSOLIDADO": grupo_lj_consolidado,
            "CONSOLIDADO - G&A": grupo_consolidado_ga,
            "DRE CONSOLIDADO": grupo_dre_consolidado,
        }
        mapa_grupos_normalizado = {_normalizar_nome_aba(k): v for k, v in mapa_grupos.items()}
        lojas_definidas = mapa_grupos_normalizado.get(_normalizar_nome_aba(nome_grupo), [])
        # Só devolve lojas que realmente existem entre as lojas individuais carregadas.
        return [l for l in lojas_definidas if l in lojas_individuais]


    # ---- Filtra a DIÁRIO para pegar só os lançamentos relevantes deste
    # relatório: linhas cuja "Linha DRE" bate com alguma conta selecionada,
    # OU cujo "Plano de Contas" bate com um dos planos forçados (ex.:
    # "Mercadorias", que não tem Linha DRE). Essa é a ÚNICA fonte usada tanto
    # pela aba "Lançamentos" quanto para montar a composição da aba
    # "Plano de Contas" -- ou seja, os dois batem entre si por construção. ----
    if diario_disponivel:
        linhas_norm = df_diario["Linha DRE"].map(_normalizar_texto)
        mask_conta = pd.Series(False, index=df_diario.index)
        for conta in contas_sel:
            conta_norm = _normalizar_texto(conta)
            m = linhas_norm == conta_norm
            if not m.any():
                m = linhas_norm.apply(lambda x: (conta_norm in x) or (x in conta_norm) if x else False)
            mask_conta |= m

        planos_norm = df_diario["Plano de Contas"].map(_normalizar_texto)
        mask_forcado = pd.Series(False, index=df_diario.index)
        for plano in forcar_planos_contas:
            plano_norm = _normalizar_texto(plano)
            m = planos_norm == plano_norm
            if not m.any():
                m = planos_norm.apply(lambda x: (plano_norm in x) or (x in plano_norm) if x else False)
            mask_forcado |= m

        df_lanc = df_diario[mask_conta | mask_forcado].copy()
        # Resolve o nome "bonito" da loja a partir do Centro de Custos (usa o
        # mapa nos dois sentidos, montado pela Tabela_Lojas/DE_PARA). Se não
        # achar correspondência, mantém o próprio Centro de Custos como
        # rótulo (mais transparente do que esconder a linha).
        df_lanc["Loja"] = df_lanc["Centro de Custos"].astype(str).str.strip().map(
            lambda cc: mapa_loja_centro_custo.get(cc, cc)
        )
    else:
        df_lanc = pd.DataFrame(
            columns=[
                "Competência", "Centro de Custos", "Linha DRE", "Plano de Contas", "Valor Bruto", "Mês",
                "Número", "Cliente / Fornecedor", "Histórico", "Loja",
            ]
        )

    # ---- Planos de contas GLOBAIS por conta (juntando TODAS as lojas) ----
    # A aba "Plano de Contas" precisa mostrar sempre o MESMO conjunto de
    # planos de contas para uma dada linha da DRE, em todas as lojas -- com
    # valor ou não. Sem isso, uma loja que não tem lançamento pra um plano
    # específico simplesmente não mostrava aquela linha, enquanto outra loja
    # (que tem) mostrava -- deixando a estrutura inconsistente entre visões.
    planos_por_conta_global = {}
    if not df_lanc.empty:
        linhas_norm_geral = df_lanc["Linha DRE"].map(_normalizar_texto)
        for conta in contas_sel:
            conta_norm_geral = _normalizar_texto(conta)
            mask_geral = linhas_norm_geral == conta_norm_geral
            if not mask_geral.any():
                mask_geral = linhas_norm_geral.apply(
                    lambda x: (conta_norm_geral in x) or (x in conta_norm_geral) if x else False
                )
            planos_conta = sorted(df_lanc.loc[mask_geral, "Plano de Contas"].dropna().unique().tolist())
            if planos_conta:
                planos_por_conta_global[conta] = planos_conta

    # ---------------- ABA "DETALHE MENSAL" (Orçado x Realizado, por loja) ----------------
    # Formato tabular "achatado" (Loja/Conta/Tipo em colunas próprias, uma
    # linha por combinação) -- isso permite usar o AutoFilter nativo do
    # Excel na coluna "Loja" para escolher qual visão ver, sem precisar abrir
    # um arquivo gigante com tudo misturado. Clique na setinha do cabeçalho
    # da coluna "Loja" (linha 4) para filtrar.
    ws2 = wb.create_sheet("Detalhe Mensal")
    n_campos_dm = 3  # Loja, Conta, Tipo
    n_col_dm = n_campos_dm + n_meses + 1
    _escrever_titulo(ws2, "Detalhe Mensal — Orçado vs. Realizado, por Loja", 1, n_col_dm)
    _escrever_legenda(
        ws2,
        f"{gerado_em} · Use o filtro (▾) no cabeçalho da coluna \"Loja\" para escolher qual visão ver.",
        2, n_col_dm,
    )

    linha_header_dm = 4
    headers_dm = ["Loja", "Conta", "Tipo"] + [m.capitalize() for m in mapa_meses.keys()] + ["Total Ano"]
    for col, texto in enumerate(headers_dm, start=1):
        cell = ws2.cell(row=linha_header_dm, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_header"]
        cell.fill = EXCEL_STYLE["fill_header"]
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col <= n_campos_dm else "center")
    ws2.row_dimensions[linha_header_dm].height = 20
    ws2.freeze_panes = f"D{linha_header_dm + 1}"

    linha = linha_header_dm + 1
    if not lojas_ordenadas:
        ws2.cell(row=linha, column=1, value="Nenhuma loja/unidade disponível para divisão.").font = EXCEL_STYLE["font_normal"]
        linha += 1
    for loja in lojas_ordenadas:
        df_o_loja, df_r_loja = dados_por_loja[loja]

        for conta in contas_sel:
            valores_real = [get_valor_consolidado_multi([df_r_loja], conta, [m_col]) for m_col in mapa_meses.values()]
            valores_orc = [get_valor_consolidado_multi([df_o_loja], conta, [m_col]) for m_col in mapa_meses.values()]
            valores_desvio = [vr - vo for vr, vo in zip(valores_real, valores_orc)]

            _escrever_linha_flat(ws2, linha, [loja, conta, "Realizado"], valores_real, sum(valores_real))
            linha += 1
            _escrever_linha_flat(ws2, linha, [loja, conta, "Orçado"], valores_orc, sum(valores_orc), fill=EXCEL_STYLE["fill_zebra"])
            linha += 1
            _escrever_linha_flat(ws2, linha, [loja, conta, "Desvio"], valores_desvio, sum(valores_desvio), colorir_por_sinal=True)
            linha += 1

    ultima_linha_dm = linha - 1
    largura_mes = 14
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 12
    for col in range(n_campos_dm + 1, n_col_dm + 1):
        ws2.column_dimensions[get_column_letter(col)].width = largura_mes
    if ultima_linha_dm >= linha_header_dm + 1:
        ws2.auto_filter.ref = f"A{linha_header_dm}:{get_column_letter(n_col_dm)}{ultima_linha_dm}"

    # ---------------- ABA "PLANO DE CONTAS" (composição de cada linha da DRE, por loja) ----------------
    # Mesmo formato tabular achatado, com AutoFilter na coluna "Loja".
    ws3 = wb.create_sheet("Plano de Contas")
    n_campos_pc = 3  # Loja, Conta, Plano de Contas
    n_col_pc = n_campos_pc + n_meses + 1
    _escrever_titulo(ws3, "Plano de Contas — Composição das Linhas da DRE, por Loja", 1, n_col_pc)
    _escrever_legenda(
        ws3,
        f"{gerado_em} · Valores de cada plano de contas puxados da DIÁRIO (Realizado), pela loja/Centro "
        f"de Custos. Visões consolidadas somam as lojas individuais correspondentes. Use o filtro (▾) no "
        f"cabeçalho da coluna \"Loja\" para escolher qual visão ver.",
        2, n_col_pc,
    )

    linha_header_pc = 4
    headers_pc = ["Loja", "Conta", "Plano de Contas"] + [m.capitalize() for m in mapa_meses.keys()] + ["Total Ano"]
    for col, texto in enumerate(headers_pc, start=1):
        cell = ws3.cell(row=linha_header_pc, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_header"]
        cell.fill = EXCEL_STYLE["fill_header"]
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col <= n_campos_pc else "center")
    ws3.row_dimensions[linha_header_pc].height = 20
    ws3.freeze_panes = f"D{linha_header_pc + 1}"

    linha = linha_header_pc + 1
    if not lojas_ordenadas:
        ws3.cell(row=linha, column=1, value="Nenhuma loja/unidade disponível para divisão.").font = EXCEL_STYLE["font_normal"]
        linha += 1

    indice_grupo = 0
    for loja in lojas_ordenadas:
        df_o_loja, df_r_loja = dados_por_loja[loja]

        # Lançamentos desta loja -- ou, se for uma visão consolidada, de
        # todas as lojas individuais que fazem parte dela.
        if _normalizar_nome_aba(loja) in _LOJAS_CONSOLIDADAS_NORM:
            lojas_do_grupo = _lojas_do_grupo_consolidado(loja)
            df_lanc_loja = df_lanc[df_lanc["Loja"].isin(lojas_do_grupo)] if lojas_do_grupo else df_lanc.iloc[0:0]
        else:
            df_lanc_loja = df_lanc[df_lanc["Loja"] == loja]

        for conta in contas_sel:
            # Cada conta (linha da DRE) forma um "grupo" -- as linhas de
            # plano de contas dela, mais o TOTAL, recebem o mesmo
            # sombreamento leve alternado, pra ficar visualmente claro onde
            # um grupo termina e o próximo começa (em vez de tudo num fundo
            # branco igual, que se perde fácil numa planilha grande).
            indice_grupo += 1
            fill_grupo = EXCEL_STYLE["fill_group"] if indice_grupo % 2 == 1 else None
            df_diario_conta = pd.DataFrame()
            if not df_lanc_loja.empty:
                conta_norm = _normalizar_texto(conta)
                linhas_norm_loja = df_lanc_loja["Linha DRE"].map(_normalizar_texto)
                mask_conta_loja = linhas_norm_loja == conta_norm
                if not mask_conta_loja.any():
                    mask_conta_loja = linhas_norm_loja.apply(lambda x: (conta_norm in x) or (x in conta_norm) if x else False)
                df_diario_conta = df_lanc_loja[mask_conta_loja]

            soma_planos_mes = [0.0] * n_meses
            planos_globais_conta = planos_por_conta_global.get(conta)
            if planos_globais_conta:
                # Fonte principal: DIÁRIO/Lançamentos. Usa o conjunto GLOBAL
                # de planos de contas dessa linha da DRE (calculado juntando
                # todas as lojas) -- assim TODA loja mostra as mesmas linhas
                # de plano de contas, mesmo que, para essa loja específica,
                # o valor seja zero (sem lançamento).
                for plano in planos_globais_conta:
                    grupo = df_diario_conta[df_diario_conta["Plano de Contas"] == plano] if not df_diario_conta.empty else df_diario_conta
                    valores_plano = (
                        [grupo.loc[grupo["Mês"] == m_col, "Valor Bruto"].sum() for m_col in mapa_meses.values()]
                        if not grupo.empty else [0.0] * n_meses
                    )
                    soma_planos_mes = [s + v for s, v in zip(soma_planos_mes, valores_plano)]
                    _escrever_linha_flat(ws3, linha, [loja, conta, plano], valores_plano, sum(valores_plano), fill=fill_grupo)
                    linha += 1
                n_planos_escritos = len(planos_globais_conta)
            elif permitir_lancamento_manual and not (mapa_planos_dre.get(str(conta).strip(), [])):
                # Linha DRE sem nenhum Plano de Contas correspondente (nem no
                # DIÁRIO, nem na Tabela_Contas) -- típico do modelo de RH.
                valores_manual = [
                    get_valor_consolidado_multi([df_r_loja], conta, [m_col]) for m_col in mapa_meses.values()
                ]
                soma_planos_mes = valores_manual
                _escrever_linha_flat(ws3, linha, [loja, conta, "Lançado Manualmente"], valores_manual, sum(valores_manual), fill=fill_grupo)
                linha += 1
                n_planos_escritos = 1
            else:
                # Fallback: Tabela_Contas + soma nas abas por loja (método antigo).
                planos = mapa_planos_dre.get(str(conta).strip(), []) or [conta]
                conta_norm_cmp = _normalizar_texto(conta)
                for plano in planos:
                    valores_plano = [
                        get_valor_consolidado_multi([df_r_loja], plano, [m_col]) for m_col in mapa_meses.values()
                    ]
                    soma_planos_mes = [s + v for s, v in zip(soma_planos_mes, valores_plano)]
                    # Quando não há um Plano de Contas distinto de verdade
                    # (caiu no fallback "ou [conta]", ou a Tabela_Contas só
                    # tem o próprio nome da linha da DRE como "plano"), o
                    # valor já vem certo, mas o rótulo repetia a conta --
                    # mostramos "Lançado Manualmente" para deixar claro que
                    # esse valor foi lançado direto na linha da DRE, sem
                    # composição por plano de contas. Vale para qualquer
                    # modelo de relatório, não só o de RH.
                    rotulo_plano = "Lançado Manualmente" if _normalizar_texto(plano) == conta_norm_cmp else plano
                    _escrever_linha_flat(ws3, linha, [loja, conta, rotulo_plano], valores_plano, sum(valores_plano), fill=fill_grupo)
                    linha += 1
                n_planos_escritos = len(planos)

            # A linha de TOTAL só faz sentido (e só é escrita) quando há MAIS
            # DE UM plano de contas para essa linha da DRE -- com um único
            # plano, o total seria idêntico à própria linha, repetindo o
            # mesmo valor à toa. O rótulo cita a própria conta ("TOTAL —
            # <conta>") para deixar explícito a quais linhas de plano de
            # contas, logo acima, aquele total se refere -- em vez de
            # depender só da cor de fundo pra não se perder na leitura.
            if n_planos_escritos > 1:
                _escrever_linha_flat(
                    ws3, linha, [loja, conta, f"TOTAL — {conta}"], soma_planos_mes, sum(soma_planos_mes),
                    negrito=True, fill=EXCEL_STYLE["fill_total"],
                )
                linha += 1

        if forcar_planos_contas:
            # Planos de contas que fazem parte do modelo mas não têm Linha DRE
            # correspondente (ex.: "Mercadorias" no modelo de Compras) --
            # puxados direto da DIÁRIO/Lançamentos pelo nome do plano,
            # ignorando a Linha DRE.
            indice_grupo += 1
            fill_grupo_forcado = EXCEL_STYLE["fill_group"] if indice_grupo % 2 == 1 else None
            soma_forcados_mes = [0.0] * n_meses
            for plano_forcado in forcar_planos_contas:
                df_plano_forcado = pd.DataFrame()
                if not df_lanc_loja.empty:
                    plano_norm = _normalizar_texto(plano_forcado)
                    planos_norm_loja = df_lanc_loja["Plano de Contas"].map(_normalizar_texto)
                    mask_plano = planos_norm_loja == plano_norm
                    if not mask_plano.any():
                        mask_plano = planos_norm_loja.apply(lambda x: (plano_norm in x) or (x in plano_norm) if x else False)
                    df_plano_forcado = df_lanc_loja[mask_plano]

                valores_plano = (
                    [df_plano_forcado.loc[df_plano_forcado["Mês"] == m_col, "Valor Bruto"].sum() for m_col in mapa_meses.values()]
                    if not df_plano_forcado.empty else [0.0] * n_meses
                )
                soma_forcados_mes = [s + v for s, v in zip(soma_forcados_mes, valores_plano)]
                _escrever_linha_flat(
                    ws3, linha, [loja, "(Fora da DRE)", plano_forcado], valores_plano, sum(valores_plano),
                    fill=fill_grupo_forcado,
                )
                linha += 1

            if len(forcar_planos_contas) > 1:
                _escrever_linha_flat(
                    ws3, linha, [loja, "(Fora da DRE)", "TOTAL — Fora da DRE"], soma_forcados_mes, sum(soma_forcados_mes),
                    negrito=True, fill=EXCEL_STYLE["fill_total"],
                )
                linha += 1

    ultima_linha_pc = linha - 1
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 42
    for col in range(n_campos_pc + 1, n_col_pc + 1):
        ws3.column_dimensions[get_column_letter(col)].width = largura_mes
    if ultima_linha_pc >= linha_header_pc + 1:
        ws3.auto_filter.ref = f"A{linha_header_pc}:{get_column_letter(n_col_pc)}{ultima_linha_pc}"

    # ---------------- ABA "LANÇAMENTOS" (cópia filtrada da DIÁRIO) ----------------
    ws4 = wb.create_sheet("Lançamentos")
    N_COL_LANC = 9
    _escrever_titulo(ws4, "Lançamentos — Cópia filtrada da aba DIÁRIO (Realizado)", 1, N_COL_LANC)
    _escrever_legenda(
        ws4,
        f"{gerado_em} · {len(df_lanc)} lançamento(s): linhas cujo Plano de Contas pertence às linhas "
        f"da DRE deste relatório, ou aos planos adicionais do modelo (ex.: Mercadorias).",
        2, N_COL_LANC,
    )

    linha = 4
    headers_lanc = [
        "Loja", "Competência", "Número", "Centro de Custos", "Linha DRE",
        "Plano de Contas", "Cliente / Fornecedor", "Histórico", "Valor Bruto (R$)",
    ]
    col_valor_lanc = len(headers_lanc)
    for col, texto in enumerate(headers_lanc, start=1):
        cell = ws4.cell(row=linha, column=col, value=texto)
        cell.font = EXCEL_STYLE["font_header"]
        cell.fill = EXCEL_STYLE["fill_header"]
        cell.border = EXCEL_STYLE["border"]
        cell.alignment = Alignment(horizontal="left" if col < col_valor_lanc else "center")
    ws4.row_dimensions[linha].height = 20
    ws4.freeze_panes = f"A{linha + 1}"
    linha += 1

    if not df_lanc.empty:
        df_lanc_ordenado = df_lanc.sort_values(["Loja", "Competência"], na_position="last")
        for i, (_, row) in enumerate(df_lanc_ordenado.iterrows()):
            data_val = row["Competência"]
            data_txt = data_val.strftime("%d/%m/%Y") if pd.notna(data_val) and hasattr(data_val, "strftime") else str(data_val)
            valores = [
                row["Loja"], data_txt, row.get("Número", ""), row["Centro de Custos"], row["Linha DRE"],
                row["Plano de Contas"], row.get("Cliente / Fornecedor", ""), row.get("Histórico", ""), row["Valor Bruto"],
            ]
            fill = EXCEL_STYLE["fill_zebra"] if i % 2 == 1 else None
            for col, val in enumerate(valores, start=1):
                cell = ws4.cell(row=linha, column=col, value=val)
                cell.border = EXCEL_STYLE["border"]
                if fill:
                    cell.fill = fill
                if col == col_valor_lanc:
                    cell.font = EXCEL_STYLE["font_normal"]
                    cell.number_format = '"R$" #,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.font = EXCEL_STYLE["font_normal"]
                    cell.alignment = Alignment(horizontal="left")
            linha += 1

        # AutoFilter nativo do Excel -- cobre só cabeçalho + linhas de dados
        # (não a linha de TOTAL, escrita logo abaixo).
        ws4.auto_filter.ref = f"A4:{get_column_letter(N_COL_LANC)}{linha - 1}"

        total_lanc = df_lanc["Valor Bruto"].sum()
        cell_total_lbl = ws4.cell(row=linha, column=1, value="TOTAL")
        cell_total_lbl.font = EXCEL_STYLE["font_bold"]
        cell_total = ws4.cell(row=linha, column=col_valor_lanc, value=total_lanc)
        cell_total.font = EXCEL_STYLE["font_bold"]
        cell_total.number_format = '"R$" #,##0.00'
        cell_total.alignment = Alignment(horizontal="right")
        for col in range(1, N_COL_LANC + 1):
            ws4.cell(row=linha, column=col).fill = EXCEL_STYLE["fill_total"]
            ws4.cell(row=linha, column=col).border = EXCEL_STYLE["border"]
    else:
        ws4.cell(
            row=linha, column=1,
            value="Nenhum lançamento encontrado (DIÁRIO indisponível ou sem correspondência para os filtros deste relatório).",
        ).font = EXCEL_STYLE["font_normal"]

    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 20
    ws4.column_dimensions["E"].width = 30
    ws4.column_dimensions["F"].width = 30
    ws4.column_dimensions["G"].width = 28
    ws4.column_dimensions["H"].width = 40
    ws4.column_dimensions["I"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# 7. FAIXA DE CONTEXTO (fina — não repete um bloco grande em toda aba)
# ============================================================================
st.markdown(
    html_compacto(f"""
    <div class="top-status-strip">
        <span class="chip">{label_visao}</span>
        <span class="sep">·</span>
        <span>Período: <b>{label_periodo_kpi}</b></span>
        <span class="sep">·</span>
        <span>Controladoria B&amp;A · Painel Financeiro</span>
    </div>
    """),
    unsafe_allow_html=True,
)


# ============================================================================
# 7.5 MOTOR DE ALERTAS (Controladoria)
# ============================================================================
# Em vez de esperar alguém abrir cada aba e comparar números, o painel avalia
# um conjunto de regras a cada carregamento e mostra o que exige atenção.
# Os limites ficam configuráveis porque o que é "grave" muda de empresa para
# empresa -- e um alerta que dispara sempre vira ruído e passa a ser ignorado.

def _avaliar_alertas_controladoria(
    list_real, list_orc, colunas_periodo, m_map_alertas, linhas_disponiveis,
    limite_desvio_ebitda, limite_estouro_conta, valor_minimo_conta,
):
    """Devolve uma lista de alertas, cada um com nível, título e detalhe.
    Nível: 'critico' (vermelho), 'atencao' (amarelo) ou 'ok' (verde)."""
    alertas = []

    # --- 1. Desvio de EBITDA contra o orçado PROPORCIONAL ---
    # Se o período inclui o mês corrente, o orçamento entra só na fração de
    # dias já decorridos -- senão o alerta dispararia todo mês, sempre.
    hoje_alerta = datetime.now(FUSO_BR).date()
    ebitda_real = get_valor_consolidado_multi(list_real, "11 - EBITDA", colunas_periodo)
    ebitda_orc = _orcado_proporcional(
        list_orc, "11 - EBITDA", colunas_periodo, list(m_map_alertas.values()), hoje_alerta
    )
    if ebitda_orc:
        desvio_pct = (ebitda_real - ebitda_orc) / abs(ebitda_orc) * 100
        if desvio_pct <= -limite_desvio_ebitda:
            alertas.append({
                "nivel": "critico",
                "titulo": f"EBITDA {abs(desvio_pct):.1f}% abaixo do ritmo orçado",
                "detalhe": (
                    f"Realizado {formata_brl(ebitda_real)} contra {formata_brl(ebitda_orc)} esperados "
                    f"até hoje (orçamento ajustado aos dias decorridos) — diferença de "
                    f"{formata_brl(ebitda_real - ebitda_orc)}."
                ),
            })

    # --- 2. Receita abaixo da meta ---
    rec_real = get_valor_consolidado_multi(list_real, "3 - Receita Operacional Liquida", colunas_periodo)
    rec_orc = _orcado_proporcional(
        list_orc, "3 - Receita Operacional Liquida", colunas_periodo,
        list(m_map_alertas.values()), hoje_alerta
    )
    if rec_orc:
        atingimento = rec_real / rec_orc * 100
        if atingimento < (100 - limite_desvio_ebitda):
            alertas.append({
                "nivel": "atencao",
                "titulo": f"Receita líquida em {atingimento:.1f}% do ritmo esperado",
                "detalhe": (
                    f"Faltam {formata_brl(rec_orc - rec_real)} para acompanhar o orçamento "
                    f"proporcional aos dias já decorridos."
                ),
            })

    # --- 3. Contas que estouraram o orçamento ---
    linhas_custo = [l for l in linhas_disponiveis if eh_linha_custos_despesas(l)]
    linhas_custo_raiz = _linhas_raiz_do_conjunto(linhas_custo)
    estouros = []
    for linha in linhas_custo_raiz:
        v_real = abs(get_valor_consolidado_multi(list_real, linha, colunas_periodo, exato_linha_sintetica=True))
        v_orc = abs(_orcado_proporcional(
            list_orc, linha, colunas_periodo, list(m_map_alertas.values()), hoje_alerta, exato=True
        ))
        if v_orc <= 0 or v_real < valor_minimo_conta:
            continue
        excesso_pct = (v_real - v_orc) / v_orc * 100
        if excesso_pct >= limite_estouro_conta:
            estouros.append({
                "nome": _nome_sem_numero_dre(linha),
                "excesso_pct": excesso_pct,
                "excesso_valor": v_real - v_orc,
            })
    estouros.sort(key=lambda e: e["excesso_valor"], reverse=True)
    for e in estouros[:5]:
        alertas.append({
            "nivel": "critico" if e["excesso_pct"] >= limite_estouro_conta * 2 else "atencao",
            "titulo": f"{e['nome']}: {e['excesso_pct']:.0f}% acima do ritmo orçado",
            "detalhe": f"Excesso de {formata_brl(e['excesso_valor'])} no período.",
        })
    if len(estouros) > 5:
        alertas.append({
            "nivel": "atencao",
            "titulo": f"Mais {len(estouros) - 5} conta(s) acima do orçamento",
            "detalhe": "Veja a lista completa na aba DRE Orçado X Realizado.",
        })

    # --- 4. Margem EBITDA em queda por 3 meses seguidos ---
    colunas_mes = list(m_map_alertas.values())
    margens = []
    for col in colunas_mes:
        rec_mes = get_valor_consolidado_multi(list_real, "3 - Receita Operacional Liquida", [col])
        ebitda_mes = get_valor_consolidado_multi(list_real, "11 - EBITDA", [col])
        if rec_mes:
            margens.append((col, ebitda_mes / rec_mes * 100))
    if len(margens) >= 3:
        ultimas = [m for _, m in margens[-3:]]
        if ultimas[0] > ultimas[1] > ultimas[2]:
            alertas.append({
                "nivel": "atencao",
                "titulo": "Margem EBITDA em queda há 3 meses seguidos",
                "detalhe": (
                    f"Caiu de {ultimas[0]:.1f}% para {ultimas[2]:.1f}% "
                    f"({ultimas[2] - ultimas[0]:+.1f} p.p. no período)."
                ),
            })

    return alertas


def _renderizar_painel_alertas(alertas, titulo="Alertas"):
    """Mostra os alertas em cartões compactos, ordenados por gravidade."""
    if not alertas:
        st.success(f"✅ {titulo}: nenhum ponto de atenção nas regras configuradas.")
        return

    criticos = [a for a in alertas if a["nivel"] == "critico"]
    atencoes = [a for a in alertas if a["nivel"] == "atencao"]

    resumo = []
    if criticos:
        resumo.append(f"{len(criticos)} crítico(s)")
    if atencoes:
        resumo.append(f"{len(atencoes)} de atenção")
    rotulo_expander = f"🔔 {titulo} — {' · '.join(resumo)}"

    with st.expander(rotulo_expander, expanded=bool(criticos)):
        for alerta in criticos + atencoes:
            if alerta["nivel"] == "critico":
                cor_borda, icone = COLORS["negative"], "🚨"
                fundo = "rgba(247,110,110,0.08)"
            else:
                cor_borda, icone = COLORS["warning"], "⚠️"
                fundo = "rgba(245,166,35,0.08)"
            st.markdown(
                html_compacto(f"""
                <div style="border-left: 3px solid {cor_borda}; background: {fundo};
                            padding: 8px 12px; border-radius: 4px; margin-bottom: 7px;">
                    <div style="color:{COLORS['text']}; font-size:13px; font-weight:600;">
                        {icone} {alerta['titulo']}
                    </div>
                    <div style="color:{COLORS['text_muted']}; font-size:11.5px; margin-top:2px;">
                        {alerta['detalhe']}
                    </div>
                </div>
                """),
                unsafe_allow_html=True,
            )


# ---- Configuração dos limites de alerta (barra lateral) ----
# Os limites valem tanto para a visão de Controladoria quanto para a de
# departamento, então ficam disponíveis para todos os perfis.
with st.sidebar.expander("🔔 Regras de alerta"):
    st.caption("Define a partir de que ponto o painel considera algo digno de atenção.")
    limite_desvio_ebitda_cfg = st.slider(
        "Desvio de EBITDA/receita que dispara alerta (%)",
        min_value=1, max_value=30, value=10, step=1, key="cfg_limite_ebitda",
    )
    limite_estouro_conta_cfg = st.slider(
        "Estouro de conta que dispara alerta (%)",
        min_value=5, max_value=100, value=20, step=5, key="cfg_limite_conta",
    )
    valor_minimo_conta_cfg = st.number_input(
        "Ignorar contas abaixo de (R$)",
        min_value=0, value=10_000, step=5_000, key="cfg_valor_minimo",
        help="Evita alerta em conta pequena que estourou percentualmente mas é irrelevante em valor.",
    )

def _avaliar_alertas_departamento(
    list_real, list_orc, colunas_periodo, m_map_alertas, linhas_dept, linhas_dept_raiz,
    nome_dept, limite_estouro, valor_minimo,
):
    """Alertas pensados para quem gere UM departamento. Aqui não faz sentido
    falar de EBITDA ou de receita da companhia -- o gestor responde pelo
    gasto das linhas dele, pelo ritmo desse gasto e pela tendência dele."""
    alertas = []
    hoje_dep = datetime.now(FUSO_BR).date()
    colunas_ano = list(m_map_alertas.values())

    # --- 1. Gasto total do departamento contra o ritmo orçado ---
    gasto_real = abs(sum(
        get_valor_consolidado_multi(list_real, l, colunas_periodo, exato_linha_sintetica=True)
        for l in linhas_dept_raiz
    ))
    gasto_orc_prop = abs(sum(
        _orcado_proporcional(list_orc, l, colunas_periodo, colunas_ano, hoje_dep, exato=True)
        for l in linhas_dept_raiz
    ))
    if gasto_orc_prop:
        excesso_pct = (gasto_real - gasto_orc_prop) / gasto_orc_prop * 100
        if excesso_pct >= limite_estouro:
            alertas.append({
                "nivel": "critico" if excesso_pct >= limite_estouro * 2 else "atencao",
                "titulo": f"{nome_dept} está {excesso_pct:.0f}% acima do ritmo orçado",
                "detalhe": (
                    f"Gasto de {formata_brl(gasto_real)} contra {formata_brl(gasto_orc_prop)} "
                    f"esperados até aqui — excesso de {formata_brl(gasto_real - gasto_orc_prop)}."
                ),
            })
        elif excesso_pct <= -limite_estouro:
            alertas.append({
                "nivel": "atencao",
                "titulo": f"{nome_dept} está {abs(excesso_pct):.0f}% abaixo do orçado",
                "detalhe": (
                    "Pode ser economia real ou lançamento que ainda não entrou — vale conferir "
                    "se há notas pendentes de registro."
                ),
            })

    # --- 2. Contas do departamento que estouraram ---
    folhas_dept = _linhas_folha_do_conjunto(linhas_dept)
    estouros_dept = []
    for linha in folhas_dept:
        v_real = abs(get_valor_consolidado_multi(list_real, linha, colunas_periodo, exato_linha_sintetica=True))
        v_orc = abs(_orcado_proporcional(list_orc, linha, colunas_periodo, colunas_ano, hoje_dep, exato=True))
        if v_orc <= 0 or v_real < valor_minimo:
            continue
        exc = (v_real - v_orc) / v_orc * 100
        if exc >= limite_estouro:
            estouros_dept.append((_nome_sem_numero_dre(linha), exc, v_real - v_orc))
    estouros_dept.sort(key=lambda e: e[2], reverse=True)
    for nome_conta, exc, valor_exc in estouros_dept[:4]:
        alertas.append({
            "nivel": "critico" if exc >= limite_estouro * 2 else "atencao",
            "titulo": f"{nome_conta}: {exc:.0f}% acima do ritmo orçado",
            "detalhe": f"Excesso de {formata_brl(valor_exc)} no período.",
        })

    # --- 3. Tendência: gasto subindo há 3 meses seguidos ---
    serie_dept = []
    for col in colunas_ano:
        valor_mes = abs(sum(
            get_valor_consolidado_multi(list_real, l, [col], exato_linha_sintetica=True)
            for l in linhas_dept_raiz
        ))
        if valor_mes > 0:
            serie_dept.append(valor_mes)
    if len(serie_dept) >= 3 and serie_dept[-3] < serie_dept[-2] < serie_dept[-1]:
        crescimento = (serie_dept[-1] / serie_dept[-3] - 1) * 100
        alertas.append({
            "nivel": "atencao",
            "titulo": f"Gasto de {nome_dept} sobe há 3 meses seguidos",
            "detalhe": (
                f"De {formata_brl(serie_dept[-3])} para {formata_brl(serie_dept[-1])} "
                f"({crescimento:+.0f}% no período)."
            ),
        })

    # --- 4. Concentração: uma conta domina o gasto do departamento ---
    if folhas_dept and gasto_real:
        valores_folhas = [
            (
                _nome_sem_numero_dre(l),
                abs(get_valor_consolidado_multi(list_real, l, colunas_periodo, exato_linha_sintetica=True)),
            )
            for l in folhas_dept
        ]
        valores_folhas = [(n, v) for n, v in valores_folhas if v > 0]
        if valores_folhas:
            valores_folhas.sort(key=lambda x: x[1], reverse=True)
            maior_nome, maior_valor = valores_folhas[0]
            peso = maior_valor / sum(v for _, v in valores_folhas) * 100
            if peso >= 60 and len(valores_folhas) > 1:
                alertas.append({
                    "nivel": "atencao",
                    "titulo": f"{peso:.0f}% do gasto está em uma única conta",
                    "detalhe": (
                        f"{maior_nome} responde por {formata_brl(maior_valor)} do total do "
                        f"departamento — qualquer variação nela move o resultado inteiro."
                    ),
                })

    return alertas


if departamento_ativo:
    if linhas_departamento_resolvidas:
        _nome_dept_alerta = _nome_departamento_curto(departamento_ativo)
        _alertas_dept = _avaliar_alertas_departamento(
            list_df_real, list_df_orc, cols_kpi, m_map,
            linhas_departamento_resolvidas, linhas_departamento_raiz,
            _nome_dept_alerta, limite_estouro_conta_cfg, valor_minimo_conta_cfg,
        )
        _renderizar_painel_alertas(
            _alertas_dept, titulo=f"Alertas de {_nome_dept_alerta} · {label_periodo_kpi}"
        )
else:
    _col_nome_alerta = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
    _linhas_alerta = list(df_ref[_col_nome_alerta].dropna().astype(str).unique())
    _alertas_controladoria = _avaliar_alertas_controladoria(
        list_df_real, list_df_orc, cols_kpi, m_map, _linhas_alerta,
        limite_desvio_ebitda_cfg, limite_estouro_conta_cfg, valor_minimo_conta_cfg,
    )
    _renderizar_painel_alertas(
        _alertas_controladoria, titulo=f"Alertas de {label_visao} · {label_periodo_kpi}"
    )


# ============================================================================
# 8. ABAS
# ============================================================================
if departamento_ativo:
    # Modo Departamento -- painel inteiro focado só no departamento
    # escolhido: 5 abas próprias (inclusive Impacto & Tendências e Emitir
    # Relatório, já filtrado pro modelo do departamento), sem Previsões
    # nem Usuários (que são conceitos de companhia toda, não de um
    # departamento específico).
    _nome_dept_abas = _nome_departamento_curto(departamento_ativo)
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        f"📊 Visão Geral — {_nome_dept_abas}",
        f"📋 Detalhe por Conta — {_nome_dept_abas}",
        f"📅 Histórico Mensal — {_nome_dept_abas}",
        f"🎯 Impacto & Tendências — {_nome_dept_abas}",
        f"📤 Emitir Relatório — {_nome_dept_abas}",
    ])
    tab6 = None
    tab_diag = None  # Diagnóstico Executivo é visão de companhia, não de departamento
else:
    _nomes_abas = [
        "📊 Visão Geral & Charts",
        "📋 DRE Orçado X Realizado",
        "📅 Histórico Mensal",
        "🔬 Diagnóstico Executivo",
        "🔮 Previsões & Trends",
        "📤 Emitir Relatório",
    ]
    if eh_admin:
        _nomes_abas.append("👥 Usuários")
        tab1, tab2, tab3, tab_diag, tab4, tab5, tab6 = st.tabs(_nomes_abas)
    else:
        tab1, tab2, tab3, tab_diag, tab4, tab5 = st.tabs(_nomes_abas)
        tab6 = None

# ---------------------------------------------------------------------------
# ABA 1: VISÃO GERAL & CHARTS
# ---------------------------------------------------------------------------
with tab1:
    if departamento_ativo:
        nome_departamento_curto = _nome_departamento_curto(departamento_ativo)
        st.markdown(
            f'<div class="section-title">🏢 Visão do Departamento — {nome_departamento_curto}</div>',
            unsafe_allow_html=True,
        )
        if not linhas_departamento_resolvidas:
            st.warning(
                f"Não encontrei nenhuma linha da DRE atual que bata com o modelo de {nome_departamento_curto}. "
                "O texto das linhas pode estar um pouco diferente do esperado."
            )
        else:
            dept_real = sum(
                get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True)
                for l in linhas_departamento_raiz
            )
            dept_orc = sum(
                get_valor_consolidado_multi(list_df_orc, l, cols_kpi, exato_linha_sintetica=True)
                for l in linhas_departamento_raiz
            )
            dept_real_abs = abs(dept_real)
            dept_orc_abs = abs(dept_orc)
            desvio_dept = dept_orc_abs - dept_real_abs  # gasto menor que orçado é favorável
            pct_atg_dept = (dept_real_abs / dept_orc_abs * 100) if dept_orc_abs else 0

            st.markdown(
                render_kpi_row([
                    dict(label="TOTAL REALIZADO (PERÍODO)", value=formata_brl(dept_real_abs),
                         value_color=COLORS["text"], subtext=f"{len(linhas_departamento_resolvidas)} linha(s) da DRE", icon="💸"),
                    dict(label="TOTAL ORÇADO (PERÍODO)", value=formata_brl(dept_orc_abs),
                         value_color=COLORS["text_muted"], subtext="Referência do orçamento", icon="🎯"),
                    dict(label="DESVIO vs. ORÇADO", value=formata_brl(desvio_dept), value_color=cor_variacao(desvio_dept),
                         subtext="Positivo = gastou menos que o orçado", subtext_color=cor_variacao(desvio_dept), icon="⚖️"),
                    dict(label="% DO ORÇADO CONSUMIDO", value=f"{pct_atg_dept:.1f}%", value_color=cor_variacao(-(pct_atg_dept - 100)),
                         subtext=f"Período: {label_periodo_kpi}", icon="📊"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ---- Impacto no resultado da empresa (Receita, EBITDA, Custos+Despesas totais) ----
            rec_liq_real_dept = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_kpi)
            ebitda_real_dept = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)
            cmv_real_dept = abs(get_valor_consolidado_multi(list_df_real, "4 - ", cols_kpi, exato_linha_sintetica=True)) \
                or abs(get_valor_consolidado_multi(list_df_real, "4 - Custo das Vendas", cols_kpi))
            dvar_real_dept = abs(get_valor_consolidado_multi(list_df_real, "6 - Despesas Variáveis", cols_kpi))
            dop_real_dept = abs(get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_kpi))
            total_custos_desp_dept = cmv_real_dept + dvar_real_dept + dop_real_dept

            pct_da_receita_dept = (dept_real_abs / rec_liq_real_dept * 100) if rec_liq_real_dept else 0.0
            pct_do_ebitda_dept = (dept_real_abs / abs(ebitda_real_dept) * 100) if ebitda_real_dept else 0.0
            pct_dos_custos_dept = (dept_real_abs / total_custos_desp_dept * 100) if total_custos_desp_dept else 0.0

            linhas_maior_conta_dept = _linhas_composicao_do_conjunto(linhas_departamento_resolvidas)
            st.markdown('<div class="section-title">🎯 Peso do Departamento no Resultado da Empresa</div>', unsafe_allow_html=True)
            st.markdown(
                render_kpi_row([
                    dict(label="% DA RECEITA LÍQUIDA", value=f"{pct_da_receita_dept:.1f}%",
                         value_color=COLORS["text"], subtext="O quanto o departamento consome da receita", icon="💰"),
                    dict(label="% DO EBITDA DA EMPRESA", value=f"{pct_do_ebitda_dept:.1f}%",
                         value_color=COLORS["warning"], subtext=f"EBITDA do período: {formata_brl(ebitda_real_dept)}", icon="📐"),
                    dict(label="% DOS CUSTOS + DESPESAS TOTAIS", value=f"{pct_dos_custos_dept:.1f}%",
                         value_color=COLORS["text"], subtext="Fatia do departamento no total de custos e despesas", icon="🧾"),
                    dict(label="MAIOR CONTA DO DEPARTAMENTO",
                         value=formata_m(max(
                             (abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True)) for l in linhas_maior_conta_dept),
                             default=0,
                         )),
                         value_color=COLORS["negative"],
                         subtext=_nome_sem_numero_dre(max(
                             linhas_maior_conta_dept,
                             key=lambda l: abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True)),
                         ))[:38] if linhas_maior_conta_dept else "",
                         icon="🔎"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            col_donut_dept, col_evol_dept = st.columns(2)
            with col_donut_dept:
                st.markdown('<div class="section-title">🍩 Composição do Departamento</div>', unsafe_allow_html=True)
                linhas_composicao_dept = _linhas_composicao_do_conjunto(linhas_departamento_resolvidas)
                pares_donut_dept = [
                    (l, abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True)))
                    for l in linhas_composicao_dept
                ]
                pares_donut_dept = [(l, v) for l, v in pares_donut_dept if v > 0]
                pares_donut_dept.sort(key=lambda par: par[1], reverse=True)
                # Com muitas contas, as fatias menores viram um amontoado de
                # rótulos ilegíveis -- mostra só as maiores e agrupa o resto
                # em "Outros", que ainda soma certinho no total do meio.
                TOP_N_DONUT = 7
                if len(pares_donut_dept) > TOP_N_DONUT:
                    soma_outros_donut = sum(v for _, v in pares_donut_dept[TOP_N_DONUT:])
                    pares_donut_dept = pares_donut_dept[:TOP_N_DONUT] + [("__outros__", soma_outros_donut)]
                labels_donut_dept = [
                    "Outros" if l == "__outros__" else _nome_sem_numero_dre(l) for l, _ in pares_donut_dept
                ]
                valores_donut_dept = [v for _, v in pares_donut_dept]
                if sum(valores_donut_dept) > 0:
                    fig_donut_dept = go.Figure(data=[go.Pie(
                        labels=labels_donut_dept, values=valores_donut_dept, hole=0.55,
                        marker=dict(colors=[COLORS["primary"], COLORS["positive"], COLORS["warning"],
                                             COLORS["negative"], COLORS["secondary"], COLORS["muted_line"], "#6B7280", "#94A3B8"]),
                        textinfo="percent", textfont=dict(color=COLORS["text"], size=11),
                        texttemplate="%{percent:.1%}",
                    )])
                    fig_donut_dept.add_annotation(
                        text=f"<b>{formata_m(dept_real_abs)}</b><br><span style='font-size:10px;color:{COLORS['text_muted']}'>Total Realizado</span>",
                        showarrow=False, font=dict(color=COLORS["text"], size=13, family=FONT_STACK),
                    )
                    estilo_grafico(
                        fig_donut_dept, height=380,
                        legend=dict(orientation="h", yanchor="top", y=-0.12, xanchor="center", x=0.5, font=dict(size=10)),
                        margin=dict(l=10, r=10, t=20, b=10),
                    )
                    st.plotly_chart(fig_donut_dept, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                else:
                    st.info("Sem valores no período selecionado para montar o gráfico.")

            with col_evol_dept:
                st.markdown('<div class="section-title">📈 Evolução Mensal (Ano Completo)</div>', unsafe_allow_html=True)
                meses_evol_real = [
                    sum(get_valor_consolidado_multi(list_df_real, l, [m_col], exato_linha_sintetica=True) for l in linhas_departamento_raiz)
                    for m_col in m_map.values()
                ]
                meses_evol_orc = [
                    sum(get_valor_consolidado_multi(list_df_orc, l, [m_col], exato_linha_sintetica=True) for l in linhas_departamento_raiz)
                    for m_col in m_map.values()
                ]
                fig_evol_dept = go.Figure()
                fig_evol_dept.add_trace(go.Scatter(
                    x=list(m_map.keys()), y=[abs(v) for v in meses_evol_real], name="Realizado",
                    mode="lines+markers", line=dict(color=COLORS["primary"], width=2.5),
                ))
                fig_evol_dept.add_trace(go.Scatter(
                    x=list(m_map.keys()), y=[abs(v) for v in meses_evol_orc], name="Orçado",
                    mode="lines+markers", line=dict(color=COLORS["muted_line"], width=2, dash="dot"),
                ))
                estilo_grafico(
                    fig_evol_dept, height=380,
                    xaxis=dict(gridcolor=COLORS["border"], fixedrange=True, tickfont=dict(size=9),
                               tickangle=-35, automargin=True),
                    yaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.32, xanchor="center", x=0.5),
                    margin=dict(l=20, r=20, t=30, b=80),
                )
                st.plotly_chart(fig_evol_dept, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="section-title">📋 Linhas da DRE do Departamento</div>', unsafe_allow_html=True)
            linhas_tabela_dept = []
            for l in linhas_departamento_resolvidas:
                v_r = abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True))
                v_o = abs(get_valor_consolidado_multi(list_df_orc, l, cols_kpi, exato_linha_sintetica=True))
                linhas_tabela_dept.append({
                    "Conta / Linha DRE": l, "Realizado (R$)": v_r, "Orçado (R$)": v_o, "Desvio (R$)": v_o - v_r,
                })
            df_tabela_dept = pd.DataFrame(linhas_tabela_dept)
            cols_num_dept = ["Realizado (R$)", "Orçado (R$)", "Desvio (R$)"]
            st.dataframe(
                df_tabela_dept.style.format(
                    {
                        "Realizado (R$)": formata_brl,
                        "Orçado (R$)": formata_brl,
                        "Desvio (R$)": formata_brl,
                    }
                ).map(cor_valor, subset=cols_num_dept),
                column_config={
                    "Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large"),
                },
                use_container_width=True,
                hide_index=True,
            )

            # ---- Linhas informativas (ex.: no MKT, a gestão GB da
            # indústria) -- não entram em nenhum KPI/gráfico/soma do
            # departamento, mas ficam visíveis aqui só pra conhecimento dos
            # valores, já que aparecem na DRE consolidada. ----
            if linhas_departamento_informativas:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(
                    '<div class="section-title">ℹ️ Outras Linhas Relacionadas (fora da gestão do departamento)</div>',
                    unsafe_allow_html=True,
                )
                st.caption(
                    "Essas linhas aparecem na DRE dentro do mesmo grupo, mas não são geridas por este "
                    "departamento -- estão aqui só para conhecimento dos valores, e não entram em nenhum "
                    "KPI, gráfico ou soma do departamento acima."
                )
                linhas_tabela_info = []
                for l in linhas_departamento_informativas:
                    v_r_info = abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True))
                    v_o_info = abs(get_valor_consolidado_multi(list_df_orc, l, cols_kpi, exato_linha_sintetica=True))
                    linhas_tabela_info.append({
                        "Conta / Linha DRE": l, "Realizado (R$)": v_r_info, "Orçado (R$)": v_o_info,
                        "Desvio (R$)": v_o_info - v_r_info,
                    })
                df_tabela_info = pd.DataFrame(linhas_tabela_info)
                st.dataframe(
                    df_tabela_info.style.format(
                        {
                            "Realizado (R$)": formata_brl,
                            "Orçado (R$)": formata_brl,
                            "Desvio (R$)": formata_brl,
                        }
                    ).map(cor_valor, subset=cols_num_dept),
                    column_config={
                        "Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large"),
                    },
                    use_container_width=True,
                    hide_index=True,
                )

    else:
        st.markdown(
            '<div class="section-title">📊 Visão Geral & Charts — Indicadores Executivos e Composição do Resultado</div>',
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # ---- KPIs executivos da visão geral ----
        rec_liq_real_kpi = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_kpi)
        rec_liq_orc_kpi = get_valor_consolidado_multi(list_df_orc, "3 - Receita Operacional Liquida", cols_kpi)


        ebitda_real_kpi = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)
        ebitda_orc_kpi = get_valor_consolidado_multi(list_df_orc, "11 - EBITDA", cols_kpi)

        margem_ebitda_kpi = (ebitda_real_kpi / rec_liq_real_kpi * 100) if rec_liq_real_kpi != 0 else 0

        diff_ebitda_kpi = ebitda_real_kpi - ebitda_orc_kpi
        pct_ebitda_kpi = (diff_ebitda_kpi / abs(ebitda_orc_kpi)) * 100 if ebitda_orc_kpi != 0 else 0

        pct_vendas_prog = min(100.0, max(0.0, (rec_liq_real_kpi / rec_liq_orc_kpi * 100))) if rec_liq_orc_kpi > 0 else 0
        pct_lucro_prog = min(100.0, max(0.0, (ebitda_real_kpi / ebitda_orc_kpi * 100))) if ebitda_orc_kpi > 0 else 0

        cor_rec = cor_variacao(rec_liq_real_kpi)
        cor_ebitda = cor_variacao(ebitda_real_kpi)
        cor_diff_eb = cor_variacao(diff_ebitda_kpi)
        cor_mg_eb = cor_variacao(margem_ebitda_kpi)

        st.markdown(
            render_kpi_row([
                dict(label="RECEITA LÍQUIDA (YTD)", value=formata_brl(rec_liq_real_kpi), value_color=cor_rec,
                     subtext=f"Orçado: {formata_brl(rec_liq_orc_kpi)}", progress_pct=pct_vendas_prog, icon="💰"),
                dict(label="EBITDA (YTD)", value=formata_brl(ebitda_real_kpi), value_color=cor_ebitda,
                     subtext=f"Orçado: {formata_brl(ebitda_orc_kpi)}", progress_pct=pct_lucro_prog, icon="📈"),
                dict(label="VARIAÇÃO EBITDA", value=formata_brl(diff_ebitda_kpi), value_color=cor_diff_eb,
                     subtext=f"{pct_ebitda_kpi:+.1f}% vs Orçamento", subtext_color=cor_diff_eb, icon="⚖️"),
                dict(label="MARGEM EBITDA %", value=f"{margem_ebitda_kpi:.1f}%", value_color=cor_mg_eb,
                     subtext="Realizada no Período", icon="🎯"),
            ]),
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # ---- Ritmo do mês corrente: orçado proporcional aos dias ----
        # Comparar o realizado parcial de um mês em andamento contra o
        # orçamento cheio faz o desvio parecer sempre catastrófico. Aqui o
        # orçamento é ajustado aos dias já decorridos, que é a comparação
        # justa enquanto o mês não fecha.
        _hoje_ritmo = datetime.now(FUSO_BR).date()
        _col_mes_atual, _frac_mes, _dias_corridos, _dias_mes = _fator_proporcional_mes_corrente(
            cols_kpi, meses_cols, _hoje_ritmo
        )
        if _col_mes_atual is not None:
            _nome_mes_atual = next(
                (nome for nome, col in m_map.items() if col == _col_mes_atual), _col_mes_atual
            )
            rec_real_mes = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", [_col_mes_atual])
            rec_orc_mes_cheio = get_valor_consolidado_multi(list_df_orc, "3 - Receita Operacional Liquida", [_col_mes_atual])
            rec_orc_mes_prop = rec_orc_mes_cheio * _frac_mes
            gap_rec_mes = rec_real_mes - rec_orc_mes_prop
            ritmo_rec_pct = (rec_real_mes / rec_orc_mes_prop * 100) if rec_orc_mes_prop else 0

            ebitda_real_mes = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", [_col_mes_atual])
            ebitda_orc_mes_prop = get_valor_consolidado_multi(list_df_orc, "11 - EBITDA", [_col_mes_atual]) * _frac_mes
            ritmo_ebitda_pct = (ebitda_real_mes / ebitda_orc_mes_prop * 100) if ebitda_orc_mes_prop else 0

            # Projeção simples do mês: mantido o ritmo diário atual
            rec_projetada_mes = (rec_real_mes / _dias_corridos * _dias_mes) if _dias_corridos else 0

            # Faixa compacta em vez de uma fileira de cards: a informação é
            # de acompanhamento, não merece o mesmo peso visual dos KPIs
            # principais.
            _cor_ritmo = COLORS["positive"] if ritmo_rec_pct >= 100 else (
                COLORS["warning"] if ritmo_rec_pct >= 90 else COLORS["negative"]
            )
            _pct_barra = min(ritmo_rec_pct, 130)
            _proj_vs_orc = rec_projetada_mes - rec_orc_mes_cheio
            _cor_proj = COLORS["positive"] if _proj_vs_orc >= 0 else COLORS["negative"]
            _cor_ritmo_ebitda = COLORS["positive"] if ritmo_ebitda_pct >= 100 else (
                COLORS["warning"] if ritmo_ebitda_pct >= 90 else COLORS["negative"]
            )

            st.markdown(
                html_compacto(f"""
                <div style="background:{COLORS['surface']}; border:1px solid {COLORS['border']};
                            border-left:3px solid {_cor_ritmo}; border-radius:8px;
                            padding:12px 16px; margin-bottom:14px;">
                    <div style="display:flex; justify-content:space-between; align-items:center;
                                flex-wrap:wrap; gap:14px;">
                        <div style="min-width:190px; padding-right:20px;
                                    border-right:1px solid {COLORS['border']};">
                            <div style="font-size:10px; color:{COLORS['text_muted']};
                                        text-transform:uppercase; letter-spacing:0.5px;">
                                ⏱️ Ritmo de {_nome_mes_atual.capitalize()}
                            </div>
                            <div style="font-size:22px; font-weight:800; color:{_cor_ritmo};
                                        margin-top:3px; line-height:1.1;">
                                {ritmo_rec_pct:.0f}%
                            </div>
                            <div style="font-size:10.5px; color:{COLORS['text_muted']}; margin-top:2px;">
                                do esperado · dia {_dias_corridos} de {_dias_mes}
                            </div>
                        </div>
                        <div style="flex:1; min-width:250px;">
                            <div style="font-size:10px; color:{COLORS['text_muted']};
                                        text-transform:uppercase; letter-spacing:0.5px;
                                        margin-bottom:6px;">
                                Receita do mês · realizado vs. meta
                            </div>
                            <div style="height:8px; background:{COLORS['surface_alt']};
                                        border-radius:4px; overflow:hidden; position:relative;">
                                <div style="width:{_pct_barra / 1.3:.0f}%; height:100%;
                                            background:{_cor_ritmo}; border-radius:4px;"></div>
                                <div style="position:absolute; left:76.9%; top:-3px; width:2px; height:14px;
                                            background:{COLORS['text_muted']};"></div>
                            </div>
                            <div style="display:flex; justify-content:space-between; margin-top:5px;
                                        font-size:11px; color:{COLORS['text_muted']};">
                                <span>Realizado <b style="color:{COLORS['text']};">{formata_m(rec_real_mes)}</b></span>
                                <span>Meta até hoje <b style="color:{COLORS['text']};">{formata_m(rec_orc_mes_prop)}</b></span>
                            </div>
                        </div>
                        <div style="display:flex; align-items:stretch; gap:20px;
                                    padding-left:20px; border-left:1px solid {COLORS['border']};">
                            <div style="text-align:right;">
                                <div style="font-size:10px; color:{COLORS['text_muted']};
                                            text-transform:uppercase; letter-spacing:0.5px;">
                                    Projeção do mês
                                </div>
                                <div style="font-size:17px; font-weight:800; color:{_cor_proj};
                                            margin-top:3px; line-height:1.1;">
                                    {formata_m(rec_projetada_mes)}
                                </div>
                                <div style="font-size:10.5px; color:{COLORS['text_muted']}; margin-top:2px;">
                                    orçado {formata_m(rec_orc_mes_cheio)}
                                </div>
                            </div>
                            <div style="text-align:right;">
                                <div style="font-size:10px; color:{COLORS['text_muted']};
                                            text-transform:uppercase; letter-spacing:0.5px;">
                                    Ritmo EBITDA
                                </div>
                                <div style="font-size:17px; font-weight:800; color:{_cor_ritmo_ebitda};
                                            margin-top:3px; line-height:1.1;">
                                    {ritmo_ebitda_pct:.0f}%
                                </div>
                                <div style="font-size:10.5px; color:{COLORS['text_muted']}; margin-top:2px;">
                                    do esperado
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                """),
                unsafe_allow_html=True,
            )

        st.caption(f"Visualização e Eficiência referente ao período: **{label_periodo_kpi}**")

        cg1, cg2 = st.columns(2)

        with cg1:
            st.markdown('<div class="section-title">Bridge de Performance (YTD)</div>', unsafe_allow_html=True)

            rec_bruta = get_valor_consolidado_multi(list_df_real, "1 - Receita Operacional Bruta", cols_kpi)
            deducoes = get_valor_consolidado_multi(list_df_real, "2 - Deduções da Receita Operacional Bruta", cols_kpi)
            rec_liq = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_kpi)

            cmv_bridge = get_valor_consolidado_multi(list_df_real, "4 - ", cols_kpi, exato_linha_sintetica=True)
            if cmv_bridge == 0:
                cmv_bridge = get_valor_consolidado_multi(list_df_real, "4 - Custo das Vendas", cols_kpi)

            margem_bruta = get_valor_consolidado_multi(list_df_real, "5 - Margem de Contribuição 1", cols_kpi)

            desp_var = get_valor_consolidado_multi(list_df_real, "6 - Despesas Variáveis", cols_kpi)
            if desp_var == 0:
                desp_var = get_valor_consolidado_multi(list_df_real, "Despesas Variáveis", cols_kpi)

            desp_op = get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_kpi)
            sga_total = desp_var + desp_op

            deprec = get_valor_consolidado_multi(list_df_real, "13 - Depreciação e Amortização", cols_kpi)

            ebitda = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)
            ebit = margem_bruta - abs(sga_total) - abs(deprec)

            base_rec = rec_liq if rec_liq != 0 else 1.0

            p_rec_bruta = round((rec_bruta / base_rec) * 100)
            p_deducoes = round((deducoes / base_rec) * 100)
            p_rec_liq = 100
            p_cmv = round((-abs(cmv_bridge) / base_rec) * 100)
            p_mb = round((margem_bruta / base_rec) * 100)
            p_sga = round((-abs(sga_total) / base_rec) * 100)

            p_ebit = round((ebit / base_rec) * 100)
            p_deprec = round((-abs(deprec) / base_rec) * 100)
            p_ebitda = round((ebitda / base_rec) * 100)

            x_bridge = [
                "Receita Bruta", "Deduções", "Receita Líquida", "CMV",
                "Margem Bruta", "SG&A", "EBIT", "D&A", "EBITDA",
            ]
            measures = ["absolute", "relative", "total", "relative", "total", "relative", "total", "relative", "total"]
            y_bridge = [p_rec_bruta, p_deducoes, 0, p_cmv, 0, p_sga, 0, abs(p_deprec), 0]
            text_labels = [
                f"{p_rec_bruta}%", f"{p_deducoes}%", f"{p_rec_liq}%",
                f"{p_cmv}%", f"{p_mb}%", f"{p_sga}%",
                f"{p_ebit}%", f"{p_deprec}%", f"{p_ebitda}%",
            ]

            fig_waterfall = go.Figure(
                go.Waterfall(
                    orientation="v",
                    measure=measures,
                    x=x_bridge,
                    y=y_bridge,
                    text=text_labels,
                    textposition="outside",
                    connector={"line": {"color": COLORS["border"], "width": 1}},
                    decreasing={"marker": {"color": COLORS["muted_line"]}},
                    increasing={"marker": {"color": COLORS["primary"]}},
                    totals={"marker": {"color": COLORS["secondary"]}},
                )
            )
            estilo_grafico(
                fig_waterfall,
                height=400,
                xaxis=dict(tickangle=-45, gridcolor="rgba(0,0,0,0)", fixedrange=True, automargin=True),
                yaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True),
            )
            st.plotly_chart(fig_waterfall, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        with cg2:
            st.markdown('<div class="section-title">Real vs. Orçado (YTD)</div>', unsafe_allow_html=True)

            cats = ["CMV", "TRF/REM", "Marg. Contrib. 2", "Despesas Fixas", "EBITDA"]

            NOME_LINHA_CMV_DETALHADO = "4.1 - Custo da Mercadoria Vendida - CMV"

            cmv_r = abs(get_valor_consolidado_multi(list_df_real, NOME_LINHA_CMV_DETALHADO, cols_kpi, exato_linha_sintetica=True))
            if cmv_r == 0:
                cmv_r = abs(get_valor_consolidado_multi(list_df_real, "4.1 - Custo da Mercadoria Vendida", cols_kpi))

            cmv_o = abs(get_valor_consolidado_multi(list_df_orc, NOME_LINHA_CMV_DETALHADO, cols_kpi, exato_linha_sintetica=True))
            if cmv_o == 0:
                cmv_o = abs(get_valor_consolidado_multi(list_df_orc, "4.1 - Custo da Mercadoria Vendida", cols_kpi))

            trf_r = abs(get_valor_consolidado_multi(list_df_real, "TRF/REM", cols_kpi))
            if trf_r == 0:
                trf_r = abs(get_valor_consolidado_multi(list_df_real, "TRF / REM", cols_kpi))

            trf_o = abs(get_valor_consolidado_multi(list_df_orc, "TRF/REM", cols_kpi))
            if trf_o == 0:
                trf_o = abs(get_valor_consolidado_multi(list_df_orc, "TRF / REM", cols_kpi))

            mc_r = get_valor_consolidado_multi(list_df_real, "Margem de Contribuição 2", cols_kpi)
            if mc_r == 0:
                mc_r = get_valor_consolidado_multi(list_df_real, "7 - Margem de Contribuição 2", cols_kpi)

            mc_o = get_valor_consolidado_multi(list_df_orc, "Margem de Contribuição 2", cols_kpi)
            if mc_o == 0:
                mc_o = get_valor_consolidado_multi(list_df_orc, "7 - Margem de Contribuição 2", cols_kpi)

            dfix_r = abs(get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_kpi))
            dfix_o = abs(get_valor_consolidado_multi(list_df_orc, "8 - Despesas Operacionais", cols_kpi))

            eb_r = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)
            eb_o = get_valor_consolidado_multi(list_df_orc, "11 - EBITDA", cols_kpi)

            val_r = [cmv_r, trf_r, mc_r, dfix_r, eb_r]
            val_o = [cmv_o, trf_o, mc_o, dfix_o, eb_o]

            labels_r = [formata_m(v) for v in val_r]
            labels_o = [formata_m(v) for v in val_o]

            fig_bar = go.Figure(
                data=[
                    go.Bar(name="Realizado (R$)", x=cats, y=val_r, text=labels_r, textposition="outside", marker_color=COLORS["primary"]),
                    go.Bar(name="Orçado (R$)", x=cats, y=val_o, text=labels_o, textposition="outside", marker_color=COLORS["secondary"]),
                ]
            )
            estilo_grafico(
                fig_bar,
                height=420,
                barmode="group",
                xaxis=dict(
                    gridcolor=COLORS["border"], zerolinecolor=COLORS["border"], fixedrange=True,
                    tickangle=-30, automargin=True,
                ),
                yaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True),
                legend=dict(orientation="h", yanchor="bottom", y=-0.32, xanchor="center", x=0.5),
                margin=dict(l=50, r=30, t=30, b=90),
            )
            st.plotly_chart(fig_bar, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        st.markdown("<br>", unsafe_allow_html=True)

        cg3, cg4 = st.columns([1.3, 0.7])

        with cg3:
            st.markdown('<div class="section-title">Evolução Mensal (Receita vs. Margem de Contribuição)</div>', unsafe_allow_html=True)

            rec_m, mc_m, rotulos_m = [], [], []
            for m_nome, c in m_map.items():
                v_rec = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", [c])
                v_mc = get_valor_consolidado_multi(list_df_real, "7 - Margem de Contribuição 2", [c])
                if v_mc == 0:
                    v_mc = get_valor_consolidado_multi(list_df_real, "Margem de Contribuição 2", [c])

                if v_rec != 0 or v_mc != 0:
                    rec_m.append(v_rec)
                    mc_m.append(v_mc)
                    rotulos_m.append(m_nome.capitalize())

            labels_rec = [formata_m(v) for v in rec_m]
            labels_mc = [formata_m(v) for v in mc_m]

            fig_line = go.Figure()
            fig_line.add_trace(
                go.Scatter(
                    x=rotulos_m, y=rec_m, mode="lines+markers+text", name="Receita (R$)",
                    text=labels_rec, textposition="top center",
                    line=dict(color=COLORS["primary"], width=2, shape="spline"),
                    marker=dict(size=6, color=COLORS["surface"], line=dict(color=COLORS["primary"], width=2)),
                    textfont=dict(color=COLORS["text"], size=11, family=FONT_STACK),
                )
            )
            fig_line.add_trace(
                go.Scatter(
                    x=rotulos_m, y=mc_m, mode="lines+markers+text", name="Margem Contrib. 2 (R$)",
                    text=labels_mc, textposition="bottom center",
                    line=dict(color=COLORS["muted_line"], width=2, shape="spline"),
                    marker=dict(size=6, color=COLORS["surface"], line=dict(color=COLORS["muted_line"], width=2)),
                    textfont=dict(color=COLORS["text_muted"], size=11, family=FONT_STACK),
                )
            )
            estilo_grafico(
                fig_line,
                height=380,
                xaxis=dict(showgrid=False, zeroline=False, tickangle=-45, tickfont=dict(size=11, color=COLORS["text_muted"]), fixedrange=True),
                yaxis=dict(showgrid=False, showticklabels=False, zeroline=False, fixedrange=True),
                legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="center", x=0.5, font=dict(color=COLORS["text_muted"])),
            )
            st.plotly_chart(fig_line, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        with cg4:
            st.markdown('<div class="section-title">Composição dos Custos & Saídas</div>', unsafe_allow_html=True)

            cmv_real_kpi = abs(get_valor_consolidado_multi(list_df_real, "4 - ", cols_kpi, exato_linha_sintetica=True))
            if cmv_real_kpi == 0:
                cmv_real_kpi = abs(get_valor_consolidado_multi(list_df_real, "4 - Custo das Vendas", cols_kpi))

            desp_op_real = abs(get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_kpi))

            v_cmv_pie = abs(cmv_real_kpi)
            v_desp_var_pie = abs(get_valor_consolidado_multi(list_df_real, "6 - Despesas Variáveis", cols_kpi))
            v_desp_op_pie = abs(desp_op_real)
            v_deprec_pie = abs(get_valor_consolidado_multi(list_df_real, "13 - Depreciação e Amortização", cols_kpi))

            total_pie = v_cmv_pie + v_desp_var_pie + v_desp_op_pie + v_deprec_pie

            fig_donut = go.Figure(
                data=[
                    go.Pie(
                        labels=["CMV / Custo", "Despesas Var.", "Despesas Op. (OpEx)", "Depreciação/Amort."],
                        values=[v_cmv_pie, v_desp_var_pie, v_desp_op_pie, v_deprec_pie],
                        hole=0.62,
                        marker=dict(colors=[COLORS["primary"], COLORS["muted_line"], COLORS["secondary"], COLORS["border_soft"]],
                                    line=dict(color=COLORS["surface"], width=2)),
                        textinfo="percent",
                        hoverinfo="label+value+percent",
                    )
                ]
            )
            fig_donut.add_annotation(
                text=f"<b>{formata_m(total_pie)}</b><br><span style='font-size:10px;color:{COLORS['text_muted']}'>Total Saídas</span>",
                showarrow=False, font=dict(color=COLORS["text"], size=13, family=FONT_STACK),
            )
            estilo_grafico(
                fig_donut,
                height=380,
                legend=dict(orientation="h", yanchor="top", y=-0.1, xanchor="center", x=0.5),
            )
            st.plotly_chart(fig_donut, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)


    # ---------------------------------------------------------------------------
    # ABA 2: DRE COMPLETA & DESVIOS
    # ---------------------------------------------------------------------------
with tab2:
    st.markdown(f'<div class="section-title">📋 Análise de DRE e Desvios — {label_visao} · {label_periodo_graf}</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    c1, _ = st.columns([2, 1])
    with c1:
        tipo_visao_dre = st.radio(
            "Filtro de Nível de Visão:",
            [
                "Apenas Grupos Principais (Sintética)",
                "Todas as Contas (Analítica)",
                "Visão Gerencial (Custos e Despesas)",
            ],
            # No Modo Departamento, a maioria das linhas já vem com "ponto"
            # (ex.: "6.24.1 - ..."), então a visão Sintética quase sempre
            # fica vazia -- por padrão, abre em Analítica nesse caso.
            index=1 if departamento_ativo else 0,
            horizontal=True,
        )
    st.caption("💡 Marque o checkbox de uma linha de grupo pra abrir (Sintética) ou fechar (Analítica/Gerencial) as contas dela.")

    if "grupos_dre_expandidos" not in st.session_state:
        st.session_state["grupos_dre_expandidos"] = set()
    if "grupos_dre_colapsados" not in st.session_state:
        st.session_state["grupos_dre_colapsados"] = set()

    col_nome = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
    linhas_dre_todas = list(df_ref[col_nome].dropna().astype(str).unique())
    if departamento_ativo and linhas_departamento_resolvidas:
        linhas_dre_todas = [l for l in linhas_dre_todas if l in linhas_departamento_resolvidas]

    is_sintetica_dre = tipo_visao_dre == "Apenas Grupos Principais (Sintética)"
    is_gerencial_dre = tipo_visao_dre == "Visão Gerencial (Custos e Despesas)"

    filtro_manual_dre_ativo = False
    if is_sintetica_dre:
        linhas_dre = _montar_linhas_com_expansao(
            linhas_dre_todas, "sintetica", st.session_state["grupos_dre_expandidos"]
        )
    elif is_gerencial_dre:
        linhas_dre_ger = [l for l in linhas_dre_todas if eh_linha_custos_despesas(l)]
        linhas_dre = _montar_linhas_com_expansao(
            linhas_dre_ger, "expandida", st.session_state["grupos_dre_colapsados"]
        )
    else:
        linhas_dre = _montar_linhas_com_expansao(
            linhas_dre_todas, "expandida", st.session_state["grupos_dre_colapsados"]
        )
        contas_filtradas_dre = st.multiselect(
            "🔍 Filtrar Contas Específicas (Estilo Excel):",
            options=linhas_dre_todas,
            default=[],
            key="filtro_contas_dre",
        )
        if contas_filtradas_dre:
            linhas_dre = contas_filtradas_dre
            filtro_manual_dre_ativo = True

    rec_liquida_real = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_graficos)
    rec_liquida_orc = get_valor_consolidado_multi(list_df_orc, "3 - Receita Operacional Liquida", cols_graficos)

    dados_dre = []
    for linha in linhas_dre:
        v_real = get_valor_consolidado_multi(list_df_real, linha, cols_graficos)
        v_orc = get_valor_consolidado_multi(list_df_orc, linha, cols_graficos)
        desvio_rs = v_real - v_orc

        av_real_pct = (v_real / rec_liquida_real * 100) if rec_liquida_real != 0 else 0.0
        av_orc_pct = (v_orc / rec_liquida_orc * 100) if rec_liquida_orc != 0 else 0.0
        ah_pct = (desvio_rs / abs(v_orc) * 100) if v_orc != 0 else 0.0

        dados_dre.append(
            {
                "Conta / Linha DRE": linha,
                "Realizado (R$)": v_real,
                "AV Real (%)": av_real_pct,
                "Orçado (R$)": v_orc,
                "AV Orçado (%)": av_orc_pct,
                "Desvio (R$)": desvio_rs,
                "AH (%)": ah_pct,
            }
        )

    df_dre_final = pd.DataFrame(dados_dre)

    # ---- KPIs contextuais de desvio ----
    if not df_dre_final.empty:
        n_favoravel = int((df_dre_final["Desvio (R$)"] > 0).sum())
        n_desfavoravel = int((df_dre_final["Desvio (R$)"] < 0).sum())
        idx_maior_desvio = df_dre_final["Desvio (R$)"].abs().idxmax()
        linha_maior_desvio = df_dre_final.loc[idx_maior_desvio, "Conta / Linha DRE"]
        valor_maior_desvio = df_dre_final.loc[idx_maior_desvio, "Desvio (R$)"]

        if departamento_ativo:
            # "Desvio EBITDA" é uma métrica de empresa toda -- sem sentido
            # pra um departamento isolado. Troca pelo desvio agregado do
            # próprio departamento (usando as linhas raiz, sem duplicar).
            card1_label = "DESVIO TOTAL DO DEPARTAMENTO"
            card1_icon = "📐"
            valor_desvio_card1 = sum(
                abs(get_valor_consolidado_multi(list_df_orc, l, cols_graficos, exato_linha_sintetica=True))
                - abs(get_valor_consolidado_multi(list_df_real, l, cols_graficos, exato_linha_sintetica=True))
                for l in linhas_departamento_raiz
            )
            card1_subtext = f"{label_periodo_graf} · positivo = gastou menos que o orçado"
        else:
            card1_label = "DESVIO EBITDA NO PERÍODO"
            card1_icon = "📐"
            valor_desvio_card1 = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_graficos) - \
                get_valor_consolidado_multi(list_df_orc, "11 - EBITDA", cols_graficos)
            card1_subtext = label_periodo_graf

        st.markdown(
            render_kpi_row([
                dict(label=card1_label, value=formata_brl(valor_desvio_card1),
                     value_color=cor_variacao(valor_desvio_card1), subtext=card1_subtext, icon=card1_icon),
                dict(label="CONTAS COM DESVIO FAVORÁVEL", value=str(n_favoravel), value_color=COLORS["positive"],
                     subtext=f"de {len(df_dre_final)} linhas analisadas", icon="✅"),
                dict(label="CONTAS COM DESVIO DESFAVORÁVEL", value=str(n_desfavoravel), value_color=COLORS["negative"],
                     subtext=f"de {len(df_dre_final)} linhas analisadas", icon="⚠️"),
                dict(label="MAIOR DESVIO INDIVIDUAL", value=formata_brl(valor_maior_desvio),
                     value_color=cor_variacao(valor_maior_desvio), subtext=linha_maior_desvio[:38], icon="🔎"),
            ]),
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

    if is_gerencial_dre:
        # Visão rápida e completa de custos e despesas: CMV, Despesas
        # Variáveis, Despesas Operacionais e o total dos três, cada um com
        # % sobre a Receita Líquida e o desvio vs. orçado -- é o resumo que
        # a visão Gerencial existe pra dar de cara, antes da tabela detalhada.
        rec_liq_ger = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_graficos)

        cmv_ger_r = abs(get_valor_consolidado_multi(list_df_real, "4 - ", cols_graficos, exato_linha_sintetica=True)) \
            or abs(get_valor_consolidado_multi(list_df_real, "4 - Custo das Vendas", cols_graficos))
        cmv_ger_o = abs(get_valor_consolidado_multi(list_df_orc, "4 - ", cols_graficos, exato_linha_sintetica=True)) \
            or abs(get_valor_consolidado_multi(list_df_orc, "4 - Custo das Vendas", cols_graficos))

        dvar_ger_r = abs(get_valor_consolidado_multi(list_df_real, "6 - Despesas Variáveis", cols_graficos))
        dvar_ger_o = abs(get_valor_consolidado_multi(list_df_orc, "6 - Despesas Variáveis", cols_graficos))

        dop_ger_r = abs(get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_graficos))
        dop_ger_o = abs(get_valor_consolidado_multi(list_df_orc, "8 - Despesas Operacionais", cols_graficos))

        total_ger_r = cmv_ger_r + dvar_ger_r + dop_ger_r
        total_ger_o = cmv_ger_o + dvar_ger_o + dop_ger_o

        def _pct_receita(valor):
            return (valor / rec_liq_ger * 100) if rec_liq_ger else 0.0

        st.markdown('<div class="section-title">💸 Resumo Gerencial — Custos e Despesas</div>', unsafe_allow_html=True)
        st.markdown(
            render_kpi_row([
                dict(label="CMV (CUSTO DAS VENDAS)", value=formata_brl(cmv_ger_r),
                     value_color=cor_variacao(cmv_ger_o - cmv_ger_r),
                     subtext=f"{_pct_receita(cmv_ger_r):.1f}% da receita líquida · Orçado: {formata_brl(cmv_ger_o)}", icon="📦"),
                dict(label="DESPESAS VARIÁVEIS", value=formata_brl(dvar_ger_r),
                     value_color=cor_variacao(dvar_ger_o - dvar_ger_r),
                     subtext=f"{_pct_receita(dvar_ger_r):.1f}% da receita líquida · Orçado: {formata_brl(dvar_ger_o)}", icon="📉"),
                dict(label="DESPESAS OPERACIONAIS", value=formata_brl(dop_ger_r),
                     value_color=cor_variacao(dop_ger_o - dop_ger_r),
                     subtext=f"{_pct_receita(dop_ger_r):.1f}% da receita líquida · Orçado: {formata_brl(dop_ger_o)}", icon="🏢"),
                dict(label="TOTAL CUSTOS + DESPESAS", value=formata_brl(total_ger_r),
                     value_color=cor_variacao(total_ger_o - total_ger_r),
                     subtext=f"{_pct_receita(total_ger_r):.1f}% da receita líquida · Orçado: {formata_brl(total_ger_o)}", icon="🧾"),
            ]),
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

    column_config_dre = {
        "Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large"),
    }

    ALTURA_17_LINHAS = 633
    cols_num_dre = ["Realizado (R$)", "AV Real (%)", "Orçado (R$)", "AV Orçado (%)", "Desvio (R$)", "AH (%)"]

    if df_dre_final.empty:
        # Nenhuma linha sobrou pra mostrar (ex.: a visão Sintética só mostra
        # linhas de grupo "sem ponto", e um recorte -- como o Modo
        # Departamento -- pode não ter nenhuma linha nesse formato). Sem essa
        # proteção, um DataFrame sem nenhuma linha também vem sem nenhuma
        # coluna, e o .style.format(...) quebra tentando achar colunas que
        # não existem.
        st.info(
            "Nenhuma linha da DRE para mostrar nessa combinação de filtros. "
            "Tente trocar para \"Todas as Contas (Analítica)\" ou revise o "
            "escopo/departamento selecionado."
        )
        evento_dre = None
    else:
        key_tabela_dre = (
            f"tabela_dre__{tipo_visao_dre}__{len(linhas_dre)}__"
            f"{','.join(sorted(st.session_state['grupos_dre_expandidos']))}__"
            f"{','.join(sorted(st.session_state['grupos_dre_colapsados']))}"
        )
        evento_dre = st.dataframe(
            df_dre_final.style.format(
                {
                    "Realizado (R$)": formata_brl,
                    "AV Real (%)": "{:.1f}%",
                    "Orçado (R$)": formata_brl,
                    "AV Orçado (%)": "{:.1f}%",
                    "Desvio (R$)": formata_brl,
                    "AH (%)": "{:.1f}%",
                }
            ).map(cor_valor, subset=cols_num_dre),
            column_config=column_config_dre,
            use_container_width=True,
            height=ALTURA_17_LINHAS,
            hide_index=True,
            on_select="rerun",
            selection_mode="multi-row",
            key=key_tabela_dre,
        )

    if not filtro_manual_dre_ativo:
        grupos_alvo_dre = (
            st.session_state["grupos_dre_expandidos"] if is_sintetica_dre
            else st.session_state["grupos_dre_colapsados"]
        )
        if _processar_clique_expansao(df_dre_final, evento_dre, grupos_alvo_dre):
            st.rerun()

    # ---- Linhas informativas (ex.: no MKT, a gestão GB da indústria) --
    # não fazem parte da gestão do departamento, mas ficam visíveis aqui só
    # pra conhecimento dos valores. ----
    if departamento_ativo and linhas_departamento_informativas:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div class="section-title">ℹ️ Outras Linhas Relacionadas (fora da gestão do departamento)</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Essas linhas aparecem na DRE dentro do mesmo grupo, mas não são geridas por este "
            "departamento -- estão aqui só para conhecimento dos valores."
        )
        linhas_tabela_info_dre = []
        for l in linhas_departamento_informativas:
            v_r_info_dre = abs(get_valor_consolidado_multi(list_df_real, l, cols_graficos, exato_linha_sintetica=True))
            v_o_info_dre = abs(get_valor_consolidado_multi(list_df_orc, l, cols_graficos, exato_linha_sintetica=True))
            linhas_tabela_info_dre.append({
                "Conta / Linha DRE": l, "Realizado (R$)": v_r_info_dre, "Orçado (R$)": v_o_info_dre,
                "Desvio (R$)": v_o_info_dre - v_r_info_dre,
            })
        st.dataframe(
            pd.DataFrame(linhas_tabela_info_dre).style.format(
                {"Realizado (R$)": formata_brl, "Orçado (R$)": formata_brl, "Desvio (R$)": formata_brl}
            ).map(cor_valor, subset=["Realizado (R$)", "Orçado (R$)", "Desvio (R$)"]),
            column_config={"Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large")},
            use_container_width=True,
            hide_index=True,
        )


# ---------------------------------------------------------------------------
# ABA 3: HISTÓRICO MENSAL
# ---------------------------------------------------------------------------
with tab3:
    st.markdown(f'<div class="section-title">📅 Histórico Mensal Mês a Mês — {label_visao}</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    ch1, ch2 = st.columns([1, 2.4])
    with ch1:
        tipo_hist = st.radio("Base de Dados:", ["Realizado", "Orçado"], horizontal=True)
    with ch2:
        visao_hist_dre = st.radio(
            "Detalhes:",
            [
                "Grupos Fechados (Sintético)",
                "Todas as Contas (Analítico)",
                "Visão Gerencial (Custos e Despesas)",
            ],
            index=1 if departamento_ativo else 0,
            horizontal=True,
        )
    st.caption("💡 Marque o checkbox de uma linha de grupo pra abrir (Sintético) ou fechar (Analítico/Gerencial) as contas dela.")

    if "grupos_hist_expandidos" not in st.session_state:
        st.session_state["grupos_hist_expandidos"] = set()
    if "grupos_hist_colapsados" not in st.session_state:
        st.session_state["grupos_hist_colapsados"] = set()

    linhas_hist_todas = list(df_ref[col_nome].dropna().astype(str).unique())
    if departamento_ativo and linhas_departamento_resolvidas:
        linhas_hist_todas = [l for l in linhas_hist_todas if l in linhas_departamento_resolvidas]
    is_sintetica_hist = visao_hist_dre == "Grupos Fechados (Sintético)"
    is_gerencial_hist = visao_hist_dre == "Visão Gerencial (Custos e Despesas)"

    filtro_manual_hist_ativo = False
    if is_sintetica_hist:
        linhas_hist = _montar_linhas_com_expansao(
            linhas_hist_todas, "sintetica", st.session_state["grupos_hist_expandidos"]
        )
    elif is_gerencial_hist:
        linhas_hist_ger = [l for l in linhas_hist_todas if eh_linha_custos_despesas(l)]
        linhas_hist = _montar_linhas_com_expansao(
            linhas_hist_ger, "expandida", st.session_state["grupos_hist_colapsados"]
        )
    else:
        linhas_hist = _montar_linhas_com_expansao(
            linhas_hist_todas, "expandida", st.session_state["grupos_hist_colapsados"]
        )
        contas_filtradas_hist = st.multiselect(
            "🔍 Filtrar Contas Específicas (Estilo Excel):",
            options=linhas_hist_todas,
            default=[],
            key="filtro_contas_hist",
        )
        if contas_filtradas_hist:
            linhas_hist = contas_filtradas_hist
            filtro_manual_hist_ativo = True

    target_dfs = list_df_real if tipo_hist == "Realizado" else list_df_orc

    # ---- KPIs contextuais do histórico (referência: Receita Operacional Líquida) ----
    if departamento_ativo:
        # Pra um departamento, a referência de "melhor/pior mês" é o custo
        # do próprio departamento (linhas raiz, sem duplicar) -- e "melhor"
        # aqui é o mês com o MENOR custo, igual na visão Gerencial.
        def _total_dept_mes(m_col):
            return abs(sum(
                get_valor_consolidado_multi(target_dfs, l, [m_col], exato_linha_sintetica=True)
                for l in linhas_departamento_raiz
            ))

        valores_ref_mensal = {m_nome: _total_dept_mes(m_col) for m_nome, m_col in m_map.items()}
        meses_com_dado = {m: v for m, v in valores_ref_mensal.items() if v != 0}

        if meses_com_dado:
            mes_menor_custo_dept = min(meses_com_dado, key=meses_com_dado.get)
            mes_maior_custo_dept = max(meses_com_dado, key=meses_com_dado.get)
            media_mensal_dept = sum(meses_com_dado.values()) / len(meses_com_dado)

            st.markdown(
                render_kpi_row([
                    dict(label=f"MENOR CUSTO MENSAL ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_menor_custo_dept]),
                         value_color=COLORS["positive"], subtext=mes_menor_custo_dept.capitalize(), icon="🏆"),
                    dict(label=f"MAIOR CUSTO MENSAL ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_maior_custo_dept]),
                         value_color=COLORS["negative"], subtext=mes_maior_custo_dept.capitalize(), icon="📈"),
                    dict(label="MÉDIA MENSAL DO DEPARTAMENTO", value=formata_brl(media_mensal_dept),
                         value_color=COLORS["text"], subtext=f"{len(meses_com_dado)} meses com dados", icon="📊"),
                    dict(label="AMPLITUDE (MAIOR - MENOR)", value=formata_brl(meses_com_dado[mes_maior_custo_dept] - meses_com_dado[mes_menor_custo_dept]),
                         value_color=COLORS["muted_line"], subtext="Variação entre extremos", icon="↕️"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)
    elif is_gerencial_hist:
        # Na visão Gerencial, a referência de "melhor/pior mês" é o total de
        # Custos + Despesas (CMV + Despesas Variáveis + Despesas
        # Operacionais) -- e, diferente da receita, aqui "melhor" é o mês
        # com o MENOR custo, não o maior.
        def _total_custos_mes(m_col):
            cmv_m = abs(get_valor_consolidado_multi(target_dfs, "4 - ", [m_col], exato_linha_sintetica=True)) \
                or abs(get_valor_consolidado_multi(target_dfs, "4 - Custo das Vendas", [m_col]))
            dvar_m = abs(get_valor_consolidado_multi(target_dfs, "6 - Despesas Variáveis", [m_col]))
            dop_m = abs(get_valor_consolidado_multi(target_dfs, "8 - Despesas Operacionais", [m_col]))
            return cmv_m + dvar_m + dop_m

        valores_ref_mensal = {m_nome: _total_custos_mes(m_col) for m_nome, m_col in m_map.items()}
        meses_com_dado = {m: v for m, v in valores_ref_mensal.items() if v != 0}

        if meses_com_dado:
            mes_menor_custo = min(meses_com_dado, key=meses_com_dado.get)
            mes_maior_custo = max(meses_com_dado, key=meses_com_dado.get)
            media_mensal_hist = sum(meses_com_dado.values()) / len(meses_com_dado)

            st.markdown(
                render_kpi_row([
                    dict(label=f"MENOR CUSTO MENSAL ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_menor_custo]),
                         value_color=COLORS["positive"], subtext=mes_menor_custo.capitalize(), icon="🏆"),
                    dict(label=f"MAIOR CUSTO MENSAL ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_maior_custo]),
                         value_color=COLORS["negative"], subtext=mes_maior_custo.capitalize(), icon="📈"),
                    dict(label="MÉDIA MENSAL (CUSTOS + DESPESAS)", value=formata_brl(media_mensal_hist),
                         value_color=COLORS["text"], subtext=f"{len(meses_com_dado)} meses com dados", icon="📊"),
                    dict(label="AMPLITUDE (MAIOR - MENOR)", value=formata_brl(meses_com_dado[mes_maior_custo] - meses_com_dado[mes_menor_custo]),
                         value_color=COLORS["muted_line"], subtext="Variação entre extremos", icon="↕️"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)
    else:
        valores_ref_mensal = {
            m_nome: get_valor_consolidado_multi(target_dfs, "3 - Receita Operacional Liquida", [m_col])
            for m_nome, m_col in m_map.items()
        }
        meses_com_dado = {m: v for m, v in valores_ref_mensal.items() if v != 0}

        if meses_com_dado:
            mes_melhor = max(meses_com_dado, key=meses_com_dado.get)
            mes_pior = min(meses_com_dado, key=meses_com_dado.get)
            media_mensal_hist = sum(meses_com_dado.values()) / len(meses_com_dado)

            st.markdown(
                render_kpi_row([
                    dict(label=f"MELHOR MÊS ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_melhor]),
                         value_color=COLORS["positive"], subtext=mes_melhor.capitalize(), icon="🏆"),
                    dict(label=f"PIOR MÊS ({tipo_hist.upper()})", value=formata_brl(meses_com_dado[mes_pior]),
                         value_color=COLORS["negative"], subtext=mes_pior.capitalize(), icon="📉"),
                    dict(label="MÉDIA MENSAL (RECEITA)", value=formata_brl(media_mensal_hist),
                         value_color=COLORS["text"], subtext=f"{len(meses_com_dado)} meses com dados", icon="📊"),
                    dict(label="AMPLITUDE (MELHOR - PIOR)", value=formata_brl(meses_com_dado[mes_melhor] - meses_com_dado[mes_pior]),
                         value_color=COLORS["muted_line"], subtext="Variação entre extremos", icon="↕️"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

    hist_data = []
    for linha in linhas_hist:
        row_dict = {"Conta / Linha DRE": linha}
        soma_linha = 0.0
        for m_nome, m_col in m_map.items():
            val_m = get_valor_consolidado_multi(target_dfs, linha, [m_col])
            row_dict[m_nome] = val_m
            soma_linha += val_m
        row_dict["Total Acumulado"] = soma_linha
        hist_data.append(row_dict)

    df_hist = pd.DataFrame(hist_data)

    col_config_hist = {
        "Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large", pinned=True),
    }

    colunas_numericas = list(m_map.keys()) + ["Total Acumulado"]
    format_dict_hist = {col: formata_brl for col in colunas_numericas}

    ALTURA_17_LINHAS = 633

    if df_hist.empty:
        st.info(
            "Nenhuma linha da DRE para mostrar nessa combinação de filtros. "
            "Tente trocar para \"Todas as Contas (Analítico)\" ou revise o "
            "escopo/departamento selecionado."
        )
        evento_hist = None
    else:
        key_tabela_hist = (
            f"tabela_hist__{visao_hist_dre}__{tipo_hist}__{len(linhas_hist)}__"
            f"{','.join(sorted(st.session_state['grupos_hist_expandidos']))}__"
            f"{','.join(sorted(st.session_state['grupos_hist_colapsados']))}"
        )
        evento_hist = st.dataframe(
            df_hist.style.format(format_dict_hist).map(cor_valor, subset=colunas_numericas),
            column_config=col_config_hist,
            use_container_width=True,
            height=ALTURA_17_LINHAS,
            hide_index=True,
            on_select="rerun",
            selection_mode="multi-row",
            key=key_tabela_hist,
        )

    if not filtro_manual_hist_ativo:
        grupos_alvo_hist = (
            st.session_state["grupos_hist_expandidos"] if is_sintetica_hist
            else st.session_state["grupos_hist_colapsados"]
        )
        if _processar_clique_expansao(df_hist, evento_hist, grupos_alvo_hist):
            st.rerun()

    # ---- Linhas informativas (ex.: no MKT, a gestão GB da indústria) --
    # não fazem parte da gestão do departamento, mas ficam visíveis aqui só
    # pra conhecimento dos valores -- no mesmo formato mensal da tabela
    # acima (colunas por mês + Total Acumulado), respeitando o mesmo
    # Realizado/Orçado escolhido no toggle "Base de Dados". ----
    if departamento_ativo and linhas_departamento_informativas:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div class="section-title">ℹ️ Outras Linhas Relacionadas (fora da gestão do departamento)</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Essas linhas aparecem na DRE dentro do mesmo grupo, mas não são geridas por este "
            "departamento -- estão aqui só para conhecimento dos valores."
        )
        hist_data_info = []
        for linha in linhas_departamento_informativas:
            row_dict_info = {"Conta / Linha DRE": linha}
            soma_linha_info = 0.0
            for m_nome, m_col in m_map.items():
                val_m_info = get_valor_consolidado_multi(target_dfs, linha, [m_col], exato_linha_sintetica=True)
                row_dict_info[m_nome] = val_m_info
                soma_linha_info += val_m_info
            row_dict_info["Total Acumulado"] = soma_linha_info
            hist_data_info.append(row_dict_info)

        df_hist_info = pd.DataFrame(hist_data_info)
        colunas_numericas_info = list(m_map.keys()) + ["Total Acumulado"]
        st.dataframe(
            df_hist_info.style.format({col: formata_brl for col in colunas_numericas_info}).map(
                cor_valor, subset=colunas_numericas_info
            ),
            column_config={"Conta / Linha DRE": st.column_config.TextColumn("Conta / Linha DRE", width="large", pinned=True)},
            use_container_width=True,
            hide_index=True,
        )


# ---------------------------------------------------------------------------
# ABA "IMPACTO & TENDÊNCIAS" (só no Modo Departamento)
# ---------------------------------------------------------------------------
if departamento_ativo:
    with tab4:
        st.markdown(
            f'<div class="section-title">🎯 Impacto & Tendências — {_nome_dept_abas}</div>',
            unsafe_allow_html=True,
        )
        if not linhas_departamento_resolvidas:
            st.warning(
                f"Não encontrei nenhuma linha da DRE atual que bata com o modelo de {_nome_dept_abas}."
            )
        else:
            st.caption(
                "Como os custos deste departamento pesam no resultado da empresa, e como eles "
                "vêm se comportando mês a mês ao longo do ano completo."
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ---- Simulação "e se": quanto o EBITDA melhoraria com um corte
            # de X% nos custos do departamento (no período selecionado na
            # barra lateral) ----
            rec_liq_impacto = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_kpi)
            ebitda_impacto = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)
            dept_real_impacto = abs(sum(
                get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True)
                for l in linhas_departamento_raiz
            ))
            margem_atual_impacto = (ebitda_impacto / rec_liq_impacto * 100) if rec_liq_impacto else 0.0

            st.markdown('<div class="section-title">🧪 Simulação: e se o departamento cortasse custos?</div>', unsafe_allow_html=True)
            pct_corte_simulado = st.slider(
                "Redução simulada nos custos do departamento (%)", min_value=0, max_value=30, value=10, step=5,
                help="Arraste para simular o efeito de uma redução de custos deste departamento sobre o EBITDA da empresa, no período selecionado na barra lateral.",
            )
            economia_simulada = dept_real_impacto * (pct_corte_simulado / 100)
            ebitda_simulado = ebitda_impacto + economia_simulada
            margem_simulada = (ebitda_simulado / rec_liq_impacto * 100) if rec_liq_impacto else 0.0

            st.markdown(
                render_kpi_row([
                    dict(label="ECONOMIA SIMULADA", value=formata_brl(economia_simulada),
                         value_color=COLORS["positive"], subtext=f"{pct_corte_simulado}% dos custos do departamento", icon="✂️"),
                    dict(label="EBITDA ATUAL → SIMULADO", value=f"{formata_m(ebitda_impacto)} → {formata_m(ebitda_simulado)}",
                         value_color=COLORS["text"], subtext="No período selecionado", icon="📐"),
                    dict(label="MARGEM EBITDA ATUAL → SIMULADA", value=f"{margem_atual_impacto:.1f}% → {margem_simulada:.1f}%",
                         value_color=COLORS["positive"], subtext=f"+{margem_simulada - margem_atual_impacto:.1f} p.p. de margem", icon="📊"),
                ]),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            col_rank_dept, col_desvio_dept = st.columns(2)
            with col_rank_dept:
                st.markdown('<div class="section-title">🏆 Ranking de Contas (Maiores Custos)</div>', unsafe_allow_html=True)
                linhas_ranking_dept = _linhas_composicao_do_conjunto(linhas_departamento_resolvidas)
                valores_rank = [
                    abs(get_valor_consolidado_multi(list_df_real, l, cols_kpi, exato_linha_sintetica=True))
                    for l in linhas_ranking_dept
                ]
                df_rank_dept = pd.DataFrame({
                    "Conta": [_nome_sem_numero_dre(l) for l in linhas_ranking_dept],
                    "Valor": valores_rank,
                }).sort_values("Valor", ascending=True).tail(8)
                fig_rank_dept = go.Figure(data=[go.Bar(
                    x=df_rank_dept["Valor"], y=df_rank_dept["Conta"], orientation="h",
                    marker_color=COLORS["primary"], text=[formata_m(v) for v in df_rank_dept["Valor"]],
                    textposition="outside", textfont=dict(color=COLORS["text_muted"], size=10),
                )])
                # Folga de 25% no eixo X pra o rótulo da maior barra (que
                # quase encosta na borda) não ficar cortado.
                maior_valor_rank = df_rank_dept["Valor"].max() if not df_rank_dept.empty else 0
                estilo_grafico(
                    fig_rank_dept, height=340,
                    xaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True,
                               range=[0, maior_valor_rank * 1.25 if maior_valor_rank else 1]),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickfont=dict(size=10)),
                    margin=dict(l=10, r=20, t=20, b=30),
                )
                st.plotly_chart(fig_rank_dept, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

            with col_desvio_dept:
                st.markdown('<div class="section-title">📐 Desvio Mensal vs. Orçado (Ano Completo)</div>', unsafe_allow_html=True)
                desvios_mensais = []
                for m_col in m_map.values():
                    v_r_mes = sum(get_valor_consolidado_multi(list_df_real, l, [m_col], exato_linha_sintetica=True) for l in linhas_departamento_raiz)
                    v_o_mes = sum(get_valor_consolidado_multi(list_df_orc, l, [m_col], exato_linha_sintetica=True) for l in linhas_departamento_raiz)
                    # positivo = gastou menos que o orçado (favorável)
                    desvios_mensais.append(abs(v_o_mes) - abs(v_r_mes))
                cores_desvio = [COLORS["positive"] if v >= 0 else COLORS["negative"] for v in desvios_mensais]
                fig_desvio_dept = go.Figure(data=[go.Bar(
                    x=list(m_map.keys()), y=desvios_mensais, marker_color=cores_desvio,
                )])
                estilo_grafico(
                    fig_desvio_dept, height=340,
                    xaxis=dict(gridcolor=COLORS["border"], fixedrange=True, tickfont=dict(size=9), automargin=True),
                    yaxis=dict(showticklabels=False, gridcolor="rgba(0,0,0,0)", fixedrange=True),
                )
                st.plotly_chart(fig_desvio_dept, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)
                st.caption("Barras verdes = departamento gastou menos que o orçado naquele mês. Vermelhas = gastou mais.")

            # ---- Insight automático: melhor/pior mês, tendência ----
            meses_com_valor = {
                nome: abs(sum(get_valor_consolidado_multi(list_df_real, l, [col], exato_linha_sintetica=True) for l in linhas_departamento_raiz))
                for nome, col in m_map.items()
            }
            meses_com_valor_positivo = {k: v for k, v in meses_com_valor.items() if v > 0}
            if meses_com_valor_positivo:
                mes_maior_gasto = max(meses_com_valor_positivo, key=meses_com_valor_positivo.get)
                mes_menor_gasto = min(meses_com_valor_positivo, key=meses_com_valor_positivo.get)
                st.markdown("<br>", unsafe_allow_html=True)
                # Escapa o "$" antes de passar pro st.info -- dois "R$" na
                # mesma string de markdown fazem o Streamlit tentar renderizar
                # tudo entre eles como fórmula (LaTeX), comendo o cifrão e
                # quebrando o **negrito** do meio.
                texto_insight_meses = (
                    f"📌 **Maior gasto do ano:** {mes_maior_gasto.capitalize()} ({formata_brl(meses_com_valor_positivo[mes_maior_gasto])}) · "
                    f"**Menor gasto do ano:** {mes_menor_gasto.capitalize()} ({formata_brl(meses_com_valor_positivo[mes_menor_gasto])})"
                ).replace("$", "\\$")
                st.info(texto_insight_meses)


# ---------------------------------------------------------------------------
# ABA: DIAGNÓSTICO EXECUTIVO (break-even, curva ABC, anomalias, forecast)
# ---------------------------------------------------------------------------
if not departamento_ativo and tab_diag is not None:
    with tab_diag:
        st.markdown(
            f'<div class="section-title">🔬 Diagnóstico Executivo — {label_visao} · {label_periodo_kpi}</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Leituras que exigiriam cruzar várias abas na mão: onde está o ponto de equilíbrio, "
            "quais contas concentram o gasto, o que fugiu do padrão histórico e onde o ano deve fechar."
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # Bases usadas pelo ponto de equilíbrio e pelo forecast.
        rec_liq_diag = get_valor_consolidado_multi(list_df_real, "3 - Receita Operacional Liquida", cols_kpi)
        ebitda_diag = get_valor_consolidado_multi(list_df_real, "11 - EBITDA", cols_kpi)

        # Universo de linhas de custo/despesa usado pela curva ABC e pela
        # detecção de anomalias.
        col_nome_diag = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
        linhas_diag = list(df_ref[col_nome_diag].dropna().astype(str).unique())
        linhas_custo_diag = [l for l in linhas_diag if eh_linha_custos_despesas(l)]
        linhas_custo_raiz_diag = _linhas_raiz_do_conjunto(linhas_custo_diag)
        # Para ABC e anomalias, o nível de topo (Custo das Vendas, Despesas
        # Operacionais, Despesas Variáveis) não serve: três contas gigantes
        # explicam tudo e a análise não aponta onde agir. As FOLHAS trazem as
        # contas de verdade (Salários, Energia, Frete...).
        linhas_custo_folha_diag = _linhas_folha_do_conjunto(linhas_custo_diag)

        # =====================================================================
        # 1. PONTO DE EQUILÍBRIO (BREAK-EVEN)
        # =====================================================================
        st.markdown('<div class="section-title">🎯 Ponto de Equilíbrio</div>', unsafe_allow_html=True)

        cmv_diag = abs(get_valor_consolidado_multi(list_df_real, "4 - ", cols_kpi, exato_linha_sintetica=True)) \
            or abs(get_valor_consolidado_multi(list_df_real, "4 - Custo das Vendas", cols_kpi))
        desp_var_diag = abs(get_valor_consolidado_multi(list_df_real, "6 - Despesas Variáveis", cols_kpi))
        desp_op_diag = abs(get_valor_consolidado_multi(list_df_real, "8 - Despesas Operacionais", cols_kpi))

        custos_variaveis_diag = cmv_diag + desp_var_diag
        custos_fixos_diag = desp_op_diag
        # Margem de contribuição = quanto sobra de cada real vendido depois
        # de pagar o que varia com a venda.
        mc_valor_diag = rec_liq_diag - custos_variaveis_diag
        mc_pct_diag = (mc_valor_diag / rec_liq_diag) if rec_liq_diag else 0
        ponto_equilibrio_diag = (custos_fixos_diag / mc_pct_diag) if mc_pct_diag > 0 else 0
        margem_seguranca_valor = rec_liq_diag - ponto_equilibrio_diag
        margem_seguranca_pct = (margem_seguranca_valor / rec_liq_diag * 100) if rec_liq_diag else 0
        lucro_operacional_diag = mc_valor_diag - custos_fixos_diag
        meses_periodo_be = max(len(cols_kpi), 1)

        if ponto_equilibrio_diag <= 0:
            st.warning(
                "Não foi possível calcular o ponto de equilíbrio: a margem de contribuição está zerada ou "
                "negativa no período (os custos variáveis consomem toda a receita)."
            )
        else:
            # ---- Números essenciais numa faixa enxuta ----
            # Quatro leituras bastam; as demais eram derivadas e poluíam.
            _cor_seg = COLORS["positive"] if margem_seguranca_pct > 0 else COLORS["negative"]
            _metricas_be = [
                ("Ponto de equilíbrio", formata_brl(ponto_equilibrio_diag), COLORS["text"],
                 f"{formata_m(ponto_equilibrio_diag / meses_periodo_be)} por mês"),
                ("Margem de contribuição", f"{mc_pct_diag * 100:.1f}%", COLORS["primary"],
                 f"R$ {mc_pct_diag:.2f} sobram de cada R$ 1,00"),
                ("Margem de segurança", f"{margem_seguranca_pct:.1f}%", _cor_seg,
                 f"{formata_m(abs(margem_seguranca_valor))} de folga"),
                ("Resultado operacional", formata_m(lucro_operacional_diag),
                 cor_variacao(lucro_operacional_diag), "Depois de fixos e variáveis"),
            ]
            _html_metricas = "".join(
                f"""
                <div style="flex:1; min-width:170px; padding:0 18px;
                            border-left:1px solid {COLORS['border']};">
                    <div style="font-size:10px; color:{COLORS['text_muted']};
                                text-transform:uppercase; letter-spacing:0.5px;">{rotulo}</div>
                    <div style="font-size:23px; font-weight:800; color:{cor};
                                margin-top:4px; line-height:1.15;">{valor}</div>
                    <div style="font-size:11.5px; color:{COLORS['text_muted']};
                                margin-top:4px;">{complemento}</div>
                </div>
                """
                for rotulo, valor, cor, complemento in _metricas_be
            )
            st.markdown(
                html_compacto(f"""
                <div style="display:flex; flex-wrap:wrap; gap:6px 0; padding:14px 4px;
                            border-top:1px solid {COLORS['border']};
                            border-bottom:1px solid {COLORS['border']}; margin-bottom:20px;">
                    {_html_metricas}
                </div>
                """),
                unsafe_allow_html=True,
            )

            # ---- Régua da receita: onde o dinheiro vai e onde vira lucro ----
            # Uma régua horizontal fina, com os marcos por fora do desenho.
            # Sem eixos, sem legenda flutuante, sem rótulo dentro da barra.
            pct_var = custos_variaveis_diag / rec_liq_diag * 100 if rec_liq_diag else 0
            pct_fix = custos_fixos_diag / rec_liq_diag * 100 if rec_liq_diag else 0
            pct_res = max(lucro_operacional_diag, 0) / rec_liq_diag * 100 if rec_liq_diag else 0
            pos_pe = ponto_equilibrio_diag / rec_liq_diag * 100 if rec_liq_diag else 0

            st.markdown(
                html_compacto(f"""
                <div style="margin:0 4px 6px 4px;">
                    <!-- marcador do ponto de equilíbrio: rótulo bem acima da
                         régua, com uma haste ligando os dois -->
                    <div style="position:relative; height:52px;">
                        <div style="position:absolute; left:{min(max(pos_pe, 8), 92):.1f}%;
                                    transform:translateX(-50%); text-align:center; white-space:nowrap;">
                            <div style="font-size:10px; color:{COLORS['text_muted']};
                                        text-transform:uppercase; letter-spacing:0.6px;">
                                ponto de equilíbrio
                            </div>
                            <div style="font-size:14px; color:{COLORS['text']}; font-weight:700;
                                        margin-top:1px;">
                                {formata_m(ponto_equilibrio_diag)}
                            </div>
                        </div>
                        <!-- haste vertical descendo até a régua -->
                        <div style="position:absolute; left:{pos_pe:.2f}%; bottom:0; width:1px; height:10px;
                                    background:{COLORS['text_muted']};"></div>
                    </div>
                    <!-- régua -->
                    <div style="position:relative; display:flex; height:38px;
                                border-radius:6px; overflow:hidden;
                                background:{COLORS['surface_alt']};">
                        <div style="width:{pct_var:.2f}%; background:rgba(247,110,110,0.55);"></div>
                        <div style="width:{pct_fix:.2f}%; background:rgba(245,166,35,0.55);"></div>
                        <div style="width:{pct_res:.2f}%; background:rgba(62,207,142,0.55);"></div>
                        <div style="position:absolute; left:{pos_pe:.2f}%; top:-5px; bottom:-5px;
                                    width:3px; background:{COLORS['text']};
                                    box-shadow:0 0 6px rgba(0,0,0,0.7);"></div>
                    </div>
                    <!-- marcos: início e receita total -->
                    <div style="display:flex; justify-content:space-between; margin-top:7px;
                                font-size:11px; color:{COLORS['text_muted']};">
                        <span>R$ 0</span>
                        <span>receita realizada <b style="color:{COLORS['text']};">
                            {formata_m(rec_liq_diag)}</b></span>
                    </div>
                    <!-- legenda: cada item num bloco próprio, com respiro -->
                    <div style="display:flex; flex-wrap:wrap; gap:10px; margin-top:16px;">
                        <div style="flex:1; min-width:190px; background:{COLORS['surface']};
                                    border:1px solid {COLORS['border']};
                                    border-left:3px solid {COLORS['negative']};
                                    border-radius:6px; padding:9px 13px;">
                            <div style="font-size:11px; color:{COLORS['text_muted']};">Custos variáveis</div>
                            <div style="font-size:15px; font-weight:700; color:{COLORS['text']}; margin-top:2px;">
                                {formata_m(custos_variaveis_diag)}
                                <span style="font-size:11.5px; font-weight:500;
                                             color:{COLORS['text_muted']};"> · {pct_var:.0f}% da receita</span>
                            </div>
                        </div>
                        <div style="flex:1; min-width:190px; background:{COLORS['surface']};
                                    border:1px solid {COLORS['border']};
                                    border-left:3px solid {COLORS['warning']};
                                    border-radius:6px; padding:9px 13px;">
                            <div style="font-size:11px; color:{COLORS['text_muted']};">Custos fixos</div>
                            <div style="font-size:15px; font-weight:700; color:{COLORS['text']}; margin-top:2px;">
                                {formata_m(custos_fixos_diag)}
                                <span style="font-size:11.5px; font-weight:500;
                                             color:{COLORS['text_muted']};"> · {pct_fix:.0f}% da receita</span>
                            </div>
                        </div>
                        <div style="flex:1; min-width:190px; background:{COLORS['surface']};
                                    border:1px solid {COLORS['border']};
                                    border-left:3px solid {COLORS['positive']};
                                    border-radius:6px; padding:9px 13px;">
                            <div style="font-size:11px; color:{COLORS['text_muted']};">Resultado</div>
                            <div style="font-size:15px; font-weight:700; color:{COLORS['text']}; margin-top:2px;">
                                {formata_m(lucro_operacional_diag)}
                                <span style="font-size:11.5px; font-weight:500;
                                             color:{COLORS['text_muted']};"> · {pct_res:.0f}% da receita</span>
                            </div>
                        </div>
                    </div>
                </div>
                """),
                unsafe_allow_html=True,
            )

            # ---- Uma frase de conclusão, sem caixa colorida ----
            _b_ini = f'<b style="color:{_cor_seg};">'
            if margem_seguranca_pct > 0:
                _texto_conclusao_html = (
                    f"A receita pode cair até {_b_ini}"
                    f"{formata_brl(margem_seguranca_valor).replace('$', '&#36;')}</b> "
                    f"({margem_seguranca_pct:.1f}%) antes da operação entrar em prejuízo."
                )
            else:
                _texto_conclusao_html = (
                    f"Faltam {_b_ini}"
                    f"{formata_brl(abs(margem_seguranca_valor)).replace('$', '&#36;')}</b> "
                    "de receita para cobrir os custos do período."
                )
            st.markdown(
                html_compacto(f"""
                <div style="margin-top:16px; padding:11px 15px; border-radius:6px;
                            background:{COLORS['surface']};
                            border-left:3px solid {_cor_seg};
                            font-size:13px; color:{COLORS['text']};">
                    {_texto_conclusao_html}
                </div>
                """),
                unsafe_allow_html=True,
            )
            st.caption(
                "Custos variáveis (CMV e despesas variáveis) acompanham a venda; custos fixos "
                "(despesas operacionais) existem independentemente dela."
            )
        st.markdown("<br>", unsafe_allow_html=True)

        # =====================================================================
        # 2. CURVA ABC DE DESPESAS
        # =====================================================================
        st.markdown('<div class="section-title">📊 Curva ABC de Custos e Despesas</div>', unsafe_allow_html=True)

        dados_abc = []
        for linha in linhas_custo_folha_diag:
            valor = abs(get_valor_consolidado_multi(list_df_real, linha, cols_kpi, exato_linha_sintetica=True))
            if valor > 0:
                dados_abc.append({"Conta": _nome_sem_numero_dre(linha), "Valor": valor})

        if not dados_abc:
            st.info("Sem custos ou despesas no período selecionado para montar a curva ABC.")
        else:
            df_abc = pd.DataFrame(dados_abc).sort_values("Valor", ascending=False).reset_index(drop=True)
            total_abc = df_abc["Valor"].sum()
            df_abc["% do total"] = df_abc["Valor"] / total_abc * 100
            df_abc["% acumulado"] = df_abc["% do total"].cumsum()
            df_abc["Classe"] = df_abc["% acumulado"].apply(
                lambda p: "A" if p <= 80 else ("B" if p <= 95 else "C")
            )
            # Custo é saída de caixa: na tabela detalhada aparece negativo.
            df_abc["Valor (R$)"] = -df_abc["Valor"]

            _cores_classe = {"A": COLORS["negative"], "B": COLORS["warning"], "C": COLORS["secondary"]}
            _info_classe = {
                "A": ("Prioridade", "Poucas contas, quase todo o gasto — é onde a gestão age"),
                "B": ("Acompanhar", "Peso intermediário, revisão periódica"),
                "C": ("Baixo impacto", "Muitas contas, pouco valor — não vale o esforço de gestão"),
            }

            qtd_a = int((df_abc["Classe"] == "A").sum())
            valor_a = df_abc.loc[df_abc["Classe"] == "A", "Valor"].sum()
            pct_valor_a = valor_a / total_abc * 100
            pct_contas_a = qtd_a / len(df_abc) * 100 if len(df_abc) else 0
            maior_conta = df_abc.iloc[0]

            # ---- Faixa de leitura, no mesmo padrão do ponto de equilíbrio ----
            _metricas_abc = [
                ("Concentração", f"{pct_valor_a:.0f}%", COLORS["negative"],
                 f"do gasto em {qtd_a} de {len(df_abc)} contas"),
                ("Maior conta", formata_m(maior_conta["Valor"]), COLORS["text"],
                 f"{maior_conta['Conta'][:30]} · {maior_conta['% do total']:.0f}%"),
                ("Gasto total analisado", formata_m(total_abc), COLORS["text"],
                 f"{len(df_abc)} contas no período"),
                ("Potencial de −5% na classe A", formata_m(valor_a * 0.05), COLORS["positive"],
                 "Economia se cortar 5% só nas prioritárias"),
            ]
            _html_abc = "".join(
                f"""
                <div style="flex:1; min-width:175px; padding:0 18px;
                            border-left:1px solid {COLORS['border']};">
                    <div style="font-size:10px; color:{COLORS['text_muted']};
                                text-transform:uppercase; letter-spacing:0.5px;">{rotulo}</div>
                    <div style="font-size:23px; font-weight:800; color:{cor};
                                margin-top:4px; line-height:1.15;">{valor}</div>
                    <div style="font-size:11.5px; color:{COLORS['text_muted']};
                                margin-top:4px;">{complemento}</div>
                </div>
                """
                for rotulo, valor, cor, complemento in _metricas_abc
            )
            st.markdown(
                html_compacto(f"""
                <div style="display:flex; flex-wrap:wrap; gap:6px 0; padding:14px 4px;
                            border-top:1px solid {COLORS['border']};
                            border-bottom:1px solid {COLORS['border']}; margin-bottom:20px;">
                    {_html_abc}
                </div>
                """),
                unsafe_allow_html=True,
            )

            # ---- Régua das classes: proporção de contas x proporção de gasto ----
            # O contraste entre as duas réguas é a mensagem da curva ABC:
            # poucas contas de um lado, quase todo o dinheiro do outro.
            _larguras_contas, _larguras_valor, _legenda_classes = [], [], []
            for classe in ["A", "B", "C"]:
                sub = df_abc[df_abc["Classe"] == classe]
                if sub.empty:
                    continue
                pct_qtd = len(sub) / len(df_abc) * 100
                pct_val = sub["Valor"].sum() / total_abc * 100
                cor_cl = _cores_classe[classe]
                _larguras_contas.append(
                    f'<div style="width:{pct_qtd:.2f}%; background:{cor_cl}; opacity:0.35;"></div>'
                )
                _larguras_valor.append(
                    f'<div style="width:{pct_val:.2f}%; background:{cor_cl}; opacity:0.75;"></div>'
                )
                titulo_cl, descricao_cl = _info_classe[classe]
                _legenda_classes.append(
                    f"""
                    <div style="flex:1; min-width:205px; background:{COLORS['surface']};
                                border:1px solid {COLORS['border']};
                                border-left:3px solid {cor_cl}; border-radius:6px; padding:9px 13px;">
                        <div style="font-size:11px; color:{COLORS['text_muted']};">
                            Classe {classe} · {titulo_cl}
                        </div>
                        <div style="font-size:15px; font-weight:700; color:{COLORS['text']}; margin-top:2px;">
                            {len(sub)} contas
                            <span style="font-size:11.5px; font-weight:500; color:{COLORS['text_muted']};">
                                · {formata_m(sub['Valor'].sum())} · {pct_val:.0f}% do gasto</span>
                        </div>
                        <div style="font-size:10.5px; color:{COLORS['text_muted']}; margin-top:3px;">
                            {descricao_cl}
                        </div>
                    </div>
                    """
                )

            st.markdown(
                html_compacto(f"""
                <div style="margin:0 4px 4px 4px;">
                    <div style="font-size:11px; color:{COLORS['text_muted']}; margin-bottom:5px;">
                        Proporção de <b style="color:{COLORS['text']};">contas</b>
                    </div>
                    <div style="display:flex; height:16px; border-radius:4px; overflow:hidden;
                                background:{COLORS['surface_alt']};">{''.join(_larguras_contas)}</div>
                    <div style="font-size:11px; color:{COLORS['text_muted']}; margin:14px 0 5px 0;">
                        Proporção do <b style="color:{COLORS['text']};">gasto</b>
                    </div>
                    <div style="display:flex; height:16px; border-radius:4px; overflow:hidden;
                                background:{COLORS['surface_alt']};">{''.join(_larguras_valor)}</div>
                    <div style="display:flex; flex-wrap:wrap; gap:10px; margin-top:16px;">
                        {''.join(_legenda_classes)}
                    </div>
                </div>
                """),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ---- Ranking das maiores contas ----
            st.markdown(
                f'<div style="font-size:12.5px; color:{COLORS["text_muted"]}; margin-bottom:8px;">'
                f'Maiores contas do período · a barra mostra o peso de cada uma no gasto total</div>',
                unsafe_allow_html=True,
            )
            qtd_ranking = min(10, len(df_abc))
            _linhas_ranking = []
            maior_valor_abc = float(df_abc["Valor"].max())
            for _, linha_abc in df_abc.head(qtd_ranking).iterrows():
                largura = linha_abc["Valor"] / maior_valor_abc * 100 if maior_valor_abc else 0
                cor_barra = _cores_classe[linha_abc["Classe"]]
                nome_conta = linha_abc["Conta"]
                nome_exibido = nome_conta if len(nome_conta) <= 42 else nome_conta[:41] + "…"
                _linhas_ranking.append(
                    f"""
                    <div style="display:flex; align-items:center; gap:12px; margin-bottom:7px;">
                        <div style="width:16px; text-align:center; font-size:10px; font-weight:700;
                                    color:{cor_barra};">{linha_abc['Classe']}</div>
                        <div style="width:250px; font-size:12px; color:{COLORS['text']};
                                    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;"
                             title="{nome_conta}">{nome_exibido}</div>
                        <div style="flex:1; background:{COLORS['surface_alt']}; border-radius:3px;
                                    height:16px; position:relative;">
                            <div style="width:{largura:.1f}%; height:100%; background:{cor_barra};
                                        opacity:0.6; border-radius:3px;"></div>
                        </div>
                        <div style="width:150px; text-align:right; font-size:12px;
                                    color:{COLORS['text']}; font-weight:600;">
                            {formata_m(linha_abc['Valor'])}
                            <span style="font-weight:400; color:{COLORS['text_muted']};
                                         font-size:11px;"> · {linha_abc['% do total']:.1f}%</span>
                        </div>
                    </div>
                    """
                )
            st.markdown(
                html_compacto(f'<div style="margin:0 4px;">{"".join(_linhas_ranking)}</div>'),
                unsafe_allow_html=True,
            )

            with st.expander(f"📋 Ver a classificação completa das {len(df_abc)} contas"):
                st.dataframe(
                    df_abc[["Conta", "Classe", "Valor (R$)", "% do total", "% acumulado"]].style.format({
                        "Valor (R$)": formata_brl,
                        "% do total": lambda v: f"{v:.1f}%".replace(".", ","),
                        "% acumulado": lambda v: f"{v:.1f}%".replace(".", ","),
                    }).map(cor_valor, subset=["Valor (R$)"]),
                    use_container_width=True, hide_index=True, height=420,
                )
            st.caption(
                f"A classificação segue a regra 80/15/5: **classe A** são as contas que somam os "
                f"primeiros 80% do gasto, **B** os 15% seguintes e **C** os últimos 5%. "
                f"Aqui, {pct_contas_a:.0f}% das contas concentram {pct_valor_a:.0f}% do dinheiro."
            )
        st.markdown("<br>", unsafe_allow_html=True)

        # =====================================================================
        # 3. DETECÇÃO AUTOMÁTICA DE ANOMALIAS
        # =====================================================================
        st.markdown('<div class="section-title">🚨 Anomalias Detectadas</div>', unsafe_allow_html=True)

        col_sens_a, col_sens_b = st.columns([1, 3])
        with col_sens_a:
            sensibilidade_anom = st.selectbox(
                "Sensibilidade:", ["Alta (1,5σ)", "Média (2σ)", "Baixa (3σ)"],
                index=1, key="diag_sensibilidade",
                help="Quanto menor o valor, mais alertas — e mais falsos positivos.",
            )
        limiar_sigma = {"Alta (1,5σ)": 1.5, "Média (2σ)": 2.0, "Baixa (3σ)": 3.0}[sensibilidade_anom]

        colunas_hist_anom = list(m_map.values())
        anomalias = []
        contas_avaliadas_anom = 0
        if len(colunas_hist_anom) >= 4:
            for linha in linhas_custo_folha_diag:
                serie_mensal = [
                    (nome, abs(get_valor_consolidado_multi(
                        list_df_real, linha, [col], exato_linha_sintetica=True)))
                    for nome, col in m_map.items()
                ]
                serie_com_dado = [(n, v) for n, v in serie_mensal if v > 0]
                if len(serie_com_dado) < 4:
                    continue
                contas_avaliadas_anom += 1
                # O último mês com movimento é comparado com os anteriores
                mes_anom, ultimo_valor = serie_com_dado[-1]
                anteriores = [v for _, v in serie_com_dado[:-1]]
                media_ant = sum(anteriores) / len(anteriores)
                variancia = sum((v - media_ant) ** 2 for v in anteriores) / len(anteriores)
                desvio_padrao = variancia ** 0.5
                if desvio_padrao == 0:
                    continue
                z = (ultimo_valor - media_ant) / desvio_padrao
                if abs(z) >= limiar_sigma:
                    anomalias.append({
                        "Conta": _nome_sem_numero_dre(linha),
                        "Mês": mes_anom.capitalize(),
                        "Valor do mês": ultimo_valor,
                        "Média anterior": media_ant,
                        "Variação": ultimo_valor - media_ant,
                        "Var. %": ((ultimo_valor - media_ant) / media_ant * 100) if media_ant else 0,
                        "Desvios": z,
                        "Historico": [v for _, v in serie_com_dado],
                    })

        if len(colunas_hist_anom) < 4:
            st.info(
                "São necessários pelo menos 4 meses de histórico para detectar anomalias com alguma "
                "confiança estatística. Com menos que isso, qualquer variação parece anormal."
            )
        elif not anomalias:
            st.success(
                f"✅ Nenhuma das {contas_avaliadas_anom} contas analisadas fugiu do próprio padrão "
                f"histórico na sensibilidade **{sensibilidade_anom}** — os gastos do último mês estão "
                "dentro do comportamento habitual."
            )
        else:
            df_anom = pd.DataFrame(anomalias)
            df_anom["Ordem"] = df_anom["Desvios"].abs()
            df_anom = df_anom.sort_values("Ordem", ascending=False).drop(columns=["Ordem"])
            n_altas = int((df_anom["Variação"] > 0).sum())
            n_baixas = int((df_anom["Variação"] < 0).sum())
            impacto_alta = df_anom.loc[df_anom["Variação"] > 0, "Variação"].sum()
            impacto_baixa = df_anom.loc[df_anom["Variação"] < 0, "Variação"].sum()

            # ---- Faixa de leitura, mesmo padrão das outras seções ----
            _metricas_anom = [
                ("Anomalias", str(len(df_anom)), COLORS["warning"],
                 f"de {contas_avaliadas_anom} contas com histórico"),
                ("Gastos acima do padrão", str(n_altas),
                 COLORS["negative"] if n_altas else COLORS["text_muted"],
                 f"{formata_m(impacto_alta)} acima da média" if n_altas else "nenhuma conta subiu fora da curva"),
                ("Gastos abaixo do padrão", str(n_baixas),
                 COLORS["positive"] if n_baixas else COLORS["text_muted"],
                 f"{formata_m(abs(impacto_baixa))} abaixo da média" if n_baixas else "nenhuma conta caiu fora da curva"),
                ("Efeito líquido", formata_m(impacto_alta + impacto_baixa),
                 cor_variacao(-(impacto_alta + impacto_baixa)),
                 "Soma dos desvios em relação ao normal"),
            ]
            _html_anom = "".join(
                f'<div style="flex:1; min-width:175px; padding:0 18px; '
                f'border-left:1px solid {COLORS["border"]};">'
                f'<div style="font-size:10px; color:{COLORS["text_muted"]}; '
                f'text-transform:uppercase; letter-spacing:0.5px;">{rotulo}</div>'
                f'<div style="font-size:23px; font-weight:800; color:{cor}; '
                f'margin-top:4px; line-height:1.15;">{valor}</div>'
                f'<div style="font-size:11.5px; color:{COLORS["text_muted"]}; '
                f'margin-top:4px;">{complemento}</div></div>'
                for rotulo, valor, cor, complemento in _metricas_anom
            )
            st.markdown(
                html_compacto(
                    f'<div style="display:flex; flex-wrap:wrap; gap:6px 0; padding:14px 4px; '
                    f'border-top:1px solid {COLORS["border"]}; '
                    f'border-bottom:1px solid {COLORS["border"]}; margin-bottom:20px;">{_html_anom}</div>'
                ),
                unsafe_allow_html=True,
            )

            # ---- Cada anomalia como um cartão com o histórico visível ----
            # Ver a série mensal ao lado do número explica na hora se aquilo
            # é um pico isolado ou uma mudança de patamar -- coisa que uma
            # tabela de números não mostra.
            _cartoes_anom = []
            for _, linha_anom in df_anom.head(12).iterrows():
                subiu = linha_anom["Variação"] > 0
                cor_anom = COLORS["negative"] if subiu else COLORS["positive"]
                seta = "▲" if subiu else "▼"
                historico = linha_anom["Historico"]
                maior_hist = max(historico) if historico else 1

                # Mini-gráfico do histórico em barras, com o último mês em destaque
                barras = "".join(
                    f'<div style="flex:1; height:{max(3, v / maior_hist * 26):.0f}px; '
                    f'background:{cor_anom if i == len(historico) - 1 else COLORS["secondary"]}; '
                    f'opacity:{1 if i == len(historico) - 1 else 0.45}; border-radius:1px;"></div>'
                    for i, v in enumerate(historico)
                )

                nome_conta_anom = linha_anom["Conta"]
                nome_exibido_anom = (
                    nome_conta_anom if len(nome_conta_anom) <= 38 else nome_conta_anom[:37] + "…"
                )
                _cartoes_anom.append(
                    f'<div style="flex:1; min-width:330px; background:{COLORS["surface"]}; '
                    f'border:1px solid {COLORS["border"]}; border-left:3px solid {cor_anom}; '
                    f'border-radius:6px; padding:11px 14px;">'
                    # cabeçalho: conta + mês
                    f'<div style="display:flex; justify-content:space-between; align-items:baseline; gap:10px;">'
                    f'<div style="font-size:12.5px; color:{COLORS["text"]}; font-weight:600; '
                    f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
                    f'title="{nome_conta_anom}">{nome_exibido_anom}</div>'
                    f'<div style="font-size:10.5px; color:{COLORS["text_muted"]}; white-space:nowrap;">'
                    f'{linha_anom["Mês"]}</div></div>'
                    # variação em destaque
                    f'<div style="display:flex; align-items:flex-end; gap:14px; margin-top:7px;">'
                    f'<div><div style="font-size:19px; font-weight:800; color:{cor_anom}; line-height:1.1;">'
                    f'{seta} {abs(linha_anom["Var. %"]):.0f}%</div>'
                    f'<div style="font-size:10.5px; color:{COLORS["text_muted"]}; margin-top:2px;">'
                    f'{formata_m(linha_anom["Valor do mês"])} vs. {formata_m(linha_anom["Média anterior"])} '
                    f'de média</div></div>'
                    # mini-histórico
                    f'<div style="flex:1; display:flex; align-items:flex-end; gap:2px; height:28px;">'
                    f'{barras}</div>'
                    f'<div style="text-align:right;">'
                    f'<div style="font-size:13px; font-weight:700; color:{cor_anom};">'
                    f'{linha_anom["Desvios"]:+.1f}σ</div>'
                    f'<div style="font-size:9.5px; color:{COLORS["text_muted"]};">do padrão</div>'
                    f'</div></div></div>'
                )

            st.markdown(
                html_compacto(
                    f'<div style="display:flex; flex-wrap:wrap; gap:10px; margin:0 4px;">'
                    f'{"".join(_cartoes_anom)}</div>'
                ),
                unsafe_allow_html=True,
            )
            if len(df_anom) > 12:
                st.caption(f"Mostrando as 12 anomalias mais fortes de {len(df_anom)} encontradas.")

            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander("📋 Ver todas as anomalias em tabela"):
                st.dataframe(
                    df_anom.drop(columns=["Historico"]).style.format({
                        "Valor do mês": formata_brl,
                        "Média anterior": formata_brl,
                        "Variação": formata_brl,
                        "Var. %": lambda v: f"{v:+.1f}%".replace(".", ","),
                        "Desvios": lambda v: f"{v:+.1f}σ".replace(".", ","),
                    }).map(cor_valor, subset=["Valor do mês", "Média anterior"])
                    .map(lambda v: f"color: {COLORS['negative']}" if v > 0 else f"color: {COLORS['positive']}",
                         subset=["Variação", "Var. %"]),
                    use_container_width=True, hide_index=True,
                    height=min(500, 60 + 35 * len(df_anom)),
                )

            with st.expander("❓ Como ler estes números"):
                st.markdown(
                    """
**A comparação é de cada conta consigo mesma**, não com o orçamento. O painel olha o histórico
mensal daquela conta, calcula o comportamento típico dela e verifica se o último mês fugiu desse
padrão.

**Média anterior** é quanto a conta costumava gastar por mês, considerando os meses *antes* do mês
sinalizado. É o "normal" dela — não uma meta.

**σ (sigma)** é o desvio-padrão: mede o quanto os valores daquela conta costumam variar em torno
da própria média. Conta estável tem sigma baixo; conta que oscila muito tem sigma alto. O número
em **σ** diz a quantos desvios-padrão o mês está da média:

- **±1σ** — variação comum, acontece na maioria dos meses
- **±2σ** — fora do padrão, cerca de 5% dos casos
- **±3σ ou mais** — bem improvável de ser só oscilação natural

Por isso o mesmo aumento em reais pode ser anomalia numa conta e rotina em outra — o que importa é
o quanto foge do comportamento *daquela* conta.

**As barrinhas** mostram o histórico mensal: a última barra (colorida) é o mês sinalizado. Elas
revelam se é um pico isolado ou uma mudança de patamar que vem se formando — e isso importa, porque
o método estatístico é bom em achar **picos**, mas tende a não sinalizar uma subida **gradual**
(os meses já elevados entram na média e "normalizam" o desvio). Nesses casos, o degrau aparece nas
barrinhas mesmo com sigma baixo.

⚠️ Anomalia não é sinônimo de erro — pode ser evento legítimo (compra sazonal, rescisão, campanha).
Ela indica **onde vale olhar primeiro**. Valor abaixo do padrão pode ser economia real ou lançamento
que ainda não entrou.
                    """
                )

        # =====================================================================
        # 4. FORECAST ROLLING — REPROJEÇÃO DO ANO
        # =====================================================================
        st.markdown('<div class="section-title">🔄 Forecast Rolling do Ano</div>', unsafe_allow_html=True)
        st.caption(
            "O orçamento foi feito antes do ano começar e vai perdendo aderência com o tempo. "
            "O forecast rolling substitui os meses que ainda não aconteceram por uma projeção baseada "
            "no que a operação vem entregando de verdade — e mostra onde o ano deve fechar."
        )

        col_fc1, col_fc2 = st.columns([1, 2])
        with col_fc1:
            base_forecast = st.selectbox(
                "Base da projeção:",
                ["Média dos últimos 3 meses", "Média dos últimos 6 meses", "Média de todo o realizado"],
                key="diag_base_forecast",
                help="Quantos meses realizados entram na média que projeta os meses restantes.",
            )
        meses_base_fc = {"Média dos últimos 3 meses": 3, "Média dos últimos 6 meses": 6,
                         "Média de todo o realizado": 99}[base_forecast]

        colunas_ano_fc = list(m_map.values())

        def _projetar_linha_fc(termo_linha, exato=False):
            """Realizado até agora + projeção dos meses restantes.
            Devolve (realizado, projetado, forecast_total, orcado_ano, meses_realizados)."""
            valores_mes = [
                get_valor_consolidado_multi(list_df_real, termo_linha, [col], exato_linha_sintetica=exato)
                for col in colunas_ano_fc
            ]
            # Mês "realizado" é aquele que tem movimento; os seguintes ainda
            # não aconteceram (ou não foram lançados).
            indices_com_dado = [i for i, v in enumerate(valores_mes) if v != 0]
            if not indices_com_dado:
                return 0.0, 0.0, 0.0, 0.0, 0
            ultimo_idx = max(indices_com_dado)
            realizados = valores_mes[: ultimo_idx + 1]
            meses_restantes = len(colunas_ano_fc) - (ultimo_idx + 1)

            base = [v for v in realizados[-meses_base_fc:] if v != 0] if meses_base_fc != 99 \
                else [v for v in realizados if v != 0]
            media_base = (sum(base) / len(base)) if base else 0.0

            total_realizado = sum(realizados)
            total_projetado = media_base * meses_restantes
            orcado_ano = get_valor_consolidado_multi(
                list_df_orc, termo_linha, colunas_ano_fc, exato_linha_sintetica=exato
            )
            return total_realizado, total_projetado, total_realizado + total_projetado, orcado_ano, ultimo_idx + 1

        linhas_forecast = [
            ("Receita Líquida", "3 - Receita Operacional Liquida", False),
            ("EBITDA", "11 - EBITDA", False),
        ]

        dados_fc = []
        meses_realizados_fc = 0
        for rotulo, termo, exato in linhas_forecast:
            real, proj, total, orc, n_meses = _projetar_linha_fc(termo, exato)
            meses_realizados_fc = max(meses_realizados_fc, n_meses)
            dados_fc.append({
                "Linha": rotulo,
                "Realizado até agora": real,
                "Projeção do restante": proj,
                "Forecast do ano": total,
                "Orçado do ano": orc,
                "Desvio vs. orçado": total - orc,
                "% do orçado": (total / orc * 100) if orc else 0,
            })

        if meses_realizados_fc == 0:
            st.info("Ainda não há meses realizados suficientes para projetar o fechamento do ano.")
        else:
            meses_restantes_fc = len(colunas_ano_fc) - meses_realizados_fc
            fc_receita = dados_fc[0]
            fc_ebitda = dados_fc[1]

            st.markdown(
                render_kpi_row([
                    dict(label="MESES JÁ REALIZADOS", value=f"{meses_realizados_fc} de {len(colunas_ano_fc)}",
                         value_color=COLORS["text"],
                         subtext=f"Faltam {meses_restantes_fc} meses para projetar", icon="📆"),
                    dict(label="FORECAST DE RECEITA", value=formata_m(fc_receita["Forecast do ano"]),
                         value_color=COLORS["primary"],
                         subtext=f"{fc_receita['% do orçado']:.0f}% do orçado do ano", icon="💰"),
                    dict(label="FORECAST DE EBITDA", value=formata_m(fc_ebitda["Forecast do ano"]),
                         value_color=cor_variacao(fc_ebitda["Desvio vs. orçado"]),
                         subtext=f"{fc_ebitda['% do orçado']:.0f}% do orçado do ano", icon="📈"),
                    dict(label="DESVIO PROJETADO DO EBITDA", value=formata_brl(fc_ebitda["Desvio vs. orçado"]),
                         value_color=cor_variacao(fc_ebitda["Desvio vs. orçado"]),
                         subtext="Forecast do ano − orçado do ano", icon="⚖️"),
                ]),
                unsafe_allow_html=True,
            )

            df_fc = pd.DataFrame(dados_fc)
            st.dataframe(
                df_fc.style.format({
                    "Realizado até agora": formata_brl,
                    "Projeção do restante": formata_brl,
                    "Forecast do ano": formata_brl,
                    "Orçado do ano": formata_brl,
                    "Desvio vs. orçado": formata_brl,
                    "% do orçado": lambda v: f"{v:.1f}%".replace(".", ","),
                }).map(cor_valor, subset=["Realizado até agora", "Projeção do restante",
                                          "Forecast do ano", "Orçado do ano", "Desvio vs. orçado"]),
                use_container_width=True, hide_index=True,
            )

            # Gráfico: realizado x projetado x orçado, mês a mês
            valores_ebitda_mes = [
                get_valor_consolidado_multi(list_df_real, "11 - EBITDA", [col]) for col in colunas_ano_fc
            ]
            valores_ebitda_orc = [
                get_valor_consolidado_multi(list_df_orc, "11 - EBITDA", [col]) for col in colunas_ano_fc
            ]
            realizados_ebitda = valores_ebitda_mes[:meses_realizados_fc]
            base_ebitda = [v for v in realizados_ebitda[-meses_base_fc:] if v != 0] if meses_base_fc != 99 \
                else [v for v in realizados_ebitda if v != 0]
            media_ebitda_fc = (sum(base_ebitda) / len(base_ebitda)) if base_ebitda else 0.0

            serie_realizada_graf = realizados_ebitda + [None] * meses_restantes_fc
            serie_projetada_graf = (
                [None] * (meses_realizados_fc - 1)
                + [realizados_ebitda[-1] if realizados_ebitda else 0]
                + [media_ebitda_fc] * meses_restantes_fc
            )
            rotulos_meses_fc = list(m_map.keys())

            fig_fc = go.Figure()
            fig_fc.add_trace(go.Scatter(
                name="Realizado", x=rotulos_meses_fc, y=serie_realizada_graf,
                mode="lines+markers", line=dict(color=COLORS["primary"], width=3),
                marker=dict(size=8), connectgaps=False,
                hovertemplate="Realizado: R$ %{y:,.2f}<extra></extra>",
            ))
            fig_fc.add_trace(go.Scatter(
                name="Projetado (forecast)", x=rotulos_meses_fc, y=serie_projetada_graf,
                mode="lines+markers", line=dict(color=COLORS["warning"], width=2.5, dash="dot"),
                marker=dict(size=7), connectgaps=False,
                hovertemplate="Projetado: R$ %{y:,.2f}<extra></extra>",
            ))
            fig_fc.add_trace(go.Scatter(
                name="Orçado", x=rotulos_meses_fc, y=valores_ebitda_orc,
                mode="lines", line=dict(color=COLORS["muted_line"], width=2, dash="dash"),
                hovertemplate="Orçado: R$ %{y:,.2f}<extra></extra>",
            ))
            # Faixa entre o pior e o melhor mês já realizado, aplicada aos
            # meses futuros: mostra a incerteza da projeção em vez de fingir
            # que o número é exato.
            if realizados_ebitda and meses_restantes_fc > 0:
                _validos_graf = [v for v in realizados_ebitda if v != 0]
                if _validos_graf:
                    _pior, _melhor = min(_validos_graf), max(_validos_graf)
                    banda_inf = [None] * (meses_realizados_fc - 1) + \
                        [realizados_ebitda[-1]] + [_pior] * meses_restantes_fc
                    banda_sup = [None] * (meses_realizados_fc - 1) + \
                        [realizados_ebitda[-1]] + [_melhor] * meses_restantes_fc
                    fig_fc.add_trace(go.Scatter(
                        name="Faixa de cenários", x=rotulos_meses_fc, y=banda_sup,
                        mode="lines", line=dict(width=0), showlegend=False,
                        hoverinfo="skip", connectgaps=False,
                    ))
                    fig_fc.add_trace(go.Scatter(
                        name="Faixa de cenários", x=rotulos_meses_fc, y=banda_inf,
                        mode="lines", line=dict(width=0), fill="tonexty",
                        fillcolor="rgba(245,166,35,0.12)", connectgaps=False,
                        hovertemplate="Faixa de cenários<extra></extra>",
                    ))
            estilo_grafico(
                fig_fc, height=380,
                xaxis=dict(gridcolor="rgba(0,0,0,0)", fixedrange=True, tickangle=-35,
                           automargin=True, tickfont=dict(size=9)),
                yaxis=dict(title=dict(text="EBITDA (R$)", font=dict(size=10, color=COLORS["text_muted"])),
                           gridcolor=COLORS["border"], fixedrange=True, tickformat=",.0f"),
                legend=dict(orientation="h", yanchor="bottom", y=-0.35, xanchor="center", x=0.5),
                margin=dict(l=80, r=30, t=25, b=95),
                hovermode="x unified",
            )
            st.plotly_chart(fig_fc, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

            # Cenários: além da média, mostra o que aconteceria no melhor e
            # no pior ritmo já observado -- uma projeção única passa falsa
            # sensação de precisão.
            if realizados_ebitda:
                meses_validos_fc = [v for v in realizados_ebitda if v != 0]
                if meses_validos_fc and meses_restantes_fc > 0:
                    pior_mes_fc = min(meses_validos_fc)
                    melhor_mes_fc = max(meses_validos_fc)
                    fc_pessimista = fc_ebitda["Realizado até agora"] + pior_mes_fc * meses_restantes_fc
                    fc_otimista = fc_ebitda["Realizado até agora"] + melhor_mes_fc * meses_restantes_fc
                    orcado_ano_fc = fc_ebitda["Orçado do ano"]

                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown('<div class="section-title">🎲 Cenários de Fechamento do EBITDA</div>',
                                unsafe_allow_html=True)
                    st.markdown(
                        render_kpi_row([
                            dict(label="CENÁRIO PESSIMISTA", value=formata_m(fc_pessimista),
                                 value_color=COLORS["negative"],
                                 subtext=f"Repetindo o pior mês ({formata_m(pior_mes_fc)}/mês)", icon="🔻"),
                            dict(label="CENÁRIO BASE", value=formata_m(fc_ebitda["Forecast do ano"]),
                                 value_color=COLORS["primary"],
                                 subtext=f"{base_forecast.lower()}", icon="🎯"),
                            dict(label="CENÁRIO OTIMISTA", value=formata_m(fc_otimista),
                                 value_color=COLORS["positive"],
                                 subtext=f"Repetindo o melhor mês ({formata_m(melhor_mes_fc)}/mês)", icon="🔺"),
                            dict(label="ORÇADO DO ANO", value=formata_m(orcado_ano_fc),
                                 value_color=COLORS["text_muted"],
                                 subtext=("Dentro da faixa de cenários"
                                          if fc_pessimista <= orcado_ano_fc <= fc_otimista
                                          else "Fora da faixa projetada"), icon="📋"),
                        ]),
                        unsafe_allow_html=True,
                    )
                    if orcado_ano_fc > fc_otimista:
                        st.error(
                            "🚨 O orçamento do ano está **acima do melhor mês já entregue** repetido até "
                            "dezembro. Sem uma mudança relevante na operação, a meta não é alcançável — "
                            "vale rediscutir o orçamento ou o plano de ação."
                        )
                    elif orcado_ano_fc > fc_ebitda["Forecast do ano"]:
                        esforco_mensal = (orcado_ano_fc - fc_ebitda["Realizado até agora"]) / meses_restantes_fc
                        media_atual_fc = media_ebitda_fc
                        aumento_necessario = (
                            (esforco_mensal / media_atual_fc - 1) * 100 if media_atual_fc else 0
                        )
                        st.warning(
                            f"⚠️ Para bater a meta, seria preciso entregar {formata_brl(esforco_mensal)} "
                            f"de EBITDA por mês nos {meses_restantes_fc} meses restantes — "
                            f"{aumento_necessario:+.0f}% em relação à média atual de "
                            f"{formata_brl(media_atual_fc)}.".replace("$", "\\$")
                        )

            desvio_fc = fc_ebitda["Desvio vs. orçado"]
            if desvio_fc < 0:
                st.warning(
                    f"⚠️ Mantido o ritmo atual, o ano deve fechar **{formata_brl(abs(desvio_fc))} abaixo** "
                    f"do EBITDA orçado ({fc_ebitda['% do orçado']:.0f}% da meta). Para chegar ao orçado, "
                    f"seria preciso gerar {formata_brl(abs(desvio_fc) / max(meses_restantes_fc, 1))} a mais "
                    f"por mês nos {meses_restantes_fc} meses restantes.".replace("$", "\\$")
                )
            else:
                st.success(
                    f"✅ Mantido o ritmo atual, o ano deve fechar **{formata_brl(desvio_fc)} acima** do "
                    f"EBITDA orçado ({fc_ebitda['% do orçado']:.0f}% da meta).".replace("$", "\\$")
                )
            st.caption(
                "A projeção repete a média dos meses escolhidos para os meses que faltam. É uma leitura de "
                "tendência, não uma previsão com sazonalidade — se o negócio tem meses tipicamente mais "
                "fortes ou fracos à frente, considere isso na análise."
            )


# ---------------------------------------------------------------------------
# ABA 4: PREVISÕES & TRENDS
# ---------------------------------------------------------------------------
if not departamento_ativo:
    with tab4:
        st.markdown('<div class="section-title">🔮 Painel Avançado de Previsões e Tendências</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        c_f1, c_f2, c_f3 = st.columns([1.2, 1.2, 1.6])

        with c_f1:
            metrica_sel = st.selectbox("Métrica de Análise:", ["Receita Operacional Líquida", "EBITDA"], index=0)
            termo_metrica = "3 - Receita Operacional Liquida" if metrica_sel == "Receita Operacional Líquida" else "11 - EBITDA"

        with c_f2:
            modelo_proj = st.selectbox(
                "Modelo de Projeção Futura:",
                ["Média Histórica (Run-Rate)", "Manter Orçamento (Budget)", "Ajustado por Sazonalidade/Performance"],
                index=0,
                help="Define como os meses futuros não realizados serão calculados.",
            )

        with c_f3:
            sensibilidade = st.slider(
                "Ajuste Fino de Cenário / Estresse (%):",
                min_value=-20.0, max_value=20.0, value=0.0, step=1.0,
                help="Aplica uma variação percentual sobre os meses projetados.",
            )

        meses_todos = list(m_map.keys())

        meses_realizados_cols = (
            cols_kpi if tipo_periodo == "Mês Selecionado"
            else [m_map[m] for m in meses_todos if get_valor_consolidado_multi(list_df_real, termo_metrica, [m_map[m]]) != 0]
        )

        if not meses_realizados_cols:
            meses_realizados_cols = list(m_map.values())[:7]

        num_meses_realizados = len(meses_realizados_cols)

        val_real_acumulado = get_valor_consolidado_multi(list_df_real, termo_metrica, meses_realizados_cols)
        val_orc_acumulado = get_valor_consolidado_multi(list_df_orc, termo_metrica, meses_realizados_cols)

        val_orc_anual_total = get_valor_consolidado_multi(list_df_orc, termo_metrica, colunas_validas)

        media_mensal_real = val_real_acumulado / num_meses_realizados if num_meses_realizados > 0 else 0
        fator_performance = (val_real_acumulado / val_orc_acumulado) if val_orc_acumulado != 0 else 1.0

        fator_sensibilidade = 1.0 + (sensibilidade / 100.0)

        valores_finais_mes = []
        valores_orcado_mes = []
        valores_real_mes = []
        tipos_serie = []

        for idx_m, m_nome in enumerate(meses_todos):
            m_col = m_map[m_nome]
            v_orc = get_valor_consolidado_multi(list_df_orc, termo_metrica, [m_col])
            valores_orcado_mes.append(v_orc)

            if idx_m < num_meses_realizados:
                v_real = get_valor_consolidado_multi(list_df_real, termo_metrica, [m_col])
                valores_finais_mes.append(v_real)
                valores_real_mes.append(v_real)
                tipos_serie.append("Realizado")
            else:
                if modelo_proj == "Média Histórica (Run-Rate)":
                    v_proj = media_mensal_real * fator_sensibilidade
                elif modelo_proj == "Manter Orçamento (Budget)":
                    v_proj = v_orc * fator_sensibilidade
                else:
                    v_proj = (v_orc * fator_performance) * fator_sensibilidade

                valores_finais_mes.append(v_proj)
                valores_real_mes.append(np.nan)
                tipos_serie.append("Projetado")

        projecao_total_anual = sum(valores_finais_mes)
        diff_anual = projecao_total_anual - val_orc_anual_total
        pct_atingimento_anual = (projecao_total_anual / val_orc_anual_total * 100) if val_orc_anual_total != 0 else 0

        c_proj = cor_variacao(projecao_total_anual)
        c_diff = cor_variacao(diff_anual)

        st.markdown("<br>", unsafe_allow_html=True)

        subtext_gap = f"{diff_anual / abs(val_orc_anual_total) * 100:+.1f}% vs Meta" if val_orc_anual_total != 0 else "—"

        st.markdown(
            render_kpi_row([
                dict(label=f"PROJEÇÃO ANUAL ({metrica_sel.upper()})", value=formata_brl(projecao_total_anual),
                     value_color=c_proj, subtext=f"Cenário: {modelo_proj.split(' ')[0]}", icon="🔮"),
                dict(label="META ANUAL (ORÇADO)", value=formata_brl(val_orc_anual_total), value_color=COLORS["text"],
                     subtext="Orçamento Fechado", icon="🎯"),
                dict(label="GAP / DESVIO ANUAL", value=formata_brl(diff_anual), value_color=c_diff,
                     subtext=subtext_gap, subtext_color=c_diff, icon="📐"),
                dict(label="ATINGIMENTO ESTIMADO", value=f"{pct_atingimento_anual:.1f}%", value_color=c_diff,
                     subtext=f"Média Mensal Real: {formata_m(media_mensal_real)}", icon="📊"),
            ]),
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)

        df_trend = pd.DataFrame({
            "Mês": [m.capitalize() for m in meses_todos],
            "Valor Projetado/Real": valores_finais_mes,
            "Orçado": valores_orcado_mes,
            "Tipo": tipos_serie,
        })

        posicoes_meta = []
        posicoes_barras = []
        for val_p, val_o in zip(valores_finais_mes, valores_orcado_mes):
            if val_p >= val_o:
                posicoes_meta.append("bottom center")
                posicoes_barras.append("outside")
            else:
                posicoes_meta.append("top center")
                posicoes_barras.append("inside")

        fig_comb = go.Figure()

        df_real_bar = df_trend[df_trend["Tipo"] == "Realizado"]
        pos_bar_real = [posicoes_barras[i] for i in df_real_bar.index]
        fig_comb.add_trace(
            go.Bar(
                x=df_real_bar["Mês"], y=df_real_bar["Valor Projetado/Real"], name="Realizado",
                marker_color=COLORS["primary"],
                text=[formata_m(v) for v in df_real_bar["Valor Projetado/Real"]],
                textposition=pos_bar_real, textfont=dict(size=11, color=COLORS["text"]),
                cliponaxis=False,
            )
        )

        df_proj_bar = df_trend[df_trend["Tipo"] == "Projetado"]
        pos_bar_proj = [posicoes_barras[i] for i in df_proj_bar.index]
        fig_comb.add_trace(
            go.Bar(
                x=df_proj_bar["Mês"], y=df_proj_bar["Valor Projetado/Real"], name="Projetado (Tendência)",
                marker_color=COLORS["border_soft"],
                text=[formata_m(v) for v in df_proj_bar["Valor Projetado/Real"]],
                textposition=pos_bar_proj, textfont=dict(size=11, color=COLORS["text_muted"]),
                cliponaxis=False,
            )
        )

        fig_comb.add_trace(
            go.Scatter(
                x=df_trend["Mês"], y=df_trend["Orçado"], name="Orçado (Meta)",
                mode="lines+markers+text",
                text=[formata_m(v) for v in df_trend["Orçado"]],
                textposition=posicoes_meta, textfont=dict(size=10, color=COLORS["warning"]),
                line=dict(color=COLORS["warning"], width=2, dash="dash"),
                marker=dict(size=6, color=COLORS["warning"]),
                cliponaxis=False,
            )
        )

        max_val = max(
            max(df_trend["Valor Projetado/Real"].dropna(), default=0),
            max(df_trend["Orçado"].dropna(), default=0),
        )

        estilo_grafico(
            fig_comb,
            height=500,
            title=f"Evolução Mensal & Projeção Run-Rate: {metrica_sel}",
            xaxis=dict(gridcolor=COLORS["border"], zerolinecolor=COLORS["border"], fixedrange=True),
            yaxis=dict(
                showticklabels=False, gridcolor="rgba(0,0,0,0)",
                range=[0, max_val * 1.35] if max_val > 0 else None,
                fixedrange=True,
            ),
            legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
            barmode="group",
        )
        st.plotly_chart(fig_comb, use_container_width=True, config=CONFIG_PLOTLY_TRAVADO)

        st.markdown('<div class="section-title">📋 Detalhamento da Projeção Mensal (R$)</div>', unsafe_allow_html=True)

        df_resumo_proj = pd.DataFrame({
            "Mês": [m.capitalize() for m in meses_todos],
            "Tipo de Dado": tipos_serie,
            "Valor Realizado / Projetado": valores_finais_mes,
            "Orçado Original": valores_orcado_mes,
            "Desvio (R$)": [v_p - v_o for v_p, v_o in zip(valores_finais_mes, valores_orcado_mes)],
        })

        ALTURA_12_LINHAS = 38 + len(df_resumo_proj) * 35

        st.dataframe(
            df_resumo_proj.style.format({
                "Valor Realizado / Projetado": formata_brl,
                "Orçado Original": formata_brl,
                "Desvio (R$)": formata_brl,
            }).map(cor_valor, subset=["Desvio (R$)"]),
            use_container_width=True,
            hide_index=True,
            height=ALTURA_12_LINHAS,
        )

        # -----------------------------------------------------------------------
        # Estresse por Linha da DRE — impacto em cascata
        # -----------------------------------------------------------------------
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div class="section-title">🧪 Estresse de Cenário por Linha da DRE — Impacto em Cascata</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Escolha uma linha da DRE e aplique uma variação percentual para ver o impacto direto no "
            "EBITDA (e na Receita Líquida, se a linha escolhida fizer parte dela). Se outras linhas devem "
            "se mover junto (ex.: um custo variável que sobe quando as vendas sobem), marque-as em "
            "\"Linhas dependentes\" -- elas recebem a MESMA variação percentual da linha principal."
        )

        col_nome_stress = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
        todas_linhas_dre = df_ref[col_nome_stress].dropna().astype(str).unique().tolist()

        ce1, ce2 = st.columns([1.3, 1.3])
        with ce1:
            idx_padrao_stress = (
                todas_linhas_dre.index("3 - Receita Operacional Liquida")
                if "3 - Receita Operacional Liquida" in todas_linhas_dre else 0
            )
            linha_estresse_sel = st.selectbox(
                "Linha da DRE a estressar:", todas_linhas_dre, index=idx_padrao_stress, key="linha_estresse_sel",
            )
            pct_estresse_linha = st.slider(
                f'Variação em "{linha_estresse_sel}" (%):',
                min_value=-50.0, max_value=50.0, value=5.0, step=1.0, key="pct_estresse_linha",
            )
        with ce2:
            linhas_dependentes = st.multiselect(
                "Linhas dependentes (variam junto, mesma variação %):",
                [l for l in todas_linhas_dre if l != linha_estresse_sel],
                key="linhas_dependentes_estresse",
                help='Ex.: se estressar "Vendas", marque aqui os custos variáveis que sobem/descem junto.',
            )

        periodo_estresse = colunas_validas  # ano completo com dados válidos no escopo atual

        def _valor_linha_stress(termo):
            return get_valor_consolidado_multi(list_df_real, termo, periodo_estresse, exato_linha_sintetica=True)

        LINHAS_RECEITA_STRESS = {"1 - Receita Operacional Bruta", "3 - Receita Operacional Liquida"}

        valor_original_linha = _valor_linha_stress(linha_estresse_sel)
        valor_estressado_linha = valor_original_linha * (1 + pct_estresse_linha / 100.0)
        delta_linha_principal = valor_estressado_linha - valor_original_linha

        delta_total_ebitda = delta_linha_principal
        delta_total_receita = delta_linha_principal if linha_estresse_sel in LINHAS_RECEITA_STRESS else 0.0

        detalhes_dependentes = []
        for linha_dep in linhas_dependentes:
            valor_dep = _valor_linha_stress(linha_dep)
            valor_dep_novo = valor_dep * (1 + pct_estresse_linha / 100.0)
            delta_dep = valor_dep_novo - valor_dep
            delta_total_ebitda += delta_dep
            if linha_dep in LINHAS_RECEITA_STRESS:
                delta_total_receita += delta_dep
            detalhes_dependentes.append((linha_dep, valor_dep, valor_dep_novo, delta_dep))

        ebitda_original_stress = _valor_linha_stress("11 - EBITDA")
        ebitda_novo_stress = ebitda_original_stress + delta_total_ebitda

        rec_liq_original_stress = _valor_linha_stress("3 - Receita Operacional Liquida")
        rec_liq_novo_stress = rec_liq_original_stress + delta_total_receita

        margem_original_stress = (ebitda_original_stress / rec_liq_original_stress * 100) if rec_liq_original_stress else 0
        margem_nova_stress = (ebitda_novo_stress / rec_liq_novo_stress * 100) if rec_liq_novo_stress else 0

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            render_kpi_row([
                dict(label=f'"{linha_estresse_sel[:28]}" — Original', value=formata_brl(valor_original_linha),
                     value_color=COLORS["text"], subtext="Valor realizado no período", icon="📍"),
                dict(label=f'"{linha_estresse_sel[:28]}" — Estressado', value=formata_brl(valor_estressado_linha),
                     value_color=cor_variacao(delta_linha_principal), subtext=f"{pct_estresse_linha:+.1f}% aplicado", icon="🧪"),
                dict(label="EBITDA — Impacto do Cenário", value=formata_brl(ebitda_novo_stress),
                     value_color=cor_variacao(delta_total_ebitda),
                     subtext=f"Original: {formata_brl(ebitda_original_stress)} · Delta: {formata_brl(delta_total_ebitda)}", icon="💹"),
                dict(label="Margem EBITDA — Impacto do Cenário", value=f"{margem_nova_stress:.1f}%",
                     value_color=cor_variacao(margem_nova_stress - margem_original_stress),
                     subtext=f"Original: {margem_original_stress:.1f}%", icon="📊"),
            ]),
            unsafe_allow_html=True,
        )

        if detalhes_dependentes:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="section-title">🔗 Linhas Dependentes — Detalhe do Impacto</div>', unsafe_allow_html=True)
            df_dependentes_stress = pd.DataFrame(
                [(l, v_o, v_n, d) for l, v_o, v_n, d in detalhes_dependentes],
                columns=["Linha da DRE", "Valor Original", "Valor Estressado", "Delta (R$)"],
            )
            st.dataframe(
                df_dependentes_stress.style.format({
                    "Valor Original": formata_brl, "Valor Estressado": formata_brl, "Delta (R$)": formata_brl,
                }).map(cor_valor, subset=["Delta (R$)"]),
                use_container_width=True, hide_index=True,
                height=38 + len(df_dependentes_stress) * 35,
            )

        st.caption(
            "⚠️ Este cenário assume que todas as demais linhas da DRE permanecem constantes, exceto a linha "
            "estressada e as linhas dependentes marcadas -- é uma simulação isolada (\"what-if\"), não uma "
            "reprojeção completa do orçamento."
        )

# ---------------------------------------------------------------------------
# ABA 5: EMISSÃO DE RELATÓRIOS
# ---------------------------------------------------------------------------

with tab5:
    st.markdown('<div class="section-title">📤 Emissão de Relatórios em Excel</div>', unsafe_allow_html=True)
    st.caption(
        "Escolha um modelo padrão ou selecione manualmente as linhas da DRE, e gere um relatório "
        "formatado com 4 planilhas: Resumo (consolidado), Detalhe Mensal (contas x meses, por loja), "
        "Plano de Contas (composição de cada linha, por loja, a partir do DIÁRIO) e Lançamentos "
        "(cópia filtrada da DIÁRIO). Dentro do Excel, use o filtro (▾) da coluna \"Loja\" para escolher "
        "qual visão ver sem precisar gerar o arquivo de novo."
    )
    st.markdown("<br>", unsafe_allow_html=True)

    linhas_relatorio = df_ref[col_nome].dropna().astype(str).unique()

    if departamento_ativo:
        # No Modo Departamento, o relatório só oferece o modelo do próprio
        # departamento -- já vem selecionado, sem precisar escolher.
        opcoes_modelo = [departamento_ativo]
        modelo_sel = st.selectbox(
            "📁 Modelo de Relatório:", opcoes_modelo,
            format_func=_nome_departamento_curto, disabled=True,
        )
    else:
        opcoes_modelo = ["Seleção manual"] + list(MODELOS_RELATORIO.keys())
        modelo_sel = st.selectbox("📁 Modelo de Relatório:", opcoes_modelo)

    default_contas = []
    termos_nao_encontrados = []
    if modelo_sel != "Seleção manual":
        termos_do_modelo = (
            MODELOS_RELATORIO[modelo_sel]["linhas_dre"]
            + MODELOS_RELATORIO[modelo_sel].get("linhas_informativas", [])
        )
        for termo in termos_do_modelo:
            encontrados = _resolver_termo_departamento(termo, linhas_relatorio)
            if encontrados:
                default_contas.extend(encontrados)
            else:
                termos_nao_encontrados.append(termo)
        default_contas = list(dict.fromkeys(default_contas))
        if termos_nao_encontrados:
            st.warning(
                "Não encontrei estas linhas do modelo na DRE atual (o texto pode estar um pouco diferente): "
                + "; ".join(termos_nao_encontrados)
                + ". Adicione-as manualmente no campo abaixo, se existirem."
            )

    contas_relatorio = st.multiselect(
        "🔍 Linhas da DRE incluídas no relatório:",
        options=linhas_relatorio,
        default=default_contas,
        key=f"contas_relatorio__{modelo_sel}",
    )

    opcoes_lojas_relatorio = list(abas_disponiveis)

    lojas_relatorio_sel = st.multiselect(
        "🏬 Lojas / Visões incluídas no relatório:",
        options=opcoes_lojas_relatorio,
        default=opcoes_lojas_relatorio,
        help=(
            "Por padrão, todas as lojas e visões consolidadas entram no relatório -- dentro do Excel "
            "gerado, use o filtro (▾) na coluna \"Loja\" das abas \"Detalhe Mensal\" e \"Plano de "
            "Contas\" para escolher qual visão ver, sem precisar gerar o arquivo de novo. Só reduza a "
            "seleção aqui se quiser um arquivo menor desde já."
        ),
    )

    col_btn, col_info = st.columns([1, 2])
    with col_btn:
        gerar_clicado = st.button(
            "📊 Gerar Relatório Excel",
            use_container_width=True,
            disabled=not contas_relatorio or not lojas_relatorio_sel,
        )

    if gerar_clicado and contas_relatorio:
        with st.spinner("Carregando dados por loja, plano de contas e DIÁRIO..."):
            # Carrega só as lojas/visões escolhidas no filtro acima -- evita
            # gerar um arquivo gigante com todas as 26 abas de uma vez.
            dados_por_loja_rel = carregar_dados_por_loja(path_orc, path_real, lojas_relatorio_sel)
            df_tabela_contas = carregar_tabela_contas(path_real)
            mapa_planos_dre_rel = montar_mapa_planos_por_dre(df_tabela_contas)
            df_diario_rel = carregar_diario(path_real)
            df_tabela_lojas_rel = carregar_tabela_lojas(path_real)
            mapa_loja_cc_rel = montar_mapa_loja_centro_custo(df_tabela_lojas_rel)

        info_modelo_sel = MODELOS_RELATORIO.get(modelo_sel, {})

        with st.spinner("Montando o relatório em Excel..."):
            excel_bytes = montar_relatorio_excel(
                contas_relatorio, list_df_real, list_df_orc, m_map, colunas_validas, label_visao,
                dados_por_loja=dados_por_loja_rel,
                mapa_planos_dre=mapa_planos_dre_rel,
                df_diario=df_diario_rel,
                forcar_planos_contas=info_modelo_sel.get("forcar_planos_contas", []),
                permitir_lancamento_manual=info_modelo_sel.get("permitir_lancamento_manual", False),
                mapa_loja_centro_custo=mapa_loja_cc_rel,
            )
        st.session_state["relatorio_excel_bytes"] = excel_bytes

        def _nome_arquivo_modelo(nome_modelo):
            """Usa o nome completo do modelo (igual aparece no seletor, só
            sem o emoji na frente) como base do nome do arquivo -- ex.:
            "🛒 Relatório de Custos - Compras" vira "Relatório de Custos -
            Compras". Só remove caracteres inválidos em nome de arquivo."""
            texto = re.sub(r"^[^\w]+", "", nome_modelo, flags=re.UNICODE).strip()
            texto = re.sub(r'[\\/*?:"<>|]', "", texto)
            return texto or "Relatório"

        st.session_state["relatorio_excel_nome"] = f"{_nome_arquivo_modelo(modelo_sel)}.xlsx"
        st.success(f"Relatório gerado com {len(contas_relatorio)} conta(s) selecionada(s), pronto para download.")

        if df_diario_rel is None or df_diario_rel.empty:
            st.warning(
                "Aba 'DIÁRIO' não encontrada (ou vazia/sem as colunas esperadas) no arquivo Realizado — "
                "a aba 'Plano de Contas' do relatório usou o método antigo (Tabela_Contas) como alternativa."
            )
        else:
            st.caption(f"📄 DIÁRIO conectado: {len(df_diario_rel)} lançamento(s) encontrados na aba do Realizado.")

    if not contas_relatorio:
        st.info("Selecione um modelo padrão acima, ou escolha manualmente ao menos uma linha da DRE.")

    if st.session_state.get("relatorio_excel_bytes"):
        st.download_button(
            label="⬇️ Baixar Relatório (.xlsx)",
            data=st.session_state["relatorio_excel_bytes"],
            file_name=st.session_state.get("relatorio_excel_nome", "relatorio_dre.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption("O relatório contém três planilhas: **Resumo** (total do ano por conta, consolidado), **Detalhe Mensal** (contas nas linhas x meses nas colunas, por loja) e **Plano de Contas** (composição de cada linha da DRE, por loja, a partir do DIÁRIO).")

# ---------------------------------------------------------------------------
# ABA 6: GESTÃO DE USUÁRIOS (somente administrador)
# ---------------------------------------------------------------------------
if eh_admin and not departamento_ativo:
    with tab6:
        st.markdown('<div class="section-title">👥 Gestão de Usuários</div>', unsafe_allow_html=True)
        st.caption(
            "Cadastre novos usuários de **visualização** (eles só podem usar os filtros, "
            "visualizar o painel e gerar relatórios — não têm acesso a esta aba). "
            "Como o painel não tem banco de dados, o cadastro gera um bloco de texto que "
            "você precisa colar nos **Secrets** do app (Configurações do app → Secrets)."
        )
        st.markdown("<br>", unsafe_allow_html=True)

        usuarios_atuais = obter_usuarios_cadastrados()
        st.markdown("**Usuários atualmente configurados nos Secrets:**")
        lista_usuarios = [
            {"E-mail": u["email"], "Perfil": u["perfil"]}
            for u in usuarios_atuais.values()
        ]
        st.dataframe(pd.DataFrame(lista_usuarios), use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("**➕ Novo usuário de visualização**")

        with st.form("form_novo_usuario"):
            col_email, col_senha = st.columns(2)
            with col_email:
                novo_email = st.text_input("E-mail do novo usuário")
            with col_senha:
                nova_senha = st.text_input("Senha do novo usuário", type="password")
            gerar_clicado_usuario = st.form_submit_button("Gerar bloco para os Secrets")

        if gerar_clicado_usuario:
            if not novo_email or not nova_senha:
                st.warning("Preencha e-mail e senha para gerar o cadastro.")
            elif novo_email.strip().lower() in usuarios_atuais:
                st.error("Já existe um usuário cadastrado com esse e-mail.")
            else:
                apelido = re.sub(r"[^a-z0-9]+", "_", novo_email.strip().lower().split("@")[0]).strip("_")
                senha_escapada = nova_senha.replace('"', '\\"')
                bloco_toml = (
                    f'[usuarios.{apelido}]\n'
                    f'email = "{novo_email.strip().lower()}"\n'
                    f'senha = "{senha_escapada}"\n'
                    f'perfil = "visualizacao"'
                )
                st.success("Copie o bloco abaixo e cole nos **Secrets** do app (em uma nova linha, mantendo os que já existem).")
                st.code(bloco_toml, language="toml")
                st.caption(
                    "Depois de colar e salvar nos Secrets, o app reinicia sozinho e o novo "
                    "usuário já consegue entrar com o e-mail e a senha cadastrados."
                )


# ============================================================================
# 9. RODAPÉ
# ============================================================================
st.markdown(
    html_compacto(f"""
    <div class="footer-note">
        Controladoria B&A · Painel Financeiro · Dados atualizados automaticamente a cada 60s
    </div>
    """),
    unsafe_allow_html=True,
)