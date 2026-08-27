"""
Testes do Painel Controladoria/Financeiro — Grupo B&A
=====================================================

Rodar fora do Streamlit, na pasta do projeto:

    python testes_painel.py            # usa app.py
    python testes_painel.py outro.py   # ou o arquivo que quiser

O que estes testes protegem, em uma frase: os erros que NÃO aparecem na tela.
Cada bloco aqui nasceu de um problema real que passou despercebido porque o
painel continuou abrindo normalmente, só com o número errado.

Como funciona: as funções são extraídas do arquivo do app pela árvore de
sintaxe (sem importar o módulo, o que exigiria o Streamlit rodando) e
executadas com dados montados à mão.
"""
import ast
import base64
import json
import hashlib
from datetime import date, datetime
import hmac
import io
import re
import sys
import time
import unittest
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd

CAMINHO_APP = sys.argv[1] if len(sys.argv) > 1 else "app.py"
try:
    FONTE = open(CAMINHO_APP, encoding="utf-8").read()
except FileNotFoundError:
    print(f"Nao encontrei {CAMINHO_APP}. Rode na pasta do projeto ou passe o caminho.")
    raise SystemExit(1)

ARVORE = ast.parse(FONTE)
FUSO_BR = ZoneInfo("America/Sao_Paulo")


# Dependencias internas: quem pede a funcao da esquerda precisa das da
# direita para rodar, mesmo sem saber disso. Sem esta tabela, cada teste teria
# de conhecer as tripas do app -- e quebrava sozinho quando o app mudava por
# dentro sem mudar de comportamento.
DEPENDENCIAS = {
    "_linhas_raiz_do_conjunto": ["_eh_linha_de_resultado", "_normalizar_texto"],
    "_eh_linha_de_resultado": ["_normalizar_texto"],
    "resolver_planos_forcados": ["_planos_sem_linha_dre", "_normalizar_texto"],
    "_planos_sem_linha_dre": ["_normalizar_texto"],
    "_assinatura_coluna_fin": ["_normalizar_coluna_fin"],
    "resolver_colunas_fluxo": ["_assinatura_coluna_fin", "_normalizar_coluna_fin"],
    "guardar_memoria": ["memoria_em_uso_mb"],
    "_avaliar_alertas_fluxo": ["_saldo_posicao_atual_fin", "_saldo_abertura_do_mes",
                               "formata_brl"],
    "meta_diaria_que_ainda_falta": [],
    "_tabela_departamento": ["_cor_valor_invertido", "cor_valor", "formata_brl"],
    "_cor_valor_invertido": ["cor_valor"],
    "_segredo": ["_segredo_com_origem"],
    "seletor_periodo_meses": ["_rotulo_mes_pt_extenso"],
    "rotulo_periodo_meses": ["_rotulo_mes_pt_extenso"],
    "prazo_do_fechamento": ["pendentes_do_fechamento", "_data_do_texto_br"],
    "panorama_do_ano": ["checklist_da_competencia", "resumo_do_fechamento"],
    "preparar_checklist_da_planilha": ["resolver_colunas_fluxo", "_assinatura_coluna_fin",
                                      "_normalizar_coluna_fin", "_texto_ou_vazio",
                                      "normalizar_status_fechamento", "_normalizar_coluna_fin"],
    "normalizar_status_fechamento": ["_normalizar_coluna_fin"],
    "chave_conta_orcamento": ["_normalizar_coluna_fin"],
    "realizado_da_dre_por_aba": ["realizado_da_linha_dre", "chave_conta_orcamento",
                                 "_normalizar_coluna_fin"],
    "realizado_da_linha_dre": ["chave_conta_orcamento", "_normalizar_coluna_fin"],
    "ler_meta_da_industria": ["chave_conta_orcamento", "_normalizar_nome_aba",
                              "_normalizar_coluna_fin", "_normalizar_texto"],
}
CONSTANTES_DE_DEPENDENCIA_CONST = {
    # Constante que depende de outra constante para ser avaliada.
    "VISOES_CONSOLIDADAS": ["GRUPO_ABPR", "GRUPO_VD", "GRUPO_LJ_GA", "GRUPO_LJ_CONSOLIDADO"],
    "RENOMEAR_MOVIMENTO_FIN": ["MOV_RECEBER_META", "MOV_RECEBER_AVENCER",
                               "MOV_RECEBER_LIQUIDADO", "MOV_PAGAR"],
}
CONSTANTES_DE_DEPENDENCIA = {
    "_eh_linha_de_resultado": ["PALAVRAS_LINHA_DE_RESULTADO"],
    "_planos_sem_linha_dre": ["MARCAS_SEM_LINHA_DRE"],
    "_assinatura_coluna_fin": ["LIGACOES_NOME_COLUNA", "_ACENTOS_FIN"],
    "_normalizar_coluna_fin": ["_ACENTOS_FIN"],
    "tabela_selecionavel": ["COLORS", "FONTE_MONO", "FONTE_PADRAO_TABELA",
                            "TETO_LINHAS_TABELA", "ALTURA_LINHA_TABELA_PX",
                            "ALTURA_CABECALHO_TABELA_PX", "ALTURA_BARRA_SOMA_PX",
                            "ALTURA_BARRA_ROLAGEM_PX", "COLUNAS_SEM_ROLAGEM_HORIZONTAL",
                            "FOLGA_ABAIXO_TABELA_PX",
                            "FUNDO_TABELA_FLUXO", "PARAM_LINHAS_ABERTAS",
                            "SEPARADOR_LINHAS_ABERTAS", "PREFIXO_BOTAO_ABRIR"],
    "botoes_de_abrir": ["PREFIXO_BOTAO_ABRIR"],
    "guardar_memoria": ["LIMITE_MEMORIA_LIMPEZA_MB"],
    "linhas_abertas_da_url": ["PARAM_LINHAS_ABERTAS", "SEPARADOR_LINHAS_ABERTAS"],
    "resolver_planos_forcados": ["MODELOS_RELATORIO"],
}


def dubla_html_embutido(capturado):
    """Dublê de html_embutido com a assinatura COMPLETA da função de
    verdade. Cada teste tinha o seu, escrito com os argumentos do dia -- e
    quando a função ganhou `redimensionavel`, catorze testes quebraram de
    uma vez por um motivo que não era o que eles testavam."""

    def _capturar(codigo, altura=0, largura=None, redimensionavel=False):
        capturado.update(codigo=codigo, altura=altura, largura=largura,
                         redimensionavel=redimensionavel)

    return _capturar


def carregar(nomes_funcoes, nomes_constantes=(), extras=None):
    """Executa apenas as funcoes e constantes pedidas, num espaco isolado."""
    nomes_funcoes = list(nomes_funcoes)
    nomes_constantes = list(nomes_constantes)
    # Fila, e nao um laco unico: a dependencia de uma dependencia tambem
    # precisa entrar. Sem isso, o detector entrava mas a constante dele nao.
    fila = list(nomes_funcoes)
    while fila:
        nome = fila.pop()
        for extra in DEPENDENCIAS.get(nome, []):
            if extra not in nomes_funcoes:
                nomes_funcoes.append(extra)
                fila.append(extra)
        for extra in CONSTANTES_DE_DEPENDENCIA.get(nome, []):
            if extra not in nomes_constantes:
                nomes_constantes.append(extra)
    for nome in list(nomes_constantes):
        for extra in CONSTANTES_DE_DEPENDENCIA_CONST.get(nome, []):
            if extra not in nomes_constantes:
                nomes_constantes.append(extra)
    pedacos = []
    for no in ARVORE.body:
        if isinstance(no, ast.FunctionDef) and no.name in nomes_funcoes:
            pedacos.append(ast.get_source_segment(FONTE, no))
        elif isinstance(no, ast.Assign):
            for alvo in no.targets:
                if isinstance(alvo, ast.Name) and alvo.id in nomes_constantes:
                    pedacos.append(ast.get_source_segment(FONTE, no))
    espaco = {
        "pd": pd, "re": re, "io": io, "time": time, "datetime": datetime,
        "base64": base64, "hashlib": hashlib, "hmac": hmac, "FUSO_BR": FUSO_BR,
    }
    espaco.update(extras or {})
    exec("\n\n".join(pedacos), espaco)
    faltando = [n for n in nomes_funcoes if n not in espaco]
    if faltando:
        raise SystemExit(f"Nao encontrei no {CAMINHO_APP}: {', '.join(faltando)}")
    return espaco


# ============================================================================
# 1. LEITURA DE DATAS - o erro que trocou dia por mes em 187 mil lancamentos
# ============================================================================
class TesteOrdemDaData(unittest.TestCase):
    """O CSV do fluxo vem em mes/dia/ano. Quem adivinha pelo primeiro valor
    erra quando ele e ambiguo (1/2/2026), e a coluna inteira vira outra coisa
    sem nenhum aviso na tela."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["_ordem_data_fin"])

    def test_mes_primeiro_quando_segundo_componente_passa_de_12(self):
        serie = pd.Series(["12/31/2026", "1/2/2026", "3/20/2026"])
        self.assertFalse(self.ns["_ordem_data_fin"](serie))

    def test_dia_primeiro_quando_primeiro_componente_passa_de_12(self):
        serie = pd.Series(["31/12/2026", "02/01/2026", "20/03/2026"])
        self.assertTrue(self.ns["_ordem_data_fin"](serie))

    def test_valor_ambiguo_no_comeco_nao_engana(self):
        serie = pd.Series(["1/2/2026", "1/5/2026", "12/29/2025"])
        self.assertFalse(self.ns["_ordem_data_fin"](serie))

    def test_tudo_ambiguo_mantem_comportamento_antigo(self):
        serie = pd.Series(["1/2/2026", "3/4/2026"])
        self.assertTrue(self.ns["_ordem_data_fin"](serie))

    def test_serie_vazia_nao_quebra(self):
        self.assertTrue(self.ns["_ordem_data_fin"](pd.Series([], dtype=object)))


# ============================================================================
# 2. SALDO INICIAL DO FLUXO DIARIO
# ============================================================================
class TesteSaldoInicialDiario(unittest.TestCase):
    """Dia sem caixa/banco preenchido herda o fechamento do dia anterior; dia
    com posicao preenchida zera o saldo inicial (senao o dinheiro e contado
    duas vezes)."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["calcular_saldo_inicial_diario"])

    def _base(self):
        def linha(dia, tipo, valor):
            return {"DiaOrd": pd.Timestamp(f"2026-08-{dia:02d}"),
                    "Tipo Movimento": tipo, "Valor.1": valor}
        return pd.DataFrame([
            linha(13, "saldo", 4_000_000.0), linha(13, "entrada", 300_000.0),
            linha(13, "saida", -55_330.09),
            linha(14, "saldo", 4_236_000.0), linha(14, "entrada", 481_290.60),
            linha(14, "saida", -236_433.50),
            linha(15, "saldo", 0.0), linha(15, "saida", -65_906.77),
            linha(16, "saldo", 0.0), linha(16, "saida", -9_195.12),
        ])

    def test_dia_com_posicao_zera_o_saldo_inicial(self):
        ini, _ = self.ns["calcular_saldo_inicial_diario"](self._base(), "Valor.1")
        self.assertEqual(ini[pd.Timestamp("2026-08-13")], 0.0)
        self.assertEqual(ini[pd.Timestamp("2026-08-14")], 0.0)

    def test_dia_sem_posicao_herda_o_fechamento_anterior(self):
        ini, tot = self.ns["calcular_saldo_inicial_diario"](self._base(), "Valor.1")
        self.assertAlmostEqual(ini[pd.Timestamp("2026-08-15")],
                               tot[pd.Timestamp("2026-08-14")], places=2)
        self.assertAlmostEqual(ini[pd.Timestamp("2026-08-16")],
                               tot[pd.Timestamp("2026-08-15")], places=2)

    def test_total_do_dia_soma_saldo_inicial_e_movimento(self):
        ini, tot = self.ns["calcular_saldo_inicial_diario"](self._base(), "Valor.1")
        d15 = pd.Timestamp("2026-08-15")
        self.assertAlmostEqual(tot[d15], ini[d15] - 65_906.77, places=2)

    def test_acumulo_atravessa_meses(self):
        linhas = [{"DiaOrd": pd.Timestamp("2026-08-14"), "Tipo Movimento": "saldo",
                   "Valor.1": 1_000_000.0}]
        for mes, dias in ((8, range(15, 32)), (9, range(1, 31))):
            for dia in dias:
                linhas.append({"DiaOrd": pd.Timestamp(2026, mes, dia),
                               "Tipo Movimento": "saida", "Valor.1": -1_000.0})
        ini, tot = self.ns["calcular_saldo_inicial_diario"](pd.DataFrame(linhas), "Valor.1")
        self.assertAlmostEqual(ini[pd.Timestamp("2026-09-01")],
                               tot[pd.Timestamp("2026-08-31")], places=2)


# ============================================================================
# 3. ALERTAS DO FLUXO DE CAIXA
# ============================================================================
class TesteAlertasDoFluxo(unittest.TestCase):
    """Alerta que dispara sempre vira ruido e passa a ser ignorado - por isso
    cada regra e testada no cenario que deve disparar E no que nao deve."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_avaliar_alertas_fluxo", "_saldo_posicao_atual_fin",
             "_saldo_abertura_do_mes", "formata_brl"],
            ["META_RESERVA_PADRAO", "HORIZONTE_ALERTA_SEMANAS", "COL_FIN_VALOR",
             "COL_FIN_VENCIMENTO", "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_LIQ_AMPLA",
             "COL_FIN_LIQ_DIARIO", "COL_FIN_CANAL"],
        )
        cls.hoje = pd.Timestamp(datetime.now(FUSO_BR).date())

    def _linha(self, canal, tipo, dias, valor):
        data = self.hoje + pd.Timedelta(days=dias)
        return {"Canal.1": canal, "Tipo Movimento": tipo, "Data Efetiva": data,
                "Valor.1": valor, "Vencimento.1": data, "Data Liquidação": pd.NaT}

    def _rodar(self, linhas, **kw):
        df = pd.DataFrame(linhas)
        df["Data Liquidação"] = pd.to_datetime(df["Data Liquidação"])
        p = dict(meta=30, vencido=50_000, concentracao=60, horizonte=30)
        p.update(kw)
        return self.ns["_avaliar_alertas_fluxo"](
            df, "Valor.1", p["meta"], p["vencido"], p["concentracao"],
            horizonte_canal_dias=p["horizonte"],
        )

    def test_empresa_saudavel_nao_dispara_nada(self):
        linhas = [self._linha("LOJA", "saldo", 0, 8_000_000.0),
                  self._linha("LOJA", "entrada", 3, 2_000_000.0)]
        linhas += [self._linha("LOJA", "saida", d, -60_000.0) for d in range(1, 13)]
        self.assertEqual(self._rodar(linhas), [])

    def test_vencido_em_aberto_dispara(self):
        linhas = [self._linha("LOJA", "saldo", 0, 5_000_000.0)]
        linhas += [self._linha("LOJA", "saida", -20, -40_000.0) for _ in range(3)]
        titulos = [a["titulo"] for a in self._rodar(linhas)]
        self.assertTrue(any("vencidos" in t for t in titulos), titulos)

    def test_vencido_abaixo_do_limite_nao_dispara(self):
        linhas = [self._linha("LOJA", "saldo", 0, 5_000_000.0),
                  self._linha("LOJA", "saida", -20, -1_000.0)]
        titulos = [a["titulo"] for a in self._rodar(linhas)]
        self.assertFalse(any("vencidos" in t for t in titulos), titulos)

    def test_pagamentos_da_semana_acima_do_caixa(self):
        linhas = [self._linha("LOJA", "saldo", 0, 100_000.0),
                  self._linha("LOJA", "saida", 3, -500_000.0)]
        niveis = {a["nivel"] for a in self._rodar(linhas)}
        self.assertIn("critico", niveis)

    def test_concentracao_exige_minimo_de_dias(self):
        linhas = [self._linha("LOJA", "saldo", 0, 9_000_000.0)]
        linhas += [self._linha("LOJA", "saida", d, -50_000.0) for d in range(1, 4)]
        titulos = [a["titulo"] for a in self._rodar(linhas)]
        self.assertFalse(any("desembolso do" in t for t in titulos), titulos)

    def test_canal_sem_caixa_proprio_e_atencao_quando_a_empresa_cobre(self):
        linhas = [self._linha("LOJA", "saldo", 0, 50_000.0),
                  self._linha("VENDA DIRETA", "saldo", 0, 5_000_000.0),
                  self._linha("LOJA", "saida", 5, -400_000.0)]
        canal = [a for a in self._rodar(linhas) if "sem recurso" in a["titulo"]]
        self.assertEqual(len(canal), 1)
        self.assertEqual(canal[0]["nivel"], "atencao")

    def test_canal_e_empresa_sem_caixa_e_critico(self):
        linhas = [self._linha("LOJA", "saldo", 0, 50_000.0),
                  self._linha("VENDA DIRETA", "saldo", 0, 80_000.0),
                  self._linha("LOJA", "saida", 5, -400_000.0)]
        canal = [a for a in self._rodar(linhas) if "sem recurso" in a["titulo"]]
        self.assertEqual(len(canal), 1)
        self.assertEqual(canal[0]["nivel"], "critico")

    def test_alerta_de_reserva_le_pela_abertura_do_mes(self):
        """25/08/2026: o alerta usava o saldo de HOJE somado ao total de
        entradas do mes -- e a posicao de hoje ja contem o que foi recebido no
        mes. O mesmo dinheiro entrava duas vezes e o alerta anunciava uma
        reserva melhor do que a tabela mostrava logo abaixo. Aqui o mes abre
        com 1 mi, recebe 4 mi e paga 4 mi: sobra 1 mi de 5 mi disponiveis, ou
        20%. Pela conta antiga (saldo de hoje = 5 mi + 4 mi de entradas) daria
        55%, e o alerta nao dispararia."""
        primeiro = self.hoje.replace(day=1)
        dias_ate_o_primeiro = (primeiro - self.hoje).days
        linhas = [
            # Abertura do mes: 1 milhao.
            {"Canal.1": "LOJA", "Tipo Movimento": "saldo", "Data Efetiva": primeiro,
             "Valor.1": 1_000_000.0, "Vencimento.1": primeiro, "Data Liquidação": pd.NaT},
            # Posicao de hoje: 5 milhoes -- ja embute os 4 recebidos.
            self._linha("LOJA", "saldo", 0, 5_000_000.0),
            self._linha("LOJA", "entrada", max(dias_ate_o_primeiro + 1, -25), 4_000_000.0),
        ]
        linhas += [self._linha("LOJA", "saida", 1, -4_000_000.0)]
        reserva = [a for a in self._rodar(linhas) if "Reserva do mês" in a["titulo"]]
        self.assertEqual(len(reserva), 1, "a reserva em 20% tem de disparar o alerta")
        self.assertIn("20%", reserva[0]["titulo"])

    def test_abertura_do_mes_pega_o_primeiro_dia_com_saldo(self):
        """E soma os canais NAQUELE dia: e a posicao da empresa na data, nao a
        soma do mes. Mes sem nenhuma posicao devolve zero, nao erro."""
        abertura = self.ns["_saldo_abertura_do_mes"]
        df = pd.DataFrame([
            {"Tipo Movimento": "saldo", "Data Efetiva": pd.Timestamp(2026, 7, 1),
             "Valor.1": 300_000.0},
            {"Tipo Movimento": "saldo", "Data Efetiva": pd.Timestamp(2026, 7, 1),
             "Valor.1": 200_000.0},
            {"Tipo Movimento": "saldo", "Data Efetiva": pd.Timestamp(2026, 7, 31),
             "Valor.1": 900_000.0},
            {"Tipo Movimento": "entrada", "Data Efetiva": pd.Timestamp(2026, 7, 2),
             "Valor.1": 50_000.0},
        ])
        valor, data = abertura(df, "Valor.1", pd.Period("2026-07", "M"))
        self.assertAlmostEqual(valor, 500_000.0, places=2)
        self.assertEqual(data, pd.Timestamp(2026, 7, 1))
        vazio, sem_data = abertura(df, "Valor.1", pd.Period("2026-09", "M"))
        self.assertEqual(vazio, 0.0)
        self.assertIsNone(sem_data)

    def test_canal_sem_posicao_propria_nao_e_avaliado(self):
        linhas = [self._linha("VENDA DIRETA", "saldo", 0, 9_000_000.0),
                  self._linha("HUB LOGISTICO", "saida", 4, -80_000.0)]
        canal = [a for a in self._rodar(linhas) if "sem recurso" in a["titulo"]]
        self.assertEqual(canal, [])


# ============================================================================
# 4. SESSAO QUE SOBREVIVE AO F5
# ============================================================================
class TesteBilheteDeSessao(unittest.TestCase):
    """O bilhete viaja na URL. Se pudesse ser forjado ou ter o prazo esticado
    a mao, qualquer pessoa entraria sem senha."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_assinar_sessao", "gerar_bilhete_sessao", "ler_bilhete_sessao"],
            ["MINUTOS_INATIVIDADE", "CHAVE_SESSAO_URL"],
            extras={"_segredo_sessao": lambda: "segredo-de-teste"},
        )

    def test_bilhete_valido_devolve_o_email(self):
        b = self.ns["gerar_bilhete_sessao"]("Controladoria@GrupoBeea.com.br ")
        self.assertEqual(self.ns["ler_bilhete_sessao"](b), "controladoria@grupobeea.com.br")

    def test_bilhete_forjado_e_recusado(self):
        falso = base64.urlsafe_b64encode(b"outro@empresa.com|9999999999|" + b"0" * 32).decode()
        self.assertIsNone(self.ns["ler_bilhete_sessao"](falso))

    def test_prazo_esticado_a_mao_e_recusado(self):
        b = self.ns["gerar_bilhete_sessao"]("controladoria@grupobeea.com.br")
        email, exp, assinatura = base64.urlsafe_b64decode(b).decode().rsplit("|", 2)
        esticado = base64.urlsafe_b64encode(
            f"{email}|{int(exp) + 99999}|{assinatura}".encode()).decode()
        self.assertIsNone(self.ns["ler_bilhete_sessao"](esticado))

    def test_bilhete_vencido_e_recusado(self):
        corpo = f"controladoria@grupobeea.com.br|{int(time.time()) - 60}"
        vencido = base64.urlsafe_b64encode(
            f"{corpo}|{self.ns['_assinar_sessao'](corpo)}".encode()).decode()
        self.assertIsNone(self.ns["ler_bilhete_sessao"](vencido))

    def test_lixo_e_vazio_nao_quebram(self):
        for valor in ("abc123", "", None, "!!!"):
            self.assertIsNone(self.ns["ler_bilhete_sessao"](valor))

    def test_bilhete_nao_carrega_a_senha(self):
        b = self.ns["gerar_bilhete_sessao"]("controladoria@grupobeea.com.br")
        self.assertNotIn("senha", base64.urlsafe_b64decode(b).decode().lower())


# ============================================================================
# 5. RELATORIOS: PLANOS DE CONTAS
# ============================================================================
class TestePlanosDeContas(unittest.TestCase):
    """Cada modelo so pode oferecer os planos que pertencem a ele - oferecer a
    lista inteira seria dar a um departamento acesso a conta de outro."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["planos_do_diario", "resolver_planos_forcados",
                           "_normalizar_texto"])
        cls.diario = pd.DataFrame({
            "Plano de Contas": ["Mercadorias", "Material de Embalagem", "Flaconetes",
                                "Benfeitorias em Imóveis de Terceiros", "Combustível"],
            "Linha DRE": ["", "6.6 - Material de Embalagem", "6.14 - Flaconetes",
                          "", "8.5.3 - Combustível"],
            "Valor Bruto": [-1.0] * 5,
        })

    def test_lista_restrita_as_linhas_do_modelo(self):
        planos = self.ns["planos_do_diario"](self.diario, ["6.14 - Flaconetes"])
        self.assertEqual(planos, ["Flaconetes"])

    def test_sem_linhas_traz_todos(self):
        self.assertEqual(len(self.ns["planos_do_diario"](self.diario)), 5)

    def test_plano_sem_linha_da_dre_e_achado_pelo_nome(self):
        achados, faltando = self.ns["resolver_planos_forcados"](self.diario, ["Mercadorias"])
        self.assertEqual(achados, ["Mercadorias"])
        self.assertEqual(faltando, [])

    def test_busca_tolera_caixa_e_nome_parcial(self):
        for termo in ("mercadorias", "MERCADORIAS", "Benfeitorias em Imóveis"):
            achados, _ = self.ns["resolver_planos_forcados"](self.diario, [termo])
            self.assertTrue(achados, termo)

    def test_termo_inexistente_e_reportado(self):
        achados, faltando = self.ns["resolver_planos_forcados"](self.diario, ["Conta Que Sumiu"])
        self.assertEqual(achados, [])
        self.assertEqual(faltando, ["Conta Que Sumiu"])


# ============================================================================
# 5b. LEITURA DE LINHA DA DRE POR CODIGO
# ============================================================================
class TesteSomaComFilhas(unittest.TestCase):
    """A linha-mae nem sempre esta preenchida: em algumas abas o valor so
    existe nas filhas. Ler so a mae devolvia zero em silencio."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["_valor_linha_por_codigo", "_somar_codigo_com_filhas"])

    def _aba(self, linhas):
        return pd.DataFrame([{"Nome": n, "01/2026": v} for n, v in linhas])

    def test_usa_a_mae_quando_ela_existe(self):
        aba = self._aba([("6.24.2 - Marketing", -90_000.0), ("6.24.2.1 - Midia", -50_000.0)])
        self.assertEqual(
            self.ns["_somar_codigo_com_filhas"](aba, "6.24.2", ["01/2026"]), -90_000.0)

    def test_desce_para_as_filhas_quando_a_mae_falta(self):
        aba = self._aba([("6.24.2.1 - Midia", -30_000.0), ("6.24.2.2 - Eventos", -20_000.0)])
        self.assertEqual(
            self.ns["_somar_codigo_com_filhas"](aba, "6.24.2", ["01/2026"]), -50_000.0)

    def test_nao_soma_neto_junto_com_filho(self):
        aba = self._aba([("6.24.2.1 - Midia", -30_000.0), ("6.24.2.1.1 - Radio", -10_000.0)])
        self.assertEqual(
            self.ns["_somar_codigo_com_filhas"](aba, "6.24.2", ["01/2026"]), -30_000.0)

    def test_sem_a_conta_devolve_zero(self):
        aba = self._aba([("9 - Resultado Operacional", 0.0)])
        self.assertEqual(self.ns["_somar_codigo_com_filhas"](aba, "6.24.2", ["01/2026"]), 0.0)


# ============================================================================
# 5c. GERENCIA COMERCIAL
# ============================================================================
class TesteGerenciaComercial(unittest.TestCase):
    """O bloco reproduz o relatorio que a Controladoria enviava na mao: o
    Quadro de Metas (3 linhas) e o grupo 6 aberto por sublinha."""

    NOME = "📈 Relatório de Custos - Gerência Comercial"

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_normalizar_texto", "_normalizar_nome_aba", "get_valor_consolidado_multi",
             "_nome_sem_numero_dre", "_resolver_termo_departamento", "cor_valor",
             "cor_variacao", "formata_brl", "formata_valor_curto", "faixa_metricas_html",
             "html_compacto", "_total_dre", "_tabela_departamento", "_painel_dept_comercial",
             "_numero_linha_dre", "_linha_pertence_ao_grupo", "_linhas_raiz_do_conjunto"],
            ["COLORS", "FONTE_MONO", "LINHA_RECEITA_LIQUIDA", "LINHA_DESPESAS_VARIAVEIS",
             "MODELOS_RELATORIO"],
            extras={"st": type("st", (), {
                "markdown": staticmethod(lambda *a, **k: None),
                "caption": staticmethod(lambda *a, **k: None),
                "dataframe": staticmethod(lambda d, **k: TesteGerenciaComercial.tabelas.append(d)),
            })()},
        )

    def setUp(self):
        TesteGerenciaComercial.tabelas = []

    def _resolver(self, universo):
        """Separa o que a gerencia gere (Quadro de Metas) do que vai junto so
        para conhecimento (grupo 6), como o painel faz."""
        modelo = self.ns["MODELOS_RELATORIO"][self.NOME]
        geridas, informativas = [], []
        for termo in modelo["linhas_dre"]:
            geridas.extend(self.ns["_resolver_termo_departamento"](termo, universo))
        for termo in modelo.get("linhas_informativas", []):
            informativas.extend(self.ns["_resolver_termo_departamento"](termo, universo))
        return list(dict.fromkeys(geridas)), list(dict.fromkeys(informativas))

    def _rodar(self):
        # Valores consolidados (jan a maio) do relatorio enviado pelo gestor.
        real = {"6 - Despesas Variáveis": -6_108_698.41,
                "6.24.2.5 - Encontro de Ciclo": -4_519.40,
                "6.24.2.6 - Outras Despesas de Marketing": -156_203.19,
                "8.3.3.7 - Prêmios / Bônus": -176_115.22,
                "6.1 - Comissões sobre Vendas": -700_432.11,
                "3 - Receita Operacional Liquida": 75_737_289.82}
        orc = {"6 - Despesas Variáveis": -7_000_276.88,
               "6.24.2.5 - Encontro de Ciclo": -28_400.00,
               "6.24.2.6 - Outras Despesas de Marketing": -393_600.00,
               "8.3.3.7 - Prêmios / Bônus": -128_357.03,
               "6.1 - Comissões sobre Vendas": -801_680.80,
               "3 - Receita Operacional Liquida": 91_012_257.49}
        aba = lambda d: pd.DataFrame([{"Nome": n, "01/2026": v} for n, v in d.items()])
        universo = list(real.keys())
        geridas, informativas = self._resolver(universo)
        self.ns["_painel_dept_comercial"]({
            "departamento": self.NOME, "dfs_real": [aba(real)], "dfs_orc": [aba(orc)],
            "colunas": ["01/2026"], "linhas_resolvidas": geridas,
            "linhas_raiz": self.ns["_linhas_raiz_do_conjunto"](geridas),
            "linhas_todas": universo, "linhas_informativas": informativas,
            "path_orc": "o", "path_real": "r"})
        return [t.data for t in TesteGerenciaComercial.tabelas]

    def test_quadro_de_metas_bate_com_o_relatorio(self):
        metas = self._rodar()[0]
        total = metas[metas["Linha da DRE"].str.startswith("TOTAL")].iloc[0]
        self.assertAlmostEqual(total["Realizado (R$)"], 336_837.81, places=2)
        self.assertAlmostEqual(total["Orçado (R$)"], 550_357.03, places=2)

    def test_quadro_de_metas_tem_as_tres_linhas(self):
        metas = self._rodar()[0]
        self.assertEqual(len(metas), 4)  # 3 linhas + TOTAL
        self.assertTrue(metas.iloc[-1]["Linha da DRE"].startswith("TOTAL"))

    def test_grupo_6_nao_soma_no_total_do_departamento(self):
        """O grupo 6 vai junto so para conhecimento. Se ele entrasse nas
        linhas geridas, as duas linhas de marketing (netas da 6) seriam
        contadas duas vezes e o total do departamento nao bateria com o
        'Quadro de Metas GER.COM.' do relatorio."""
        universo = ["6 - Despesas Variáveis", "6.1 - Comissões sobre Vendas",
                    "6.24.2.5 - Encontro de Ciclo",
                    "6.24.2.6 - Outras Despesas de Marketing",
                    "8.3.3.7 - Prêmios / Bônus"]
        geridas, informativas = self._resolver(universo)
        self.assertEqual(len(geridas), 3, geridas)
        self.assertNotIn("6 - Despesas Variáveis", geridas)
        self.assertIn("6 - Despesas Variáveis", informativas)

    def test_grupo_6_fecha_com_o_total(self):
        grupo = self._rodar()[1]
        total = grupo[grupo["Linha da DRE"].str.startswith("TOTAL")].iloc[0]
        self.assertAlmostEqual(total["Realizado (R$)"], 6_108_698.41, places=2)
        self.assertAlmostEqual(total["Orçado (R$)"], 7_000_276.88, places=2)

    def test_modelo_seleciona_o_recorte_do_relatorio(self):
        """Por padrao o relatorio sai com o grupo 6 e as sublinhas diretas
        (6.1 a 6.25) mais as tres linhas do Quadro de Metas - o mesmo recorte
        do arquivo que a Controladoria envia. A 6.24.2 (mae das duas linhas de
        marketing) NAO entra: ela nao esta no relatorio."""
        ns = carregar(["_numero_linha_dre", "_linha_pertence_ao_grupo",
                       "_resolver_termo_departamento", "_linhas_raiz_do_conjunto"])
        dre = ["1 - Receita Operacional Bruta", "6 - Despesas Variáveis"]
        dre += [f"6.{i} - Linha {i}" for i in range(1, 26)]
        dre += ["6.24.2 - Marketing Regional - Gestão CP", "6.24.2.5 - Encontro de Ciclo",
                "6.24.2.6 - Outras Despesas de Marketing", "8.3.3.7 - Prêmios / Bônus"]
        modelo = ["FILHAS:6 - Despesas Variáveis", "6.24.2.5 - Encontro de Ciclo",
                  "6.24.2.6 - Outras Despesas de Marketing", "8.3.3.7 - Prêmios / Bônus"]
        sel = []
        for termo in modelo:
            sel.extend(ns["_resolver_termo_departamento"](termo, dre))
        sel = list(dict.fromkeys(sel))
        self.assertIn("6 - Despesas Variáveis", sel)
        self.assertEqual(sum(1 for l in sel if l.startswith("6.") and l.count(".") == 1), 25)
        self.assertFalse(any(l.startswith("6.24.2 ") for l in sel),
                         "a linha-mae 6.24.2 nao faz parte do relatorio da gerencia")
        self.assertIn("8.3.3.7 - Prêmios / Bônus", sel)
        self.assertEqual(
            sorted(ns["_linhas_raiz_do_conjunto"](sel)),
            sorted(["6 - Despesas Variáveis", "8.3.3.7 - Prêmios / Bônus"]),
            "os totais do painel contariam o mesmo custo duas vezes")

    def test_bloco_do_departamento_avisa_o_que_ja_mostrou(self):
        """Cada bloco personalizado devolve o que ja desenhou, e o bloco
        generico pula isso. Sem essa combinacao, a Gerencia Comercial exibia
        as mesmas duas tabelas duas vezes na mesma tela."""
        ns = carregar(
            ["_normalizar_texto", "_normalizar_nome_aba", "get_valor_consolidado_multi",
             "_nome_sem_numero_dre", "_resolver_termo_departamento", "_numero_linha_dre",
             "_linha_pertence_ao_grupo", "_linhas_raiz_do_conjunto", "cor_valor",
             "cor_variacao", "formata_brl", "formata_valor_curto", "faixa_metricas_html",
             "html_compacto", "_total_dre", "_tabela_departamento", "_painel_dept_comercial"],
            ["COLORS", "FONTE_MONO", "LINHA_RECEITA_LIQUIDA", "LINHA_DESPESAS_VARIAVEIS",
             "MODELOS_RELATORIO"],
            extras={"st": type("st", (), {
                "markdown": staticmethod(lambda *a, **k: None),
                "caption": staticmethod(lambda *a, **k: None),
                "dataframe": staticmethod(lambda *a, **k: None),
            })()},
        )
        nome = "📈 Relatório de Custos - Gerência Comercial"
        valores = {"6 - Despesas Variáveis": -9_363_906.38,
                   "6.1 - Comissões sobre Vendas": -935_381.52,
                   "6.24.2.5 - Encontro de Ciclo": -4_519.40,
                   "6.24.2.6 - Outras Despesas de Marketing": -415_547.83,
                   "8.3.3.7 - Prêmios / Bônus": -301_779.55,
                   "3 - Receita Operacional Liquida": 75_000_000}
        aba = pd.DataFrame([{"Nome": n, "01/2026": v} for n, v in valores.items()])
        universo = list(valores.keys())
        modelo = ns["MODELOS_RELATORIO"][nome]
        geridas, informativas = [], []
        for termo in modelo["linhas_dre"]:
            geridas.extend(ns["_resolver_termo_departamento"](termo, universo))
        for termo in modelo.get("linhas_informativas", []):
            informativas.extend(ns["_resolver_termo_departamento"](termo, universo))
        cobertura = ns["_painel_dept_comercial"]({
            "departamento": nome, "dfs_real": [aba], "dfs_orc": [aba], "colunas": ["01/2026"],
            "linhas_resolvidas": geridas, "linhas_raiz": ns["_linhas_raiz_do_conjunto"](geridas),
            "linhas_todas": universo, "linhas_informativas": informativas,
            "path_orc": "o", "path_real": "r"})
        self.assertEqual(cobertura, {"linhas_dept", "informativas"})
        self.assertIn('if "linhas_dept" not in _ja_mostrado_dept:', FONTE)
        self.assertIn('"informativas" not in _ja_mostrado_dept', FONTE)

    def test_relatorio_agrupa_o_quadro_de_metas(self):
        """No Excel o Quadro de Metas tem de sair como um bloco proprio, com
        subtotal e as tres linhas recuadas abaixo -- foi assim que o
        relatorio sempre foi enviado. Sem isso, o Excel lista tudo numa
        sequencia so e o total do quadro nao aparece em lugar nenhum."""
        self.assertIn('"bloco_relatorio"', FONTE)
        self.assertIn("Despesa Variavel - Quadro de Metas GER.COM.", FONTE)
        i = FONTE.index("def montar_relatorio_excel(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("blocos_agrupados", trecho,
                      "o gerador do Excel voltou a ignorar os blocos do modelo")
        self.assertIn("contas_em_bloco", trecho)

    def test_relatorio_tem_opcao_de_puxar_a_dre_inteira(self):
        self.assertIn("Puxar todas as linhas da DRE", FONTE)
        self.assertIn("puxar_dre_completa", FONTE)

    def test_modelo_e_email_cadastrados(self):
        self.assertIn(self.NOME, FONTE)
        self.assertIn("gerente.comercial@grupobeea.com.br", FONTE)
        self.assertIn('"linhas_informativas"', FONTE)


# ============================================================================
# 5d. ADM/FINANCEIRO
# ============================================================================
class TesteAdmFinanceiro(unittest.TestCase):
    """O ADM/Financeiro fica com o que sobra da DRE. O recorte e calculado,
    nao listado a mao -- entao precisa de trava, senao um departamento novo
    passa a levar linhas embora sem ninguem perceber."""

    DRE = [
        "1 - Receita Operacional Bruta", "2 - Deduções da Receita Operacional Bruta",
        "3 - Receita Operacional Liquida", "4 - Custo das Vendas",
        "4.1 - Custo da Mercadoria Vendida - CMV", "5 - Margem de Contribuição 1",
        "6 - Despesas Variáveis", "6.1 - Comissões sobre Vendas",
        "6.2 - Taxa com Cartão de Crédito / Débito", "6.5 - Taxa de Emissão de Boleto",
        "6.6 - Material de Embalagem", "6.8 - Serviço de Entrega",
        "6.11 - Catálogos e Revistas", "6.13 - Amostras", "6.14 - Flaconetes",
        "6.16 - Vitrines", "6.24 - Esforços de Marketing",
        "6.24.1 - Marketing Regional - Gestão GB", "6.24.2 - Marketing Regional - Gestão CP",
        "6.24.2.5 - Encontro de Ciclo", "7 - Margem de Contribuição 2",
        "8 - Despesas Operacionais", "8.1 - Ocupação", "8.1.1 - Aluguel",
        "8.1.2 - Energia Elétrica", "8.1.3 - Limpeza e Conservação",
        "8.1.4 - Manutenção e Reparos", "8.2 - Tributos e Taxas", "8.3 - Pessoal",
        "8.3.1 - Salários", "8.3.3.7 - Prêmios / Bônus", "8.5.3 - Combustível",
        "8.6.1 - Material de Escritório", "8.6.6 - Outras Despesas Administrativas",
        "8.6.7 - Tarifas Bancárias", "8.8.1 - Contabilidade",
        "8.8.2 - Auditoria / Consultoria", "11 - EBITDA",
        "12 - Resultado Financeiro", "12.1 - Despesas Financeiras",
        "13 - Depreciação e Amortização",
        "14 - Outras Receitas e Despesas não Operacionais",
        "14.3 - Outras Receitas não Operacionais",
        "15 - Resultado Antes do Imposto", "16 - Impostos sobre o Lucro", "16.1 - IRPJ",
        "17 - Resultado Gerencial do Período",
    ]

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_normalizar_texto", "_numero_linha_dre", "_linha_pertence_ao_grupo",
             "eh_linha_custos_despesas", "_resolver_termo_departamento",
             "_linhas_com_dono_de_departamento", "_linhas_restantes_da_dre",
             "_linhas_raiz_do_conjunto",
             "resolver_planos_forcados"],
            ["MODELOS_RELATORIO", "PALAVRAS_LINHA_DE_RESULTADO"],
        )

    def _geridas(self):
        return self.ns["_resolver_termo_departamento"]("RESTANTE", self.DRE)

    def test_nao_leva_linha_de_outro_departamento(self):
        geridas = self._geridas()
        for alheia in ["6.1 - Comissões sobre Vendas", "6.6 - Material de Embalagem",
                       "6.8 - Serviço de Entrega", "8.1.3 - Limpeza e Conservação",
                       "8.3 - Pessoal", "8.3.1 - Salários", "8.5.3 - Combustível",
                       "8.6.1 - Material de Escritório", "8.8.2 - Auditoria / Consultoria",
                       "6.24.2 - Marketing Regional - Gestão CP",
                       "6.24.1 - Marketing Regional - Gestão GB"]:
            self.assertNotIn(alheia, geridas, f"{alheia} ja tem dono")

    def test_nao_leva_linha_ancestral(self):
        """A linha de grupo carrega o custo dos outros departamentos dentro:
        se ela entrasse, a folha do RH viraria despesa do ADM."""
        geridas = self._geridas()
        for ancestral in ["6 - Despesas Variáveis", "8 - Despesas Operacionais",
                          "8.1 - Ocupação", "6.24 - Esforços de Marketing"]:
            self.assertNotIn(ancestral, geridas, f"{ancestral} soma custo de outra area")

    def test_fica_com_o_que_sobra(self):
        geridas = self._geridas()
        for propria in ["4 - Custo das Vendas", "6.2 - Taxa com Cartão de Crédito / Débito",
                        "6.5 - Taxa de Emissão de Boleto", "6.16 - Vitrines",
                        "8.1.1 - Aluguel", "8.1.2 - Energia Elétrica",
                        "8.2 - Tributos e Taxas", "8.6.7 - Tarifas Bancárias",
                        "8.8.1 - Contabilidade"]:
            self.assertIn(propria, geridas, f"{propria} nao tem dono e sumiu do ADM")

    def test_departamento_leva_a_dre_inteira_que_sobra(self):
        """Nao operacionais, impostos sobre o lucro e ate as linhas de
        resultado fazem parte do departamento -- em tela e no relatorio."""
        geridas = self._geridas()
        for linha in ["13 - Depreciação e Amortização", "12.1 - Despesas Financeiras",
                      "14 - Outras Receitas e Despesas não Operacionais",
                      "14.3 - Outras Receitas não Operacionais",
                      "16 - Impostos sobre o Lucro", "16.1 - IRPJ",
                      "15 - Resultado Antes do Imposto",
                      "17 - Resultado Gerencial do Período",
                      "1 - Receita Operacional Bruta", "11 - EBITDA"]:
            self.assertIn(linha, geridas, f"{linha} sumiu do ADM/Financeiro")

    def test_linha_calculada_nunca_entra_numa_soma(self):
        """15 e 17 sao a DRE inteira fechada: soma-las junto com as contas
        que as formam multiplicaria o total do departamento."""
        calculadas = ["1 - Receita Operacional Bruta",
                      "2 - Deduções da Receita Operacional Bruta",
                      "3 - Receita Operacional Liquida", "5 - Margem de Contribuição 1",
                      "11 - EBITDA", "12 - Resultado Financeiro",
                      "15 - Resultado Antes do Imposto",
                      "17 - Resultado Gerencial do Período"]
        for linha in calculadas:
            self.assertTrue(self.ns["_eh_linha_de_resultado"](linha), linha)
        raizes = self.ns["_linhas_raiz_do_conjunto"](self._geridas())
        for linha in calculadas:
            self.assertNotIn(linha, raizes, f"{linha} entrou no total")

    def test_imposto_sobre_o_lucro_e_conta_de_verdade(self):
        """A palavra "lucro" ja esteve na lista de linhas calculadas e
        derrubava os impostos do total do departamento."""
        self.assertFalse(self.ns["_eh_linha_de_resultado"]("16 - Impostos sobre o Lucro"))
        raizes = self.ns["_linhas_raiz_do_conjunto"](self._geridas())
        self.assertIn("16 - Impostos sobre o Lucro", raizes)
        self.assertNotIn("16.1 - IRPJ", raizes, "a filha nao pode somar junto com a mae")

    def test_gerencia_comercial_nao_tira_linha_do_adm(self):
        """As tres linhas da Gerencia Comercial sao acompanhamento sobre ramos
        de outras areas -- desconta-las abriria um buraco na DRE."""
        donas = self.ns["_linhas_com_dono_de_departamento"](self.DRE)
        self.assertNotIn("6.16 - Vitrines", donas)
        modelo = self.ns["MODELOS_RELATORIO"]["📈 Relatório de Custos - Gerência Comercial"]
        self.assertFalse(modelo.get("linhas_exclusivas", False))

    def test_planos_que_sobram(self):
        diario = pd.DataFrame([
            {"Plano de Contas": "Mercadorias", "Linha DRE": ""},
            {"Plano de Contas": "Benfeitorias em imóvel próprio", "Linha DRE": ""},
            {"Plano de Contas": "Aplicação Financeira", "Linha DRE": ""},
            {"Plano de Contas": "Energia Elétrica", "Linha DRE": "8.1.2 - Energia Elétrica"},
        ])
        achados, _ = self.ns["resolver_planos_forcados"](diario, ["RESTANTE"])
        self.assertIn("Aplicação Financeira", achados)
        self.assertNotIn("Mercadorias", achados, "Mercadorias e de Compras")
        self.assertNotIn("Benfeitorias em imóvel próprio", achados, "benfeitorias sao de Suprimentos")
        self.assertNotIn("Energia Elétrica", achados, "esse ja tem linha da DRE")

    def test_mapa_de_email_aponta_para_departamento_existente(self):
        """Um e-mail apontando para um departamento com nome errado nao da
        erro: a pessoa simplesmente abre na Controladoria e ninguem descobre
        o porque. Por isso a conferencia e automatica."""
        trecho = FONTE[FONTE.index("MAPA_EMAIL_DEPARTAMENTO = {"):]
        trecho = trecho[:trecho.index("\n}")]
        mapeados = re.findall(r'"([^"]+@[^"]+)":\s*"([^"]+)"', trecho)
        self.assertGreaterEqual(len(mapeados), 8)
        for email, departamento in mapeados:
            self.assertIn(departamento, self.ns["MODELOS_RELATORIO"],
                          f"{email} aponta para um departamento que nao existe")
        for email, departamento in [
            ("coordenador.financeiro@grupobeea.com.br", "🏦 Relatório de Custos - ADM/Financeiro"),
            ("coordenador.loja@grupobeea.com.br", "🏬 Relatório de Custos - Coordenação de Loja"),
            ("coordenador.vd@grupobeea.com.br", "🚗 Relatório de Custos - Coordenação de VD"),
        ]:
            self.assertIn((email, departamento), mapeados)

    def test_celula_vazia_de_linha_dre_e_reconhecida(self):
        """A leitura converte a coluna inteira para texto, entao o NaN do
        pandas vira a palavra "nan". Testar so por string vazia deixava
        quase todos os planos fora da DRE passarem batido."""
        ns = carregar(["_normalizar_texto", "_planos_sem_linha_dre"],
                      ["MARCAS_SEM_LINHA_DRE"])
        diario = pd.DataFrame({
            "Plano de Contas": ["Aplicação Financeira", "Empréstimos", "Energia Elétrica",
                                "Distribuição de Lucros", "Transferência"],
            "Linha DRE": [float("nan"), "", "8.1.2 - Energia Elétrica", "-", "nan"],
        })
        diario["Linha DRE"] = diario["Linha DRE"].astype(str).str.strip()
        achados = ns["_planos_sem_linha_dre"](diario)
        self.assertIn("Aplicação Financeira", achados)
        self.assertIn("Distribuição de Lucros", achados)
        self.assertIn("Transferência", achados)
        self.assertNotIn("Energia Elétrica", achados, "esse tem linha da DRE")

    def test_planos_de_outro_departamento_ficam_de_fora(self):
        ns = carregar(["_normalizar_texto", "_planos_sem_linha_dre",
                       "resolver_planos_forcados"],
                      ["MARCAS_SEM_LINHA_DRE", "MODELOS_RELATORIO"])
        diario = pd.DataFrame({
            "Plano de Contas": ["Aplicação Financeira", "Mercadorias",
                                "Adiantamento de Benfeitorias em Imóvel Próprio"],
            "Linha DRE": ["", "", ""],
        })
        achados, _ = ns["resolver_planos_forcados"](diario, ["RESTANTE"])
        self.assertIn("Aplicação Financeira", achados)
        self.assertNotIn("Mercadorias", achados, "Mercadorias e de Compras")
        self.assertNotIn("Adiantamento de Benfeitorias em Imóvel Próprio", achados,
                         "benfeitorias sao de Suprimentos")

    def test_bloco_de_planos_aparece_no_painel(self):
        self.assertIn("Planos de Contas Fora da DRE", FONTE)
        i = FONTE.index("def _painel_dept_adm(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("_planos_sem_linha_dre", trecho)
        self.assertNotIn('"% do bloco"', trecho,
                         "percentual sobre total com sinais opostos nao significa nada")

    def test_todos_passam_pela_tela_de_escolha(self):
        """A tela de escolha entre Controladoria e Financeiro aparece para
        TODO usuario logado. O que muda por pessoa e so o acesso ao
        Financeiro: quem nao esta na lista ve o cartao bloqueado, com o
        botao desligado. Nao pode existir atalho que mande alguem direto
        para um painel sem passar por aqui."""
        i = FONTE.index('if st.session_state["painel_escolhido"] is None:')
        trecho = FONTE[i:i + 12000]
        # A unica condicao que governa a tela e "ainda nao escolheu" -- nao
        # pode haver desvio por e-mail antes dela.
        antes = FONTE[:i]
        self.assertNotIn('st.session_state["painel_escolhido"] = "controladoria"', antes,
                         "alguem esta escolhendo o painel antes da tela aparecer")
        self.assertIn("_pode_financeiro = _email_hub in EMAILS_FINANCEIRO_PERMITIDOS", trecho)
        self.assertIn("disabled=not _pode_financeiro", trecho,
                      "o botao do Financeiro precisa nascer desligado para quem nao tem acesso")
        self.assertIn("st.stop()", trecho, "a tela precisa parar o resto do app")

    def test_acesso_ao_financeiro_e_por_lista(self):
        i = FONTE.index("EMAILS_FINANCEIRO_PERMITIDOS = {")
        trecho = FONTE[i:FONTE.index("}", i)]
        emails = re.findall(r'"([^"]+@[^"]+)"', trecho)
        self.assertIn("controladoria@grupobeea.com.br", emails)
        for de_fora in ["coordenador.loja@grupobeea.com.br",
                        "coordenador.vd@grupobeea.com.br",
                        "gerente.comercial@grupobeea.com.br"]:
            self.assertNotIn(de_fora, emails,
                             f"{de_fora} nao foi autorizado para o Painel Financeiro")

    def test_modelo_e_painel_cadastrados(self):
        self.assertIn("🏦 Relatório de Custos - ADM/Financeiro", self.ns["MODELOS_RELATORIO"])
        self.assertIn("_painel_dept_adm", FONTE)
        self.assertIn('"🏦 Relatório de Custos - ADM/Financeiro": _painel_dept_adm', FONTE)


# ============================================================================
# 5d-bis. COORDENACOES DE LOJA E DE VD
# ============================================================================
class TesteCoordenacoes(unittest.TestCase):
    """As duas coordenacoes leem da receita ao EBITDA e respondem so pelas
    abas delas."""

    DRE = ["1 - Receita Operacional Bruta", "2 - Deduções da Receita Operacional Bruta",
           "3 - Receita Operacional Liquida", "4 - Custo das Vendas",
           "5 - Margem de Contribuição 1", "6 - Despesas Variáveis",
           "6.1 - Comissões sobre Vendas", "7 - Margem de Contribuição 2",
           "8 - Despesas Operacionais", "9 - Resultado Operacional", "11 - EBITDA",
           "12 - Resultado Financeiro", "13 - Depreciação e Amortização",
           "15 - Resultado Antes do Imposto", "16 - Impostos sobre o Lucro",
           "17 - Resultado Gerencial do Período"]
    LOJA = "🏬 Relatório de Custos - Coordenação de Loja"
    VD = "🚗 Relatório de Custos - Coordenação de VD"

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_normalizar_texto", "_numero_linha_dre", "_linha_pertence_ao_grupo",
             "_linhas_ate_ebitda", "_resolver_termo_departamento"],
            ["MODELOS_RELATORIO", "GRUPO_EBITDA_DRE"],
        )

    def test_para_no_ebitda(self):
        """Abaixo do 11 fica o que a operacao da loja nao decide: resultado
        financeiro, depreciacao, imposto sobre o lucro."""
        achadas = self.ns["_resolver_termo_departamento"]("ATE_EBITDA", self.DRE)
        numeros = [self.ns["_numero_linha_dre"](l) for l in achadas]
        self.assertIn("11", numeros)
        self.assertIn("6.1", numeros, "sublinha do grupo 6 tem de entrar")
        for abaixo in ["12", "13", "15", "16", "17"]:
            self.assertNotIn(abaixo, numeros, f"a linha {abaixo} vem depois do EBITDA")

    def test_cada_coordenacao_so_ve_as_abas_dela(self):
        loja = self.ns["MODELOS_RELATORIO"][self.LOJA]
        vd = self.ns["MODELOS_RELATORIO"][self.VD]
        self.assertEqual(len(loja["unidades_permitidas"]), 13)
        self.assertEqual(len(vd["unidades_permitidas"]), 7)  # 5 VD + 2 ABPR
        self.assertNotIn("DRE CONSOLIDADO", loja["visoes_permitidas"])
        self.assertNotIn("DRE CONSOLIDADO", vd["visoes_permitidas"])
        # Nenhuma unidade pode aparecer nas duas coordenacoes.
        self.assertFalse(
            set(loja["unidades_permitidas"]) & set(vd["unidades_permitidas"]),
            "a mesma loja apareceria nas duas coordenacoes",
        )

    def test_abas_batem_com_as_visoes_consolidadas(self):
        """As listas do modelo tem de bater com VISOES_CONSOLIDADAS, que e a
        fonte unica usada pelo painel e pelo Excel -- se uma loja nova entrar
        so num lugar, o consolidado e o ranking param de fechar."""
        ns = carregar([], ["VISOES_CONSOLIDADAS"])
        visoes = ns["VISOES_CONSOLIDADAS"]
        loja = self.ns["MODELOS_RELATORIO"][self.LOJA]
        vd = self.ns["MODELOS_RELATORIO"][self.VD]
        self.assertEqual(sorted(loja["unidades_permitidas"]), sorted(visoes["LJ - G&A"]))
        self.assertEqual(sorted(vd["unidades_permitidas"]),
                         sorted(visoes["VD CONSOLIDADO"] + visoes["ABPR CONSOLIDADO"]))
        for nome_visao in loja["visoes_permitidas"] + vd["visoes_permitidas"]:
            self.assertIn(nome_visao, visoes, f"{nome_visao} nao existe no mapa de visoes")

    def test_diario_e_recortada_pela_visao(self):
        """Numero vindo da DIARIO (Mercadorias, benfeitorias, planos fora da
        DRE) tem de respeitar a visao escolhida -- senao aparece o valor da
        empresa inteira ao lado de linhas da DRE de uma loja so."""
        ns = carregar(["_normalizar_nome_aba", "_normalizar_texto",
                       "_lojas_individuais_das_abas", "_recortar_diario_por_loja"],
                      ["VISOES_CONSOLIDADAS"])
        self.assertEqual(ns["_lojas_individuais_das_abas"](["LJ PVH1 11927"]), ["LJ PVH1 11927"])
        self.assertEqual(len(ns["_lojas_individuais_das_abas"](["LJ - G&A"])), 13)
        self.assertEqual(ns["_lojas_individuais_das_abas"]([]), [])
        diario = pd.DataFrame({
            "Loja": ["LJ PVH1 11927", "LJ SETE 6052", "ABPR 23427"],
            "Plano de Contas": ["Mercadorias"] * 3,
            "Valor Bruto": [-100.0, -200.0, -300.0],
        })
        recorte = ns["_recortar_diario_por_loja"](diario, ["LJ PVH1 11927"])
        self.assertEqual(len(recorte), 1)
        self.assertEqual(recorte["Valor Bruto"].sum(), -100.0)
        # Nome que nao existe na DIARIO: devolve tudo, nunca zero.
        self.assertEqual(len(ns["_recortar_diario_por_loja"](diario, ["LOJA QUE NAO EXISTE"])), 3)

    def test_desvio_e_realizado_menos_orcado(self):
        """Um so significado para "Desvio" na tela inteira: realizado menos
        orcado, igual ao relatorio em Excel. Positivo = gastou mais."""
        self.assertIn("Realizado menos orçado · positivo = gastou mais", FONTE)
        self.assertNotIn('subtext="Positivo = gastou menos que o orçado"', FONTE)
        self.assertIn("def _cor_valor_invertido(", FONTE)
        # Nenhum bloco pode ter voltado para orcado menos realizado.
        for invertido in ['"Desvio (R$)": orc - real,', '"Desvio (R$)": orcado - valor,',
                          '"Desvio (R$)": v_o - v_r,']:
            self.assertNotIn(invertido, FONTE, f"convencao antiga de volta: {invertido}")

    def test_linha_de_percentual_do_ebitda(self):
        for nome in (self.LOJA, self.VD):
            self.assertTrue(self.ns["MODELOS_RELATORIO"][nome].get("linha_percentual_ebitda"))
        i = FONTE.index("if percentual_ebitda is not None:")
        trecho = FONTE[i:i + 900]
        self.assertIn("11 - EBITDA (%)", trecho)
        self.assertIn("pontos percentuais", FONTE,
                      "a legenda precisa avisar que o desvio da linha e em p.p.")

    def test_uma_aba_anual_por_visao_consolidada(self):
        """Loja tem 2 visoes consolidadas, VD tem 3 -- e cada uma vira uma aba
        de resumo anual no Excel."""
        loja = self.ns["MODELOS_RELATORIO"][self.LOJA]
        vd = self.ns["MODELOS_RELATORIO"][self.VD]
        self.assertTrue(loja.get("resumos_anuais_por_visao"))
        self.assertTrue(vd.get("resumos_anuais_por_visao"))
        self.assertEqual(len(loja["visoes_permitidas"]), 2)
        self.assertEqual(len(vd["visoes_permitidas"]), 3)
        self.assertIn("def _criar_aba_resumo_anual(", FONTE)
        i = FONTE.index("def _criar_aba_resumo_anual(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("ORÇADO", trecho)
        self.assertIn("REALIZADO", trecho)
        self.assertIn("EBITDA (%)", trecho)

    def test_percentual_anual_e_mes_a_mes(self):
        """O EBITDA % de cada mes sai do proprio mes, e o total sai do
        acumulado -- media dos meses daria outro numero quando o faturamento
        e desigual entre eles."""
        ebitda = [419_362.50, 382_194.63, 355_756.86]
        receita = [1_934_279.55, 1_708_908.50, 2_097_982.47]
        por_mes = [e / r * 100 for e, r in zip(ebitda, receita)]
        total = sum(ebitda) / sum(receita) * 100
        self.assertAlmostEqual(por_mes[0], 21.68, places=1)
        self.assertNotAlmostEqual(total, sum(por_mes) / 3, places=1)
        i = FONTE.index("def _criar_aba_resumo_anual(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("pares.append((sum(ebitda[indice]), sum(receita[indice])))", trecho)

    def test_relatorio_so_traz_plano_com_linha_da_dre(self):
        for nome in (self.LOJA, self.VD):
            self.assertTrue(
                self.ns["MODELOS_RELATORIO"][nome].get("apenas_planos_com_linha_dre"))
        i = FONTE.index("if apenas_planos_com_linha_dre")
        trecho = FONTE[i:i + 400]
        self.assertIn("MARCAS_SEM_LINHA_DRE", trecho,
                      "o filtro precisa usar o mesmo helper de celula vazia")

    def test_relatorio_so_oferece_as_abas_do_departamento(self):
        i = FONTE.index("opcoes_lojas_relatorio = _filtrar_abas_permitidas")
        trecho = FONTE[i - 600:i + 200]
        self.assertIn('_modelo_rel.get("visoes_permitidas"', trecho)
        self.assertIn('_modelo_rel.get("unidades_permitidas"', trecho)

    def test_barra_lateral_filtra_as_abas(self):
        self.assertIn("_filtrar_abas_permitidas", FONTE)
        self.assertIn('_modelo_dept_ativo.get("visoes_permitidas")', FONTE)
        self.assertIn('_modelo_dept_ativo.get("unidades_permitidas")', FONTE)


# ============================================================================
# 5e. PRAZOS DO FLUXO
# ============================================================================
class TestePrazosDoFluxo(unittest.TestCase):
    """A baixa dos titulos nao vem no CSV do fluxo; ela e buscada na aba
    DIARIO pelo numero do documento. O cruzamento precisa cobrir os DOIS
    lados -- quando olhava so o "pagar", o atraso medio no recebimento
    aparecia vazio na tela."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["_normalizar_coluna_fin", "_chave_numero_fin",
                           "_classificar_movimento_fin"])

    def _cruzar(self):
        df = pd.DataFrame({
            "Movimento": ["4 - Contas a Pagar", "2 - Contas a Receber", "1 - Banco"],
            "Número": ["1001", "5001", ""],
            "Data Liquidação": [pd.NaT, pd.NaT, pd.NaT],
            "Vencimento.1": pd.to_datetime(["2026-08-05", "2026-08-03", None]),
        })
        diario = pd.DataFrame({
            "chave_num": ["1001", "5001"],
            "data_liq": pd.to_datetime(["2026-08-07", "2026-08-06"]),
        })
        pagar = df["Movimento"].astype(str).str.contains("pagar", case=False, na=False)
        receber = df["Movimento"].astype(str).str.contains("receber", case=False, na=False)
        faltando = (pagar | receber) & df["Data Liquidação"].isna()
        a_cruzar = pd.DataFrame({
            "_idx": df.index[faltando],
            "chave_num": self.ns["_chave_numero_fin"](df.loc[faltando, "Número"]).values,
        })
        casados = a_cruzar.merge(diario, on="chave_num", how="left").dropna(subset=["data_liq"])
        df["Liquidação Efetiva"] = pd.NaT
        df.loc[casados["_idx"].values, "Liquidação Efetiva"] = casados["data_liq"].values
        df["Tipo Movimento"] = df["Movimento"].map(self.ns["_classificar_movimento_fin"])
        return df

    def test_o_receber_tambem_ganha_data_de_baixa(self):
        df = self._cruzar()
        entradas = df[df["Tipo Movimento"] == "entrada"]
        self.assertTrue(entradas["Liquidação Efetiva"].notna().all(),
                        "titulo a receber ficou sem data de baixa")

    def test_atraso_sai_dos_dois_lados(self):
        df = self._cruzar()
        prazo = df[df["Liquidação Efetiva"].notna() & df["Vencimento.1"].notna()].copy()
        prazo["DiasAteLiquidar"] = (prazo["Liquidação Efetiva"] - prazo["Vencimento.1"]).dt.days
        for tipo in ("saida", "entrada"):
            lado = prazo[prazo["Tipo Movimento"] == tipo]
            self.assertFalse(lado.empty, f"nenhum titulo de {tipo} com prazo calculado")
            self.assertEqual(lado["DiasAteLiquidar"].mean(), 2 if tipo == "saida" else 3)

    def test_as_duas_pontas_olham_a_mesma_coluna(self):
        """A pagar e a receber precisam ler a MESMA coluna de baixa; se uma
        delas olhar outra fonte, os dois lados param de conversar."""
        i = FONTE.index("mask_receber = df[COL_FIN_MOVIMENTO]")
        trecho = FONTE[i - 300:i + 600]
        self.assertIn("pagar_sem_liq_no_csv", trecho)
        self.assertIn("receber_sem_liq_no_csv", trecho)
        self.assertIn("COL_FIN_DATA_LIQUIDACAO", trecho)

    def test_baixa_vem_da_propria_aba_do_fluxo(self):
        """A aba Fluxo de Caixa 2026 passou a trazer a Data de Liquidacao
        preenchida, entao o confronto vencimento x liquidacao e feito ali
        mesmo. O cruzamento com a aba DIARIO -- que casava por numero,
        vencimento e valor -- foi removido e nao pode voltar sem decisao."""
        for resto in ("carregar_liquidacoes_diario", "COL_FIN_LIQ_DIARIO",
                      "liq_do_diario", "por_num_venc_valor", "pagar_sem_match"):
            self.assertNotIn(resto, FONTE, f"sobrou pedaco do cruzamento: {resto}")
        i = FONTE.index("df[COL_FIN_LIQ_EFETIVA] = df[COL_FIN_DATA_LIQUIDACAO]")
        trecho = FONTE[i:i + 300]
        self.assertIn("COL_FIN_LIQ_AMPLA", trecho,
                      "a leitura tolerante da MESMA coluna continua valendo")

    def test_titulo_sem_baixa_fica_no_vencimento(self):
        """Sem data de liquidacao o titulo esta em aberto: continua
        posicionado no vencimento, que e a programacao."""
        df = pd.DataFrame({
            "Movimento": ["4 - Contas a Pagar"] * 3,
            "Vencimento.1": pd.to_datetime(["2026-08-31"] * 3),
            "Data Liquidação": pd.to_datetime(["2026-08-24", None, None]),
            "Valor.1": [-300_000.0, -60_000.0, -40_000.0],
        })
        df["Liquidação Efetiva"] = df["Data Liquidação"]
        df["Data Efetiva"] = df["Vencimento.1"]
        com_baixa = df["Liquidação Efetiva"].notna()
        df.loc[com_baixa, "Data Efetiva"] = df.loc[com_baixa, "Liquidação Efetiva"]
        por_dia = df.groupby(df["Data Efetiva"].dt.date)["Valor.1"].sum()
        self.assertAlmostEqual(por_dia[date(2026, 8, 24)], -300_000.0, places=2)
        self.assertAlmostEqual(por_dia[date(2026, 8, 31)], -100_000.0, places=2)

    def test_indicador_nao_se_chama_prazo_medio(self):
        """O calculo e liquidacao menos vencimento, ou seja ATRASO. Chamar de
        prazo medio faz ler "+5 dias" como se a empresa pagasse em 5 dias."""
        self.assertIn("ATRASO MÉDIO NO PAGAMENTO", FONTE)
        self.assertIn("ATRASO MÉDIO NO RECEBIMENTO", FONTE)
        self.assertNotIn('label="PRAZO MÉDIO DE PAGAMENTO"', FONTE)
        self.assertNotIn('label="PRAZO MÉDIO DE RECEBIMENTO"', FONTE)


# ============================================================================
# 5f. REVISAO DOS PAINEIS (17/08/2026)
# ============================================================================
class TesteRevisaoDosPaineis(unittest.TestCase):
    """Travas do que a revisao de 17/08 corrigiu -- sao regras que somem
    facil na proxima mexida e que produzem numero errado sem dar erro."""

    def test_orcado_do_mes_corrente_entra_proporcional(self):
        ns = carregar(["_fator_proporcional_mes_corrente", "_escalar_orcado_mes_corrente"])
        orc = [pd.DataFrame({"Nome": ["11 - EBITDA"], "07/2026": [100_000.0],
                             "08/2026": [310_000.0]})]
        cols = ["07/2026", "08/2026"]
        saida, aviso = ns["_escalar_orcado_mes_corrente"](orc, cols, cols, date(2026, 8, 18))
        self.assertEqual(saida[0]["07/2026"].iloc[0], 100_000.0, "mes fechado nao pode encolher")
        self.assertAlmostEqual(saida[0]["08/2026"].iloc[0], 310_000 * 18 / 31, places=2)
        self.assertIn("18 de 31 dias", aviso)

    def test_periodo_fechado_nao_mexe_no_orcado(self):
        ns = carregar(["_fator_proporcional_mes_corrente", "_escalar_orcado_mes_corrente"])
        orc = [pd.DataFrame({"Nome": ["11 - EBITDA"], "07/2026": [100_000.0]})]
        saida, aviso = ns["_escalar_orcado_mes_corrente"](orc, ["07/2026"], ["07/2026"],
                                                          date(2026, 8, 18))
        self.assertIs(saida, orc)
        self.assertEqual(aviso, "")

    def test_mkt_acumula_o_fora_do_1pct(self):
        """Se um segundo canal passar a descontar, o valor do primeiro nao
        pode sumir -- o bug seria silencioso."""
        i = FONTE.index("def _painel_dept_mkt(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("fora_do_1pct += fora_do_canal", trecho)
        self.assertNotIn("fora_do_1pct = sum(", trecho)

    def test_reguas_do_rh_estao_declaradas(self):
        ns = carregar([], ["LIMITE_HORA_EXTRA_PCT_FOLHA", "LIMITE_RESCISOES_PCT_FOLHA"])
        self.assertEqual(ns["LIMITE_HORA_EXTRA_PCT_FOLHA"], 5.0)
        self.assertEqual(ns["LIMITE_RESCISOES_PCT_FOLHA"], 3.0)
        i = FONTE.index("def _painel_dept_rh(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertNotIn("* 100 - 5)", trecho, "limite voltou a ficar cravado no codigo")
        self.assertIn("régua", trecho, "a regua precisa aparecer na tela")

    def test_rh_nao_zera_o_orcado_do_resto(self):
        i = FONTE.index("def _painel_dept_rh(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("outros_orc = folha_orc_total", trecho)
        self.assertNotIn('"Orçado (R$)": 0.0,', trecho)

    def test_adm_ranqueia_por_valor_e_por_percentual(self):
        """So por valor, o custo da venda lidera sempre e enterra as
        distorcoes cronicas pequenas."""
        i = FONTE.index("def _painel_dept_adm(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("por_percentual", trecho)
        self.assertIn("piso_relevante", trecho)

    def test_bloco_que_carrega_planilha_usa_a_mesma_regua_de_orcado(self):
        """MKT (tabela por canal) e as coordenacoes (ranking por unidade)
        carregam planilha por conta propria. Se nao aplicarem o mesmo ajuste
        de mes corrente dos cartoes do topo, a tabela mostra o orcamento
        cheio e o cabecalho a parte decorrida -- foi o que apareceu na tela
        em 18/08/2026, com R$ 64 mil de diferenca."""
        self.assertIn('"escalar_orcado": lambda dfs: _escalar_orcado_mes_corrente(', FONTE)
        i = FONTE.index("def _painel_dept_mkt(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn('escalar = ctx.get("escalar_orcado")', trecho)
        self.assertIn("df_o = escalar([df_o])[0]", trecho)
        j = FONTE.index("def _painel_dept_coordenacao(")
        corpo = FONTE[j:FONTE.index("\ndef ", j + 10)]
        self.assertIn('dfs_orc_un = ctx["escalar_orcado"](dfs_orc_un)', corpo)

    def test_mkt_avisa_quando_a_tabela_nao_fecha(self):
        """A tabela por canal e os cartoes do topo saem por caminhos
        diferentes e por isso um valida o outro. Quando divergem, a tela tem
        de dizer -- melhor que descobrir na reuniao."""
        i = FONTE.index("def _painel_dept_mkt(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn('_total_linha = next((l for l in linhas_tabela if l["Canal"] == "TOTAL")', trecho)
        self.assertIn("não está fechando com os cartões do topo", trecho)
        self.assertIn("st.warning(", trecho)

    def test_soma_por_plano_passa_pelo_recorte(self):
        """A funcao que soma planos da DIARIO tem de aplicar o recorte por
        loja. Sem esta trava, tirar a chamada nao quebra teste nenhum: o
        numero fica so maior, e maior parece plausivel."""
        i = FONTE.index("def _total_planos(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("_recortar_diario_por_loja(df_diario, lojas)", trecho)
        # E os tres blocos que a usam precisam passar as lojas do contexto.
        for bloco in ("_painel_dept_compras", "_painel_dept_suprimentos"):
            j = FONTE.index(f"def {bloco}(")
            corpo = FONTE[j:FONTE.index("\ndef ", j + 10)]
            self.assertIn('ctx.get("lojas")', corpo, f"{bloco} soma a empresa inteira")
        j = FONTE.index("def _painel_dept_adm(")
        corpo = FONTE[j:FONTE.index("\ndef ", j + 10)]
        self.assertIn("_recortar_diario_por_loja(df_diario_adm", corpo)

    def test_coordenacao_compara_com_a_mediana(self):
        """Mediana, nao media: com uma unidade fora da curva a media se
        desloca e todas as outras parecem boas."""
        i = FONTE.index("def _painel_dept_coordenacao(")
        trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("p.p. vs mediana", trecho)
        self.assertNotIn("sum(margens) / len(margens)", trecho, "virou media")


# ============================================================================
# 5g. FLUXO MENSAL — SALDO DE ABERTURA
# ============================================================================
class TesteSaldoDeAberturaMensal(unittest.TestCase):
    """Em "Movimentos por Mes", caixa e banco mostram o saldo do PRIMEIRO dia
    do mes. Todo o resto do painel segue com a posicao de fechamento."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["_normalizar_texto", "_classificar_movimento_fin",
                           "_pivot_fluxo_fin"])
        linhas = []
        for dia, saldo in [(1, 500_000.0), (15, 620_000.0), (31, 480_000.0)]:
            linhas.append({"Data Efetiva": pd.Timestamp(2026, 7, dia),
                           "Movimento": "1 - Banco", "Valor.1": saldo})
        for dia, saldo in [(1, 480_000.0), (10, 700_000.0), (18, 655_000.0)]:
            linhas.append({"Data Efetiva": pd.Timestamp(2026, 8, dia),
                           "Movimento": "1 - Banco", "Valor.1": saldo})
        for mes in (7, 8):
            for dia in (5, 20):
                linhas.append({"Data Efetiva": pd.Timestamp(2026, mes, dia),
                               "Movimento": "4 - Contas a Pagar", "Valor.1": -250_000.0})
        cls.df = pd.DataFrame(linhas)
        cls.df["PeriodoMes"] = cls.df["Data Efetiva"].dt.to_period("M")
        cls.meses = sorted(cls.df["PeriodoMes"].unique())

    def _pivo(self, posicao):
        return self.ns["_pivot_fluxo_fin"](
            self.df, "PeriodoMes", "Valor.1", "Movimento", self.meses, posicao_saldo=posicao)

    def test_saldo_do_primeiro_dia_do_mes(self):
        abertura = self._pivo("primeira")
        self.assertEqual(abertura.loc["1 - Banco"].iloc[0], 500_000.0)
        self.assertEqual(abertura.loc["1 - Banco"].iloc[1], 480_000.0)

    def test_padrao_continua_sendo_o_fechamento(self):
        """Quem nao pedir nada tem de continuar recebendo a posicao do ultimo
        dia -- o resto do painel depende disso."""
        fechamento = self.ns["_pivot_fluxo_fin"](
            self.df, "PeriodoMes", "Valor.1", "Movimento", self.meses)
        self.assertEqual(fechamento.loc["1 - Banco"].iloc[0], 480_000.0)
        self.assertEqual(fechamento.loc["1 - Banco"].iloc[1], 655_000.0)

    def test_movimentacao_nao_muda_de_base(self):
        """A troca vale so para linha de SALDO: a pagar e a receber continuam
        somando o mes inteiro nas duas leituras."""
        for posicao in ("primeira", "ultima"):
            pivo = self._pivo(posicao)
            self.assertEqual(pivo.loc["4 - Contas a Pagar"].iloc[0], -500_000.0)

    def test_coluna_final_do_mensal_traz_a_ultima_posicao(self):
        """Caixa e banco na coluna final mostram a posicao MAIS RECENTE. A
        versao anterior pegava o primeiro mes com saldo e exibia julho
        quando agosto ja tinha posicao. Meses de previsao vem zerados, e
        zero e ausencia de dado, nao posicao -- por isso nao contam."""
        caixa = pd.Series({"JUL": 14_184.65, "AGO": 20_318.50, "SET": 0.0, "OUT": 0.0})
        nao_zerados = caixa[caixa != 0]
        self.assertEqual(nao_zerados.iloc[-1], 20_318.50)
        i = FONTE.index("totais_finais_m = {}")
        trecho = FONTE[i:i + 900]
        self.assertIn("nao_zerados.iloc[-1]", trecho,
                      "a coluna final tem de pegar a ultima posicao, nao a primeira")
        self.assertNotIn("nao_zerados.iloc[0]", trecho)

    def test_escrita_do_mapa_nunca_sai_vazia_em_silencio(self):
        """O mapa chegava VAZIO ao navegador porque a escrita falhava e o
        erro morria numa variavel. Agora ha uma ultima defesa para tipos que
        o JSON nao conhece, e o erro viaja ate a tela."""
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("default=_valor_para_json", corpo,
                      "sem isto, um tipo desconhecido zera o mapa inteiro")
        self.assertIn('data-erro-mapa=', corpo, "o erro precisa chegar na tela")
        self.assertIn("erroDoMapa", corpo)

    def test_saldo_inicial_mensal_so_vale_para_previsao(self):
        """Mes que ja passou, e o mes corrente, tem as posicoes reais de caixa
        e banco DENTRO deles: somar um saldo de abertura por cima contaria o
        mesmo dinheiro duas vezes. Da frente em diante e previsao, e ai o mes
        parte do que sobrou do anterior."""
        mes_corrente = pd.Period("2026-08", "M")
        meses = [pd.Period(m, "M") for m in ("2026-07", "2026-08", "2026-09", "2026-10")]
        movimentos = {"JUL": 4_200.0, "AGO": 5_037.0, "SET": -300.0, "OUT": -400.0}
        colunas = list(movimentos)

        saldos, acumulado = {}, 0.0
        for periodo, coluna in zip(meses, colunas):
            if periodo <= mes_corrente:
                saldos[coluna] = 0.0
                acumulado = movimentos[coluna]
            else:
                saldos[coluna] = acumulado
                acumulado += movimentos[coluna]

        self.assertEqual(saldos["JUL"], 0.0, "mes realizado nao tem abertura")
        self.assertEqual(saldos["AGO"], 0.0, "mes corrente nao tem abertura")
        self.assertEqual(saldos["SET"], 5_037.0, "previsao parte do fechamento anterior")
        self.assertEqual(saldos["OUT"], 4_737.0, "e segue em cadeia")

        i = FONTE.index("📋 Movimentos por Mês")
        trecho = FONTE[max(0, i - 6000):i + 3000]
        self.assertIn("if _periodo <= _mes_corrente_m:", trecho,
                      "a regra precisa distinguir passado/corrente de previsao")
        self.assertIn('index=["SALDO INICIAL"]', trecho)

    def test_mensal_e_reserva_leem_pela_mesma_abertura(self):
        """25/08/2026: a Reserva de Caixa passou a ler caixa e banco pela
        ABERTURA, igual a tabela de cima. Com isso o pivo de fechamento, que
        so ela consumia, deixou de existir -- e nao pode voltar por descuido,
        porque a volta dele traz a dupla contagem junto."""
        # Janela ampla: o bloco cresceu quando o SALDO INICIAL entrou, e uma
        # janela curta faz o teste falhar por posicao, nao por defeito.
        i = FONTE.index("📋 Movimentos por Mês")
        trecho = FONTE[max(0, i - 6000):i + 3000]
        self.assertIn('posicao_saldo="primeira"', trecho)
        self.assertNotIn("pivot_m_fechamento", FONTE,
                         "o pivo de fechamento voltou -- a Reserva usa a abertura")
        i_reserva = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        bloco_reserva = FONTE[max(0, i_reserva - 4500):i_reserva]
        # O disponivel sai das linhas de SALDO e ENTRADA (caixa, banco,
        # a receber, liquidado) -- nunca de TODAS, o que arrastaria a META.
        self.assertIn("_saldos_abertura_m", bloco_reserva)
        self.assertIn("_entradas_do_mes_m", bloco_reserva)
        # 25/08/2026: a heranca do mes anterior e a sobra CONTRATADA -- a meta
        # nao pode entrar nela. Ver test_reserva_herda_so_a_sobra_contratada.
        self.assertIn("_heranca_m", bloco_reserva)
        self.assertIn("_periodo > _mes_corrente_m", bloco_reserva,
                      "so mes de previsao recebe meta e heranca")

    def test_reserva_monta_cada_parte_da_sua_fonte(self):
        """Regra da area, LEITURA DE FLUXO (25/08/2026):
          caixa e banco  -> ABERTURA do mes (primeiro dia)
          a receber e liquidado -> TOTAL do mes
          a pagar        -> TOTAL do mes
          meta           -> SO nos meses de previsao (ver o teste da meta)
        A abertura e o unico saldo que conversa com os totais do mes sem contar
        o mesmo dinheiro duas vezes."""
        # Recorte entre MARCOS, nao por janela de N caracteres: a janela
        # quebrou quando a linha de liquidez entrou entre o calculo e o titulo.
        i = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        ini = FONTE.rindex("_saldos_abertura_m = ", 0, i)
        bloco = FONTE[ini:i]
        self.assertIn("pivot_m.loc[_saldos_abertura_m, coluna]", bloco,
                      "caixa e banco vem da ABERTURA")
        self.assertIn("pivot_m.loc[_entradas_do_mes_m, coluna]", bloco,
                      "a receber e liquidado somam o mes")
        self.assertIn('_classificar_movimento_fin(m) == "saldo"', bloco)
        self.assertIn('_classificar_movimento_fin(m) == "entrada"', bloco)
        self.assertNotIn("pivot_m_fechamento", bloco,
                         "o pivo de fechamento traz a dupla contagem de volta")

    def test_a_reserva_tem_a_linha_de_liquidez_imediata(self):
        """Disponivel / a pagar, mostrado como o que EXCEDE a divida: 30% quer
        dizer R$ 1,30 disponivel para cada R$ 1,00 devido. Assim os 30% da meta
        significam a mesma coisa nas duas linhas, e ninguem precisa lembrar que
        numa delas 30 quer dizer 130."""
        self.assertIn('LINHA_LIQUIDEZ = "Índice de liquidez imediata"', FONTE)
        self.assertIn("((disp / abs(pagar) - 1) * 100)", FONTE,
                      "o indice tem de descontar o 1,00 da divida")
        # Mes sem nada a pagar NAO tem indice: dividir por zero daria folga
        # infinita, e a tela mostraria isso como se fosse conquista.
        i = FONTE.index("serie_liquidez = pd.Series(")
        self.assertIn("if (abs(pagar) > 0 and disp > 0) else float(\"nan\")",
                      FONTE[i:i + 500])
        # E ela e cobrada contra a MESMA meta da linha de cima.
        self.assertIn("v >= META_RESERVA_PADRAO", FONTE)
        # Calcular e nao mostrar deixava a trava verde com a linha ausente da
        # tela: a serie tem de entrar na tabela E no indice dela.
        self.assertIn("serie_liquidez.map(_formata_pct_sobra)", FONTE,
                      "a linha foi calculada mas nao entra na tabela")
        self.assertIn("LINHA_PCT_SOBRA, LINHA_LIQUIDEZ]", FONTE,
                      "a linha nao entra no indice da tabela")

    def test_a_liquidez_entra_no_grafico_da_reserva(self):
        """A linha existia so na tabela. No grafico ela e tracejada de
        proposito: divide pela DIVIDA enquanto a laranja divide pelo
        disponivel, e dois tracos iguais no mesmo eixo fariam quem olha de
        longe somar as duas leituras numa so."""
        self.assertIn('name="Liquidez imediata"', FONTE)
        self.assertIn("liquidez_grafico = [float(serie_liquidez[c]) for c in rotulos_x_m]",
                      FONTE)
        i = FONTE.index('name="Liquidez imediata"')
        self.assertIn('dash="dot"', FONTE[i:i + 400],
                      "a linha de liquidez ficou igual a de sobra")
        # E a faixa do eixo tem de considerar as DUAS linhas, senao a de
        # liquidez sai do desenho nos meses em que ela e mais alta.
        self.assertIn("pct_sobra_grafico + liquidez_grafico", FONTE)
        # Cobrada contra a MESMA meta.
        self.assertIn("v < META_RESERVA_PADRAO", FONTE)


    def test_o_grafico_da_reserva_nao_tem_mais_a_linha_de_resultado(self):
        """Ela era a diferenca entre as duas barras que estao logo ali, entao
        nao acrescentava leitura -- so puxava a escala da esquerda para baixo
        (ia a -10 milhoes) e achatava as barras contra o topo do quadro."""
        self.assertNotIn('name="Resultado do mês"', FONTE)
        self.assertNotIn("resultado_mes = [e - s for e, s in zip(", FONTE)
        # E a faixa da esquerda passa a comecar em ZERO.
        self.assertIn("range=[0, teto_barras * 1.20]", FONTE)

    def test_os_rotulos_das_duas_linhas_nao_se_cobrem(self):
        """Com dois tracados no mesmo eixo, "top center" punha o numero da
        sobra em cima do ponto da liquidez em todo mes em que as duas se
        cruzam."""
        # Os rotulos viraram ANOTACOES com fundo: texto de trace nao aceita
        # cor de fundo, e sem fundo o numero sumia sempre que a linha passava
        # por cima de uma barra -- foi o que escondeu o "60%" de agosto.
        self.assertIn("anotacoes_pct = []", FONTE)
        # AS DUAS anotacoes precisam do fundo: cobrar so uma deixava a outra
        # ser mutilada em silencio.
        self.assertEqual(FONTE.count('bgcolor="rgba(36,44,60,0.82)"'), 2,
                         "algum rotulo perdeu o fundo e volta a sumir atras da barra")
        self.assertIn("annotations=anotacoes_pct", FONTE,
                      "as anotacoes foram montadas mas nao entram no grafico")
        # Sobra EMBAIXO do ponto, liquidez em CIMA: assim as duas nunca se
        # cobrem nos meses em que os tracados se cruzam.
        i = FONTE.index("for _x, _v in zip(rotulos_x_m, pct_sobra_grafico):")
        self.assertIn("yshift=-15", FONTE[i:i + 500])
        j = FONTE.index("for _x, _v in zip(rotulos_x_m, liquidez_grafico):")
        self.assertIn("yshift=15", FONTE[j:j + 500])


    def test_a_liquidez_nao_repete_a_conta_da_sobra(self):
        """As duas chegam a 30% na meta, mas dividem por coisas diferentes: a
        de cima pelo DISPONIVEL, a de baixo pela DIVIDA. Modelo da conta."""
        disp, pagar = 130.0, 100.0
        pct_sobra = (disp - pagar) / disp * 100          # 23,1%
        liquidez = (disp / pagar - 1) * 100              # 30,0%
        self.assertAlmostEqual(liquidez, 30.0, places=6)
        self.assertNotAlmostEqual(pct_sobra, liquidez, places=1)


    def test_meta_entra_no_disponivel_so_na_previsao(self):
        """25/08/2026, EXCECAO UNICA a regra de que a meta nunca entra em
        total. Mes que ainda nao aconteceu tem so a cauda das parcelas ja
        emitidas no a receber (dezembro tinha R$ 1,3 mi), enquanto o a pagar ja
        esta quase completo -- a tabela comparava entrada incompleta com saida
        completa e chamava de rombo. Mes realizado e mes corrente seguem sem
        meta: la existe realizado de verdade, e a meta so inflaria."""
        # Recorte entre MARCOS, nao por janela de N caracteres: a janela
        # quebrou quando a linha de liquidez entrou entre o calculo e o titulo.
        i = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        ini = FONTE.rindex("_saldos_abertura_m = ", 0, i)
        bloco = FONTE[ini:i]
        self.assertIn("serie_meta_a_realizar", bloco)
        self.assertIn('_classificar_movimento_fin(m) == "meta"', bloco)
        # A condicao de previsao tem de estar na propria montagem da serie.
        trecho_meta = bloco[bloco.index("serie_meta_a_realizar = pd.Series("):]
        self.assertIn("_periodo > _mes_corrente_m", trecho_meta[:600],
                      "a meta nao pode entrar em mes realizado nem no corrente")
        # E tem de entrar DE FATO na soma do disponivel, nao so existir: o
        # trecho entre a montagem do disponivel e a da sobra e o unico lugar
        # onde ela pode somar.
        soma_disponivel = bloco[bloco.index("serie_disponivel_total = pd.Series("):
                                bloco.index("serie_sobra = ")]
        self.assertIn("serie_meta_a_realizar[coluna]", soma_disponivel,
                      "a meta existe mas nao esta entrando no disponivel")
        # E a linha tem de ficar VISIVEL na tabela, nao embutida no disponivel.
        depois = FONTE[i:i + 6000]
        self.assertIn("LINHA_META_RESERVA", depois)
        self.assertIn("meta ainda a realizar", depois)

    def test_heranca_e_a_sobra_do_mes_anterior(self):
        """A heranca de um mes de previsao e EXATAMENTE a Sobra da coluna
        anterior -- a mesma que a tabela mostra. A versao anterior passava
        adiante so a sobra "contratada" (sem a meta), e o resultado ficava
        contraditorio na tela: outubro mostrava sobra de R$ 4,38 mi e novembro
        herdava R$ 0,00."""
        i = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        bloco = FONTE[max(0, i - 9000):i]
        # A conta que avanca a cadeia tem de conter TODAS as parcelas da sobra.
        # Recorte por marco do codigo, e nao pelo primeiro ")" -- a expressao
        # tem parenteses internos e o corte ingenuo mutilava o trecho.
        avanco = bloco[bloco.index("_sobra_anterior = ("):
                       bloco.index("serie_heranca = pd.Series(")]
        for parcela in ("_heranca_m[_coluna]", "_abertura_col", "_entradas_col",
                        "serie_meta_a_realizar[_coluna]", "serie_a_pagar"):
            self.assertIn(parcela, avanco,
                          f"a sobra que passa adiante tem de incluir {parcela}")
        # Piso em zero: mes que nao se paga passa ZERO, nao o buraco.
        self.assertIn("max(_sobra_anterior, 0.0) if _periodo > _mes_corrente_m else 0.0",
                      bloco, "faltou o piso em zero ou a regra de so-previsao")
        # E a heranca tem de somar DE FATO no disponivel, nao so existir.
        soma_disponivel = bloco[bloco.index("serie_disponivel_total = pd.Series("):
                                bloco.index("serie_sobra = ")]
        self.assertIn("serie_heranca[coluna]", soma_disponivel,
                      "a heranca existe mas nao esta entrando no disponivel")
        # E ficar VISIVEL na tabela, como a meta.
        depois = FONTE[i:i + 8000]
        self.assertIn("LINHA_HERANCA", depois)
        self.assertIn("sobra herdada do mês anterior", depois)

    def test_previsao_herda_a_sobra_com_piso_em_zero(self):
        """O numero que a decisao produz, mes a mes. Valores reais de
        25/08/2026. A heranca de cada mes tem de ser IGUAL a sobra do mes
        anterior -- e essa igualdade e a melhor trava que existe aqui, porque
        e ela que o usuario le na tela, coluna contra coluna."""
        mes_corrente = pd.Period("2026-08", "M")
        meses = [pd.Period(f"2026-{m:02d}", "M") for m in range(7, 13)]
        # abertura de caixa/banco + a receber + a receber liquidado do mes
        proprio = [11_978_614.40, 13_567_638.60, 8_350_298.44,
                   5_504_942.44, 3_019_686.31, 1_295_467.83]
        metas = [4_300_482.26, 7_776_935.86, 10_462_221.58,
                 12_111_597.35, 13_128_339.38, 14_685_042.65]
        a_pagar = [9_704_926.16, 8_625_812.77, 9_336_790.37,
                   11_689_048.42, 13_785_703.74, 11_298_162.96]

        def roda(proprio, metas, a_pagar, com_piso=True):
            heranca, disponiveis, sobras, anterior = [], [], [], 0.0
            for periodo, entrada, meta, pagar in zip(meses, proprio, metas, a_pagar):
                previsao = periodo > mes_corrente
                base = (max(anterior, 0.0) if com_piso else anterior) if previsao else 0.0
                falta = max(meta - entrada, 0.0) if previsao else 0.0
                disponivel = base + entrada + falta
                heranca.append(base)
                disponiveis.append(disponivel)
                sobras.append(disponivel - pagar)
                anterior = sobras[-1]      # a sobra CHEIA, com a meta dentro
            return heranca, disponiveis, sobras

        heranca, disponiveis, sobras = roda(proprio, metas, a_pagar)

        # Mes realizado e corrente: sem meta e sem heranca.
        self.assertEqual(heranca[0], 0.0)
        self.assertEqual(heranca[1], 0.0)
        self.assertAlmostEqual(sobras[0], 2_273_688.24, places=2)
        self.assertAlmostEqual(sobras[1], 4_941_825.83, places=2)
        # A REGRA, coluna contra coluna: a heranca de cada mes de previsao e a
        # sobra do mes anterior, sem nada tirado no caminho.
        for i in range(2, 6):
            self.assertAlmostEqual(heranca[i], sobras[i - 1], places=2,
                                   msg="a heranca tem de bater com a coluna anterior")
        self.assertAlmostEqual(sobras[2], 6_067_257.04, places=2)
        self.assertAlmostEqual(sobras[3], 6_489_805.97, places=2)
        self.assertAlmostEqual(sobras[4], 5_832_441.61, places=2)
        self.assertAlmostEqual(sobras[5], 9_219_321.30, places=2)

        pcts = [s / d * 100 for s, d in zip(sobras, disponiveis)]
        for esperado, obtido in zip([19.0, 36.4, 39.4, 35.7, 29.7, 44.9], pcts):
            self.assertAlmostEqual(obtido, esperado, places=1)

        # PISO: mes que nao se paga passa ZERO adiante, nao o buraco. Aqui um
        # a pagar monstruoso em outubro derruba o mes -- novembro herda zero e
        # nao carrega o deficit junto.
        pagar_ruim = list(a_pagar)
        pagar_ruim[3] = 40_000_000.0
        heranca_ruim, _, sobras_ruim = roda(proprio, metas, pagar_ruim)
        self.assertLess(sobras_ruim[3], 0.0, "outubro tem de ficar negativo no cenario")
        self.assertEqual(heranca_ruim[4], 0.0, "novembro nao pode herdar o buraco")
        # Sem o piso, o buraco cascatearia para novembro.
        heranca_sem, _, _ = roda(proprio, metas, pagar_ruim, com_piso=False)
        self.assertLess(heranca_sem[4], 0.0)

    def test_meta_ja_batida_nao_soma_nada(self):
        """Piso em zero: se o a receber ja emitido passou da meta do mes, o
        excedente ja esta nas linhas de a receber e a parcela da meta e zero.
        Somar a meta cheia por cima contaria o mesmo recebimento duas vezes."""
        entrada, meta = 12_000_000.0, 10_000_000.0
        self.assertEqual(max(meta - entrada, 0.0), 0.0)
        # A funcao que produz esse numero e a mesma do resto do painel.
        self.assertIn(".clip(lower=0)", FONTE[FONTE.index("def _aplicar_meta_como_falta("):
                                              FONTE.index("def _total_geral_sem_meta(")])

    def test_fechamento_com_total_do_mes_conta_o_dinheiro_duas_vezes(self):
        """O defeito que motivou a mudanca, em numeros. O saldo do ULTIMO dia
        JA E a abertura mais tudo que entrou menos tudo que saiu. Somar esse
        fechamento COM o total de a receber liquidado conta o mesmo dinheiro
        duas vezes; e descontar o total de a pagar desconta segunda vez o que
        ja tinha saido. O excesso na sobra e exatamente o movimento liquido de
        caixa do mes."""
        abertura, liquidado, aberto = 2_000_000.0, 12_000_000.0, 500_000.0
        pago, pagar_aberto = 9_200_000.0, 500_000.0
        fechamento = abertura + liquidado - pago
        self.assertEqual(fechamento, 4_800_000.0)

        sobra_fluxo = (abertura + aberto + liquidado) - (pago + pagar_aberto)
        sobra_posicao = (fechamento + aberto) - pagar_aberto
        sobra_antiga = (fechamento + aberto + liquidado) - (pago + pagar_aberto)

        # As duas leituras honestas dao a MESMA sobra em reais: e algebra,
        # porque o fechamento e a abertura mais o liquido. So o denominador
        # muda -- e com ele a porcentagem.
        self.assertAlmostEqual(sobra_fluxo, sobra_posicao, places=2)
        self.assertAlmostEqual(sobra_fluxo, 4_800_000.0, places=2)
        # A antiga inflava a sobra pelo liquido de caixa do mes.
        self.assertAlmostEqual(sobra_antiga - sobra_fluxo, liquidado - pago, places=2)
        self.assertAlmostEqual(sobra_antiga, 7_600_000.0, places=2)

    def test_disponivel_do_mes_sai_da_abertura_mais_o_que_passou(self):
        """Mesma conta, agora saindo do pivo de verdade: o disponivel de um
        mes fechado tem de dar abertura + a receber + liquidado, e a sobra tem
        de bater com o fechamento menos o que sobrou em aberto."""
        classificar = self.ns["_classificar_movimento_fin"]
        linhas = [
            {"Data Efetiva": pd.Timestamp(2026, 7, 1),
             "Movimento": "1.Banco", "Valor.1": 2_000_000.0},
            {"Data Efetiva": pd.Timestamp(2026, 7, 31),
             "Movimento": "1.Banco", "Valor.1": 4_800_000.0},
            {"Data Efetiva": pd.Timestamp(2026, 7, 10),
             "Movimento": "2.2 - Contas a Receber Liquidado", "Valor.1": 12_000_000.0},
            {"Data Efetiva": pd.Timestamp(2026, 7, 28),
             "Movimento": "2.1 - Contas a Receber", "Valor.1": 500_000.0},
            {"Data Efetiva": pd.Timestamp(2026, 7, 12),
             "Movimento": "3 - Contas a Pagar", "Valor.1": -9_200_000.0},
            {"Data Efetiva": pd.Timestamp(2026, 7, 30),
             "Movimento": "3 - Contas a Pagar", "Valor.1": -500_000.0},
        ]
        df = pd.DataFrame(linhas)
        df["PeriodoMes"] = df["Data Efetiva"].dt.to_period("M")
        meses = sorted(df["PeriodoMes"].unique())
        pivo = self.ns["_pivot_fluxo_fin"](
            df, "PeriodoMes", "Valor.1", "Movimento", meses, posicao_saldo="primeira")
        coluna = meses[0]

        saldos = [m for m in pivo.index if classificar(m) == "saldo"]
        entradas = [m for m in pivo.index if classificar(m) == "entrada"]
        saidas = [m for m in pivo.index if classificar(m) == "saida"]

        disponivel = (float(pivo.loc[saldos, coluna].sum())
                      + float(pivo.loc[entradas, coluna].sum()))
        a_pagar = abs(float(pivo.loc[saidas, coluna].sum()))
        sobra = disponivel - a_pagar

        self.assertAlmostEqual(disponivel, 14_500_000.0, places=2)
        self.assertAlmostEqual(a_pagar, 9_700_000.0, places=2)
        self.assertAlmostEqual(sobra, 4_800_000.0, places=2)
        self.assertAlmostEqual(sobra / disponivel * 100, 33.10, places=2)
        # E a mesma sobra da leitura de posicao: fechamento (4,8 mi) mais o
        # que ainda ha para receber, menos o que ainda ha para pagar.
        self.assertAlmostEqual(sobra, (4_800_000.0 + 500_000.0) - 500_000.0, places=2)

    def test_sobra_da_reserva_e_o_total_geral_do_mensal(self):
        """As duas tabelas leem pela mesma abertura, entao a linha Sobra e o
        TOTAL GERAL da tabela de cima sao o MESMO numero -- nos meses
        REALIZADOS e no CORRENTE. Enquanto divergiam ali, a diferenca era o
        sintoma da dupla contagem.

        Nos meses de PREVISAO elas divergem de proposito, exatamente pelo tanto
        que falta para a meta: a tabela de cima mostra so o contratado, a
        Reserva mostra o cenario com a meta cumprida."""
        abertura, a_receber, liquidado, a_pagar = 2_000_000.0, 500_000.0, 12_000_000.0, -9_700_000.0

        # --- Mes realizado: sem meta, os dois numeros batem. ---
        total_geral = abertura + a_receber + liquidado + a_pagar
        disponivel = abertura + a_receber + liquidado
        self.assertAlmostEqual(disponivel - abs(a_pagar), total_geral, places=2)

        # --- Mes de previsao: divergem, e por dois motivos somados. ---
        # A Reserva poe a meta e NAO herda nada; a tabela de cima nao poe a
        # meta e herda o contratado do mes anterior.
        falta_meta, herdado = 3_000_000.0, 800_000.0
        total_geral_prev = herdado + a_receber + liquidado + a_pagar
        sobra_prev = (a_receber + liquidado + falta_meta) - abs(a_pagar)
        self.assertAlmostEqual(sobra_prev - total_geral_prev, falta_meta - herdado, places=2)

        # E o codigo tem de continuar montando as duas do mesmo pivo.
        self.assertIn("movimentos_do_mes_m = _total_geral_sem_meta(pivot_m)", FONTE)
        self.assertNotIn("pivot_m_fechamento", FONTE)

    def test_pct_sem_disponivel_fica_vazia_e_nao_inventa_numero(self):
        """Com disponivel zero ou negativo nao ha porcentagem. A divisao com
        os dois lados negativos devolve numero POSITIVO (197% de sobra num mes
        sem dinheiro), e o -100% que ficava no lugar era invencao igual.

        25/08/2026: dezembro apareceu escrito "None" na tela, com o ausente
        atravessando a formatacao -- e nao reproduziu em pandas 2.2 nem 3.0,
        nem na funcao do Streamlit que monta as celulas. Como nao da para
        consertar a camada que nao se sabe qual e, o ausente deixou de chegar
        la: a linha de % entra na tabela JA COMO TEXTO."""
        # Recorte entre MARCOS, nao por janela de N caracteres: a janela
        # quebrou quando a linha de liquidez entrou entre o calculo e o titulo.
        i = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        ini = FONTE.rindex("_saldos_abertura_m = ", 0, i)
        bloco = FONTE[ini:i]
        self.assertIn('if disp > 0 else float("nan")', bloco,
                      "sem disponivel a porcentagem tem de ficar vazia")
        self.assertNotIn("else (0.0 if sobra == 0 else -100.0)", bloco,
                         "o -100% inventado voltou")
        depois = FONTE[i:i + 7000]
        self.assertIn("def _formata_pct_sobra(", depois, "falta o formatador do traco")
        self.assertIn('return "—"', depois)
        # O ausente nao pode voltar a chegar na tabela como numero.
        self.assertIn("serie_pct_sobra.map(_formata_pct_sobra)", depois,
                      "a linha de % tem de entrar na tabela ja como texto")
        self.assertNotIn("subset=pd.IndexSlice[[LINHA_PCT_SOBRA], :]", depois,
                         "formatar a linha de % no Styler deixa o ausente passar")
        # A cor sai da serie NUMERICA, porque na tabela a linha ja e texto.
        self.assertIn("for v in serie_pct_sobra.reindex(linha.index)", depois)
        self.assertIn("not pd.isna(v) and v >= 30", depois)

    def test_map_aplica_o_formatador_tambem_no_ausente(self):
        """O conserto acima depende de Series.map CHAMAR a funcao no valor
        ausente (na_action=None e o padrao). Se isso mudar no pandas, o ausente
        volta a escapar para a tabela -- e este teste avisa antes da tela."""
        def formata(valor):
            return "—" if pd.isna(valor) else f"{valor:.1f}%"

        serie = pd.Series([39.4, float("nan"), 29.7], index=["SET", "OUT", "NOV"])
        texto = serie.map(formata)
        self.assertEqual(list(texto), ["39.4%", "—", "29.7%"])
        self.assertFalse(texto.isna().any(), "nenhum ausente pode sobreviver ao map")

    def test_realizado_e_corrente_nao_herdam_nada(self):
        """Mes que ja passou e o mes corrente tem as posicoes reais de caixa e
        banco DENTRO deles: herdar por cima contaria o mesmo dinheiro duas
        vezes. E a mesma regra do saldo inicial do diario. So a previsao
        herda."""
        i = FONTE.index("Reserva de Caixa — sobra depois de pagar tudo")
        bloco = FONTE[max(0, i - 9000):i]
        self.assertIn("serie_sobra = serie_disponivel_total - serie_a_pagar.abs()", bloco)
        self.assertIn("if _periodo > _mes_corrente_m else 0.0", bloco,
                      "mes realizado e corrente tem de herdar zero")
        self.assertNotIn("saldos_iniciais_m", bloco,
                         "a Reserva monta a propria heranca, com piso em zero")

    def test_faixa_do_eixo_da_pct_tem_trava(self):
        """Um mes fora da curva achatava o grafico inteiro: com novembro em
        -1643%, os outros meses viravam uma linha reta colada no topo e o
        tracejado da meta sumia. A faixa sai so dos meses com resposta e nunca
        abre mais que o limite para cada lado."""
        limite = 150
        for valores, esperado_dentro in [
            ([19.0, 36.4, 29.8, 35.7], True),
            ([19.0, -1643.0, 29.8], False),
        ]:
            piso, teto = min(valores), max(valores)
            faixa = [max(-limite, min(0, piso - 10)), min(limite, max(60, teto + 15))]
            self.assertGreaterEqual(faixa[0], -limite)
            self.assertLessEqual(faixa[1], limite)
            self.assertEqual(piso >= faixa[0], esperado_dentro)
        # Mes sem resposta nao pode entrar na conta da faixa.
        self.assertIn("pct_com_resposta = [v for v in (pct_sobra_grafico + liquidez_grafico)",
                      FONTE)
        self.assertIn("LIMITE_EIXO_PCT_SOBRA = 150", FONTE)
        self.assertIn("range=faixa_pct,", FONTE)
        self.assertNotIn("range=[0, max(60, teto_pct * 1.4)]", FONTE,
                         "a faixa sem trava voltou")
        # E o ausente nao pode quebrar a cor nem o rotulo dos pontos. O
        # rotulo virou ANOTACAO, entao o mes sem resposta e pulado no laco em
        # vez de virar texto vazio.
        self.assertIn("COLORS[\"negative\"] if (pd.isna(v) or v < 30)", FONTE)
        i = FONTE.index("for _x, _v in zip(rotulos_x_m, pct_sobra_grafico):")
        self.assertIn("if pd.isna(_v):", FONTE[i:i + 200],
                      "mes sem resposta voltaria a escrever um rotulo vazio")

    def test_disponivel_nunca_fica_negativo(self):
        """Mes que fecha no vermelho nao entrega caixa negativo ao seguinte:
        o proximo comeca com o que ele mesmo recebe. O buraco continua
        visivel na linha de sobra do mes em que aconteceu."""
        mes_corrente = pd.Period("2026-08", "M")
        meses = [pd.Period(f"2026-{m:02d}", "M") for m in range(7, 13)]
        base = [13_964_192.05, 14_568_560.54, 8_142_648.34,
                5_327_020.43, 2_879_489.29, 1_226_104.71]
        pagar = [9_704_926.13, 9_111_756.02, 9_225_374.37,
                 11_612_832.02, 13_663_715.37, 11_207_121.40]
        disponiveis, anterior = [], 0.0
        for periodo, propria, a_pagar in zip(meses, base, pagar):
            disponivel = propria if periodo <= mes_corrente else propria + anterior
            disponiveis.append(disponivel)
            anterior = max(disponivel - a_pagar, 0.0)
        for valor in disponiveis:
            self.assertGreaterEqual(valor, 0.0, "disponivel e caixa: nao pode ser negativo")
        # Setembro pega a sobra de agosto, como pedido.
        self.assertAlmostEqual(disponiveis[2], 8_142_648.34 + 5_456_804.52, places=2)
        # E nenhuma outra chamada do pivo pode ter mudado de base.
        self.assertEqual(FONTE.count('posicao_saldo="primeira"'), 1)


# ============================================================================
# 5h. METAS DE RECEBIMENTO E AS TRES LINHAS DE A RECEBER
# ============================================================================
class TesteMetasDeRecebimento(unittest.TestCase):
    """A meta e balizador: aparece como QUANTO FALTA e nunca soma em total."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_normalizar_texto", "_classificar_movimento_fin", "_dias_da_meta_no_mes",
             "montar_linhas_de_meta", "_aplicar_meta_como_falta", "_total_geral_sem_meta",
             "_ordenar_movimentos_fin", "_peso_ordem_movimento_fin"],
            ["METAS_RECEBER", "DIAS_DA_SEMANA_POR_MODALIDADE", "MOV_RECEBER_META",
             "MOV_RECEBER_AVENCER", "MOV_RECEBER_LIQUIDADO", "RENOMEAR_MOVIMENTO_FIN",
             "COL_FIN_MOVIMENTO", "COL_FIN_CANAL",
             "COL_FIN_MODALIDADE", "COL_FIN_VALOR", "COL_FIN_VENCIMENTO",
             "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_LIQ_EFETIVA"],
        )

    def test_meta_mensal_bate_com_a_planilha_do_gestor(self):
        """A soma dos dias tem de devolver a meta do mes -- os centavos do
        arredondamento vao no ultimo dia, entao nada se perde."""
        metas = self.ns["METAS_RECEBER"]
        # Valores de 21/08/2026, conferidos contra a linha de Total de cada
        # canal na planilha de faturamento.
        esperado = {("HUB LOGISTICO", 7): 868_869.98, ("HUB LOGISTICO", 9): 2_731_272.77,
                    ("HUB LOGISTICO", 12): 3_577_081.73,
                    ("LOJA", 7): 1_731_735.90, ("LOJA", 9): 2_634_495.90,
                    ("LOJA", 12): 4_253_680.96,
                    ("VENDA DIRETA", 7): 1_699_876.38, ("VENDA DIRETA", 9): 5_096_452.91,
                    ("VENDA DIRETA", 12): 6_854_279.96}
        for (canal, mes), alvo in esperado.items():
            soma = sum(v.get((2026, mes), 0.0) for v in metas[canal].values())
            self.assertAlmostEqual(soma, alvo, delta=0.02, msg=f"{canal} mes {mes}")

    def test_soma_dos_dias_devolve_o_mes(self):
        colunas = [self.ns["COL_FIN_MOVIMENTO"], self.ns["COL_FIN_CANAL"],
                   self.ns["COL_FIN_MODALIDADE"], self.ns["COL_FIN_VALOR"], "Data Efetiva",
                   self.ns["COL_FIN_VENCIMENTO"], self.ns["COL_FIN_DATA_LIQUIDACAO"],
                   self.ns["COL_FIN_LIQ_EFETIVA"], "Liquidado", "Tipo Movimento"]
        diarias = self.ns["montar_linhas_de_meta"](colunas)
        setembro = diarias[diarias["Data Efetiva"].dt.to_period("M") == pd.Period("2026-09")]
        self.assertAlmostEqual(setembro[self.ns["COL_FIN_VALOR"]].sum(), 10_462_221.58, delta=0.05)

    def test_boleto_garantido_so_cai_em_terca_e_quinta(self):
        dias = self.ns["_dias_da_meta_no_mes"](2026, 9, "Boleto Garantido")
        self.assertEqual({d.weekday() for d in dias}, {1, 3})
        # As demais modalidades usam dia util, sem fim de semana.
        uteis = self.ns["_dias_da_meta_no_mes"](2026, 9, "Débito")
        self.assertEqual({d.weekday() for d in uteis}, {0, 1, 2, 3, 4})

    def test_meta_e_um_tipo_proprio_fora_dos_totais(self):
        """O nome contem "receber"; sem tratamento proprio ela seria
        classificada como entrada e entraria em todo indicador do painel."""
        self.assertEqual(self.ns["_classificar_movimento_fin"]("2 - Contas a Receber Meta"), "meta")
        self.assertEqual(self.ns["_classificar_movimento_fin"]("3 - Contas a Receber"), "entrada")
        self.assertEqual(
            self.ns["_classificar_movimento_fin"]("3.1 - Contas a Receber Liquidado"), "entrada")

    def _pivo(self, avencer, liquidado, meta=10_462_221.58):
        return pd.DataFrame({"Setembro": {
            self.ns["MOV_RECEBER_META"]: meta,
            self.ns["MOV_RECEBER_AVENCER"]: avencer,
            self.ns["MOV_RECEBER_LIQUIDADO"]: liquidado,
            "4 - Contas a Pagar": -3_500_000.00,
        }})

    def test_meta_mostra_quanto_falta(self):
        ajustado, cheia = self.ns["_aplicar_meta_como_falta"](
            self._pivo(4_000_000.0, 6_000_000.0))
        self.assertAlmostEqual(ajustado.loc[self.ns["MOV_RECEBER_META"], "Setembro"],
                               462_221.58, places=2)
        self.assertAlmostEqual(cheia["Setembro"], 10_462_221.58, places=2)

    def test_meta_zera_quando_batida_e_nao_fica_negativa(self):
        batida, _ = self.ns["_aplicar_meta_como_falta"](self._pivo(4_000_000.0, 6_462_221.58))
        self.assertAlmostEqual(batida.loc[self.ns["MOV_RECEBER_META"], "Setembro"], 0.0, places=2)
        # Um real abaixo da meta: e um real que tem de aparecer.
        quase, _ = self.ns["_aplicar_meta_como_falta"](self._pivo(4_000_000.0, 6_462_220.58))
        self.assertAlmostEqual(quase.loc[self.ns["MOV_RECEBER_META"], "Setembro"], 1.0, places=2)
        # Passando da meta continua zero, nunca negativo.
        passou, _ = self.ns["_aplicar_meta_como_falta"](self._pivo(4_000_000.0, 9_000_000.0))
        self.assertEqual(passou.loc[self.ns["MOV_RECEBER_META"], "Setembro"], 0.0)

    def test_total_geral_ignora_a_meta(self):
        pivo = self._pivo(4_000_000.0, 6_000_000.0)
        total = self.ns["_total_geral_sem_meta"](pivo)["Setembro"]
        self.assertAlmostEqual(total, 4_000_000.0 + 6_000_000.0 - 3_500_000.0, places=2)

    def test_ordem_de_leitura_das_linhas(self):
        """Sequencia fixa: o dinheiro que ja esta, o que deve entrar, o que
        sai. Os nomes vem da planilha como "1.1.Caixa" e "1.Banco", entao
        ordenar pelo numero colocaria o Banco na frente do Caixa."""
        embaralhado = ["2 - Contas a Receber Meta", "3 - Contas a Receber",
                       "3.1 - Contas a Receber Liquidado", "1.1.Caixa", "1.Banco",
                       "4 - Contas a Pagar"]
        self.assertEqual(
            self.ns["_ordenar_movimentos_fin"](embaralhado),
            ["1.1.Caixa", "1.Banco", "2 - Contas a Receber Meta", "3 - Contas a Receber",
             "3.1 - Contas a Receber Liquidado", "4 - Contas a Pagar"],
        )

    def test_ordem_tolera_outras_grafias(self):
        """A ordenacao olha o conteudo do nome, nao o texto exato -- se a
        planilha mudar "1.Banco" para "1 - Banco Bradesco", a ordem segue."""
        ordem = self.ns["_ordenar_movimentos_fin"](
            ["4 - Contas a Pagar", "9 - Aplicação Financeira", "1 - Banco Bradesco",
             "1.1 - Caixa Geral", "3 - Contas a Receber"])
        self.assertEqual(ordem[0], "1.1 - Caixa Geral")
        self.assertEqual(ordem[1], "1 - Banco Bradesco")
        self.assertEqual(ordem[-1], "9 - Aplicação Financeira",
                         "movimento desconhecido tem de cair no fim, nunca no meio")

    def test_liquidado_nao_troca_de_lugar_com_a_vencer(self):
        """O nome do liquidado tambem contem "receber": se a checagem vier
        na ordem errada, as duas linhas trocam de posicao."""
        self.assertEqual(self.ns["_peso_ordem_movimento_fin"]("3 - Contas a Receber"), 3)
        self.assertEqual(
            self.ns["_peso_ordem_movimento_fin"]("3.1 - Contas a Receber Liquidado"), 4)

    def test_diario_ordena_a_meta_junto_com_as_outras(self):
        """No diario a meta e montada a parte, entao ela poderia ser
        empilhada por cima de caixa e banco. Este teste roda a MESMA
        expressao de ordenacao do painel e confere onde a meta cai."""
        achado = re.search(
            r"linhas_do_canal\.sort\(\s*key=(lambda item:.+?)\s*\n\s*\)", FONTE, re.S)
        self.assertIsNotNone(achado, "o diario deixou de ordenar as linhas do canal")
        chave = eval(  # noqa: S307 - a expressao vem do proprio app
            achado.group(1),
            {"_peso_ordem_movimento_fin": self.ns["_peso_ordem_movimento_fin"], "str": str},
        )
        # Nomes SEM numero de propósito: com "1.1.Caixa" e "1.Banco" a ordem
        # alfabetica coincide com a certa por acaso, e o teste passaria mesmo
        # se a ordenacao voltasse a ser por nome. Aqui alfabetico poria Banco
        # antes de Caixa e "A pagar" antes de tudo.
        linhas = [("Contas a Pagar", None), ("Banco Itaú", None),
                  ("Contas a Receber", None), ("Caixa Geral", None),
                  (self.ns["MOV_RECEBER_META"], None),
                  ("Contas a Receber Liquidado", None)]
        linhas.sort(key=chave)
        self.assertEqual(
            [nome for nome, _ in linhas],
            ["Caixa Geral", "Banco Itaú", "2 - Contas a Receber Meta", "Contas a Receber",
             "Contas a Receber Liquidado", "Contas a Pagar"],
        )

    def test_as_duas_linhas_de_a_receber_nao_se_repetem(self):
        """Titulo com baixa vai para a linha de liquidado, na data da baixa;
        sem baixa fica na linha a vencer, na data de vencimento. Se as duas
        seguissem o mesmo eixo, o mesmo titulo apareceria duas vezes."""
        i = FONTE.index("_e_receber = df[COL_FIN_MOVIMENTO] == MOV_RECEBER_AVENCER")
        trecho = FONTE[i:i + 900]
        self.assertIn("MOV_RECEBER_LIQUIDADO", trecho)
        self.assertIn('_e_receber & _tem_baixa, "Data Efetiva"', trecho)
        self.assertIn('_e_receber & ~_tem_baixa, "Data Efetiva"', trecho)

    def test_meta_da_planilha_e_descartada(self):
        i = FONTE.index("_metas_antigas = int(")
        trecho = FONTE[i:i + 300]
        self.assertIn("df = df[df[COL_FIN_MOVIMENTO] != MOV_RECEBER_META]", trecho)
        self.assertIn("montar_linhas_de_meta(df.columns)", FONTE)

    def test_tabela_permite_selecionar_celulas_soltas(self):
        """O st.dataframe so seleciona linha ou coluna inteira. A area precisa
        do comportamento do Excel: clicar numa celula, segurar Ctrl, clicar em
        outras e ver a soma. Por isso a tabela e HTML propria."""
        ns_local = carregar(["formata_brl", "tabela_selecionavel"],
                            ["COLORS", "FONTE_MONO", "ALTURA_LINHA_TABELA_PX",
                             "ALTURA_CABECALHO_TABELA_PX", "ALTURA_BARRA_SOMA_PX"])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        df = pd.DataFrame([[10.0, -20.0], [30.0, 40.0]],
                          index=["1.1.Caixa", "4 - Contas a Pagar"], columns=["18/08", "19/08"])
        ns_local["tabela_selecionavel"](df, chave="t", tipos_linha=["movimento", "movimento"],
                                        linhas_visiveis=21)
        codigo = capturado["codigo"]
        self.assertEqual(len(re.findall(r'data-v="', codigo)), 4, "faltou celula clicavel")
        self.assertIn("evento.ctrlKey || evento.metaKey", codigo, "sem Ctrl+clique")
        self.assertIn("evento.shiftKey", codigo, "sem Shift+clique para intervalo")
        for identificador in ("soma", "media", "qtd"):
            self.assertIn(f'id="{identificador}"', codigo)

    def _altura_para(self, n_linhas):
        ns_local = carregar(["formata_brl", "tabela_selecionavel"],
                            ["COLORS", "FONTE_MONO", "FONTE_PADRAO_TABELA",
                             "TETO_LINHAS_TABELA", "ALTURA_LINHA_TABELA_PX",
                             "ALTURA_CABECALHO_TABELA_PX", "ALTURA_BARRA_SOMA_PX",
                             "ALTURA_BARRA_ROLAGEM_PX",
                             "FOLGA_ABAIXO_TABELA_PX"])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        # Colunas suficientes para haver rolagem horizontal, que e o caso das
        # tabelas de verdade (14 a 31 dias no diario).
        colunas = [f"C{i}" for i in range(20)]
        df = pd.DataFrame([[1.0] * len(colunas)] * n_linhas,
                          index=[f"L{i}" for i in range(n_linhas)], columns=colunas)
        ns_local["tabela_selecionavel"](df, chave="t")
        return capturado, ns_local

    def _medir(self, n_linhas, n_colunas):
        ns_local = carregar(["formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        df = pd.DataFrame([[1.0] * n_colunas] * n_linhas,
                          index=[f"L{i}" for i in range(n_linhas)],
                          columns=[f"C{i}" for i in range(n_colunas)])
        ns_local["tabela_selecionavel"](df, chave="t")
        return capturado, ns_local

    def test_altura_cabe_a_tabela_inteira_sem_rolagem(self):
        """A barra de rolagem HORIZONTAL fica dentro da area que rola e come
        altura: sem reservar esse espaco sobra sempre um filete de rolagem
        vertical, que foi o que apareceu na tela."""
        for n_linhas in (6, 21, 25):
            capturado, ns_local = self._altura_para(n_linhas)
            esperado = (ns_local["ALTURA_CABECALHO_TABELA_PX"]
                        + ns_local["ALTURA_LINHA_TABELA_PX"] * n_linhas
                        + ns_local["ALTURA_BARRA_ROLAGEM_PX"] + 2
                        + ns_local["ALTURA_BARRA_SOMA_PX"]
                        + ns_local["FOLGA_ABAIXO_TABELA_PX"])
            self.assertEqual(capturado["altura"], esperado, f"{n_linhas} linhas")

    def test_tabela_pequena_nao_ganha_espaco_vazio(self):
        """Altura fixa em 21 deixaria 15 linhas de vazio numa tabela de 6."""
        pequena, _ = self._altura_para(6)
        grande, _ = self._altura_para(21)
        self.assertLess(pequena["altura"], grande["altura"])

    def test_altura_tem_teto_para_o_caso_extremo(self):
        capturado, ns_local = self._altura_para(60)
        teto = ns_local["TETO_LINHAS_TABELA"]
        esperado = (ns_local["ALTURA_CABECALHO_TABELA_PX"]
                    + ns_local["ALTURA_LINHA_TABELA_PX"] * teto
                    + ns_local["ALTURA_BARRA_ROLAGEM_PX"] + 2
                    + ns_local["ALTURA_BARRA_SOMA_PX"]
                    + ns_local["FOLGA_ABAIXO_TABELA_PX"])
        self.assertEqual(capturado["altura"], esperado)

    def test_altura_da_caixa_fecha_com_o_conteudo(self):
        """A conta em Python e o CSS tem de falar a mesma altura de
        cabecalho. Enquanto o cabecalho herdava os 34px das linhas e a conta
        somava 40, sobrava um filete vazio embaixo da ultima linha."""
        # A folga da barra de rolagem sai da CONSTANTE, nao cravada: ela subiu
        # de 18 para 26 quando a tela em escala de 125% cortou a ultima linha,
        # e um numero cravado aqui faria o teste falhar por isso.
        _reserva = carregar([], ["ALTURA_BARRA_ROLAGEM_PX"])["ALTURA_BARRA_ROLAGEM_PX"]
        for n_linhas, n_colunas, folga_esperada in [(7, 7, 0), (21, 31, _reserva)]:
            capturado, ns_local = self._medir(n_linhas, n_colunas)
            css = capturado["codigo"]
            altura_cabecalho = int(
                re.search(r"th\.cabecalho, th\.canto \{ height:(\d+)px", css).group(1))
            altura_linha = int(re.search(r"th, td \{ height:(\d+)px", css).group(1))
            altura_caixa = int(re.search(r"\.rolagem \{ height:(\d+)px", css).group(1))
            conteudo = altura_cabecalho + altura_linha * n_linhas + 2
            self.assertEqual(altura_cabecalho, ns_local["ALTURA_CABECALHO_TABELA_PX"],
                             "CSS e conta discordam da altura do cabecalho")
            self.assertEqual(altura_caixa - conteudo, folga_esperada,
                             f"{n_linhas} linhas x {n_colunas} colunas")

    @staticmethod
    def _fundo_do_seletor(css, seletor):
        """Le o background de UM bloco do CSS. Casar por regex solta pega a
        regra vizinha e da falso positivo -- foi o que aconteceu quando este
        teste dizia que o cabecalho era preto sem ele ser."""
        # Tira os comentarios antes: um deles tem virgula, e a virgula e o
        # separador de seletores -- sem limpar, o bloco certo nao e achado.
        css = re.sub(r"/\*.*?\*/", "", css, flags=re.S)
        for bloco in css.split("}"):
            if "{" not in bloco:
                continue
            alvos = [s.strip() for s in bloco.split("{")[0].split(",")]
            if seletor not in alvos:
                continue
            achado = re.search(r"background:([^;]+);", bloco)
            if achado:
                return achado.group(1).strip()
        return None

    def test_corpo_da_tabela_e_preto(self):
        """O painel tem fundo azulado. Deixar a tabela transparente fazia as
        linhas comuns puxarem esse azul e o destaque das consolidacoes se
        perder no meio -- o corpo tem de ser preto."""
        capturado, ns_local = self._medir(5, 5)
        css = capturado["codigo"].split("<div")[0]
        preto = ns_local["FUNDO_TABELA_FLUXO"]
        self.assertEqual(preto, "#000000")
        self.assertEqual(self._fundo_do_seletor(css, ".rolagem"), preto)
        self.assertEqual(self._fundo_do_seletor(css, ".linha-movimento td"), preto)
        self.assertEqual(self._fundo_do_seletor(css, "th.rotulo"), preto)

    def test_regras_de_destaque_encontram_alvo_no_html(self):
        """Ler o CSS nao basta: a regra pode estar escrita e nao valer para
        nada. Foi o que aconteceu -- a classe do tipo estava na CELULA e o
        seletor procurava "td DENTRO de algo com essa classe", entao nenhuma
        linha de consolidacao ficava destacada."""
        from html.parser import HTMLParser

        ns_local = carregar(["formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        df = pd.DataFrame([[1.0, -2.0]] * 5,
                          index=["SALDO INICIAL", "HUB", "1.1.Caixa", "1.Banco", "TOTAL GERAL"],
                          columns=["18/08", "19/08"])
        ns_local["tabela_selecionavel"](
            df, chave="t",
            tipos_linha=["saldo_inicial", "canal", "movimento", "movimento", "total"])

        class Leitor(HTMLParser):
            def __init__(self):
                super().__init__()
                self.linha = None
                self.pares = []

            def handle_starttag(self, tag, attrs):
                classes = dict(attrs).get("class", "")
                if tag == "tr":
                    self.linha = classes
                elif tag in ("td", "th") and self.linha is not None:
                    self.pares.append((self.linha, tag))

        leitor = Leitor()
        leitor.feed(capturado["codigo"])
        for classe in ("linha-saldo_inicial", "linha-canal", "linha-total", "linha-movimento"):
            self.assertIn((classe, "td"), leitor.pares,
                          f"nenhuma celula dentro de tr.{classe} -- a regra do CSS fica morta")
            self.assertIn((classe, "th"), leitor.pares, f"rotulo fora de tr.{classe}")

    def test_consolidacoes_e_cabecalho_ficam_no_tom_claro(self):
        capturado, ns_local = self._medir(5, 5)
        css = capturado["codigo"].split("<div")[0]
        claro = ns_local["COLORS"]["surface_alt"]
        self.assertEqual(self._fundo_do_seletor(css, "th.cabecalho"), claro)
        self.assertEqual(self._fundo_do_seletor(css, "th.canto"), claro)
        for destaque in (".linha-canal td", ".linha-total td", ".linha-saldo_inicial td"):
            self.assertEqual(self._fundo_do_seletor(css, destaque), claro, destaque)

    def test_destaque_nao_pinta_a_coluna_de_nomes(self):
        """O preenchimento fica nos VALORES. Pintar tambem a coluna de nomes
        criava uma faixa clara atravessando a tabela, competindo com os
        numeros -- que sao o que se le ali. O nome se distingue pelo negrito."""
        capturado, _ = self._medir(5, 5)
        css = capturado["codigo"].split("<div")[0]
        for tipo in ("canal", "total", "saldo_inicial"):
            seletor = f".linha-{tipo} th.rotulo"
            self.assertIsNone(self._fundo_do_seletor(css, seletor),
                              f"{seletor} nao pode ter preenchimento")
            self.assertIn(f"{seletor} {{", re.sub(r"/\*.*?\*/", "", css, flags=re.S),
                          f"{seletor} precisa existir para manter o negrito")

    def test_folga_da_rolagem_so_quando_ha_rolagem(self):
        """Com poucas colunas nao aparece barra horizontal, e reservar o
        espaco dela deixava uma faixa vazia embaixo da ultima linha."""
        ns_local = carregar(["formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)

        def altura(n_colunas):
            df = pd.DataFrame([[1.0] * n_colunas] * 7,
                              index=[f"L{i}" for i in range(7)],
                              columns=[f"C{i}" for i in range(n_colunas)])
            ns_local["tabela_selecionavel"](df, chave=f"t{n_colunas}")
            return capturado["altura"]

        poucas = altura(7)   # mensal
        muitas = altura(31)  # diario
        self.assertEqual(muitas - poucas, ns_local["ALTURA_BARRA_ROLAGEM_PX"])

    def test_tabela_usa_a_mesma_fonte_das_outras(self):
        """O iframe nao herda a fonte da pagina. Com monoespacada nos valores
        esta tabela destoava de todas as outras da tela."""
        capturado, _ = self._altura_para(3)
        codigo = capturado["codigo"]
        self.assertNotIn("monospace", codigo, "voltou a fonte monoespacada")
        self.assertIn("Source Sans Pro", codigo, "sem a fonte do app declarada no iframe")
        corpo = codigo.split(".barra")[0]
        self.assertNotIn("text-transform:uppercase", corpo,
                         "cabecalho da tabela nao e caixa alta nas outras telas")

    def test_rotulo_nao_mostra_o_espaco_invisivel(self):
        """O indice do diario usa espacos invisiveis para nao repetir rotulo;
        eles nao podem chegar na tela."""
        ns_local = carregar(["formata_brl", "tabela_selecionavel"],
                            ["COLORS", "FONTE_MONO", "ALTURA_LINHA_TABELA_PX",
                             "ALTURA_CABECALHO_TABELA_PX", "ALTURA_BARRA_SOMA_PX"])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        df = pd.DataFrame([[1.0]], index=["\u200b\u200b    1.1.Caixa"], columns=["18/08"])
        ns_local["tabela_selecionavel"](df, chave="t")
        self.assertNotIn("\u200b", capturado["codigo"])
        self.assertIn(">1.1.Caixa<", capturado["codigo"])


# ============================================================================
# 5n. MÉTRICAS DA ABA ANÁLISES
# ============================================================================
class TestePrazoMedioECicloFinanceiro(unittest.TestCase):
    """PMR, PMP e ciclo financeiro (25/08/2026), e os dois cartoes da aba
    Analises cujo calculo nao media o que o rotulo prometia."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["dias_de_giro", "formata_dias"])

    # -- PMR / PMP -------------------------------------------------------
    def test_prazo_medio_e_saldo_em_aberto_sobre_movimento(self):
        """R$ 300 mil parados contra R$ 900 mil movimentados em 90 dias sao 30
        dias de giro. E a formula da Controladoria na versao que esta base
        permite."""
        self.assertAlmostEqual(
            self.ns["dias_de_giro"](300_000, 900_000, 90), 30.0, places=6)

    def test_prazo_medio_ignora_o_sinal(self):
        """Saida vem negativa no CSV e entrada positiva. Sinais MISTOS sao o
        caso que morde: sem o abs, um saldo negativo sobre um movimento
        positivo daria "-30 dias de prazo de pagamento". Com os dois negativos
        o erro se esconde, porque eles se cancelam na divisao."""
        giro = self.ns["dias_de_giro"]
        self.assertAlmostEqual(giro(-300_000, 900_000, 90), 30.0, places=6)
        self.assertAlmostEqual(giro(300_000, -900_000, 90), 30.0, places=6)
        self.assertAlmostEqual(giro(-300_000, -900_000, 90), 30.0, places=6)

    def test_sem_movimento_no_periodo_nao_ha_prazo(self):
        """Dividir por zero aqui produziria um prazo infinito que a tela
        mostraria como numero."""
        self.assertIsNone(self.ns["dias_de_giro"](300_000, 0, 90))
        self.assertIsNone(self.ns["dias_de_giro"](300_000, 900_000, 0))

    def test_nada_em_aberto_da_prazo_zero_e_nao_vazio(self):
        """Zero em aberto e uma resposta: recebe a vista."""
        self.assertEqual(self.ns["dias_de_giro"](0, 900_000, 90), 0.0)

    def test_ciclo_financeiro_e_a_diferenca_dos_dois(self):
        pmr = self.ns["dias_de_giro"](300_000, 900_000, 90)     # 30
        pmp = self.ns["dias_de_giro"](600_000, 900_000, 90)     # 60
        self.assertAlmostEqual(pmr - pmp, -30.0, places=6)

    def test_dias_corridos_e_nao_dias_com_movimento(self):
        """Prazo medio se mede contra o CALENDARIO. Usar "dias com movimento"
        encurtaria o prazo em todo mes com feriado, e o indicador melhoraria
        por motivo nenhum."""
        i = FONTE.index("# ---- PMR, PMP e ciclo financeiro ----")
        trecho = FONTE[i:i + 1500]
        self.assertIn('(df_a["DiaOrd"].max() - df_a["DiaOrd"].min()).days + 1', trecho)
        self.assertNotIn("dias_com_mov_a", trecho,
                         "o prazo medio nao pode usar dias com movimento")

    # -- formatacao ------------------------------------------------------
    def test_dias_saem_com_virgula_decimal(self):
        """O painel inteiro escreve numero em portugues; estes cartoes
        escapavam com ponto ("-8.4 dias") por usarem f-string crua."""
        formata = self.ns["formata_dias"]
        self.assertEqual(formata(19.7), "19,7 dias")
        self.assertEqual(formata(-8.4, com_sinal=True), "-8,4 dias")
        self.assertEqual(formata(0.0, com_sinal=True), "+0,0 dias")
        self.assertEqual(formata(None), "—")
        self.assertEqual(formata(float("nan")), "—")

    # -- os dois cartoes corrigidos --------------------------------------
    def test_liquidez_imediata_divide_pelo_que_falta_pagar(self):
        """Ate 25/08/2026 dividia pelo TOTAL DE SAIDAS do periodo, que inclui
        tudo que ja foi pago: com "todo o periodo" o denominador era R$ 133 mi
        e o cartao marcava 4%, um numero que nao mede liquidez nenhuma."""
        i = FONTE.index("liquidez_imediata_a = ")
        trecho = FONTE[i:i + 260]
        self.assertIn("valor_aberto_a", trecho)
        self.assertNotIn("abs(saidas_a)", trecho,
                         "voltou a dividir pelo total de saidas do periodo")
        # E o subtexto tem de dizer qual e o denominador.
        j = FONTE.index('label="LIQUIDEZ IMEDIATA"')
        self.assertIn("EM ABERTO", FONTE[j:j + 500])

    def test_recebiveis_realizados_usam_os_movimentos_que_existem(self):
        """Ate 25/08/2026 procurava a palavra "projetado" no nome do movimento
        -- que NAO EXISTE nesta base. Dava sempre zero projetado e sempre 100%
        realizado: um cartao que nunca variava e por isso nunca informava."""
        i = FONTE.index("projetado_a = ")
        trecho = FONTE[i:i + 400]
        self.assertNotIn("projetad", trecho.replace("projetado_a", ""),
                         "voltou a procurar um nome de movimento inexistente")
        self.assertIn("valor_receber_aberto_a", trecho)
        self.assertIn("valor_recebido_a", trecho)

    def test_contas_a_receber_separa_pelo_nome_do_movimento(self):
        """Aqui o que diz se o titulo foi recebido e o NOME DO MOVIMENTO, nao a
        data de liquidacao -- e a mesma leitura do Fluxo Mensal. Ler pela data
        aqui e pelo movimento la faria as duas telas discordarem."""
        i = FONTE.index("# ---- Contas a RECEBER em aberto ----")
        trecho = FONTE[i:i + 1600]
        self.assertIn("MOV_RECEBER_AVENCER", trecho)
        self.assertIn("MOV_RECEBER_LIQUIDADO", trecho)

    def test_filtro_de_mes_abre_no_mes_corrente(self):
        """Em "todo o periodo" a base junta de outubro/2023 a fevereiro/2027, e
        metrica de ritmo perde o sentido: concentracao nos 3 maiores dias sobre
        484 dias dava 7% e parecia baixa quando e altissima, e o PMR ia a 418
        dias porque o denominador diluia o recebimento em 1400 dias."""
        i = FONTE.index("meses_disp_a = sorted(")
        trecho = FONTE[i:FONTE.index('if canal_sel_a != "Todos":', i)]
        self.assertIn("_mes_corrente_a", trecho)
        self.assertIn("seletor_periodo_meses(", trecho)
        # Mes corrente ausente da base nao pode estourar: vira None, e o
        # seletor cai no primeiro e no ultimo mes existentes.
        self.assertIn("if _mes_corrente_a in meses_disp_a else None", trecho)
        # E o recorte tem de usar o INTERVALO. Filtrar so pelo mes inicial
        # deixaria o seletor de fim desenhado na tela sem efeito nenhum --
        # pior que nao ter o campo.
        self.assertIn(">= periodo_ini_a", trecho)
        self.assertIn("<= periodo_fim_a", trecho)

    def test_as_duas_abas_usam_o_MESMO_seletor_de_meses(self):
        """Duas copias da mesma escolha sao duas chances de elas divergirem --
        e a semeadura na sessao, que o defeito do filtro revelou, precisa valer
        nas duas telas."""
        self.assertEqual(FONTE.count("seletor_periodo_meses("), 3,
                         "1 definicao + 2 usos (Fluxo Mensal e Analises)")
        # O Fluxo Mensal nao pode ter voltado a montar os seletores na mao.
        self.assertNotIn('key="fin_mes_ini"', FONTE)
        self.assertNotIn('key="fin_mes_fim"', FONTE)


class TesteSeletorDeIntervaloDeMeses(unittest.TestCase):
    """O seletor "Do mes / Ate o mes" compartilhado pelo Fluxo Mensal e pela
    aba Analises (25/08/2026)."""

    class _StFalso:
        """Dubla o que o seletor usa do Streamlit. selectbox devolve o que
        estiver na sessao -- que e exatamente como o Streamlit se comporta
        quando a chave ja existe, e e esse comportamento que o teste precisa
        reproduzir para o `index` nao mascarar a semeadura."""

        class _Coluna:
            def __enter__(self):
                return self

            def __exit__(self, *_):
                return False

        def __init__(self):
            self.session_state = {}

        def columns(self, quantas):
            n = quantas if isinstance(quantas, int) else len(quantas)
            return [self._Coluna() for _ in range(n)]

        def selectbox(self, _rotulo, opcoes, index=0, key=None, help=None):
            if key in self.session_state:
                return self.session_state[key]
            return opcoes[index]

    @classmethod
    def setUpClass(cls):
        cls.st_falso = cls._StFalso()
        cls.ns = carregar(
            ["seletor_periodo_meses", "rotulo_periodo_meses", "_rotulo_mes_pt_extenso"],
            ["MESES_PT_FIN"],
            extras={"st": cls.st_falso},
        )

    def setUp(self):
        self.st_falso.session_state.clear()
        self.meses = [pd.Period(f"2026-{m:02d}", "M") for m in range(1, 13)]

    def _rodar(self, ini=None, fim=None):
        return self.ns["seletor_periodo_meses"](
            self.meses, "teste", periodo_ini_padrao=ini, periodo_fim_padrao=fim)

    def test_um_mes_so_quando_os_dois_padroes_sao_iguais(self):
        agosto = pd.Period("2026-08", "M")
        self.assertEqual(self._rodar(agosto, agosto), (agosto, agosto))

    def test_intervalo_de_varios_meses(self):
        ini, fim = pd.Period("2026-03", "M"), pd.Period("2026-07", "M")
        self.assertEqual(self._rodar(ini, fim), (ini, fim))

    def test_padrao_ausente_cai_no_primeiro_e_no_ultimo(self):
        """Mes corrente fora da base nao pode estourar nem abrir vazio."""
        self.assertEqual(self._rodar(None, None), (self.meses[0], self.meses[-1]))
        fora = pd.Period("2019-05", "M")
        self.assertEqual(self._rodar(fora, fora), (self.meses[0], self.meses[-1]))

    def test_meses_invertidos_sao_ordenados_e_nao_viram_erro(self):
        """O painel entende a intencao em vez de exigir a ordem certa."""
        self.st_falso.session_state["teste_ini"] = "Setembro/2026"
        self.st_falso.session_state["teste_fim"] = "Março/2026"
        self.st_falso.session_state["_teste_iniciado"] = True
        ini, fim = self._rodar()
        self.assertEqual((ini, fim), (pd.Period("2026-03", "M"), pd.Period("2026-09", "M")))

    def test_padrao_vence_sessao_antiga_uma_vez_so(self):
        """O `index=` do selectbox NAO basta: o Streamlit o IGNORA quando a
        chave ja existe na sessao. Foi assim que o padrao novo do filtro da
        Analises nunca apareceu, em 25/08/2026, com o codigo certo no ar."""
        agosto = pd.Period("2026-08", "M")
        # Sessao velha, com escolha antiga guardada e SEM a marca.
        self.st_falso.session_state["teste_ini"] = "Janeiro/2026"
        self.st_falso.session_state["teste_fim"] = "Dezembro/2026"
        self.assertEqual(self._rodar(agosto, agosto), (agosto, agosto),
                         "o padrao novo tem de vencer a sessao antiga")
        # Semeado. Agora a escolha da pessoa manda e sobrevive ao redesenho.
        self.st_falso.session_state["teste_ini"] = "Abril/2026"
        self.st_falso.session_state["teste_fim"] = "Junho/2026"
        self.assertEqual(self._rodar(agosto, agosto),
                         (pd.Period("2026-04", "M"), pd.Period("2026-06", "M")))

    def test_rotulo_que_sumiu_da_lista_volta_ao_padrao(self):
        """Base que muda de recorte deixaria o widget travado num rotulo que
        nao existe mais."""
        agosto = pd.Period("2026-08", "M")
        self.st_falso.session_state["_teste_iniciado"] = True
        self.st_falso.session_state["teste_ini"] = "Dezembro/2019"
        self.st_falso.session_state["teste_fim"] = "Dezembro/2019"
        self.assertEqual(self._rodar(agosto, agosto), (agosto, agosto))

    def test_lista_vazia_nao_quebra(self):
        self.assertEqual(self.ns["seletor_periodo_meses"]([], "vazio"), (None, None))

    def test_rotulo_do_recorte(self):
        rotulo = self.ns["rotulo_periodo_meses"]
        um = pd.Period("2026-08", "M")
        self.assertEqual(rotulo(um, um), "Agosto/2026")
        self.assertEqual(rotulo(um, pd.Period("2026-10", "M")), "Agosto/2026 a Outubro/2026")
        self.assertEqual(rotulo(None, None), "")

    def test_recorte_longo_avisa_que_o_prazo_nao_significa_nada(self):
        """Com o recorte inteiro o valor em aberto e a foto de hoje e o
        movimentado soma o intervalo todo: a razao entre os dois vira artefato
        do tamanho do recorte, nao prazo."""
        ns = carregar([], ["LIMITE_DIAS_PRAZO_CONFIAVEL"])
        self.assertEqual(ns["LIMITE_DIAS_PRAZO_CONFIAVEL"], 120)
        i = FONTE.index('label="CICLO FINANCEIRO"')
        trecho = FONTE[i:i + 3000]
        self.assertIn("if _dias_corridos_a > LIMITE_DIAS_PRAZO_CONFIAVEL:", trecho)
        self.assertIn("Escolha um mês", trecho)


class TesteMetricasDeAnalise(unittest.TestCase):
    """Com vencimento e liquidacao completos, os prazos passaram a ser
    ponderados PELO VALOR. A media simples por titulo trata um pagamento de
    R$ 1 milhao atrasado igual a um de R$ 10 -- e e o de R$ 1 milhao que
    move o caixa."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["atraso_ponderado_por_valor", "pontualidade_por_valor",
                           "aging_de_vencidos"])

    def test_atraso_e_ponderado_pelo_valor(self):
        df = pd.DataFrame({"dias": [10, 0, 0, 0],
                           "valor": [-1_000_000.0, -10.0, -10.0, -10.0]})
        self.assertAlmostEqual(df["dias"].mean(), 2.5, places=2)   # a leitura antiga
        self.assertAlmostEqual(
            self.ns["atraso_ponderado_por_valor"](df, "dias", "valor"), 10.0, places=2)

    def test_atraso_sem_titulo_liquidado_devolve_nada(self):
        """Sem baixa nao ha prazo -- e zero seria mentira, nao ausencia."""
        df = pd.DataFrame({"dias": [None, None], "valor": [-10.0, -20.0]})
        self.assertIsNone(self.ns["atraso_ponderado_por_valor"](df, "dias", "valor"))
        self.assertIsNone(self.ns["atraso_ponderado_por_valor"](pd.DataFrame(), "dias", "valor"))

    def test_pontualidade_separa_antecipado_de_em_dia(self):
        """Antecipar e decisao de tesouraria; pagar no dia e cumprir o
        combinado. Juntar os dois num "pagos ate o vencimento" escondia qual
        dos dois estava acontecendo."""
        df = pd.DataFrame({"dias": [-5, 0, 3], "valor": [-100.0, -100.0, -200.0]})
        antecipado, em_dia, em_atraso = self.ns["pontualidade_por_valor"](
            df, "dias", "valor")
        self.assertAlmostEqual(antecipado, 25.0, places=1)
        self.assertAlmostEqual(em_dia, 25.0, places=1)
        self.assertAlmostEqual(em_atraso, 50.0, places=1)

    def test_aging_separa_atraso_curto_de_longo(self):
        """Um total de vencidos nao diz se e de ontem ou de tres meses --
        e a diferenca e o tamanho do problema."""
        hoje = pd.Timestamp("2026-08-20")
        df = pd.DataFrame({
            "venc": pd.to_datetime(["2026-08-18", "2026-08-10", "2026-05-01", "2026-08-25"]),
            "valor": [-500.0, -2_000.0, -30_000.0, -100.0],
        })
        faixas = self.ns["aging_de_vencidos"](df, "venc", "valor", hoje)
        por_rotulo = {rotulo: (valor, qtd) for rotulo, valor, qtd in faixas}
        self.assertEqual(por_rotulo["1 a 7 dias"], (500.0, 1))
        self.assertEqual(por_rotulo["8 a 15 dias"], (2_000.0, 1))
        self.assertEqual(por_rotulo["mais de 60 dias"], (30_000.0, 1))
        # O de 25/08 ainda NAO venceu: nao pode aparecer em faixa nenhuma,
        # nem no valor total. Sem esta conferencia, incluir o que esta a
        # vencer passava despercebido.
        self.assertEqual(sum(qtd for _r, _v, qtd in faixas), 3)
        self.assertAlmostEqual(sum(v for _r, v, _q in faixas), 32_500.0, places=2)
        self.assertNotIn(100.0, [v for _r, v, _q in faixas],
                         "o titulo a vencer entrou no aging")

    def test_aba_usa_as_metricas_ponderadas(self):
        # Recorte por MARCOS, nao por janela de N caracteres: a janela fixa
        # quebrou quando o bloco cresceu, e o teste passou a falhar por
        # posicao em vez de por defeito.
        i = FONTE.index("# ---- Prazos e cobertura ----")
        trecho = FONTE[i:FONTE.index("with sub_aberto:")]
        self.assertIn("atraso_ponderado_por_valor(", trecho)
        self.assertIn("pontualidade_por_valor(", trecho)
        self.assertIn("Ponderado por valor", trecho,
                      "o cartao precisa dizer que e ponderado, senao engana")
        j = FONTE.index("with sub_aberto:")
        self.assertIn("aging_de_vencidos(", FONTE[j:FONTE.index("with sub_receber:")])
        # O contas a RECEBER tem de usar a MESMA funcao de faixas -- duas
        # construcoes da mesma conta sao duas chances de elas discordarem.
        k = FONTE.index("with sub_receber:")
        self.assertIn("aging_de_vencidos(", FONTE[k:FONTE.index("with sub_estrutura:")])


# ============================================================================
# 5m. NOMES DAS LINHAS E A VIRADA DA META
# ============================================================================
class TesteNomesEVirada(unittest.TestCase):

    def test_nomes_das_linhas_do_fluxo(self):
        ns = carregar([], ["MOV_RECEBER_META", "MOV_RECEBER_AVENCER",
                           "MOV_RECEBER_LIQUIDADO", "MOV_PAGAR",
                           "RENOMEAR_MOVIMENTO_FIN"])
        self.assertEqual(ns["MOV_RECEBER_META"], "2 - Contas a Receber Meta")
        self.assertEqual(ns["MOV_RECEBER_AVENCER"], "2.1 - Contas a Receber")
        self.assertEqual(ns["MOV_RECEBER_LIQUIDADO"], "2.2 - Contas a Receber Liquidado")
        self.assertEqual(ns["MOV_PAGAR"], "3 - Contas a Pagar")
        # A planilha continua escrevendo do jeito antigo: a traducao precisa
        # existir, senao o contas a pagar entra com o numero velho.
        self.assertEqual(ns["RENOMEAR_MOVIMENTO_FIN"]["4 - contas a pagar"], "3 - Contas a Pagar")

    def test_ordem_continua_valendo_com_os_nomes_novos(self):
        ns = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                       "_ordenar_movimentos_fin"])
        self.assertEqual(
            ns["_ordenar_movimentos_fin"](
                ["3 - Contas a Pagar", "2.2 - Contas a Receber Liquidado", "1.Banco",
                 "2.1 - Contas a Receber", "2 - Contas a Receber Meta", "1.1.Caixa"]),
            ["1.1.Caixa", "1.Banco", "2 - Contas a Receber Meta", "2.1 - Contas a Receber",
             "2.2 - Contas a Receber Liquidado", "3 - Contas a Pagar"])

    def _falta(self, meta, realizado, dias_visiveis=None, meta_alvo=None):
        """Roda o par de funcoes como as telas rodam: o fator sai da EMPRESA
        e a meta mostrada pode ser a de um canal."""
        ns = carregar(["fracao_da_meta_ainda_nao_coberta", "meta_diaria_que_ainda_falta"])
        fator = ns["fracao_da_meta_ainda_nao_coberta"](meta, realizado)
        return ns["meta_diaria_que_ainda_falta"](
            meta if meta_alvo is None else meta_alvo, fator, dias_visiveis)

    def test_recebivel_do_fim_do_mes_ja_cobre_os_dias_anteriores(self):
        """O contas a receber e posicionado pelo VENCIMENTO. Comparando o
        acumulado ATE o dia, um titulo que vence dia 28 so contava no dia 28
        e os dias anteriores seguiam cobrando meta com o mes ja garantido --
        era o que deixava valor futuro na tela."""
        dias = pd.date_range("2026-08-01", "2026-08-31", freq="D")
        meta = pd.Series([100.0] * 31, index=dias)           # meta do mes: 3.100
        real = pd.Series(0.0, index=dias)
        real[pd.Timestamp("2026-08-28")] = 3500.0            # tudo vence no fim
        falta = self._falta(meta, real)
        self.assertEqual((falta > 0).sum(), 0, "o mes inteiro esta coberto")

    def test_mes_parcialmente_coberto_cobra_o_que_sobra(self):
        """Cobrindo 1.250 de 3.100, os primeiros dias zeram, um fica parcial
        e o resto continua cheio -- e a soma bate com o que falta."""
        dias = pd.date_range("2026-08-01", "2026-08-31", freq="D")
        meta = pd.Series([100.0] * 31, index=dias)
        real = pd.Series(0.0, index=dias)
        real[pd.Timestamp("2026-08-28")] = 1250.0
        falta = self._falta(meta, real)
        self.assertEqual(falta[pd.Timestamp("2026-08-01")], 0.0)
        self.assertEqual(falta[pd.Timestamp("2026-08-12")], 0.0)
        self.assertAlmostEqual(falta[pd.Timestamp("2026-08-13")], 50.0, places=2)
        self.assertAlmostEqual(falta[pd.Timestamp("2026-08-14")], 100.0, places=2)
        self.assertAlmostEqual(falta.sum(), 3100.0 - 1250.0, places=2)

    def test_canal_que_nao_bateu_zera_quando_a_empresa_cobriu(self):
        """A decisao e da EMPRESA: um canal pode nao ter batido a sua parte
        enquanto a empresa ja cobriu o mes."""
        dias = pd.date_range("2026-08-01", "2026-08-31", freq="D")
        meta_empresa = pd.Series([100.0] * 31, index=dias)
        real_empresa = pd.Series(0.0, index=dias)
        real_empresa[pd.Timestamp("2026-08-28")] = 3500.0
        meta_loja = pd.Series([40.0] * 31, index=dias)
        falta_loja = self._falta(meta_empresa, real_empresa, meta_alvo=meta_loja)
        self.assertEqual(falta_loja.sum(), 0.0)

    def test_empresa_longe_da_meta_ninguem_zera(self):
        dias = pd.date_range("2026-08-01", "2026-08-10", freq="D")
        meta = pd.Series([100.0] * 10, index=dias)
        real = pd.Series(0.0, index=dias)
        real.iloc[2] = 50.0
        falta = self._falta(meta, real)
        self.assertAlmostEqual(falta.sum(), 950.0, places=2)

    def test_conta_usa_o_mes_inteiro_e_nao_o_recorte(self):
        """Com o periodo comecando no dia 18, olhar so o recorte esconde o
        que entrou do dia 1 ao 17."""
        dias = pd.date_range("2026-08-01", "2026-08-31", freq="D")
        meta = pd.Series([100.0] * 31, index=dias)
        real = pd.Series(0.0, index=dias)
        real[pd.Timestamp("2026-08-05")] = 3500.0
        visiveis = pd.date_range("2026-08-18", "2026-08-31", freq="D")

        so_recorte = self._falta(meta[visiveis], real[visiveis])
        self.assertTrue((so_recorte > 0).all(),
                        "o cenario precisa mesmo falhar sem o mes inteiro")
        com_mes = self._falta(meta, real, visiveis)
        self.assertEqual(list(com_mes.index), list(visiveis))
        self.assertFalse(com_mes.any())

    def test_cada_mes_e_avaliado_por_si(self):
        """Agosto coberto nao pode zerar setembro."""
        dias_ago = pd.date_range("2026-08-01", "2026-08-31", freq="D")
        dias_set = pd.date_range("2026-09-01", "2026-09-30", freq="D")
        meta = pd.concat([pd.Series([100.0] * 31, index=dias_ago),
                          pd.Series([200.0] * 30, index=dias_set)])
        real = pd.Series(0.0, index=list(dias_ago) + list(dias_set))
        real[pd.Timestamp("2026-08-05")] = 3500.0
        falta = self._falta(meta, real)
        self.assertFalse(falta[:pd.Timestamp("2026-08-31")].any(), "agosto esta coberto")
        self.assertTrue((falta[pd.Timestamp("2026-09-01"):] > 0).all(),
                        "setembro ainda tem meta a cobrar")

    def test_diario_decide_o_fator_uma_vez_no_geral(self):
        i = FONTE.index("with tab_fin_diario:")
        trecho = FONTE[i:FONTE.index("with tab_fin_consolidado:")]
        self.assertIn("_fator_meta_d = fracao_da_meta_ainda_nao_coberta(", trecho)
        self.assertIn("serie_meta, _fator_meta_d, dias_ordenados_d", trecho)
        j = trecho.index("_fator_meta_d = fracao_da_meta_ainda_nao_coberta(")
        self.assertIn("_agrega_no_mes_cheio(df_d_metas)", trecho[j:j + 300],
                      "o fator precisa usar as metas de TODOS os canais")

    def test_as_duas_telas_montam_a_meta_com_o_mes_cheio(self):
        """Trava estrutural: se qualquer uma voltar a somar so o recorte, o
        numero volta a discordar do mensal."""
        i = FONTE.index("with tab_fin_diario:")
        diario = FONTE[i:FONTE.index("with tab_fin_consolidado:")]
        self.assertIn("_agrega_no_mes_cheio(df_meta_canal)", diario)
        self.assertIn("serie_meta, _fator_meta_d, dias_ordenados_d", diario)
        j = FONTE.index("with tab_fin_consolidado:")
        consolidado = FONTE[j:FONTE.index("# ---------------- TESOURARIA", j)]
        self.assertIn("_por_dia_mes_cheio(", consolidado)
        self.assertIn("fracao_da_meta_ainda_nao_coberta(", consolidado)
        self.assertIn("dias_dc)", consolidado,
                      "o consolidado precisa cortar o resultado nos dias da tela")

    def test_linha_da_meta_some_quando_zerada_no_intervalo(self):
        """Se no recorte visivel a meta ja foi batida em todos os dias, a
        linha nao acrescenta nada -- e some. Volta se o periodo alcancar um
        mes em que ainda falta."""
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertIn("serie_falta_dc = falta if falta.any() else None", trecho)
        self.assertIn("if serie_falta_dc is not None:", trecho)
        j = FONTE.index("falta_meta_canal = meta_diaria_que_ainda_falta(")
        self.assertIn("if falta_meta_canal.any():", FONTE[j:j + 500])


# ============================================================================
# 5l. NOME DAS COLUNAS DA PLANILHA
# ============================================================================
class TesteNomeDasColunas(unittest.TestCase):
    """A planilha muda de escrita de vez em quando ("Data de Liquidação" no
    lugar de "Data Liquidação"). Para o painel e a mesma coluna, e uma
    preposicao a mais nao pode derrubar a tela inteira."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["_normalizar_coluna_fin", "_assinatura_coluna_fin", "resolver_colunas_fluxo"],
            ["_ACENTOS_FIN", "LIGACOES_NOME_COLUNA", "COL_FIN_VALOR", "COL_FIN_MODALIDADE",
             "COL_FIN_CANAL", "COL_FIN_MOVIMENTO", "COL_FIN_DATA_LIQUIDACAO",
             "COL_FIN_VENCIMENTO"],
        )
        cls.esperadas = [cls.ns[k] for k in (
            "COL_FIN_VALOR", "COL_FIN_MODALIDADE", "COL_FIN_CANAL", "COL_FIN_MOVIMENTO",
            "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VENCIMENTO")]

    def _resolver(self, colunas):
        df = pd.DataFrame({c: [1] for c in colunas})
        return self.ns["resolver_colunas_fluxo"](df, self.esperadas)

    def test_aceita_as_escritas_que_a_planilha_ja_usou(self):
        base = ["Valor.1", "Modalidade", "Canal.1", "Movimento", "Vencimento.1"]
        for escrita in ["Data Liquidação", "Data de Liquidação", "DATA DE LIQUIDACAO",
                        "  Data  de  Liquidação "]:
            _saida, faltando, _ren = self._resolver(base + [escrita])
            self.assertEqual(faltando, [], f"nao reconheceu: {escrita!r}")

    def test_coluna_que_falta_de_verdade_continua_faltando(self):
        """Tolerancia nao pode virar adivinhacao: sem a coluna, o painel tem
        de dizer o que falta, e nao inventar um substituto."""
        _saida, faltando, _ren = self._resolver(
            ["Valor.1", "Modalidade", "Canal.1", "Movimento", "Vencimento.1"])
        self.assertEqual(faltando, [self.ns["COL_FIN_DATA_LIQUIDACAO"]])

    def test_nao_confunde_valor_com_valor_ponto_um(self):
        """A planilha tem pares como "Valor" e "Valor.1". Trocar uma pela
        outra em silencio e bem pior do que a tela de erro."""
        saida, faltando, _ren = self._resolver(
            ["Valor", "Valor.1", "Modalidade", "Canal.1", "Movimento",
             "Data de Liquidação", "Vencimento.1"])
        self.assertEqual(faltando, [])
        self.assertIn("Valor.1", saida.columns)
        self.assertIn("Valor", saida.columns)

    def test_reconhece_os_nomes_antes_de_descartar_colunas(self):
        """ORDEM IMPORTA. A economia de memoria descarta as colunas que o
        painel nao usa; se ela rodar ANTES do reconhecimento, joga fora
        justamente a coluna que ainda nao tem o nome canonico ("Data de
        Liquidacao") e o painel acusa que ela esta faltando. Foi o que
        aconteceu em 20/08/2026."""
        i = FONTE.index("df_fluxo, faltando, colunas_renomeadas = resolver_colunas_fluxo(")
        j = FONTE.index("_descartar = [c for c in df_fluxo.columns")
        self.assertLess(i, j, "descartar colunas antes de reconhecer os nomes quebra o painel")

    def test_a_ordem_certa_salva_a_coluna_com_outra_escrita(self):
        """Prova pelo resultado, e nao so pela posicao no arquivo."""
        colunas_csv = ["Movimento", "Valor.1", "Vencimento.1", "Data de Liquidação",
                       "Canal.1", "Modalidade", "GRUPO DESPESA", "Coluna Extra"]
        uteis = set(self.esperadas) | {"GRUPO DESPESA"}
        df = pd.DataFrame({c: [1] for c in colunas_csv})

        # Ordem errada: descarta primeiro.
        errada = df.drop(columns=[c for c in df.columns if c not in uteis])
        _s, faltando_errada, _r = self.ns["resolver_colunas_fluxo"](errada, self.esperadas)
        self.assertEqual(faltando_errada, [self.ns["COL_FIN_DATA_LIQUIDACAO"]],
                         "o cenario precisa mesmo falhar na ordem errada")

        # Ordem certa: reconhece primeiro.
        certa, faltando_certa, _r = self.ns["resolver_colunas_fluxo"](df, self.esperadas)
        certa = certa.drop(columns=[c for c in certa.columns if c not in uteis])
        self.assertEqual(faltando_certa, [])
        self.assertIn(self.ns["COL_FIN_DATA_LIQUIDACAO"], certa.columns)
        self.assertNotIn("Coluna Extra", certa.columns, "a economia tem de continuar valendo")

    def test_checagem_inicial_tambem_tolera_a_escrita(self):
        i = FONTE.index("ausentes = [")
        trecho = FONTE[i - 400:i + 400]
        self.assertIn("_assinatura_coluna_fin(c) not in _assinaturas", trecho)

    def test_nome_exato_tem_prioridade(self):
        """Se a coluna certa existe com o nome exato, ninguem mexe nela."""
        _saida, faltando, renomeadas = self._resolver(
            ["Valor.1", "Modalidade", "Canal.1", "Movimento",
             "Data Liquidação", "Vencimento.1"])
        self.assertEqual(faltando, [])
        self.assertEqual(renomeadas, {}, "renomeou uma coluna que ja estava certa")


# ============================================================================
# 5j. CONTAS A PAGAR — PROGRAMADO NO FUTURO, EFETIVO NO PASSADO
# ============================================================================
class TesteContasAPagarEfetivo(unittest.TestCase):
    """Titulo ja pago entra no dia do pagamento; em aberto fica no
    vencimento. Sem isso, um dia com R$ 1 milhao vencendo aparecia com o
    milhao inteiro mesmo tendo sido pago dias antes."""

    COL_VAL, COL_VENC, COL_LIQ = "Valor.1", "Vencimento.1", "Data Liquidação"
    COL_EFET, COL_MOV = "Liquidação Efetiva", "Movimento"

    def _base(self):
        return pd.DataFrame({
            self.COL_MOV: ["4 - Contas a Pagar"] * 5,
            self.COL_VENC: pd.to_datetime(["2026-07-31"] * 4 + ["2026-09-15"]),
            self.COL_LIQ: pd.to_datetime(["2026-07-31", "2026-07-28", "2026-07-24", None, None]),
            self.COL_VAL: [-290_000.0, -350_000.0, -300_000.0, -60_000.0, -400_000.0],
        })

    def _aplicar(self, df):
        df[self.COL_EFET] = df[self.COL_LIQ]
        df["Data Efetiva"] = df[self.COL_VENC].fillna(df[self.COL_LIQ])
        e_pagar = df[self.COL_MOV].str.contains("pagar", case=False)
        com_baixa = e_pagar & df[self.COL_EFET].notna()
        df.loc[com_baixa, "Data Efetiva"] = df.loc[com_baixa, self.COL_EFET]
        return df.groupby(df["Data Efetiva"].dt.date)[self.COL_VAL].sum()

    def test_pago_antes_sai_do_dia_do_vencimento(self):
        por_dia = self._aplicar(self._base())
        self.assertAlmostEqual(por_dia[date(2026, 7, 31)], -350_000.0, places=2)
        self.assertAlmostEqual(por_dia[date(2026, 7, 28)], -350_000.0, places=2)
        self.assertAlmostEqual(por_dia[date(2026, 7, 24)], -300_000.0, places=2)

    def test_em_aberto_continua_no_vencimento(self):
        por_dia = self._aplicar(self._base())
        self.assertAlmostEqual(por_dia[date(2026, 9, 15)], -400_000.0, places=2)

    def test_total_do_periodo_nao_muda(self):
        """Mover de dia nao pode criar nem sumir com dinheiro."""
        df = self._base()
        total_antes = df[self.COL_VAL].sum()
        self.assertAlmostEqual(self._aplicar(df).sum(), total_antes, places=2)

    def test_regra_esta_no_preparo_e_usa_a_liquidacao_efetiva(self):
        i = FONTE.index("_pagar_com_baixa = _e_pagar")
        trecho = FONTE[i - 900:i + 600]
        self.assertIn('df.loc[_pagar_com_baixa, "Data Efetiva"] = df.loc[_pagar_com_baixa, COL_FIN_LIQ_EFETIVA]',
                      trecho)
        self.assertIn("COL_FIN_LIQ_EFETIVA].notna()", trecho)
        # A liquidacao efetiva vale para o painel todo, nao so na aba
        # Analises, e junta as duas leituras da coluna do CSV.
        j = FONTE.index("df[COL_FIN_LIQ_EFETIVA] = df[COL_FIN_DATA_LIQUIDACAO]")
        montagem = FONTE[j:j + 400]
        self.assertIn("COL_FIN_LIQ_AMPLA", montagem)

    def test_a_receber_usa_a_mesma_fonte_de_baixa(self):
        """As duas pontas precisam olhar a mesma coluna; se o a receber
        olhasse so o CSV, perderia as baixas que vieram da DIARIO."""
        i = FONTE.index("_e_receber = df[COL_FIN_MOVIMENTO] == MOV_RECEBER_AVENCER")
        trecho = FONTE[i:i + 700]
        self.assertIn("_tem_baixa = df[COL_FIN_LIQ_EFETIVA].notna()", trecho)


# ============================================================================
# 5k. DIARIO CONSOLIDADO — HIERARQUIA COM EXPANDIR
# ============================================================================
class TesteDiarioConsolidado(unittest.TestCase):
    """A empresa inteira dia a dia, com as linhas abrindo em modalidade,
    grupo de despesa ou canal."""

    INDICES = ["SALDO INICIAL", "4 - Contas a Pagar", "Frete", "Energia", "1.1.Caixa",
               "3 - Contas a Receber", "Débito", "Boleto Garantido", "TOTAL GERAL"]
    PAIS = [None, None, 1, 1, None, None, 5, 5, None]
    TIPOS = ["saldo_inicial", "movimento", "movimento", "movimento", "movimento",
             "movimento", "movimento", "movimento", "total"]

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["_normalizar_texto", "_classificar_movimento_fin",
                           "_peso_ordem_movimento_fin", "_rotulo_unico_tabela",
                           "_ordenar_com_filhas", "_pais_reordenados"])

    def test_filha_fica_logo_abaixo_da_mae(self):
        """Ordenar so as maes embaralharia as filhas: a filha iria para junto
        de outra linha e passaria a compor um total que nao e o dela."""
        ordem = self.ns["_ordenar_com_filhas"](self.INDICES, self.PAIS, self.TIPOS)
        nomes = [self.INDICES[p] for p in ordem]
        self.assertEqual(nomes[0], "SALDO INICIAL")
        self.assertEqual(nomes[-1], "TOTAL GERAL")
        self.assertEqual(nomes[nomes.index("3 - Contas a Receber") + 1:][:2],
                         ["Débito", "Boleto Garantido"])
        self.assertEqual(nomes[nomes.index("4 - Contas a Pagar") + 1:][:2],
                         ["Frete", "Energia"])

    def test_referencia_da_mae_acompanha_a_reordenacao(self):
        """Depois de reordenar, o indice da mae muda de lugar: sem reescrever
        a referencia, o clique de abrir mostraria as filhas de outra linha."""
        ordem = self.ns["_ordenar_com_filhas"](self.INDICES, self.PAIS, self.TIPOS)
        novos = self.ns["_pais_reordenados"](self.PAIS, ordem)
        for nova, antiga in enumerate(ordem):
            if novos[nova] is None:
                self.assertIsNone(self.PAIS[antiga])
            else:
                mae_antes = self.INDICES[self.PAIS[antiga]]
                mae_depois = self.INDICES[ordem[novos[nova]]]
                self.assertEqual(mae_antes, mae_depois, self.INDICES[antiga])

    def test_rotulo_repetido_nao_quebra_o_indice(self):
        """A mesma modalidade pode aparecer sob duas maes; rotulo repetido
        quebra o indice do DataFrame."""
        rotulos = [self.ns["_rotulo_unico_tabela"]("Débito", i) for i in range(3)]
        self.assertEqual(len(set(rotulos)), 3)
        self.assertTrue(all(r.replace("\u200b", "") == "Débito" for r in rotulos))

    def _montar(self, abertas=()):
        """Monta a tabela como a aba monta: as filhas so entram para as
        linhas marcadas como abertas."""
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "_ordenar_com_filhas",
                             "_pais_reordenados", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        filhas_de = {"4 - Contas a Pagar": ["Frete", "Energia"],
                     "3 - Contas a Receber": ["Débito", "Boleto Garantido"]}
        indices, tipos, pais = [], [], []
        for mae in ["SALDO INICIAL", "1.1.Caixa", "3 - Contas a Receber",
                    "4 - Contas a Pagar", "TOTAL GERAL"]:
            posicao = len(indices)
            indices.append(mae)
            tipos.append("saldo_inicial" if mae == "SALDO INICIAL"
                         else "total" if mae == "TOTAL GERAL" else "movimento")
            pais.append(None)
            if mae in abertas:
                for filha in filhas_de.get(mae, []):
                    indices.append(filha)
                    tipos.append("movimento")
                    pais.append(posicao)
        ordem = ns_local["_ordenar_com_filhas"](indices, pais, tipos)
        df = pd.DataFrame([[1.0] * 31] * len(indices),
                          index=[ns_local["_rotulo_unico_tabela"](indices[p], p) for p in ordem],
                          columns=[f"D{i}" for i in range(31)])
        ns_local["tabela_selecionavel"](
            df, chave="t", tipos_linha=[tipos[p] for p in ordem],
            pais=ns_local["_pais_reordenados"](pais, ordem), filhas_abertas=True,
            comandos_abrir={mae: (mae in abertas, i)
                            for i, mae in enumerate(filhas_de)})
        return capturado, ns_local

    def test_nenhum_metodo_de_teste_duplicado(self):
        """Dois metodos com o mesmo nome numa classe: o ultimo apaga o
        primeiro em silencio. Foi o que aconteceu com um `_falta` antigo,
        que passou a responder pelas chamadas do novo."""
        arvore = ast.parse(open(__file__, encoding="utf-8").read())
        repetidos = []
        for no in ast.walk(arvore):
            if not isinstance(no, ast.ClassDef):
                continue
            vistos = set()
            for item in no.body:
                if isinstance(item, ast.FunctionDef):
                    if item.name in vistos:
                        repetidos.append(f"{no.name}.{item.name}")
                    vistos.add(item.name)
        self.assertEqual(repetidos, [], "metodo definido duas vezes na mesma classe")

    def test_desempacotar_tupla_bate_com_o_que_foi_guardado(self):
        """Uma lista passou a guardar 3 itens por linha e um leitor continuou
        desempacotando em 2. O Python compila, a suite passa, e o app quebra
        NO AR com ValueError -- o erro so existe em tempo de execucao.
        Aconteceu em 20/08/2026 com estilo_linhas_d."""
        from collections import defaultdict

        arvore = ast.parse(FONTE)
        guardados = defaultdict(set)
        for no in ast.walk(arvore):
            if (isinstance(no, ast.Call) and isinstance(no.func, ast.Attribute)
                    and no.func.attr == "append"
                    and isinstance(no.func.value, ast.Name)
                    and len(no.args) == 1 and isinstance(no.args[0], ast.Tuple)):
                guardados[no.func.value.id].add(len(no.args[0].elts))

        problemas = []
        for no in ast.walk(arvore):
            if not isinstance(no, (ast.comprehension, ast.For)):
                continue
            alvo, fonte_iter = no.target, no.iter
            if isinstance(fonte_iter, ast.Call) and fonte_iter.args:
                fonte_iter = fonte_iter.args[0]          # enumerate(lista)
                if isinstance(alvo, ast.Tuple) and len(alvo.elts) == 2:
                    alvo = alvo.elts[1]
            if not isinstance(fonte_iter, ast.Name) or not isinstance(alvo, ast.Tuple):
                continue
            tamanhos = guardados.get(fonte_iter.id)
            if tamanhos and len(alvo.elts) not in tamanhos:
                problemas.append(
                    f"{fonte_iter.id}: guarda {sorted(tamanhos)}, "
                    f"desempacota em {len(alvo.elts)} (linha {no.iter.lineno})")
        self.assertEqual(sorted(set(problemas)), [],
                         "desempacotamento que nao bate com o que foi guardado")

    def test_toda_funcao_chamada_existe(self):
        """Teste que so procura o TEXTO da chamada passa mesmo quando a
        funcao nunca foi definida -- e o app quebra no ar com NameError.
        Aconteceu com _agrega_no_mes_cheio em 19/08/2026: o script que a
        criaria abortou, so a chamada entrou, e a trava estrutural (que
        procurava a chamada) continuou verde."""
        import builtins

        arvore = ast.parse(FONTE)
        definidos = set(dir(builtins))
        for no in ast.walk(arvore):
            if isinstance(no, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                definidos.add(no.name)
                definidos.update(a.arg for a in getattr(no.args, "args", []))
                definidos.update(a.arg for a in getattr(no.args, "kwonlyargs", []))
            elif isinstance(no, ast.Name) and isinstance(no.ctx, ast.Store):
                definidos.add(no.id)
            elif isinstance(no, (ast.Import, ast.ImportFrom)):
                for alias in no.names:
                    definidos.add((alias.asname or alias.name).split(".")[0])
            elif isinstance(no, ast.ExceptHandler) and no.name:
                definidos.add(no.name)

        faltando = {}
        for no in ast.walk(arvore):
            if isinstance(no, ast.Call) and isinstance(no.func, ast.Name):
                if no.func.id not in definidos:
                    faltando.setdefault(no.func.id, no.lineno)
        self.assertEqual(
            faltando, {},
            "funcao chamada e nunca definida: "
            + ", ".join(f"{n}() na linha {l}" for n, l in faltando.items()))

    def test_javascript_gerado_nao_tem_erro_de_sintaxe(self):
        """Um erro de sintaxe no JS nao quebra so o trecho errado: impede o
        script INTEIRO de rodar, e a tabela vira uma imagem -- sem clique,
        sem soma, sem seta. Foi o que aconteceu quando dois blocos escritos
        em momentos diferentes declararam `const ALTURA_LINHA` no mesmo
        escopo. Nenhum teste de Python pegava isso, porque o Python estava
        certo; o defeito nascia no texto que ele montava."""
        for descricao, argumentos in [
            ("tabela simples", dict(tipos_linha=["movimento", "movimento", "total"])),
            ("com hierarquia", dict(tipos_linha=["movimento", "movimento", "total"],
                                    pais=[None, 0, None], filhas_abertas=True,
                                    comandos_abrir={"1.1.Caixa": (True, 0)})),
        ]:
            ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                                 "_rotulo_unico_tabela", "formata_brl",
                                 "tabela_selecionavel"], [])
            capturado = {}
            ns_local["html_embutido"] = dubla_html_embutido(capturado)
            df = pd.DataFrame([[1.0, -2.0]] * 3,
                              index=["1.1.Caixa", "HUB", "TOTAL GERAL"],
                              columns=["18/08", "19/08"])
            ns_local["tabela_selecionavel"](df, chave="t", **argumentos)
            codigo = capturado["codigo"]
            # Do PRIMEIRO <script> ao ULTIMO </script>: a pagina de detalhe
            # que o codigo monta contem um <script> dentro dela, e recortar
            # no primeiro fechamento truncava o texto no meio.
            js = codigo[codigo.index("<script>") + len("<script>"):
                        codigo.rindex("</script>")]

            self.assertEqual(js.count("{"), js.count("}"), f"{descricao}: chaves")
            self.assertEqual(js.count("("), js.count(")"), f"{descricao}: parenteses")

    # Uma string aberta e nao fechada na mesma linha e erro de sintaxe -- e
            # nenhuma contagem de chaves pega isso. Foi o que aconteceu com um
            # "\\r\\n" que o Python transformou em quebra de linha DE VERDADE dentro
            # do JavaScript: a tabela inteira parou de responder e o verificador
            # continuou dizendo que estava tudo certo.
            #
            # Precisa ser um leitor de verdade, e nao contagem de aspas: aspa dentro
            # de aspa e comum ('"' + valor) e a contagem sozinha da falso positivo.
            dentro_de_crase = False
            for numero, linha in enumerate(js.split("\n"), 1):
                aberta = None
                pulo = False
                saltar_ate = -1
                for posicao, caractere in enumerate(linha):
                    if posicao <= saltar_ate:
                        continue
                    if pulo:
                        pulo = False
                        continue
                    if caractere == "\\":
                        pulo = True
                        continue
                    if dentro_de_crase:
                        if caractere == "`":
                            dentro_de_crase = False
                        continue
                    if aberta:
                        if caractere == aberta:
                            aberta = None
                        continue
                    if caractere == "`":
                        dentro_de_crase = True
                    elif caractere in "'\"":
                        aberta = caractere
                    elif caractere == "/":
                        if linha[posicao + 1:posicao + 2] == "/":
                            break
                        # Expressao regular como /"/g: as aspas ali dentro nao abrem
                        # string nenhuma. Sem pular isso, o leitor acusa erro onde
                        # nao ha -- e um verificador que mente e pior que nenhum.
                        anterior = linha[:posicao].rstrip()
                        if not anterior or anterior[-1] in "(,=:[!&|?{;+":
                            fecha = posicao + 1
                            while fecha < len(linha):
                                if linha[fecha] == "\\":
                                    fecha += 2
                                    continue
                                if linha[fecha] == "/":
                                    break
                                fecha += 1
                            saltar_ate = fecha
                if aberta:
                    self.fail(
                        f"{descricao}: linha {numero} abre uma string com {aberta} e nao fecha "
                        f"-- {linha.strip()[:60]}")


            declaracoes = {}
            for linha in js.split("\n"):
                achado = re.match(r"(\s*)(?:const|let)\s+([A-Za-z_$][\w$]*)\s*=", linha)
                if achado:
                    chave = (len(achado.group(1)), achado.group(2))
                    declaracoes[chave] = declaracoes.get(chave, 0) + 1
            repetidas = [nome for (_nivel, nome), vezes in declaracoes.items() if vezes > 1]
            self.assertEqual(repetidas, [],
                             f"{descricao}: declarado mais de uma vez no mesmo escopo "
                             f"-- isso e erro de sintaxe e mata o script inteiro")

    def test_a_altura_da_tabela_e_MEDIDA_e_nao_chutada(self):
        """26/08/2026: em tela menor a ultima linha (o TOTAL GERAL) sumia
        debaixo da barra de soma. A conta era em pixels cravados -- cabecalho
        + altura de linha x quantidade + 18px reservados para a barra de
        rolagem. Com o Windows em 125% ou 150% de escala, o padrao em notebook
        de tela pequena, a barra passa dos 18px; o excedente vazava e o
        overflow-y:hidden cortava.

        Deixar o navegador MEDIR resolve para qualquer escala, zoom, fonte do
        sistema ou espessura de barra -- sem ninguem ter de adivinhar nenhuma
        delas."""
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("rolagem.style.height = 'auto'", corpo,
                      "a altura voltou a ser calculada em vez de medida")
        self.assertIn("rolagem.offsetHeight", corpo)
        # E a conta antiga tem de continuar como plano B: medicao que falha
        # (quadro escondido, tela ainda desenhando) nao pode virar altura zero.
        self.assertIn("medida > 0 ? medida : alturaCaixa", corpo)

    def test_a_reserva_da_barra_cobre_tela_com_escala(self):
        """Este numero vale so ate o JavaScript medir, no primeiro desenho.
        Sobrar alguns pixels por um instante e invisivel; faltar corta a linha
        do TOTAL GERAL, que e justo a que se olha."""
        ns = carregar([], ["ALTURA_BARRA_ROLAGEM_PX", "FOLGA_ABAIXO_TABELA_PX"])
        self.assertGreaterEqual(ns["ALTURA_BARRA_ROLAGEM_PX"], 24,
                                "18px nao cobre tela em 125% ou 150% de escala")
        self.assertGreaterEqual(ns["FOLGA_ABAIXO_TABELA_PX"], 12,
                                "sem folga a caixa e a barra de soma se tocam")
        # A folga tem de ser USADA nos dois lugares: no HTML servido e no JS.
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("+ FOLGA_ABAIXO_TABELA_PX", corpo)
        self.assertIn("+ FOLGA_ABAIXO_PX", corpo)

    def test_uma_conta_de_altura_so_no_javascript(self):
        """Duas versoes da mesma conta no mesmo script foi o que gerou a
        duplicacao. Uma so, e com nome unico."""
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertEqual(corpo.count("function ajustarCaixa()"), 1)
        self.assertNotIn("function avisarAltura()", corpo,
                         "voltou a segunda funcao de altura")

    def _mapa_de_lancamentos(self):
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO",
                       "TETO_LANCAMENTOS_NA_PAGINA"])
        dia = pd.Timestamp("2026-08-21")
        df = pd.DataFrame({
            "Data Efetiva": [dia] * 3,
            ns["COL_FIN_MOVIMENTO"]: ["3 - Contas a Pagar"] * 3,
            ns["COL_FIN_NUMERO"]: ["NF 1001", "NF 1002", "NF 1003"],
            ns["COL_FIN_HISTORICO"]: ["FORNECEDOR A LTDA"] * 3,
            ns["COL_FIN_CANAL"]: ["LOJA", "HUB LOGISTICO", "LOJA"],
            ns["COL_FIN_MODALIDADE"]: ["", "", ""],
            ns["COL_FIN_GRUPO_DESPESA"]: ["Ativo Permanente"] * 3,
            ns["COL_FIN_VENCIMENTO"]: [dia] * 3,
            ns["COL_FIN_DATA_LIQUIDACAO"]: [dia, pd.NaT, dia],
            ns["COL_FIN_VALOR"]: [-10_000.00, -4_426.75, -10_000.00],
        })
        mapa, _cortou = ns["montar_lancamentos_por_celula"](
            df, {dia: "21/08"}, df[ns["COL_FIN_GRUPO_DESPESA"]].astype(str))
        return mapa, ns

    def test_lancamentos_da_celula_somam_o_valor_mostrado(self):
        """A conta da aba nova tem de bater com a celula clicada -- e nao
        adianta ela abrir se mostrar outro numero."""
        mapa, _ns = self._mapa_de_lancamentos()
        chave = "Ativo Permanente||21/08"
        self.assertIn(chave, mapa)
        self.assertEqual(len(mapa[chave]), 3)
        self.assertAlmostEqual(sum(l["Valor"] for l in mapa[chave]), -24_426.75, places=2)
        # Procura pelo documento, e nao pela posicao: a lista vem ordenada
        # por valor, para o corte por teto preservar o que explica a celula.
        em_aberto = [l for l in mapa[chave] if l["Número"] == "NF 1002"]
        self.assertEqual(len(em_aberto), 1)
        self.assertEqual(em_aberto[0]["Liquidação"], "",
                         "titulo em aberto tem de aparecer sem data, nao como NaT")

    def test_lancamento_de_outro_dia_nao_entra_na_celula(self):
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO",
                       "TETO_LANCAMENTOS_NA_PAGINA"])
        dias = [pd.Timestamp("2026-08-21"), pd.Timestamp("2026-08-22")]
        df = pd.DataFrame({
            "Data Efetiva": dias,
            ns["COL_FIN_MOVIMENTO"]: ["3 - Contas a Pagar"] * 2,
            ns["COL_FIN_NUMERO"]: ["NF 1", "NF 2"],
            ns["COL_FIN_HISTORICO"]: ["FORNECEDOR A", "FORNECEDOR B"],
            ns["COL_FIN_CANAL"]: ["LOJA", "LOJA"],
            ns["COL_FIN_GRUPO_DESPESA"]: ["Ativo Permanente"] * 2,
            ns["COL_FIN_VENCIMENTO"]: dias,
            ns["COL_FIN_DATA_LIQUIDACAO"]: dias,
            ns["COL_FIN_VALOR"]: [-100.0, -200.0],
        })
        # Só o dia 21 está na tela: o lançamento do 22 fica de fora.
        mapa, _cortou = ns["montar_lancamentos_por_celula"](
            df, {dias[0]: "21/08"}, df[ns["COL_FIN_GRUPO_DESPESA"]].astype(str))
        self.assertEqual(list(mapa), ["Ativo Permanente||21/08"])
        self.assertEqual(len(mapa["Ativo Permanente||21/08"]), 1)

    def test_teto_de_lancamentos_por_celula(self):
        """Eles viajam junto com a pagina: sem teto, uma celula com milhares
        de lancamentos pesaria no carregamento."""
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO",
                       "TETO_LANCAMENTOS_NA_PAGINA"])
        teto = ns["TETO_LANCAMENTOS_POR_CELULA"]
        dia = pd.Timestamp("2026-08-21")
        n = teto + 50
        df = pd.DataFrame({
            "Data Efetiva": [dia] * n,
            ns["COL_FIN_MOVIMENTO"]: ["3 - Contas a Pagar"] * n,
            ns["COL_FIN_NUMERO"]: [f"NF {i}" for i in range(n)],
            ns["COL_FIN_HISTORICO"]: ["FORNECEDOR A"] * n,
            ns["COL_FIN_CANAL"]: ["LOJA"] * n,
            ns["COL_FIN_GRUPO_DESPESA"]: ["Frete"] * n,
            ns["COL_FIN_VENCIMENTO"]: [dia] * n,
            ns["COL_FIN_DATA_LIQUIDACAO"]: [dia] * n,
            ns["COL_FIN_VALOR"]: [-1.0] * n,
        })
        mapa, cortou = ns["montar_lancamentos_por_celula"](
            df, {dia: "21/08"}, df[ns["COL_FIN_GRUPO_DESPESA"]].astype(str))
        self.assertEqual(len(mapa["Frete||21/08"]), teto)
        self.assertTrue(cortou, "a tela precisa saber que cortou, para avisar")

    def test_nenhuma_celula_fica_sem_chave(self):
        """Uma celula sem chave e a falha mais irritante possivel: o valor
        continua sublinhado prometendo detalhe e o duplo clique nao abre.
        Ja aconteceu duas vezes por causa de teto de pagina cortando linhas
        soltas -- por isso o unico limite agora e POR CELULA."""
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO"])
        self.assertNotIn("TETO_LANCAMENTOS_NA_PAGINA", FONTE,
                         "o teto de pagina ja apagou celulas duas vezes")
        n = 6000
        grupos = [f"Grupo {i}" for i in range(40)]
        dias = pd.to_datetime("2026-08-01") + pd.to_timedelta(
            [i % 13 for i in range(n)], unit="D")
        # O "Grupo 0" so tem centavos: por valor, seria o primeiro a cair.
        df = pd.DataFrame({
            "Data Efetiva": dias,
            ns["COL_FIN_MOVIMENTO"]: "3 - Contas a Pagar",
            ns["COL_FIN_NUMERO"]: [f"NF {i}" for i in range(n)],
            ns["COL_FIN_HISTORICO"]: "FORN",
            ns["COL_FIN_CANAL"]: "LOJA",
            ns["COL_FIN_GRUPO_DESPESA"]: [grupos[i % 40] for i in range(n)],
            ns["COL_FIN_VENCIMENTO"]: dias,
            ns["COL_FIN_DATA_LIQUIDACAO"]: dias,
            ns["COL_FIN_VALOR"]: [-0.01 if i % 40 == 0 else -1000.0 - i for i in range(n)],
        })
        rotulos = {d: pd.Timestamp(d).strftime("%d/%m")
                   for d in df["Data Efetiva"].dt.normalize().unique()}
        mapa, _cortou = ns["montar_lancamentos_por_celula"](
            df, rotulos, df[ns["COL_FIN_GRUPO_DESPESA"]].astype(str))
        nos_dados = df.groupby(
            [df[ns["COL_FIN_GRUPO_DESPESA"]], df["Data Efetiva"].dt.normalize()]).ngroups
        self.assertEqual(len(mapa), nos_dados,
                         "toda celula que existe nos dados precisa ter chave")

    def test_detalhe_cobre_linha_fechada_e_linha_aberta(self):
        """A linha aberta mostra o grupo de despesa; a fechada mostra o
        movimento. O detalhe precisa existir para os DOIS rotulos -- antes
        ele dependia de saber quais linhas estavam abertas, e quando essa
        informacao nao chegava, a linha filha ficava sublinhada e o duplo
        clique nao abria nada."""
        ns = carregar(["_normalizar_texto", "_classificar_movimento_fin",
                       "montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "TETO_LANCAMENTOS_NA_PAGINA",
                       "COL_FIN_CANAL", "COL_FIN_MODALIDADE", "COL_FIN_GRUPO_DESPESA",
                       "COL_FIN_VENCIMENTO", "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR",
                       "COL_FIN_MOVIMENTO", "COL_FIN_NUMERO", "COL_FIN_HISTORICO"])
        dia = pd.Timestamp("2026-08-25")
        df = pd.DataFrame({
            "Data Efetiva": [dia] * 2,
            ns["COL_FIN_MOVIMENTO"]: ["3 - Contas a Pagar"] * 2,
            ns["COL_FIN_NUMERO"]: ["NF 1", "NF 2"],
            ns["COL_FIN_HISTORICO"]: ["FORN A", "FORN B"],
            ns["COL_FIN_CANAL"]: ["LOJA", "HUB"],
            ns["COL_FIN_MODALIDADE"]: ["", ""],
            ns["COL_FIN_GRUPO_DESPESA"]: ["Custo s/ Venda", "Custo - Mercadoria"],
            ns["COL_FIN_VENCIMENTO"]: [dia] * 2,
            ns["COL_FIN_DATA_LIQUIDACAO"]: [dia] * 2,
            ns["COL_FIN_VALOR"]: [-3000.0, -1_070_186.40],
        })
        mapa = {}
        por_movimento = df[ns["COL_FIN_MOVIMENTO"]].astype(str)
        for rotulagem in (por_movimento,
                          df[ns["COL_FIN_GRUPO_DESPESA"]].astype(str).str.strip()):
            parcial, _c = ns["montar_lancamentos_por_celula"](df, {dia: "25/08"}, rotulagem)
            for chave, linhas in parcial.items():
                mapa.setdefault(chave, linhas)
        # Fechada: o movimento inteiro. Aberta: cada grupo.
        self.assertIn("3 - Contas a Pagar||25/08", mapa)
        self.assertIn("Custo s/ Venda||25/08", mapa)
        self.assertIn("Custo - Mercadoria||25/08", mapa)
        self.assertAlmostEqual(
            sum(l["Valor"] for l in mapa["Custo s/ Venda||25/08"]), -3000.0, places=2)

    def test_codigo_monta_o_detalhe_para_as_duas_rotulagens(self):
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertIn("_rotulagens = [_por_movimento]", trecho)
        self.assertIn("lancamentos_dc.setdefault(_chave, _linhas)", trecho)
        self.assertIn('lancamentos_dc.pop("||", None)', trecho,
                      "a chave vazia (linhas sem grupo) nao pode virar celula")

    def test_categoria_com_valor_ausente_nao_derruba_o_mapa(self):
        """astype(str) numa coluna CATEGORIA devolve o ausente como nan de
        verdade, nao como o texto "nan" -- entao um replace depois nao pega,
        e esse nan derruba a escrita do mapa INTEIRO. Foi a mensagem que o
        painel finalmente mostrou: "Out of range float values are not JSON
        compliant: nan". Canal, Modalidade e Grupo de Despesa sao categorias."""
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO"])
        dia = pd.Timestamp("2026-08-24")
        df = pd.DataFrame({
            "Data Efetiva": [dia] * 2,
            ns["COL_FIN_MOVIMENTO"]: pd.Categorical(["3 - Contas a Pagar"] * 2),
            ns["COL_FIN_NUMERO"]: ["NF 1", None],
            ns["COL_FIN_HISTORICO"]: pd.Categorical(["FORN A", None]),
            ns["COL_FIN_CANAL"]: pd.Categorical(["LOJA", None]),
            ns["COL_FIN_MODALIDADE"]: pd.Categorical([None, None]),
            ns["COL_FIN_GRUPO_DESPESA"]: pd.Categorical(["Frete", None]),
            ns["COL_FIN_VENCIMENTO"]: [dia, pd.NaT],
            ns["COL_FIN_DATA_LIQUIDACAO"]: [pd.NaT] * 2,
            ns["COL_FIN_VALOR"]: [-79_843.01, float("nan")],
        })
        mapa, _cortou = ns["montar_lancamentos_por_celula"](
            df, {dia: "24/08"}, df[ns["COL_FIN_MOVIMENTO"]].astype(str))
        # allow_nan=False e o que o navegador faz: recusa nan.
        texto = json.dumps(mapa, ensure_ascii=False, allow_nan=False)
        self.assertIn("3 - Contas a Pagar||24/08", json.loads(texto))
        for lancamento in json.loads(texto)["3 - Contas a Pagar||24/08"]:
            for campo, valor in lancamento.items():
                self.assertNotEqual(valor, "nan", f"{campo} virou o texto 'nan'")

    def test_fluxo_mensal_nao_tem_coluna_de_total(self):
        """So os meses. Uma coluna que ora soma (a receber, a pagar) ora
        mostra posicao (caixa, banco) confunde mais do que ajuda."""
        self.assertNotIn('pivot_m["TOTAL / SALDO DE ABERTURA"]', FONTE)
        i = FONTE.index("📋 Movimentos por Mês")
        trecho = FONTE[max(0, i - 6000):i + 2000]
        self.assertNotIn("TOTAL / SALDO DE ABERTURA", trecho)

    def test_mapa_e_sempre_legivel_pelo_navegador(self):
        """UM lancamento com valor vazio bastava para o navegador nao
        conseguir ler o mapa INTEIRO -- Python escreve NaN, que nao e JSON
        valido -- e ai TODA celula respondia "sem detalhe". Foi o defeito
        mais caro desta sessao: o servidor mostrava as chaves certas
        enquanto o navegador estava com o mapa vazio."""
        ns = carregar(["montar_lancamentos_por_celula"],
                      ["TETO_LANCAMENTOS_POR_CELULA", "COL_FIN_CANAL", "COL_FIN_MODALIDADE",
                       "COL_FIN_GRUPO_DESPESA", "COL_FIN_VENCIMENTO",
                       "COL_FIN_DATA_LIQUIDACAO", "COL_FIN_VALOR", "COL_FIN_MOVIMENTO",
                       "COL_FIN_NUMERO", "COL_FIN_HISTORICO"])
        dia = pd.Timestamp("2026-08-24")
        df = pd.DataFrame({
            "Data Efetiva": [dia] * 3,
            ns["COL_FIN_MOVIMENTO"]: ["3 - Contas a Pagar"] * 3,
            ns["COL_FIN_NUMERO"]: ["NF 1", "NF 2", "NF 3"],
            ns["COL_FIN_HISTORICO"]: ["FORN A", "FORN B", "FORN C"],
            ns["COL_FIN_CANAL"]: ["LOJA"] * 3,
            ns["COL_FIN_MODALIDADE"]: [""] * 3,
            ns["COL_FIN_GRUPO_DESPESA"]: ["Frete"] * 3,
            ns["COL_FIN_VENCIMENTO"]: [dia] * 3,
            ns["COL_FIN_DATA_LIQUIDACAO"]: [dia] * 3,
            ns["COL_FIN_VALOR"]: [-79843.01, float("nan"), float("inf")],
        })
        mapa, _cortou = ns["montar_lancamentos_por_celula"](
            df, {dia: "24/08"}, df[ns["COL_FIN_MOVIMENTO"]].astype(str))
        # allow_nan=False e o que o navegador faz: recusa NaN e Infinity.
        texto = json.dumps(mapa, ensure_ascii=False, allow_nan=False)
        self.assertIn("3 - Contas a Pagar||24/08", json.loads(texto))
        self.assertNotIn("NaN", texto)
        self.assertNotIn("Infinity", texto)

    def test_codigo_recusa_escrever_mapa_ilegivel(self):
        """E, ao escrever, o Python tem de RECUSAR NaN em vez de gerar um
        texto que o navegador nao le. Se escapar, a tela avisa."""
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("allow_nan=False", corpo)
        self.assertIn("não consegui ler os lançamentos", corpo,
                      "o navegador precisa avisar quando a leitura falha")

    def test_celula_com_detalhe_e_clicavel_no_html(self):
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        # A marca tem de estar na PRÓPRIA célula, não só na folha de estilo:
        # sem ela, nada indica que aquele número abre alguma coisa.
        self.assertRegex(html, r'<td class="[^"]*com-detalhe[^"]*"[^>]*data-k=',
                         "a celula com detalhe precisa se mostrar clicavel")
        self.assertIn('data-k="Ativo Permanente||21/08"', html)
        self.assertIn("dblclick", html, "um clique so continua servindo para somar")
        self.assertIn("painel-detalhe", html)
        embutido = json.loads(html.split('id="detalhes">')[1].split("</script>")[0])
        self.assertIn("Ativo Permanente||21/08", embutido)

    def test_aba_de_lancamentos_segue_o_padrao_do_painel(self):
        """A leitura ali e longa: linhas alternadas para o olho nao pular de
        linha, cabecalho discreto em vez de branco forte, realce sob o
        cursor e o total fixo no rodape."""
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        for marca, motivo in [
            ("nth-child(even)", "linhas alternadas"),
            ("tbody tr:hover", "realce sob o cursor"),
            ("td.pos", "valor colorido por sinal"),
            ("td.neg", "valor colorido por sinal"),
            ("white-space:normal", "o historico precisa poder quebrar linha"),
            ("position:sticky; bottom:0", "total fixo no rodape"),
            ("FONTE_PADRAO_TABELA" if False else "Source Sans Pro", "mesma fonte do painel"),
        ]:
            self.assertIn(marca, html, motivo)

    def test_aba_tem_um_unico_botao_de_download(self):
        """Um botao so, e o CSV no formato que o Excel em portugues abre sem
        pedir nada: ponto e virgula, virgula decimal e BOM para os acentos."""
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        self.assertEqual(html.count('id="baixar"'), 1, "tem de ser um botao so")
        self.assertIn("join(';')", html, "separador ponto e virgula")
        self.assertIn("replace('.', ',')", html, "virgula decimal")
        self.assertIn("\ufeff", html, "BOM: sem ele os acentos saem trocados no Excel")
        self.assertIn("URL.createObjectURL", html)
        self.assertIn("link.download", html)

    def test_duplo_clique_abre_a_aba_com_window_open(self):
        """ABRIR JANELA e permitido; LEVAR PARA OUTRA PAGINA nao e. O console
        acusou "Unsafe attempt to initiate navigation" quando troquei
        window.open por um link -- eu confundi as duas coisas e quebrei o que
        funcionava. Este teste RODA o JavaScript para garantir que o caminho
        e o permitido."""
        import shutil
        import subprocess
        import tempfile

        node = shutil.which("node") or shutil.which("nodejs")
        if not node:
            self.skipTest("sem motor JavaScript nesta maquina")

        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        inicio = html.index("<script>", html.index('id="detalhes"')) + len("<script>")
        js = html[inicio:html.index("</script>", inicio)]

        # ABRIR JANELA e permitido ("allow-popups"); LEVAR PARA OUTRA PAGINA
        # nao e ("allow-top-navigation" fica de fora). Sao coisas diferentes,
        # e confundi-las custou varios dias: window.open funciona, link e
        # window.parent.open viram navegacao e o navegador barra.
        self.assertIn("window.open('', '_blank')", js,
                      "abrir janela e o caminho permitido")
        # Sem comentarios: eles CITAM os padroes proibidos de proposito, para
        # que o proximo a mexer saiba por que nao usar.
        sem_comentario = re.sub(r"//[^\n]*", "", js)
        for proibido in ("location.href =", "target = '_blank'", "window.parent.open("):
            self.assertNotIn(proibido, sem_comentario,
                             f"{proibido} vira navegacao e o navegador barra")
        # Ler o documento de fora NAO e navegacao -- e o que faz a seta
        # funcionar. Por isso nao entra na lista acima.

        roteiro = """
        let ouvinteDaCelula = null;
        const celula = {dataset:{k:'Ativo Permanente||21/08',v:'-1',l:'0',c:'0'},
          style:{}, classList:{add(){},remove(){},toggle(){},contains(){return false}},
          addEventListener(t,f){ if (t === 'dblclick') ouvinteDaCelula = f; },
          closest(s){return s==='td[data-k]'?this:null}};
        const jsonTag = {textContent: JSON.stringify(
          {'Ativo Permanente||21/08':[{'Valor':-1}]})};
        const doc = {}; const corpo = {innerHTML:'', style:{cssText:''}};
        const elemento = () => ({style:{cssText:''}, textContent:'', dataset:{},
          classList:{add(){},remove(){},toggle(){},contains(){return false}},
          addEventListener(){}, click(){}, appendChild(){}, querySelector(){return null}});
        global.document = {addEventListener(t,f){doc[t]=f},
          getElementById(id){return id==='detalhes'?jsonTag:elemento()},
          querySelector(){return elemento()},
          querySelectorAll(s){return (s.includes('data-v')||s.includes('data-k'))?[celula]:[]},
          createElement: elemento, body: corpo, documentElement:{scrollHeight:1}};
        let abriuJanela = false;
        global.window = {parent:{postMessage(){}}, top:null, frameElement:null,
          location:{href:'http://x/', reload(){}},
          open(){abriuJanela = true; return {document:{write(){},close(){}}}}};
        global.URL={createObjectURL(){return 'blob:'},revokeObjectURL(){}};
        global.Blob=function(){};
        eval(JS_DA_TABELA);
        ouvinteDaCelula({target: celula, preventDefault(){}});
        console.log(abriuJanela ? 'abriu janela' : 'nao abriu');
        """
        with tempfile.TemporaryDirectory() as pasta:
            caminho = f"{pasta}/t.js"
            with open(caminho, "w", encoding="utf-8") as arquivo:
                arquivo.write("const JS_DA_TABELA = " + json.dumps(js) + ";\n" + roteiro)
            saida = subprocess.run([node, caminho], capture_output=True, text=True)
        self.assertEqual(saida.returncode, 0, f"o JavaScript nao rodou: {saida.stderr[:300]}")
        self.assertEqual(saida.stdout.strip(), "abriu janela",
                         "o duplo clique tem de abrir a aba com os lancamentos")

    def test_abertura_da_aba_e_o_minimo_possivel(self):
        """Tres linhas: abre, escreve, fecha -- exatamente a versao em que a
        aba abria. Cada verificacao que acrescentei aqui (document.open
        antes, reconferir se a janela seguia aberta) foi uma chance a mais de
        o caminho ser abandonado no meio, e nenhuma resolveu nada."""
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        # CAMINHO COMPROVADO POR PRINT (20/08/2026): a aba abriu com o
        # endereco "about:blank" -- janela em branco com a pagina escrita
        # dentro. Cada variacao que tentei aqui quebrou o que funcionava:
        #   - arquivo em memoria: devolve janela mesmo quando o navegador
        #     barra o conteudo, entao o codigo nunca cai no jeito certo
        #   - link ou window.parent.open: viram navegacao da pagina, que o
        #     quadro nao tem permissao de fazer
        #   - document.open() antes de escrever, reconferir se a janela
        #     seguia aberta: so criam caminhos para desistir no meio
        self.assertIn("const aba = window.open('', '_blank');", html)
        self.assertIn("aba.document.write(pagina);", html)
        for proibido in ("window.open(enderecoPagina", "aba.document.open()",
                         "!aba.closed", "window.parent.open("):
            self.assertNotIn(proibido, html,
                             f"{proibido} ja quebrou a abertura da aba antes")

    def test_duplo_clique_tem_caminho_alternativo(self):
        """O navegador pode recusar a abertura da aba vinda de dentro do
        quadro -- e recusa em SILENCIO. Sem um plano B, o duplo clique
        simplesmente nao fazia nada e nao havia como saber por que."""
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        self.assertIn("window._corpoOriginal", html)
        self.assertIn("Voltar para a tabela", html, "precisa de volta para a tabela")
        self.assertIn("URL.createObjectURL", html, "o CSV continua disponivel")

    def test_duplo_clique_usa_delegacao_e_nunca_fica_mudo(self):
        """Ouvinte por celula depende do momento em que o script roda; a
        delegacao no documento nao. E o caminho de "sem detalhe" precisa
        DIZER alguma coisa: o silencio ali atrasou o diagnostico por dois
        turnos seguidos."""
        mapa, _ns = self._mapa_de_lancamentos()
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[-24_426.75]], index=["Ativo Permanente"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula=mapa)
        html = capturado["codigo"]
        self.assertIn("querySelectorAll('td[data-k]')", html,
                      "um ouvinte por celula, como na versao em que a aba abria")
        self.assertIn("sem detalhe para", html,
                      "o caminho sem detalhe precisa avisar, nao ficar mudo")

    def test_tela_mostra_quantas_celulas_tem_detalhe(self):
        """Quando o duplo clique nao responde, a primeira pergunta e se o
        detalhe chegou a ser montado."""
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertIn("Detalhe por célula — diagnóstico", trecho)
        # E so para quem administra: e uma ferramenta de investigacao, nao
        # informacao para quem usa o painel no dia a dia.
        posicao = trecho.index("Detalhe por célula — diagnóstico")
        self.assertIn("if eh_admin:", trecho[max(0, posicao - 300):posicao])
        self.assertIn("células com detalhe:", trecho)
        self.assertIn("com dia reconhecido:", trecho,
                      "o diagnostico precisa dizer onde o detalhe se perde")

    def test_celula_sem_detalhe_nao_finge_ser_clicavel(self):
        ns_local = carregar(["_normalizar_texto", "_peso_ordem_movimento_fin",
                             "_rotulo_unico_tabela", "formata_brl", "tabela_selecionavel"], [])
        capturado = {}
        ns_local["html_embutido"] = dubla_html_embutido(capturado)
        tabela = pd.DataFrame([[1.0]], index=["Sem detalhe"], columns=["21/08"])
        ns_local["tabela_selecionavel"](tabela, chave="t", detalhes_por_celula={})
        self.assertNotIn("com-detalhe\"", capturado["codigo"])
        self.assertNotIn("data-k=", capturado["codigo"])

    def test_seta_fica_na_propria_linha(self):
        """A escolha do que abrir fica DENTRO da tabela, na setinha da
        linha -- nao num campo separado acima dela."""
        capturado, _ = self._montar()
        html = capturado["codigo"]
        self.assertEqual(len(re.findall(r'class="seta comando', html)), 2,
                         "so as maes com detalhe podem ter seta")
        self.assertIn("data-linha=", html)
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertNotIn("st.multiselect", trecho,
                         "voltou o campo separado para escolher o que abrir")

    def test_clique_na_seta_faz_o_quadro_crescer(self):
        """A altura do quadro e decidida ao desenhar: um pedido de crescer
        vindo de dentro dele e ignorado. Por isso o clique navega a pagina e
        a tabela volta redesenhada -- com as filhas contando na altura."""
        fechado, ns_local = self._montar()
        aberto, _ = self._montar({"4 - Contas a Pagar"})
        self.assertEqual(aberto["altura"] - fechado["altura"],
                         ns_local["ALTURA_LINHA_TABELA_PX"] * 2,
                         "abrir uma linha com 2 filhas tem de crescer 2 linhas")
        self.assertNotIn('style="display:none"', aberto["codigo"],
                         "filha de linha aberta nao pode nascer escondida")
        self.assertIn("alvo.click()", aberto["codigo"],
                      "o clique precisa acionar o botao da pagina")

    def test_seta_da_linha_aberta_vem_girada(self):
        aberto, _ = self._montar({"4 - Contas a Pagar"})
        self.assertIn("seta comando aberta", aberto["codigo"])

    def test_separador_da_url_nao_quebra_nome_com_virgula(self):
        """Nome de modalidade tem virgula; com virgula de separador, uma
        linha aberta viraria duas e nenhuma seria encontrada."""
        ns_local = carregar(["linhas_abertas_da_url"],
                            ["PARAM_LINHAS_ABERTAS", "SEPARADOR_LINHAS_ABERTAS"])
        self.assertNotEqual(ns_local["SEPARADOR_LINHAS_ABERTAS"], ",")
        ns_local["st"] = type("st", (), {"query_params": {
            ns_local["PARAM_LINHAS_ABERTAS"]:
                "3 - Contas a Receber~Crédito à Vista, parcelado"}})()
        self.assertEqual(ns_local["linhas_abertas_da_url"](),
                         ["3 - Contas a Receber", "Crédito à Vista, parcelado"])

    def test_seta_aciona_um_botao_da_pagina(self):
        """O quadro da tabela NAO pode trocar o endereco da pagina -- o
        navegador recusa a navegacao sem avisar, e foi por isso que a seta
        ficou muda mesmo com o resto do script funcionando. O que ele pode e
        apertar um botao que ja existe na pagina, porque as duas sao da
        mesma origem."""
        ns_local = carregar(["botoes_de_abrir"], ["PREFIXO_BOTAO_ABRIR"])
        estado, estilos = {}, []

        class FakeST:
            session_state = estado
            clicado = None

            def markdown(self, html, **kw):
                estilos.append(html)

            def button(self, rotulo, key=None, **kw):
                return key == FakeST.clicado

        ns_local["st"] = FakeST()
        linhas = ["1.1.Caixa", "4 - Contas a Pagar"]

        comandos = ns_local["botoes_de_abrir"](linhas, "dc")
        self.assertEqual(comandos, {"1.1.Caixa": (False, 0), "4 - Contas a Pagar": (False, 1)})
        self.assertIn("left:-9999px", estilos[0], "o botao tem de ficar fora da tela")

        FakeST.clicado = "dcabrir_0"
        comandos = ns_local["botoes_de_abrir"](linhas, "dc")
        self.assertTrue(comandos["1.1.Caixa"][0], "o clique tinha de abrir a linha")
        comandos = ns_local["botoes_de_abrir"](linhas, "dc")
        self.assertFalse(comandos["1.1.Caixa"][0], "clicar de novo tinha de fechar")

    def test_js_procura_o_botao_e_tem_plano_b(self):
        i = FONTE.index("def tabela_selecionavel(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("{PREFIXO_BOTAO_ABRIR}' + indice + ' button'", corpo,
                      "o JS precisa achar o botao pela classe da chave")
        self.assertIn("innerText.trim() ===", corpo,
                      "sem a classe por chave, o plano B e achar pelo texto")
        self.assertIn("alvo.click()", corpo)

    def test_botao_e_desenhado_antes_da_tabela(self):
        """O Streamlit conta o clique no ponto em que o botao aparece: se ele
        vier depois, a tabela e montada com o estado velho e o clique so faz
        efeito na proxima interacao."""
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        posicao_botoes = trecho.index("botoes_de_abrir(")
        posicao_tabela = trecho.index("tabela_selecionavel(")
        self.assertLess(posicao_botoes, posicao_tabela)

    def test_periodo_sobrevive_ao_clique_na_seta(self):
        """Abrir uma linha recarrega a pagina, e recarregar apaga o que
        estava nos campos. Sem guardar o periodo na URL, cada abertura
        jogaria as datas de volta para o padrao."""
        ns_local = carregar(["guardar_periodo_na_url"], [])
        guardadas = {}
        ns_local["st"] = type("st", (), {"query_params": guardadas})()
        ns_local["datetime"] = datetime
        ns_local["guardar_periodo_na_url"]("dc", date(2026, 7, 20), date(2026, 8, 15))
        self.assertEqual(guardadas, {"dc_ini": "2026-07-20", "dc_fim": "2026-08-15"})
        self.assertEqual(ns_local["guardar_periodo_na_url"]("dc"),
                         (date(2026, 7, 20), date(2026, 8, 15)))
        guardadas.clear()
        self.assertEqual(ns_local["guardar_periodo_na_url"]("dc"), (None, None),
                         "URL sem o periodo nao pode explodir")
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertIn('guardar_periodo_na_url("dc", data_ini_dc, data_fim_dc)', trecho)
        # Sem esta condicao, TODAS as linhas abririam sempre -- a tabela
        # nasceria com 60 linhas e o clique na seta nao faria diferenca.
        self.assertIn("if movimento in abertas_dc and coluna_abertura in recorte.columns:",
                      trecho, "a abertura deixou de respeitar a URL")
        self.assertIn("if MOV_RECEBER_META in abertas_dc else []", trecho)

    def test_total_do_fluxo_diario_soma_o_mes_e_nao_o_recorte(self):
        """Esta aba mostra UM mes de cada vez (misturar meses e o papel do
        Diario Consolidado). A leitura da coluna final e "quanto este
        movimento soma no mes", nao "quanto soma nos dias que sobraram na
        tela": olhando de 19 a 31/08, o total traz agosto inteiro."""
        i = FONTE.index("with tab_fin_diario:")
        trecho = FONTE[i:FONTE.index("with tab_fin_consolidado:")]
        self.assertIn("def _total_do_mes_d(", trecho)
        self.assertIn("_total_do_mes_d(canal_da_linha, nome_limpo)", trecho,
                      "a linha de movimento tem de somar o mes")
        self.assertIn("_fluxo_do_mes_por_canal_d", trecho, "a linha de canal tambem")
        # A base e a COMPLETA, e o calculo e UM agrupamento -- nao um filtro
        # da base inteira por linha da tabela, que foi como fiz primeiro.
        self.assertIn("_mes_cheio_d = df_d_completo[", trecho)
        self.assertIn("_soma_por_canal_e_movimento_d = _mes_cheio_d.groupby(", trecho)
        j = trecho.index("def _total_do_mes_d(")
        corpo = trecho[j:j + 900]
        self.assertNotIn("df_d_completo[", corpo,
                         "filtrar a base dentro da funcao volta a varrer tudo por linha")
        self.assertIn("TOTAL DO MÊS / ÚLT. POSIÇÃO", trecho,
                      "o titulo precisa dizer que e do mes, senao engana")

    def test_saldo_continua_sendo_posicao_e_nao_soma(self):
        """Saldo de caixa e banco e POSICAO num dia: somar dias diferentes
        nao faz sentido, entao essas linhas seguem mostrando a ultima."""
        i = FONTE.index("with tab_fin_diario:")
        trecho = FONTE[i:FONTE.index("with tab_fin_consolidado:")]
        self.assertIn('if _classificar_movimento_fin(nome_limpo) in ("saldo", "aplicacao"):',
                      trecho)
        posicao = trecho.index('in ("saldo", "aplicacao"):')
        self.assertIn("nao_zerados.iloc[-1]", trecho[posicao:posicao + 400],
                      "saldo tem de continuar pegando a ultima posicao")

    def test_as_duas_abas_diarias_abrem_no_mesmo_dia(self):
        """Ontem, nas duas. O movimento do dia corrente costuma estar
        incompleto; e, se as abas abrissem em dias diferentes, a mesma
        pergunta daria respostas diferentes conforme a aba."""
        for ancora, variavel in [("with tab_fin_diario:", "data_ontem_d"),
                                 ("with tab_fin_consolidado:", "_ontem_dc")]:
            i = FONTE.index(ancora)
            trecho = FONTE[i:i + 2600]
            self.assertIn(f"{variavel} = ", trecho, ancora)
            self.assertIn("Timedelta(days=1)", trecho, f"{ancora}: nao comeca em ontem")
            self.assertIn("MonthEnd(1)", trecho, f"{ancora}: nao vai ate o fim do mes seguinte")
            # Nao basta calcular "ontem": o seletor tem de RECEBER esse
            # valor. Calcular e nao usar passaria despercebido.
            self.assertRegex(trecho, rf"padrao_ini=[^,]*{re.escape(variavel)}",
                             f"{ancora}: o padrao nao usa o dia de ontem")

    def test_periodo_por_datas_nas_duas_abas_diarias(self):
        """As duas abas diarias usam o mesmo seletor de duas datas, que pode
        atravessar meses -- julho e agosto na mesma tela."""
        self.assertIn("def seletor_periodo_dias(", FONTE)
        for ancora in ("with tab_fin_diario:", "with tab_fin_consolidado:"):
            i = FONTE.index(ancora)
            trecho = FONTE[i:i + 3000]
            self.assertIn("seletor_periodo_dias(", trecho, ancora)
        i = FONTE.index("def seletor_periodo_dias(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("min_value=limite_ini", corpo)
        self.assertIn("data_ini, data_fim = data_fim, data_ini", corpo,
                      "datas invertidas tem de ser ordenadas, nao virar erro")

    def test_aba_existe_com_as_duas_formas_de_abrir(self):
        self.assertIn("📆 Diário Consolidado", FONTE)
        i = FONTE.index("with tab_fin_consolidado:")
        trecho = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertIn("COL_FIN_GRUPO_DESPESA", trecho, "contas a pagar abre por grupo")
        self.assertIn("COL_FIN_MODALIDADE", trecho, "contas a receber abre por modalidade")
        self.assertIn("COL_FIN_CANAL", trecho, "falta a opcao de abrir por canal")
        self.assertIn('"Detalhe da linha", "Canal"', trecho)
        # A meta nao pode entrar no total desta aba tambem.
        self.assertIn('rotulo != MOV_RECEBER_META', trecho)


# ============================================================================
# 5i. MEMORIA E LEITURA DAS PLANILHAS
# ============================================================================
class TesteMemoriaELeitura(unittest.TestCase):
    """O servidor derruba o app quando a memoria estoura, e a tela e um erro
    generico sem causa. Estas travas protegem as duas frentes: ler menos e
    soltar memoria antes do limite."""

    def test_abas_sao_lidas_com_a_planilha_aberta_uma_vez(self):
        """Cada `pd.read_excel(caminho, sheet_name=...)` reprocessa o arquivo
        INTEIRO. Carregar 13 lojas dos dois arquivos eram 26 leituras
        completas -- 26 picos de memoria."""
        for funcao in ("carregar_dados_abas", "carregar_dados_por_loja"):
            i = FONTE.index(f"def {funcao}(")
            trecho = FONTE[i:FONTE.index("\ndef ", i + 10)]
            self.assertIn("_planilha_aberta(path_o)", trecho, funcao)
            self.assertIn("_planilha_aberta(path_r)", trecho, funcao)
            self.assertNotIn("_ler_aba_ou_vazio(path_o", trecho,
                             f"{funcao} voltou a reabrir o arquivo por aba")

    def test_caches_pesados_tem_teto_de_entradas(self):
        """Sem max_entries, cada combinacao de abas guarda uma copia inteira
        dos dados e elas se acumulam ate o app cair."""
        import re as _re
        for funcao in ("carregar_dados_abas", "carregar_dados_por_loja", "carregar_diario",
                       "preparar_fluxo_caixa"):
            i = FONTE.index(f"def {funcao}(")
            decorador = FONTE[max(0, i - 400):i]
            achado = _re.search(r"@st\.cache_\w+\([^)]*\)\s*$", decorador)
            self.assertIsNotNone(achado, f"{funcao} sem decorador de cache")
            self.assertIn("max_entries", achado.group(0), f"{funcao} sem teto de entradas")

    def test_csv_do_fluxo_e_lido_em_formato_economico(self):
        """O servidor corta o app perto de 1 GB. As colunas de texto que se
        repetem ao longo das ~650 mil linhas entram como CATEGORIA: cada
        valor distinto e guardado uma vez. Medido em 19/08/2026: 57% menos
        memoria por copia do fluxo."""
        ns = carregar([], ["TIPOS_ECONOMICOS_FLUXO"])
        for coluna in ("Movimento", "Canal.1", "Modalidade", "GRUPO DESPESA"):
            self.assertEqual(ns["TIPOS_ECONOMICOS_FLUXO"].get(coluna), "category", coluna)
        self.assertIn("dtype=TIPOS_ECONOMICOS_FLUXO", FONTE)
        self.assertIn("low_memory=False", FONTE)

    def test_colunas_sem_uso_sao_descartadas(self):
        i = FONTE.index("_descartar = [c for c in df_fluxo.columns")
        self.assertIn("df_fluxo.drop(columns=_descartar)", FONTE[i:i + 300])

    def test_nao_converte_coluna_inteira_para_texto(self):
        """`astype(str)` numa coluna categoria desfaz a economia: cria uma
        copia inteira em texto. Dentro de um laco por canal isso acontece
        dezenas de vezes por tela."""
        self.assertNotIn(".dropna().astype(str).unique()", FONTE,
                         "converte a coluna toda antes do unique() -- inverta a ordem")
        i = FONTE.index("with tab_fin_diario:")
        bloco = FONTE[i:FONTE.index("# ---------------- TESOURARIA", i)]
        self.assertNotIn("].astype(str) ==", bloco,
                         "comparacao com categoria nao precisa virar texto")
        self.assertNotIn("].astype(str).isin(", bloco)

    def test_plano_de_contas_continua_fora(self):
        """Nenhuma tela do fluxo usa o Plano de Contas, e ele e um texto
        diferente por lancamento, 650 mil vezes."""
        i = FONTE.index("_uteis = {")
        bloco = FONTE[i:FONTE.index("}", i)]
        self.assertNotIn("COL_FIN_PLANO_CONTAS", bloco,
                         "voltou a ser carregada sem nenhuma tela usar")
        self.assertLessEqual(FONTE.count("COL_FIN_PLANO_CONTAS"), 1,
                             "passou a ser usada -- precisa voltar para _uteis")

    def test_historico_e_numero_so_existem_para_o_detalhamento(self):
        """Elas custam memoria (cerca de 39 MB por copia) e so se justificam
        por causa do detalhamento por celula, onde respondem "de quem e este
        valor". Se o uso sumir, precisam sair de novo."""
        i = FONTE.index("_uteis = {")
        bloco = FONTE[i:FONTE.index("}", i)]
        for coluna in ("COL_FIN_HISTORICO", "COL_FIN_NUMERO"):
            self.assertIn(coluna, bloco, f"{coluna} precisa ser carregada para o detalhamento")
        j = FONTE.index("def montar_lancamentos_por_celula(")
        corpo = FONTE[j:FONTE.index("\ndef ", j + 10)]
        self.assertIn("COL_FIN_NUMERO", corpo, "o Numero precisa aparecer no detalhamento")
        self.assertIn("COL_FIN_HISTORICO", corpo)
        self.assertIn('"Fornecedor / Histórico"', corpo)

    def test_historico_entra_como_categoria_e_numero_nao(self):
        """O mesmo fornecedor aparece em centenas de lancamentos (42 MB viram
        1,4 MB como categoria). O Numero e unico por lancamento, e ai a
        categoria custa MAIS que o texto puro."""
        ns = carregar([], ["TIPOS_ECONOMICOS_FLUXO"])
        self.assertEqual(ns["TIPOS_ECONOMICOS_FLUXO"].get("Histórico"), "category")
        self.assertNotIn("Número", ns["TIPOS_ECONOMICOS_FLUXO"])

    def test_diario_guarda_as_repetitivas_como_categoria(self):
        i = FONTE.index("def carregar_diario(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn('astype("category")', corpo)
        for coluna in ("Plano de Contas", "Centro de Custos", "Linha DRE"):
            self.assertIn(f'"{coluna}"', corpo)

    def test_caches_pesados_guardam_poucas_copias(self):
        """Cada entrada guarda uma copia inteira dos dados de 13 lojas. Oito
        entradas eram oito copias."""
        import re as _re
        for funcao in ("carregar_dados_abas", "carregar_dados_por_loja"):
            i = FONTE.index(f"def {funcao}(")
            decorador = FONTE[max(0, i - 400):i]
            achado = _re.search(r"max_entries=(\d+)", decorador)
            self.assertIsNotNone(achado, f"{funcao} sem teto")
            self.assertLessEqual(int(achado.group(1)), 3, f"{funcao} guarda copias demais")

    def test_guarda_de_memoria_limpa_antes_do_corte(self):
        """850 MB era tarde: o servidor corta perto de 1 GB, manda e-mail e
        bloqueia o acesso. Reler uma planilha custa segundos."""
        ns = carregar([], ["LIMITE_MEMORIA_ALERTA_MB", "LIMITE_MEMORIA_LIMPEZA_MB"])
        self.assertLessEqual(ns["LIMITE_MEMORIA_LIMPEZA_MB"], 700)
        self.assertLess(ns["LIMITE_MEMORIA_ALERTA_MB"], ns["LIMITE_MEMORIA_LIMPEZA_MB"])

    def test_guarda_de_memoria_existe_e_tem_limite(self):
        ns = carregar(["memoria_em_uso_mb"],
                      ["LIMITE_MEMORIA_ALERTA_MB", "LIMITE_MEMORIA_LIMPEZA_MB"])
        self.assertLess(ns["LIMITE_MEMORIA_ALERTA_MB"], ns["LIMITE_MEMORIA_LIMPEZA_MB"],
                        "o alerta tem de vir antes da limpeza")
        # A medicao nao pode derrubar nada onde nao houver /proc.
        valor = ns["memoria_em_uso_mb"]()
        self.assertTrue(valor is None or valor > 0)

    def test_guarda_so_limpa_acima_do_limite(self):
        ns = carregar(["guardar_memoria"], ["LIMITE_MEMORIA_LIMPEZA_MB"])
        limpezas = []
        ns["st"] = type("st", (), {"cache_data": type("c", (), {
            "clear": staticmethod(lambda: limpezas.append(1))})()})()
        ns["gc"] = type("gc", (), {"collect": staticmethod(lambda: None)})()

        ns["memoria_em_uso_mb"] = lambda: 100.0
        mb, limpou = ns["guardar_memoria"]()
        self.assertFalse(limpou, "limpou com a memoria baixa")

        ns["memoria_em_uso_mb"] = lambda: ns["LIMITE_MEMORIA_LIMPEZA_MB"] + 50
        mb, limpou = ns["guardar_memoria"]()
        self.assertTrue(limpou, "nao limpou mesmo acima do limite")
        self.assertEqual(len(limpezas), 1)

        # Sem medicao possivel, nao age -- e nao quebra.
        ns["memoria_em_uso_mb"] = lambda: None
        self.assertEqual(ns["guardar_memoria"](), (None, False))


# ============================================================================
# 5o. CHECKLIST DE FECHAMENTO MENSAL (lido da Planilha Google)
# ============================================================================
class TesteEnderecoDaPlanilhaDeFechamento(unittest.TestCase):
    """O que a pessoa cola e o que o painel precisa buscar sao coisas
    diferentes. Aqui entra o link do botao Compartilhar; tem de sair o
    endereco que devolve CSV."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["url_csv_do_fechamento"])

    def _url(self, texto):
        return self.ns["url_csv_do_fechamento"](texto)

    def test_link_de_compartilhamento_vira_endereco_de_csv(self):
        ident = "1Qfg95yYd-6J55drs5p4lMgGF6SVAV6vH"
        for link in (
            f"https://docs.google.com/spreadsheets/d/{ident}/edit?usp=sharing",
            f"https://docs.google.com/spreadsheets/d/{ident}/edit#gid=0",
            f"https://docs.google.com/spreadsheets/d/{ident}",
        ):
            self.assertEqual(
                self._url(link),
                f"https://docs.google.com/spreadsheets/d/{ident}/export?format=csv",
                f"nao reconheceu {link}",
            )

    def test_endereco_sai_sem_gid(self):
        """Sem gid de proposito: sem ele o Google devolve a PRIMEIRA aba, e a
        aba de dados e a primeira da planilha. No painel financeiro o gid mudou
        sozinho tres vezes num dia e derrubou a leitura toda vez."""
        url = self._url("https://docs.google.com/spreadsheets/d/1AbC_def-123456789012/edit#gid=847362")
        self.assertNotIn("gid", url)

    def test_endereco_de_csv_pronto_passa_direto(self):
        """Quem ja tem o link de publicacao na web nao pode ter o endereco
        reescrito -- ele nao tem /spreadsheets/d/ e seria perdido."""
        publicado = ("https://docs.google.com/spreadsheets/d/e/2PACX-1vQB1ygqIm/pub?output=csv")
        self.assertEqual(self._url(publicado), publicado)

    def test_id_solto_tambem_serve(self):
        ident = "1Qfg95yYd-6J55drs5p4lMgGF6SVAV6vH"
        self.assertEqual(
            self._url(ident),
            f"https://docs.google.com/spreadsheets/d/{ident}/export?format=csv")

    def test_texto_que_nao_e_link_devolve_nada(self):
        """Melhor a tela pedir o link do que sair buscando um endereco
        inventado e mostrar erro de rede."""
        for lixo in ("", "   ", None, "planilha do fechamento", "abc123"):
            self.assertIsNone(self._url(lixo), f"aceitou {lixo!r}")


class TesteLeituraDoChecklist(unittest.TestCase):
    """Do CSV cru da planilha ate a tabela que a tela mostra."""

    CABECALHO = "Ano,Mês,Competência,Processo,Status,Responsável,Concluído em,Observação"

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["preparar_checklist_da_planilha", "checklist_da_competencia",
             "normalizar_status_fechamento", "_texto_ou_vazio",
             "resumo_do_fechamento", "pendentes_do_fechamento",
             "competencias_da_planilha", "competencia_padrao",
             "rotulo_competencia", "resolver_colunas_fluxo",
             "_assinatura_coluna_fin", "_normalizar_coluna_fin", "_normalizar_texto"],
            ["PROCESSOS_FECHAMENTO", "COL_FECH_ANO", "COL_FECH_MES",
             "COL_FECH_PROCESSO", "COL_FECH_STATUS", "COL_FECH_QUEM",
             "COL_FECH_QUANDO", "COL_FECH_OBS", "COLUNAS_FECHAMENTO",
             "COLUNAS_VISIVEIS_FECHAMENTO", "STATUS_OK", "STATUS_PENDENTE",
             "STATUS_EM_ANDAMENTO", "STATUS_NAO_SE_APLICA", "STATUS_FECHAMENTO",
             "LIGACOES_NOME_COLUNA", "_ACENTOS_FIN"],
        )

    def _csv(self, linhas, cabecalho=None):
        texto = (cabecalho or self.CABECALHO) + "\n" + "\n".join(linhas)
        return pd.read_csv(io.StringIO(texto), dtype=str,
                           keep_default_na=False, na_values=[""])

    def _preparar(self, linhas, cabecalho=None):
        return self.ns["preparar_checklist_da_planilha"](self._csv(linhas, cabecalho))

    def _linhas_de_um_mes(self, ano, mes, statuses):
        return [
            f'{ano},{mes},{mes:02d}/{ano},"{nome}",{status},,,'
            for nome, status in zip(self.ns["PROCESSOS_FECHAMENTO"], statuses)
        ]

    # -- leitura --------------------------------------------------------
    def test_le_a_planilha_e_devolve_ano_e_mes_como_numero(self):
        """Ano e Mes sao numeros separados de proposito, nunca uma data: o CSV
        do Google entrega data em mes/dia/ano e isso ja trocou dia por mes em
        187 mil lancamentos do fluxo. Numero nao tem idioma."""
        limpo, faltando = self._preparar(
            self._linhas_de_um_mes(2026, 7, ["OK"] * 8))
        self.assertEqual(faltando, [])
        self.assertEqual(len(limpo), 8)
        self.assertEqual(limpo[self.ns["COL_FECH_ANO"]].iloc[0], 2026)
        self.assertEqual(limpo[self.ns["COL_FECH_MES"]].iloc[0], 7)
        self.assertTrue(pd.api.types.is_integer_dtype(limpo[self.ns["COL_FECH_MES"]]))

    def test_coluna_com_escrita_diferente_ainda_e_encontrada(self):
        """Mesmo resolvedor do fluxo: acento, caixa e espaco sobrando nao
        podem derrubar a leitura. Quem edita a planilha e gente.

        O que NAO e tolerado, de proposito: trocar a palavra. "Observacoes"
        no plural nao casa com "Observacao", e a tela aponta a coluna pelo
        nome -- melhor pedir o nome certo do que adivinhar qual coluna a
        pessoa quis dizer e ler a errada em silencio."""
        cabecalho = "ANO,mes,Competencia,PROCESSO, status ,Responsavel,Concluido em,Observacao"
        limpo, faltando = self._preparar(
            self._linhas_de_um_mes(2026, 7, ["OK"] * 8), cabecalho=cabecalho)
        self.assertEqual(faltando, [], "a leitura ficou presa na escrita exata")
        self.assertEqual(len(limpo), 8)

    def test_coluna_que_falta_de_verdade_e_apontada_pelo_nome(self):
        """Erro generico manda a pessoa adivinhar. O nome da coluna que falta
        resolve em segundos."""
        cabecalho = "Ano,Mês,Competência,Processo,Responsável,Concluído em,Observação"
        linhas = [f'2026,7,07/2026,"{n}",,,' for n in self.ns["PROCESSOS_FECHAMENTO"]]
        limpo, faltando = self._preparar(linhas, cabecalho=cabecalho)
        self.assertIn(self.ns["COL_FECH_STATUS"], faltando)
        self.assertTrue(limpo.empty)

    def test_linha_sem_ano_ou_mes_e_descartada(self):
        """Linha em branco no fim da planilha e o caso mais comum -- nao pode
        virar uma competencia fantasma no seletor."""
        linhas = self._linhas_de_um_mes(2026, 7, ["OK"] * 8)
        linhas += [",,,,,,,", "ano,mes,,Contas a Pagar,OK,,,"]
        limpo, _ = self._preparar(linhas)
        self.assertEqual(len(limpo), 8)

    def test_celula_vazia_nao_vira_o_texto_nan(self):
        """Em 20/08 um ausente convertido para texto virou 'nan' e zerou o
        mapa das tabelas do fluxo. O tratamento vem ANTES da conversao."""
        limpo, _ = self._preparar(self._linhas_de_um_mes(2026, 7, ["OK"] * 8))
        for coluna in (self.ns["COL_FECH_QUEM"], self.ns["COL_FECH_QUANDO"],
                       self.ns["COL_FECH_OBS"]):
            self.assertEqual(set(limpo[coluna]), {""}, f"{coluna} veio suja")

    def test_planilha_vazia_nao_quebra(self):
        limpo, faltando = self.ns["preparar_checklist_da_planilha"](None)
        self.assertTrue(limpo.empty)
        self.assertEqual(faltando, [])

    # -- status ---------------------------------------------------------
    def test_status_aceita_variacao_de_quem_digita(self):
        normalizar = self.ns["normalizar_status_fechamento"]
        self.assertEqual(normalizar("ok"), self.ns["STATUS_OK"])
        self.assertEqual(normalizar(" OK "), self.ns["STATUS_OK"])
        self.assertEqual(normalizar("Concluído"), self.ns["STATUS_OK"])
        self.assertEqual(normalizar("N/A"), self.ns["STATUS_NAO_SE_APLICA"])
        self.assertEqual(normalizar("n.a."), self.ns["STATUS_NAO_SE_APLICA"])
        self.assertEqual(normalizar("não se aplica"), self.ns["STATUS_NAO_SE_APLICA"])
        self.assertEqual(normalizar("Em Andamento"), self.ns["STATUS_EM_ANDAMENTO"])

    def test_status_desconhecido_vira_pendente_e_nao_categoria_nova(self):
        """Devolver o texto original faria a contagem ter tantas categorias
        quanto erros de digitacao -- e 'okk' contaria como concluido em nenhum
        lugar e apareceria na tela em todos."""
        normalizar = self.ns["normalizar_status_fechamento"]
        for esquisito in ("okk", "?", "talvez", "", None, float("nan")):
            self.assertEqual(normalizar(esquisito), self.ns["STATUS_PENDENTE"],
                             f"{esquisito!r} escapou")

    # -- recorte por competencia ----------------------------------------
    def test_processo_que_falta_no_mes_entra_pendente(self):
        """A planilha pode nao ter todos os processos num mes (linha apagada,
        mes recem-criado). A tela mostra os oito, e o que falta e pendente --
        nunca some da lista."""
        limpo, _ = self._preparar(self._linhas_de_um_mes(2026, 7, ["OK"] * 8)[:3])
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        self.assertEqual(list(tabela[self.ns["COL_FECH_PROCESSO"]]),
                         self.ns["PROCESSOS_FECHAMENTO"])
        self.assertEqual(tabela[self.ns["COL_FECH_STATUS"]].iloc[0], self.ns["STATUS_OK"])
        self.assertEqual(tabela[self.ns["COL_FECH_STATUS"]].iloc[5],
                         self.ns["STATUS_PENDENTE"])

    def test_processo_a_mais_na_planilha_e_ignorado(self):
        """Se a planilha ganhar um processo antes do painel, a tela nao pode
        quebrar -- ele so nao aparece ate a lista daqui ser atualizada."""
        linhas = self._linhas_de_um_mes(2026, 7, ["OK"] * 8)
        linhas.append("2026,7,07/2026,Processo Novo Que O Painel Nao Conhece,OK,,,")
        limpo, _ = self._preparar(linhas)
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        self.assertEqual(list(tabela[self.ns["COL_FECH_PROCESSO"]]),
                         self.ns["PROCESSOS_FECHAMENTO"])

    def test_ordem_e_a_oficial_nao_a_da_planilha(self):
        """A ordem e a que a area segue no fechamento. Reordenar a planilha
        por engano nao pode reordenar a tela."""
        linhas = self._linhas_de_um_mes(2026, 7, ["OK"] * 8)[::-1]
        limpo, _ = self._preparar(linhas)
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        self.assertEqual(list(tabela[self.ns["COL_FECH_PROCESSO"]]),
                         self.ns["PROCESSOS_FECHAMENTO"])

    def test_um_mes_nao_contamina_o_outro(self):
        """Erro classico de filtro: julho marcado aparecendo em agosto."""
        linhas = (self._linhas_de_um_mes(2026, 7, ["OK"] * 8)
                  + self._linhas_de_um_mes(2026, 8, ["Pendente"] * 8))
        limpo, _ = self._preparar(linhas)
        julho = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        agosto = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-08", "M"))
        self.assertTrue((julho[self.ns["COL_FECH_STATUS"]] == self.ns["STATUS_OK"]).all())
        self.assertTrue((agosto[self.ns["COL_FECH_STATUS"]] == self.ns["STATUS_PENDENTE"]).all())

    def test_mesmo_mes_e_ano_de_anos_diferentes_nao_se_misturam(self):
        """Julho de 2026 e julho de 2027 sao meses diferentes -- filtrar so
        pelo mes juntaria os dois."""
        linhas = (self._linhas_de_um_mes(2026, 7, ["OK"] * 8)
                  + self._linhas_de_um_mes(2027, 7, ["Pendente"] * 8))
        limpo, _ = self._preparar(linhas)
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2027-07", "M"))
        self.assertTrue((tabela[self.ns["COL_FECH_STATUS"]] == self.ns["STATUS_PENDENTE"]).all())

    def test_linha_repetida_vale_a_de_baixo(self):
        """Se a mesma linha aparecer duas vezes, vale a ultima: e a que a
        pessoa acabou de preencher."""
        linhas = self._linhas_de_um_mes(2026, 7, ["Pendente"] * 8)
        linhas.append("2026,7,07/2026,Contas a Pagar,OK,,,")
        limpo, _ = self._preparar(linhas)
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        self.assertEqual(tabela[self.ns["COL_FECH_STATUS"]].iloc[0], self.ns["STATUS_OK"])

    def test_observacao_e_responsavel_chegam_na_tela(self):
        linhas = ["2026,7,07/2026,Contas a Pagar,OK,Richards,05/08/2026,Lancado no dia 5"]
        limpo, _ = self._preparar(linhas)
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        self.assertEqual(tabela[self.ns["COL_FECH_QUEM"]].iloc[0], "Richards")
        self.assertEqual(tabela[self.ns["COL_FECH_QUANDO"]].iloc[0], "05/08/2026")
        self.assertEqual(tabela[self.ns["COL_FECH_OBS"]].iloc[0], "Lancado no dia 5")

    # -- contagem -------------------------------------------------------
    def test_na_sai_da_conta_dos_dois_lados(self):
        """Processo que nao se aplica ao mes nao e trabalho feito nem
        pendencia. Conta-lo como feito inflaria o percentual; como pendente
        prenderia o fechamento em algo que nao existe."""
        statuses = ["OK", "OK", "N/A", "Pendente", "Pendente", "Pendente", "Pendente", "Pendente"]
        limpo, _ = self._preparar(self._linhas_de_um_mes(2026, 7, statuses))
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        concluidos, aplicaveis, pct = self.ns["resumo_do_fechamento"](tabela)
        self.assertEqual(concluidos, 2)
        self.assertEqual(aplicaveis, 7, "o N/A tem de sair do denominador")
        self.assertAlmostEqual(pct, 2 / 7 * 100, places=6)
        faltando = self.ns["pendentes_do_fechamento"](tabela)
        self.assertEqual(len(faltando), 5)
        self.assertNotIn(self.ns["PROCESSOS_FECHAMENTO"][2], faltando)

    def test_em_andamento_nao_conta_como_concluido(self):
        statuses = ["Em andamento"] * 8
        limpo, _ = self._preparar(self._linhas_de_um_mes(2026, 7, statuses))
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        concluidos, aplicaveis, pct = self.ns["resumo_do_fechamento"](tabela)
        self.assertEqual((concluidos, aplicaveis), (0, 8))
        self.assertEqual(pct, 0.0)
        self.assertEqual(len(self.ns["pendentes_do_fechamento"](tabela)), 8)

    def test_mes_inteiro_em_na_nao_divide_por_zero(self):
        limpo, _ = self._preparar(self._linhas_de_um_mes(2026, 7, ["N/A"] * 8))
        tabela = self.ns["checklist_da_competencia"](limpo, pd.Period("2026-07", "M"))
        concluidos, aplicaveis, pct = self.ns["resumo_do_fechamento"](tabela)
        self.assertEqual((concluidos, aplicaveis), (0, 0))
        self.assertEqual(pct, 100.0)
        self.assertEqual(self.ns["pendentes_do_fechamento"](tabela), [])

    # -- seletor de competencia -----------------------------------------
    def test_competencias_saem_da_planilha_da_mais_nova_para_tras(self):
        linhas = (self._linhas_de_um_mes(2026, 7, ["OK"] * 8)
                  + self._linhas_de_um_mes(2026, 9, ["OK"] * 8)
                  + self._linhas_de_um_mes(2026, 8, ["OK"] * 8))
        limpo, _ = self._preparar(linhas)
        comps = self.ns["competencias_da_planilha"](limpo)
        self.assertEqual(comps, [pd.Period("2026-09", "M"), pd.Period("2026-08", "M"),
                                 pd.Period("2026-07", "M")])
        self.assertEqual(self.ns["rotulo_competencia"](comps[-1]), "07/2026")

    def test_tela_abre_no_mes_anterior(self):
        """Fechamento se faz do mes que acabou: abrir no corrente faria a tela
        aparecer zerada todo dia 1o."""
        comps = [pd.Period(f"2026-{m:02d}", "M") for m in range(12, 0, -1)]
        self.assertEqual(self.ns["competencia_padrao"](comps, date(2026, 8, 25)),
                         pd.Period("2026-07", "M"))

    def test_mes_anterior_ausente_cai_no_disponivel_mais_proximo(self):
        """Planilha que so tem ate maio nao pode abrir vazia em agosto."""
        comps = [pd.Period("2026-05", "M"), pd.Period("2026-04", "M")]
        self.assertEqual(self.ns["competencia_padrao"](comps, date(2026, 8, 25)),
                         pd.Period("2026-05", "M"))

    def test_so_ha_meses_futuros_cai_no_mais_recente(self):
        comps = [pd.Period("2027-03", "M"), pd.Period("2027-02", "M")]
        self.assertEqual(self.ns["competencia_padrao"](comps, date(2026, 8, 25)),
                         pd.Period("2027-03", "M"))

    def test_planilha_sem_competencia_nenhuma_devolve_nada(self):
        self.assertIsNone(self.ns["competencia_padrao"]([], date(2026, 8, 25)))

    def test_virada_de_ano_anda_para_tras(self):
        comps = [pd.Period("2026-01", "M"), pd.Period("2025-12", "M")]
        self.assertEqual(self.ns["competencia_padrao"](comps, date(2026, 1, 3)),
                         pd.Period("2025-12", "M"))


class TestePrazoEPanorama(unittest.TestCase):
    """As duas informacoes que a tela de um mes nao dava: ha quanto tempo o
    fechamento esta aberto, e qual processo atrasa sempre."""

    CABECALHO = "Ano,Mês,Competência,Processo,Status,Responsável,Concluído em,Observação"

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["preparar_checklist_da_planilha", "checklist_da_competencia",
             "normalizar_status_fechamento", "_texto_ou_vazio", "resumo_do_fechamento",
             "pendentes_do_fechamento", "prazo_do_fechamento", "_data_do_texto_br",
             "panorama_do_ano", "anos_da_planilha", "url_para_editar_planilha",
             "resolver_colunas_fluxo", "_assinatura_coluna_fin", "_normalizar_coluna_fin"],
            ["PROCESSOS_FECHAMENTO", "COL_FECH_ANO", "COL_FECH_MES", "COL_FECH_PROCESSO",
             "COL_FECH_STATUS", "COL_FECH_QUEM", "COL_FECH_QUANDO", "COL_FECH_OBS",
             "COLUNAS_FECHAMENTO", "COLUNAS_VISIVEIS_FECHAMENTO", "STATUS_OK",
             "STATUS_PENDENTE", "STATUS_EM_ANDAMENTO", "STATUS_NAO_SE_APLICA",
             "SIMBOLOS_FECHAMENTO", "LINHA_PANORAMA_PCT", "LIGACOES_NOME_COLUNA",
             "_ACENTOS_FIN"],
        )

    def _tabela(self, ano, mes, statuses, datas=None):
        datas = datas or [""] * len(statuses)
        linhas = [self.CABECALHO] + [
            f'{ano},{mes},{mes:02d}/{ano},"{nome}",{st},,{quando},'
            for nome, st, quando in zip(self.ns["PROCESSOS_FECHAMENTO"], statuses, datas)
        ]
        bruto = pd.read_csv(io.StringIO("\n".join(linhas)), dtype=str,
                            keep_default_na=False, na_values=[""])
        limpo, _ = self.ns["preparar_checklist_da_planilha"](bruto)
        return limpo, self.ns["checklist_da_competencia"](
            limpo, pd.Period(f"{ano}-{mes:02d}", "M"))

    # -- data ------------------------------------------------------------
    def test_data_so_aceita_dia_mes_ano(self):
        """Sem `dayfirst` nem inferencia: a coluna e digitada a mao e exportada
        por um Google que pode estar em local dos Estados Unidos. Um parser
        esperto leria 05/08 como 8 de maio e ninguem perceberia."""
        ler = self.ns["_data_do_texto_br"]
        self.assertEqual(ler("05/08/2026"), date(2026, 8, 5))
        self.assertEqual(ler("5/8/2026"), date(2026, 8, 5))
        for lixo in ("", None, "2026-08-05", "08/05/26", "amanha", "32/08/2026", "05/13/2026"):
            self.assertIsNone(ler(lixo), f"aceitou {lixo!r}")

    # -- prazo -----------------------------------------------------------
    def test_mes_em_curso_nao_tem_prazo(self):
        """Contar dias de atraso de um mes que ainda nao acabou seria acusar
        atraso do que nem venceu."""
        _, tabela = self._tabela(2026, 8, ["Pendente"] * 8)
        estado, dias, rotulo = self.ns["prazo_do_fechamento"](
            pd.Period("2026-08", "M"), tabela, date(2026, 8, 25))
        self.assertEqual(estado, "em_curso")
        self.assertIsNone(dias)
        self.assertIn("curso", rotulo)

    def test_mes_fechado_conta_do_fim_do_mes_ate_a_ultima_conclusao(self):
        datas = ["05/08/2026"] * 7 + ["07/08/2026"]
        _, tabela = self._tabela(2026, 7, ["OK"] * 8, datas)
        estado, dias, rotulo = self.ns["prazo_do_fechamento"](
            pd.Period("2026-07", "M"), tabela, date(2026, 8, 25))
        self.assertEqual(estado, "concluido")
        self.assertEqual(dias, 7, "31/07 -> 07/08 sao 7 dias")
        self.assertIn("7 dias", rotulo)

    def test_mes_em_aberto_conta_do_fim_do_mes_ate_hoje(self):
        """O numero que interessa numa Controladoria: ha quanto tempo este
        fechamento esta em aberto."""
        _, tabela = self._tabela(2026, 7, ["OK"] * 6 + ["Pendente"] * 2)
        estado, dias, rotulo = self.ns["prazo_do_fechamento"](
            pd.Period("2026-07", "M"), tabela, date(2026, 8, 25))
        self.assertEqual(estado, "aberto")
        self.assertEqual(dias, 25, "31/07 -> 25/08 sao 25 dias")
        self.assertIn("25 dias", rotulo)

    def test_mes_so_com_na_conta_como_concluido(self):
        """N/A nao e pendencia -- um mes inteiro sem aplicacao esta fechado."""
        _, tabela = self._tabela(2026, 7, ["N/A"] * 8)
        estado, _, _ = self.ns["prazo_do_fechamento"](
            pd.Period("2026-07", "M"), tabela, date(2026, 8, 25))
        self.assertEqual(estado, "concluido")

    def test_concluido_sem_data_preenchida_nao_inventa_prazo(self):
        """Marcar OK sem preencher a data e comum. O prazo fica sem numero em
        vez de virar zero, que leria como "fechou no mesmo dia"."""
        _, tabela = self._tabela(2026, 7, ["OK"] * 8)
        estado, dias, _ = self.ns["prazo_do_fechamento"](
            pd.Period("2026-07", "M"), tabela, date(2026, 8, 25))
        self.assertEqual(estado, "concluido")
        self.assertIsNone(dias)

    def test_um_dia_fica_no_singular(self):
        _, tabela = self._tabela(2026, 7, ["OK"] * 6 + ["Pendente"] * 2)
        _, _, rotulo = self.ns["prazo_do_fechamento"](
            pd.Period("2026-07", "M"), tabela, date(2026, 8, 1))
        self.assertIn("1 dia", rotulo)
        self.assertNotIn("1 dias", rotulo)

    # -- panorama --------------------------------------------------------
    def test_panorama_tem_um_processo_por_linha_e_um_mes_por_coluna(self):
        linhas = [self.CABECALHO]
        for mes, status in ((7, "OK"), (8, "Pendente")):
            linhas += [f'2026,{mes},{mes:02d}/2026,"{n}",{status},,,'
                       for n in self.ns["PROCESSOS_FECHAMENTO"]]
        bruto = pd.read_csv(io.StringIO("\n".join(linhas)), dtype=str,
                            keep_default_na=False, na_values=[""])
        limpo, _ = self.ns["preparar_checklist_da_planilha"](bruto)
        matriz = self.ns["panorama_do_ano"](limpo, 2026)
        self.assertEqual(list(matriz.columns), ["07", "08"])
        self.assertEqual(list(matriz.index)[:-1], self.ns["PROCESSOS_FECHAMENTO"])
        self.assertEqual(list(matriz.index)[-1], self.ns["LINHA_PANORAMA_PCT"])
        self.assertEqual(matriz.loc[self.ns["PROCESSOS_FECHAMENTO"][0], "07"],
                         self.ns["SIMBOLOS_FECHAMENTO"][self.ns["STATUS_OK"]])
        self.assertEqual(matriz.loc[self.ns["LINHA_PANORAMA_PCT"], "07"], "100%")
        self.assertEqual(matriz.loc[self.ns["LINHA_PANORAMA_PCT"], "08"], "0%")

    def test_panorama_mostra_o_processo_que_atrasa_sempre(self):
        """E para isto que a matriz existe: quatro meses seguidos de pendencia
        na mesma linha salta aos olhos aqui e e invisivel na tela de mes."""
        linhas = [self.CABECALHO]
        for mes in range(1, 5):
            for i, nome in enumerate(self.ns["PROCESSOS_FECHAMENTO"]):
                status = "Pendente" if i == 6 else "OK"
                linhas.append(f'2026,{mes},{mes:02d}/2026,"{nome}",{status},,,')
        bruto = pd.read_csv(io.StringIO("\n".join(linhas)), dtype=str,
                            keep_default_na=False, na_values=[""])
        limpo, _ = self.ns["preparar_checklist_da_planilha"](bruto)
        matriz = self.ns["panorama_do_ano"](limpo, 2026)
        atrasado = self.ns["PROCESSOS_FECHAMENTO"][6]
        pendente = self.ns["SIMBOLOS_FECHAMENTO"][self.ns["STATUS_PENDENTE"]]
        self.assertEqual(list(matriz.loc[atrasado]), [pendente] * 4)
        self.assertTrue(all(v == "88%" for v in matriz.loc[self.ns["LINHA_PANORAMA_PCT"]]))

    def test_panorama_de_mes_todo_na_nao_divide_por_zero(self):
        linhas = [self.CABECALHO] + [
            f'2026,7,07/2026,"{n}",N/A,,,' for n in self.ns["PROCESSOS_FECHAMENTO"]]
        bruto = pd.read_csv(io.StringIO("\n".join(linhas)), dtype=str,
                            keep_default_na=False, na_values=[""])
        limpo, _ = self.ns["preparar_checklist_da_planilha"](bruto)
        matriz = self.ns["panorama_do_ano"](limpo, 2026)
        self.assertEqual(matriz.loc[self.ns["LINHA_PANORAMA_PCT"], "07"], "—")

    def test_panorama_de_ano_inexistente_volta_vazio(self):
        limpo, _ = self._tabela(2026, 7, ["OK"] * 8)
        self.assertTrue(self.ns["panorama_do_ano"](limpo, 2030).empty)

    def test_anos_saem_do_mais_recente_para_tras(self):
        linhas = [self.CABECALHO]
        for ano in (2026, 2027):
            linhas += [f'{ano},7,07/{ano},"{n}",OK,,,'
                       for n in self.ns["PROCESSOS_FECHAMENTO"]]
        bruto = pd.read_csv(io.StringIO("\n".join(linhas)), dtype=str,
                            keep_default_na=False, na_values=[""])
        limpo, _ = self.ns["preparar_checklist_da_planilha"](bruto)
        self.assertEqual(self.ns["anos_da_planilha"](limpo), [2027, 2026])

    # -- botao de abrir --------------------------------------------------
    def test_link_de_edicao_sai_do_mesmo_valor_guardado(self):
        ident = "1p1T_2VWR4RY6lWq1_cn3omo7Iss_uX_l9D8j6_Sl8cA"
        alvo = f"https://docs.google.com/spreadsheets/d/{ident}/edit"
        self.assertEqual(
            self.ns["url_para_editar_planilha"](f"{alvo}?usp=sharing"), alvo)
        self.assertEqual(self.ns["url_para_editar_planilha"](ident), alvo)

    def test_link_publicado_nao_tem_edicao(self):
        """O /d/e/ e um documento publicado -- nao existe tela de edicao para
        ele, e um botao que leva a lugar nenhum e pior que botao nenhum."""
        self.assertIsNone(self.ns["url_para_editar_planilha"](
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vQB1ygqIm/pub?output=csv"))
        self.assertIsNone(self.ns["url_para_editar_planilha"](""))


class TesteTelaDoFechamento(unittest.TestCase):
    """Travas da aba: o que a tela tem de dizer quando algo nao esta no lugar."""

    def _bloco(self, tamanho=None):
        """O bloco INTEIRO da aba, recortado entre marcos do codigo.

        Antes era uma janela de N caracteres, e ela ficou pequena quando a aba
        cresceu: a legenda saiu da janela e o teste passou a falhar por
        posicao, nao por defeito. Marco nao encolhe."""
        i = FONTE.index("Fechamento Mensal — lançamentos e conferências")
        # Ate o proximo cabecalho de aba, seja qual for: cravar "# ABA 6:"
        # quebrou quando a aba de Orcamento 2027 entrou entre as duas.
        fim = min(p for p in (FONTE.find("# ABA: ORÇAMENTO 2027", i),
                              FONTE.find("# ABA 6: GESTÃO DE USUÁRIOS", i))
                  if p > 0)
        bloco = FONTE[i:fim]
        return bloco if tamanho is None else bloco[:tamanho]

    def test_sem_planilha_ligada_a_tela_explica_o_que_fazer(self):
        """Mostrar oito linhas pendentes sem planilha ligada seria mentira: a
        tela nao sabe se esta pendente ou se nao ha dado."""
        bloco = self._bloco()
        self.assertIn("if not _url_fech:", bloco)
        self.assertIn("FECHAMENTO_CSV_URL", bloco)
        self.assertIn("Qualquer pessoa com o link", bloco)

    def test_erro_de_leitura_chega_na_tela_com_o_motivo(self):
        """Licao de 20/08: erro guardado em variavel e nunca exibido custou
        muitos turnos de palpite."""
        bloco = self._bloco()
        self.assertIn("_dados_fech, _erro_fech = carregar_planilha_fechamento(", bloco)
        self.assertIn("st.error(_erro_fech)", bloco)

    def test_a_aba_nao_finge_que_grava(self):
        """A marcacao e na planilha. Um editor aqui daria a entender que o
        painel salva -- e ele nao tem onde."""
        bloco = self._bloco()
        self.assertNotIn("st.data_editor", bloco,
                         "a aba voltou a ter editor, mas o painel nao grava")
        self.assertIn("A marcação é feita **na planilha**", bloco)

    def test_a_leitura_da_planilha_tem_validade(self):
        """Cache sem ttl deixaria a tela mostrando o fechamento de ontem."""
        i = FONTE.index("def carregar_planilha_fechamento(")
        cabecalho = FONTE[max(0, i - 200):i]
        self.assertIn("@st.cache_data(ttl=", cabecalho)
        self.assertIn("max_entries=", cabecalho)


class TesteBuscaDeSegredo(unittest.TestCase):
    """A armadilha do TOML que custou uma tarde em 25/08/2026: o arquivo de
    Secrets e TOML, e toda chave escrita DEPOIS de um cabecalho [secao]
    pertence aquela secao. Colar a linha no fim do arquivo -- o lugar mais
    natural do mundo -- faz a chave sumir do topo, e st.secrets.get devolve
    vazio como se ela nunca tivesse sido escrita."""

    LINK = "https://docs.google.com/spreadsheets/d/1p1T_2VWR4RY6lWq1_cn3omo7Iss_uX_l9D8j6_Sl8cA/edit"

    class _SecretsFalso:
        """Dubla o objeto de Secrets do Streamlit: tabela que responde a
        keys(), `in` e indexacao, com sub-tabelas aninhadas."""

        def __init__(self, dados):
            self._dados = dados

        def keys(self):
            return self._dados.keys()

        def __contains__(self, chave):
            return chave in self._dados

        def __getitem__(self, chave):
            valor = self._dados[chave]
            if isinstance(valor, dict):
                return TesteBuscaDeSegredo._SecretsFalso(valor)
            if isinstance(valor, list):
                return [TesteBuscaDeSegredo._SecretsFalso(v) if isinstance(v, dict) else v
                        for v in valor]
            return valor

    def _buscar(self, dados, nome="FECHAMENTO_CSV_URL"):
        ns = carregar(["_segredo_com_origem", "_segredo"],
                      extras={"st": type("St", (), {"secrets": self._SecretsFalso(dados)})})
        return ns["_segredo_com_origem"](nome), ns["_segredo"](nome)

    def test_acha_no_topo(self):
        (valor, onde), simples = self._buscar(
            {"email": "a@b.c", "FECHAMENTO_CSV_URL": self.LINK})
        self.assertEqual(valor, self.LINK)
        self.assertEqual(onde, "topo")
        self.assertEqual(simples, self.LINK)

    def test_acha_dentro_de_secao(self):
        (valor, onde), _ = self._buscar(
            {"email": "a@b.c", "usuarios": {"FECHAMENTO_CSV_URL": self.LINK}})
        self.assertEqual(valor, self.LINK, "chave dentro de secao continua invisivel")
        self.assertIn("usuarios", onde, "a tela precisa dizer ONDE achou")

    def test_acha_dentro_de_secao_aninhada(self):
        (valor, onde), _ = self._buscar(
            {"usuarios": {"fulano": {"email": "y", "FECHAMENTO_CSV_URL": self.LINK}}})
        self.assertEqual(valor, self.LINK)
        self.assertIn("fulano", onde)

    def test_o_que_nao_existe_volta_vazio(self):
        (valor, onde), simples = self._buscar({"email": "a@b.c", "usuarios": {"x": {"email": "y"}}})
        self.assertEqual((valor, onde), ("", ""))
        self.assertEqual(simples, "")

    def test_acha_dentro_de_lista_de_tabelas(self):
        """[[users]] no TOML vira uma LISTA de tabelas, e a chave colada
        depois de um cabecalho desses cai dentro do ultimo item. Sem descer em
        lista, o segredo continuaria invisivel exatamente no formato que o
        arquivo de Secrets deles ja usa."""
        (valor, onde), _ = self._buscar(
            {"email": "a@b.c",
             "users": [{"email": "x@y.z"},
                       {"email": "w@y.z", "FECHAMENTO_CSV_URL": self.LINK}]})
        self.assertEqual(valor, self.LINK)
        self.assertIn("users", onde)
        self.assertIn("[1]", onde, "a tela precisa dizer QUAL item da lista")

    def test_nao_desce_em_texto(self):
        """Texto tambem responde a `in`. Sem a checagem de tabela, um trecho
        de e-mail poderia ser lido como se fosse o segredo."""
        (valor, _), _ = self._buscar({"email": "FECHAMENTO_CSV_URL@grupobeea.com.br"})
        self.assertEqual(valor, "")

    def test_lista_de_texto_nao_quebra_a_busca(self):
        """Nem toda lista e lista de tabelas -- uma lista de e-mails no meio do
        caminho nao pode derrubar a leitura dos Secrets."""
        (valor, _), _ = self._buscar(
            {"permitidos": ["a@b.c", "d@e.f"], "FECHAMENTO_CSV_URL": self.LINK})
        self.assertEqual(valor, self.LINK)

    def test_ambiente_sem_secrets_nao_quebra(self):
        """Rodar fora do Streamlit Cloud nao pode derrubar o app."""
        class Explode:
            @property
            def secrets(self):
                raise RuntimeError("sem Secrets aqui")
        ns = carregar(["_segredo_com_origem", "_segredo"], extras={"st": Explode()})
        self.assertEqual(ns["_segredo"]("QUALQUER"), "")

    def test_a_tela_distingue_ausente_de_nao_reconhecido(self):
        """Antes a mesma mensagem servia para "nao configurei" e para
        "configurei e nao funcionou" -- e quem configurou ficava sem saber
        onde procurar."""
        i = FONTE.index("Fechamento Mensal — lançamentos e conferências")
        bloco = FONTE[i:i + 4000]
        self.assertIn("_valor_secret, _onde_secret = _segredo_com_origem(", bloco)
        self.assertIn("if _valor_secret:", bloco)
        self.assertIn("nao reconheci o conteudo".replace("nao", "não").replace("conteudo", "conteúdo"),
                      bloco)
        self.assertIn("caiu dentro de uma seção", bloco,
                      "a tela precisa avisar da armadilha do TOML")


class TesteNomesIndefinidos(unittest.TestCase):
    """Nome usado e nunca definido — de VARIAVEL, nao so de funcao.

    Nasceu de um defeito real em 25/08/2026: ao trocar os dois selectbox do
    Fluxo Mensal pela funcao compartilhada, as variaveis mes_ini_sel_fin e
    mes_fim_sel_fin deixaram de existir, mas o bloco "Ajuste fino por dia"
    continuava usando as duas. A tela quebraria com NameError ao abrir o
    expansor, e a suite inteira ficou VERDE: a trava que existia
    (test_toda_funcao_chamada_existe) so olha CHAMADAS DE FUNCAO.

    Escrever um analisador de escopo aqui seria reinventar mal uma roda que ja
    existe, entao a checagem usa o pyflakes. Sem ele instalado o teste PULA
    com instrucao -- pular avisando e melhor que passar em silencio."""

    def test_nenhum_nome_indefinido_no_app(self):
        try:
            from pyflakes.api import check
            from pyflakes.reporter import Reporter
        except ImportError:
            self.skipTest(
                "pyflakes não instalado — rode `pip install pyflakes` para ligar "
                "esta trava (ela pega variável usada e nunca definida)")

        saida, erros = io.StringIO(), io.StringIO()
        check(FONTE, CAMINHO_APP, Reporter(saida, erros))
        # So os nomes indefinidos: import sem uso e f-string sem placeholder
        # sao ruido de estilo, e reprovar por eles faria a trava ser ignorada.
        indefinidos = [linha for linha in saida.getvalue().splitlines()
                       if "undefined name" in linha]
        self.assertEqual(indefinidos, [], "\n".join(indefinidos))


# ============================================================================
# 5p. ORÇAMENTO 2027 — por plano de contas
# ============================================================================
class TesteOrcamento2027(unittest.TestCase):
    """O motor que propoe o orcamento de 2027 conta a conta."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["classificar_nome_orcamento", "chave_conta_orcamento", "curva_do_ano",
             "aplicar_direcionador_2027", "distribuir_no_ano", "anualizar_realizado",
             "resumo_da_proposta", "_normalizar_coluna_fin"],
            ["PADRAO_LINHA_DRE", "DIRECIONADORES_2027", "_ACENTOS_FIN"],
        )

    def test_linha_da_dre_e_plano_de_contas_sao_separados_pelo_numero(self):
        """O criterio e o NUMERO na frente, nao a indentacao: a indentacao
        varia entre niveis e um plano fundo na hierarquia tem o mesmo recuo de
        uma linha da DRE de nivel parecido."""
        classificar = self.ns["classificar_nome_orcamento"]
        for nome in ("1 - Receita Operacional Bruta", "  2.1.1 - COFINS sobre Receita Bruta",
                     "6.24.2.6 - Outras Despesas", "11 - EBITDA"):
            self.assertEqual(classificar(nome), "dre", nome)
        for nome in ("    COFINS sobre Receita", "  PIS", "  Taxa com Cartao de Credito / Debito",
                     "Aluguel de POS / Outras Taxas"):
            self.assertEqual(classificar(nome), "plano", nome)
        self.assertEqual(classificar(""), "vazio")
        self.assertEqual(classificar(None), "vazio")

    def test_chave_de_conta_ignora_indentacao_acento_e_caixa(self):
        """O mesmo nome aparece indentado na planilha modelo e sem indentacao
        no DIARIO. Se a chave nao normalizar, nada casa e a tela mostra tudo
        zerado sem dizer por que."""
        chave = self.ns["chave_conta_orcamento"]
        self.assertEqual(chave("    Taxa com Cartão de Crédito / Débito"),
                         chave("taxa com cartao de credito / debito"))
        self.assertEqual(chave("  PIS  "), chave("Pis"))

    def test_curva_usa_valor_absoluto(self):
        """Despesa vem negativa na planilha. Sem o absoluto, o sinal inverteria
        a curva e jogaria o peso nos meses de MENOR gasto."""
        curva = self.ns["curva_do_ano"]([-100]*11 + [-1200])
        self.assertAlmostEqual(sum(curva), 1.0, places=9)
        self.assertGreater(curva[11], curva[0])
        self.assertAlmostEqual(curva[11] / curva[0], 12.0, places=6)

    def test_curva_sem_historico_cai_em_linear(self):
        """Nao ha curva a extrair de uma linha sem historico, e inventar uma
        seria pior do que assumir o mes medio."""
        for entrada in ([0]*12, [], None, [1, 2, 3]):
            self.assertEqual(self.ns["curva_do_ano"](entrada), [1/12]*12)

    def test_direcionador_em_branco_nao_vira_zero(self):
        """Conta sem direcionador e conta que ninguem decidiu; zero mentiria
        dizendo que foi decidida por zero. Quem quer zerar escolhe 'Nao orcar',
        que e uma decisao."""
        aplicar = self.ns["aplicar_direcionador_2027"]
        prem = {"ipca": 0.0425, "reajuste_sm": 0.074, "receita_2027": 1_000_000}
        self.assertIsNone(aplicar(1000, "", None, prem))
        self.assertIsNone(aplicar(1000, None, None, prem))
        self.assertEqual(aplicar(1000, "Não orçar", None, prem), 0.0)

    def test_cada_direcionador_faz_a_conta_certa(self):
        aplicar = self.ns["aplicar_direcionador_2027"]
        prem = {"ipca": 0.0425, "reajuste_sm": 0.074, "receita_2027": 1_000_000}
        self.assertAlmostEqual(aplicar(1000, "Inflação (IPCA)", None, prem), 1042.5)
        self.assertAlmostEqual(aplicar(1000, "Inflação + ganho real", 0.02, prem), 1062.5)
        self.assertAlmostEqual(aplicar(1000, "Crescimento %", -0.10, prem), 900.0)
        self.assertAlmostEqual(aplicar(1000, "% da receita", 0.0035, prem), 3500.0)
        self.assertAlmostEqual(aplicar(1000, "Valor fixo (ano)", 50_000, prem), 50_000)
        self.assertAlmostEqual(aplicar(1000, "Repetir 2026", None, prem), 1000.0)
        self.assertAlmostEqual(aplicar(1000, "Salário mínimo", None, prem), 1074.0)

    def test_os_doze_meses_somam_o_total_ao_centavo(self):
        """Diferenca de centavo entre a linha do plano e a soma dela vira meia
        hora de procura na planilha do usuario."""
        distribuir, curva = self.ns["distribuir_no_ano"], self.ns["curva_do_ano"]
        for total in (10_000.0, 33_333.33, 1.0, -7_777.77):
            meses = distribuir(total, curva([1]*12))
            self.assertAlmostEqual(sum(meses), total, places=2, msg=str(total))
        meses = distribuir(100_000.0, curva([1]*11 + [5]))
        self.assertAlmostEqual(sum(meses), 100_000.0, places=2)

    def test_total_ausente_nao_vira_doze_zeros(self):
        self.assertEqual(self.ns["distribuir_no_ano"](None, [1/12]*12), [None]*12)

    def test_anualizar_usa_so_os_meses_fechados(self):
        """Em agosto, somar os 12 valores incluiria cinco meses vazios e
        devolveria uma base 40% menor que a real, sem nada na tela avisando."""
        anualizar = self.ns["anualizar_realizado"]
        self.assertAlmostEqual(anualizar([100]*7 + [0]*5, 7), 1200.0)
        self.assertAlmostEqual(anualizar([100]*12, 12), 1200.0)
        self.assertEqual(anualizar([100]*7, 0), 0.0)

    def test_resumo_conta_as_pendentes_separado(self):
        """Num orcamento de 219 planos ninguem percebe uma faltando olhando a
        tela: o contador de pendentes e o que trava a entrega."""
        linhas = [{"direcionador": "Inflação (IPCA)", "total_2027": 100.0},
                  {"direcionador": "Não orçar", "total_2027": 0.0},
                  {"direcionador": "", "total_2027": None}]
        self.assertEqual(self.ns["resumo_da_proposta"](linhas), (2, 1, 100.0))


    def test_anualizar_ignora_o_que_vem_depois_dos_meses_fechados(self):
        """Mes ainda nao fechado pode ter numero na planilha (lancamento
        adiantado, rateio ja provisionado). A base tem de parar no mes fechado,
        senao ela mistura o que aconteceu com o que ainda vai acontecer."""
        anualizar = self.ns["anualizar_realizado"]
        self.assertAlmostEqual(anualizar([100]*7 + [999]*5, 7), 1200.0)
        self.assertAlmostEqual(anualizar([100]*3 + [999]*9, 3), 1200.0)


class TesteGeradorDoOrcamento2027(unittest.TestCase):
    """O gerador do Excel, contra um arquivo de verdade.

    Existe porque as duas travas mais perigosas do motor NAO eram cobertas por
    teste nenhum: trocar a deteccao de formula por "texto que comeca com =" e
    o gerador APAGA as formulas de matriz. As 6 abas consolidadas da planilha
    real sao 100% formula de matriz -- 419 linhas cada uma -- e o openpyxl
    devolve isso como objeto ArrayFormula, nao como texto. A suite ficava
    verde enquanto a planilha do usuario era destruida."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["ler_estrutura_orcamento", "gerar_excel_orcamento",
             "classificar_nome_orcamento"],
            ["PADRAO_LINHA_DRE"],
        )

    def _modelo(self):
        """Monta um modelo com as tres naturezas da planilha real: aba de
        unidade (formula normal + valor), aba consolidada (formula de MATRIZ
        em tudo) e linhas de plano de contas sob a linha da DRE."""
        from openpyxl import Workbook
        from openpyxl.worksheet.formula import ArrayFormula
        wb = Workbook()
        cons = wb.active
        cons.title = "CONSOLIDADO"
        uni = wb.create_sheet("UNIDADE A")
        for ws in (cons, uni):
            ws["A1"] = "Nome"
            ws["B1"] = "01/2027"
            ws["C1"] = "02/2027"
            for linha, nome in ((2, "1 - Receita"), (3, "  Venda de mercadoria"),
                                (4, "  Venda de servico"), (5, "2 - Despesas"),
                                (6, "  Aluguel")):
                ws.cell(row=linha, column=1, value=nome)
        # Unidade: linha da DRE por formula normal, plano por valor.
        for col in ("B", "C"):
            uni[f"{col}2"] = f"=SUM({col}3:{col}4)"
            uni[f"{col}5"] = f"=SUM({col}6)"
            for linha in (3, 4, 6):
                uni[f"{col}{linha}"] = 0
        # Consolidada: TUDO formula de matriz, como na planilha real.
        for col in ("B", "C"):
            for linha in range(2, 7):
                ref = f"{col}{linha}"
                cons[ref] = ArrayFormula(ref, f"=SUMPRODUCT('UNIDADE A'!{ref})")
        caminho = "/tmp/modelo_orcamento_teste.xlsx"
        wb.save(caminho)
        return caminho

    def test_le_a_estrutura_e_separa_formula_de_valor(self):
        estrutura, abas = self.ns["ler_estrutura_orcamento"](self._modelo())
        tipos = {}
        for e in estrutura:
            tipos[e["tipo"]] = tipos.get(e["tipo"], 0) + 1
        self.assertEqual(tipos.get("formula"), 2, "as 2 linhas da DRE sao formula")
        self.assertEqual(tipos.get("plano"), 3, "os 3 planos sao de preencher")
        self.assertEqual([e["nome"] for e in estrutura if e["editavel"]],
                         ["Venda de mercadoria", "Venda de servico", "Aluguel"])

    def test_a_aba_toda_calculada_fica_fora_da_lista_de_preencher(self):
        """Escrever nela apagaria a soma das unidades. Ela se resolve sozinha
        quando as unidades forem preenchidas."""
        _, abas = self.ns["ler_estrutura_orcamento"](self._modelo())
        self.assertEqual(abas, ["UNIDADE A"])

    def test_a_referencia_nao_e_a_primeira_aba(self):
        """A primeira aba da planilha real e "DRE CONSOLIDADO", 100% calculada:
        dela nao se extrai linha de valor nenhuma, e a estrutura sairia vazia."""
        estrutura, _ = self.ns["ler_estrutura_orcamento"](self._modelo())
        self.assertTrue(any(e["editavel"] for e in estrutura),
                        "leu a aba calculada e nao achou linha de valor")

    def test_o_plano_aponta_para_a_linha_da_dre_dele(self):
        estrutura, _ = self.ns["ler_estrutura_orcamento"](self._modelo())
        por_nome = {e["nome"]: e for e in estrutura}
        self.assertEqual(por_nome["Venda de mercadoria"]["linha_dre"], "1 - Receita")
        self.assertEqual(por_nome["Aluguel"]["linha_dre"], "2 - Despesas")

    def test_nenhuma_formula_e_destruida_nem_a_de_matriz(self):
        """A trava que faltava. Sem ela, as 6 abas consolidadas da planilha
        real perdiam as 419 formulas cada uma -- e o arquivo voltava para o
        usuario parecendo certo, so que morto."""
        from openpyxl import load_workbook
        caminho = self._modelo()
        estrutura, abas = self.ns["ler_estrutura_orcamento"](caminho)
        # De proposito manda escrever em TODA linha, inclusive nas de formula
        # e na aba calculada: o gerador tem de se recusar sozinho.
        todas_as_linhas = {e["linha"]: [111.0, 222.0] for e in estrutura}
        valores = {"UNIDADE A": todas_as_linhas, "CONSOLIDADO": todas_as_linhas}
        saida, _ = self.ns["gerar_excel_orcamento"](caminho, valores)
        with open("/tmp/saida_orcamento_teste.xlsx", "wb") as arq:
            arq.write(saida)
        antes = load_workbook(caminho)
        depois = load_workbook("/tmp/saida_orcamento_teste.xlsx")
        destruidas = 0
        for aba in antes.sheetnames:
            a, b = antes[aba], depois[aba]
            for linha in range(2, a.max_row + 1):
                for col in (2, 3):
                    if a.cell(row=linha, column=col).data_type == "f":
                        if b.cell(row=linha, column=col).data_type != "f":
                            destruidas += 1
        self.assertEqual(destruidas, 0, "o gerador apagou formula")
        # E o valor TEM de ter entrado nas linhas de plano.
        self.assertEqual(depois["UNIDADE A"]["B3"].value, 111.0)
        self.assertEqual(depois["UNIDADE A"]["C6"].value, 222.0)



class TesteTelaDoOrcamento(unittest.TestCase):
    """Travas da aba: o que a tela tem de fazer antes de mostrar numero."""

    def _bloco(self):
        i = FONTE.index("Orçamento — por plano de contas")
        return FONTE[i:FONTE.index("# ABA 6: GESTÃO DE USUÁRIOS", i)]

    def test_o_casamento_dos_nomes_vem_antes_dos_numeros(self):
        """Nome que nao casa vira base zero, e a tela mostraria "R$ 0,00"
        igualzinho a uma conta que de fato nao teve gasto -- quem olha nao tem
        como distinguir. Por isso a conferencia aparece ANTES, e nao escondida
        num expander fechado."""
        bloco = self._bloco()
        self.assertIn("conferir_casamento_dos_planos(", bloco)
        posicao_conferencia = bloco.index("conferir_casamento_dos_planos(")
        posicao_editor = bloco.index("st.data_editor(")
        self.assertLess(posicao_conferencia, posicao_editor,
                        "o casamento tem de ser conferido antes da tabela de valores")
        self.assertIn("não tiveram", bloco)

    def test_a_estrutura_e_lida_do_arquivo_e_nao_cravada(self):
        """A lista de planos muda de um ano para o outro. Cravada no codigo,
        ela viraria mentira em silencio na primeira mudanca -- escrevendo
        valores na linha errada da planilha do usuario."""
        bloco = self._bloco()
        self.assertIn("st.file_uploader(", bloco)
        self.assertIn("ler_estrutura_orcamento(", bloco)

    def test_o_ano_sai_do_modelo_e_nao_do_codigo(self):
        """A aba serve para 2027, 2028 e os que vierem. Ano cravado faria a
        tela buscar o realizado do ano errado, sem nada avisando."""
        bloco = self._bloco()
        self.assertIn("anos_do_modelo_orcamento(", bloco)
        self.assertIn("_ano_base_orc", bloco)
        self.assertNotIn('f"{m:02d}/2026"', bloco, "o ano voltou a ser cravado")

    def test_o_realizado_das_linhas_da_dre_e_buscado(self):
        """Elas nao estao no DIARIO -- ele tem plano de contas, nao linha. Sem
        esta busca apareciam TODAS zeradas, e o zero se confundia com "nao teve
        gasto". A cascata cai na aba DRE MENSAL para as linhas que a planilha
        de Realizado guarda so la."""
        bloco = self._bloco()
        self.assertIn("realizado_da_dre_por_aba(", bloco)
        self.assertIn("DRE MENSAL", bloco)
        self.assertIn('"Origem"', bloco,
                      "a tela precisa dizer de onde veio a base de cada linha")

    def test_plano_sem_movimento_nao_e_tratado_como_erro(self):
        """Plano sem lancamento no DIARIO simplesmente nao foi usado no ano. O
        aviso vermelho tratava fato normal como defeito e mandava conferir nome
        que estava certo."""
        bloco = self._bloco()
        # A frase quebra entre duas linhas do codigo -- procurar o texto
        # inteiro falharia por formatacao, nao por defeito.
        self.assertIn("não tiveram ", bloco)
        self.assertIn("movimento no ano anterior", bloco)
        self.assertNotIn("não foram encontrados no DIÁRIO", bloco,
                         "o aviso de erro voltou para um fato normal")
        self.assertIn("Valor fixo (ano)", bloco,
                      "a tela precisa dizer COMO orcar uma conta de base zero")

    def test_a_tela_avisa_das_contas_sem_direcionador(self):
        """Conta em branco nao e conta zerada, e conta esquecida. Num
        orcamento de 219 planos ninguem percebe uma faltando olhando a tela."""
        bloco = self._bloco()
        self.assertIn("resumo_da_proposta(", bloco)
        self.assertIn("SEM DIRECIONADOR", bloco)
        self.assertIn("Não orçar", bloco)



class TesteRealizadoDaDreNoOrcamento(unittest.TestCase):
    """A cascata que busca o realizado das LINHAS DA DRE.

    Elas nao estao no DIARIO -- ele tem plano de contas, nao linha -- e
    algumas nem estao nas abas de unidade: a planilha de Realizado as guarda
    numa aba propria, a DRE MENSAL. Sem a cascata, essas linhas apareciam
    TODAS zeradas na tela, e o zero se confundia com "nao teve gasto"."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["realizado_da_dre_por_aba", "realizado_da_linha_dre",
             "chave_conta_orcamento", "_normalizar_coluna_fin"],
            ["_ACENTOS_FIN"],
        )
        cls.meses = ["01/2026", "02/2026"]

    def _aba(self, pares):
        return pd.DataFrame([{"Nome": n, "01/2026": v, "02/2026": v} for n, v in pares])

    def test_acha_na_aba_da_unidade(self):
        dados = {"LOJA A": (pd.DataFrame(), self._aba([("4.1 - CMV", -100)]))}
        saida = self.ns["realizado_da_dre_por_aba"](
            dados, pd.DataFrame(), ["4.1 - CMV"], self.meses)
        valor, origem = saida[("LOJA A", self.ns["chave_conta_orcamento"]("4.1 - CMV"))]
        self.assertAlmostEqual(valor, -200.0)
        self.assertEqual(origem, "unidade")

    def test_cai_na_dre_mensal_quando_a_unidade_nao_tem(self):
        """O caso real: 1.1 - Vendas de mercadorias, ICMS, CMV e devolucoes
        nao existem nas abas de unidade da planilha de Realizado."""
        dados = {"LOJA A": (pd.DataFrame(), self._aba([("outra linha", -5)]))}
        mensal = self._aba([("1.1 - Vendas de mercadorias", 900)])
        saida = self.ns["realizado_da_dre_por_aba"](
            dados, mensal, ["1.1 - Vendas de mercadorias"], self.meses)
        valor, origem = saida[("LOJA A", self.ns["chave_conta_orcamento"]("1.1 - Vendas de mercadorias"))]
        self.assertAlmostEqual(valor, 1800.0)
        self.assertEqual(origem, "DRE MENSAL",
                         "a origem precisa viajar junto: esse numero e da empresa "
                         "inteira, nao daquela unidade")

    def test_sem_movimento_em_lugar_nenhum_e_dito_como_tal(self):
        """Zero por falta de lancamento e zero por nome que nao casa parecem
        iguais na tela. A origem e o que separa os dois."""
        dados = {"LOJA A": (pd.DataFrame(), self._aba([("outra", -5)]))}
        saida = self.ns["realizado_da_dre_por_aba"](
            dados, self._aba([("mais outra", 1)]), ["4.1 - CMV"], self.meses)
        valor, origem = saida[("LOJA A", self.ns["chave_conta_orcamento"]("4.1 - CMV"))]
        self.assertEqual(valor, 0.0)
        self.assertEqual(origem, "sem movimento")

    def test_a_unidade_tem_prioridade_sobre_a_dre_mensal(self):
        """Quando a unidade tem o numero, ele vale: a DRE MENSAL e da empresa
        inteira e sobrescreveria a unidade com um valor 21 vezes maior."""
        dados = {"LOJA A": (pd.DataFrame(), self._aba([("4.1 - CMV", -100)]))}
        mensal = self._aba([("4.1 - CMV", -99999)])
        saida = self.ns["realizado_da_dre_por_aba"](
            dados, mensal, ["4.1 - CMV"], self.meses)
        valor, origem = saida[("LOJA A", self.ns["chave_conta_orcamento"]("4.1 - CMV"))]
        self.assertAlmostEqual(valor, -200.0)
        self.assertEqual(origem, "unidade")



class TesteDirecionadorPorLoja(unittest.TestCase):
    """A camada de excecao por loja (25/08/2026).

    Sao 219 planos x 21 lojas -- quase 4.600 decisoes. Ninguem preenche isso
    numa tabela, entao a regra e definida uma vez e sobreposta so onde a loja
    foge dela. A tela entrega as contas que merecem a segunda olhada."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["direcionador_efetivo", "peso_da_loja_no_total",
             "desvio_do_padrao_da_loja", "contas_fora_do_padrao",
             "resumo_das_excecoes"])

    def test_a_excecao_da_loja_manda_sobre_a_regra_geral(self):
        efetivo = self.ns["direcionador_efetivo"]
        geral = {"Aluguel": ("Inflação (IPCA)", 0.0)}
        excecoes = {("LOJA A", "Aluguel"): ("Crescimento %", 0.15)}
        self.assertEqual(efetivo("Aluguel", "LOJA A", geral, excecoes),
                         ("Crescimento %", 0.15, "exceção da loja"))
        self.assertEqual(efetivo("Aluguel", "LOJA B", geral, excecoes),
                         ("Inflação (IPCA)", 0.0, "regra geral"))

    def test_excecao_vazia_faz_a_conta_voltar_a_herdar(self):
        """E assim que se desfaz uma excecao: apagar o direcionador da loja."""
        efetivo = self.ns["direcionador_efetivo"]
        geral = {"Aluguel": ("Inflação (IPCA)", 0.0)}
        self.assertEqual(efetivo("Aluguel", "LOJA A", geral, {("LOJA A", "Aluguel"): ("", 0.0)}),
                         ("Inflação (IPCA)", 0.0, "regra geral"))

    def test_a_origem_viaja_junto(self):
        """Na reuniao a primeira pergunta e "por que esta loja esta
        diferente" -- a resposta tem de estar na tela."""
        efetivo = self.ns["direcionador_efetivo"]
        self.assertEqual(efetivo("X", "LOJA A", {}, {})[2], "sem direcionador")

    def test_peso_da_loja_usa_valor_absoluto(self):
        """Receita e positiva e despesa e negativa. Somar sem modulo daria um
        total perto de zero e faria qualquer loja parecer ter peso infinito."""
        pesos = self.ns["peso_da_loja_no_total"](
            {"A": {"receita": 1000, "despesa": -900},
             "B": {"receita": 100, "despesa": -0}})
        self.assertAlmostEqual(pesos["A"], 1900 / 2000, places=6)
        self.assertAlmostEqual(pesos["B"], 100 / 2000, places=6)
        self.assertAlmostEqual(sum(pesos.values()), 1.0, places=9)

    def test_desvio_compara_a_conta_com_o_padrao_da_loja(self):
        """A loja que responde por 8% da empresa deveria responder por perto
        de 8% de cada conta. 30% numa conta e onde a regra geral nao serve."""
        desvio = self.ns["desvio_do_padrao_da_loja"]
        self.assertAlmostEqual(desvio(30, 100, 0.08), 0.30 / 0.08 - 1, places=6)
        self.assertAlmostEqual(desvio(8, 100, 0.08), 0.0, places=9)
        self.assertAlmostEqual(desvio(2, 100, 0.08), 0.02 / 0.08 - 1, places=6)

    def test_sem_como_comparar_devolve_nada_e_nao_um_numero(self):
        """Inventar um numero aqui encheria o topo da lista de contas que nao
        significam nada -- exatamente as que a tela existe para esconder."""
        desvio = self.ns["desvio_do_padrao_da_loja"]
        self.assertIsNone(desvio(10, 0, 0.08))
        self.assertIsNone(desvio(10, 100, 0))

    def test_a_lista_vem_ordenada_pelo_tamanho_do_desvio(self):
        """O objetivo e entregar as poucas contas que merecem uma segunda
        olhada, nao devolver as 219 numa ordem qualquer."""
        # A ordem de ENTRADA e proposital: o desvio maior vem por ULTIMO no
        # dicionario. Com os dados ja na ordem certa, tirar o sorted nao muda
        # nada e o teste nao provaria a ordenacao.
        fora = self.ns["contas_fora_do_padrao"](
            {"desvio pequeno": 2, "no padrao": 10, "desvio enorme": 90},
            {"desvio pequeno": 100, "no padrao": 100, "desvio enorme": 100},
            0.10, minimo=0.5)
        self.assertEqual([a["conta"] for a in fora], ["desvio enorme", "desvio pequeno"],
                         "a lista tem de vir do maior desvio para o menor")
        self.assertNotIn("no padrao", [a["conta"] for a in fora])
        self.assertGreater(abs(fora[0]["desvio"]), abs(fora[1]["desvio"]))

    def test_o_corte_controla_o_tamanho_da_lista(self):
        """Sem corte a lista viria com tudo e nao pouparia trabalho nenhum."""
        bases = {f"c{i}": v for i, v in enumerate((11, 12, 30, 5))}
        empresa = {f"c{i}": 100 for i in range(4)}
        self.assertEqual(len(self.ns["contas_fora_do_padrao"](bases, empresa, 0.10, 0.5)), 2)
        self.assertEqual(len(self.ns["contas_fora_do_padrao"](bases, empresa, 0.10, 3.0)), 0)

    def test_resumo_conta_so_as_excecoes_de_verdade(self):
        """Direcionador vazio nao e excecao: e heranca da regra geral."""
        self.assertEqual(self.ns["resumo_das_excecoes"]({
            ("A", "x"): ("Crescimento %", 0.1),
            ("A", "y"): ("", 0.0),
            ("B", "z"): ("Repetir 2026", 0.0)}), (2, 2))
        self.assertEqual(self.ns["resumo_das_excecoes"]({}), (0, 0))


class TesteTelaDasExcecoesPorLoja(unittest.TestCase):
    """Travas da tela de excecoes."""

    def _bloco(self):
        i = FONTE.index("🏪 Exceções por loja")
        return FONTE[i:FONTE.index("# ---- Resumo do que já foi decidido", i)]

    def test_a_tela_mostra_o_que_esta_valendo_e_de_onde_vem(self):
        bloco = self._bloco()
        self.assertIn("direcionador_efetivo(", bloco)
        self.assertIn("Valendo hoje", bloco)

    def test_a_lista_e_so_do_que_foge_do_padrao(self):
        """Listar as 219 contas de cada loja nao pouparia trabalho nenhum."""
        bloco = self._bloco()
        self.assertIn("contas_fora_do_padrao(", bloco)
        self.assertIn("peso_da_loja_no_total(", bloco)

    def test_a_proposta_e_calculada_UMA_vez_e_reusada(self):
        """A memoria de calculo, o confronto com a meta e o gerador do Excel
        precisam do MESMO numero. Calcular em tres lugares seriam tres chances
        de eles discordarem -- e a discordancia entre a tela e o arquivo so
        apareceria depois de o Excel ja estar colado."""
        i = FONTE.index("_proposta_por_loja, _memoria_itens = {}, []")
        trecho = FONTE[i:i + 3000]
        self.assertIn("direcionador_efetivo(", trecho,
                      "o calculo unico tem de respeitar a excecao da loja")
        self.assertIn("_excecoes_orc", trecho)
        # O gerador NAO pode refazer a conta: ele le a proposta pronta.
        j = FONTE.index("Gerar a planilha preenchida")
        gerador = FONTE[j:FONTE.index("gerar_excel_orcamento(", j)]
        self.assertIn("_proposta_por_loja.get(", gerador)
        self.assertNotIn("aplicar_direcionador_2027(", gerador,
                         "o gerador voltou a refazer a conta por conta propria")



class TesteConfrontoComAMeta(unittest.TestCase):
    """A comparacao entre a proposta do painel e a meta da industria, que
    chega por loja e por mes."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["ler_meta_da_industria", "confrontar_com_a_meta", "memoria_de_calculo",
             "chave_conta_orcamento", "_normalizar_nome_aba", "_normalizar_coluna_fin",
             "_normalizar_texto"],
            ["_ACENTOS_FIN"])

    def _planilha_meta(self, linhas, cabecalho=None):
        colunas = cabecalho or (["Loja", "Conta"] + [f"{m:02d}/2027" for m in range(1, 13)])
        caminho = "/tmp/meta_teste.xlsx"
        pd.DataFrame(linhas, columns=colunas).to_excel(caminho, index=False)
        return caminho

    def test_le_a_planilha_padrao(self):
        arq = self._planilha_meta([["LJ MARECHAL 6039", "1.1 - Vendas de mercadorias"] + [1000] * 12])
        meta, meses, erro = self.ns["ler_meta_da_industria"](arq)
        self.assertEqual(erro, "")
        self.assertEqual(len(meses), 12)
        self.assertEqual(len(meta), 1)
        self.assertEqual(meta["chave_loja"].iloc[0],
                         self.ns["_normalizar_nome_aba"]("LJ MARECHAL 6039"))

    def test_planilha_sem_as_colunas_obrigatorias_e_apontada(self):
        """Erro generico manda a pessoa adivinhar; o nome da coluna resolve em
        segundos."""
        arq = self._planilha_meta([["x", 1]], cabecalho=["Unidade", "Valor"])
        _meta, _meses, erro = self.ns["ler_meta_da_industria"](arq)
        self.assertIn("Loja", erro)
        self.assertIn("Conta", erro)

    def test_mes_fora_do_formato_e_apontado(self):
        """Os meses sao TEXTO 01/2027. Data viraria mes/dia/ano na exportacao
        e trocaria janeiro por outra coisa."""
        arq = self._planilha_meta([["LOJA", "conta", 1]],
                                  cabecalho=["Loja", "Conta", "Janeiro"])
        _meta, _meses, erro = self.ns["ler_meta_da_industria"](arq)
        self.assertIn("01/2027", erro)

    def test_confronto_soma_o_ano_e_calcula_a_diferenca(self):
        arq = self._planilha_meta([["LJ MARECHAL 6039", "1.1 - Vendas de mercadorias"] + [1000] * 12])
        meta, meses, _ = self.ns["ler_meta_da_industria"](arq)
        chave = (self.ns["_normalizar_nome_aba"]("LJ MARECHAL 6039"),
                 self.ns["chave_conta_orcamento"]("1.1 - Vendas de mercadorias"))
        confronto, so_meta, so_prop = self.ns["confrontar_com_a_meta"](
            {chave: [1100.0] * 12}, meta, meses)
        self.assertEqual(len(confronto), 1)
        self.assertAlmostEqual(confronto["Proposta do painel"].iloc[0], 13200.0)
        self.assertAlmostEqual(confronto["Meta da indústria"].iloc[0], 12000.0)
        self.assertAlmostEqual(confronto["Diferença"].iloc[0], 1200.0)
        self.assertAlmostEqual(confronto["Diferença %"].iloc[0], 0.10, places=6)
        self.assertEqual((so_meta, so_prop), ([], []))

    def test_o_que_ficou_de_um_lado_so_e_devolvido(self):
        """Quase sempre e nome escrito diferente, nao meta faltando. Sem esta
        lista a pessoa compara um subconjunto achando que comparou tudo."""
        arq = self._planilha_meta([["LOJA QUE NAO EXISTE", "conta"] + [10] * 12])
        meta, meses, _ = self.ns["ler_meta_da_industria"](arq)
        confronto, so_meta, so_prop = self.ns["confrontar_com_a_meta"](
            {("outra", "outra conta"): [1.0] * 12}, meta, meses)
        self.assertTrue(confronto.empty)
        self.assertEqual(len(so_meta), 1)
        self.assertEqual(len(so_prop), 1)

    def test_meta_zerada_nao_vira_divisao_por_zero(self):
        arq = self._planilha_meta([["LJ A", "conta"] + [0] * 12])
        meta, meses, _ = self.ns["ler_meta_da_industria"](arq)
        chave = (self.ns["_normalizar_nome_aba"]("LJ A"),
                 self.ns["chave_conta_orcamento"]("conta"))
        confronto, _sm, _sp = self.ns["confrontar_com_a_meta"](
            {chave: [100.0] * 12}, meta, meses)
        self.assertIsNone(confronto["Diferença %"].iloc[0])

    def test_a_memoria_diz_de_onde_cada_numero_saiu(self):
        """"O painel calculou" nao e resposta quando um departamento questiona
        o valor dele. A defesa tem de caber numa linha."""
        tabela = self.ns["memoria_de_calculo"]([{
            "loja": "LJ A", "linha_dre": "6 - Despesas", "conta": "Aluguel",
            "base": 1000.0, "origem_base": "DIÁRIO", "direcionador": "Inflação (IPCA)",
            "parametro": 0.0, "origem_direcionador": "regra geral", "proposto": 1042.5,
        }])
        linha = tabela.iloc[0]
        self.assertEqual(linha["Origem da base"], "DIÁRIO")
        self.assertEqual(linha["Vem de"], "regra geral")
        self.assertAlmostEqual(linha["Variação %"], 0.0425, places=6)

    def test_memoria_com_base_zero_nao_inventa_variacao(self):
        """Conta que nao teve movimento no ano anterior: dividir por zero daria
        infinito, e a tela mostraria isso como se fosse informacao."""
        tabela = self.ns["memoria_de_calculo"]([{
            "loja": "LJ A", "conta": "Nova", "base": 0.0, "proposto": 5000.0}])
        self.assertIsNone(tabela.iloc[0]["Variação %"])



class TestePainelTV(unittest.TestCase):
    """A rosca de composicao e o ranque de despesas operacionais."""

    def _bloco_tv(self):
        i = FONTE.index("def renderizar_painel_tv(")
        return FONTE[i:FONTE.index("\ndef ", i + 10)]

    def test_a_rosca_tem_um_cinza_so(self):
        """Eram TRES cinzas em 4 fatias: a rosca virava um borrao e so o azul
        se distinguia."""
        ns = carregar([], ["COLORS"])
        cores = ns["COLORS"]
        bloco = self._bloco_tv()
        i = bloco.index("marker=dict(colors=[COLORS")
        # So a LISTA de cores. Recortar ate o primeiro "]" nao serve: ele
        # fecha COLORS["primary"], nao a lista. O fim e a linha da borda, que
        # vem logo depois e usa uma cor que NAO e fatia.
        trecho = bloco[i:bloco.index("line=dict(", i)]
        usadas = re.findall(r'COLORS\["(\w+)"\]', trecho)
        self.assertEqual(len(usadas), 4, "a rosca tem 4 fatias")
        cinzentas = [c for c in usadas if c in ("muted_line", "secondary", "border_soft")]
        self.assertEqual(len(cinzentas), 1,
                         f"a rosca voltou a ter mais de um cinza: {cinzentas}")
        self.assertIn("warning", usadas,
                      "a Operacional tem de ser ambar, a mesma cor das barras "
                      "da lista que detalha justamente ela")
        self.assertIn("accent", cores, "sumiu a cor neutra da paleta")

    def test_a_lista_ao_lado_usa_a_MESMA_cor_da_fatia(self):
        """Divergir aqui faria o quadradinho de uma linha apontar para outra
        fatia da rosca."""
        bloco = self._bloco_tv()
        i = bloco.index("categorias_custo = [")
        trecho = bloco[i:i + 700]
        # A tupla ganhou a linha da DRE no fim (o modo mes a mes precisa saber
        # o que buscar), entao o teste olha o par nome/cor, nao a tupla inteira.
        self.assertIn('("Despesas Operacionais", desp_op_tv_kpi, desp_op_tv_o, COLORS["warning"]',
                      trecho)

    def test_o_detalhe_nao_disputa_lugar_no_ranque(self):
        """A linha de detalhe e um recorte de DENTRO do pai. Se disputasse
        posicao, o mesmo dinheiro apareceria duas vezes e empurraria um grupo
        de verdade para fora do top 5."""
        bloco = self._bloco_tv()
        self.assertIn("top_despop = detalhe_despop[:5]", bloco)
        i = bloco.index("ranque_despop = []")
        trecho = bloco[i:i + 1400]
        # O detalhe e pendurado DEPOIS do corte do top 5, nunca antes.
        self.assertLess(bloco.index("top_despop = detalhe_despop[:5]"),
                        bloco.index("ranque_despop = []"))
        self.assertIn("DETALHES_DO_RANQUE_TV", trecho)

    def test_o_detalhe_se_mede_contra_o_pai(self):
        """O detalhe ja esta DENTRO do grupo. Mostrar "4% das despesas
        operacionais" ao lado dos outros convidava a somar a coluna, e a soma
        daria mais que o total do cartao acima -- o mesmo dinheiro contado duas
        vezes. Contra o pai, o numero responde a pergunta certa: quanto de
        Servicos de Terceiros e transporte."""
        b = self._bloco_tv()
        self.assertIn("pct_do_despop = (v_grp / v_pai * 100) if v_pai else 0", b)
        # Recorte da LINHA do ranque: "% do grupo" tambem aparece noutro
        # ponto do painel, e procurar no bloco inteiro deixava mutilar este.
        i_rot = b.index('<div class="tv-rank-val">')
        rotulo = b[i_rot:i_rot + 700]
        # O rotulo distingue detalhe de grupo pela COR, nao mais pelo texto
        # "do grupo" -- ele nao cabia na coluna e quebrava a fileira em duas
        # linhas. O recuo e a seta ja dizem que a linha e um detalhe, e a
        # legenda do bloco explica contra o que o percentual e medido.
        self.assertIn("if eh_detalhe else", rotulo,
                      "o rotulo deixou de distinguir detalhe de grupo")
        self.assertIn("não entra na soma", b,
                      "a legenda tem de dizer que o detalhe nao soma")
        # A barra tambem: contra o pai, senao ela some ao lado dos grupos.
        self.assertIn("v_grp / (v_pai if eh_detalhe else max_despop)", b)

    def test_a_lista_tem_o_total_dos_principais_grupos(self):
        """O cartao de cima traz o total das despesas operacionais INTEIRAS,
        que inclui grupos que nem aparecem na lista. Sem um total dos cinco,
        nao da para saber quanto eles representam do bloco."""
        b = self._bloco_tv()
        self.assertIn("Total dos principais grupos", b)
        # E o detalhe NAO pode entrar nessa soma: ele ja esta dentro do pai.
        self.assertIn("_soma_grupos = sum(v for _n, v, _d, _l, _p, _lp in ranque_despop if not _d)", b)

    def test_os_grupos_mudam_de_coluna_no_mes_a_mes(self):
        """A tabela e larga; a coluna da esquerda fica vazia abaixo dos
        graficos enquanto a direita empilha barras, cartoes e tabela numa
        fileira so."""
        b = self._bloco_tv()
        # DOIS espacos reservados, criados sempre, um em cada coluna: so um
        # recebe conteudo e o outro e esvaziado por ordem. Sem isso, ao trocar
        # de modo o bloco mudava de coluna e o Streamlit deixava o desenho
        # anterior preso na coluna antiga -- um fantasma que nao sai sozinho.
        self.assertIn("_vaga_grupos_esq = cgtv1.empty()", b)
        self.assertIn("_vaga_grupos_dir = cgtv2.empty()", b)
        self.assertIn("_vaga_limpa.empty()", b,
                      "sem limpar a outra vaga, o bloco aparece duas vezes")
        self.assertIn("_html_grupos_tv", b)

    def test_os_iframes_do_rodape_nao_rolam(self):
        """Conteudo alguns pixels mais alto que a altura pedida faz o navegador
        desenhar uma barra de rolagem ao lado do relogio e do botao de tela
        cheia. Aumentar a altura nao basta: a barra volta em qualquer tela com
        fonte ou escala diferente."""
        b = self._bloco_tv()
        # As chaves sao DUPLAS no codigo: o trecho vive dentro de uma
        # f-string, e no arquivo aparece {{ }} para render como { }.
        self.assertEqual(b.count("html,body{{margin:0;padding:0;overflow:hidden;}}"), 2,
                         "os DOIS iframes precisam travar a rolagem")


    def test_a_depreciacao_aparece_na_lista_de_composicao(self):
        """A rosca mostrava quatro fatias e a lista explicava tres: a fatia de
        1,37% nao tinha nome em lugar nenhum."""
        b = self._bloco_tv()
        i = b.index("categorias_custo = [")
        trecho = b[i:b.index("]", b.index("13 - Depreciação", i))]
        self.assertIn('"Depreciação / Amort."', trecho)
        self.assertIn("deprec_tv_o", trecho, "sem o orcado nao ha desvio para mostrar")

    def test_a_lista_de_composicao_tem_barra(self):
        """A escala e a MAIOR categoria, nao o total das saidas: com o total, o
        CMV encostaria em 60% e as outras tres virariam tracinhos."""
        b = self._bloco_tv()
        self.assertIn("_maior_cat = max(", b)
        i = b.index("_maior_cat = max(")
        self.assertIn("tv-rank-bar-fill", b[i:i + 1400])


    def test_a_escala_da_barra_segue_o_maior_GRUPO(self):
        """O detalhe e sempre menor que o pai; deixa-lo definir a escala
        encolheria todas as barras de uma vez."""
        bloco = self._bloco_tv()
        self.assertIn("max(v for _n, v, _d, _l, _p, _lp in ranque_despop if not _d)", bloco)

    def test_o_detalhe_aparece_recuado_e_marcado(self):
        """Quem bate o olho tem de ver na hora que aquilo esta DENTRO da linha
        de cima, e nao somando com ela."""
        bloco = self._bloco_tv()
        self.assertIn('_prefixo = "↳ " if eh_detalhe else ""', bloco)
        self.assertIn("padding-left:18px", bloco)

    def test_o_filtro_permite_escolher_meses_e_o_tipo_de_visao(self):
        """Antes era um mes so, e a tela SEMPRE mostrava o acumulado ate ele --
        nao havia como ver um mes isolado nem um trimestre."""
        bloco = self._bloco_tv()
        self.assertIn("st.multiselect(", bloco)
        # E o seletor tem de morar DENTRO de um popover: solto no cabecalho,
        # oito meses viravam oito pilulas que cobriam os cartoes de KPI.
        self.assertIn("with st.popover(", bloco)
        self.assertIn('["Acumulado", "Mês a mês"]', bloco)
        # O padrao reproduz o comportamento antigo, para quem ja usava nao
        # estranhar a tela.
        # O padrao e semeado na sessao (o multiselect vive dentro do popover,
        # e `default=` nao vale quando a chave ja existe).
        self.assertIn('st.session_state["tv_sel_meses"] = nomes_meses_tv[: idx_mes_atual + 1]',
                      bloco)

    def test_os_meses_saem_em_ordem_de_calendario(self):
        """O multiselect devolve na ordem em que a pessoa CLICOU. Escolher
        "marco, janeiro" faria as colunas do mes a mes sairem fora de ordem."""
        bloco = self._bloco_tv()
        # Tem de percorrer a lista OFICIAL de meses e filtrar pelos clicados --
        # percorrer os clicados preserva a ordem do clique, que e o defeito.
        self.assertIn("meses_ativos_tv = [n for n in nomes_meses_tv", bloco)
        self.assertNotIn("for n in meses_escolhidos_tv if n in m_map_tv", bloco,
                         "voltou a percorrer a ordem em que a pessoa clicou")
        # Modelo do que o codigo faz.
        nomes = ["janeiro", "fevereiro", "março", "abril"]
        clicados = {"março", "janeiro"}
        self.assertEqual([n for n in nomes if n in clicados], ["janeiro", "março"])

    def test_nenhum_mes_escolhido_nao_zera_a_tela(self):
        """Multiselect vazio zeraria o painel inteiro sem dizer por que."""
        bloco = self._bloco_tv()
        self.assertIn("if not meses_ativos_tv:", bloco)
        i = bloco.index("if not meses_ativos_tv:")
        self.assertIn("nomes_meses_tv[: idx_mes_atual + 1]", bloco[i:i + 400])

    def test_a_legenda_diz_qual_recorte_esta_na_tela(self):
        """Um mes so, acumulado ate X ou meses avulsos sao tres coisas
        diferentes, e a tela de parede nao tem quem pergunte."""
        bloco = self._bloco_tv()
        self.assertIn('legenda_periodo_tv = f"{meses_ativos_tv[0].capitalize()}', bloco)
        self.assertIn('f"Acumulado até {meses_ativos_tv[-1].capitalize()}', bloco)
        self.assertIn('legenda_periodo_tv += " · mês a mês"', bloco)

    def test_no_mes_a_mes_a_rosca_vira_barra_empilhada(self):
        """Rosca compara PROPORCAO, nao evolucao -- e tres roscas lado a lado
        nao cabem numa tela de parede."""
        bloco = self._bloco_tv()
        i = bloco.index("_cores_composicao_tv = [")
        trecho = bloco[i:bloco.index("fig_tv_donut", i)]
        self.assertIn("_abrir_por_mes_tv_kpi", trecho,
                      "a rosca deixou de reagir ao tipo de visao")
        self.assertIn('barmode="stack"', trecho)
        self.assertIn("go.Bar(", trecho)

    def test_o_modo_e_so_o_que_a_pessoa_escolheu(self):
        """A versao anterior tinha um "and len(meses_ativos_tv) > 1" grudado no
        modo: enquanto se editava a lista de meses, a selecao passava por UM mes
        e a tela voltava sozinha para consolidado -- com o seletor ainda escrito
        "Mes a mes". Tela que troca de modo por conta propria faz quem esta
        olhando duvidar do numero, que e o pior que um painel executivo causa."""
        bloco = self._bloco_tv()
        self.assertIn('_abrir_por_mes_tv_kpi = tipo_visao_tv == "Mês a mês"', bloco)
        self.assertNotIn('tipo_visao_tv == "Mês a mês" and len(meses_ativos_tv)', bloco,
                         "o modo voltou a depender da quantidade de meses")
        # UMA marca so, reusada: duas condicoes iguais em dois lugares seriam
        # duas chances de a tela ficar meio num modo e meio no outro.
        self.assertIn("_abrir_por_mes_tv = _abrir_por_mes_tv_kpi", bloco)
        # E o tipo tambem e semeado na sessao, como os meses.
        # O modo vive numa chave COMUM, que o Streamlit nao recolhe: qualquer
        # st.rerun() disparado ANTES do seletor no cabecalho apagava o estado
        # do widget e a tela voltava a Acumulado sozinha.
        self.assertIn('st.session_state["_tv_modo"] = tipo_visao_tv', bloco)
        # A copia SO e usada quando o estado do widget sumiu. Escrever a copia
        # por cima a cada execucao apagava a escolha no instante seguinte ao
        # clique -- ficou pior que o defeito original, nao dava nem para trocar.
        self.assertIn('if "tv_sel_tipo_visao" not in st.session_state:', bloco)
        i_seed = bloco.index('if "tv_sel_tipo_visao" not in st.session_state:')
        self.assertIn('st.session_state.get("_tv_modo"', bloco[i_seed:i_seed + 400])
        self.assertNotIn('        st.session_state["tv_sel_tipo_visao"] = _modo_guardado', bloco,
                         "a copia voltou a ser escrita por cima da escolha")
        # E o atalho NAO pode chamar st.rerun(): era o que abortava o script
        # antes de o seletor existir.
        i = bloco.index('key=f"tv_atalho_')
        # Ignora COMENTARIO: o trecho tem um comentario longo explicando por
        # que o rerun nao esta ali, e procurar no texto cru encontrava a
        # propria explicacao. Ja aconteceu antes com o verificador de
        # JavaScript, e e o mesmo engano.
        codigo_atalho = "\n".join(
            linha for linha in bloco[i:i + 900].split("\n")
            if not linha.strip().startswith("#"))
        self.assertNotIn("st.rerun()", codigo_atalho,
                         "o atalho voltou a chamar rerun e derruba o modo")

    def test_as_duas_listas_abrem_por_mes_com_cabecalho(self):
        """Sem cabecalho a pessoa ve quatro numeros numa linha e nao sabe qual
        e de qual mes."""
        bloco = self._bloco_tv()
        # Contar ocorrencia exata e fragil -- muda quando eu acrescento um uso
        # legitimo. O que importa e que as DUAS listas usem a mesma marca, e
        # nao cada uma a sua regra.
        self.assertGreaterEqual(bloco.count("_abrir_por_mes_tv"), 4)
        i_custo = bloco.index("linhas_custo = [")
        i_ranque = bloco.index("linhas_despop = [")
        self.assertIn("_abrir_por_mes_tv", bloco[i_custo:i_ranque],
                      "a lista de composicao nao respeita o tipo de visao")
        self.assertIn("_abrir_por_mes_tv", bloco[i_ranque:i_ranque + 3000],
                      "o ranque de grupos nao respeita o tipo de visao")
        self.assertIn("_cab_grp", bloco)


    def test_o_seletor_de_meses_cabe_num_botao(self):
        """Num painel de parede o filtro e o que menos importa: ele tem de
        caber num botao e sair da frente. A primeira versao era um multiselect
        aberto -- com oito meses, as pilulas cobriam os cartoes de KPI."""
        bloco = self._bloco_tv()
        self.assertIn("with st.popover(", bloco)
        self.assertIn("_rotulo_botao", bloco, "o botao precisa mostrar o recorte resolvido")
        # E atalhos, para o caso comum nao exigir clicar mes a mes.
        self.assertIn("tv_atalho_", bloco)
        for atalho in ('"Mês"', '"Tri"', '"YTD"', '"Ano"'):
            self.assertIn(atalho, bloco, f"sumiu o atalho {atalho}")

    def test_as_faiscas_dos_cartoes_sairam(self):
        """Elas nao ficaram boas e o espaco vale mais para o subtexto. A
        informacao mensal vive na matriz, que a mostra por extenso."""
        bloco = self._bloco_tv()
        self.assertNotIn("def _tv_faisca(", bloco)
        self.assertNotIn("tv-faisca", bloco)

    def test_o_cmv_nao_zera_no_mes_a_mes(self):
        """O CMV mora numa linha sintetica "4 - " em algumas planilhas e
        "4 - Custo das Vendas" em outras. O acumulado ja tratava os dois casos;
        a leitura mes a mes nao -- e o CMV apareceu R$ 0M em TODOS os meses
        enquanto o acumulado mostrava R$ 37,7M, na mesma tela."""
        self.assertIn("def valor_da_linha_tv(", FONTE)
        bloco = self._bloco_tv()
        self.assertGreaterEqual(bloco.count("valor_da_linha_tv("), 2,
                                "os dois blocos do mes a mes tem de usar a mesma leitura")
        i = FONTE.index("def valor_da_linha_tv(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn("Custo das Vendas", corpo, "sumiu o caminho alternativo do CMV")

    def test_a_pilha_e_ordenada_e_mostra_o_total(self):
        """A base da pilha e a que se compara de relance entre os meses, e tem
        de ser a parcela que manda no custo. E sem o total escrito em cima, a
        pessoa compara alturas no olho -- que e o que uma tela de parede nao
        permite fazer com precisao."""
        bloco = self._bloco_tv()
        self.assertIn("sorted(_series_tv, key=lambda t: sum(t[2])", bloco)
        self.assertIn("_totais_mes_tv", bloco)
        self.assertIn('textposition="top center"', bloco)


    def test_os_graficos_param_no_mes_corrente(self):
        """Desenhar de janeiro a dezembro com R$ 0M de setembro em diante
        enchia meia tela de nada e ainda achatava a escala dos meses que
        existem. Conforme o ano anda, eles crescem sozinhos."""
        b = self._bloco_tv()
        self.assertIn("m_map_ate_hoje = {n: c for i, (n, c) in enumerate(m_map_tv.items())", b)
        self.assertIn("for m_nome, c in m_map_ate_hoje.items():", b)
        self.assertEqual(b.count("for m_nome, c in m_map_ate_hoje.items():"), 2,
                         "os DOIS graficos da esquerda tem de usar o recorte")
        # E o recorte nao pode zerar a tela se o mes corrente faltar na base.
        i = b.index("m_map_ate_hoje = ")
        self.assertIn("if not m_map_ate_hoje:", b[i:i + 400])

    def test_o_espaco_foi_redistribuido_para_a_direita(self):
        """Com os graficos menores, a coluna da direita -- composicao e grupos,
        que e onde estao os numeros que se le -- ganha o espaco."""
        b = self._bloco_tv()
        largura = re.search(r"cgtv1, cgtv2 = st\.columns\(\[([\d.]+), 1\]\)", b)
        self.assertIsNotNone(largura, "sumiu a divisao de colunas")
        self.assertLessEqual(float(largura.group(1)), 1.3,
                             "a esquerda voltou a comer a largura da direita")

    def test_o_cartao_do_mes_e_so_custos_e_saidas(self):
        """Sairam dali a receita liquida, o EBITDA e a variacao contra o
        orcado: o bloco se chama "Composicao de Custos & Saidas", e resultado
        nao e composicao de custo. Repetir o EBITDA disputava a atencao com o
        cartao do topo, que ja o mostra."""
        b = self._bloco_tv()
        i = b.index("def _cartao_mes(")
        corpo_bruto = b[i:b.index("_partes_mes = [", i)]
        # So o CODIGO: a docstring explica por que o EBITDA saiu, e procurar no
        # texto cru encontrava a propria explicacao. Terceira vez que caio
        # nisso -- comentario nao e comportamento.
        _dentro_doc = False
        _linhas_codigo = []
        for _l in corpo_bruto.split("\n"):
            if _l.count('"""') == 1:
                _dentro_doc = not _dentro_doc
                continue
            if _dentro_doc or _l.strip().startswith("#"):
                continue
            _linhas_codigo.append(_l)
        corpo = "\n".join(_linhas_codigo)
        self.assertNotIn("EBITDA", corpo, "o EBITDA voltou para o cartao de custos")
        self.assertNotIn("vs. orçado", corpo, "a variacao voltou para o cartao")
        self.assertNotIn("Receita líquida", corpo)
        # Fica a receita bruta, que e a referencia dos percentuais.
        self.assertIn("receita bruta {formata_m(rec_bru)}", corpo)
        # E as QUATRO parcelas, cada uma com valor, % do custo e % da receita.
        for parcela in ('"CMV"', '"Variáveis"', '"Operacionais"', '"Deprec./Amort."'):
            self.assertIn(parcela, corpo, f"sumiu a parcela {parcela}")
        # A CONTA, nao o nome da variavel: renomear nao e defeito, mas trocar
        # a divisao e.
        self.assertIn("(_v / total * 100) if total else 0", corpo,
                      "sumiu o % que a parcela representa do custo")
        self.assertIn("(_v / rec_liq * 100) if rec_liq else 0", corpo,
                      "sumiu o % que a parcela representa da receita")
        self.assertIn("Total de saídas", corpo)

    def test_o_cartao_e_um_so_para_mes_e_periodo_v2(self):
        """Antes o cartao do periodo era escrito a parte, com o HTML repetido:
        mudar o desenho exigia mexer nos dois, e eles ja tinham divergido."""
        b = self._bloco_tv()
        i = b.index("# ---- Cartões de mês (visão MÊS A MÊS)")
        cartoes = b[i:b.index("# ---- Detalhamento de Despesas Operacionais", i)]
        self.assertEqual(cartoes.count("def _cartao_mes("), 1)
        self.assertEqual(cartoes.count('class="tv-mescard"'), 1,
                         "o HTML do cartao voltou a ser escrito em dois lugares")
        # O cartao do PERIODO existe e ocupa a largura INTEIRA da ultima
        # fileira: sem ele a comparacao exigiria somar de cabeca, e sem a
        # largura total sobrava uma faixa vazia embaixo dos meses.
        self.assertIn('"Período"', cartoes)
        self.assertIn("flex:1 1 100%;background:", cartoes,
                      "o cartao do periodo perdeu a largura total")

    def test_a_barra_do_cartao_e_proporcao_do_custo(self):
        """Proporcao se compara de relance entre cartoes; valor absoluto nao --
        um mes que vendeu o dobro teria todas as barras maiores sem que a
        composicao tenha mudado."""
        b = self._bloco_tv()
        i = b.index("def _cartao_mes(")
        corpo = b[i:b.index("_partes_mes = [", i)]
        # A largura da fatia sai da MESMA conta de proporcao. Cobrar so a
        # existencia da expressao no bloco nao servia: ela aparece tambem na
        # linha de parcela, e mutilar a barra passava batido.
        i_barra = corpo.index("fatias += (")
        self.assertIn("(_v / total * 100) if total else 0", corpo[i_barra:i_barra + 200],
                      "a barra deixou de ser proporcao do custo")

    def test_o_peso_nas_saidas_op_mais_var_existe_nos_dois_modos(self):
        """Terceiro denominador, e ele responde uma pergunta que os outros dois
        nao respondem: quanto o grupo pesa em TUDO o que a operacao gasta fora
        do CMV. O % do bloco compara grupos entre si; o da receita diz se cabe
        no faturamento; este dimensiona o grupo no gasto total."""
        b = self._bloco_tv()
        # Consolidado: coluna propria na lista.
        self.assertIn("_saidas_op_var_tv = abs(desp_op_tv_kpi) + abs(desp_var_tv)", b)
        # DENTRO da linha de grupo: a linha de TOTAL tambem tem a coluna, e
        # procurar no bloco todo deixava mutilar a das linhas em silencio.
        # Ancora no desenho da linha de grupo, que vem DEPOIS do ramo do mes
        # a mes: o "if _abrir_por_mes_tv" aparece antes e o recorte parava ali.
        i_linha = b.index('<div class="tv-rank-bar-bg">')
        linha = b[i_linha:i_linha + 900]
        self.assertIn('class="tv-rank-pct-sai"', linha,
                      "a coluna de saidas sumiu das linhas de grupo")
        self.assertIn("_pct_saidas_linha", linha)
        # Mes a mes: terceiro percentual na celula, com a base do PROPRIO mes.
        self.assertIn("_saidas_mes_tv = [", b)
        i = b.index("_saidas_mes_tv = [")
        self.assertIn("[m_map_tv[_m]]", b[i:i + 400],
                      "a base tem de ser a saida do proprio mes, nao a do periodo")
        self.assertIn('_pct_sai_m = (f"{_v_m / _base_sai * 100:.0f}% saí"', b)


    def test_o_detalhe_mostra_o_pct_do_grupo_em_cada_mes(self):
        """A pergunta na linha recuada e "quanto de Servicos de Terceiros foi
        transporte em marco". O % da receita ja aparece nas linhas de grupo."""
        b = self._bloco_tv()
        self.assertIn("_linha_pai_grp", b, "a linha do pai precisa viajar junto")
        self.assertIn('_primeiro = (f"{_v_m / _base_pai * 100:.0f}% grupo"', b)
        # E a linha recuada tambem mostra o % da RECEITA -- ela era a unica
        # que nao trazia, e a comparacao com as outras linhas ficava capenga.
        # O % da receita e calculado ANTES do if do detalhe, entao vale para
        # as duas -- e essa ordem que garante que a linha recuada tenha o
        # numero. Calcular dentro do else deixaria ela sem, de novo.
        i_sai = b.index("_pct_sai_m = ")
        i_rec = b.index("_pct_rec_m = ", i_sai)
        i_if = b.index("if eh_detalhe:", i_sai)
        self.assertLess(i_rec, i_if,
                        "o % da receita voltou para dentro do ramo e a linha "
                        "recuada ficou sem ele")


    def test_o_nome_do_detalhe_tem_fonte_propria(self):
        """A linha de detalhe ja entra recuada, entao perde largura logo de
        saida, e era o unico nome que nao cabia inteiro. Menor e melhor que
        reticencias: a conta chama "Servicos de Transporte", nao
        "Servicos de Trans..."."""
        b = self._bloco_tv()
        self.assertIn(".tv-matriz td.rot.detalhe {{ font-size:10.5px", b)
        self.assertIn('_classe_rot = "rot detalhe" if eh_detalhe else "rot"', b)


    def test_a_coluna_do_grupo_fica_travada_na_rolagem(self):
        """Com 12 meses a tabela rola para o lado e o nome da linha saia da
        vista junto: sobravam seis colunas de numeros sem dizer de que conta
        eles eram.

        Fundo OPACO e z-index sao obrigatorios e nao obvios: sem o fundo, as
        celulas que rolam aparecem por baixo da travada; sem o z-index, e a
        travada que passa por baixo."""
        b = self._bloco_tv()
        self.assertIn('class="tv-matriz tv-matriz-fixa"', b)
        # O CSS vive no bloco de estilo, FORA da funcao do painel: procurar
        # no corpo da funcao nao encontrava a regra.
        self.assertIn(".tv-matriz-fixa td.rot, .tv-matriz-fixa th.rot {{", FONTE)
        i = FONTE.index(".tv-matriz-fixa td.rot")
        # Janela LARGA: os comentarios que explicam cada linha ficam entre a
        # ancora e as declaracoes, e uma janela curta parava no meio deles.
        # Ja errei nisso tres vezes hoje -- comentario ocupa espaco no recorte.
        regra = FONTE[i:i + 1400]
        self.assertIn("position:sticky", regra)
        self.assertIn("left:0", regra)
        self.assertIn("z-index:20", regra)
        # Cor CRAVADA e opaca, nao a variavel do tema: a celula que rola por
        # baixo tem fundo ambar SEMITRANSPARENTE, e qualquer transparencia
        # aqui deixa o ambar aparecer atraves -- foi o vazamento visto no
        # print de 27/08/2026.
        self.assertIn("background-color:#242C3C", regra,
                      "sem fundo opaco cravado o ambar vaza pela coluna travada")
        self.assertIn("background-clip:padding-box", regra,
                      "sem isto sobra uma faixa de um pixel por onde o ambar passa")
        # Separador de VERDADE: sombra e desenhada FORA da celula e nao impede
        # nada de passar por baixo.
        self.assertIn("border-right:1px solid", regra)
        # E a celula de calor nao pode subir na pilha.
        self.assertIn(".tv-matriz-fixa td.calor {{ position:relative; z-index:1; }}",
                      FONTE, "a celula de calor pode voltar a passar por cima")

    def test_so_a_tabela_de_grupos_trava_a_coluna(self):
        """A tabela do detalhamento de despesas usa a MESMA classe tv-matriz e
        nao deve travar nada -- por isso a trava mora numa classe propria."""
        self.assertEqual(FONTE.count('class="tv-matriz tv-matriz-fixa"'), 1)
        # A do detalhamento continua sem a classe.
        i = FONTE.index("def _corpo_despesas_tv(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn('class="tv-matriz"', corpo)
        self.assertNotIn("tv-matriz-fixa", corpo,
                         "a trava vazou para a tabela do detalhamento")


    def test_a_tabela_mensal_mostra_o_peso_do_grupo(self):
        """No mes a mes a linha do detalhe nao dizia quanto ela representa do
        grupo -- a informacao existia so no consolidado."""
        b = self._bloco_tv()
        # A coluna PESO saiu: com 12 meses eram 15 colunas e a tabela
        # transbordava POR CIMA da coluna vizinha. O peso mora nas celulas.
        self.assertNotIn('<th class="tot">Peso</th>', b,
                         "a coluna Peso voltou e a tabela transborda de novo")
        self.assertIn('_primeiro_t = f"{_peso_t:.0f}% grupo"', b)
        self.assertIn('_primeiro_t = f"{_peso_t:.0f}% op"', b)
        self.assertIn('_pct_tot = f"{_primeiro_t} · {_pct_sai_t} · {_pct_tot}"', b)
        # E a tabela ganha rolagem propria, para nunca escrever por cima do
        # bloco vizinho.
        self.assertIn("'<div style=\"overflow-x:auto;\">'", b)

    def test_o_titulo_dos_grupos_aparece_uma_vez_so(self):
        """O bloco muda de coluna conforme o modo, e o titulo viaja junto. O
        titulo antigo ficou para tras e a tela mostrava os dois."""
        # O titulo virou LINK, entao ele nao fecha mais com </div> logo
        # depois do texto. Conta o TEXTO, que e o que nao pode aparecer duas
        # vezes na tela.
        self.assertEqual(FONTE.count("🏢 Despesas Operacionais — Principais Grupos"), 1)


    def test_os_grupos_viram_mapa_de_calor(self):
        """Ler doze numeros e achar o maior e trabalho; ver a celula mais
        escura nao e -- e num painel de parede ninguem compara digitos."""
        b = self._bloco_tv()
        self.assertIn('class="calor"', b)
        self.assertIn("_teto_calor", b)
        # Piso na intensidade: transparente demais some no fundo.
        self.assertIn("0.06 + (_pct_num / _teto_calor)", b)


    def test_o_grafico_compara_realizado_e_orcado(self):
        """Antes mostrava SO o desvio: uma barra vermelha de R$ -0,8M sem dizer
        se era 0,8 sobre 1,5 (metade da meta perdida) ou 0,8 sobre 20 (um
        arranhao). Com as duas barras lado a lado e os dois valores escritos, a
        diferenca se le pela distancia entre elas -- e por isso a LINHA de
        desvio saiu: ela so cruzava o desenho por cima das duas."""
        b = self._bloco_tv()
        i = b.index("eb_real_m_tv, eb_orc_m_tv, desvio_m_tv")
        trecho = b[i:b.index("st.plotly_chart(fig_tv_desvio", i)]
        self.assertIn('name="Orçado"', trecho)
        self.assertIn('name="Realizado"', trecho)
        self.assertIn('barmode="group"', trecho, "as barras voltaram a se sobrepor")
        self.assertNotIn('name="Desvio"', trecho, "a linha de desvio voltou")
        self.assertNotIn("yaxis2=", trecho,
                         "sem a linha de desvio nao ha o que por no eixo da direita")
        # O ORCADO usa o AZUL da paleta: o cinza de antes nao era cor da casa,
        # e a barra parecia de outro painel.
        i_orc = trecho.index('name="Orçado"')
        self.assertIn('line=dict(color=COLORS["primary"], width=1.6)',
                      trecho[i_orc:i_orc + 800],
                      "o orcado deixou de usar o azul da paleta")
        self.assertIn("text=[formata_m(v) for v in eb_orc_m_tv]", trecho,
                      "o orcado precisa mostrar o valor: e por ele que se compara")
        # O REALIZADO sai verde ou vermelho conforme bate a meta -- mesma
        # convencao dos cartoes de KPI e da tabela de reserva.
        self.assertIn("_preenche_real = [", trecho)
        self.assertIn("line=dict(color=cores_desvio_tv, width=1.6)", trecho)
        # Legenda EMBAIXO, como no grafico da Reserva de Caixa.
        self.assertIn('yanchor="top", y=-0.16', trecho)

    def test_o_peso_do_grupo_aparece_em_cada_mes(self):
        """O peso so existia na ultima coluna, para o periodo inteiro -- e
        periodo inteiro nao responde "em qual mes essa conta saiu da curva",
        que e a pergunta que se faz olhando uma tabela mensal."""
        b = self._bloco_tv()
        self.assertIn("_despop_mes = [valor_da_linha_tv(", b)
        self.assertIn('_primeiro = (f"{_v_m / _base_bloco * 100:.0f}% op"', b)
        # E ele tem de CHEGAR na celula: calcular e nao usar deixava a trava
        # verde com a coluna mostrando so o % da receita.
        self.assertIn('_pct_m = f"{_primeiro} · {_pct_sai_m} · {_pct_rec_m}"', b,
                      "o peso do bloco foi calculado mas nao chega na celula")
        # A base e a despesa operacional DAQUELE mes, nao a do periodo: senao
        # janeiro sairia medido contra oito meses de despesa.
        i = b.index("_despop_mes = [valor_da_linha_tv(")
        self.assertIn("[m_map_tv[_m]]", b[i:i + 300])

    def test_as_parcelas_do_cartao_nao_quebram_linha(self):
        """O nome empurrava "R$ 0,1M" para a linha de baixo, e o cartao ficava
        com uma parcela alta e as outras baixas."""
        b = self._bloco_tv()
        i = b.index(".tv-mescard .parc em {{")
        self.assertIn("white-space:nowrap", b[i:i + 300])
        self.assertIn("min-width:56px", b[i:i + 300])


    def test_o_grafico_de_desvio_ocupa_a_lacuna(self):
        """A coluna da direita (rosca + as duas listas) e mais alta que a
        esquerda, e sobrava uma faixa morta embaixo do desvio ate o letreiro.
        O teto existe para a altura nao crescer a ponto de empurrar o rodape:
        letreiro e botao de tela cheia tem de continuar visiveis sem rolagem,
        que e o ponto de um painel de parede."""
        bloco = self._bloco_tv()
        i = bloco.index("fig_tv_desvio, height=")
        altura = int(re.search(r"height=(\d+)", bloco[i:i + 40]).group(1))
        self.assertGreaterEqual(altura, 180, "voltou a sobrar lacuna embaixo")
        self.assertLessEqual(altura, 300, "alto demais: empurra o letreiro para fora")

    def test_a_linha_de_transporte_esta_configurada(self):
        ns = carregar([], ["DETALHES_DO_RANQUE_TV"])
        self.assertEqual(ns["DETALHES_DO_RANQUE_TV"].get("8.8.10"),
                         "Serviços de Transporte")

    def test_o_detalhe_so_pendura_no_pai_certo(self):
        """Modelo do que o codigo faz: 8.8.10 pendura em 8.8, nunca em 8.5 nem
        em 8 -- e a comparacao usa o PONTO, senao "8.8" casaria com "8.80"."""
        detalhes = {"8.8.10": "Serviços de Transporte"}
        for pai, esperado in (("8.8", True), ("8.5", False), ("8", False), ("8.80", False)):
            achou = any(n.rsplit(".", 1)[0] == pai for n in detalhes)
            self.assertEqual(achou, esperado, pai)
        # E o codigo tem de usar a comparacao exata, nao startswith: "8.8.10"
        # comeca com "8." e se penduraria no grupo "8" tambem.
        self.assertIn('_num_det.rsplit(".", 1)[0] != _num_grp', FONTE)



class TesteMemoriaDoPainel(unittest.TestCase):
    """Travas de memoria (27/08/2026).

    O app chegou a 1039 MB em producao e os caches foram esvaziados sozinhos
    para ele nao cair. O servidor derruba o processo quando a memoria estoura,
    e o que aparece na tela e um erro generico, sem causa -- por isso cada
    decisao daqui vira trava: elas sao invisiveis e voltam sozinhas na
    proxima vez que alguem "melhorar" um carregamento."""

    def test_as_colunas_repetitivas_do_diario_sao_categoria(self):
        """Medido numa DIARIO com a forma da real: o Historico sozinho ocupava
        12,9 MB como texto e cai para menos de 1 MB como categoria, porque o
        mesmo texto se repete em centenas de lancamentos. O conjunto economiza
        cerca de 64% da tabela."""
        i = FONTE.index("def carregar_diario(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        for coluna in ('"Plano de Contas", "Centro de Custos", "Linha DRE"',):
            self.assertIn(coluna, corpo)
        self.assertIn('df[_coluna_repetitiva].astype(str).str.strip().astype("category")', corpo)
        self.assertIn('df[col_extra] = df[col_extra].astype("category")', corpo,
                      "as colunas extras voltaram a ser texto puro")
        self.assertIn('mes_formatado.astype("category")', corpo,
                      "a coluna Mes voltou a ser texto")

    def test_numero_fica_como_texto_de_proposito(self):
        """Categoria com uma categoria POR LINHA gasta mais que o texto puro:
        guarda o dicionario inteiro e ainda um indice para cada linha."""
        i = FONTE.index("def carregar_diario(")
        corpo = FONTE[i:FONTE.index("\ndef ", i + 10)]
        self.assertIn('if col_extra != "Número":', corpo)

    def test_categoria_nao_quebra_o_que_o_app_faz_com_a_coluna(self):
        """Modelo: comparacao, isin e groupby precisam continuar iguais. O
        groupby com observed=False geraria todas as combinacoes possiveis de
        categorias, mesmo as que nao existem -- explosao combinatoria."""
        df = pd.DataFrame({
            "Mês": pd.Categorical(["01/2026"] * 3 + ["02/2026"] * 2),
            "Centro de Custos": pd.Categorical(["A", "A", "B", "B", "A"]),
            "Valor Bruto": [1.0, 2, 3, 4, 5],
        })
        self.assertEqual(df.loc[df["Mês"] == "01/2026", "Valor Bruto"].sum(), 6.0)
        self.assertEqual(df[df["Mês"].isin(["01/2026"])]["Valor Bruto"].sum(), 6.0)
        # Filtra para sobrarem categorias sem uso -- e nelas que a diferenca
        # aparece. O padrao do pandas para `observed` MUDOU entre 2.x e 3.x,
        # entao o app passa o argumento EXPLICITO: sem ele o resultado depende
        # da versao instalada no servidor.
        so_um = df[df["Centro de Custos"] == "A"]
        com = so_um.groupby(["Centro de Custos", "Mês"], observed=True)["Valor Bruto"].sum()
        sem = so_um.groupby(["Centro de Custos", "Mês"], observed=False)["Valor Bruto"].sum()
        self.assertLess(len(com), len(sem),
                        "observed=True e o que evita a explosao combinatoria")

    def test_o_groupby_do_diario_usa_observed(self):
        """Sem observed=True, agrupar tres colunas de categoria geraria o
        produto cartesiano das categorias -- centenas de milhares de linhas
        vazias so para serem descartadas depois."""
        # Varre com find em cadeia, e nao com um laco sobre CADA caractere do
        # arquivo: o `for i in range(len(FONTE))` fazia 900 mil iteracoes e
        # sozinho triplicou o tempo da suite inteira.
        pos, faltando = 0, []
        while True:
            pos = FONTE.find(".groupby(", pos)
            if pos < 0:
                break
            trecho = FONTE[pos:pos + 300]
            if (("Plano de Contas" in trecho or "Centro de Custos" in trecho)
                    and "observed=True" not in trecho):
                faltando.append(trecho.split("\n")[0][:90])
            pos += 9
        self.assertEqual(faltando, [], "groupby de categoria sem observed=True")

    def test_os_caches_grandes_guardam_uma_entrada_so(self):
        """Cada entrada guarda as 21 abas ou a base inteira do fluxo. A segunda
        so serviria se a pessoa trocasse de visao e voltasse dentro do prazo do
        cache, e custava o dobro da memoria o tempo todo."""
        for funcao in ("carregar_dados_abas", "carregar_dados_por_loja",
                       "carregar_diario", "preparar_fluxo_caixa"):
            i = FONTE.index(f"def {funcao}(")
            # SO o decorador imediatamente acima: uma janela de 700 caracteres
            # alcancava o decorador de OUTRA funcao e passava mesmo com este
            # errado. O recorte vai do ultimo "@st.cache" ate a definicao.
            ini = FONTE.rindex("@st.cache", 0, i)
            decorador = FONTE[ini:i]
            self.assertIn("max_entries=1", decorador,
                          f"{funcao} voltou a guardar mais de uma entrada")

    def test_a_base_do_fluxo_nao_e_copiada_inteira(self):
        """Era uma copia da base inteira -- centenas de milhares de linhas --
        que a primeira filtragem logo abaixo ja substituia por um recorte novo.
        A copia so existia para ser jogada fora, e enquanto isso dobrava a
        memoria do fluxo."""
        self.assertNotIn("df_d_completo = df_fin.copy()", FONTE)
        self.assertIn("df_d_completo = df_fin\n", FONTE)
        # E a copia passou a ser feita DEPOIS da filtragem, so no recorte.
        i = FONTE.index("df_d_completo = df_fin\n")
        self.assertIn('!= "aplicacao"].copy()', FONTE[i:i + 1400],
                      "sem a copia no recorte, criar DiaOrd tocaria a base original")



class TesteRamificacaoDeDespesas(unittest.TestCase):
    """A tela que abre variaveis e operacionais ate o plano de contas."""

    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(
            ["arvore_de_despesas", "ofensores_por_desvio", "ofensores_por_salto",
             "_subgrupos_nivel2", "_filhos_diretos_do_conjunto",
             "_linha_pertence_ao_grupo",
             "_numero_linha_dre", "_nome_sem_numero_dre"])

    LINHAS = ["8 - Despesas Operacionais", "8.3 - Pessoal", "8.3.1 - Salários",
              "8.3.2 - Encargos", "8.8 - Serviços de Terceiros",
              "8.8.10 - Serviços de Transporte", "6 - Despesas Variáveis",
              "6.2 - Taxa com Cartão"]

    def test_a_arvore_desce_ate_as_sublinhas(self):
        valores = {"8.3 - Pessoal": 100.0, "8.3.1 - Salários": 70.0,
                   "8.3.2 - Encargos": 30.0, "8.8 - Serviços de Terceiros": 40.0,
                   "8.8.10 - Serviços de Transporte": 40.0}
        arvore = self.ns["arvore_de_despesas"](
            self.LINHAS, "8", lambda l: valores.get(l, 0.0))
        self.assertEqual([g["nome"] for g in arvore],
                         ["Pessoal", "Serviços de Terceiros"],
                         "os subgrupos vem do maior para o menor")
        self.assertEqual([f["nome"] for f in arvore[0]["filhos"]],
                         ["Salários", "Encargos"])

    def test_a_arvore_avisa_quando_o_detalhe_nao_fecha(self):
        """Subgrupo cujo detalhe nao soma com ele tem lancamento direto no pai,
        e mostrar a lista incompleta faria a pessoa procurar um dinheiro que a
        tela escondeu."""
        valores = {"8.3 - Pessoal": 100.0, "8.3.1 - Salários": 70.0}
        arvore = self.ns["arvore_de_despesas"](
            self.LINHAS, "8", lambda l: valores.get(l, 0.0))
        pessoal = next(g for g in arvore if g["nome"] == "Pessoal")
        self.assertAlmostEqual(pessoal["cobertura"], 0.70, places=6)
        self.assertLess(pessoal["cobertura"], 0.99, "tem de acender o aviso")

    def test_o_valor_vem_de_fora_da_arvore(self):
        """A arvore nao sabe de onde o numero vem -- realizado, orcado ou um mes
        so. E a mesma funcao para as tres leituras que a tela faz."""
        chamadas = []
        self.ns["arvore_de_despesas"](self.LINHAS, "6",
                                      lambda l: chamadas.append(l) or 1.0)
        self.assertIn("6.2 - Taxa com Cartão", chamadas)

    def test_os_estouros_saem_em_REAIS(self):
        """Uma conta de R$ 500 que gastou o dobro aparece como 100% de estouro e
        nao muda nada no mes; uma de R$ 200 mil que passou 12% e a que precisa
        de conversa. Percentual sozinho no topo de uma lista engana."""
        itens = [
            {"conta": "pequena dobrou", "realizado": 1000.0, "orcado": 500.0},
            {"conta": "grande passou pouco", "realizado": 224_000.0, "orcado": 200_000.0},
            {"conta": "dentro do orçado", "realizado": 90.0, "orcado": 100.0},
        ]
        saida = self.ns["ofensores_por_desvio"](itens)
        self.assertEqual([o["conta"] for o in saida],
                         ["grande passou pouco", "pequena dobrou"])
        self.assertAlmostEqual(saida[0]["desvio"], 24_000.0)

    def test_conta_dentro_do_orcado_nao_e_ofensor(self):
        self.assertEqual(self.ns["ofensores_por_desvio"](
            [{"conta": "ok", "realizado": 90.0, "orcado": 100.0}]), [])
        self.assertEqual(self.ns["ofensores_por_desvio"]([]), [])

    def test_o_salto_precisa_de_tres_meses(self):
        """Com dois, a "media dos anteriores" e um mes so, e qualquer oscilacao
        normal vira alarme."""
        # Serie que SUBIU no segundo mes: com dois meses fechados a "media dos
        # anteriores" e um mes so, e essa conta viraria alarme. Precisa sair
        # vazia mesmo tendo subido.
        serie = {"conta": [30.0, 90.0, 0.0, 0.0]}
        self.assertEqual(self.ns["ofensores_por_salto"](serie, 2), [],
                         "com dois meses qualquer oscilacao normal vira alarme")
        self.assertNotEqual(self.ns["ofensores_por_salto"](
            {"conta": [30.0, 30.0, 90.0]}, 3), [])

    def test_o_salto_compara_com_a_propria_media(self):
        # "quase parada" sobe 2%: ela ENTRA na lista (subiu), mas fica atras.
        # Excluir quem subiu pouco exigiria um corte arbitrario, e um corte
        # arbitrario esconde justamente a conta que comecou a escorregar.
        saida = self.ns["ofensores_por_salto"](
            {"triplicou": [30.0, 30.0, 90.0], "quase parada": [50.0, 50.0, 51.0]}, 3)
        self.assertEqual(saida[0]["conta"], "triplicou")
        self.assertEqual([o["conta"] for o in saida],
                         ["triplicou", "quase parada"])
        self.assertAlmostEqual(saida[0]["salto"], 2.0, places=6)
        self.assertAlmostEqual(saida[0]["diferenca"], 60.0, places=6)

    def test_o_salto_tambem_ordena_por_reais(self):
        """Triplicar R$ 300 nao e noticia; subir 20% numa conta de milhoes e."""
        saida = self.ns["ofensores_por_salto"]({
            "pequena que triplicou": [100.0, 100.0, 300.0],
            "grande que subiu pouco": [100_000.0, 100_000.0, 130_000.0],
        }, 3)
        self.assertEqual(saida[0]["conta"], "grande que subiu pouco")

    def test_conta_que_caiu_nao_entra(self):
        self.assertEqual(self.ns["ofensores_por_salto"](
            {"caiu": [90.0, 90.0, 30.0]}, 3), [])
        self.assertEqual(self.ns["ofensores_por_salto"](
            {"do zero": [0.0, 0.0, 50.0]}, 3), [],
            "media zero daria salto infinito")


class TesteTelaDaRamificacao(unittest.TestCase):
    """Travas do roteamento e do corpo da ramificacao."""

    def test_o_titulo_dos_grupos_leva_a_ramificacao(self):
        """Botao solto ao lado ocuparia espaco numa tela que ja disputa cada
        pixel, e o titulo ja diz do que a outra tela trata."""
        self.assertIn('href="?modo=tv&foco=despesas"', FONTE)
        self.assertIn("abrir detalhe ›", FONTE)
        self.assertIn('href="?modo=tv"', FONTE, "sumiu o caminho de volta")

    def test_a_ramificacao_reusa_o_cabecalho_e_o_rodape(self):
        """Duplicar o cabecalho numa funcao nova seria manter dois filtros que
        precisam concordar para sempre -- e eles nao concordariam por muito
        tempo."""
        i = FONTE.index("def renderizar_painel_tv(")
        corpo = FONTE[i:FONTE.index("\nif ", i)]
        self.assertIn('if foco == "despesas":', corpo)
        # E a funcao tem de ser CHAMADA de verdade: comentar a chamada deixava
        # a tela abrir vazia, com cabecalho e rodape e nada no meio.
        i_foco = corpo.index('if foco == "despesas":')
        chamada = corpo[i_foco:corpo.index("else:", i_foco)]
        self.assertIn("_corpo_despesas_tv(", chamada)
        self.assertNotIn("pass", chamada, "a chamada foi desligada")
        # A bifurcacao vem DEPOIS dos filtros e ANTES do letreiro.
        self.assertLess(corpo.index("meses_ativos_tv = "),
                        corpo.index('if foco == "despesas":'))
        self.assertLess(corpo.index('if foco == "despesas":'),
                        corpo.index("# ---------------- Ticker de destaques"))

    def test_o_foco_vem_da_url(self):
        self.assertIn('foco=str(st.query_params.get("foco") or "geral")', FONTE)

    def test_a_ramificacao_cobre_os_dois_grupos(self):
        self.assertIn('("6", "6 - Despesas Variáveis")', FONTE)
        self.assertIn('("8", "8 - Despesas Operacionais")', FONTE)


# ============================================================================
# 6. FORMATACAO
# ============================================================================
class TesteFormatacao(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.ns = carregar(["formata_brl", "formata_valor_curto", "formata_sigma_curto"])

    def test_reais_no_padrao_brasileiro(self):
        self.assertEqual(self.ns["formata_brl"](1234567.89), "R$ 1.234.567,89")

    def test_valor_curto_muda_de_escala(self):
        self.assertEqual(self.ns["formata_valor_curto"](1_200_000), "R$ 1,2M")
        self.assertEqual(self.ns["formata_valor_curto"](12_000), "R$ 12 mil")
        self.assertEqual(self.ns["formata_valor_curto"](480), "R$ 480")

    def test_sigma_tem_teto(self):
        self.assertEqual(self.ns["formata_sigma_curto"](25), "+10σ+")
        # o sinal negativo aqui e o MENOS tipografico (U+2212), nao o hifen
        self.assertEqual(self.ns["formata_sigma_curto"](-25), "\u221210\u03c3+")
        self.assertEqual(self.ns["formata_sigma_curto"](5.4), "+5,4σ")


# ============================================================================
# 7. TRAVAS ESTRUTURAIS - as regressoes que ja aconteceram
# ============================================================================
class TesteTravasEstruturais(unittest.TestCase):
    """Estes testes nao olham resultado: olham o codigo. Cada um trava uma
    decisao que ja foi desfeita sem querer e custou horas para achar."""

    def test_login_falha_fechado(self):
        # Olha SO o bloco do "if not usuarios": e ali que estava o return True
        # que abria o painel quando a leitura dos Secrets falhava.
        i = FONTE.index("def checar_login(")
        trecho = FONTE[i:i + 6000]
        inicio = trecho.index("if not usuarios:")
        # o bloco termina na proxima linha com 4 espacos de recuo
        resto = trecho[inicio:]
        fim = len(resto)
        for linha_inicio in range(1, len(resto)):
            if resto[linha_inicio - 1] == "\n" and resto[linha_inicio:linha_inicio + 5] == "    i":
                fim = linha_inicio
                break
        bloco = resto[:fim]
        self.assertNotIn("return True", bloco,
                         "checar_login voltou a liberar acesso quando nao ha usuarios")
        self.assertIn("return False", bloco,
                      "o bloco de 'sem usuarios' precisa barrar o acesso")

    def test_colunas_do_csv_usam_a_ordem_detectada(self):
        i = FONTE.index("df[COL_FIN_DATA_LIQUIDACAO] = pd.to_datetime(")
        self.assertIn("dayfirst=dayfirst_liq", FONTE[i:i + 300])
        j = FONTE.index("df[COL_FIN_VENCIMENTO] = pd.to_datetime(")
        self.assertIn("dayfirst=dayfirst_venc", FONTE[j:j + 300])

    def test_baixa_da_diario_nao_entra_na_coluna_do_csv(self):
        self.assertNotIn("df[COL_FIN_DATA_LIQUIDACAO] = df[COL_FIN_LIQ_DIARIO]", FONTE)

    def test_caches_de_dados_tem_validade(self):
        for funcao in ("obter_caminhos_excel", "obter_dados_fluxo_caixa",
                       "preparar_fluxo_caixa"):
            i = FONTE.index(f"def {funcao}(")
            decorador = FONTE[max(0, i - 260):i]
            self.assertIn("ttl=", decorador, f"{funcao} ficou sem ttl no cache")

    def test_fluxo_tenta_mais_de_uma_fonte(self):
        i = FONTE.index("def fontes_csv_fluxo(")
        trecho = FONTE[i:i + 2000]
        self.assertIn("pub?output=csv", trecho)
        self.assertIn("FLUXO_CAIXA_FILE_ID", trecho)
        # A validacao evita aceitar em silencio uma fonte que devolveu
        # outra aba - o pior desfecho possivel: numero errado sem aviso.
        i_leitura = FONTE.index("def obter_dados_fluxo_caixa(")
        leitura = FONTE[i_leitura:i_leitura + 2200]
        self.assertIn("COLUNAS_MINIMAS_FLUXO", leitura,
                      "a leitura do fluxo parou de conferir as colunas da fonte")
        self.assertIn("veio outra aba", leitura)

    def test_chamadas_depreciadas_do_streamlit_nao_voltam(self):
        self.assertNotIn("use_container_width", FONTE,
                         "use_container_width foi removido do Streamlit em 31/12/2025")
        self.assertLessEqual(
            FONTE.count("components.html("), 2,
            "components.html so pode aparecer dentro de html_embutido")

    def test_orcado_por_canal_desce_para_as_linhas_filhas(self):
        """Em VD CONSOLIDADO o orcamento de marketing esta lancado nas linhas
        filhas, nao na 6.24.2. Ler so a mae devolvia zero e o canal aparecia
        sem orcamento - com o desvio inteiro contado como estouro."""
        i = FONTE.index("def _painel_dept_mkt(")
        trecho = FONTE[i:i + 3500]
        self.assertIn("_somar_codigo_com_filhas(df_o, CODIGO_MKT_GESTAO_CP", trecho,
                      "o orcado por canal voltou a ler so a linha-mae")
        self.assertIn("_somar_codigo_com_filhas(df_r, CODIGO_MKT_GESTAO_CP", trecho)

    def test_tabela_por_canal_fecha_com_o_cabecalho(self):
        """Somar a coluna da tabela por canal dava um numero diferente do
        cartao 'Desvio vs. Orcado' do topo - a diferenca era o que sai do 1%,
        descontado do investido da Loja mas presente no total do departamento.
        A linha 'Fora do 1%' e a linha TOTAL fazem os dois blocos conciliarem."""
        i = FONTE.index("def _painel_dept_mkt(")
        trecho = FONTE[i:i + 5000]
        self.assertIn('"Canal": "Fora do 1% (Loja)"', trecho,
                      "sumiu a linha que concilia o que sai do 1%")
        self.assertIn('"Canal": "TOTAL"', trecho, "sumiu a linha de total da tabela por canal")

    def test_nomes_de_parametro_nao_vazam_entre_funcoes(self):
        """Uma troca de nome em massa ja renomeou height=/width= para
        altura=/largura= em 29 chamadas de grafico e tabela, que esperam os
        nomes em ingles - o painel so quebrava ao abrir a aba afetada."""
        arvore = ast.parse(FONTE)
        # Funcoes escritas por nos, cujos parametros sao em portugues.
        FUNCOES_EM_PORTUGUES = {"html_embutido", "tabela_selecionavel"}

        def nome(no):
            f = no.func
            return f.id if isinstance(f, ast.Name) else (f.attr if isinstance(f, ast.Attribute) else "")

        errados = []
        for no in ast.walk(arvore):
            if not isinstance(no, ast.Call):
                continue
            args = [kw.arg for kw in no.keywords]
            if nome(no) in FUNCOES_EM_PORTUGUES:
                # Funcoes nossas: aqui o portugues e o certo, e o ingles e
                # que seria o erro.
                if "height" in args or "width" in args:
                    errados.append(f"linha {no.lineno}: {nome(no)} com height/width")
            elif "altura" in args or "largura" in args:
                errados.append(f"linha {no.lineno}: {nome(no)}(...) com altura/largura")
        self.assertEqual(errados, [], "argumentos com o nome trocado: " + "; ".join(errados[:5]))

    def test_regra_do_1pct_usa_o_orcado_do_canal_loja(self):
        i = FONTE.index("def _painel_dept_mkt(")
        # a janela cobre a funcao inteira: ela ja cresceu duas vezes
        fim = FONTE.index("\ndef ", i + 10)
        trecho = FONTE[i:fim]
        self.assertIn('teto_loja = loja["Orçado (R$)"]', trecho,
                      "o teto do 1% voltou a ser calculado sobre a receita")


if __name__ == "__main__":
    print(f"Testando: {CAMINHO_APP}\n" + "=" * 62)
    unittest.main(argv=[sys.argv[0]], verbosity=2, exit=True)