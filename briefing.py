# -*- coding: utf-8 -*-
"""Briefing executivo das 8h -- roda no GitHub Actions e envia por e-mail.

Reaproveita as funções do app.py SEM importar o Streamlit: extrai o fonte de
cada função e constante por AST (o mesmo truque da suíte de testes), fecha as
dependências automaticamente e executa tudo num espaço onde `st` é um objeto
de mentira que aceita qualquer chamada e não faz nada. Resultado: a tela e o
e-mail contam a MESMA história, porque saem das MESMAS funções -- nada foi
duplicado, e quando o app mudar a regra, o e-mail muda junto.

Uso:
    python briefing.py                    calcula e envia (precisa das variáveis abaixo)
    python briefing.py --forcar           reenvia mesmo que o de hoje já tenha saído
    python briefing.py --teste            calcula e imprime o e-mail em texto, sem enviar
    python briefing.py --salvar b.html    grava o HTML do e-mail para conferir no navegador

Variáveis de ambiente (no GitHub ficam em Settings > Secrets and variables > Actions):
    SMTP_USUARIO        e-mail que envia (Gmail)
    SMTP_SENHA          SENHA DE APP do Gmail (não é a senha da conta)
    EMAIL_DESTINO       quem recebe; vários separados por vírgula
    FECHAMENTO_CSV_URL  link da planilha de Fechamento Mensal (o mesmo do app; opcional)
    LINK_PAINEL         link do painel para o rodapé do e-mail (opcional)
    URL_ORCADO, URL_REALIZADO, ABA_DRE, SMTP_SERVIDOR, SMTP_PORTA   opcionais (têm padrão)
"""
import ast
import base64
import contextlib
import gc
import hashlib
import hmac
import io
import json
import math
import os
import re
import smtplib
import ssl
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from email.message import EmailMessage
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

CAMINHO_APP = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.py")

# Os mesmos endereços do obter_caminhos_excel do app; podem ser trocados por
# variável de ambiente sem mexer no código.
URL_ORCADO_PADRAO = "https://docs.google.com/spreadsheets/d/1x68Eg_6LlSKeFJEGmfhyBfcGgheSrVsl/export?format=xlsx"
URL_REALIZADO_PADRAO = "https://docs.google.com/spreadsheets/d/12I0vGpYU_KNhGxAHOMHWAQu3Xkz_EsUZ/export?format=xlsx"

# O que o briefing precisa do app. As dependências de cada uma (funções e
# constantes que elas chamam) são descobertas sozinhas pela árvore do fonte.
SEMENTES = [
    "carregar_dados_abas", "get_valor_consolidado_multi",
    "fatos_do_ritmo", "montar_fatos_executivos", "montar_narrativa_executiva",
    "narrativa_em_texto", "projetar_margem_fechamento",
    "_deltas_pendentes_do_fechamento", "url_csv_do_fechamento",
    "carregar_planilha_fechamento", "formata_valor_curto", "_pct_br",
    "DIAS_DEFASAGEM_DADOS", "FUSO_BR", "LOGO_BEEA_B64", "MODELOS_RELATORIO",
    "MAPA_EMAIL_DEPARTAMENTO", "EMAILS_TRAVADOS_NO_DEPARTAMENTO", "_subgrupos_nivel2",
    "ofensores_por_desvio", "_nome_sem_numero_dre",
]

NOMES_MESES = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
               "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
DIAS_SEMANA = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
               "sexta-feira", "sábado", "domingo"]


class _StreamlitDeMentira:
    """Aceita qualquer atributo, chamada, índice ou iteração e não faz nada.

    Serve para o `st` (e o `components`, e o `go`) dentro das funções
    extraídas: um `@st.cache_data(ttl=None)` vira decorador que devolve a
    própria função; um `st.warning(...)` dentro de um loader vira nada. Só o
    que o briefing PRECISA acontece -- ler planilha e fazer conta."""

    def __getattr__(self, nome):
        return self

    def __call__(self, *args, **kwargs):
        if len(args) == 1 and not kwargs and callable(args[0]):
            return args[0]          # uso como decorador: devolve a função intacta
        return self

    def __getitem__(self, chave):
        return self

    def __contains__(self, item):
        return False

    def __bool__(self):
        return False

    def __iter__(self):
        return iter(())

    def __str__(self):
        return ""


def carregar_funcoes_do_app(caminho=CAMINHO_APP, sementes=SEMENTES):
    """Extrai do app.py as funções/constantes pedidas MAIS tudo o que elas
    usam, e executa num espaço isolado. Devolve o espaço (dict nome -> objeto).

    A busca de dependências é por NOME: qualquer identificador dentro de uma
    função que seja função ou constante de nível de módulo entra também. Pode
    trazer um pouco a mais (nome local homônimo), nunca a menos."""
    with open(caminho, encoding="utf-8") as arquivo:
        fonte = arquivo.read()
    arvore = ast.parse(fonte)
    # Entram no catálogo as FUNÇÕES e as CONSTANTES (nome em maiúsculas, com
    # ou sem _ na frente). Variável de execução do app -- list_df_real, m_map,
    # df_ref -- fica de fora de propósito: ela também é nome de parâmetro
    # das funções, e incluí-la arrastaria meio app (e o Streamlit) junto.
    eh_constante = re.compile(r"_?[A-Z][A-Z0-9_]*").fullmatch
    catalogo = {}
    for no in arvore.body:
        if isinstance(no, ast.FunctionDef):
            catalogo[no.name] = no
        elif isinstance(no, ast.Assign):
            for alvo in no.targets:
                if isinstance(alvo, ast.Name) and eh_constante(alvo.id):
                    catalogo[alvo.id] = no
        elif (isinstance(no, ast.AnnAssign) and isinstance(no.target, ast.Name)
              and eh_constante(no.target.id)):
            catalogo[no.target.id] = no
    faltam = [s for s in sementes if s not in catalogo]
    if faltam:
        raise SystemExit(f"O app.py não tem mais: {', '.join(faltam)} -- o briefing precisa acompanhar.")
    incluidos, fila = set(), list(sementes)
    while fila:
        nome = fila.pop()
        if nome in incluidos:
            continue
        incluidos.add(nome)
        for sub in ast.walk(catalogo[nome]):
            if isinstance(sub, ast.Name) and sub.id in catalogo and sub.id not in incluidos:
                fila.append(sub.id)
    nos = sorted((catalogo[n] for n in incluidos), key=lambda n: n.lineno)
    st = _StreamlitDeMentira()
    espaco = {
        "pd": pd, "np": np, "re": re, "io": io, "os": os, "math": math, "time": time,
        "base64": base64, "hashlib": hashlib, "hmac": hmac, "contextlib": contextlib,
        "datetime": datetime, "timedelta": timedelta, "ZoneInfo": ZoneInfo,
        "gc": gc, "urllib": urllib, "Workbook": Workbook, "Font": Font, "PatternFill": PatternFill,
        "Alignment": Alignment, "Border": Border, "Side": Side,
        "st": st, "components": st, "go": st,
    }
    exec("\n\n".join(ast.unparse(no) for no in nos), espaco)   # noqa: S102 -- fonte do próprio repositório
    return espaco


def montar_briefing(ns, hoje=None, aba=None, url_orc=None, url_real=None, url_fech=""):
    """Carrega as planilhas e devolve (fatos, itens_da_narrativa, contexto)."""
    hoje = hoje or datetime.now(ns["FUSO_BR"]).date()
    hoje_dados = hoje - timedelta(days=ns["DIAS_DEFASAGEM_DADOS"])
    aba = aba or os.environ.get("ABA_DRE", "DRE CONSOLIDADO")
    url_orc = url_orc or urls_das_planilhas()[0]
    url_real = url_real or urls_das_planilhas()[1]

    list_df_orc, list_df_real = ns["carregar_dados_abas"](url_orc, url_real, [aba])
    df_ref = next((d for d in list_df_real if d is not None and not d.empty), None)
    if df_ref is None:
        raise SystemExit(f"Não consegui ler a aba '{aba}' do Realizado.")
    meses_cols = [f"{m:02d}/{hoje.year}" for m in range(1, 13)]
    m_map = {n: c for n, c in zip(NOMES_MESES, meses_cols) if c in df_ref.columns}
    cols_kpi = [c for c in m_map.values() if int(c[:2]) <= hoje.month]
    if not cols_kpi:
        raise SystemExit("A planilha não tem as colunas de mês deste ano.")
    col_nome = "Nome" if "Nome" in df_ref.columns else df_ref.columns[0]
    linhas = list(df_ref[col_nome].dropna().unique().astype(str))
    gv = ns["get_valor_consolidado_multi"]

    def valor(lado, linha, cols, exato=False):
        return gv(list_df_orc if lado == "orc" else list_df_real, linha, cols,
                  exato_linha_sintetica=exato)

    ritmo = ns["fatos_do_ritmo"](list_df_real, list_df_orc, cols_kpi, meses_cols, m_map, hoje_dados)

    pendencias, margem_proj = [], None
    if url_fech:
        df_fech, erro = ns["carregar_planilha_fechamento"](ns["url_csv_do_fechamento"](url_fech))
        if not erro and df_fech is not None and not df_fech.empty:
            d_rec, d_eb, pendencias = ns["_deltas_pendentes_do_fechamento"](
                df_fech, cols_kpi, meses_cols, (hoje.year, hoje.month),
                lambda linha, col: gv(list_df_real, linha, [col]))
            margem_proj = ns["projetar_margem_fechamento"](
                valor("real", "3 - Receita Operacional Liquida", cols_kpi),
                valor("real", "11 - EBITDA", cols_kpi), d_rec, d_eb)

    col_corrente = f"{hoje.month:02d}/{hoje.year}"
    nome_corrente = next((n for n, c in m_map.items() if c == col_corrente), col_corrente)
    rotulo = f"Acumulado YTD até {str(nome_corrente).capitalize()}"
    fatos = ns["montar_fatos_executivos"](
        valor, linhas, cols_kpi, col_corrente, nome_corrente,
        pendencias=pendencias, ritmo=ritmo, margem_proj=margem_proj, rotulo_periodo=rotulo)
    itens = ns["montar_narrativa_executiva"](fatos)
    # Meses FECHADOS: bateu (True) ou não (False) a receita orçada. É o gabarito
    # do placar da chance.
    resultado_por_mes = {}
    for c in meses_cols:
        if int(c[:2]) >= hoje.month:
            continue
        o = gv(list_df_orc, "3 - Receita Operacional Liquida", [c])
        r_ = gv(list_df_real, "3 - Receita Operacional Liquida", [c])
        if o > 0 and r_ > 0:
            resultado_por_mes[c] = r_ >= o
    contexto = {"hoje": hoje, "hoje_dados": hoje_dados, "aba": aba, "rotulo": rotulo,
                "resultado_por_mes": resultado_por_mes, "list_df_real": list_df_real, "linhas": linhas,
                "list_df_orc": list_df_orc, "meses_cols": meses_cols, "m_map": m_map}
    return fatos, itens, contexto


# ---------------------------------------------------------------------------
# HISTÓRICO: o que cada briefing viu, para o próximo dizer "desde ontem"
# ---------------------------------------------------------------------------
# O GitHub Actions grava este arquivo de volta no repositório a cada envio.
# Duas coisas nascem dele: a linha "desde o último briefing" (chefe lê
# delta, não foto) e o placar da chance de bater a meta (a previsão do dia
# 15 conferida contra o fechamento real -- é o que dá credibilidade ao
# número).
CAMINHO_HISTORICO = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "dados", "historico_briefing.json")
LIMITE_HISTORICO = 400
DIA_DE_CORTE_DO_PLACAR = 15


def carregar_historico(caminho=CAMINHO_HISTORICO):
    try:
        with open(caminho, encoding="utf-8") as arquivo:
            dados = json.load(arquivo)
        return dados if isinstance(dados, list) else []
    except (OSError, ValueError):
        return []


def salvar_historico(historico, caminho=CAMINHO_HISTORICO):
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    with open(caminho, "w", encoding="utf-8") as arquivo:
        json.dump(historico, arquivo, ensure_ascii=False, indent=1)


def entrada_do_dia(fatos, hoje):
    """O que vale guardar de um briefing: pouco, e sempre no mesmo formato."""
    r = fatos.get("ritmo") or {}
    return {
        "data": hoje.isoformat(),
        "mes_col": r.get("col"), "mes": r.get("mes"), "dia": r.get("dia"),
        "rec_mes": r.get("rec_real"), "ritmo_pct": r.get("pct"), "chance": r.get("chance"),
        "rec_ytd": fatos.get("rec_real"), "ebitda_ytd": fatos.get("ebitda_real"),
        "margem_proj": fatos.get("margem_proj"),
        "estouros": [e["conta"] for e in (fatos.get("estouros") or [])],
    }


def registrar_no_historico(historico, entrada, limite=LIMITE_HISTORICO):
    """Substitui a entrada do mesmo dia (rodar duas vezes não duplica) e
    mantém só as últimas `limite`."""
    novo = [h for h in (historico or []) if h.get("data") != entrada["data"]]
    novo.append(entrada)
    novo.sort(key=lambda h: h.get("data", ""))
    return novo[-limite:]


def entrada_anterior(historico, hoje):
    """O último briefing ANTES de hoje -- na segunda, é o de sexta."""
    alvo = hoje.isoformat()
    anteriores = [h for h in (historico or []) if h.get("data", "") < alvo]
    return anteriores[-1] if anteriores else None


def comparar_com_ontem(atual, anterior, fmt=None):
    """Frases curtas do que mudou. Vazio sem base; se o mês virou, diz isso
    em vez de comparar receita de setembro com receita de agosto."""
    if not anterior:
        return []
    fmt = fmt or (lambda v: f"R$ {v:,.0f}")
    frases = []
    mesmo_mes = atual.get("mes_col") and atual.get("mes_col") == anterior.get("mes_col")
    if mesmo_mes:
        if atual.get("rec_mes") is not None and anterior.get("rec_mes") is not None:
            d = atual["rec_mes"] - anterior["rec_mes"]
            frases.append(f"receita de {str(atual.get('mes', '')).lower()} "
                          f"{'+' if d >= 0 else '−'}{fmt(abs(d))}")
        if atual.get("ritmo_pct") is not None and anterior.get("ritmo_pct") is not None:
            frases.append(f"ritmo {anterior['ritmo_pct']:.0f}% → {atual['ritmo_pct']:.0f}%")
        if atual.get("chance") is not None and anterior.get("chance") is not None:
            frases.append(f"chance de bater a meta {anterior['chance'] * 100:.0f}% → "
                          f"{atual['chance'] * 100:.0f}%")
    elif anterior.get("mes_col") and atual.get("mes_col"):
        frases.append(f"o mês de referência virou: {str(anterior.get('mes', '')).capitalize()} → "
                      f"{str(atual.get('mes', '')).capitalize()}")
    novos = [c for c in (atual.get("estouros") or []) if c not in (anterior.get("estouros") or [])]
    if novos:
        frases.append("conta nova acima do orçado: " + ", ".join(novos))
    elif atual.get("estouros") is not None:
        frases.append("nenhuma conta nova acima do orçado")
    return frases


def placar_da_chance(historico, resultado_por_mes, dia_corte=DIA_DE_CORTE_DO_PLACAR):
    """Confere a chance registrada até o dia de corte de cada mês com o
    resultado real (bateu ou não). Devolve {meses, acertos, detalhes}."""
    por_mes = {}
    for h in (historico or []):
        col, dia, chance = h.get("mes_col"), h.get("dia"), h.get("chance")
        if not col or dia is None or chance is None or dia > dia_corte:
            continue
        if col not in por_mes or dia >= por_mes[col]["dia"]:
            por_mes[col] = {"dia": dia, "chance": chance}
    detalhes = []
    for col, bateu in (resultado_por_mes or {}).items():
        if col in por_mes and bateu is not None:
            previu = por_mes[col]["chance"] >= 0.5
            detalhes.append({"mes": col, "dia": por_mes[col]["dia"], "chance": por_mes[col]["chance"],
                             "bateu": bool(bateu), "acertou": previu == bool(bateu)})
    return {"meses": len(detalhes), "acertos": sum(1 for d in detalhes if d["acertou"]),
            "detalhes": detalhes}


def series_mensais(ns, list_df_real, list_df_orc, meses_cols, m_map, ate_mes):
    """Receita e EBITDA por mês, realizado e orçado, até `ate_mes` (1-12).
    Alimenta os gráficos do board pack."""
    gv = ns["get_valor_consolidado_multi"]
    saida = {"rotulos": [], "rec_real": [], "rec_orc": [], "eb_real": [], "eb_orc": []}
    for nome, col in m_map.items():
        if int(col[:2]) > ate_mes:
            continue
        saida["rotulos"].append(nome.capitalize()[:3])
        saida["rec_real"].append(gv(list_df_real, "3 - Receita Operacional Liquida", [col]))
        saida["rec_orc"].append(gv(list_df_orc, "3 - Receita Operacional Liquida", [col]))
        saida["eb_real"].append(gv(list_df_real, "11 - EBITDA", [col]))
        saida["eb_orc"].append(gv(list_df_orc, "11 - EBITDA", [col]))
    return saida


# ---------------------------------------------------------------------------
# DEPARTAMENTOS: o escopo de cada gestor, para alertas e board pack por área
# ---------------------------------------------------------------------------
def urls_das_planilhas():
    return (os.environ.get("URL_ORCADO", URL_ORCADO_PADRAO),
            os.environ.get("URL_REALIZADO", URL_REALIZADO_PADRAO))


def ja_enviado_hoje(historico, hoje):
    """Idempotência do dia: o agendador do GitHub atrasa e às vezes falha, então
    há DOIS horários agendados; o segundo vê que já saiu e fica quieto."""
    alvo = hoje.isoformat()
    return any(h.get("data") == alvo for h in (historico or []))


def emails_do_departamento(departamento, mapa_email_departamento, emails_travados=None):
    """Quem é gestor daquele departamento, pelos mapas de acesso do próprio
    app (e-mail -> departamento). Vale para alerta e para o board pack."""
    juntos = {**(emails_travados or {}), **(mapa_email_departamento or {})}
    return sorted({e for e, d in juntos.items() if d == departamento})


def linhas_do_departamento(modelo, linhas_da_visao, ns):
    """Resolve o escopo de um departamento em linhas da DRE.
    - ["ATE_EBITDA"]: todos os subgrupos de despesa (nível 2 de 6 e 8) da visão;
    - lista explícita: ela mesma (as linhas que existem na visão);
    - "RESTANTE" / só planos de contas forçados: None -- esse recorte vive no
      plano de contas dentro do app, e fora dele seria chute."""
    linhas = [str(l) for l in (modelo.get("linhas_dre") or [])]
    existentes = set(linhas_da_visao)
    if linhas == ["ATE_EBITDA"]:
        return [l for g in ("6", "8") for l in ns["_subgrupos_nivel2"](list(linhas_da_visao), g)]
    explicitas = [l for l in linhas if l in existentes]
    if explicitas:
        return explicitas
    return None


def fatos_do_departamento(valor, linhas, cols, col_corrente, nome, ns, rotulo_periodo=""):
    """Gasto realizado x orçado do departamento nos meses fechados, com o
    pódio de contas. `valor(lado, linha, cols, exato)` injetado, como sempre."""
    cols_fech = [c for c in cols if c != col_corrente] or list(cols)
    itens = []
    for linha in linhas:
        r = abs(valor("real", linha, cols_fech, True))
        o = abs(valor("orc", linha, cols_fech, True))
        if not r and not o:
            continue
        itens.append({"conta": ns["_nome_sem_numero_dre"](linha), "linha": linha, "realizado": r, "orcado": o})
    acima = ns["ofensores_por_desvio"](itens)
    folgas = sorted(({"conta": i["conta"], "folga": i["orcado"] - i["realizado"],
                      "pct": ((i["orcado"] - i["realizado"]) / i["orcado"] * 100) if i["orcado"] else None}
                     for i in itens if i["orcado"] - i["realizado"] > 0), key=lambda x: -x["folga"])
    return {
        "departamento": nome, "periodo": rotulo_periodo,
        "gasto_real": sum(i["realizado"] for i in itens), "gasto_orc": sum(i["orcado"] for i in itens),
        "estouros": [{"conta": e["conta"], "desvio": e["desvio"],
                      "pct": (e["desvio"] / e["orcado"] * 100) if e["orcado"] else None} for e in acima[:5]],
        "folgas": folgas[:5], "n_contas": len(itens), "n_acima": len(acima),
        "mes_excluido": col_corrente if col_corrente in cols and len(cols) > 1 else None,
    }


def narrativa_departamento(f, ns=None):
    fmt = (ns or {}).get("formata_valor_curto", lambda v: f"R$ {v:,.0f}")
    pct = (ns or {}).get("_pct_br", lambda v, casas=1: f"{v:.{casas}f}".replace(".", ","))
    itens = []
    if f.get("gasto_orc"):
        var = (f["gasto_real"] / f["gasto_orc"] - 1) * 100
        itens.append({"rotulo": "Gasto", "tom": "negativo" if var > 0 else "positivo",
                      "texto": (f"<b>{fmt(f['gasto_real'])}</b> nos meses fechados, "
                                f"<b>{pct(abs(var))}% {'acima' if var > 0 else 'abaixo'}</b> do orçado "
                                f"({fmt(f['gasto_orc'])}).")})
    if f.get("estouros"):
        partes = [f"<b>{e['conta']}</b> (+{fmt(e['desvio'])}"
                  + (f", +{e['pct']:.0f}%" if e.get("pct") is not None else ", sem orçamento") + ")"
                  for e in f["estouros"][:3]]
        itens.append({"rotulo": "Estouros", "tom": "negativo",
                      "texto": f"Quem mais passou do orçado: {_lista_pt(partes)}."})
    if f.get("folgas"):
        partes = [f"<b>{c['conta']}</b> (−{fmt(c['folga'])}"
                  + (f", −{c['pct']:.0f}%" if c.get("pct") is not None else "") + ")" for c in f["folgas"][:2]]
        itens.append({"rotulo": "Folgas", "tom": "positivo", "texto": f"Sobrou em {_lista_pt(partes)}."})
    if f.get("n_contas"):
        itens.append({"rotulo": "Cobertura", "tom": "neutro",
                      "texto": f"<b>{f['n_acima']} de {f['n_contas']}</b> contas do departamento acima do orçado."})
    return itens


def _lista_pt(partes):
    partes = [p for p in partes if p]
    return "".join(partes) if len(partes) <= 1 else ", ".join(partes[:-1]) + " e " + partes[-1]


def series_departamento(valor, linhas, m_map, ate_mes):
    saida = {"rotulos": [], "gasto_real": [], "gasto_orc": []}
    for nome, col in m_map.items():
        if int(col[:2]) > ate_mes:
            continue
        saida["rotulos"].append(nome.capitalize()[:3])
        saida["gasto_real"].append(sum(abs(valor("real", l, [col], True)) for l in linhas))
        saida["gasto_orc"].append(sum(abs(valor("orc", l, [col], True)) for l in linhas))
    return saida


# ---------------------------------------------------------------------------
# O E-MAIL
# ---------------------------------------------------------------------------
# Regras de e-mail, que não são as da web: tabela para tudo (Outlook não
# entende flex/grid), estilo INLINE em cada célula (o Gmail apaga <style>),
# fundo claro (cliente de celular ignora tema escuro), fonte segura, largura
# de 640px e uma coluna só no miolo. Logo por CID (imagem anexada ao próprio
# e-mail): o Gmail bloqueia imagem embutida em base64 e a de link externo
# chega "bloqueada" até a pessoa clicar.
CORES = {
    "fundo": "#EEF1F5", "cartao": "#FFFFFF", "borda": "#E3E7ED", "texto": "#1F2937",
    "apagado": "#6B7280", "marca": "#1B2A41", "marca_claro": "#9FB3D1",
    "positivo": "#1E8449", "negativo": "#C0392B", "alerta": "#B9770E", "neutro": "#6B7280",
}
FONTE = "'Segoe UI', Helvetica, Arial, sans-serif"
CID_LOGO = "logo-grupo-beea"


def _tirar_tags(texto):
    return re.sub(r"<[^>]+>", "", str(texto or ""))


def status_geral(fatos):
    """Semáforo do topo. Mede o EBITDA na BASE FECHADA (estrutural), não o
    ritmo de um dia: no dia 1 do mês o ritmo é ruído, e um semáforo que
    acende vermelho por dois dias de venda perde a credibilidade na terceira
    manhã. Devolve (rótulo, cor, frase curta para o assunto)."""
    fech = fatos.get("fechado") or {}
    eb_r = fech.get("ebitda_real", fatos.get("ebitda_real"))
    eb_o = fech.get("ebitda_orc", fatos.get("ebitda_orc"))
    if not eb_o or eb_r is None:
        return "SEM ORÇADO", CORES["neutro"], ""
    gap = eb_r / eb_o - 1
    pct = f"{abs(gap) * 100:.1f}".replace(".", ",")
    frase = f"EBITDA {pct}% {'abaixo' if gap < 0 else 'acima'} do orçado"
    if gap >= 0:
        return "NO ORÇADO", CORES["positivo"], frase
    if gap > -0.05:
        return "OBSERVAR", CORES["alerta"], frase
    return "ATENÇÃO", CORES["negativo"], frase


def assunto_do_briefing(fatos, hoje):
    """O assunto já conta a história: quem não abrir o e-mail sabe o essencial."""
    partes = [f"Briefing {hoje.strftime('%d/%m')}"]
    _, _, frase = status_geral(fatos)
    if frase:
        partes.append(frase)
    r = fatos.get("ritmo")
    if r:
        partes.append(f"{str(r['mes']).capitalize()} a {r['pct']:.0f}% do ritmo")
    return " · ".join(partes)


def _cartao_kpi(rotulo, valor, sub, cor_topo, cor_sub):
    return (
        '<td width="50%" valign="top" style="width:50%; padding:6px; vertical-align:top;">'
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
        f'style="width:100%; border:1px solid {CORES["borda"]}; border-top:3px solid {cor_topo}; '
        f'border-radius:6px; background:{CORES["cartao"]};">'
        f'<tr><td style="padding:13px 15px 12px 15px; font-family:{FONTE};">'
        f'<div style="font-size:10px; letter-spacing:1.2px; text-transform:uppercase; color:{CORES["apagado"]};">{rotulo}</div>'
        f'<div style="font-size:25px; font-weight:700; color:{CORES["texto"]}; margin-top:4px; line-height:1.1;">{valor}</div>'
        f'<div style="font-size:12px; color:{cor_sub}; margin-top:6px; line-height:1.4;">{sub}</div>'
        "</td></tr></table></td>")


def montar_email(itens, fatos, hoje, link_painel="", ns=None, logo_src="",
                 desde_ontem=None, data_anterior=None, placar=None):
    """Devolve (html, texto). `logo_src` é "cid:..." no envio e um data-URI
    na prévia gravada em disco; vazio, o cabeçalho usa um selo com as
    iniciais."""
    fmt = (ns or {}).get("formata_valor_curto", lambda v: f"R$ {v:,.0f}")
    pct = (ns or {}).get("_pct_br", lambda v, casas=1: f"{v:.{casas}f}".replace(".", ","))
    dia = f"{DIAS_SEMANA[hoje.weekday()]}, {hoje.strftime('%d/%m/%Y')}"
    status, cor_status, _ = status_geral(fatos)
    r = fatos.get("ritmo")
    mc = fatos.get("mes_corrente") or {}
    tons = {"negativo": CORES["negativo"], "positivo": CORES["positivo"],
            "alerta": CORES["alerta"], "neutro": CORES["neutro"]}

    def _delta(real, orc):
        if not orc or real is None:
            return "", CORES["apagado"]
        var = (real / orc - 1) * 100
        seta = "▲" if var >= 0 else "▼"
        return (f"{seta} {pct(abs(var))}% vs orçado ({fmt(orc)})",
                CORES["positivo"] if var >= 0 else CORES["negativo"])

    # ---- cartões (2 x 2) ----
    cartoes = []
    if fatos.get("rec_real") is not None:
        sub, cor = _delta(fatos["rec_real"], fatos.get("rec_orc"))
        cartoes.append(_cartao_kpi("Receita líquida YTD", fmt(fatos["rec_real"]), sub, cor, cor))
    if fatos.get("ebitda_real") is not None:
        sub, cor = _delta(fatos["ebitda_real"], fatos.get("ebitda_orc"))
        cartoes.append(_cartao_kpi("EBITDA YTD", fmt(fatos["ebitda_real"]), sub, cor, cor))
    if fatos.get("rec_real"):
        margem = fatos["ebitda_real"] / fatos["rec_real"] * 100
        if fatos.get("margem_proj") is not None:
            sub, cor = (f"fecha em <b>{pct(fatos['margem_proj'])}%</b> com os lançamentos pendentes",
                        CORES["alerta"])
        else:
            sub, cor = "realizada no período", CORES["apagado"]
        cartoes.append(_cartao_kpi("Margem EBITDA", f"{pct(margem)}%", sub, CORES["marca"], cor))
    if r:
        cor_r = (CORES["positivo"] if r["pct"] >= 100 else
                 CORES["alerta"] if r["pct"] >= 90 else CORES["negativo"])
        sub = f"do esperado · dia {r['dia']} de {r['dias']}"
        if r.get("chance") is not None:
            sub += f" · chance de bater a meta: <b>{r['chance'] * 100:.0f}%</b>"
        cartoes.append(_cartao_kpi(f"Ritmo de {str(r['mes']).capitalize()}", f"{r['pct']:.0f}%",
                                   sub, cor_r, CORES["apagado"]))
    linhas_kpi = ""
    for i in range(0, len(cartoes), 2):
        par = cartoes[i:i + 2]
        if len(par) == 1:
            par.append('<td width="50%"></td>')
        linhas_kpi += "<tr>" + "".join(par) + "</tr>"

    # ---- desde o último briefing ----
    bloco_delta = ""
    if desde_ontem:
        rotulo_ant = (f"Desde o briefing de {data_anterior[8:10]}/{data_anterior[5:7]}"
                      if data_anterior else "Desde o último briefing")
        bloco_delta = (
            f'<tr><td style="background:{CORES["cartao"]}; padding:4px 26px 8px 26px; font-family:{FONTE};">'
            '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%;"><tr>'
            f'<td style="background:#F4F7FB; border:1px solid {CORES["borda"]}; border-radius:6px; padding:10px 14px;">'
            f'<div style="font-size:10px; letter-spacing:1.2px; text-transform:uppercase; color:{CORES["apagado"]};">{rotulo_ant}</div>'
            f'<div style="font-size:13px; line-height:1.5; color:{CORES["texto"]}; margin-top:3px;">{" · ".join(desde_ontem)}</div>'
            '</td></tr></table></td></tr>')

    # ---- barra de ritmo ----
    bloco_ritmo = ""
    if r and r.get("rec_orc_prop") and r.get("data_dados"):
        cheio = max(2, min(100, int(round(r["pct"]))))
        cor_r = (CORES["positivo"] if r["pct"] >= 100 else
                 CORES["alerta"] if r["pct"] >= 90 else CORES["negativo"])
        bloco_ritmo = (
            f'<tr><td style="background:{CORES["cartao"]}; padding:6px 26px 18px 26px; font-family:{FONTE};">'
            '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%;"><tr>'
            f'<td style="font-size:10px; letter-spacing:1.2px; text-transform:uppercase; color:{CORES["apagado"]};">'
            f'Ritmo de {str(r["mes"]).capitalize()} · realizado vs. meta até {r["data_dados"]}</td>'
            f'<td align="right" style="font-size:12px; color:{CORES["texto"]};">'
            f'<b>{fmt(r.get("rec_real", 0))}</b> de {fmt(r["rec_orc_prop"])}</td></tr></table>'
            '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%; margin-top:7px;"><tr>'
            f'<td width="{cheio}%" bgcolor="{cor_r}" style="width:{cheio}%; background:{cor_r};">'
            '<div style="height:10px; line-height:10px; font-size:1px;">&nbsp;</div></td>'
            f'<td bgcolor="{CORES["borda"]}" style="background:{CORES["borda"]};">'
            '<div style="height:10px; line-height:10px; font-size:1px;">&nbsp;</div></td>'
            "</tr></table>"
            f'<div style="font-size:11px; color:{CORES["apagado"]}; margin-top:6px;">'
            f'A barra cheia é a meta até {r["data_dados"]} (dados chegam D+2). '
            f'Projeção do mês no ritmo atual: <b>{fmt(r.get("rec_proj", 0))}</b> contra {fmt(r.get("rec_orc_cheio", 0))} orçados.</div>'
            "</td></tr>")

    # ---- narrativa ----
    linhas_narrativa = ""
    for i in itens:
        cor = tons.get(i.get("tom"), CORES["neutro"])
        linhas_narrativa += (
            "<tr>"
            f'<td width="4" bgcolor="{cor}" style="width:4px; background:{cor}; font-size:0; line-height:0;">&nbsp;</td>'
            f'<td width="92" valign="top" style="vertical-align:top; padding:13px 8px 11px 12px; font-family:{FONTE}; font-size:10px; '
            f'letter-spacing:1.1px; text-transform:uppercase; font-weight:700; color:{cor}; '
            f'border-bottom:1px solid {CORES["borda"]};">{i["rotulo"]}</td>'
            f'<td valign="top" style="vertical-align:top; padding:11px 0 11px 6px; font-family:{FONTE}; font-size:14px; line-height:1.55; '
            f'color:{CORES["texto"]}; border-bottom:1px solid {CORES["borda"]};">{i["texto"]}</td>'
            "</tr>")

    # ---- cabeçalho ----
    if logo_src:
        selo = (f'<img src="{logo_src}" width="46" height="46" alt="Grupo B&amp;A" '
                'style="display:block; width:46px; height:46px; border-radius:23px;">')
    else:
        selo = ('<div style="width:46px; height:46px; border-radius:23px; background:#FFFFFF; '
                f'color:{CORES["marca"]}; font-family:{FONTE}; font-weight:700; font-size:15px; '
                'text-align:center; line-height:46px;">B&amp;A</div>')
    base = (f"Análise sobre os meses fechados · {str(mc['mes']).lower()} entra só como explicação do gap · "
            if mc else "")
    preheader = (_tirar_tags(itens[0]["texto"])[:140] + "…") if itens else ""
    botao = (
        '<table role="presentation" align="center" cellpadding="0" cellspacing="0" style="margin:6px auto 0 auto;">'
        f'<tr><td bgcolor="{CORES["marca"]}" style="background:{CORES["marca"]}; border-radius:6px;">'
        f'<a href="{link_painel}" style="display:inline-block; padding:12px 28px; font-family:{FONTE}; font-size:14px; '
        'font-weight:600; color:#FFFFFF; text-decoration:none;">Abrir o painel &rarr;</a></td></tr></table>'
        if link_painel else "")

    html = (
        '<meta charset="utf-8">'
        f'<div style="display:none; max-height:0; overflow:hidden; opacity:0;">{preheader}</div>'
        f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" bgcolor="{CORES["fundo"]}" '
        f'style="background:{CORES["fundo"]};"><tr><td align="center" style="padding:24px 12px;">'
        '<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="width:640px; max-width:100%;">'
        # cabeçalho
        f'<tr><td bgcolor="{CORES["marca"]}" style="background:{CORES["marca"]}; padding:22px 26px; border-radius:10px 10px 0 0;">'
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%;"><tr>'
        f'<td width="46" valign="middle">{selo}</td>'
        f'<td valign="middle" style="padding-left:14px; font-family:{FONTE};">'
        f'<div style="font-size:10px; letter-spacing:1.6px; text-transform:uppercase; color:{CORES["marca_claro"]};">'
        'Controladoria B&amp;A · Briefing executivo</div>'
        f'<div style="font-size:20px; font-weight:700; color:#FFFFFF; margin-top:3px; line-height:1.2;">{fatos.get("periodo", "")}</div>'
        f'<div style="font-size:12px; color:{CORES["marca_claro"]}; margin-top:3px;">{dia}'
        + (f' · dados até {r["data_dados"]} (D+2)' if r and r.get("data_dados") else "") + "</div></td>"
        f'<td align="right" valign="top" style="font-family:{FONTE};">'
        f'<span style="display:inline-block; padding:5px 11px; border-radius:12px; background:{cor_status}; '
        f'color:#FFFFFF; font-size:10px; font-weight:700; letter-spacing:1.2px;">{status}</span></td>'
        "</tr></table></td></tr>"
        # KPIs
        f'<tr><td style="background:{CORES["cartao"]}; padding:14px 20px 8px 20px;">'
        f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%;">{linhas_kpi}</table></td></tr>'
        + bloco_delta + bloco_ritmo +
        # narrativa
        f'<tr><td style="background:{CORES["cartao"]}; padding:8px 26px 4px 26px; font-family:{FONTE};">'
        f'<div style="font-size:10px; letter-spacing:1.4px; text-transform:uppercase; color:{CORES["apagado"]}; '
        f'padding-bottom:8px; border-bottom:2px solid {CORES["marca"]};">O que aconteceu e por quê</div>'
        f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="width:100%;">{linhas_narrativa}</table></td></tr>'
        # rodapé
        f'<tr><td align="center" style="background:{CORES["cartao"]}; padding:18px 26px 24px 26px; '
        f'border-radius:0 0 10px 10px; font-family:{FONTE};">{botao}'
        f'<div style="font-size:11px; line-height:1.5; color:{CORES["apagado"]}; margin-top:16px;">'
        f'{base}lançamentos chegam D+2 · a chance de bater a meta é uma estimativa a partir do histórico do ano.<br>'
        f'{texto_placar(placar)}<br>'
        'Gerado automaticamente pelo painel a partir das planilhas de Orçado, Realizado e Fechamento.</div>'
        "</td></tr></table></td></tr></table>")

    kpis_texto = []
    for c in cartoes:
        partes = [_tirar_tags(p) for p in re.findall(r"<div[^>]*>(.*?)</div>", c)]
        if len(partes) >= 3:
            kpis_texto.append(f"{partes[0]}: {partes[1]} ({partes[2]})")
    texto = (f"Controladoria B&A · briefing executivo · {status}\n{fatos.get('periodo', '')} · {dia}\n\n"
             + "\n".join(kpis_texto)
             + ("\n\nDesde o último briefing: " + " · ".join(desde_ontem) if desde_ontem else "")
             + "\n\nO que aconteceu e por quê\n"
             + (ns or {}).get("narrativa_em_texto", lambda it: "\n".join(
                 f"{i['rotulo']}: {_tirar_tags(i['texto'])}" for i in it))(itens)
             + (f"\n\nPainel: {link_painel}" if link_painel else "")
             + "\n\n" + texto_placar(placar)
             + "\n\nGerado automaticamente pelo painel. Lançamentos chegam D+2.")
    return html, texto


def texto_placar(placar):
    """Uma linha honesta sobre o placar: nos primeiros meses ele ainda não
    existe, e dizer isso vale mais do que esconder."""
    if not placar or not placar.get("meses"):
        return ("Placar da chance de bater a meta: começa a contar no primeiro mês fechado "
                "com histórico de briefings.")
    return (f"Placar da chance de bater a meta: acertou {placar['acertos']} de {placar['meses']} "
            f"meses (previsão registrada até o dia {DIA_DE_CORTE_DO_PLACAR}).")


def enviar_email(assunto, html, texto, logo_b64="", anexos=None, destinos=None, credenciais=None):
    """`destinos` e `credenciais` ({"usuario", "senha"}) sobrepõem as variáveis de
    ambiente: é assim que o app envia pelo st.secrets e os alertas falam com
    cada gestor."""
    credenciais = credenciais or {}
    usuario = credenciais.get("usuario") or os.environ["SMTP_USUARIO"]
    senha = credenciais.get("senha") or os.environ["SMTP_SENHA"]
    destinos = list(destinos or [d.strip() for d in os.environ.get("EMAIL_DESTINO", "").split(",") if d.strip()])
    if not destinos:
        raise SystemExit("EMAIL_DESTINO está vazio.")
    servidor = os.environ.get("SMTP_SERVIDOR", "smtp.gmail.com")
    porta = int(os.environ.get("SMTP_PORTA", "465"))
    mensagem = EmailMessage()
    mensagem["Subject"] = assunto
    mensagem["From"] = f"Controladoria B&A <{usuario}>"
    mensagem["To"] = ", ".join(destinos)
    mensagem.set_content(texto)
    mensagem.add_alternative(html, subtype="html")
    if logo_b64:
        # O logo viaja DENTRO do e-mail, ligado ao HTML pelo Content-ID.
        parte_html = mensagem.get_payload()[1]
        parte_html.add_related(base64.b64decode(logo_b64), maintype="image", subtype="jpeg",
                               cid=f"<{CID_LOGO}>")
    for nome, conteudo, maintype, subtype in (anexos or []):
        mensagem.add_attachment(conteudo, maintype=maintype, subtype=subtype, filename=nome)
    with smtplib.SMTP_SSL(servidor, porta, context=ssl.create_default_context()) as conexao:
        conexao.login(usuario, senha)
        conexao.send_message(mensagem)
    return destinos


def main(argv):
    ns = carregar_funcoes_do_app()
    fatos, itens, ctx = montar_briefing(ns, url_fech=os.environ.get("FECHAMENTO_CSV_URL", ""))
    if not itens:
        raise SystemExit("A narrativa saiu vazia -- confira as planilhas.")
    link = os.environ.get("LINK_PAINEL", "")
    logo_b64 = str(ns.get("LOGO_BEEA_B64") or "")
    assunto = assunto_do_briefing(fatos, ctx["hoje"])
    historico = carregar_historico()
    if ja_enviado_hoje(historico, ctx["hoje"]) and "--forcar" not in argv and "--teste" not in argv:
        print(f"O briefing de {ctx['hoje'].strftime('%d/%m')} já foi enviado; nada a fazer (use --forcar para reenviar).")
        return
    anterior = entrada_anterior(historico, ctx["hoje"])
    atual = entrada_do_dia(fatos, ctx["hoje"])
    desde_ontem = comparar_com_ontem(atual, anterior, ns.get("formata_valor_curto"))
    placar = placar_da_chance(historico, ctx["resultado_por_mes"])
    extras = dict(desde_ontem=desde_ontem, data_anterior=(anterior or {}).get("data"), placar=placar)
    if "--salvar" in argv:
        # Na prévia em disco o logo vai embutido (data-URI), que o navegador
        # aceita; no e-mail vai por CID, que os clientes de e-mail aceitam.
        html_previa, _ = montar_email(itens, fatos, ctx["hoje"], link, ns,
                                      logo_src=f"data:image/jpeg;base64,{logo_b64}" if logo_b64 else "",
                                      **extras)
        caminho = argv[argv.index("--salvar") + 1]
        with open(caminho, "w", encoding="utf-8") as arquivo:
            arquivo.write(html_previa)
        print(f"HTML gravado em {caminho}")
    html, texto = montar_email(itens, fatos, ctx["hoje"], link, ns,
                               logo_src=f"cid:{CID_LOGO}" if logo_b64 else "", **extras)
    if "--teste" in argv:
        print(assunto)
        print(texto)
        return
    destinos = enviar_email(assunto, html, texto, logo_b64)
    print(f"Briefing enviado para {', '.join(destinos)}: {assunto}")
    # O histórico só avança depois do envio: um envio que falhou não vira
    # o "ontem" de amanhã.
    salvar_historico(registrar_no_historico(historico, atual))


if __name__ == "__main__":
    main(sys.argv[1:])
